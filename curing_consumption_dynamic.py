"""
B2C Pipeline — Phase 0 Extended: 31-Day Dynamic Curing Consumption
===================================================================
Pre-computes a 31-sheet Excel (one sheet per day) showing how curing
consumption evolves across the May planning horizon as changeovers execute.

Approach: two-pass pre-computation (fully independent of the building scheduler)
  Pass 1 — Build CO schedule from Day 0 data:
    - Runner-Out presses: CO to highest-urgency NRI target, Day 1+
    - Runner-In presses : CO fires on the day demand is fulfilled (instantly)
    - NRI SKUs          : receive a press via CO, ranked by urgency score
    - Max 8 COs per day (plant-wide hard limit)
    - CO target urgency  = f(Priority_Score, production_days vs horizon remaining)
      Class A (CRITICAL): current_production_days > horizon_left  → can't meet demand without CO
      Class B (HELPFUL) : current_production_days ≤ horizon_left  → already fulfillable
      Sort: Class A first, then −Priority_Score, then after_CO_days ASC

  Pass 2 — Simulate 31 days using that CO schedule:
    - Running_Press_Count updated per CO event (new SKU from Shift C of CO day)
    - Updated_Demand_Qty decremented daily by Total_GT_Per_Shift × 3 shifts
    - production_days recomputed each day from remaining demand and press count
    - NRI SKUs before their CO fires: Running_Press_Count=0, Total_GT=0, production_days=blank

Outputs
  data/output/curing_consumption_<days>day_<plan_start>.xlsx  (bc_config.DYNAMIC_CC_OUTPUT)
    - Sheets Day_01 … Day_31  : per-day consumption table
    - Sheet  CO_Schedule       : full changeover plan (press, day, old_sku → new_sku)
    - Sheet  Day0_Summary      : same as existing curing_consumption_table.xlsx (for reference)

Standalone usage:
    python curing_consumption_dynamic.py
"""

from __future__ import annotations

import math
import os
import sys
import warnings
from datetime import datetime, timedelta
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

import bc_config

HERE    = os.path.dirname(os.path.abspath(__file__))
IN_DIR  = bc_config.INPUT_DIR
OUT_DIR = bc_config.OUTPUT_DIR


# Now, this is the inpput and output sheet, i have production data and receipe master and i wants final output in the output. We have to take each unique value from the recipeID and then count how many times occuring in the production data sheet, and then search this recipeID in the receipe master "id" column, and then extract SKUCode from this. Then, i wants this SKUCode and Requirement as the total count. You can refer from the 2 output samples as well. I wants the python script as well as the complete output sheet within the excel file

# Demand file- June (production data- SKUCOde , requirement, prioirty score)

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

# ── All scheduling params imported from bc_config (single source of truth) ───
from bc_config import (
    PLAN_START,
    PLANNING_DAYS,
    MAX_CHANGEOVERS_PER_DAY as MAX_CO_PER_DAY,
    SHIFTS_PER_DAY,
    CO_CLASS_B_THRESHOLD,
    DYNAMIC_CC_OUTPUT,
)
import bc_config as _bc
from bc_config import RUNNING_MOULDS_TABLE, RUNNING_MOULDS_MONTH, PLAN_MONTH


# ══════════════════════════════════════════════════════════════════════════════
# ETL + CONFIG (merged from the former curing_consumption.py — single source now)
# ══════════════════════════════════════════════════════════════════════════════

class ConsumptionConfig:
    # ── planning horizon ──────────────────────────────────────────────────────
    PLANNING_DAYS      = 30
    HOURS_PER_SHIFT    = 8
    SHIFT_MINS         = 480        # minutes per shift

    # ── press physics ─────────────────────────────────────────────────────────
    CAVITIES_PER_MOULD    = 2
    LOAD_UNLOAD_BUFFER_MIN = 0
    PRESS_EFFICIENCY       = 0.94
    # Default EFFECTIVE cycle time for SKUs missing from the CT master.
    # Already includes buffer + efficiency — do NOT re-apply the formula.
    DEFAULT_CYCLE_TIME_MIN = 17.0

    # ── downtime (minutes) — used for press exclusion logic ───────────────────

    # ── changeover cap ────────────────────────────────────────────────────────

    # ── database ──────────────────────────────────────────────────────────────
    DB_NAME = bc_config.ENV.get("JKT_DB_DATABASE", "jkplanningV1")


# ══════════════════════════════════════════════════════════════════════════════
# ETL  (adapted from curing_lp.ETL — loads the three DB tables we need)
# ══════════════════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════════════════
# SKU CLASSIFIER
# ══════════════════════════════════════════════════════════════════════════════

class SKUClassifier:
    """Classify demand SKUs into Runner-In and Non-Runner-In only."""

    def classify(
        self,
        df_demand: pd.DataFrame,
        df_running_moulds: pd.DataFrame,
    ) -> pd.DataFrame:
        """
        Returns DataFrame with columns:
          [SKUCode, Category, RunningPressCount, MouldLife_min]

        Only demand SKUs are classified — presses running non-demand SKUs
        (Runner-Out) are excluded from the consumption table entirely.
        """
        demand_skus  = set(df_demand["SKUCode"].str.strip())
        # Group running moulds by SKU to count active presses
        press_count = (
            df_running_moulds.groupby("SKUCode")
            .agg(
                RunningPressCount=("Machine", "count"),
                MouldLife_min=("MouldLife_remaining", "min"),
            )
            .reset_index()
        )
        press_count["SKUCode"] = press_count["SKUCode"].str.strip()
        running_skus = set(press_count["SKUCode"])

        # Only iterate demand SKUs — Runner-Out (non-demand) are not included
        rows = []
        for sku in sorted(demand_skus):
            is_running = sku in running_skus
            cat = "Runner-In" if is_running else "Non-Runner-In"

            pc_row = press_count[press_count["SKUCode"] == sku]
            run_count  = int(pc_row["RunningPressCount"].values[0]) if len(pc_row) else 0
            mould_life = int(pc_row["MouldLife_min"].values[0])     if len(pc_row) else 0
            rows.append({
                "SKUCode":           sku,
                "Category":          cat,
                "RunningPressCount": run_count,
                "MouldLife_min":     mould_life,
            })

        return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# CYCLE TIME RESOLVER
# ══════════════════════════════════════════════════════════════════════════════

class CycleTimeResolver:
    """Resolve effective cycle time per SKU; fall back to DEFAULT_CYCLE_TIME_MIN."""

    def resolve(
        self,
        skus: list[str],
        df_cycle_times: pd.DataFrame,
    ) -> dict[str, float]:
        """Returns {SKUCode: effective_CT_min}."""
        ct_lookup = dict(zip(
            df_cycle_times["SKUCode"].str.strip(),
            df_cycle_times["CycleTime_min"].astype(float),
        ))
        result = {}
        for sku in skus:
            ct = ct_lookup.get(sku)
            if ct is None or (isinstance(ct, float) and math.isnan(ct)) or ct <= 0:
                ct = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
            result[sku] = float(ct)
        return result


# ══════════════════════════════════════════════════════════════════════════════
# CONSUMPTION CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

class ConsumptionCalculator:
    """Compute GT consumption per shift per SKU and build the planning-horizon table."""

    def compute(
        self,
        df_classify: pd.DataFrame,
        ct_map: dict[str, float],
        df_demand: pd.DataFrame,
        plan_start: datetime,
        planning_days: int = 30,
    ) -> pd.DataFrame:
        """
        Returns the full consumption DataFrame covering all planning shifts.

        Columns:
          SKUCode, Category, Running_Press_Count, Effective_CT_Min,
          Qty_Per_Press_Per_Shift, Total_GT_Per_Shift_Day0,
          Demand_Qty, Priority_Score
        """
        # Merge demand info
        demand_lookup  = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Quantity"]))
        priority_lookup = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Priority"]))

        records = []
        for _, row in df_classify.iterrows():
            sku        = row["SKUCode"]
            category   = row["Category"]
            press_count = int(row["RunningPressCount"])

            ct = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
            qty_per_press = math.floor(ConsumptionConfig.SHIFT_MINS / ct) \
                            * ConsumptionConfig.CAVITIES_PER_MOULD
            total_gt = press_count * qty_per_press

            records.append({
                "SKUCode":                  sku,
                "Category":                 category,
                "Running_Press_Count":       press_count,
                "MouldLife_min":             int(row["MouldLife_min"]),
                "Effective_CT_Min":          ct,
                "Qty_Per_Press_Per_Shift":   qty_per_press,
                "Total_GT_Per_Shift_Day0":   total_gt,
                "Demand_Qty":                demand_lookup.get(sku, 0),
                "Priority_Score":            priority_lookup.get(sku, 0),
            })

        df = pd.DataFrame(records)
        # Sort: Runner-In first, then Runner-Out, then Non-Runner-In; within each by priority desc
        cat_order = {"Runner-In": 0, "Runner-Out": 1, "Non-Runner-In": 2}
        df["_cat_ord"] = df["Category"].map(cat_order)
        df = df.sort_values(["_cat_ord", "Priority_Score"], ascending=[True, False]) \
               .drop(columns=["_cat_ord"]) \
               .reset_index(drop=True)
        return df

    def build_shift_index(
        self,
        plan_start: datetime,
        planning_days: int,
    ) -> list[tuple[datetime, str, int]]:
        """
        Returns list of (shift_start_dt, shift_label, shift_idx) for all planning shifts.
        Shift labels: A (07-15), B (15-23), C (23-07+1).
        """
        shifts = []
        shift_hours = [7, 15, 23]
        shift_labels = ["A", "B", "C"]
        # Start one shift before plan_start (building pre-start shift)
        pre_start = plan_start - timedelta(hours=ConsumptionConfig.HOURS_PER_SHIFT)
        for day_offset in range(-1, planning_days):  # -1 = pre-start day
            base_date = plan_start.date() + timedelta(days=day_offset)
            for sh, label in zip(shift_hours, shift_labels):
                dt = datetime(base_date.year, base_date.month, base_date.day, sh, 0, 0)
                if dt >= pre_start:
                    shifts.append((dt, label, len(shifts)))
        return shifts


# ══════════════════════════════════════════════════════════════════════════════
# SKU ELIGIBILITY FILTER
# ══════════════════════════════════════════════════════════════════════════════

class SKUEligibilityFilter:
    """
    Checks each demand SKU against building and curing master + history data.

    Eligibility rules:
      - Building OK  : SKU in (Master_Building_Allowable OR Building_Stage1/2_History)
      - Curing OK    : SKU in (Master_Curing_Allowable  OR testing_Daily_Running_Moulds history)
      - BOTH must be OK; failing either → excluded with remark

    CT missing is NOT an exclusion criterion — default CT = 17 min is used instead.
    """

    def filter(
        self,
        df_demand: pd.DataFrame,
        bld_master_skus: set,
        bld_history_skus: set,
        cur_master_skus: set,
        cur_history_skus: set,
    ) -> tuple:
        """
        Returns (df_eligible, df_excluded).

        df_excluded columns:
          SKUCode, Demand_Qty, Priority_Score, Remark
        """
        bld_pool = {s.upper() for s in bld_master_skus}
        cur_pool = {s.upper() for s in cur_master_skus}

        eligible_rows: list[dict] = []
        excluded_rows: list[dict] = []

        for _, row in df_demand.iterrows():
            sku    = str(row["SKUCode"]).strip()
            sku_up = sku.upper()
            in_bld = sku_up in bld_pool
            in_cur = sku_up in cur_pool

            if in_bld and in_cur:
                eligible_rows.append(row.to_dict())
            else:
                missing = []
                if not in_bld:
                    missing.append("building machine")
                if not in_cur:
                    missing.append("curing mould")
                excluded_rows.append({
                    "SKUCode":        sku,
                    "Demand_Qty":     float(row.get("Quantity", 0)),
                    "Priority_Score": float(row.get("Priority", 0)),
                    "Remark": f"No PDE & master data- {', '.join(missing)}",
                })

        df_eligible = (
            pd.DataFrame(eligible_rows)
            if eligible_rows
            else df_demand.iloc[0:0].copy()
        )
        df_excluded = (
            pd.DataFrame(excluded_rows)
            if excluded_rows
            else pd.DataFrame(
                columns=["SKUCode", "Demand_Qty", "Priority_Score", "Remark"]
            )
        )
        return df_eligible, df_excluded



# ── Plant-holiday support: working-days-left urgency horizon ──────────────────
# PLANT_HOLIDAYS (bc_config) is a list of "YYYY-MM-DD" strings (empty by default).
# Urgency (Class A/B CO gating) must count WORKING days remaining, not calendar days,
# so a holiday-shortened horizon fires COs on time. Empty holidays ⇒ working==calendar
# ⇒ byte-for-byte identical to today (the parity guarantee).
def _holiday_day_index_set(plan_start=None):
    """1-based day indices (relative to plan_start) that are plant holidays.
    #5 robustness: read the module PLAN_START at CALL time (not a def-time default) so a
    run that syncs PLAN_START gets the right holiday indices — the default-arg gotcha would
    otherwise pin the import-time value and diverge from the main loop's plan_start."""
    if plan_start is None:
        plan_start = PLAN_START
    out = set()
    base = plan_start.date() if hasattr(plan_start, "date") else plan_start
    for _h in (getattr(_bc, "PLANT_HOLIDAYS", []) or []):
        try:
            _idx = (datetime.strptime(str(_h).strip(), "%Y-%m-%d").date() - base).days + 1
            out.add(_idx)
        except Exception:
            pass
    return out


def _working_days_left(day, planning_days, holiday_set):
    """Count working (non-holiday) days in [day, planning_days] inclusive."""
    if not holiday_set:
        return planning_days - day + 1
    return sum(1 for d in range(day, planning_days + 1) if d not in holiday_set)

# Curing-side ratio alignment: replaces Priority_Score in _urgency_sort_key with
# static demand[target]/press_total_demand[press] (never decremented), mirroring
# the building-side _BUILDING_RATIO_ENABLED mechanism in b2c_pipeline.py. Class
# A/B urgency gating is untouched — this only changes ranking within a class.
_CURING_RATIO_ENABLED = True

# Surplus RI-press early-release (env SURPLUS_RELEASE=1, default OFF). Detects
# over-provisioned Runner-In SKUs (press_count > presses_needed to meet demand by
# the horizon) and releases the SURPLUS presses EARLY, spread across days, instead
# of dumping them all at once when demand finally hits 0 (the month-end CO cliff).
# Guards: (5a) only to a compatible needy SKU; (5b) only if building can supply the
# target (buildable_rate passed in from b2c_pipeline); n-1 RI-protection preserved.
# Flip True/False to turn surplus-release on/off (env SURPLUS_RELEASE also works):
# _SURPLUS_RELEASE_ENABLED  = os.environ.get("SURPLUS_RELEASE") == "1"
_SURPLUS_RELEASE_ENABLED  = True
# (Deprecated by the global-pairing redesign — surplus is now uncapped; the daily
# CO cap + global pairing + n-1 RI-protection decide how many presses actually move.)
_SURPLUS_PER_SKU_PER_DAY  = int(os.environ.get("SURPLUS_PER_DAY", "2"))

# #4 Mould-aware Phase-0 (P0_MOULD_GATE, default OFF = current plan bit-for-bit). Phase-0 is
# otherwise mould-BLIND (it ranks CO targets by allowable-machine eligibility only), so ~half
# its planned COs get retargeted downstream by the rolling 2-mould gate to a DIFFERENT SKU.
# When ON, Phase-0 keeps a day-granularity free-mould tracker (seeded Day-0 from the running
# moulds) and only fires a CO to a target that has >= 2 eligible FREE moulds — so it picks a
# mould-feasible target itself instead of one that will be redirected. The rolling gate in
# b2c_pipeline.py stays as the AUTHORITATIVE final filter (Phase-0 can only APPROXIMATE the
# shift-by-shift mould contention). Must A/B — may move KPI either way.
_P0_MOULD_GATE = os.environ.get("P0_MOULD_GATE", "0") != "0"

# DELIVERY_PRIORITY Phase-0 sub-levers (bisection A/B; all default ON when the feature is
# active). DP_ACQUIRE=0 → drop the EDF ordering + deadline Class-A + deadline RI-skip (stop
# forcing committed presses to be acquired). DP_RESERVE=0 → drop the "never CO a committed
# press away pre-deadline" guards. DP_MOULDCAP=1 → cap committed-SKU acquisition at its
# mould-pair count (needs sku_moulds). Used only when priority_deadline_map is non-empty.
_DP_ACQUIRE  = os.environ.get("DP_ACQUIRE", "1") != "0"
_DP_RESERVE  = os.environ.get("DP_RESERVE", "1") != "0"
_DP_MOULDCAP = os.environ.get("DP_MOULDCAP", "1") != "0"
# Acquisition pacing margin: a committed SKU targets (JIT-needed presses + margin), bounded by
# its mould-pair cap. margin=0 = pure just-in-time (fewest presses that finish by the deadline —
# most KPI-efficient, least slack); a large margin = fill to the mould-pair cap (max delivery
# insurance, most collateral). Default 99 = "fill to the mould cap" (client priority: deliver
# on date). Swept per month; env DP_PACE_MARGIN overrides.
_DP_PACE_MARGIN = int(os.environ.get("DP_PACE_MARGIN", "99"))

# #3 Press-swap hysteresis (SURPLUS_HYST, default OFF = current surplus-release bit-for-bit).
# P1 (month-end CO-cost awareness): don't RELEASE surplus presses in the last SURPLUS_P1_DAYS
# days — a swap loses a full 480-min shift + a mould reset and can't be amortized that late
# (an over-provisioned SKU's presses free themselves via demand_done at month-end anyway).
# P2 (release dead-band): size presses_needed with a small safety margin so each SKU keeps ~1
# buffer press instead of being stripped to the razor-thin minimum — prevents over-release and
# the CO→re-acquire ping-pong (each round-trip = 2 lost shifts + 2 mould resets).
_SURPLUS_HYST        = os.environ.get("SURPLUS_HYST", "0") != "0"
_SURPLUS_P1_MIN_DAYS = int(os.environ.get("SURPLUS_P1_DAYS", "2"))
_SURPLUS_P2_MARGIN   = float(os.environ.get("SURPLUS_P2_MARGIN", "0.15"))

# Press-stability guard (PRESS_STABLE, ADOPTED — default ON; PRESS_STABLE=0 reverts bit-for-bit).
# Fixes the day-to-day curing press-count churn (e.g. 1225170015010LSTL0 swinging 2→10→2→6→2). Root
# cause: the memoryless daily "surplus" arithmetic (presses_needed = ceil(rem/(rate·horizon_left)))
# wobbles as horizon and rem shrink, so a press is RELEASED off a live-demand SKU one day and
# RE-ACQUIRED the next — each round-trip = 2 curing COs = 2×480 min + 2 mould resets, and the
# mid-life dip starves the SKU. ON = "hold": presses are NOT voluntarily released for surplus; they
# free only via demand_done_free (SKU demand met) + RO, and the pairing loop's n-1 protection already
# prevents stripping the last covering press. Ramp-up is unchanged (a fresh 0/low-press or starving
# target is hungriest → still acquires presses as they free). Delivers points 1 (memory-aware, no
# mid-life dip), 2 (release only when demand done) and 3 (smooth demand-done drain) of the churn fix.
# Measured 3-month vs baseline: curing COs 846→720 (−126), starvation 8,814→7,898 (−916), press-count
# churn −57%; cost cured 1,984,410→1,982,292 (−2,118, the price of a stable plan on press-tight months).
# Feasibility clean (demand-cap R8, mould R17, CO-cap R10 all PASS). The near-done release
# (PRESS_RELEASE_DAYS) and the starvation valve (PRESS_VALVE) were both measured WORSE and left OFF.
_PRESS_STABLE = os.environ.get("PRESS_STABLE", "1") != "0"
# Rule 2(b) near-done early-release window (days). Under PRESS_STABLE a held press MAY be released
# early once its SKU is winding down (remaining demand clearable by n-1 presses within this window).
# Default 0 = pure hold: sweeping 3 REGRESSED all 3 months (reintroduced churn + starvation, July
# −7,434) because near-done releases cascade into fresh disruption. Kept as a tunable for the record.
_PRESS_RELEASE_DAYS = int(os.environ.get("PRESS_RELEASE_DAYS", "0"))
# Monotone press-count ratchet (PRESS_RATCHET, default OFF — MEASURED REDUNDANT, kept for record).
# Idea: once an SKU sheds a press, cap it at that reduced count so it can't re-acquire (forbid the
# up-leg of the ping-pong). Implemented in the SOFT form the user asked for (yield the cap to a
# genuine re-ramp, not a marginal top-up) it is a NO-OP: the only targets it would block are already
# on-track, which the pairing loop rejects upstream → byte-identical to baseline on all 3 months.
# The STRICT form (block Class-A re-acquire too) does bind but over-constrains legitimate re-ramp.
# Either way PRESS_STABLE subsumes it — hold prevents the mid-life SHED, so there is no down-leg to
# ratchet (stable+ratchet == stable, byte-identical). Tracks _pc_ceiling[sku] = count at last decrease.
_PRESS_RATCHET = os.environ.get("PRESS_RATCHET", "0") != "0"
# Narrow starvation valve (PRESS_VALVE, default OFF — MEASURED WORSE THAN PURE HOLD, kept for record).
# Idea: pure hold never redistributes, so scarce-inch SKUs starve on tight months (Aug −5,700); the
# valve releases a held press ONLY to feed a SEVERELY-STARVED, SERVABLE target (≥2 free moulds) from a
# donor over-supplied WITH MARGIN, capped by #starved targets + CO headroom, ≤1 press/donor-SKU/day.
# Measured 3-month vs pure hold: June +524 / July −3,530 / Aug +1,568 = NET −1,438, and July churn
# 72→114. It helps loose months but TANKS the tight one — on July there is no genuine surplus (donors
# are themselves building-limited on the same scarce 15"/13" inches, so shedding them re-creates the
# ping-pong). Same structural finding as the REJECTED global mould optimiser (−38k): July's gap is true
# mould/press scarcity, not misallocation. Pure hold (PRESS_STABLE, valve OFF) is the adopted mechanism.
_PRESS_VALVE        = os.environ.get("PRESS_VALVE", "0") != "0"
_PRESS_VALVE_MARGIN = int(os.environ.get("PRESS_VALVE_MARGIN", "2"))  # donor surplus above n-1 need
# MARGIN=2 (not 1): a shed donor keeps ONE press of genuine slack, so a building-limited donor
# (scarce inch, produces below rate) won't immediately flip Class-A and re-acquire (the ping-pong).

# Curing CO same-inch alignment (env CURING_INCH_ALIGN, default OFF). When on (and a sku_inch
# map is supplied), a press changing over PREFERS a target SKU of the SAME inch as its current
# SKU — a tiebreak placed AFTER urgency_class + constraint, so demand-critical (Class-A) and
# sole-supplier needs are never sacrificed. Keeps each press on one inch across COs, so the
# building side can feed it without different-size changeovers (esp. BJ/US single-inch machines).
_CURING_INCH_ALIGN = os.environ.get("CURING_INCH_ALIGN", "1") != "0"   # ADOPTED: default ON (match b2c_pipeline)

# Size-balanced Phase-0 allocation (env SIZE_BAL, default OFF). Promotes the same-inch supply
# signal from a soft tiebreak (CO_SUPPLY_MATCH) to a HARD per-inch cap: a curing press only CO's
# to an inch that building can still SUPPLY (building_inch_capacity[inch] − live per-inch draw > 0).
# Over-cap targets are removed from the candidate set BEFORE urgency ranking (the cap dominates
# urgency — firing a press onto an inch building can't feed produces a RUNNING-but-starved press
# and steals building from inches it can feed; that "coverage" is illusory — the plant deliberately
# under-draws building-limited 15"). Day-0 presses inherited on an over-supplied inch are migrated
# off it toward under-supplied inches building can feed. Independent of CO_SUPPLY_MATCH (that env
# still governs only the legacy soft tiebreak). OFF → today's plan bit-for-bit.
_SIZE_BALANCED_ALLOC = os.environ.get("SIZE_BAL", "1") != "0"   # ADOPTED (hard filter): default ON

# Sub-lever of SIZE_BAL: proactively MIGRATE presses inherited on an over-cap inch toward
# under-supplied inches (vs only BLOCKING new over-cap fires). MEASURED WORSE on all 3 months
# (May -6.5k, July -8.1k, starvation up) — each migration is a curing CO (a lost press-shift +
# mould reset), and churning the inherited Day-0 state costs more than the rebalance gains.
# Default OFF (rejected experiment, kept for the record). SIZE_BAL alone = hard filter only.
_SIZE_BAL_MIGRATE = os.environ.get("SIZE_BAL_MIGRATE", "0") != "0"

# Same-inch-FIRST priority (env SAME_INCH_FIRST). Today a freed press's CO target is ranked by
# urgency FIRST and same-inch only 4th, so it still pulls presses across inches (forcing building
# diff-size COs). ON promotes same-inch ABOVE urgency so a press prefers a target on its OWN inch
# (building feeds it with no diff-size CO) — urgent cross-inch SKUs are served by presses already
# on that inch (plant behaviour). SAME_INCH_RANK: "safe" = Class-A (can't-meet-demand) still fires
# first, same-inch 2nd (protects critical cross-inch demand); "top" = same-inch beats everything.
# SIZE_BAL hard pre-filter stays on top regardless. OFF → current order (same-inch 4th), bit-for-bit.
_SAME_INCH_FIRST = os.environ.get("SAME_INCH_FIRST", "0") != "0"
_SAME_INCH_RANK  = os.environ.get("SAME_INCH_RANK", "safe")   # "safe" | "top"

# ── Press-swap DWELL / anti-boomerang (PRESS_DWELL, default OFF = bit-for-bit) ──────────
# RCA (July): 68% of curing COs are BOOMERANG (a SKU loses AND regains presses) and 64% are
# cross-inch — the Phase-0 pairing loop is memoryless per-day, so a press CO's away from an SKU
# one day and (a different) press CO's back the next. That thrash (a) manufactures ~half the
# starvation (a cross-inch press ARRIVES on an SKU whose GT building hasn't pre-fed → runs dry),
# (b) wastes 480 min + a mould reset per CO, and (c) produces the [2,8,4,0,0,6,2] press-count
# jaggedness the plant rejects (wants gradual [2,4,6,4,2,1]). Three coupled rate-limits smooth it:
#   • GAIN cap  — an SKU may ACQUIRE ≤ PRESS_MAX_GAIN_PER_DAY presses/day (kills the +6 spikes).
#   • SHED cap  — an SKU may LOSE   ≤ PRESS_MAX_SHED_PER_DAY presses/day (gradual down-ramp, no 3→0).
#   • COOLDOWN  — after an SKU loses a press, no press may CO back onto it for PRESS_BOOMERANG_COOLDOWN
#                 days (kills the re-acquire leg of the boomerang; the SKU it was starving on is
#                 building-limited anyway, so re-adding a press just starves it → coverage-neutral).
# All three yield to genuine INFEASIBLE need is NOT applied — the caps are hard (stability first,
# per plant-expert guidance: killing churn frees the wasted CO/starved-shift capacity → KPI up).
_PRESS_DWELL = os.environ.get("PRESS_DWELL", "0") != "0"
_PRESS_MAX_GAIN_PER_DAY = int(os.environ.get("PRESS_MAX_GAIN_PER_DAY", "2"))
_PRESS_MAX_SHED_PER_DAY = int(os.environ.get("PRESS_MAX_SHED_PER_DAY", "2"))
_PRESS_BOOMERANG_COOLDOWN = int(os.environ.get("PRESS_BOOMERANG_COOLDOWN", "3"))

# ── Month-end tail damper (TAIL_NO_COLD, default OFF = bit-for-bit) ─────────────────────
# RCA (July): unique curing SKUs/day rises 47→68 across the month while building SKUs/day FALLS
# 44→20; starvation tracks this (corr +0.69) and spikes to 20-33% on days 22-31. Mechanism: as
# SKUs finish demand, their freed presses CO onto ever-MORE different SKUs, but building has wound
# down and can only feed ~20 of them → the rest run dry. Fix: in the last TAIL_NO_COLD_DAYS working
# days, DON'T let a freed press start a COLD sku (press_count==0) — building won't ramp a fresh SKU
# that late, so the press would only starve. Keeps the press on its current SKU or idle instead of
# spreading curing thinner than building can feed. Bounded to the tail so early campaigns are intact.
_TAIL_NO_COLD = os.environ.get("TAIL_NO_COLD", "0") != "0"
_TAIL_NO_COLD_DAYS = int(os.environ.get("TAIL_NO_COLD_DAYS", "6"))

# ── Anti-boomerang levers (the ONLY true monotonic-rule violations, ~7-13/month) ────────
# RCA (July): the visible press-count jaggedness is ~93% GT-STARVATION representation (a press
# stays COMMITTED but produces 0 when building under-feeds), NOT CO churn. The genuine monotonic
# violation is small: a committed-press count that sheds then RE-GAINS. Two targeted, KPI-neutral
# levers remove it; both default OFF = bit-for-bit. Both act TARGET-side only (never touch the n-1
# donor guard, CO-cap, or mould gate) and EXEMPT delivery-priority + fully-abandoned (0-press) SKUs
# so no covering path is ever removed.
# L2 — PRESS_RETURN_BLOCK: a specific press never CO's back to a SKU it already left (same-press
#   round-trip). A DIFFERENT eligible press serves that SKU instead → coverage unchanged.
_PRESS_RETURN_BLOCK = os.environ.get("PRESS_RETURN_BLOCK", "1") != "0"
# L2_STRICT: also block a return to a FULLY-ABANDONED target (0 presses). Off by default because
# such a return is a demand-driven RESTART — blocking it can strand the SKU (coverage loss). A/B only.
_L2_STRICT = os.environ.get("L2_STRICT", "0") != "0"
# L1 — TAIL_DAMP: in the last TD_TAIL_DAYS working days, block a MARGINAL warm re-acquire (target
#   already has >=1 press AND its residual demand < TD_MIN_RESID_DAYS press-days) — the month-end
#   cascade top-up that just re-adds a press onto a nearly-done SKU (starves anyway on the tight
#   inches). Cold starts (0-press) are handled by TAIL_NO_COLD, not here.
_TAIL_DAMP = os.environ.get("TAIL_DAMP", "0") != "0"
_TD_TAIL_DAYS = int(os.environ.get("TD_TAIL_DAYS", "3"))
_TD_MIN_RESID_DAYS = float(os.environ.get("TD_MIN_RESID_DAYS", "1.0"))

# ── Supply-aware curing draw (SUPPLY_ALIGN, default OFF = bit-for-bit) ─────────────────
# RCA: idle building (Stage-2/VMI) + unmet demand coexist because building only builds what
# curing DRAWS (no-waste-GT). A freed press that CO's onto an SKU whose building is ALREADY
# fully drawn just starves; a press CO'd onto an SKU with idle building capacity (e.g. HTORE:
# 6001 idle, free moulds, demand, 0 presses) ACTIVATES that idle building → real coverage.
# This ranks CO targets by per-SKU building HEADROOM (buildable_rate − current draw) right after
# urgency: presses are pulled toward SKUs idle building can supply, breaking the NRI-bootstrap
# deadlock and cutting starvation coverage-POSITIVELY (unlike inch-freedom, which regressed 7×).
_SUPPLY_ALIGN = os.environ.get("SUPPLY_ALIGN", "0") != "0"
# how many presses' worth building supply must fall SHORT of a SKU's draw before we donate its
# marginal (starving) press to a supply-rich SKU. 1.0 = donate only when ≥1 full press is unfed.
_SUPPLY_OVERFEED_MARGIN = float(os.environ.get("SUPPLY_OVERFEED_MARGIN", "1.0"))

# ── STATEFUL / horizon press planner (#10, env STATEFUL_PLAN, default OFF = memoryless bit-for-bit) ──
# The memoryless pairing loop re-decides press moves every day → boomerang (68%), 278 COs, churn that
# propagates to building. v1 replaces the daily `presses_needed = ceil(rem/(rate·horizon_left))` wobble
# with an UPFRONT per-SKU press TARGET (demand-sized, capped by the SKU's eligible presses/moulds) and
# a MONOTONE rule: a SKU ramps UP toward its target, then once it sheds ANY press it may never re-gain
# (unimodal ramp up→peak→down). No boomerang and no CO-to-undo by construction. STATEFUL_PLAN=0 → the
# current memoryless behaviour bit-for-bit.
_STATEFUL_PLAN = os.environ.get("STATEFUL_PLAN", "0") != "0"
_SP_FILL = float(os.environ.get("SP_FILL", "0.85"))   # steady-fill factor sizing the peak vs a flat month
_SP_MONOTONE = os.environ.get("SP_MONOTONE", "1") != "0"   # unimodal ramp (no re-gain after a shed)
_SP_CAP = os.environ.get("SP_CAP", "1") != "0"             # enforce the per-SKU target peak cap
_SP_V2 = os.environ.get("SP_V2", "0") != "0"               # v2 supply-cap+per-inch norm: MEASURED WORSE (off)
_SP_WARM_FIRST = os.environ.get("SP_WARM_FIRST", "1") != "0"  # v3: fill warm SKUs before opening cold ones

# ── v4 CAMPAIGN / active-set planner (env CAMPAIGN_PLAN, default OFF) ──────────────────
# Phase 1 = build the abstract per-SKU-per-day campaign target (Stage 1) + LOG diagnostics only
# (co_events unchanged). Packs each SKU as ONE contiguous campaign into a per-inch × per-day grid
# bounded by building supply (concentrate, not spread). Wiring to co_events is Phase 2.
_CAMPAIGN_PLAN = os.environ.get("CAMPAIGN_PLAN", "0") != "0"
_CAMPAIGN_DEBUG = os.environ.get("CAMPAIGN_DEBUG", "0") != "0"
# EFF = fraction of theoretical building_inch_capacity that is really achievable (COs/idle) — tightens
# the per-inch press slots to what building ACTUALLY feeds. WINDOW = max campaign length (days): smaller
# → bigger P per SKU → fewer concurrent SKUs (concentrate) → later SKUs stagger into freed slots.
_CAMPAIGN_EFF = float(os.environ.get("CAMPAIGN_EFF", "0.8"))
_CAMPAIGN_WINDOW = int(os.environ.get("CAMPAIGN_WINDOW", "10"))
_CAMPAIGN_FRONTLOAD = os.environ.get("CAMPAIGN_FRONTLOAD", "0") != "0"  # Step 1: grab spare slots early

# Per-SKU shared-capacity feed filter (env PERSKU_FEED, default OFF). The 5b guard blocks
# CO'ing another press onto a target only when (n+1)·draw > buildable_rate[target]. Today
# buildable_rate is the FULL un-apportioned sum of the target's eligible in-inch machines'
# GT/day — but those machines are SHARED and build one SKU at a time, so within an
# oversubscribed inch a press can still be parked on a specific SKU whose machines are busy
# building OTHER same-inch SKUs. When ON, the guard replaces the static sum with a DRAW-WEIGHTED
# per-machine share: each of target's machines is split across the same-inch SKUs currently
# drawing it, weighted by draw, so target's feed = its realistic slice. FLOORED at one full
# machine's GT/day (building can always dedicate a machine to a target). Needs the co-plan
# feed_ctx (sku_machines + machine_gtday, lock-aware); inert without it. OFF → the static
# buildable_rate, bit-for-bit.
# REJECTED (measured, default OFF, code retained for the record). Draw-weighted sharing assumes
# building splits a machine PROPORTIONALLY to current draw among same-inch SKUs — but building
# serves DEFICIT-FIRST and can dedicate a machine to a starving SKU (giving it MORE than its draw
# share), so the estimate is too pessimistic and OVER-BLOCKS valid COs. The one-machine floor kept
# it from the 44x catastrophe but not from a net loss: June -9,992 / July -9,433 / Aug -1,680
# (= -21,105) with starvation UP all 3 months (2-seed deterministic, OFF-parity 664,345 bit-for-bit).
# Confirms the July gap is building/mould CAPACITY, not CO mis-prioritization (only ~4.5k July was
# ever curing-side recoverable). Env PERSKU_FEED=1 to re-enable for experiments.
_PERSKU_FEED = os.environ.get("PERSKU_FEED", "0") != "0"
# PERSKU_FEED_V2 — the BETTER per-SKU feasibility model. The v1 above split a shared machine
# PROPORTIONAL to draw (too pessimistic → over-blocked, −21k). v2 uses DEFICIT-FIRST allocation:
# contending SKUs (most-constrained = fewest eligible machines first) claim their required draw,
# preferring machines OUTSIDE the target's set — so a flexible SKU is served elsewhere and only a
# CAPTIVE SKU consumes the target's machines. The target's feasible draw is then the residual on
# its machines. Blocks only GENUINELY infeasible COs. Default OFF; PERSKU_FEED_V2=1 enables.
# MEASURED (ADOPT): fixes the v1 −21k bug — Jul +7,747 alone (679,904), and as part of the
# adopted 1+2 config lifts Jun/Aug most. Validated the thesis that a CORRECT building-aware
# gate improves KPIs (the v1 regression was the proportional model, not the idea).
_PERSKU_FEED_V2 = True
_NAVY  = "1F3864"
_WHITE = "FFFFFF"
_BLUE  = "D6E4F0"
_YELL  = "FFF2CC"
_GREEN = "E2EFDA"
_ORNG  = "FCE4D6"


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _qty_per_press_per_shift(ct_min: float) -> int:
    return math.floor(ConsumptionConfig.SHIFT_MINS / ct_min) \
           * ConsumptionConfig.CAVITIES_PER_MOULD


def _qty_per_press_per_day(ct_min: float) -> float:
    return _qty_per_press_per_shift(ct_min) * SHIFTS_PER_DAY


def _production_days(remaining_demand: float, press_count: int, rate_per_day: float) -> Optional[float]:
    """Days needed to fulfill remaining demand at current press rate. None if no press."""
    if press_count <= 0 or rate_per_day <= 0:
        return None
    return remaining_demand / (press_count * rate_per_day)


def _urgency_sort_key(
    priority_score: float,
    current_press_count: int,
    updated_demand: float,
    rate_per_day: float,
    horizon_left: int,
    deadline_days: float | None = None,
) -> tuple:
    """
    Two-level urgency sort key (sort ascending = highest urgency first).

    Class A (CRITICAL, key=0): current_production_days > horizon_left
        Demand CANNOT be met without this CO.
    Class B (HELPFUL, key=1) : current_production_days <= horizon_left
        Demand can be met with existing presses.

    Within class: highest Priority_Score first, then fewest after-CO days.

    deadline_days (DELIVERY_PRIORITY): a committed-delivery SKU is measured against
    its OWN (nearer) deadline, not the whole horizon — so it flips to Class A exactly
    when its current presses can't finish by its date, and drops back once enough
    presses are on it (self-pacing acquisition). None (every non-priority caller) →
    the horizon test, byte-identical. The returned tuple shape is UNCHANGED so every
    caller (and downstream sort) behaves exactly as before.
    """
    if current_press_count <= 0 or rate_per_day <= 0:
        current_days = float("inf")
    else:
        current_days = updated_demand / (current_press_count * rate_per_day)

    if rate_per_day > 0:
        after_days = updated_demand / ((current_press_count + 1) * rate_per_day)
    else:
        after_days = float("inf")

    _h = horizon_left if deadline_days is None else min(horizon_left, deadline_days)
    cls = 0 if current_days > _h * CO_CLASS_B_THRESHOLD else 1
    return (cls, -priority_score, after_days)


# ══════════════════════════════════════════════════════════════════════════════
# CO SCHEDULER  (Pass 1)
# ══════════════════════════════════════════════════════════════════════════════

class COScheduler:
    """
    Compute the full 31-day changeover schedule from Day 0 data alone.

    Returns a list of CO events:
        [{"day": int, "press": str, "old_sku": str, "new_sku": str}, ...]

    Rules:
    - Runner-Out presses: CO on earliest available day (Day 1+)
    - Runner-In presses : CO on the day Updated_Demand_Qty first reaches 0
    - NRI target        : highest urgency_sort_key among eligible allowable presses
    - Max MAX_CO_PER_DAY COs per day; excess deferred to next day
    - CO timing: CO fires on Day D; new SKU production from Day D (Shift C = same day)
      For daily-level modelling, press count update takes effect Day D+1 onwards
    """

    def schedule(
        self,
        df_day0: pd.DataFrame,
        df_demand: pd.DataFrame,
        df_allowable: pd.DataFrame,
        df_running_moulds: pd.DataFrame,
        ct_map: dict[str, float],
        max_co_per_day: int = MAX_CO_PER_DAY,
        planning_days: int = PLANNING_DAYS,
        ratio_demand_map: dict | None = None,
        buildable_rate: dict | None = None,
        sku_inch: dict | None = None,
        building_inch_capacity: dict | None = None,
        feed_ctx: dict | None = None,            # PERSKU_FEED: {sku_machines, machine_skus, machine_gtday}
        sku_moulds: dict | None = None,          # #4: {sku: set(eligible mould IDs)} or None
        priority_deadline_map: dict | None = None,  # DELIVERY_PRIORITY: {sku: deadline_day} or None
    ) -> list[dict]:
        """Returns sorted list of CO events.

        planning_days: horizon length for THIS call (defaults to the module
        constant, preserving legacy behavior exactly). Rolling-horizon callers
        pass a shrinking remaining-horizon length instead of the fixed project
        total.

        ratio_demand_map: when provided, used instead of demand_map (built
        from df_demand["Quantity"]) for the two ratio computations only
        (press_total_demand, _priority_signal's numerator) — lets a caller
        keep ratio ranking anchored to static/original total demand even when
        df_demand["Quantity"] itself carries live/remaining demand. Defaults
        to None, preserving legacy behavior exactly (both read demand_map).
        """

        # ── press → current SKU map ───────────────────────────────────────────
        press_to_sku: dict[str, str] = {}
        for _, r in df_running_moulds.iterrows():
            press_to_sku[str(r["Machine"])] = str(r["SKUCode"])

        # ── DELIVERY_PRIORITY: committed-delivery deadlines (Phase-0 EDF + reservation) ──
        # Empty map → _prio_on False → every insertion below is identity (bit-for-bit).
        _pdm: dict = {str(k): int(v) for k, v in (priority_deadline_map or {}).items()}
        _prio_on  = bool(_pdm)
        _prio_acq = _prio_on and _DP_ACQUIRE     # force acquisition (EDF + deadline Class-A)?
        _prio_res = _prio_on and _DP_RESERVE     # reserve committed presses pre-deadline?
        # Mould-pair ceiling: a committed SKU cannot cure on more concurrent presses than it
        # has mould pairs (2 moulds/press). Forcing acquisition beyond that only churns COs
        # and steals presses from other SKUs (those extra presses can never get moulds). None
        # → no cap (older behaviour). Computed from the eligibility passed in for the P0 gate.
        _prio_cap: dict = {}
        if _prio_on and _DP_MOULDCAP and sku_moulds:
            for _s in _pdm:
                _prio_cap[_s] = len(sku_moulds.get(_s, ()) or ()) // 2
        def _dd_days(sku: str, day: int):
            dd = _pdm.get(sku)
            return None if dd is None else max(1, dd - day + 1)   # shifts→days to this SKU's deadline
        def _prio_target_presses(sku: str, day: int) -> int:
            """How many concurrent presses this committed SKU should hold TODAY: the fewest
            that still finish its remaining demand by its deadline at the current rate
            (JIT pacing — don't front-load capacity earlier than the deadline requires),
            bounded by its mould-pair ceiling (extra presses can never get moulds)."""
            _rate = _qty_per_press_per_day(ct_map.get(sku, _dct))
            _rem  = updated_demand.get(sku, 0.0)
            _ddl  = max(1, _pdm[sku] - day + 1)
            _need = 1
            if _rate > 0 and _rem > 0:
                _need = max(1, math.ceil(_rem / (_rate * _ddl)))
            _need += _DP_PACE_MARGIN           # slack against build-lag / mould-contention / CO time
            _cap = _prio_cap.get(sku)          # mould-pair ceiling (None/0 = unknown → no cap)
            return min(_cap, _need) if _cap else _need
        def _prio_wants(sku: str, day: int) -> bool:
            """A committed SKU still needs another press today? (below its JIT pace target)."""
            if sku not in _pdm:
                return False
            return press_count.get(sku, 0) < _prio_target_presses(sku, day)

        # ── #4 Mould-aware Phase-0: day-granularity free-mould tracker (mirrors the
        # rolling gate's Day-0 seed). A CO target is mountable only if the press can
        # secure 2 eligible FREE moulds; else Phase-0 picks a different feasible target
        # instead of one the rolling gate would later redirect. OFF → no filtering. ─────
        _p0_gate = _P0_MOULD_GATE and bool(sku_moulds)
        _p0_sm: dict[str, set] = {}
        _p0_owner: dict[str, str] = {}
        _p0_pm: dict[str, set] = {}
        if _p0_gate:
            _p0_sm = {str(k): set(v) for k, v in sku_moulds.items()}
            for _, r in df_running_moulds.iterrows():          # seed Day-0 mounted moulds
                _p = str(r["Machine"]); _sku0 = str(r["SKUCode"])
                for _m in (r.get("MouldNos", []) or []):
                    _m = str(_m).strip()
                    if not _m or _m.lower() == "nan":
                        continue
                    _p0_owner[_m] = _p
                    _p0_pm.setdefault(_p, set()).add(_m)
                    _p0_sm.setdefault(_sku0, set()).add(_m)     # orphan-fold (seed never self-violates)
            for _, r in df_running_moulds.iterrows():           # Day-0 second-mould top-up
                _p = str(r["Machine"]); _sku0 = str(r["SKUCode"])
                if len(_p0_pm.get(_p, set())) >= 2:
                    continue
                for _m in sorted(_x for _x in _p0_sm.get(_sku0, set()) if _p0_owner.get(_x) is None):
                    if len(_p0_pm.get(_p, set())) >= 2:
                        break
                    _p0_owner[_m] = _p
                    _p0_pm.setdefault(_p, set()).add(_m)

        def _n_free_for(target: str, press: str) -> int:
            """# moulds eligible for `target` that are FREE or already on `press` (>=2 ⇒ mountable)."""
            _e = _p0_sm.get(str(target), set())
            if len(_e) < 2:
                return 0
            _ps = str(press)
            return sum(1 for _m in _e if _p0_owner.get(_m) in (None, _ps))

        def _p0_mount(press: str, sku: str) -> None:
            """Claim 2 eligible moulds for `sku` on `press`; release the press's now-ineligible ones."""
            press = str(press); _e = _p0_sm.get(str(sku), set())
            _have = _p0_pm.get(press, set())
            _keep = {_x for _x in _have if _x in _e}
            for _x in (_have - _keep):
                if _p0_owner.get(_x) == press:
                    _p0_owner[_x] = None
            _p0_pm[press] = set(_keep)
            for _x in sorted(_y for _y in _e if _p0_owner.get(_y) is None and _y not in _keep):
                if len(_p0_pm[press]) >= 2:
                    break
                _p0_owner[_x] = press
                _p0_pm[press].add(_x)

        # Same-inch alignment (Part 1): 0 if the CO target's inch matches the press's CURRENT
        # SKU inch, else 1 — used as a low-priority tiebreak so a press keeps its inch across COs.
        _sku_inch = sku_inch or {}
        _align = _CURING_INCH_ALIGN and bool(_sku_inch)
        def _same_inch(press: str, target: str) -> int:
            if not _align:
                return 0                                   # OFF → constant → no ordering change
            cur = _sku_inch.get(str(press_to_sku.get(press, "")), "")
            return 0 if (cur and _sku_inch.get(str(target), "") == cur) else 1

        # Supply-matched draw (Part C): pull the press DRAW toward inches the building side can
        # SUPPLY (building_inch_capacity), so locked BJ/US machines stay fed and never idle. Live
        # per-inch consumption is derived from press_count (updated as COs fire). A CO target whose
        # inch still has building headroom sorts before one whose inch is already over-drawn.
        _bic = building_inch_capacity or {}
        # Supply-migration term measured a slight net NEGATIVE (it over-diverts the draw); default
        # OFF. The KPI recovery comes from the lock-aware buildable_rate + the demand-optimal lock.
        _supply_on = bool(_bic) and bool(_sku_inch) and os.environ.get("CO_SUPPLY_MATCH", "0") != "0"
        def _inch_consumption() -> dict:
            _c: dict = {}
            for _s, _n in press_count.items():
                if _n > 0:
                    _i = _sku_inch.get(str(_s), "")
                    if _i:
                        _c[_i] = _c.get(_i, 0.0) + _n * _qty_per_press_per_day(
                            ct_map.get(str(_s), ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
            return _c
        def _supply_pref(target: str, cons: dict) -> int:
            if not _supply_on:
                return 0                                   # OFF → constant → no ordering change
            _i = _sku_inch.get(str(target), "")
            return 0 if (_bic.get(_i, 0.0) - cons.get(_i, 0.0)) > 0 else 1   # 0 = building has headroom

        # Size-balanced HARD cap (SIZE_BAL). _over_cap(target, cons) is True when the target's inch
        # has NO building headroom left (live per-inch draw ≥ building_inch_capacity[inch]). Computed
        # unconditionally; callers apply it only when _size_bal is on. Same predicate as _supply_pref,
        # but a hard filter rather than a tiebreak — and gated by SIZE_BAL, not CO_SUPPLY_MATCH.
        _size_bal = _SIZE_BALANCED_ALLOC and bool(_bic) and bool(_sku_inch)
        _size_bal_migrate = _size_bal and _SIZE_BAL_MIGRATE   # rejected sub-lever, default OFF
        def _over_cap(target: str, cons: dict) -> bool:
            _i = _sku_inch.get(str(target), "")
            return (_bic.get(_i, 0.0) - cons.get(_i, 0.0)) <= 0

        # PERSKU_FEED: refine the 5b guard's per-SKU buildable rate. The static buildable_rate is
        # the FULL sum of a SKU's in-inch machines' GT/day — but those machines are shared. This
        # returns target's DRAW-WEIGHTED slice: each machine split across the same-inch SKUs
        # currently drawing it, weighted by draw. Floored at one full machine's GT/day (building
        # can dedicate a machine to a target) so it can only ever TIGHTEN below the static sum for
        # a genuinely contended SKU, never over-block. OFF / no ctx → the static rate (bit-for-bit).
        _perSKU_feed_on = (_PERSKU_FEED or _PERSKU_FEED_V2) and bool(feed_ctx) and bool(_sku_inch)
        _feed_sm = (feed_ctx or {}).get("sku_machines", {})
        _feed_ms = (feed_ctx or {}).get("machine_skus", {})
        _feed_gt = (feed_ctx or {}).get("machine_gtday", {})

        def _perSKU_feed(target, n_t, rate_t):
            _static = buildable_rate.get(target) if buildable_rate is not None else None
            if not _perSKU_feed_on:
                return _static
            _tgt = str(target)
            _ms = _feed_sm.get(_tgt)
            if not _ms:
                return _static
            _newdraw = (n_t + 1) * rate_t                    # target's draw WITH the added press
            _dctm = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN

            if _PERSKU_FEED_V2:
                # DEFICIT-FIRST feasibility. Contending SKUs (most-constrained = fewest eligible
                # machines first) claim their REQUIRED draw, preferring machines OUTSIDE the
                # target's set (a flexible SKU is served elsewhere; only a captive SKU consumes
                # target's machines). Target's feasible = residual capacity on its machines.
                _tset = set(_ms)
                _others = []                                 # (n_elig, sku, req, machines)
                _seen = set()
                for _m in _ms:
                    for _s in _feed_ms.get(_m, ()):
                        if _s == _tgt or _s in _seen:
                            continue
                        _seen.add(_s)
                        _ns = press_count.get(_s, 0)
                        if _ns <= 0:
                            continue
                        _sm = list(_feed_sm.get(_s) or ())
                        if _sm:
                            _others.append((len(_sm), _s,
                                            _ns * _qty_per_press_per_day(ct_map.get(_s, _dctm)), _sm))
                # residual capacity over EVERY machine any contender (or target) can use
                _residual = {}
                for _m in _ms:
                    _residual[_m] = _feed_gt.get(_m, 0.0)
                for (_nm, _s, _req, _sm) in _others:
                    for _m in _sm:
                        _residual.setdefault(_m, _feed_gt.get(_m, 0.0))
                _others.sort(key=lambda x: (x[0], x[1]))     # most-constrained first, deterministic
                for (_nm, _s, _req, _sm) in _others:
                    _order = sorted(_sm, key=lambda m: (m in _tset, m))   # outside-target machines first
                    _left = _req
                    for _m in _order:
                        if _left <= 0:
                            break
                        _take = min(_residual.get(_m, 0.0), _left)
                        _residual[_m] -= _take
                        _left -= _take
                _feasible = min(_newdraw, sum(_residual.get(_m, 0.0) for _m in _ms))
                _floor = max((_feed_gt.get(_m, 0.0) for _m in _ms), default=0.0)
                return max(_feasible, _floor)

            # v1 (proportional-to-draw — retained for A/B; documented net −21k)
            feasible = 0.0
            for _m in _ms:
                _tot = 0.0                                   # total same-inch draw contending for _m
                for _s in _feed_ms.get(_m, ()):
                    if _s == _tgt:
                        _tot += _newdraw
                    else:
                        _ns = press_count.get(_s, 0)
                        if _ns > 0:
                            _tot += _ns * _qty_per_press_per_day(ct_map.get(_s, _dctm))
                _mg = _feed_gt.get(_m, 0.0)
                feasible += _mg if _tot <= 0 else _mg * (_newdraw / _tot)   # target's slice of _m
            _floor = max((_feed_gt.get(_m, 0.0) for _m in _ms), default=0.0)
            return max(feasible, _floor)

        ro_skus  = set(df_day0.loc[df_day0["Category"] == "Runner-Out",     "SKUCode"])
        ri_skus  = set(df_day0.loc[df_day0["Category"] == "Runner-In",      "SKUCode"])
        nri_skus = set(df_day0.loc[df_day0["Category"] == "Non-Runner-In",  "SKUCode"])

        all_demand_skus = set(df_demand["SKUCode"].str.strip())
        demand_map   = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Quantity"]))
        priority_map = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Priority"]))

        # Per-inch total demand (SIZE_BAL) — sizes the Day-0 over-cap migration only.
        inch_demand: dict[str, float] = {}
        if _size_bal:
            for _s, _q in demand_map.items():
                _i = _sku_inch.get(str(_s), "")
                if _i:
                    inch_demand[_i] = inch_demand.get(_i, 0.0) + float(_q)

        # ── Build press → ALL compatible demand SKUs (NRI + RI) ─────────────
        # Fixes bug where RO presses were only matched against NRI targets.
        sku_to_presses: dict[str, set] = {}
        for _, r in df_allowable.iterrows():
            sku = str(r["SKUCode"]).strip()
            if sku in all_demand_skus:
                machines = r.get("Machines", [])
                if machines:
                    sku_to_presses[sku] = {str(p) for p in machines}

        press_to_demand_targets: dict[str, list] = {}
        for sku, presses in sku_to_presses.items():
            for p in presses:
                press_to_demand_targets.setdefault(p, []).append(sku)

        # Static per-press total demand for _CURING_RATIO_ENABLED — computed once
        # from the fixed demand file, never decremented (mirrors machine_total_demand
        # in b2c_pipeline.py's _priority_tier). ratio_demand_map lets a rolling-horizon
        # caller keep this anchored to static/original total demand even when
        # demand_map itself (from df_demand["Quantity"]) carries live/remaining
        # demand for THIS call — defaults to demand_map, matching legacy behavior.
        _ratio_map = ratio_demand_map if ratio_demand_map is not None else demand_map
        press_total_demand: dict[str, float] = {
            p: sum(_ratio_map.get(s, 0.0) for s in targets)
            for p, targets in press_to_demand_targets.items()
        }

        def _priority_signal(target: str, p: str) -> float:
            # RI keeps existing Priority_Score-driven urgency untouched — its
            # eligibility/urgency logic already reflects real running-press state.
            # Ratio only re-ranks NRI candidates (mirrors _priority_tier on the
            # building side: RI stays governed by its existing signal, NRI gets ratio).
            if _CURING_RATIO_ENABLED and target in nri_skus:
                return _ratio_map.get(target, 0.0) / press_total_demand.get(p, 1e-9)
            return float(priority_map.get(target, 0))

        def _sku_priority_signal(s: str) -> float:
            # Rescue pass has no single press yet — use the most-favorable ratio
            # across the SKU's own compatible press pool (still fully static).
            if _CURING_RATIO_ENABLED:
                presses = sku_to_presses.get(s, set())
                if not presses:
                    return 0.0
                return max(_priority_signal(s, p) for p in presses)
            return float(priority_map.get(s, 0))

        # ── Running state ─────────────────────────────────────────────────────
        press_count: dict[str, int] = {}
        for _, r in df_day0.iterrows():
            sku = str(r["SKUCode"])
            press_count[sku] = int(r.get("Running_Press_Count", 0))

        updated_demand: dict[str, float] = {
            sku: float(demand_map.get(sku, 0)) for sku in all_demand_skus
        }

        # ── Track eligible presses ────────────────────────────────────────────
        # pending_ro_presses: RO presses carried forward every day until they CO.
        # Fixes bug where RO presses were only offered on Day 1; any that didn't
        # fit in the 8/day cap were silently dropped.
        runner_out_presses: set = {p for p, s in press_to_sku.items() if s in ro_skus}
        pending_ro_presses: set = runner_out_presses.copy()

        # demand_running_presses: presses running a demand SKU (RI + CO'd NRI/RO).
        # When their SKU's demand = 0, the press is freed for CO.
        demand_running_presses: set = {p for p, s in press_to_sku.items() if s in ri_skus}

        co_events: list[dict] = []
        daily_co_used: dict[int, int] = {}
        _hol = _holiday_day_index_set()   # plant holidays → working-day urgency horizon
        # PRESS_RATCHET: per-SKU press-count ceiling — set to the count at the last DECREASE; the
        # SKU may never re-acquire above it. inf until the SKU first sheds a press.
        _pc_ceiling: dict[str, float] = {}
        # PRESS_DWELL: last day a press CO'd AWAY from each SKU (anti-boomerang cooldown key).
        _sku_lost_day: dict[str, int] = {}
        # PRESS_RETURN_BLOCK (L2): per-press set of SKUs this press has CO'd AWAY from — it may not
        # return to any of them (no same-press round-trip). Empty at run start (Day-0 carried SKU is
        # not "left", so returning to it after an excursion is not blocked); reset per Run in 2pass.
        _press_left: dict[str, set] = {}

        # ── STATEFUL_PLAN (#10): upfront per-SKU press TARGET + monotone tracker ───
        # target_peak = presses to meet the SKU's whole-month demand at steady rate (÷ _SP_FILL to
        # allow ramp/downtime), capped by the SKU's eligible-press count (physical bound). _peaked =
        # SKUs that have shed a press → may not re-gain (unimodal ramp). Computed once; no per-day wobble.
        _target_peak: dict[str, int] = {}
        _peaked: set[str] = set()
        if _STATEFUL_PLAN:
            _spdays = max(1, planning_days)
            _rate_of: dict[str, float] = {}
            for _s in all_demand_skus:
                _dem = float(demand_map.get(_s, 0) or 0)
                _rt = _qty_per_press_per_day(ct_map.get(_s, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                _rate_of[_s] = _rt
                if _dem <= 0 or _rt <= 0:
                    _target_peak[_s] = 0
                    continue
                _need = math.ceil(_dem / (_rt * _spdays * _SP_FILL))       # presses to meet demand
                _capp = len(sku_to_presses.get(_s, ())) or 10**6          # eligible-press physical cap
                if _SP_V2 and buildable_rate is not None:
                    # v2: never target more presses than BUILDING can feed this SKU (supply cap) —
                    # the press-cap done at plan time, so the ramp never over-provisions into starvation.
                    _sup = _perSKU_feed(_s, 0, _rt)
                    if _sup is not None and _sup > 0:
                        _capp = min(_capp, max(1, math.floor(_sup / _rt)))
                _target_peak[_s] = max(1, min(_need, _capp))
            if _SP_V2 and _bic:
                # v2 per-INCH supply normalisation: scale each inch's SKU targets so the planned
                # per-inch DRAW never exceeds building_inch_capacity (SIZE_BAL at plan time). Removes
                # the flat-`fill` fragility — the plan matches curing draw to building supply per inch.
                _inch_draw: dict[str, float] = {}
                _inch_skus: dict[str, list] = {}
                for _s, _tp in _target_peak.items():
                    if _tp <= 0:
                        continue
                    _i = _sku_inch.get(str(_s), "")
                    if not _i:
                        continue
                    _inch_draw[_i] = _inch_draw.get(_i, 0.0) + _tp * _rate_of.get(_s, 0.0)
                    _inch_skus.setdefault(_i, []).append(_s)
                for _i, _draw in _inch_draw.items():
                    _cap_i = _bic.get(_i, 0.0)
                    if _cap_i > 0 and _draw > _cap_i:
                        _scale = _cap_i / _draw
                        for _s in _inch_skus[_i]:
                            _target_peak[_s] = max(1, math.floor(_target_peak[_s] * _scale))

        # ── v4 CAMPAIGN PLAN (Stage 1): abstract per-SKU-per-day press target ──────
        # Packs each SKU as ONE contiguous campaign into a per-inch × per-day capacity grid bounded
        # by building supply (building_inch_capacity). CONCENTRATE (full-size campaigns, sequential)
        # not SPREAD (v2 scaled peaks down → cureRUN up). Pure/deterministic. Phase 1 = build + LOG.
        _campaign_target: dict[str, dict[int, int]] = {}
        if _CAMPAIGN_PLAN and _bic:
            _rate_c: dict[str, float] = {
                _s: _qty_per_press_per_day(ct_map.get(_s, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                for _s in all_demand_skus}
            _work_days = [d for d in range(1, planning_days + 1) if d not in _hol]
            # per-inch concurrent-press capacity = building GT/day ÷ a typical SKU rate on that inch
            _inch_slots: dict[str, int] = {}
            for _i, _cap in _bic.items():
                _rs = sorted(_rate_c[_s] for _s in all_demand_skus
                             if _sku_inch.get(str(_s), "") == _i
                             and demand_map.get(_s, 0) > 0 and _rate_c[_s] > 0)
                _typ = _rs[len(_rs) // 2] if _rs else 0.0            # median rate on the inch
                # tighten to REAL achievable building throughput (EFF), not theoretical capacity
                _inch_slots[_i] = max(0, math.floor(_cap * _CAMPAIGN_EFF / _typ)) if _typ > 0 else 0
            _used: dict[str, dict[int, int]] = {}                    # inch -> day -> slots used
            def _cap_free(_i, _d, _p):
                return _used.get(_i, {}).get(_d, 0) + _p <= _inch_slots.get(_i, 0)
            # order: Day-0 running-in SKUs first (continuity), then EDF, then biggest demand
            _plan_skus = sorted(
                (_s for _s in all_demand_skus if demand_map.get(_s, 0) > 0 and _rate_c[_s] > 0
                 and _sku_inch.get(str(_s), "") in _inch_slots),
                key=lambda _s: (0 if (_s in ri_skus and press_count.get(_s, 0) > 0) else 1,
                                _pdm.get(_s, planning_days + 1), -float(demand_map[_s]), str(_s)))
            _deferred: list[str] = []
            for _s in _plan_skus:
                _i = _sku_inch.get(str(_s), "")
                _rate = _rate_c[_s]
                _slots_i = _inch_slots.get(_i, 0)
                if _slots_i <= 0:
                    _deferred.append(_s); continue
                _dl = _pdm.get(_s, planning_days)
                _elig = [d for d in _work_days if d <= _dl]           # the SKU's servable window
                if not _elig:
                    _deferred.append(_s); continue
                _press_days = demand_map[_s] / _rate
                _mcap = (len(sku_moulds.get(_s, ())) // 2) if sku_moulds else len(sku_to_presses.get(_s, ()))
                if _CAMPAIGN_FRONTLOAD:
                    # FRONT-LOADED trapezoid: from its earliest day, grab as many presses as the inch
                    # has FREE that day (up to mould-cap), day by day, until demand is cleared → high P
                    # early, ramps down as slots fill/demand depletes. Fills spare-inch slots that flat-P
                    # leaves empty (more cured on building-sufficient inches).
                    _P_peak = max(1, min(_mcap or 10**6, _slots_i))
                    _rem = _press_days
                    _tgt: dict = {}
                    for d in _elig:
                        if _rem <= 0:
                            break
                        _free_amt = _slots_i - _used.get(_i, {}).get(d, 0)
                        _p = min(_P_peak, _free_amt, max(1, math.ceil(_rem)))
                        if _p < 1:
                            continue
                        _tgt[d] = _p
                        _used.setdefault(_i, {})[d] = _used.get(_i, {}).get(d, 0) + _p
                        _rem -= _p
                    if not _tgt:
                        _deferred.append(_s); continue
                    _campaign_target[_s] = _tgt
                    continue
                # P = presses to clear demand within a CAMPAIGN_WINDOW-day campaign → CONCENTRATE on
                # fewer SKUs at higher P (vs steady/whole-month spread). Capped by mould-pairs + inch
                # slots. Shorter window ⇒ bigger P ⇒ later SKUs stagger into freed slots (rotation).
                _wmax = min(len(_elig), max(1, _CAMPAIGN_WINDOW))
                _P = max(1, min(_mcap or 10**6, _slots_i,
                                math.ceil(_press_days / _wmax)))
                _D = min(len(_elig), max(1, math.ceil(_press_days / _P)))   # days needed at P presses
                # earliest contiguous window of _D working days ending <= deadline with free capacity
                _win = None
                for _k in range(0, len(_elig) - _D + 1):
                    _cand = _elig[_k:_k + _D]
                    if all(_cap_free(_i, d, _P) for d in _cand):
                        _win = _cand; break
                if _win is None:                                     # best-effort: any free days (may under-serve)
                    _free = [d for d in _elig if _cap_free(_i, d, _P)]
                    if _free:
                        _win = _free[:_D]
                    else:
                        _deferred.append(_s); continue
                _campaign_target[_s] = {}
                for d in _win:
                    _campaign_target[_s][d] = _P
                    _used.setdefault(_i, {})[d] = _used.get(_i, {}).get(d, 0) + _P
            # diagnostics
            _act_per_day = {d: sum(1 for _s in _campaign_target if d in _campaign_target[_s])
                            for d in range(1, planning_days + 1)}
            _avg_act = sum(_act_per_day.values()) / max(1, len(_act_per_day))
            _over = sum(1 for _i in _used for d in _used[_i] if _used[_i][d] > _inch_slots.get(_i, 0))
            print(f"  [CAMPAIGN] planned {len(_campaign_target)}/{len(_plan_skus)} SKUs, "
                  f"deferred {len(_deferred)} | avg active SKUs/day = {_avg_act:.0f} "
                  f"(target ~building) | per-inch over-capacity days = {_over}")
            if _CAMPAIGN_DEBUG:
                print(f"  [CAMPAIGN] inch_slots={dict(sorted(_inch_slots.items()))}")
                print(f"  [CAMPAIGN] active/day={[_act_per_day[d] for d in range(1, planning_days+1)]}")
                print(f"  [CAMPAIGN] deferred eg={[str(s)[-8:] for s in _deferred[:10]]}")

        # ── Day-by-day simulation ─────────────────────────────────────────────
        for day in range(1, planning_days + 1):
            horizon_left = _working_days_left(day, planning_days, _hol)
            co_used = daily_co_used.get(day, 0)

            # Drain demand by previous day's production. If the previous day was a plant
            # holiday, nothing was produced → no drain (keeps the CO-planner's demand
            # projection consistent with the holiday-idle plant).
            if day > 1 and (day - 1) not in _hol:
                for sku in all_demand_skus:
                    n = press_count.get(sku, 0)
                    if n <= 0:
                        continue
                    rate = _qty_per_press_per_day(
                        ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                    updated_demand[sku] = max(0.0, updated_demand[sku] - n * rate)

            if co_used >= max_co_per_day:
                continue

            # Identify free presses this day:
            # 1. All pending RO presses (carried forward until they CO)
            # 2. Demand-running presses whose SKU demand just hit 0
            #    → tracked in demand_done_free: these presses are producing NOTHING useful,
            #      so any CO (Class A or B) is strictly better than sitting idle.
            newly_free: list[str] = sorted(pending_ro_presses)
            demand_done_free: set[str] = set()
            for p in sorted(demand_running_presses):
                current_sku = press_to_sku.get(p)
                if current_sku and updated_demand.get(current_sku, 0) <= 0:
                    newly_free.append(p)
                    demand_done_free.add(p)

            # ── CAMPAIGN ramp-down: free presses on SKUs now OVER their per-day campaign target
            # (window ending / declining ramp). The freed press flows to an active-campaign SKU or
            # idles. Keeps per-inch draw ≤ building supply as campaigns wind down. Monotone-safe
            # (a shed SKU is _peaked → can't re-gain). Deterministic (sorted).
            campaign_free: set[str] = set()
            if _CAMPAIGN_PLAN:
                _by_sku: dict[str, list] = {}
                for p in sorted(demand_running_presses):
                    if p in demand_done_free:
                        continue
                    s = press_to_sku.get(p)
                    if s:
                        _by_sku.setdefault(s, []).append(p)
                for s, ps in _by_sku.items():
                    _tgt = _campaign_target.get(s, {}).get(day, 0)
                    _excess = len(ps) - _tgt
                    for p in sorted(ps)[:max(0, _excess)]:           # shed the over-target presses
                        if p not in newly_free:
                            newly_free.append(p)
                            campaign_free.add(p)

            # ── SUPPLY-DONATION (SUPPLY_ALIGN) ────────────────────────────────
            # RCA: 51 SKUs/day have demand + free moulds + IDLE building that could supply them,
            # but get 0 curing presses — no free press is curing-allowable for them, and under the
            # press-HOLD no press leaves a demand SKU. Meanwhile OVER-fed SKUs run more presses than
            # building can supply (the marginal press just STARVES). Fix: free that starving marginal
            # press so the pairing loop can route it (headroom-first) to a supply-rich SKU, activating
            # idle building. Coverage-POSITIVE: the donor loses a press building couldn't feed anyway.
            supply_free: set[str] = set()
            if _SUPPLY_ALIGN and horizon_left > 0:
                _presses_on: dict[str, list] = {}
                for _p in demand_running_presses:
                    _s = press_to_sku.get(_p)
                    if _s and _p not in demand_done_free:
                        _presses_on.setdefault(_s, []).append(_p)
                # supply-rich UNDER-SERVED targets: 0-press SKUs with demand + idle-building capacity.
                _rich: set[str] = set()
                for _t in all_demand_skus:
                    if press_count.get(_t, 0) == 0 and updated_demand.get(_t, 0) > 50:
                        _rt = _qty_per_press_per_day(ct_map.get(_t, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                        _bt = _perSKU_feed(_t, 0, _rt) if buildable_rate is not None else None
                        if _bt is not None and _bt >= _rt and _n_free_for(_t, "") >= 0:
                            _rich.add(_t)
                for _s, _ps in _presses_on.items():
                    n_s = len(_ps)
                    if n_s < 2 or updated_demand.get(_s, 0) <= 0:
                        continue                       # keep ≥1 covering press
                    rate_s = _qty_per_press_per_day(ct_map.get(_s, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                    if rate_s <= 0:
                        continue
                    # DEMAND-SURPLUS donor: _s runs more presses than it needs to meet its own demand
                    # by the horizon → a press is spare. (Phase-0 has demand, not actual building.)
                    # SA_RELAX=1 drops the surplus gate (n≥2 + reachable-rich) to test reachability.
                    if os.environ.get("SA_RELAX") != "1":
                        _need_s = math.ceil(updated_demand[_s] / (rate_s * max(1e-9, horizon_left)))
                        if n_s <= _need_s:
                            continue
                    # donate the press of _s that CAN reach a supply-rich under-served SKU with 2 free
                    # moulds (else the move is pointless churn). Pick the specific rich-allowable press.
                    for _don in sorted(_ps):
                        if _don in newly_free:
                            continue
                        if any((_t in _rich and (not _p0_gate or _n_free_for(_t, _don) >= 2))
                               for _t in press_to_demand_targets.get(_don, [])):
                            newly_free.append(_don)
                            supply_free.add(_don)
                            break

            # ── Surplus RI-press early-release (spread across days) ───────────
            # An RI SKU running more presses than it needs to meet demand by the
            # horizon is over-provisioned; release the surplus EARLY (a few/day)
            # so freed presses reassign to needy SKUs with many production days
            # left — instead of dumping all at once when demand hits 0 (the cliff).
            # These presses bypass the Class-B gate (like demand_done_free) and are
            # subject to the 5b building-supply guard in the firing loop.
            surplus_free: set[str] = set()
            # #3 P1: month-end guard — a surplus swap can't be amortized in the last few days.
            _surplus_ok = not (_SURPLUS_HYST and horizon_left <= _SURPLUS_P1_MIN_DAYS)
            # PRESS_STABLE: hold presses on their SKU (rule 2a) — do NOT do the aggressive
            # month-end surplus release that causes the ping-pong. Instead release ONLY near-done
            # presses (rule 2b, below): the block still runs, but the release test is replaced.
            _stable_hold = _PRESS_STABLE and _PRESS_RELEASE_DAYS <= 0   # pure hold if window=0
            if _stable_hold:
                _surplus_ok = False
            if _SURPLUS_RELEASE_ENABLED and horizon_left > 0 and _surplus_ok:
                _sku_presses: dict[str, list] = {}
                for p in sorted(demand_running_presses):
                    cs = press_to_sku.get(p)
                    if (cs is not None and cs in ri_skus
                            and updated_demand.get(cs, 0.0) > 0
                            and p not in demand_done_free):
                        _sku_presses.setdefault(cs, []).append(p)
                for sku in sorted(_sku_presses):
                    n = press_count.get(sku, 0)
                    if n <= 1:
                        continue
                    rate = _qty_per_press_per_day(
                        ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                    if rate <= 0:
                        continue
                    rem = updated_demand.get(sku, 0.0)
                    if _PRESS_STABLE:
                        # Rule 2(b): release ONLY if the SKU is near-done — its remaining demand
                        # can be cleared by n-1 presses within _PRESS_RELEASE_DAYS. Winding-down
                        # SKUs won't re-demand the press → redistribution without the ping-pong.
                        # Shed at most ONE press/SKU/day (gentle, monotone down-ramp).
                        if rem > (n - 1) * rate * _PRESS_RELEASE_DAYS:
                            continue
                        release = 1
                    else:
                        # #3 P2: size presses_needed against a slightly SHORTENED horizon so the SKU
                        # keeps ~1 buffer press (dead-band) — avoids over-release + CO/re-acquire ping-pong.
                        _hl = (horizon_left * (1.0 - _SURPLUS_P2_MARGIN)) if _SURPLUS_HYST else horizon_left
                        presses_needed = max(1, math.ceil(rem / (rate * max(1e-9, _hl))))
                        surplus = n - presses_needed
                        if surplus <= 0:
                            continue
                        # Uncapped: pool the TRUE surplus (may be all-but-one of the
                        # SKU's presses). The global pairing + n-1 RI-protection decide
                        # how many actually move — no arbitrary per-SKU-per-day throttle.
                        release = min(surplus, len(_sku_presses[sku]))
                    for p in _sku_presses[sku][:release]:
                        newly_free.append(p)
                        surplus_free.add(p)

            # ── Narrow starvation valve (PRESS_VALVE, under PRESS_STABLE hold only) ──
            # Inject a few held presses into the free pool ONLY when severe starvation exists,
            # taken from the most-over-supplied donors (with margin so they won't re-acquire).
            # The global pairing loop then routes them to the neediest (Class-A) targets. This
            # recovers the coverage pure hold gives up, without the general surplus ping-pong.
            if (_PRESS_STABLE and _PRESS_VALVE and _SURPLUS_RELEASE_ENABLED
                    and horizon_left > 0):
                _dctv = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
                _n_starved = 0
                for t in all_demand_skus:
                    rt = updated_demand.get(t, 0.0)
                    if rt <= 0:
                        continue
                    ratet = _qty_per_press_per_day(ct_map.get(t, _dctv))
                    if ratet <= 0:
                        continue
                    nt = press_count.get(t, 0)
                    # severely starved: current presses can't clear demand by month-end (nt=0 ⇒ yes)
                    if rt <= nt * ratet * horizon_left:
                        continue
                    # SERVABLE only: an extra press can help this target only if it has ≥2 FREE
                    # eligible moulds. Counting mould-scarce/unservable targets (July's real gap)
                    # would inflate the release cap → strip donors that can't be usefully placed.
                    if _p0_gate:
                        _fm = sum(1 for _m in _p0_sm.get(str(t), set())
                                  if _p0_owner.get(_m) is None)
                        if _fm < 2:
                            continue
                    _n_starved += 1
                if _n_starved > 0:
                    _cap = min(_n_starved, max_co_per_day - co_used)
                    _donors = []
                    for p in sorted(demand_running_presses):
                        cs = press_to_sku.get(p)
                        if (cs is None or cs not in ri_skus
                                or updated_demand.get(cs, 0.0) <= 0
                                or p in demand_done_free or p in surplus_free):
                            continue
                        n = press_count.get(cs, 0)
                        if n <= 1:
                            continue
                        rate = _qty_per_press_per_day(ct_map.get(cs, _dctv))
                        if rate <= 0:
                            continue
                        rem = updated_demand.get(cs, 0.0)
                        need = math.ceil(rem / (rate * horizon_left))
                        surplus = n - need
                        if surplus >= _PRESS_VALVE_MARGIN:
                            _donors.append((surplus, cs, p))
                    # most-over-supplied first; ≤1 press per donor SKU per day (gentle)
                    _donors.sort(key=lambda d: (-d[0], d[1], d[2]))
                    _used_donor_sku: set = set()
                    _rel = 0
                    for surplus, cs, p in _donors:
                        if _rel >= _cap:
                            break
                        if cs in _used_donor_sku:
                            continue
                        _used_donor_sku.add(cs)
                        newly_free.append(p)
                        surplus_free.add(p)
                        _rel += 1

            # ── Size-balanced over-cap press MIGRATION (SIZE_BAL) ────────────────
            # SIZE_BAL blocks NEW over-cap changeovers, but presses INHERITED on an
            # over-supplied inch (e.g. July starting on June's 15"-heavy Day-0 state)
            # would sit there until demand-done. Proactively free them toward inches
            # building can feed: for each over-cap inch, release presses (preferring
            # those whose SKU has the most under-cap buildable alternatives) until the
            # inch is back to building capacity — but NEVER drain an inch whose over-cap
            # is real coverable demand (n-1 RI-protection). Freed presses flow through
            # the same capped global-pairing loop → throttled to MAX_CHANGEOVERS_PER_DAY,
            # spread across days. Migration only OFFERS presses; the hard filter routes
            # each to an under-cap inch, or it stays put if none is better.
            overcap_free: set = set()
            if _size_bal_migrate:
                _cons0 = _inch_consumption()
                _dct0 = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
                def _n_alt(p: str) -> int:
                    return sum(1 for t in press_to_demand_targets.get(p, [])
                               if updated_demand.get(t, 0) > 0 and not _over_cap(t, _cons0))
                for _i in sorted(_cons0):
                    _excess = _cons0[_i] - _bic.get(_i, 0.0)
                    if _excess <= 0:
                        continue
                    _on_inch = [p for p in sorted(demand_running_presses)
                                if _sku_inch.get(str(press_to_sku.get(p, "")), "") == _i
                                and p not in newly_free and p not in overcap_free]
                    _on_inch.sort(key=lambda p: (-_n_alt(p), str(p)))
                    _released = 0.0
                    for p in _on_inch:
                        if _released >= _excess:
                            break
                        if _n_alt(p) == 0:
                            continue                     # nowhere better → leave it put
                        _old = press_to_sku.get(p, "")
                        # DELIVERY_PRIORITY: never release a committed-delivery press pre-deadline
                        # (same reservation as the global-pairing donor guard). Identity when off.
                        if (_prio_res and _old in _pdm
                                and updated_demand.get(_old, 0) > 0 and day <= _pdm[_old]):
                            continue
                        if _old in ri_skus:              # n-1 RI-protection on the donor inch
                            _nold   = press_count.get(_old, 0) - 1
                            _remold = updated_demand.get(_old, 0.0)
                            if _remold > 0 and _nold > 0:
                                _rold = _qty_per_press_per_day(ct_map.get(_old, _dct0))
                                if _rold > 0 and _remold / (_nold * _rold) > horizon_left:
                                    continue             # real demand needs this press
                        newly_free.append(p)
                        overcap_free.add(p)
                        _released += _qty_per_press_per_day(ct_map.get(_old, _dct0))

            if not newly_free:
                continue

            if _SURPLUS_RELEASE_ENABLED:
                # ══ Global (free-press, target-SKU) pair assignment ══════════════
                # Pool = RO + demand-done + surplus presses. Each iteration scores
                # EVERY eligible (press, target) pair globally by
                # (urgency_class, constraint=min(flex_press, flex_target), need, …)
                # and fires the single best, up to the daily cap — mirroring the
                # building-side _GLOBAL_ASSIGN. Building-supply is a hard filter;
                # n-1 RI-protection guards the donor. A press with no eligible
                # buildable target is simply not freed (stays on its donor SKU).
                _dct = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
                _pool = list(dict.fromkeys(newly_free))
                _assigned: set = set()
                # PRESS_DWELL per-day rate-limit counters (reset each day): how many presses each
                # SKU has GAINED / SHED so far today, capping the daily press-count change so the
                # per-SKU series ramps gradually instead of spiking.
                _gain_today: dict[str, int] = {}
                _shed_today: dict[str, int] = {}
                while co_used < max_co_per_day:
                    _pairs: list = []
                    _flex_p: dict = {}
                    _flex_t: dict = {}
                    for p in _pool:
                        if p in _assigned:
                            continue
                        old_sku = press_to_sku.get(p, "")
                        # PRESS_DWELL shed cap: this SKU already shed its daily quota → don't move
                        # another of its presses today (gradual down-ramp, no 3→0 cliff).
                        if (_PRESS_DWELL and old_sku
                                and _shed_today.get(old_sku, 0) >= _PRESS_MAX_SHED_PER_DAY):
                            continue
                        # DELIVERY_PRIORITY reservation: a press running a committed SKU is
                        # NEVER CO'd away while that SKU still has demand and its deadline has
                        # not passed — the committed SKU's GT feed must not wobble. Stricter
                        # than the n-1 RI-protection below (protects the whole press, not just
                        # when n-1 can't cover). Inactive/empty map → skipped (identity).
                        if (_prio_res and old_sku in _pdm
                                and updated_demand.get(old_sku, 0) > 0 and day <= _pdm[old_sku]):
                            continue
                        for target in press_to_demand_targets.get(p, []):
                            if target == old_sku:
                                continue
                            # L2 PRESS_RETURN_BLOCK: this press already CO'd AWAY from `target` — don't
                            # let it boomerang back (same-press round-trip). A DIFFERENT eligible press
                            # serves `target`. EXEMPT a fully-abandoned target (0 presses → must be able
                            # to restart) and a delivery-priority target that still wants it (deadline).
                            if (_PRESS_RETURN_BLOCK and target in _press_left.get(p, ())
                                    and (_L2_STRICT or press_count.get(target, 0) > 0)
                                    and not (_prio_acq and _prio_wants(target, day))):
                                continue
                            rem = updated_demand.get(target, 0)
                            if rem <= 0:
                                continue
                            # TAIL damper: in the last few working days, don't START a cold SKU
                            # (no presses on it) — building has wound down and can't ramp a fresh
                            # SKU that late, so the press would just starve. Keeps curing SKUs/day
                            # aligned to what building actually feeds. (Warm SKUs still gain presses.)
                            if (_TAIL_NO_COLD and horizon_left <= _TAIL_NO_COLD_DAYS
                                    and press_count.get(target, 0) == 0):
                                continue
                            # CAMPAIGN_PLAN: a SKU may gain presses only up to TODAY's campaign target
                            # (0 outside its window → cold/deferred/finished SKUs get no press). This is
                            # the per-day generalization of _target_peak, sized to per-inch building supply.
                            if _CAMPAIGN_PLAN and (
                                    target in _peaked
                                    or press_count.get(target, 0)
                                    >= _campaign_target.get(target, {}).get(day, 0)):
                                continue
                            # STATEFUL_PLAN: unimodal ramp — a SKU may gain presses only up to its
                            # planned target and only while it has NOT started shedding (no re-gain
                            # after a decrease → no boomerang, smooth up→peak→down).
                            if _STATEFUL_PLAN and (
                                    (_SP_MONOTONE and target in _peaked)
                                    or (_SP_CAP and press_count.get(target, 0)
                                        >= _target_peak.get(target, 10**6))):
                                continue
                            if _PRESS_DWELL:
                                # GAIN cap: target already took its daily quota of new presses.
                                if _gain_today.get(target, 0) >= _PRESS_MAX_GAIN_PER_DAY:
                                    continue
                                # Anti-boomerang COOLDOWN: this target lost a press within the last
                                # PRESS_BOOMERANG_COOLDOWN days → don't bring one back (the re-acquire
                                # leg of the ping-pong). It is a building-limited inch anyway; a
                                # re-added press just runs dry. EXEMPT a target at 0 presses (a
                                # fully-abandoned SKU must be allowed to restart) so we never strand it.
                                if (n_t := press_count.get(target, 0)) > 0 and (
                                        day - _sku_lost_day.get(target, -10**9)
                                        < _PRESS_BOOMERANG_COOLDOWN):
                                    continue
                            n_t = press_count.get(target, 0)
                            rate_t = _qty_per_press_per_day(ct_map.get(target, _dct))
                            # L1 TAIL_DAMP: in the last TD_TAIL_DAYS working days, block a MARGINAL
                            # warm re-acquire — target already has presses AND < TD_MIN_RESID_DAYS
                            # press-days of demand remain (the month-end cascade top-up that just
                            # re-adds a press onto a nearly-done SKU, which starves anyway). Cold
                            # (0-press) restarts are TAIL_NO_COLD's job; priority targets exempt.
                            if (_TAIL_DAMP and horizon_left <= _TD_TAIL_DAYS
                                    and n_t > 0 and rate_t > 0
                                    and rem < rate_t * _TD_MIN_RESID_DAYS
                                    and not (_prio_acq and _prio_wants(target, day))):
                                continue
                            # PRESS_RATCHET (soft): a target that already shed a press is capped at
                            # its ceiling (count at the last decrease) — BUT the cap YIELDS to a
                            # genuine re-ramp. Block re-acquire above the ceiling only when the
                            # target's current presses can still clear its remaining demand by
                            # month-end (a marginal Class-B top-up = the ping-pong). If the target
                            # truly can't keep up (rem > n_t·rate·horizon_left → Class-A need) or has
                            # 0 presses, the cap is overridden so real demand is still served.
                            if (_PRESS_RATCHET and n_t > 0
                                    and n_t >= _pc_ceiling.get(target, float("inf"))
                                    and rate_t > 0
                                    and rem <= n_t * rate_t * horizon_left):
                                continue
                            is_nri = target in nri_skus
                            is_ri = target in ri_skus
                            if is_nri:
                                pass
                            elif is_ri and n_t > 0:
                                cur_days = rem / (n_t * rate_t) if rate_t > 0 else float("inf")
                                # A committed target is "on track" only vs its OWN (nearer)
                                # deadline, so we keep acquiring presses for a priority SKU that
                                # is fine for month-end but behind for its earlier date.
                                _rihz = horizon_left
                                _ddt = (_dd_days(target, day)
                                        if (_prio_acq and _prio_wants(target, day)) else None)
                                if _ddt is not None:
                                    _rihz = min(horizon_left, _ddt)
                                if cur_days <= _rihz:
                                    continue   # RI already on track
                            else:
                                continue
                            # 5b building-supply hard filter (PERSKU_FEED refines per-SKU)
                            if buildable_rate is not None:
                                _br = _perSKU_feed(target, n_t, rate_t)
                                if _br is not None and (n_t + 1) * rate_t > _br:
                                    continue
                            # #4 mould-availability hard filter: skip a target the press
                            # can't mount 2 eligible free moulds for (mould-feasible only).
                            if _p0_gate and _n_free_for(target, p) < 2:
                                continue
                            # n-1 RI-protection on the donor (if freeing an RI press).
                            # SUPPLY_ALIGN donors are EXEMPT: they were selected precisely because
                            # building can't supply their full draw (the freed press was starving),
                            # so moving it to a supply-rich SKU never loses real coverage.
                            if old_sku in ri_skus and p not in supply_free and p not in campaign_free:
                                n_old = press_count.get(old_sku, 0) - 1
                                rem_old = updated_demand.get(old_sku, 0)
                                if rem_old > 0 and n_old > 0:
                                    rate_old = _qty_per_press_per_day(ct_map.get(old_sku, _dct))
                                    if rate_old > 0 and rem_old / (n_old * rate_old) > horizon_left:
                                        continue   # can't spare this press
                            key0 = _urgency_sort_key(
                                priority_score=_priority_signal(target, p),
                                current_press_count=n_t, updated_demand=rem,
                                rate_per_day=rate_t, horizon_left=horizon_left,
                                deadline_days=(_dd_days(target, day)
                                               if (_prio_acq and _prio_wants(target, day)) else None))
                            _pairs.append((p, target, old_sku, key0))
                            _flex_p[p] = _flex_p.get(p, 0) + 1
                            _flex_t[target] = _flex_t.get(target, 0) + 1
                    if not _pairs:
                        break
                    _cons_now = _inch_consumption() if (_supply_on or _size_bal) else {}
                    # Size-balanced HARD pre-filter (SIZE_BAL): keep only pairs whose target inch still
                    # has building headroom — the cap DOMINATES urgency (an over-cap fire is a
                    # RUNNING-but-starved press that steals building from inches it CAN feed; that
                    # coverage is illusory). If nothing is under-cap this iteration, only a FREED press
                    # (demand-done / surplus / RO — produces nothing where it sits) may still move, and
                    # only onto a buildable inch (_bic[i] > 0); else stop firing (draw stays ≤ supply,
                    # the plant's deliberate under-draw of building-limited inches).
                    _cand = _pairs
                    if _size_bal:
                        _under = [pr for pr in _pairs if not _over_cap(pr[1], _cons_now)]
                        if _under:
                            _cand = _under
                        else:
                            _freed = demand_done_free | surplus_free | pending_ro_presses
                            _cand = [pr for pr in _pairs
                                     if pr[0] in _freed
                                     and _bic.get(_sku_inch.get(str(pr[1]), ""), 0.0) > 0]
                            if not _cand:
                                break
                    # Global key: urgency class → constraint → BUILDING SUPPLY → same-inch → need → ties
                    def _supply_headroom_bucket(target: str) -> int:
                        # 0 = idle building can supply the NEXT press for this SKU (buildable_rate −
                        # current draw ≥ one press's rate) → CO here activates idle building; else 1.
                        if buildable_rate is None:
                            return 0
                        _rt = _qty_per_press_per_day(ct_map.get(target, _dct))
                        _nt = press_count.get(target, 0)
                        _br = _perSKU_feed(target, _nt, _rt)
                        if _br is None:
                            return 0
                        return 0 if (_br - _nt * _rt) >= _rt else 1
                    def _cokey(pr):
                        _cls = pr[3][0]                             # urgency_class (0=Class-A,1=Class-B)
                        _con = min(_flex_p[pr[0]], _flex_t[pr[1]])  # constraint
                        _sup = _supply_pref(pr[1], _cons_now)       # building supply headroom
                        _si  = _same_inch(pr[0], pr[1])             # 0 = same inch as press
                        _hb  = _supply_headroom_bucket(pr[1]) if _SUPPLY_ALIGN else 0
                        # v3 WARM-FIRST: within the same urgency class, fill a SKU that ALREADY has
                        # presses (warm) before opening a COLD one — concentrates presses on fewer SKUs
                        # so cureRUN stays close to what building feeds (fewer starved presses).
                        _warm = (0 if press_count.get(pr[1], 0) > 0 else 1) if (
                            _STATEFUL_PLAN and _SP_WARM_FIRST) else 0
                        _tail = (pr[3][1], pr[3][2],                # -priority, after_days
                                 ct_map.get(pr[1], _dct), pr[0], pr[1])
                        if _STATEFUL_PLAN and _SP_WARM_FIRST:
                            _base = (_cls, _warm, _con, _sup, _si, *_tail)  # warm SKUs fill before cold
                        elif _SUPPLY_ALIGN:
                            # supply-headroom right after urgency: pull presses to SKUs idle
                            # building can actually feed (activates idle Stage-2/VMI).
                            _base = (_cls, _hb, _con, _sup, _si, *_tail)
                        elif _SAME_INCH_FIRST and _SAME_INCH_RANK == "top":
                            _base = (_si, _cls, _con, _sup, *_tail)  # same-inch beats everything
                        elif _SAME_INCH_FIRST:
                            _base = (_cls, _si, _con, _sup, *_tail)  # "safe": Class-A first, then same-inch
                        else:
                            _base = (_cls, _con, _sup, _si, *_tail)  # OFF: current order (same-inch 4th)
                        if not _prio_acq:
                            return _base
                        # DELIVERY_PRIORITY: committed targets fire FIRST, EARLIEST-DEADLINE-FIRST.
                        # (1,0.0) is the identical constant for every pair when no committed target
                        # is in the candidate set → order-preserving; a committed target still BELOW
                        # its mould-pair cap gets (0,dd) so it wins, smaller deadline-day first = EDF.
                        _dd = _pdm.get(pr[1])
                        _pk = ((0, float(_dd)) if (_dd is not None
                                                   and updated_demand.get(pr[1], 0) > 0
                                                   and _prio_wants(pr[1], day)) else (1, 0.0))
                        return (_pk, *_base)
                    best = min(_cand, key=_cokey)
                    p, new_sku, old_sku, _ = best
                    co_events.append(
                        {"day": day, "press": p, "old_sku": old_sku, "new_sku": new_sku})
                    press_to_sku[p] = new_sku
                    if _p0_gate:
                        _p0_mount(p, new_sku)                  # #4: claim the target's 2 moulds
                    press_count[old_sku] = max(0, press_count.get(old_sku, 0) - 1)
                    press_count[new_sku] = press_count.get(new_sku, 0) + 1
                    if (_STATEFUL_PLAN or _CAMPAIGN_PLAN) and old_sku:
                        _peaked.add(old_sku)   # any shed → SKU has peaked, may not re-gain (monotone)
                    if _PRESS_RETURN_BLOCK and old_sku:
                        _press_left.setdefault(p, set()).add(old_sku)  # L2: press p has left old_sku
                    if _PRESS_DWELL:        # record the swap for the rate-limit + cooldown
                        _gain_today[new_sku] = _gain_today.get(new_sku, 0) + 1
                        if old_sku:
                            _shed_today[old_sku] = _shed_today.get(old_sku, 0) + 1
                            _sku_lost_day[old_sku] = day
                    if _PRESS_RATCHET:      # lock old_sku's ceiling at its new (reduced) count
                        _pc_ceiling[old_sku] = press_count[old_sku]
                    pending_ro_presses.discard(p)
                    demand_running_presses.add(p)
                    _assigned.add(p)
                    co_used += 1
                    daily_co_used[day] = co_used
                continue   # global path handled this day; skip the greedy fire-loop

            # Score candidates: target = NRI (any) OR under-supplied RI
            # Under-supplied RI: current press count cannot meet demand in time
            candidates: list[tuple] = []
            for p in dict.fromkeys(newly_free):  # deduplicate, order-preserving
                old_sku = press_to_sku.get(p, "")
                for target in press_to_demand_targets.get(p, []):
                    if target == old_sku:
                        continue               # don't CO to the same SKU
                    rem = updated_demand.get(target, 0)
                    if rem <= 0:
                        continue               # demand already fulfilled

                    n_t  = press_count.get(target, 0)
                    ct_t = ct_map.get(target, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                    rate_t = _qty_per_press_per_day(ct_t)

                    is_nri = target in nri_skus
                    is_ri  = target in ri_skus

                    if is_nri:
                        pass   # always eligible
                    elif is_ri and n_t > 0:
                        # Only eligible if under-supplied: needs more than horizon_left days
                        current_days = rem / (n_t * rate_t) if rate_t > 0 else float("inf")
                        if current_days <= horizon_left:
                            continue   # RI is already on track — skip
                    else:
                        continue

                    key = _urgency_sort_key(
                        priority_score=_priority_signal(target, p),
                        current_press_count=n_t,
                        updated_demand=rem,
                        rate_per_day=rate_t,
                        horizon_left=horizon_left,
                    )
                    candidates.append((key, p, old_sku, target))

            _dct = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
            _cons_fb = _inch_consumption() if _supply_on else {}
            def _fbkey(x):
                _cls = x[0][0]                                     # Class A (0) before Class B (1)
                _sup = _supply_pref(x[3], _cons_fb)               # building supply headroom
                _si  = _same_inch(x[1], x[3])                     # 0 = same inch
                _tail = (x[0][1], x[0][2], ct_map.get(x[3], _dct),
                         len(press_to_demand_targets.get(x[1], [])), x[1], x[3])
                if _SAME_INCH_FIRST and _SAME_INCH_RANK == "top":
                    return (_si, _cls, _sup, *_tail)
                if _SAME_INCH_FIRST:
                    return (_cls, _si, _sup, *_tail)              # "safe": Class-A first, then same-inch
                return (_cls, _sup, _si, *_tail)                  # OFF: current order
            candidates.sort(key=_fbkey)

            assigned: set = set()
            for key, p, old_sku, new_sku in candidates:
                if co_used >= max_co_per_day:
                    break
                if p in assigned:
                    continue

                # Plant limit = MAX_CO_PER_DAY (hard).  For normal presses, only fire
                # Class A COs — those where demand CANNOT be met in the remaining
                # horizon without this additional press.
                # EXCEPTION: presses freed because their RI demand just hit 0
                # (demand_done_free) — these produce NOTHING useful, so fire their CO
                # immediately regardless of urgency class. Any production > 0 is better.
                urgency_class = key[0]   # 0 = Class A (critical), 1 = Class B
                is_demand_done = p in demand_done_free
                if urgency_class != 0 and not is_demand_done and p not in surplus_free:
                    continue  # Class B — skip (unless demand-done or surplus-release)

                # Re-check with CURRENT press_count — earlier COs this day may have
                # already satisfied this target's demand. Without this guard, the same
                # low-demand NRI SKU can absorb the entire daily CO budget (10 presses
                # all see n_t=0 in the pre-built candidates list, all appear Class A).
                cur_n = press_count.get(new_sku, 0)
                cur_rem = updated_demand.get(new_sku, 0)
                if cur_rem <= 0:
                    continue   # demand fulfilled by earlier CO today
                if cur_n > 0:
                    rate_recheck = _qty_per_press_per_day(
                        ct_map.get(new_sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                    )
                    if rate_recheck > 0 and cur_rem / (cur_n * rate_recheck) <= horizon_left:
                        continue  # downgraded — existing presses now sufficient

                # Guard: don't CO an RI press if remaining presses can't cover its demand.
                # This prevents early CO'ing of RI presses whose SKU demand hasn't been met
                # when the theoretical drain races ahead of actual building output.
                if old_sku in ri_skus:
                    n_old_remaining = press_count.get(old_sku, 0) - 1
                    rem_old = updated_demand.get(old_sku, 0)
                    if rem_old > 0 and n_old_remaining > 0:
                        rate_old = _qty_per_press_per_day(
                            ct_map.get(old_sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                        )
                        if rate_old > 0 and rem_old / (n_old_remaining * rate_old) > horizon_left:
                            continue  # remaining n-1 presses cannot cover old_sku demand

                # 5b building-supply guard (surplus releases only): don't move a
                # surplus press onto a target that building cannot feed — otherwise
                # the reassigned press just starves (RUNNING, no GT). buildable_rate
                # is the per-SKU sustainable GT/day building can produce.
                if p in surplus_free and buildable_rate is not None:
                    _nr = _qty_per_press_per_day(
                        ct_map.get(new_sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN))
                    _br = _perSKU_feed(new_sku, cur_n, _nr)     # PERSKU_FEED refines per-SKU
                    if _br is not None:
                        if (cur_n + 1) * _nr > _br:
                            continue  # building can't supply the extra press → skip

                co_events.append(
                    {"day": day, "press": p, "old_sku": old_sku, "new_sku": new_sku}
                )
                press_to_sku[p]  = new_sku
                press_count[old_sku] = max(0, press_count.get(old_sku, 0) - 1)
                press_count[new_sku] = press_count.get(new_sku, 0) + 1

                pending_ro_presses.discard(p)       # no longer stranded RO
                demand_running_presses.add(p)       # now running a demand SKU

                assigned.add(p)
                co_used += 1
                daily_co_used[day] = co_used

        # ── Rescue pass: NRI SKUs still without any CO ────────────────────────────
        # Main loop only frees presses when (a) RO presses are stranded or
        # (b) demand-running presses fulfil their SKU's demand completely.
        # Some NRI SKUs never match either condition because their compatible
        # curing presses are busy with RI SKUs that never fully drain demand.
        # Solution: donate one press from any RI SKU that has n_presses > 1
        # AND can still meet its own demand with n−1 presses.
        scheduled_nri = {ev["new_sku"] for ev in co_events if ev["new_sku"] in nri_skus}
        rescue_nri = sorted(
            nri_skus - scheduled_nri,
            key=lambda s: (
                *_urgency_sort_key(
                    _sku_priority_signal(s),
                    0,
                    float(updated_demand.get(s, 0)),
                    _qty_per_press_per_day(ct_map.get(s, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)),
                    0,
                ),
                s,  # final deterministic tiebreak
            ),
        )
        n_rescued = 0
        if rescue_nri:
            print(f"  [CO Rescue] {len(rescue_nri)} NRI SKUs without CO — attempting rescue …")
            # Inverse map: nri_sku → set of presses that CAN produce it
            sku_to_compat: dict[str, set] = {}
            for press, targets in press_to_demand_targets.items():
                for t in targets:
                    sku_to_compat.setdefault(t, set()).add(press)

            for nri_sku in rescue_nri:
                rem = updated_demand.get(nri_sku, 0)
                if rem <= 0:
                    continue  # demand already zero — no point scheduling CO
                compatible = sorted(sku_to_compat.get(nri_sku, set()))
                scheduled = False
                for press in compatible:
                    current_sku = press_to_sku.get(press, "")
                    if current_sku == nri_sku:
                        continue
                    # Only donate from RI SKUs that have a spare press (n > 1)
                    if current_sku not in ri_skus:
                        continue
                    n_ri = press_count.get(current_sku, 0)
                    if n_ri <= 1:
                        continue  # can't spare — only press for that RI SKU
                    # Verify that RI SKU can still meet its FULL demand with n−1 presses.
                    # Previous bug: used updated_demand (= 0 after simulation with all
                    # n presses) which always passed the check — allowing COs even when
                    # n−1 presses cannot cover full demand across the horizon.
                    # Fix: use original demand_map value and compute actual capacity
                    # accounting for when the CO fires (earliest budget-available day).
                    ri_ct          = ct_map.get(current_sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                    ri_rate        = _qty_per_press_per_day(ri_ct)
                    ri_full_demand = float(demand_map.get(current_sku, 0))

                    # Find earliest day with CO budget (needed for capacity check)
                    _co_day = next(
                        (d for d in range(1, planning_days + 1)
                         if daily_co_used.get(d, 0) < max_co_per_day),
                        None,
                    )
                    if _co_day is None:
                        continue  # no CO budget anywhere in horizon

                    # Capacity: n_ri presses run days 1.._co_day, then n_ri−1 for rest.
                    # Count WORKING days only (holidays produce nothing). Empty holidays
                    # ⇒ identical to the plain day counts (parity).
                    cap_before = n_ri * ri_rate * _working_days_left(1, _co_day, _hol)
                    cap_after  = max(0, n_ri - 1) * ri_rate * _working_days_left(_co_day + 1, planning_days, _hol)
                    if (cap_before + cap_after) < ri_full_demand:
                        continue  # CO would leave RI SKU demand unmet

                    # Schedule CO on the earliest available day found above
                    co_events.append({
                        "day":     _co_day,
                        "press":   press,
                        "old_sku": current_sku,
                        "new_sku": nri_sku,
                    })
                    press_to_sku[press] = nri_sku
                    press_count[current_sku] = max(0, n_ri - 1)
                    press_count[nri_sku] = press_count.get(nri_sku, 0) + 1
                    daily_co_used[_co_day] = daily_co_used.get(_co_day, 0) + 1
                    demand_running_presses.add(press)
                    pending_ro_presses.discard(press)
                    scheduled = True
                    n_rescued += 1
                    break

            still_missing = len(rescue_nri) - n_rescued
            print(f"  [CO Rescue] Rescued {n_rescued} NRI SKUs via spare-press donation"
                  + (f" | {still_missing} still without CO (no compatible spare press)"
                     if still_missing else ""))

        # ── Summary ───────────────────────────────────────────────────────────
        total_slots = max_co_per_day * planning_days
        used_slots  = len(co_events)
        co_by_day   = {}
        for ev in co_events:
            co_by_day[ev["day"]] = co_by_day.get(ev["day"], 0) + 1
        peak_day = max(co_by_day, key=co_by_day.get) if co_by_day else 0

        print(f"  [CO Scheduler] {used_slots} COs used / {total_slots} available "
              f"({used_slots/total_slots*100:.1f}%)")
        print(f"  [CO Scheduler] Peak: Day {peak_day} "
              f"({co_by_day.get(peak_day,0)} COs)  |  "
              f"Zero-CO days: {sum(1 for d in range(1,planning_days+1) if d not in co_by_day)}")

        if pending_ro_presses:
            print(f"  [WARN] {len(pending_ro_presses)} RO presses still stranded at Day 31 "
                  f"(no compatible demand SKU found after filter):")
            for p in sorted(pending_ro_presses):
                n_compat = len(press_to_demand_targets.get(p, []))
                print(f"    Press {p} ({press_to_sku.get(p,'?')}): "
                      f"{n_compat} compatible targets in allowable")

        # expose the campaign plan so b2c can enforce the active-set (idle out-of-plan presses)
        self.campaign_target = _campaign_target if _CAMPAIGN_PLAN else {}
        return co_events


# ══════════════════════════════════════════════════════════════════════════════
# DAY SIMULATOR  (Pass 2)
# ══════════════════════════════════════════════════════════════════════════════

class DaySimulator:
    """
    Simulate 31 days of curing consumption using the pre-computed CO schedule.

    For each day D, the sheet contains:
      SKUCode, Category, Running_Press_Count, Total_Available_Moulds,
      Effective_CT_Min, Qty_Per_Press_Per_Shift, Total_GT_Per_Shift_DayN,
      Updated_Demand_Qty, Production_Days, Priority_Score
    """

    def simulate(
        self,
        df_day0: pd.DataFrame,
        df_demand: pd.DataFrame,
        df_allowable: pd.DataFrame,
        ct_map: dict[str, float],
        co_events: list[dict],
        planning_days: int = PLANNING_DAYS,
    ) -> list[pd.DataFrame]:
        """Returns one DataFrame per planning day (index 0 = Day 1).

        planning_days: horizon for THIS call (defaults to the module constant).
        Must be honoured — reading the module global here silently used the
        bc_config horizon whenever a caller passed a different one.
        """

        demand_map   = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Quantity"]))
        priority_map = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Priority"]))

        # Allowable moulds count per SKU (total eligible presses in master)
        allowable_count: dict[str, int] = {}
        for _, r in df_allowable.iterrows():
            sku = str(r["SKUCode"]).strip()
            machines = r.get("Machines", [])
            allowable_count[sku] = len(machines) if machines else 0

        # Build CO lookup: day → list of (press, old_sku, new_sku)
        co_by_day: dict[int, list] = {}
        for ev in co_events:
            co_by_day.setdefault(ev["day"], []).append(ev)

        # Universe of SKUs (demand ∪ running presses)
        all_skus = sorted(set(df_day0["SKUCode"].tolist()) | set(demand_map.keys()))

        # Initial state from Day 0 table
        press_count: dict[str, int] = {}
        category_map: dict[str, str] = {}
        for _, r in df_day0.iterrows():
            sku = str(r["SKUCode"])
            press_count[sku] = int(r.get("Running_Press_Count", 0))
            category_map[sku] = str(r.get("Category", "Non-Runner-In"))

        # Running demand
        updated_demand: dict[str, float] = {
            sku: float(demand_map.get(sku, 0)) for sku in all_skus
        }

        daily_sheets: list[pd.DataFrame] = []
        _hol = _holiday_day_index_set()   # plant holidays → working-day urgency horizon

        for day in range(1, planning_days + 1):
            horizon_left = _working_days_left(day, planning_days, _hol)
            _is_hol = day in _hol    # plant holiday: no curing, no demand drain this day

            # Apply COs for this day (press count update effective from Shift C same day).
            # Planned COs still execute on a holiday (setup crew works) so the press is
            # ready to run the new SKU on the first working day — consistent with the
            # rolling pipeline's holiday CO handling.
            for ev in co_by_day.get(day, []):
                old = ev["old_sku"]
                new = ev["new_sku"]
                press_count[old] = max(0, press_count.get(old, 0) - 1)
                press_count[new] = press_count.get(new, 0) + 1
                if category_map.get(new) == "Non-Runner-In":
                    category_map[new] = "Runner-In"

            # Snapshot today's GT output BEFORE draining (used in the sheet). On a plant
            # holiday NOTHING cures → gt_today stays empty → every day-sheet's "GT / Shift"
            # column is 0 AND the drain loop below (which iterates gt_today) makes no change,
            # so remaining demand carries across the holiday unchanged. This keeps the curing
            # consumption output consistent with the building/curing schedules (holiday idle).
            gt_today: dict[str, int] = {}
            if not _is_hol:
                for sku in all_skus:
                    n = press_count.get(sku, 0)
                    if n <= 0:
                        continue
                    ct  = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                    qps = _qty_per_press_per_shift(ct)
                    gt_today[sku] = n * qps   # per-shift; × SHIFTS_PER_DAY = day total

            # Drain demand by today's full production FIRST so the day sheet
            # shows "Updated_Demand_Qty" = remaining demand AFTER today runs.
            # This ensures Day 31's row correctly reflects the closing balance.
            for sku, qps in gt_today.items():
                ct   = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                rate = _qty_per_press_per_day(ct)
                updated_demand[sku] = max(0.0, updated_demand.get(sku, 0) - rate * press_count.get(sku, 0))

            # Build day sheet (demand figures are closing balances for the day)
            rows = []
            for sku in all_skus:
                dem = demand_map.get(sku, 0)
                if dem <= 0 and press_count.get(sku, 0) == 0:
                    continue  # not in demand and not running — skip

                cat = category_map.get(sku, "Non-Runner-In")
                n   = press_count.get(sku, 0)
                ct  = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
                qps = _qty_per_press_per_shift(ct)

                rem_demand = max(0.0, updated_demand.get(sku, 0))

                # production_days from closing balance: how many more days at current rate
                if n > 0 and rem_demand > 0:
                    rate_day = _qty_per_press_per_day(ct)
                    prod_days = round(_production_days(rem_demand, n, rate_day), 1)
                else:
                    prod_days = None   # blank in Excel

                rows.append({
                    "SKUCode":                  sku,
                    "Category":                 cat,
                    "Running_Press_Count":       n,
                    "Total_Available_Moulds":    allowable_count.get(sku, 0),
                    "Effective_CT_Min":          round(ct, 2),
                    "Qty_Per_Press_Per_Shift":   qps,
                    "Total_GT_Per_Shift_DayN":   gt_today.get(sku, 0),
                    "Updated_Demand_Qty":        int(rem_demand),
                    "Production_Days":           prod_days,
                    "Priority_Score":            priority_map.get(sku, 0),
                })

            df_day = pd.DataFrame(rows)
            _ord = {"Runner-In": 0, "Runner-Out": 1, "Non-Runner-In": 2}
            df_day["_o"] = df_day["Category"].map(_ord).fillna(3)
            df_day = (df_day
                      .sort_values(["_o", "Priority_Score"], ascending=[True, False])
                      .drop(columns=["_o"])
                      .reset_index(drop=True))

            daily_sheets.append(df_day)

        return daily_sheets


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════

class DynamicExporter:
    """Write the 31-sheet Excel file."""

    _COLS = [
        "SKUCode", "Category", "Running_Press_Count", "Total_Available_Moulds",
        "Effective_CT_Min", "Qty_Per_Press_Per_Shift", "Total_GT_Per_Shift_DayN",
        "Updated_Demand_Qty", "Production_Days", "Priority_Score",
    ]
    # Human-readable column headers for day sheets
    _COL_HEADERS = {
        "Updated_Demand_Qty":    "Remaining Demand (after day)",
        "Total_GT_Per_Shift_DayN": "GT / Shift (this day)",
    }

    _CAT_FILL = {
        "Runner-In":     PatternFill("solid", fgColor="E2EFDA"),
        "Runner-Out":    PatternFill("solid", fgColor="FCE4D6"),
        "Non-Runner-In": PatternFill("solid", fgColor="FFF2CC"),
    }

    def _hdr_style(self):
        return {
            "fill": PatternFill("solid", fgColor=_NAVY),
            "font": Font(bold=True, color=_WHITE, size=10),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                bottom=Side(style="thin", color=_WHITE),
                right=Side(style="thin", color=_WHITE),
            ),
        }

    def _apply_hdr(self, ws, row=1):
        hdr = self._hdr_style()
        for col_idx, col_name in enumerate(self._COLS, start=1):
            label = self._COL_HEADERS.get(col_name, col_name.replace("_", " "))
            cell = ws.cell(row=row, column=col_idx, value=label)
            for k, v in hdr.items():
                setattr(cell, k, v)

    def _write_day_sheet(self, ws, df: pd.DataFrame, day: int):
        plan_date = PLAN_START + timedelta(days=day - 1)
        ws.title = f"Day_{day:02d}"

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self._COLS))
        title_cell = ws.cell(row=1, column=1,
                             value=f"Curing Consumption — Day {day:02d} "
                                   f"({plan_date.strftime('%d-%b-%Y')})")
        title_cell.font = Font(bold=True, size=11, color=_WHITE)
        title_cell.fill = PatternFill("solid", fgColor=_NAVY)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Header row
        self._apply_hdr(ws, row=2)
        ws.row_dimensions[2].height = 30

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
            cat = str(row.get("Category", ""))
            fill = self._CAT_FILL.get(cat)
            for c_idx, col in enumerate(self._COLS, start=1):
                val = row.get(col)
                # Production_Days: leave blank (None) when no press
                if col == "Production_Days" and val is None:
                    val = ""
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if fill:
                    cell.fill = fill
                if col in ("Updated_Demand_Qty", "Total_GT_Per_Shift_DayN",
                           "Running_Press_Count", "Qty_Per_Press_Per_Shift"):
                    cell.number_format = "#,##0"
                elif col in ("Effective_CT_Min", "Production_Days"):
                    cell.number_format = "0.0"
                elif col == "Priority_Score":
                    cell.number_format = "0.00"

        # Column widths
        _widths = [16, 16, 18, 20, 15, 20, 22, 20, 14, 14]
        for i, w in enumerate(_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Legend below data
        legend_row = len(df) + 4
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
        for cat, fill in self._CAT_FILL.items():
            legend_row += 1
            c = ws.cell(row=legend_row, column=1, value=cat)
            c.fill = fill
            c.alignment = Alignment(horizontal="left")

    def _write_co_sheet(self, ws, co_events: list[dict]):
        ws.title = "CO_Schedule"
        headers = ["Day", "Press", "Old_SKU", "New_SKU", "Plan_Date", "CO_Type"]
        hdr = self._hdr_style()
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h.replace("_", " "))
            for k, v in hdr.items():
                setattr(cell, k, v)

        for r_idx, ev in enumerate(co_events, start=2):
            plan_date = (PLAN_START + timedelta(days=ev["day"] - 1)).strftime("%d-%b-%Y")
            ws.cell(row=r_idx, column=1, value=ev["day"])
            ws.cell(row=r_idx, column=2, value=ev["press"])
            ws.cell(row=r_idx, column=3, value=ev["old_sku"])
            ws.cell(row=r_idx, column=4, value=ev["new_sku"])
            ws.cell(row=r_idx, column=5, value=plan_date)
            ws.cell(row=r_idx, column=6, value="curing_CO")
            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).alignment = Alignment(horizontal="center")

        for c_idx, w in enumerate([8, 12, 16, 16, 14, 12], start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        # Day-level summary below
        from collections import Counter
        co_by_day = Counter(ev["day"] for ev in co_events)
        summary_row = len(co_events) + 3
        ws.cell(row=summary_row, column=1, value="Day-level CO count:").font = Font(bold=True)
        for day in sorted(co_by_day):
            summary_row += 1
            ws.cell(row=summary_row, column=1, value=f"Day {day:02d}").alignment = Alignment(horizontal="center")
            ws.cell(row=summary_row, column=2, value=co_by_day[day]).alignment = Alignment(horizontal="center")

    def _cell(self, ws, row, col, value="", bold=False, fill=None,
              align="center", num_fmt=None, font_size=10, color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = Alignment(horizontal=align, vertical="center")
        f = Font(bold=bold, size=font_size)
        if color:
            f = Font(bold=bold, size=font_size, color=color)
        c.font = f
        if fill:
            c.fill = fill
        if num_fmt:
            c.number_format = num_fmt
        return c

    def _section_header(self, ws, row, col, text, n_cols=6):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + n_cols - 1)
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, color=_WHITE, size=10)
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    def _write_summary_sheet(
        self,
        ws,
        df_day0: pd.DataFrame,
        df_excluded: pd.DataFrame,
        daily_sheets: list,
        co_events: list,
        n_demand_raw: int,
        demand_path: str,
        planning_days: int,
    ):
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 22

        _LABEL_FILL = PatternFill("solid", fgColor="EBF3FA")
        _VAL_FILL   = PatternFill("solid", fgColor="FFFFFF")
        _WARN_FILL  = PatternFill("solid", fgColor="FFE0CC")
        _GOOD_FILL  = PatternFill("solid", fgColor="E2EFDA")
        _TH_FILL    = PatternFill("solid", fgColor="D6E4F0")

        hdr = self._hdr_style()
        r = 1

        # ── Title ──────────────────────────────────────────────────────────────
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1,
                    value="B2C Curing Consumption — 31-Day Plan Summary (May 2026)")
        c.font = Font(bold=True, size=13, color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1,
                    value=f"Demand file: {os.path.basename(demand_path)}   |   "
                          f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}")
        c.font = Font(italic=True, size=9, color="555555")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 16
        r += 2

        # ── Section A: Demand File Overview ───────────────────────────────────
        self._section_header(ws, r, 1, "A.  Demand File Overview", 6)
        r += 1

        demand_rows = (df_day0[df_day0["Category"] != "Runner-Out"])
        ri_rows  = df_day0[df_day0["Category"] == "Runner-In"]
        ro_rows  = df_day0[df_day0["Category"] == "Runner-Out"]
        nri_rows = df_day0[df_day0["Category"] == "Non-Runner-In"]

        # Exclude Runner-Out (non-demand) AND excluded SKUs (no master data) from
        # total_demand so that demand_left_day31 (which also excludes them) stays
        # on the same basis — avoids phantom fulfillment of 62,802 unproducible tyres.
        _excl_codes = set(df_excluded["SKUCode"].astype(str).str.strip()) if len(df_excluded) else set()
        total_demand = float(
            df_day0[
                (df_day0["Category"] != "Runner-Out") &
                (~df_day0["SKUCode"].astype(str).str.strip().isin(_excl_codes))
            ]["Demand_Qty"].sum()
        )
        n_eligible   = len(demand_rows)
        n_excluded   = len(df_excluded)

        overview = [
            ("Total SKUs in demand file",    n_demand_raw,            "#,##0",  None),
            ("Eligible SKUs (pass filter)",  n_eligible,              "#,##0",  _GOOD_FILL),
            ("Excluded SKUs (no data)",      n_excluded,              "#,##0",  _WARN_FILL if n_excluded else None),
            ("Total demand quantity (tyres)", int(total_demand),      "#,##0",  None),
        ]
        for label, val, fmt, fill in overview:
            self._cell(ws, r, 1, label, align="left",  fill=_LABEL_FILL)
            self._cell(ws, r, 2, val,   num_fmt=fmt,   fill=fill or _VAL_FILL)
            ws.row_dimensions[r].height = 16
            r += 1
        r += 1

        # ── Section B: Day 0 Category Breakdown ───────────────────────────────
        self._section_header(ws, r, 1, "B.  Day 0 Category Breakdown", 6)
        r += 1

        # sub-header
        for ci, htext in enumerate(
            ["Category", "SKU Count", "Press Count", "Demand Qty", "GT / Shift (Day 0)", ""],
            start=1
        ):
            c = ws.cell(row=r, column=ci, value=htext)
            for k, v in hdr.items():
                setattr(c, k, v)
        ws.row_dimensions[r].height = 20
        r += 1

        cat_data = [
            ("Runner-In",
             len(ri_rows),
             int(ri_rows["Running_Press_Count"].sum()),
             int(ri_rows["Demand_Qty"].sum()),
             int(ri_rows["Total_GT_Per_Shift_Day0"].sum()),
             ""),
            ("Runner-Out  ⚠ non-demand SKUs — CO candidates only",
             len(ro_rows),
             int(ro_rows["Running_Press_Count"].sum()),
             "—",
             int(ro_rows["Total_GT_Per_Shift_Day0"].sum()),
             "These presses will CO to demand SKUs"),
            ("Non-Runner-In",
             len(nri_rows),
             0,
             int(nri_rows["Demand_Qty"].sum()),
             0,
             "Awaiting curing press via CO"),
        ]
        for cat, sku_cnt, press_cnt, dem_qty, gt_shift, note in cat_data:
            fill = self._CAT_FILL.get(cat.split("  ")[0])
            self._cell(ws, r, 1, cat,       align="left", fill=fill)
            self._cell(ws, r, 2, sku_cnt,   num_fmt="#,##0", fill=fill)
            self._cell(ws, r, 3, press_cnt, num_fmt="#,##0", fill=fill)
            self._cell(ws, r, 4, dem_qty,   num_fmt="#,##0" if dem_qty != "—" else "@", fill=fill)
            self._cell(ws, r, 5, gt_shift,  num_fmt="#,##0", fill=fill)
            self._cell(ws, r, 6, note,      align="left",    fill=fill)
            ws.row_dimensions[r].height = 16
            r += 1

        # Total row
        total_fill = PatternFill("solid", fgColor="D6E4F0")
        self._cell(ws, r, 1, "TOTAL (demand SKUs)", bold=True, align="left", fill=total_fill)
        self._cell(ws, r, 2, n_eligible,                        num_fmt="#,##0", fill=total_fill, bold=True)
        self._cell(ws, r, 3, int(ri_rows["Running_Press_Count"].sum()),
                              num_fmt="#,##0", fill=total_fill, bold=True)
        self._cell(ws, r, 4, int(total_demand),                 num_fmt="#,##0", fill=total_fill, bold=True)
        self._cell(ws, r, 5, int(ri_rows["Total_GT_Per_Shift_Day0"].sum()),
                              num_fmt="#,##0", fill=total_fill, bold=True)
        ws.row_dimensions[r].height = 16
        r += 1

        # Note about Runner-Out
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        note_cell = ws.cell(
            row=r, column=1,
            value="ℹ  Runner-Out rows are NOT from the demand file. "
                  "They are presses currently curing a non-demanded SKU. "
                  "They appear in Day0 only for CO planning; their rows disappear "
                  "from each day sheet once the CO executes."
        )
        note_cell.font = Font(italic=True, size=9, color="555555")
        note_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        note_cell.fill = PatternFill("solid", fgColor="F5F5F5")
        ws.row_dimensions[r].height = 28
        r += 2

        # ── Section C: Demand Coverage over 31 days ───────────────────────────
        self._section_header(ws, r, 1, "C.  Demand Coverage — 31-Day Horizon", 6)
        r += 1

        total_gt_capacity = sum(
            df_d["Total_GT_Per_Shift_DayN"].sum() * SHIFTS_PER_DAY
            for df_d in daily_sheets
        )
        demand_left_day31 = int(daily_sheets[-1]["Updated_Demand_Qty"].sum())
        gt_produced       = int(total_demand) - demand_left_day31
        coverage_pct      = (gt_produced / total_demand * 100) if total_demand else 0

        cov_rows = [
            ("Total demand quantity (tyres)",          int(total_demand),     "#,##0"),
            ("Total GT capacity across 31 days",       int(total_gt_capacity),"#,##0"),
            ("Demand fulfilled by Day 31 (tyres)",     gt_produced,           "#,##0"),
            ("Demand remaining after Day 31 (tyres)",  demand_left_day31,     "#,##0"),
            ("Demand coverage %",                      round(coverage_pct, 1),"0.0\"%\""),
        ]
        for label, val, fmt in cov_rows:
            self._cell(ws, r, 1, label, align="left", fill=_LABEL_FILL)
            fill = _GOOD_FILL if label == "Demand coverage %" and coverage_pct >= 90 else _VAL_FILL
            self._cell(ws, r, 2, val, num_fmt=fmt, fill=fill, bold=(label == "Demand coverage %"))
            ws.row_dimensions[r].height = 16
            r += 1
        r += 1

        # ── Section D: CO Schedule Summary ────────────────────────────────────
        self._section_header(ws, r, 1, "D.  Changeover Schedule Summary", 6)
        r += 1

        from collections import Counter
        co_by_day = Counter(ev["day"] for ev in co_events)
        total_cos = len(co_events)

        self._cell(ws, r, 1, "Total changeover events",    align="left", fill=_LABEL_FILL)
        self._cell(ws, r, 2, total_cos, num_fmt="#,##0",   fill=_VAL_FILL, bold=True)
        ws.row_dimensions[r].height = 16
        r += 1
        self._cell(ws, r, 1, "Peak COs in a single day",  align="left", fill=_LABEL_FILL)
        self._cell(ws, r, 2, max(co_by_day.values()) if co_by_day else 0,
                   num_fmt="#,##0", fill=_VAL_FILL)
        ws.row_dimensions[r].height = 16
        r += 1

        # mini day table
        for ci, htext in enumerate(["Day", "COs", ""], start=1):
            c = ws.cell(row=r, column=ci, value=htext)
            for k, v in hdr.items():
                setattr(c, k, v)
        ws.row_dimensions[r].height = 18
        r += 1
        for day in sorted(co_by_day):
            self._cell(ws, r, 1, f"Day {day:02d}", fill=_TH_FILL)
            self._cell(ws, r, 2, co_by_day[day], num_fmt="#,##0", fill=_VAL_FILL)
            ws.row_dimensions[r].height = 15
            r += 1
        r += 1

        # ── Section E: Excluded SKUs ───────────────────────────────────────────
        self._section_header(ws, r, 1, f"E.  Excluded SKUs ({n_excluded}) — Reason for Skip", 6)
        r += 1

        if df_excluded.empty:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(row=r, column=1, value="None — all demand SKUs are eligible.").font = \
                Font(italic=True, color="555555")
            ws.row_dimensions[r].height = 16
        else:
            excl_cols = ["SKUCode", "Demand_Qty", "Priority_Score", "Remark"]
            for ci, htext in enumerate(excl_cols, start=1):
                c = ws.cell(row=r, column=ci, value=htext.replace("_", " "))
                for k, v in hdr.items():
                    setattr(c, k, v)
            ws.row_dimensions[r].height = 20
            r += 1
            for _, row in df_excluded.iterrows():
                self._cell(ws, r, 1, str(row.get("SKUCode", "")), align="left",
                           fill=_WARN_FILL)
                self._cell(ws, r, 2, int(row.get("Demand_Qty", 0)),
                           num_fmt="#,##0", fill=_WARN_FILL)
                self._cell(ws, r, 3, float(row.get("Priority_Score", 0)),
                           num_fmt="0.0000", fill=_WARN_FILL)
                self._cell(ws, r, 4, str(row.get("Remark", "")), align="left",
                           fill=_WARN_FILL)
                ws.row_dimensions[r].height = 16
                r += 1

            # total excluded demand
            r += 1
            excl_dem = int(df_excluded["Demand_Qty"].sum())
            self._cell(ws, r, 1, "Total demand in excluded SKUs", align="left",
                       fill=_WARN_FILL, bold=True)
            self._cell(ws, r, 2, excl_dem, num_fmt="#,##0", fill=_WARN_FILL, bold=True)
            excl_pct = excl_dem / (total_demand + excl_dem) * 100 if (total_demand + excl_dem) else 0
            self._cell(ws, r, 3, f"{excl_pct:.1f}% of gross demand",
                       align="left", fill=_WARN_FILL)

    def export(
        self,
        daily_sheets: list[pd.DataFrame],
        co_events: list[dict],
        df_day0: pd.DataFrame,
        df_excluded: pd.DataFrame,
        n_demand_raw: int,
        demand_path: str,
        planning_days: int,
        output_path: str,
    ):
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        # Sheet 1: Summary (new)
        ws_sum = wb.create_sheet("Summary")
        self._write_summary_sheet(
            ws_sum, df_day0, df_excluded, daily_sheets, co_events,
            n_demand_raw, demand_path, planning_days,
        )

        # Sheet 2: Day0 snapshot
        ws0 = wb.create_sheet("Day0_Summary")
        day0_cols = [
            "SKUCode", "Category", "Running_Press_Count", "MouldLife_min",
            "Effective_CT_Min", "Qty_Per_Press_Per_Shift", "Total_GT_Per_Shift_Day0",
            "Demand_Qty", "Priority_Score", "Skip_Reason",
        ]
        day0_cols_present = [c for c in day0_cols if c in df_day0.columns]

        # Title row for Day0 that explains the Runner-Out rows
        n_ri  = (df_day0["Category"] == "Runner-In").sum()
        n_ro  = (df_day0["Category"] == "Runner-Out").sum()
        n_nri = (df_day0["Category"] == "Non-Runner-In").sum()
        ws0.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=len(day0_cols_present))
        title = ws0.cell(
            row=1, column=1,
            value=(f"Day 0 Snapshot — {n_ri} Runner-In + {n_nri} Non-Runner-In "
                   f"(demand SKUs)  |  {n_ro} Runner-Out (non-demand, CO candidates only)")
        )
        title.font  = Font(bold=True, size=10, color=_WHITE)
        title.fill  = PatternFill("solid", fgColor=_NAVY)
        title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws0.row_dimensions[1].height = 20

        hdr = self._hdr_style()
        for c_idx, h in enumerate(day0_cols_present, start=1):
            cell = ws0.cell(row=2, column=c_idx, value=h.replace("_", " "))
            for k, v in hdr.items():
                setattr(cell, k, v)
        ws0.row_dimensions[2].height = 20

        for r_idx, (_, row) in enumerate(df_day0.iterrows(), start=3):
            cat = str(row.get("Category", ""))
            fill = self._CAT_FILL.get(cat)
            for c_idx, col in enumerate(day0_cols_present, start=1):
                cell = ws0.cell(row=r_idx, column=c_idx, value=row.get(col))
                cell.alignment = Alignment(horizontal="center")
                if fill:
                    cell.fill = fill
        for i, w in enumerate([16, 16, 18, 14, 15, 20, 22, 16, 14, 40], start=1):
            ws0.column_dimensions[get_column_letter(i)].width = w

        # CO Schedule sheet
        ws_co = wb.create_sheet("CO_Schedule")
        self._write_co_sheet(ws_co, co_events)

        # Day sheets (Day_01 … Day_31)
        for day_idx, df_day in enumerate(daily_sheets, start=1):
            ws = wb.create_sheet()
            self._write_day_sheet(ws, df_day, day_idx)

        # demand_drawdown — demand remaining after each day's curing (RI + NRI only)
        # Day 0 = opening available demand (excludes excluded NRI + Runner-Out)
        # Daily_Consumed = prev_remaining - curr_remaining
        ws_dd = wb.create_sheet("demand_drawdown")
        hdr_dd = self._hdr_style()
        for c_idx, label in enumerate(["Day", "Remaining_Demand", "Daily_Consumed"], start=1):
            cell = ws_dd.cell(row=1, column=c_idx, value=label)
            for k, v in hdr_dd.items():
                setattr(cell, k, v)
        ws_dd.row_dimensions[1].height = 20

        demand_cats = {"Runner-In", "Non-Runner-In"}
        # Day 0 opening: sum Demand_Qty for RI + NRI, exclude rows with Skip_Reason
        d0_mask = df_day0["Category"].isin(demand_cats)
        if "Skip_Reason" in df_day0.columns:
            d0_mask &= df_day0["Skip_Reason"].isna() | (df_day0["Skip_Reason"].astype(str).str.strip() == "")
        day0_remaining = int(df_day0.loc[d0_mask, "Demand_Qty"].sum())

        ws_dd.cell(row=2, column=1, value="Opening (Day 0)").alignment = Alignment(horizontal="center")
        ws_dd.cell(row=2, column=2, value=day0_remaining).alignment = Alignment(horizontal="center")
        ws_dd.cell(row=2, column=3, value="—").alignment = Alignment(horizontal="center")

        prev_remaining = day0_remaining
        total_consumed = 0
        for day_idx, df_day in enumerate(daily_sheets, start=1):
            dmask = df_day["Category"].isin(demand_cats)
            curr_remaining = int(df_day.loc[dmask, "Updated_Demand_Qty"].sum())
            consumed = prev_remaining - curr_remaining
            total_consumed += consumed
            r = day_idx + 2
            ws_dd.cell(row=r, column=1, value=f"Day {day_idx:02d}").alignment = Alignment(horizontal="center")
            ws_dd.cell(row=r, column=2, value=curr_remaining).alignment = Alignment(horizontal="center")
            ws_dd.cell(row=r, column=3, value=consumed).alignment = Alignment(horizontal="center")
            prev_remaining = curr_remaining

        total_row = len(daily_sheets) + 3
        for col, val in [(1, "Total Consumed"), (2, ""), (3, total_consumed)]:
            cell = ws_dd.cell(row=total_row, column=col, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws_dd.column_dimensions["A"].width = 18
        ws_dd.column_dimensions["B"].width = 20
        ws_dd.column_dimensions["C"].width = 18

        # curing_daily_cons — total curing production per day across all SKUs
        ws_dc = wb.create_sheet("curing_daily_cons")
        hdr = self._hdr_style()
        for c_idx, label in enumerate(["Day", "Total_Curing_Production"], start=1):
            cell = ws_dc.cell(row=1, column=c_idx, value=label)
            for k, v in hdr.items():
                setattr(cell, k, v)
        ws_dc.row_dimensions[1].height = 20
        grand_total = 0
        for day_idx, df_day in enumerate(daily_sheets, start=1):
            demand_mask = df_day["Category"].isin({"Runner-In", "Non-Runner-In"})
            daily_total = int(df_day.loc[demand_mask, "Total_GT_Per_Shift_DayN"].sum() * SHIFTS_PER_DAY)
            grand_total += daily_total
            ws_dc.cell(row=day_idx + 1, column=1, value=f"Day {day_idx:02d}").alignment = Alignment(horizontal="center")
            ws_dc.cell(row=day_idx + 1, column=2, value=daily_total).alignment = Alignment(horizontal="center")
        total_row = len(daily_sheets) + 2
        total_label = ws_dc.cell(row=total_row, column=1, value="Total")
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal="center")
        total_val = ws_dc.cell(row=total_row, column=2, value=grand_total)
        total_val.font = Font(bold=True)
        total_val.alignment = Alignment(horizontal="center")
        ws_dc.column_dimensions["A"].width = 12
        ws_dc.column_dimensions["B"].width = 24

        wb.save(output_path)
        print(f"  [Export] Saved → {output_path}")
        print(f"  [Export] Sheets: Summary + Day0_Summary + CO_Schedule + "
              f"{len(daily_sheets)} day sheets + demand_drawdown + curing_daily_cons")


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_dynamic_consumption(
    demand_path: str | None = None,
    output_path: str | None = None,
    plan_start: datetime = PLAN_START,
    planning_days: int = PLANNING_DAYS,
    max_co_per_day: int = MAX_CO_PER_DAY,
    buildable_rate: dict | None = None,
    sku_inch: dict | None = None,
    building_inch_capacity: dict | None = None,
    feed_ctx: dict | None = None,                # PERSKU_FEED: {sku_machines, machine_skus, machine_gtday}
    priority_deadline_map: dict | None = None,   # DELIVERY_PRIORITY: {sku: deadline_day} or None
    reactive_only: bool = False,                 # Part B: skip planned schedule + CC workbook, keep ETL
    initial_press_state: dict | None = None,     # MID-MONTH: {press_to_sku, mould_life, press_moulds}
) -> dict:
    """
    Build the 31-day dynamic curing consumption file.

    Returns dict with keys:
        daily_sheets, co_events, ct_map, df_day0
    """
    from connection import ConsumptionETL  # ETL lives in the DB layer (connection.py)
    if demand_path is None:
        # Auto-detect demand file in input dir
        for fname in sorted(os.listdir(IN_DIR), reverse=True):
            if fname.lower().endswith((".xlsx", ".csv")) and "demand" in fname.lower():
                demand_path = os.path.join(IN_DIR, fname)
                break
        if demand_path is None:
            candidates = [f for f in os.listdir(IN_DIR)
                          if f.lower().endswith((".xlsx", ".csv"))]
            if candidates:
                demand_path = os.path.join(IN_DIR, candidates[0])
            else:
                raise FileNotFoundError(f"No demand file found in {IN_DIR}")

    if output_path is None:
        output_path = DYNAMIC_CC_OUTPUT

    print("\n" + "=" * 70)
    print(f"  B2C Phase 0 Extended — {planning_days}-Day Dynamic Curing Consumption and CO plan")
    print("=" * 70)
    print(f"  Demand file : {os.path.basename(demand_path)}")
    print(f"  Plan start  : {plan_start.strftime('%d-%b-%Y')} | Days: {planning_days}")

    engine = bc_config.make_engine()
    etl = ConsumptionETL(engine)

    print("\n  [ETL] Loading demand …")
    df_demand = etl.load_demand(demand_path)
    n_demand_raw = len(df_demand)          # total before eligibility filter
    print(f"        {n_demand_raw} demanded SKUs")

    print("  [ETL] Loading cycle times …")
    df_ct = etl.load_cycle_times()

    print("  [ETL] Loading running moulds …")
    df_running = etl.load_running_moulds()
    print(f"        {len(df_running)} active press rows")
    if initial_press_state is not None:
        # STAGE 3 (mid-month): reseed the CO planner from a carried day-K press state instead of
        # the Day-0 DB snapshot, so the plan is built against the SAME press positions the day-by-day
        # simulation starts from. Overriding df_running here (its sole upstream source) propagates
        # through classification, RI/RO/NRI, Running_Press_Count and the mould gate automatically.
        _pts = initial_press_state.get("press_to_sku", {})
        _mlf = initial_press_state.get("mould_life", {})
        _pm  = initial_press_state.get("press_moulds", {})
        _rows = []
        for _p, _sku in _pts.items():
            _mn = list(_pm.get(_p, []))
            _rows.append({"Machine": str(_p), "SKUCode": str(_sku), "MouldNos": _mn,
                          "MouldLife_remaining": int(_mlf.get(_p, 3000)), "Num_Moulds": len(_mn)})
        df_running = pd.DataFrame(
            _rows, columns=["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"])
        print(f"        [midmonth] CO planner reseeded from carried day-K state: "
              f"{len(df_running)} presses")

    print("  [ETL] Loading curing allowable machines …")
    df_allowable = etl.load_curing_allowable()
    print(f"        {len(df_allowable)} SKUs with allowable presses")

    # Eligibility filter
    print("  [ETL] Loading eligibility sources …")
    bld_master  = etl.load_building_allowable_skus()
    bld_history = etl.load_building_history_skus()
    cur_master  = etl.load_curing_allowable_skus()
    cur_history = etl.load_curing_history_skus()
    filt = SKUEligibilityFilter()
    df_demand, df_excluded = filt.filter(
        df_demand, bld_master, bld_history, cur_master, cur_history
    )
    print(f"  [Eligible] {len(df_demand)} SKUs pass | {len(df_excluded)} excluded")

    # Classify & resolve CT
    classifier = SKUClassifier()
    df_classify = classifier.classify(df_demand, df_running)
    ri  = (df_classify["Category"] == "Runner-In").sum()
    nri = (df_classify["Category"] == "Non-Runner-In").sum()
    print(f"  [Classify] Runner-In: {ri} | Non-Runner-In: {nri}")

    ct_resolver = CycleTimeResolver()
    ct_map = ct_resolver.resolve(df_classify["SKUCode"].tolist(), df_ct)
    n_default = sum(1 for v in ct_map.values()
                    if v == ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
    print(f"  [CT] {len(ct_map)} SKUs | {n_default} using default {ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN} min")

    # Build Day 0 consumption table (same as curing_consumption.py output)
    calc = ConsumptionCalculator()
    df_day0 = calc.compute(df_classify, ct_map, df_demand, plan_start, planning_days)
    df_day0["Skip_Reason"] = ""   # eligible demand SKUs — no skip

    # Append excluded demand SKUs back into Day0 for display (with Skip_Reason).
    # They are kept at zero press count; scheduling still uses only eligible SKUs.
    if not df_excluded.empty:
        excl_rows = []
        for _, row in df_excluded.iterrows():
            sku = str(row["SKUCode"])
            ct  = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
            qps = _qty_per_press_per_shift(ct)
            excl_rows.append({
                "SKUCode":                  sku,
                "Category":                 "Non-Runner-In",
                "Running_Press_Count":       0,
                "MouldLife_min":             0,
                "Effective_CT_Min":          ct,
                "Qty_Per_Press_Per_Shift":   qps,
                "Total_GT_Per_Shift_Day0":   0,
                "Demand_Qty":                int(row.get("Demand_Qty", 0)),
                "Priority_Score":            float(row.get("Priority_Score", 0)),
                "Skip_Reason":               str(row.get("Remark", "Not eligible")),
            })
        df_day0 = pd.concat([df_day0, pd.DataFrame(excl_rows)], ignore_index=True)
        print(f"  [Day0] Re-added {len(excl_rows)} excluded SKUs with Skip_Reason")

    # Include Runner-Out in Day 0 table for the dynamic file.
    # Runner-Out = presses currently running a NON-demand SKU.
    # Group df_running by SKU to get press count (one row per machine).
    demand_sku_set = set(df_classify["SKUCode"])
    df_ro_running = df_running[~df_running["SKUCode"].isin(demand_sku_set)].copy()

    if not df_ro_running.empty:
        ro_grouped = (
            df_ro_running.groupby("SKUCode")
            .agg(
                RunningPressCount=("Machine", "count"),
                MouldLife_min=("MouldLife_remaining", "min"),
            )
            .reset_index()
        )
        ct_map_ro = ct_resolver.resolve(ro_grouped["SKUCode"].tolist(), df_ct)
        ct_map.update(ct_map_ro)
        demand_lookup_all = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Quantity"]))
        priority_lookup_all = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Priority"]))
        ro_rows = []
        for _, r in ro_grouped.iterrows():
            sku = str(r["SKUCode"])
            ct  = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
            qps = _qty_per_press_per_shift(ct)
            n   = int(r["RunningPressCount"])
            ro_rows.append({
                "SKUCode":                  sku,
                "Category":                 "Runner-Out",
                "Running_Press_Count":       n,
                "MouldLife_min":             int(r["MouldLife_min"]),
                "Effective_CT_Min":          ct,
                "Qty_Per_Press_Per_Shift":   qps,
                "Total_GT_Per_Shift_Day0":   n * qps,
                "Demand_Qty":                demand_lookup_all.get(sku, 0),
                "Priority_Score":            priority_lookup_all.get(sku, 0),
                "Skip_Reason":               "Non-demand SKU — CO candidate",
            })
        df_day0 = pd.concat([df_day0, pd.DataFrame(ro_rows)], ignore_index=True)
        print(f"  [Day0] Added {len(ro_rows)} Runner-Out SKUs | "
              f"Total rows: {len(df_day0)} "
              f"({n_demand_raw} demand + {len(ro_rows)} non-demand CO candidates)")

    # Part B (REACTIVE_ONLY): the single reactive CO arbiter replaces the planned schedule
    # entirely — skip Pass 1 (COScheduler), Pass 2 (DaySimulator) and the
    # curing_consumption_*.xlsx export. df_day0 + ct_map (+ df_excluded) are already built
    # by the Day-0 ETL/classify above and are all the rolling pipeline needs.
    if reactive_only:
        print("  [reactive_only] planned CO schedule + consumption workbook SKIPPED")
        return {
            "daily_sheets": [], "co_events": [],
            "ct_map": ct_map, "df_day0": df_day0, "df_excluded": df_excluded,
        }

    # Pass 1: CO schedule
    print("\n  [Pass 1] Computing CO schedule …")
    scheduler = COScheduler()
    # #4: mould eligibility for the Phase-0 mould-aware gate (same source as the rolling gate).
    # Also loaded when DELIVERY_PRIORITY is active (the mould-pair cap needs it) even if the
    # P0 gate itself is off.
    _p0_sku_moulds = None
    if _P0_MOULD_GATE or (priority_deadline_map and _DP_MOULDCAP):
        try:
            _p0_sku_moulds = {k: set(v) for k, v in
                              etl.load_mould_eligibility()["sku_moulds"].items()}
        except Exception as _e:
            print(f"  [Phase-0] mould eligibility load failed ({_e}); P0 mould gate disabled")
            _p0_sku_moulds = None
    co_events = scheduler.schedule(
        df_day0, df_demand, df_allowable, df_running, ct_map, max_co_per_day,
        planning_days=planning_days,
        buildable_rate=buildable_rate,
        sku_inch=sku_inch,
        building_inch_capacity=building_inch_capacity,
        feed_ctx=feed_ctx,
        sku_moulds=_p0_sku_moulds,
        priority_deadline_map=priority_deadline_map,
    )

    # Pass 2: 31-day simulation
    print(f"\n  [Pass 2] Simulating {planning_days} days …")
    simulator = DaySimulator()
    daily_sheets = simulator.simulate(
        df_day0, df_demand, df_allowable, ct_map, co_events,
        planning_days=planning_days,
    )

    # Print day-by-day summary
    print(f"\n  {'Day':<6} {'RI Presses':>11} {'NRI Presses':>12} "
          f"{'Total GT/Shift':>15} {'Demand Left':>12}")
    print("  " + "-" * 60)
    for d_idx, df_d in enumerate(daily_sheets, start=1):
        ri_presses  = df_d.loc[df_d["Category"] == "Runner-In",  "Running_Press_Count"].sum()
        nri_presses = df_d.loc[df_d["Category"] == "Non-Runner-In", "Running_Press_Count"].sum()
        total_gt    = df_d["Total_GT_Per_Shift_DayN"].sum()
        dem_left    = df_d["Updated_Demand_Qty"].sum()
        print(f"  Day {d_idx:02d} {int(ri_presses):>11,} {int(nri_presses):>12,} "
              f"{int(total_gt):>15,} {int(dem_left):>12,}")

    # Export
    print(f"\n  [Export] Writing Excel …")
    exporter = DynamicExporter()
    exporter.export(
        daily_sheets=daily_sheets,
        co_events=co_events,
        df_day0=df_day0,
        df_excluded=df_excluded,
        n_demand_raw=n_demand_raw,
        demand_path=demand_path,
        planning_days=planning_days,
        output_path=output_path,
    )

    return {
        "daily_sheets": daily_sheets,
        "co_events":    co_events,
        "ct_map":       ct_map,
        "df_day0":      df_day0,
        "df_excluded":  df_excluded,
        "campaign_target": getattr(scheduler, "campaign_target", {}),
    }


if __name__ == "__main__":
    import glob

    # Usage: python curing_consumption_dynamic.py [demand_file_path]
    # If no argument given, auto-detect from data/input/.
    if len(sys.argv) > 1:
        demand_path = sys.argv[1]
        if not os.path.exists(demand_path):
            raise SystemExit(f"Demand file not found: {demand_path}")
    else:
        # Auto-pick demand file.
        # Preference order: normalized XLSX with "demand" in name > other demand XLSX
        # > demand CSV > any XLSX > any CSV.  Within each tier, sort descending by
        # name so date-stamped files resolve to the most recent.
        # Exclude files whose name contains "BACKUP" or "backup".
        def _is_backup(p: str) -> bool:
            return "backup" in os.path.basename(p).lower()

        tiers = [
            [p for p in glob.glob(os.path.join(IN_DIR, "*may*.xlsx"))
             if not _is_backup(p)],
            [p for p in glob.glob(os.path.join(IN_DIR, "*demand*normalized*.xlsx"))
             if not _is_backup(p)],
            [p for p in glob.glob(os.path.join(IN_DIR, "*demand*.xlsx"))
             if not _is_backup(p)],
            [p for p in glob.glob(os.path.join(IN_DIR, "*demand*.csv"))
             if not _is_backup(p)],
            [p for p in glob.glob(os.path.join(IN_DIR, "*.xlsx"))
             if not _is_backup(p)],
            [p for p in glob.glob(os.path.join(IN_DIR, "*.csv"))
             if not _is_backup(p)],
        ]
        demand_path = None
        for tier in tiers:
            if tier:
                demand_path = sorted(tier, reverse=True)[0]
                break
        if demand_path is None:
            raise SystemExit(f"No demand file found in {IN_DIR}. "
                             "Place a demand .xlsx/.csv there and re-run.")

    run_dynamic_consumption(
        demand_path=demand_path,
        output_path=DYNAMIC_CC_OUTPUT,
    )
