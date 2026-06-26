
# """
# JK Tyre BTP — PCR Building Hybrid GA + LP Scheduler v8  (INDEPENDENT)
# =================================================================
# Standalone — no imports from buildingSchedule5/6/7.

# Same hybrid GA + LP architecture as v7, tuned to LIFT GT production
# qty closer to curing demand:

#   • Larger end-of-day GT inventory headroom (10 000 vs v7's 7 500) so
#     the LP isn't forced to stop building once it has just covered the
#     day's curing.
#   • Shorter MIN_CAMPAIGN_MINS (45 vs 60) so the GA can pack more SKUs
#     into the same machine-shift without idle slack.
#   • Stronger per-unit production weight in the LP objective so HiGHS
#     pushes harder on filling capacity once WIP is met.
#   • Larger GA effort (more generations + population) to converge on
#     higher-fill chromosomes that the new LP rewards.

# Pipeline per day (day-by-day rolling) — unchanged from v7:
#   1. Smart inch-lock for Stage1 + Unistage (coverage-then-balance).
#   2. GA outer loop evolves binary y[s,m] (SKU → machine assignment),
#      seeded from the 3-month history map.
#   3. For each chromosome, LP solves continuous x[s,m,t] minutes.
#   4. Fitness = LP obj + CO penalty − distinct-SKU reward.
#   5. Best chromosome's LP solution → allocation → sequence → schedule.
#   6. Roll GT inv, carcass inv, running machines into next day.
# """

import math
import os
import random
import warnings
from collections import defaultdict
from datetime import datetime, timedelta
from itertools import permutations

import numpy as np
import pandas as pd
from scipy.optimize import linprog
from scipy.sparse import lil_matrix, vstack as sp_vstack
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from sqlalchemy import create_engine
except ImportError:
    create_engine = None

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════
# All tunables for a run. Plant-specific knobs (machine groupings,
# cycle times, shift definitions, caps) live here as class attributes.
# DB credentials come from cbc_env (.env) — never hardcoded. If the helper or
# .env is missing, fail loudly rather than fall back to a baked-in secret.
from cbc_env import db_config as _cbc_db_config
_DB = _cbc_db_config()


class Config:
    DB_SERVER   = _DB["host"]
    DB_NAME     = _DB["database"]
    DB_USER     = _DB["user"]
    DB_PASSWORD = _DB["password"]

    # Curing plan that building schedules against. Under CBC this is set to the
    # Phase-C feed-aware bridge file at runtime (see cbc.py / load_curing_schedule).
    CURING_PLAN_FILE = None

    PLANNING_DAYS    = 8
    SHIFTS_PER_DAY   = 3
    HOURS_PER_SHIFT  = 8
    SHIFT_START_HOUR = 7
    SHIFT_MINS       = HOURS_PER_SHIFT * 60    # 480

    # ── Campaign / lock rules ─────────────────────────────────────────
    LOCKED_MIN_SHIFT0_MINS  = 240
    MIN_CAMPAIGN_MINS       = 45       # shorter campaigns → more SKUs/machine, higher fill
    MIN_CAMPAIGN_UNITS      = 40
    BUFFER_SHIFTS           = 1
    URGENT_THRESHOLD_HOURS  = 16.0
    # LP per-SKU production cap = net_demand × (1 + OVERBUILD_BUFFER_FRAC).
    # 0.0 was causing LP cap to collapse to 0 on Days 2+ whenever TopUp or
    # prior-day carry-over partially covered the lead-window demand.
    # 0.2 gives a 20% headroom so the LP stays active even when some WIP already
    # exists, without violating the hard "total build ≤ 30-day demand" ceiling
    # enforced by the TopUp target (gt_topup_target is already capped to unmet demand).
    OVERBUILD_BUFFER_FRAC   = 0.2

    # ── Build-ahead lead time ─────────────────────────────────────────
    # Building targets curing demand this many shifts ahead of today.
    # 3 shifts = 1 full day: building on day D covers day D+1 curing,
    # so finished GTs sit in WIP for ~1 day before curing consumes them.
    # This eliminates Day-2+ LP cap collapse (WIP from previous day is
    # earmarked for today's curing, so the LP sees net-zero inventory
    # against the lead window and builds the full tomorrow's demand).
    # Set to 0 to revert to same-day matching (original behaviour).
    BUILD_LEAD_SHIFTS = 3

    # ── Hard caps used by the LP and reports ─────────────────────────
    QTY_MAX_OVER_DEMAND    = 5000
    GT_QTY_CAP             = 70000
    CO_BUDGET_GT           = 99999
    CO_BUDGET_S1           = 99999
    MAX_GT_INVENTORY       = 10000    # end-of-day GT inventory cap (more headroom → higher GT qty)

    # ── TopUp pre-build horizon (bounded by shelf life) ──────────────
    # Idle machine tails only pre-build demand due within this many days
    # AHEAD (net of inventory + today's build). These are HARD shelf-life
    # limits — a unit cannot be built earlier than its max age before use:
    #   • GT (green tyre): max age 3 days  → build at most 3 days pre-cure
    #   • Carcass (Stage1): max age 1 day  → build at most 1 day pre-use
    # Capping per-SKU cover to these windows also guarantees no on-hand
    # unit ever ages past its shelf life (FIFO: inv ≤ N-day demand ⇒ age ≤ N).
    #
    # LP now uses cap_on_gross=True, so the LP cap = gross×buf regardless of
    # rolling WIP. TopUp and LP no longer compete for the same demand window —
    # LP covers Day+1 and TopUp builds ahead for Days+2,+3,+4 in idle machine
    # tails, reducing late-horizon starvation risk.
    TOPUP_LOOKAHEAD_DAYS_GT      = 3
    TOPUP_LOOKAHEAD_DAYS_CARCASS = 1

    # ── Per-day SKU diversity reporting target ───────────────────────
    DAILY_SKU_TARGET       = 48

    # ── Machine groupings ────────────────────────────────────────────
    STAGE1 = frozenset({
        "6801","6802","6803","6909","6911","7601","7701",
        "7801","7802","7803","7804","8001","8002","8003","8101"
    })
    STAGE2 = frozenset({"8201","8301","8302","8501","8502","7301"})
    UNISTAGE = frozenset({
        "7001","7002","7003","7004","6001","6002","6003","6004",
        "7101","7102","7103","7104","7105","7106",
        "7201","7501","7502","7503",
    })

    # Authoritative source: Master_Building_Machine_Design_cycleTime in DB.
    # building_b2c.py loads from DB at runtime and overrides this dict.
    _CT_SEC = {
        "7001":57.6,  "7002":57.6,  "7003":57.6,  "7004":57.6,
        "6001":60.0,  "6002":60.0,  "6003":60.0,  "6004":60.0,
        "7101":102.0, "7102":102.0, "7103":78.0,  "7104":108.0,
        "7105":108.0, "7106":57.6,  "7201":66.0,
        "7501":108.0, "7502":108.0, "7503":108.0,
        "8201":72.0,  "8301":78.0,  "8302":78.0,
        "8501":108.0, "8502":120.0, "7301":90.0,
        "6801":150,   "6802":218,   "6803":262,
        "6909":187,   "6911":150,   "7601":253,
        "7701":267,   "7801":163,   "7802":182,
        "7803":261,   "7804":257,   "8001":114,
        "8002":169,   "8003":113,   "8101":300,
    }

    @classmethod
    def ct_min(cls, machine: str) -> float:
        return cls._CT_SEC.get(str(machine), 120) / 60.0

    @classmethod
    def total_shifts(cls) -> int:
        return cls.PLANNING_DAYS * cls.SHIFTS_PER_DAY

    PLAN_DATE   = datetime(2026, 5, 31, 7, 0, 0)
    OUTPUT_FILE = (
        f"BTP_PCR_Building_Hybrid_v8_{PLAN_DATE.date()}_{PLANNING_DAYS}Days.xlsx"
    )

    # ── v8 LP weights ────────────────────────────────────────────────
    # Per-unit production reward inside the LP objective. Higher values
    # push HiGHS to fill capacity once WIP / starvation is satisfied.
    # v7 implicitly used 1.0 / ct (i.e. UNIT_BONUS_MULT = 1.0).
    UNIT_BONUS_MULT        = 3.0       # 3× v7 — production wins more vs slack noise


# ══════════════════════════════════════════════════════════════════════
# SHIFT HELPERS 
# ══════════════════════════════════════════════════════════════════════
def _shift_fn(dt: datetime):
    h    = dt.hour
    base = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    sh   = Config.SHIFT_START_HOUR
    if sh <= h < sh + 8:
        return "A", base + timedelta(hours=sh + 8)
    elif sh + 8 <= h < sh + 16:
        return "B", base + timedelta(hours=sh + 16)
    else:
        return "C", (
            base + timedelta(days=1, hours=sh)
            if h >= sh + 16
            else base + timedelta(hours=sh)
        )


def _shift_start(plan_start: datetime, shift_idx: int) -> datetime:
    day_offset   = shift_idx // Config.SHIFTS_PER_DAY
    shift_in_day = shift_idx % Config.SHIFTS_PER_DAY
    base = plan_start + timedelta(days=day_offset)
    return base + timedelta(hours=shift_in_day * Config.HOURS_PER_SHIFT)


def is_urgent(sku, urgency_map, plan_start):
    first = urgency_map.get(sku)
    if first is None:
        return False
    return ((first - plan_start).total_seconds() / 3600.0
            < Config.URGENT_THRESHOLD_HOURS)


# ══════════════════════════════════════════════════════════════════════
# ETL
# ══════════════════════════════════════════════════════════════════════
# Loads every input the scheduler needs: curing demand, GT/carcass
# inventory, SKU→eligible-machines matrix, changeover times, SKU→size,
# current running-machine snapshot, and a 3-month history map that
# seeds the GA so random chromosomes resemble real plant practice.
class ETL:
    S1_NAME_MAP = {
        "midland4stage1":"7804","midland2stage1":"7802","bj2stage1":"6802",
        "bj8stage1":"7201","sai3stage1":"8003","bj7stage1":"7104",
        "bj9stage1":"7105","bj3stage1":"6803","bj4stage1":"7101",
        "bj5stage1":"7102","88d1stage1":"8101","ltmstage1":"7601",
        "midland5stage1":"7701","midland3stage1":"7803","bj10stage1":"7106",
        "sai1stage1":"8001","sai2stage1":"8002","midland1stage1":"7801",
        "nrm11stage1":"6911","nrm9stage1":"6909","bj1stage1":"6801",
        "bj6stage1":"7103",
    }
    S2_NAME_MAP = {
        "bj8":"7201","bj7":"7104","bj9":"7105","vmi1":"8501","vmi2":"8502",
        "bj4":"7101","bj5":"7102","newirm":"7301","bj6":"7103","oldirm":"8201",
        "vmi2Maxx":"7002","gtic1":"8301","vmi3Maxx":"7003","gtic2":"8302",
        "us1":"7501","us2":"7502","bj10":"7106","us3":"7503",
        "vmi4Maxx":"7004","vmi1Maxx":"7001",
        "VMIExxium01":"6001","VMIExxium02":"6002",
        "VMIExxium03":"6003","VMIExxium04":"6004",
    }

    def __init__(self, engine=None):
        self.engine = engine

    def _sql(self, q):
        return pd.read_sql(q, self.engine)

    def load_curing_schedule(self):
        """Read the curing plan building schedules against.

        Source = Config.CURING_PLAN_FILE (set by cbc.py to the Phase-C feed-aware
        bridge). Handles both shapes:
          • a flat file (csv/xlsx) with SKUCode/StartTime/EndTime/Qty columns, and
          • a curing-scheduler workbook with a 'Shift Schedule' sheet (title rows
            above the header, e.g. the provided jkt_plan.xlsx).
        Falls back to the legacy jkt_plan.csv path if nothing is configured.
        """
        src = getattr(Config, "CURING_PLAN_FILE", None) or \
            r'/Users/ajaygour/Downloads/jkt_plan.csv'

        def _has_cols(d):
            cols = {str(c).strip().lower() for c in d.columns}
            return {"skucode", "starttime", "endtime"}.issubset(cols)

        df = None
        if str(src).lower().endswith((".xlsx", ".xls", ".xlsm")):
            xl = pd.ExcelFile(src)
            # Prefer a 'Shift Schedule' sheet; else the first sheet that parses.
            sheets = (["Shift Schedule"] if "Shift Schedule" in xl.sheet_names
                      else []) + list(xl.sheet_names)
            for sh in sheets:
                for hdr in (0, 1, 2, 3):
                    try:
                        cand = pd.read_excel(src, sheet_name=sh, header=hdr)
                    except Exception:
                        continue
                    if _has_cols(cand):
                        df = cand
                        break
                if df is not None:
                    break
            if df is None:
                raise ValueError(
                    f"No sheet with SKUCode/StartTime/EndTime found in {src}")
        else:
            df = pd.read_csv(src)

        # Normalise lowercase/variant column names to the canonical ones.
        canon = {"skucode": "SKUCode", "starttime": "StartTime",
                 "endtime": "EndTime", "qty": "Qty"}
        df = df.rename(columns={c: canon[str(c).strip().lower()]
                                for c in df.columns
                                if str(c).strip().lower() in canon})
        df["StartTime"] = pd.to_datetime(df["StartTime"])
        df["EndTime"]   = pd.to_datetime(df["EndTime"])
        df["SKUCode"]   = df["SKUCode"].astype(str)
        df["Qty"]       = pd.to_numeric(df.get("Qty", 0), errors="coerce").fillna(0)
        # Drop changeover/cleaning placeholder rows and zero-demand rows.
        df = df[(df["SKUCode"] != "CHANGEOVER") & (df["Qty"] > 0)].copy()
        print(f"  [Curing plan] {len(df)} rows from {os.path.basename(str(src))}")
        return df

    def load_gt_inventory(self):
        # df = pd.read_csv(r"/Users/ajaygour/Downloads/BTP_3April_LP.csv")
        # df["StartTime"] = pd.to_datetime(df["StartTime"])
        # df["EndTime"]   = pd.to_datetime(df["EndTime"])
        # return df
        return self._sql(
            f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory "
            f"FROM {Config.DB_NAME}.gt_inventory_manual"
        )

    def load_carcass_inventory(self):
        try:
            return self._sql(
                f"SELECT sizeCode AS SKUCode, CarcassInv AS Carcass_Inventory "
                f"FROM {Config.DB_NAME}.carcass_inventory_manual"
            )
        except Exception:
            return pd.DataFrame(columns=["SKUCode","Carcass_Inventory"])

    def load_machine_allowable(self):
        df = self._sql(
            f"SELECT * FROM {Config.DB_NAME}.Master_Building_Allowable_Machines_source"
        )
        mcols = [c for c in df.columns if str(c).isdigit()]
        df["Machines"] = df.apply(
            lambda r: [str(int(c)) for c in mcols
                       if str(r[c]).strip().upper() in
                       {"Y","YES","1","ONLY ONE M/C RUN"}],
            axis=1,
        )
        return df.rename(columns={"SKU Code":"SKUCode"})[["SKUCode","Machines"]]

    def load_changeover_map(self):
        df = self._sql(f"SELECT * FROM {Config.DB_NAME}.Master_Building_ChangeoverTime")
        co_map = {}
        for _, r in df.iterrows():
            co_map[str(r["MachineCode"])] = {
                "same": float(r["Same Size(Minutes)"]) + 10,
                "diff": float(r["Different Size(Minutes)"]) + 10,
            }
        return co_map

    def load_sku_sizes(self):
        df = self._sql(
            f"SELECT SKUCode, Size "
            f"FROM {Config.DB_NAME}.Master_Curing_Allowable_Machines"
        )
        return dict(zip(df["SKUCode"].astype(str), df["Size"].astype(str)))

    def load_history_map(self):
        """3-month run counts → {(machine, sku): count}.

        Source: DB tables Building_Stage1_Best_Machines + Building_Stage2_Best_Machines
        (cols: sizeCode, MachineName, count, MachineNo). Falls back to the legacy
        master_building_stage1/2_best_machine.csv files if the DB is unavailable.
        """
        frames = []
        if self.engine is not None:
            for table in ("Building_Stage1_Best_Machines",
                          "Building_Stage2_Best_Machines"):
                try:
                    frames.append(self._sql(
                        f"SELECT MachineNo, sizeCode, count "
                        f"FROM {Config.DB_NAME}.{table}"))
                except Exception as e:  # noqa: BLE001
                    print(f"  ⚠️  history table {table}: {e}")
        if not frames:
            here = os.path.dirname(os.path.abspath(__file__))
            for fp in (os.path.join(here, "master_building_stage1_best_machine.csv"),
                       os.path.join(here, "master_building_stage2_best_machine.csv")):
                if os.path.exists(fp):
                    try:
                        frames.append(pd.read_csv(fp))
                    except Exception as e:  # noqa: BLE001
                        print(f"  ⚠️  {fp}: {e}")
                else:
                    print(f"  ⚠️  history source missing (DB + file): {fp}")

        hist = {}
        for df in frames:
            for _, r in df.iterrows():
                try:
                    m = str(int(r["MachineNo"]))
                    s = str(r["sizeCode"])
                    c = float(r["count"])
                except (ValueError, TypeError, KeyError):
                    continue
                if not m or not s or c <= 0:
                    continue
                hist[(m, s)] = hist.get((m, s), 0.0) + c
        print(f"  [History] Loaded {len(hist)} (machine, SKU) pairs.")
        return hist

    def load_running_machines(self):
        rows = []
        for table, name_map in [
            ("TBMStage1_ProductionEventData", self.S1_NAME_MAP),
            ("TBMStage2_ProductionEventData", self.S2_NAME_MAP),
        ]:
            try:
                df = self._sql(
                    f"SELECT WorkCenter, RecipeCode "
                    f"FROM {Config.DB_NAME}.{table} "
                    f"ORDER BY DtAndTime DESC"
                )
                df = df.drop_duplicates(subset=["WorkCenter"])
                for _, r in df.iterrows():
                    mid = name_map.get(str(r["WorkCenter"]))
                    if mid:
                        rows.append({"Machine":str(mid),"SKUCode":str(r["RecipeCode"])})
            except Exception as e:
                print(f"  ⚠️  {table}: {e}")
        return (
            pd.DataFrame(rows) if rows
            else pd.DataFrame(columns=["Machine","SKUCode"])
        )


# ══════════════════════════════════════════════════════════════════════
# DEMAND DERIVER
# ══════════════════════════════════════════════════════════════════════
# Converts raw curing demand rows into a per-(SKU, shift) matrix and a
# net-demand-per-SKU table (curing demand − opening GT inventory).
class DemandDeriver:
    def derive(self, df_curing, df_gt_inv, plan_start):
        gt_inv_map = dict(zip(df_gt_inv["SKUCode"], df_gt_inv["GT_Inventory"]))
        exclude = {"CHANGEOVER","MOULD_CLEAN","C/O","CLEANING"}
        prod = df_curing[~df_curing["SKUCode"].isin(exclude)].copy()
        if "Shift" not in prod.columns:
            prod["Shift"] = prod["StartTime"].apply(lambda dt: _shift_fn(dt)[0])
        if "Date" not in prod.columns:
            prod["Date"] = prod["StartTime"].dt.date

        T = Config.total_shifts()
        shift_starts = [_shift_start(plan_start, t) for t in range(T)]
        shift_ends = [
            _shift_start(plan_start, t) + timedelta(hours=Config.HOURS_PER_SHIFT)
            for t in range(T)
        ]
        plan_end = plan_start + timedelta(days=Config.PLANNING_DAYS)

        sku_shift_demand = defaultdict(float)
        for _, row in prod.iterrows():
            st, en, qty = row["StartTime"], row["EndTime"], float(row["Qty"])
            for t in range(T):
                overlap_start = max(st, shift_starts[t])
                overlap_end   = min(en, shift_ends[t])
                if overlap_end > overlap_start and (en - st).total_seconds() > 0:
                    frac = ((overlap_end - overlap_start).total_seconds() /
                            (en - st).total_seconds())
                    sku_shift_demand[(row["SKUCode"], t)] += qty * frac

        all_skus = sorted(set(k[0] for k in sku_shift_demand.keys()))
        S = len(all_skus)
        sku_idx = {s: i for i, s in enumerate(all_skus)}
        curing_matrix = np.zeros((S, T))
        for (sku, t), qty in sku_shift_demand.items():
            curing_matrix[sku_idx[sku], t] = qty

        total_gt = curing_matrix.sum(axis=1)
        rows = []
        for si, sku in enumerate(all_skus):
            gt_inv  = gt_inv_map.get(sku, 0)
            net_dem = max(0, int(total_gt[si]) - int(gt_inv))
            first_t = next((t for t in range(T) if curing_matrix[si, t] > 0), T)
            first_start = shift_starts[first_t] if first_t < T else plan_end
            rows.append({
                "SKUCode":             sku,
                "GT_Demand":           int(total_gt[si]),
                "GT_Inventory":        int(gt_inv),
                "Net_GT_Demand":       net_dem,
                "LP_Demand":           net_dem,
                "First_Curing_Start":  first_start,
                "Burn_Rate_Per_Shift": round(total_gt[si] / T, 1),
                "Active_Shifts":       int((curing_matrix[si] > 0).sum()),
            })
        df_sku_demand = pd.DataFrame(rows).sort_values("First_Curing_Start")

        sd_rows = []
        for (sku, t), qty in sku_shift_demand.items():
            if qty > 0:
                s_start = shift_starts[t]
                shift_lbl, _ = _shift_fn(s_start)
                sd_rows.append({
                    "Date": s_start.date(), "Shift": shift_lbl,
                    "SKUCode": sku, "Curing_Qty": int(qty),
                })
        df_shift_demand = pd.DataFrame(sd_rows)
        print(f"  [Demand] SKUs: {S} | Shifts: {T} | "
              f"Total curing: {total_gt.sum():,.0f}")
        return df_sku_demand, curing_matrix, all_skus, df_shift_demand


# ══════════════════════════════════════════════════════════════════════
# CAMPAIGN SEQUENCER
# ══════════════════════════════════════════════════════════════════════
# Given the LP's (SKU, machine) allocation, decides the ORDER in which
# SKUs run on each machine so that changeovers between consecutive SKUs
# are minimised (same-size preferred, urgent SKUs bumped first).
class CampaignSequencer:
    def __init__(self, co_map, sku_to_size, locked_skus, urgency_map, plan_start):
        self.co_map      = co_map
        self.sku_to_size = sku_to_size
        self.locked_skus = locked_skus
        self.urgency_map = urgency_map
        self.plan_start  = plan_start

    def _co_cost(self, machine, sku_from, sku_to):
        if sku_from is None or sku_from == sku_to:
            return 0.0
        sz_f = self.sku_to_size.get(str(sku_from))
        sz_t = self.sku_to_size.get(str(sku_to))
        restr = (machine in Config.STAGE1) or (machine in Config.UNISTAGE)
        if restr and sz_f != sz_t:
            return float("inf")
        co = self.co_map.get(str(machine), {"same":40,"diff":60})
        return co["same"] if sz_f == sz_t else co["diff"]

    def _total_co(self, machine, first_sku, order):
        cost = self._co_cost(machine, first_sku, order[0])
        for i in range(len(order) - 1):
            c = self._co_cost(machine, order[i], order[i+1])
            if c == float("inf"):
                return float("inf")
            cost += c
        return cost

    def _best_order(self, machine, first_sku, skus):
        if not skus:
            return []
        if len(skus) == 1:
            return list(skus)
        if len(skus) <= 6:
            best, best_cost = list(skus), float("inf")
            for perm in permutations(skus):
                cost = self._total_co(machine, first_sku, list(perm))
                if cost < best_cost:
                    best_cost, best = cost, list(perm)
            return best
        remaining, order, cur = list(skus), [], first_sku
        while remaining:
            nxt = min(remaining, key=lambda s: self._co_cost(machine, cur, s))
            order.append(nxt); remaining.remove(nxt); cur = nxt
        return order

    def _urgency_key(self, sku):
        return self.urgency_map.get(sku, datetime(2099, 1, 1))

    def sequence(self, df_alloc):
        if df_alloc.empty:
            return {}
        result, total_cos = {}, 0
        for mach, grp in df_alloc.groupby("Machine"):
            grp = grp.sort_values("ShiftIdx").reset_index(drop=True)
            by_sku = {}
            for _, row in grp.iterrows():
                sku = row["SKUCode"]
                if sku in by_sku:
                    by_sku[sku]["units"]     += row["Units"]
                    by_sku[sku]["mins_used"] += row["Mins_Used"]
                    by_sku[sku]["first_shift"] = min(
                        by_sku[sku]["first_shift"], int(row["ShiftIdx"])
                    )
                else:
                    by_sku[sku] = {
                        "sku": sku, "units": row["Units"],
                        "mins_used": row["Mins_Used"], "ct": row["CT_Min"],
                        "first_shift": int(row["ShiftIdx"]),
                    }
            merged = sorted(by_sku.values(), key=lambda c: c["first_shift"])
            first_sku = self.locked_skus.get(str(mach))
            is_restr = (mach in Config.STAGE1) or (mach in Config.UNISTAGE)
            if is_restr:
                if first_sku:
                    lock_sz = self.sku_to_size.get(str(first_sku))
                else:
                    size_mins = defaultdict(float)
                    for c in merged:
                        size_mins[self.sku_to_size.get(c["sku"])] += c["mins_used"]
                    lock_sz = (max(size_mins, key=size_mins.get)
                               if size_mins else None)
                same  = [c for c in merged
                         if self.sku_to_size.get(c["sku"]) == lock_sz]
                other = [c for c in merged
                         if self.sku_to_size.get(c["sku"]) != lock_sz]
                if other:
                    dropped = [c["sku"] for c in other]
                    print(f"    [Seq] {mach} (restricted, lock_sz={lock_sz}): "
                          f"dropping {len(dropped)} wrong-size: {dropped}")
                same.sort(key=lambda c: self._urgency_key(c["sku"]))
                ordered = same
            else:
                ordered = sorted(
                    merged,
                    key=lambda c: (self._urgency_key(c["sku"]), c["first_shift"])
                )
            result[str(mach)] = [
                (c["sku"], c["units"], c["ct"]) for c in ordered
            ]
            total_cos += max(0, len(result[str(mach)]) - 1)
        print(f"  [Seq] Machines: {len(result)} | Total changeovers: {total_cos}")
        return result


# ══════════════════════════════════════════════════════════════════════
# SCHEDULE BUILDER
# ══════════════════════════════════════════════════════════════════════
# Walks each machine's sequenced campaigns from plan start, inserts
# explicit CHANGEOVER rows between SKUs, and splits long campaigns at
# shift boundaries so the output is a row-per-shift-per-campaign.
class ScheduleBuilder:
    def __init__(self, plan_start, co_map):
        self.plan_start = plan_start
        self.plan_end   = plan_start + timedelta(days=Config.PLANNING_DAYS)
        self.co_map     = co_map

    def _co_mins(self, machine, sku_from, sku_to, sku_to_size):
        if sku_from is None or sku_from == sku_to:
            return 0.0
        co = self.co_map.get(str(machine), {"same":40,"diff":60})
        sz_f = sku_to_size.get(str(sku_from))
        sz_t = sku_to_size.get(str(sku_to))
        return co["same"] if sz_f == sz_t else co["diff"]

    def build(self, machine_sequences, locked_skus, locked_end_times,
              sku_to_size, stage_label="PRODUCTION"):
        rows = []
        for mach, campaigns in machine_sequences.items():
            cursor   = locked_end_times.get(str(mach), self.plan_start)
            last_sku = locked_skus.get(str(mach))
            for sku, units, ct_min in campaigns:
                co_dur = self._co_mins(mach, last_sku, sku, sku_to_size)
                if co_dur > 0:
                    co_end = cursor + timedelta(minutes=co_dur)
                    rows.extend(self._split_row(
                        cursor, co_end, mach, "CHANGEOVER", 0, 0.0,
                        f"C/O → {sku}"
                    ))
                    cursor = co_end
                prod_mins = units * ct_min
                prod_end  = min(cursor + timedelta(minutes=prod_mins),
                                self.plan_end)
                if prod_end <= cursor:
                    continue
                actual = (
                    int((prod_end - cursor).total_seconds() / 60.0 / ct_min)
                    if ct_min > 0 else units
                )
                rows.extend(self._split_row(
                    cursor, prod_end, mach, sku, actual, ct_min, stage_label
                ))
                cursor   = prod_end
                last_sku = sku
        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values(["Machine","StartTime"]).reset_index(drop=True)
        return df

    def _split_row(self, start, end, machine, sku, total_units, ct_min, remarks):
        rows, curr, rem = [], start, total_units
        total_dur = max((end - start).total_seconds() / 60.0, 1e-6)
        while curr < end:
            shift, shift_end = _shift_fn(curr)
            slice_end = min(shift_end, end)
            dur = (slice_end - curr).total_seconds() / 60.0
            if dur <= 0:
                curr = slice_end
                continue
            if total_units == 0 or sku == "CHANGEOVER":
                qty = 0
            elif slice_end == end:
                qty = rem
            else:
                qty = (
                    int(dur / ct_min) if ct_min > 0
                    else int(total_units * dur / total_dur)
                )
                qty = min(qty, rem)
            rows.append({
                "Date": curr.date(), "Shift": shift,
                "Machine": machine, "SKUCode": sku,
                "StartTime": curr, "EndTime": slice_end,
                "Qty": qty, "CT_Min": round(ct_min, 2),
                "Remarks": remarks,
            })
            rem -= qty
            curr = slice_end
        return rows


# ══════════════════════════════════════════════════════════════════════
# STARVATION VALIDATOR
# ══════════════════════════════════════════════════════════════════════
# Per-(Date, Shift, SKU) WIP balance: opening inv + built so far −
# cured so far. Negative balance = STARVATION (curing will starve),
# low-positive = WARNING, healthy = OK.
class StarvationValidator:
    def validate(self, df_building, df_shift_demand, df_gt_inv):
        gt_inv = dict(zip(df_gt_inv["SKUCode"], df_gt_inv["GT_Inventory"]))
        # A building day may produce no GT at all (e.g. the last days feed a
        # near-empty curing day). Then df_building is empty with no columns —
        # treat as zero build (everything that shift reads as starvation).
        if (df_building is None or df_building.empty
                or "SKUCode" not in df_building.columns):
            bld_grp = pd.DataFrame(columns=["Date", "Shift", "SKUCode", "Build_Qty"])
        else:
            bld = df_building[~df_building["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"])]
            bld_grp = (
                bld.groupby(["Date","Shift","SKUCode"])["Qty"].sum().reset_index()
                .rename(columns={"Qty":"Build_Qty"})
            )
        bld_grp["Date"] = pd.to_datetime(bld_grp["Date"])
        cur_grp = df_shift_demand.copy()
        if "Curing_Qty" not in cur_grp.columns:
            cur_grp.columns = ["Date","Shift","SKUCode","Curing_Qty"]
        cur_grp["Date"] = pd.to_datetime(cur_grp["Date"])
        all_events = (
            pd.concat([
                bld_grp[["Date","Shift","SKUCode"]],
                cur_grp[["Date","Shift","SKUCode"]]
            ])
            .drop_duplicates()
            .sort_values(["SKUCode","Date","Shift"])
        )
        all_events = (
            all_events.merge(bld_grp, on=["Date","Shift","SKUCode"], how="left")
                       .merge(cur_grp, on=["Date","Shift","SKUCode"], how="left")
                       .fillna(0)
        )
        rows = []
        for sku, grp in all_events.groupby("SKUCode"):
            grp = grp.sort_values(["Date","Shift"]).reset_index(drop=True)
            wip = float(gt_inv.get(sku, 0))
            for _, r in grp.iterrows():
                wip += r["Build_Qty"] - r["Curing_Qty"]
                burn = max(r["Curing_Qty"], 1)
                status = (
                    "STARVATION" if wip < 0
                    else "WARNING" if wip < burn * Config.BUFFER_SHIFTS
                    else "OK"
                )
                rows.append({
                    "Date": r["Date"], "Shift": r["Shift"], "SKUCode": sku,
                    "Build_Qty": int(r["Build_Qty"]),
                    "Cure_Qty":  int(r["Curing_Qty"]),
                    "WIP_Balance": int(wip),
                    "Status": status,
                })
        df = pd.DataFrame(rows)
        s = (df["Status"]=="STARVATION").sum()
        w = (df["Status"]=="WARNING").sum()
        print(f"  [Validate] Rows: {len(df)} | Starv: {s} | Warn: {w} | "
              f"OK: {(df['Status']=='OK').sum()}")
        return df


# ══════════════════════════════════════════════════════════════════════
# REPORT BUILDERS  (extracted from v5's scheduler private methods)
# ══════════════════════════════════════════════════════════════════════
def build_summary(df_sku_demand, df_all):
    prod = (
        df_all[~df_all["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])]
        if not df_all.empty else pd.DataFrame(columns=["SKUCode","Qty","Machine"])
    )
    # GT demand is satisfied ONLY by GT machines (Stage-2 + Unistage). Stage-1
    # carcass output is a separate upstream WIP stage that FEEDS green-tyre
    # building — counting it as "Planned_GT" double-counts and inflates the
    # built/fulfilment numbers, so restrict to GT machines here.
    gt_machines = set(map(str, Config.STAGE2 | Config.UNISTAGE))
    if not prod.empty and "Machine" in prod.columns:
        prod = prod[prod["Machine"].astype(str).isin(gt_machines)]
    planned = prod.groupby("SKUCode")["Qty"].sum().to_dict()
    rows = []
    for _, r in df_sku_demand.iterrows():
        sku  = r["SKUCode"]
        dem  = int(r.get("Net_GT_Demand", 0))
        plan = int(planned.get(sku, 0))
        gap  = max(dem - plan, 0)
        pct  = round(plan / dem * 100, 1) if dem > 0 else 100.0
        rows.append({
            "SKUCode":             sku,
            "GT_Demand":           int(r.get("GT_Demand",0)),
            "GT_Inventory":        int(r.get("GT_Inventory",0)),
            "Net_GT_Demand":       dem,
            "Planned_GT":          plan,
            "Gap":                 gap,
            "Fulfillment_Pct":     pct,
            "Status": ("FULLY MET" if gap<=0
                       else "PARTIAL" if plan>0 else "UNMET"),
            "Burn_Rate_Per_Shift": round(r.get("Burn_Rate_Per_Shift",0),1),
            "Active_Shifts":       int(r.get("Active_Shifts",0)),
            "First_Curing_Start":  r.get("First_Curing_Start",""),
        })
    return pd.DataFrame(rows).sort_values("First_Curing_Start")


def build_util(df_sched, machines):
    if not df_sched.empty:
        prod = df_sched[~df_sched["SKUCode"].isin(["CHANGEOVER"])].copy()
        prod["Elapsed"] = (
            pd.to_datetime(prod["EndTime"]) - pd.to_datetime(prod["StartTime"])
        ).dt.total_seconds() / 60.0
        grp = prod.groupby("Machine").agg(
            Used_Mins  = ("Elapsed","sum"),
            Total_Units= ("Qty","sum"),
            SKUs_Count = ("SKUCode","nunique"),
        ).reset_index()
        co_grp = (
            df_sched[df_sched["SKUCode"]=="CHANGEOVER"]
            .groupby("Machine").size().reset_index(name="Changeovers")
        )
        grp = grp.merge(co_grp, on="Machine", how="left").fillna(0)
    else:
        grp = pd.DataFrame(
            columns=["Machine","Used_Mins","Total_Units","SKUs_Count","Changeovers"]
        )
    cap  = float(Config.SHIFT_MINS * Config.total_shifts())
    df_u = (
        pd.DataFrame({"Machine":[str(m) for m in machines]})
        .merge(grp, on="Machine", how="left").fillna(0)
    )
    df_u["Available_Mins"]  = cap
    df_u["Idle_Mins"]       = cap - df_u["Used_Mins"]
    df_u["Utilization_Pct"] = (df_u["Used_Mins"] / cap * 100).round(2)
    return df_u[[
        "Machine","Available_Mins","Used_Mins","Idle_Mins",
        "Utilization_Pct","SKUs_Count","Total_Units","Changeovers"
    ]].sort_values("Utilization_Pct", ascending=False)




def build_daywise_report(df_shift_demand, df_gt_sched, df_s1_sched):
    def agg(df, val_col, out_col):
        if df is None or df.empty:
            return pd.DataFrame(columns=["Date","Shift","SKUCode",out_col])
        prod = df[~df["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])].copy()
        prod["Date"] = pd.to_datetime(prod["Date"]).dt.date
        g = prod.groupby(["Date","Shift","SKUCode"])[val_col].sum().reset_index()
        return g.rename(columns={val_col: out_col})

    gt_agg = agg(df_gt_sched, "Qty", "GT_Produced")
    s1_agg = agg(df_s1_sched, "Qty", "Carcass_Produced")
    cd = (df_shift_demand.copy()
          if df_shift_demand is not None and not df_shift_demand.empty
          else pd.DataFrame(columns=["Date","Shift","SKUCode","Curing_Qty"]))
    if not cd.empty:
        cd["Date"] = pd.to_datetime(cd["Date"]).dt.date
        cd = cd.rename(columns={"Curing_Qty": "Curing_Demand"})
    else:
        cd["Curing_Demand"] = 0
    all_keys = (
        pd.concat([
            cd[["Date","Shift","SKUCode"]] if not cd.empty
            else pd.DataFrame(columns=["Date","Shift","SKUCode"]),
            gt_agg[["Date","Shift","SKUCode"]],
            s1_agg[["Date","Shift","SKUCode"]],
        ], ignore_index=True)
        .drop_duplicates()
    )
    out = (
        all_keys
        .merge(cd[["Date","Shift","SKUCode","Curing_Demand"]] if not cd.empty
               else all_keys.assign(Curing_Demand=0),
               on=["Date","Shift","SKUCode"], how="left")
        .merge(gt_agg, on=["Date","Shift","SKUCode"], how="left")
        .merge(s1_agg, on=["Date","Shift","SKUCode"], how="left")
        .fillna(0)
    )
    for col in ["Curing_Demand","GT_Produced","Carcass_Produced"]:
        out[col] = out[col].astype(int)
    out["GT_vs_Demand"]  = out["GT_Produced"]      - out["Curing_Demand"]
    out["Carcass_vs_GT"] = out["Carcass_Produced"] - out["GT_Produced"]
    shift_order = {"A":0,"B":1,"C":2}
    out["_sh"] = out["Shift"].map(shift_order).fillna(9)
    out = out.sort_values(["Date","_sh","SKUCode"]).drop(columns=["_sh"])
    return out.reset_index(drop=True)


def build_daily_sku_counts(df_gt_sched, df_s1_sched):
    def per_day(df):
        if df is None or df.empty:
            return {}
        prod = df[~df["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])].copy()
        prod["Date"] = pd.to_datetime(prod["Date"]).dt.date
        return prod.groupby("Date")["SKUCode"].nunique().to_dict()

    gt_cnt = per_day(df_gt_sched)
    s1_cnt = per_day(df_s1_sched)
    all_dates = sorted(set(gt_cnt.keys()) | set(s1_cnt.keys()))
    rows = []
    for d in all_dates:
        gt = gt_cnt.get(d, 0)
        rows.append({
            "Date":                  d,
            "GT_Distinct_SKUs":      gt,
            "Carcass_Distinct_SKUs": s1_cnt.get(d, 0),
            "Target":                Config.DAILY_SKU_TARGET,
            "Met":                   "✓" if gt >= Config.DAILY_SKU_TARGET else "✗",
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════
# Writes every report sheet (Shift Schedule, GT/Carcass Machine
# Schedules, Changeover Analysis, Starvation, Utilization, etc.) plus
# two flat _raw.xlsx / _curing.xlsx companion files. Pure presentation
# layer — all numbers come from the scheduler's results dict.
class ExcelExporter:
    _C = {"navy":"1F3864","teal":"1F6B75","green":"C6EFCE","amber":"FFEB9C",
          "red":"FFC7CE","grey":"F2F2F2","white":"FFFFFF","lgrey":"E8E8E8",
          "orange":"F4B942"}
    STATUS_FC = {"FULLY MET":"green","PARTIAL":"amber","UNMET":"red"}
    STARV_FC  = {"STARVATION":"red","WARNING":"amber","OK":"green"}
    SHIFT_FC  = {"CHANGEOVER":"orange","A":"E8F4F8","B":"FFF8E8","C":"F0F0F0"}
    MET_FC    = {"✓":"green","✗":"red"}

    def __init__(self, path):
        self.path = path

    def _F(self, c): return PatternFill("solid", fgColor=self._C.get(c, c))
    def _b(self):
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def _hf(self): return Font(bold=True, name="Arial", size=10, color="FFFFFF")
    def _bf(self, bold=False): return Font(bold=bold, name="Arial", size=9)

    def _cell(self, ws, r, c, v, fc="white", bold=False, aln="center"):
        cell = ws.cell(r, c, v)
        cell.font = self._bf(bold); cell.fill = self._F(fc); cell.border = self._b()
        cell.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=True)

    def _hdr(self, ws, row, ncols, fc="navy"):
        for c in range(1, ncols+1):
            cell = ws.cell(row, c)
            cell.font = self._hf(); cell.fill = self._F(fc); cell.border = self._b()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    def _title(self, ws, text, sub, ncols):
        ws.insert_rows(1); ws.insert_rows(1)
        cl = get_column_letter(ncols)
        ws.merge_cells(f"A1:{cl}1"); ws["A1"] = text
        ws["A1"].font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
        ws["A1"].fill = self._F("navy")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.merge_cells(f"A2:{cl}2"); ws["A2"] = sub
        ws["A2"].font = Font(italic=True, name="Arial", size=9, color="FFFFFF")
        ws["A2"].fill = self._F("teal")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

    def export(self, results: dict):
        # Pull each artefact out of the results dict produced by
        # HybridDailyScheduler.run(). Each gets its own sheet below.
        df_gt   = results["gt_allocation"]
        df_s1   = results["s1_allocation"]
        df_sch  = results["shift_schedule"].copy()
        df_sum  = results["demand_summary"].copy()
        df_stv  = results["starvation_report"].copy()
        df_ug   = results["gt_utilization"]
        df_us   = results["s1_utilization"]
        df_cm   = results["curing_matrix"]
        df_dw   = results["daywise_report"].copy()
        df_dcnt = results["daily_sku_counts"].copy()

        for col in ("Date",):
            for d in (df_sch, df_stv, df_dw, df_dcnt):
                if col in d.columns:
                    d[col] = pd.to_datetime(d[col])
        for col in ("StartTime","EndTime"):
            if col in df_sch.columns:
                df_sch[col] = pd.to_datetime(df_sch[col])
        # Plant-day: shift A starts at 07:00, so times before 07:00 belong to
        # the previous calendar day. Mirrors the user's Excel formula
        # =TEXT(IF(HOUR(F)<7,INT(F)-1,INT(F)),"yyyy-mm-dd").
        if "StartTime" in df_sch.columns:
            df_sch["Date2"] = (
                df_sch["StartTime"] - pd.Timedelta(hours=7)
            ).dt.floor("D").dt.date
        if "First_Curing_Start" in df_sum.columns:
            df_sum["First_Curing_Start"] = pd.to_datetime(
                df_sum["First_Curing_Start"], errors="coerce"
            )

        # Headline KPIs printed at the top of every sheet title bar.
        td  = int(df_sum["Net_GT_Demand"].sum()) if not df_sum.empty else 0
        tp  = int(df_sum["Planned_GT"].sum())    if not df_sum.empty else 0
        # Fulfilment = demand actually COVERED, capped per SKU at its own demand.
        # Total built (tp) can exceed demand via shelf-life topup pre-build and
        # whole-cycle rounding, but that surplus does not "fulfil" demand, so it
        # must not inflate the % past 100.
        met = (int(np.minimum(df_sum["Planned_GT"], df_sum["Net_GT_Demand"]).sum())
               if not df_sum.empty else 0)
        pct = round(met/td*100,1) if td else 0
        stv = int((df_stv["Status"]=="STARVATION").sum()) if not df_stv.empty else 0
        kpi = (f"GT Demand: {td:,}  |  Built: {tp:,}  |  "
               f"Fulfillment: {pct}%  |  Starvation: {stv}")

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            # 0b. Daily SKU Counts
            cd0 = ["Date","GT_Distinct_SKUs","Carcass_Distinct_SKUs","Target","Met"]
            df_dcnt[cd0].to_excel(writer, sheet_name="Daily SKU Counts", index=False)
            wsd = writer.book["Daily SKU Counts"]
            self._title(wsd, "PER-DAY DISTINCT SKUs — GT ONLY",
                        f"Target: {Config.DAILY_SKU_TARGET}", len(cd0))
            self._hdr(wsd, 3, len(cd0))
            for ci, w in enumerate([14,18,20,10,8], 1):
                wsd.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, wsd.max_row+1):
                met = str(wsd.cell(ri, 5).value)
                fc  = self.MET_FC.get(met, "white")
                bf  = "grey" if ri%2==0 else "white"
                for ci in range(1, len(cd0)+1):
                    self._cell(wsd, ri, ci, wsd.cell(ri, ci).value,
                               fc=fc if ci in (2,5) else bf,
                               bold=(ci in (2,5)))
                wsd.cell(ri, 1).number_format = "yyyy-mm-dd"

            # 0c. Daywise Report
            cdw = ["Date","Shift","SKUCode","Curing_Demand",
                   "GT_Produced","Carcass_Produced","GT_vs_Demand","Carcass_vs_GT"]
            df_dw[cdw].to_excel(writer, sheet_name="Daywise Report", index=False)
            wsdw = writer.book["Daywise Report"]
            self._title(wsdw, "DAYWISE DEMAND vs PRODUCTION",
                        "Per (Date, Shift, SKU)", len(cdw))
            self._hdr(wsdw, 3, len(cdw))
            for ci, w in enumerate([12,8,26,14,14,16,14,14], 1):
                wsdw.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, wsdw.max_row+1):
                bf = "grey" if ri%2==0 else "white"
                for ci in range(1, len(cdw)+1):
                    val = wsdw.cell(ri, ci).value
                    fc = bf
                    if ci == 7:
                        fc = "red" if isinstance(val,(int,float)) and val < 0 else "green"
                    elif ci == 8:
                        fc = "red" if isinstance(val,(int,float)) and val < 0 else "green"
                    self._cell(wsdw, ri, ci, val,
                               fc=fc, bold=(ci in (4,5,6)),
                               aln="left" if ci == 3 else "center")
                wsdw.cell(ri, 1).number_format = "yyyy-mm-dd"

            # 1. Demand Summary
            c1 = ["SKUCode","GT_Demand","GT_Inventory","Net_GT_Demand",
                  "Planned_GT","Gap","Fulfillment_Pct","Status",
                  "Burn_Rate_Per_Shift","Active_Shifts","First_Curing_Start"]
            df_sum[c1].to_excel(writer, sheet_name="Demand Summary", index=False)
            ws = writer.book["Demand Summary"]
            self._title(ws, "PCR BUILDING — Hybrid v8", kpi, len(c1))
            self._hdr(ws, 3, len(c1))
            for ci, w in enumerate([22,13,13,13,13,10,10,12,18,12,20], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, ws.max_row+1):
                st = str(ws.cell(ri,8).value)
                sf = self.STATUS_FC.get(st,"white")
                bf = "grey" if ri%2==0 else "white"
                for ci in range(1, len(c1)+1):
                    self._cell(ws,ri,ci,ws.cell(ri,ci).value,
                               fc=sf if ci in (7,8) else bf,
                               bold=(ci==5),
                               aln="left" if ci==1 else "center")
                fp = ws.cell(ri,7)
                if isinstance(fp.value,(int,float)):
                    fp.value = fp.value/100
                fp.number_format = "0.0%"
                ws.cell(ri, 11).number_format = "yyyy-mm-dd hh:mm:ss"

            # 2. Shift Schedule
            c4 = ["Date","Date2","Shift","Machine","MachineType","SKUCode",
                  "StartTime","EndTime","Qty","CT_Min","Remarks"]
            out4 = df_sch[[c for c in c4 if c in df_sch.columns]]
            out4.to_excel(writer, sheet_name="Shift Schedule", index=False)
            ws4 = writer.book["Shift Schedule"]
            self._title(ws4, "SHIFT-WISE BUILDING SCHEDULE — Hybrid v8", kpi, len(c4))
            self._hdr(ws4, 3, len(c4))
            for ci,w in enumerate([12,12,8,12,12,26,18,18,10,10,26], 1):
                ws4.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, ws4.max_row+1):
                sku  = str(ws4.cell(ri,6).value)
                shft = str(ws4.cell(ri,3).value)
                fc   = self.SHIFT_FC.get(sku) or self.SHIFT_FC.get(shft,"white")
                for ci in range(1, len(c4)+1):
                    self._cell(ws4, ri, ci, ws4.cell(ri, ci).value,
                               fc=fc, bold=(sku=="CHANGEOVER"))
                ws4.cell(ri, 1).number_format = "yyyy-mm-dd"
                ws4.cell(ri, 2).number_format = "yyyy-mm-dd"
                ws4.cell(ri, 7).number_format = "yyyy-mm-dd hh:mm:ss"
                ws4.cell(ri, 8).number_format = "yyyy-mm-dd hh:mm:ss"

            # 3. GT Machine Schedule
            c2 = ["Machine","SKUCode","Units","Mins_Used","CT_Min","ShiftIdx"]
            # cat() returns a column-less empty frame when NO LP allocation rows
            # were produced across the whole horizon (e.g. GT GA returned no
            # matched allocation and the schedule was filled by topup only).
            # Guarantee the expected columns so the sort/sheet-write doesn't
            # KeyError — the sheet is simply empty in that case.
            if df_gt.empty or "Machine" not in df_gt.columns:
                df_gt = pd.DataFrame(columns=c2)
            if df_s1.empty or "Machine" not in df_s1.columns:
                df_s1 = pd.DataFrame(columns=c2)
            (df_gt.sort_values(["Machine","ShiftIdx","SKUCode"])
             .to_excel(writer, sheet_name="GT Machine Schedule", index=False))
            ws2 = writer.book["GT Machine Schedule"]
            self._title(ws2, "GT MACHINE SCHEDULE — Hybrid v8", kpi, len(c2))
            self._hdr(ws2, 3, len(c2))
            for ci,w in enumerate([12,26,12,14,12,10], 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w
            prev = None
            for ri in range(4, ws2.max_row+1):
                m = ws2.cell(ri,1).value
                bf = "lgrey" if m != prev else ("grey" if ri%2==0 else "white")
                prev = m
                for ci in range(1, len(c2)+1):
                    self._cell(ws2, ri, ci, ws2.cell(ri, ci).value,
                               fc=bf, bold=(ci==3))

            # 4. Carcass Machine Schedule
            (df_s1.sort_values(["Machine","ShiftIdx","SKUCode"])
             .to_excel(writer, sheet_name="Carcass Machine Schedule", index=False))
            ws3 = writer.book["Carcass Machine Schedule"]
            self._title(ws3, "CARCASS MACHINE SCHEDULE — Hybrid v8", kpi, len(c2))
            self._hdr(ws3, 3, len(c2))
            for ci,w in enumerate([12,26,12,14,12,10], 1):
                ws3.column_dimensions[get_column_letter(ci)].width = w
            prev = None
            for ri in range(4, ws3.max_row+1):
                m = ws3.cell(ri,1).value
                bf = "lgrey" if m != prev else ("grey" if ri%2==0 else "white")
                prev = m
                for ci in range(1, len(c2)+1):
                    self._cell(ws3, ri, ci, ws3.cell(ri, ci).value,
                               fc=bf, bold=(ci==3))

            # 4b. Changeover Analysis — per-CO row with inch-diff indicator.
            # Big=1 if From/To sizes differ (different inch), else 0.
            # Small=1 if From/To sizes match (same inch), else 0.
            sku_to_size_map = results.get("sku_to_size", {}) or {}
            co_rows = []
            if not df_sch.empty and "SKUCode" in df_sch.columns:
                co_src = df_sch.sort_values(["Machine","StartTime"]).reset_index(drop=True)
                for mach, grp in co_src.groupby("Machine"):
                    grp = grp.reset_index(drop=True)
                    prev_sku = None
                    skus = grp["SKUCode"].astype(str).tolist()
                    for i in range(len(grp)):
                        code = skus[i]
                        if code == "CHANGEOVER":
                            to_sku = ""
                            for j in range(i+1, len(grp)):
                                nxt = skus[j]
                                if nxt not in ("CHANGEOVER","MOULD_CLEAN"):
                                    to_sku = nxt
                                    break
                            from_sku = prev_sku or ""
                            from_sz = sku_to_size_map.get(str(from_sku), "") if from_sku else ""
                            to_sz   = sku_to_size_map.get(str(to_sku), "")   if to_sku   else ""
                            same = bool(from_sz and to_sz and from_sz == to_sz)
                            co_rows.append({
                                "Machine": mach,
                                "Date": (pd.to_datetime(grp.loc[i,"Date"]).date()
                                         if pd.notna(grp.loc[i,"Date"]) else None),
                                "Changeover_From": from_sku,
                                "Changeover_To":   to_sku,
                                "Big":   0 if same else 1,
                                "Small": 1 if same else 0,
                            })
                        elif code != "MOULD_CLEAN":
                            prev_sku = code
            c_co = ["Machine","Date","Changeover_From","Changeover_To","Big","Small"]
            df_co = (pd.DataFrame(co_rows, columns=c_co) if co_rows
                     else pd.DataFrame(columns=c_co))
            df_co.to_excel(writer, sheet_name="Changeover Analysis", index=False)
            wsco = writer.book["Changeover Analysis"]
            tot_big   = int(df_co["Big"].sum())   if not df_co.empty else 0
            tot_small = int(df_co["Small"].sum()) if not df_co.empty else 0
            self._title(
                wsco, "CHANGEOVER ANALYSIS — Hybrid v8",
                f"Total: {len(df_co)}  |  Big (diff inch): {tot_big}  |  "
                f"Small (same inch): {tot_small}",
                len(c_co),
            )
            self._hdr(wsco, 3, len(c_co))
            for ci, w in enumerate([12, 12, 26, 26, 8, 8], 1):
                wsco.column_dimensions[get_column_letter(ci)].width = w
            prev_m = None
            for ri in range(4, wsco.max_row+1):
                m = wsco.cell(ri,1).value
                bf = "lgrey" if m != prev_m else ("grey" if ri%2==0 else "white")
                prev_m = m
                for ci in range(1, len(c_co)+1):
                    self._cell(wsco, ri, ci, wsco.cell(ri, ci).value,
                               fc=bf, bold=(ci in (5,6)))
                wsco.cell(ri, 2).number_format = "yyyy-mm-dd"

            # 5. Starvation
            c5 = ["Date","Shift","SKUCode","Build_Qty","Cure_Qty","WIP_Balance","Status"]
            df_stv[c5].to_excel(writer, sheet_name="Starvation Report", index=False)
            ws5 = writer.book["Starvation Report"]
            stv_kpi = (f"Starvation: {stv}  |  "
                       f"Warning: {(df_stv['Status']=='WARNING').sum()}  |  "
                       f"OK: {(df_stv['Status']=='OK').sum()}")
            self._title(ws5,"ANTI-STARVATION — Hybrid v8", stv_kpi, len(c5))
            self._hdr(ws5, 3, len(c5))
            for ci,w in enumerate([12,8,26,14,14,14,14], 1):
                ws5.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, ws5.max_row+1):
                st = str(ws5.cell(ri,7).value)
                fc = self.STARV_FC.get(st,"white")
                bf = "grey" if ri%2==0 else "white"
                for ci in range(1, len(c5)+1):
                    self._cell(ws5, ri, ci, ws5.cell(ri, ci).value,
                               fc=fc if ci in (6,7) else bf,
                               bold=(st in ("STARVATION","WARNING")))
                ws5.cell(ri, 1).number_format = "yyyy-mm-dd"

            # 6. Utilization
            df_util_all = pd.concat([
                df_ug.assign(Stage="GT/Unistage"),
                df_us.assign(Stage="Stage1"),
            ], ignore_index=True)
            c6 = ["Stage","Machine","Available_Mins","Used_Mins","Idle_Mins",
                  "Utilization_Pct","SKUs_Count","Total_Units","Changeovers"]
            df_util_all[[c for c in c6 if c in df_util_all.columns]].to_excel(
                writer, sheet_name="Machine Utilization", index=False
            )
            ws6 = writer.book["Machine Utilization"]
            avg_u = round(df_util_all["Utilization_Pct"].mean(), 1)
            self._title(ws6, "PRESS UTILIZATION — Hybrid v8",
                        f"Avg: {avg_u}% | Machines: {len(df_util_all)}", len(c6))
            self._hdr(ws6, 3, len(c6))
            for ci,w in enumerate([14,12,15,14,14,14,12,14,14], 1):
                ws6.column_dimensions[get_column_letter(ci)].width = w
            for ri in range(4, ws6.max_row+1):
                u  = ws6.cell(ri,6).value or 0
                uf = "green" if u>=90 else "amber" if u>=60 else "red"
                bf = "grey" if ri%2==0 else "white"
                for ci in range(1, len(c6)+1):
                    self._cell(ws6, ri, ci, ws6.cell(ri, ci).value,
                               fc=uf if ci==6 else bf,
                               bold=(ci in (2,6)))
                fp = ws6.cell(ri,6)
                if isinstance(fp.value,(int,float)):
                    fp.value = fp.value/100
                fp.number_format = "0.0%"

            # 7. Curing Demand Matrix
            df_cm.to_excel(writer, sheet_name="Curing Demand Matrix")
            ws7 = writer.book["Curing Demand Matrix"]
            self._title(ws7, "SHIFT-WISE CURING DEMAND",
                        "Rows=SKUs | Cols=Shift idx", df_cm.shape[1]+1)
            ws7.column_dimensions["A"].width = 26

        print(f"\n  [Export] Saved → {self.path}")

        raw = df_sch.rename(columns={"MachineType": "Stage"})
        raw_cols = ["Date", "Date2", "Machine", "SKUCode", "Qty",
                    "StartTime", "EndTime", "Stage"]
        raw = raw[[c for c in raw_cols if c in raw.columns]]
        raw_path = os.path.splitext(self.path)[0] + "_raw.xlsx"
        raw.to_excel(raw_path, index=False)
        print(f"  [Export] Raw schedule → {raw_path}")

        df_cur = results.get("curing_input")
        if df_cur is not None and not df_cur.empty:
            cur = df_cur.copy()
            cur = cur.rename(columns={"machineName": "Machine"})
            cur["Stage"] = "Curing"
            if "Date" not in cur.columns and "StartTime" in cur.columns:
                cur["Date"] = pd.to_datetime(cur["StartTime"]).dt.date
            if "StartTime" in cur.columns:
                cur["Date2"] = (
                    pd.to_datetime(cur["StartTime"]) - pd.Timedelta(hours=7)
                ).dt.floor("D").dt.date
            cur_cols = ["Date", "Date2", "Machine", "SKUCode", "Qty",
                        "StartTime", "EndTime", "Stage"]
            cur = cur[[c for c in cur_cols if c in cur.columns]]
            cur_path = os.path.splitext(self.path)[0] + "_curing.xlsx"
            cur.to_excel(cur_path, index=False)
            print(f"  [Export] Curing demand → {cur_path}")


# ══════════════════════════════════════════════════════════════════════
# GA HYPERPARAMETERS
# ══════════════════════════════════════════════════════════════════════
# Hyperparameters for the outer GA (POPULATION, GENERATIONS, etc.)
# and its fitness penalties / bonuses.
class GAConfig:
    POPULATION       = 32         # v8: wider exploration
    GENERATIONS      = 30         # v8: more refinement passes
    ELITE            = 4
    TOURNAMENT       = 3
    CROSSOVER_RATE   = 0.85
    MUTATION_RATE    = 0.10
    MIN_MACH_PER_SKU = 2
    MAX_MACH_PER_SKU = 5          # v8: allow seeded chromosomes to spread further

    SLACK_PENALTY    = 1_000_000.0
    CO_PENALTY       = 30.0
    DIVERSITY_BONUS  = 2500.0


# ══════════════════════════════════════════════════════════════════════
# LP MINUTE SOLVER  (continuous, given fixed binary y)
# ══════════════════════════════════════════════════════════════════════
# Inner optimiser — given a fixed binary y[SKU, machine] matrix from the
# GA, solves a linear program for continuous minute allocations
# x[SKU, machine, shift] that (1) meet cumulative curing WIP via
# starvation slack, (2) respect per-shift machine capacity, (3) enforce
# MIN_CAMPAIGN_MINS for every active pair, and (4) reserve changeover
# time proportional to how many SKUs share a machine. Solver: HiGHS
# (scipy.optimize.linprog).
class LPMinuteSolver:
    """Given fixed y[s,m] ∈ {0,1}, solve LP for x[s,m,t] minutes (1 day)."""

    def __init__(self):
        self.shift_mins = float(Config.SHIFT_MINS)
        self.T = Config.SHIFTS_PER_DAY

    def solve(self, y, curing_matrix, ct_map, sku_list, machines,
              inv_init, locked_pairs, inv_cap=None, co_time_map=None,
              cap_on_gross=False):
        S, M, T = len(sku_list), len(machines), self.T
        if S == 0 or M == 0:
            return None, float("inf"), {}

        n_x  = S * M * T
        n_sl = S * T
        n_vars = n_x + n_sl

        def xi(s, m, t): return s * M * T + m * T + t
        def sli(s, t):   return n_x + s * T + t

        sku_idx = {s: i for i, s in enumerate(sku_list)}
        m_idx   = {m: i for i, m in enumerate(machines)}

        # Objective
        unit_w = float(getattr(Config, "UNIT_BONUS_MULT", 1.0))
        c = np.zeros(n_vars)
        for si in range(S):
            for mi in range(M):
                ct = ct_map.get(machines[mi], 2.0)
                if ct > 0:
                    for t in range(T):
                        c[xi(si, mi, t)] = -unit_w / ct  # max units (v8: ×UNIT_BONUS_MULT)
            for t in range(T):
                c[sli(si, t)] = GAConfig.SLACK_PENALTY

        # Bounds
        ub = np.full(n_vars, np.inf)
        for si in range(S):
            for mi in range(M):
                hi = self.shift_mins if y[si, mi] >= 0.5 else 0.0
                for t in range(T):
                    ub[xi(si, mi, t)] = hi

        rows, bvals = [], []
        def new_row(): return lil_matrix((1, n_vars))

        # CAP per (m,t)
        for mi in range(M):
            for t in range(T):
                r = new_row()
                for si in range(S):
                    r[0, xi(si, mi, t)] = 1.0
                rows.append(r); bvals.append(self.shift_mins)

        # CO-RESERVE per machine over the day (anti-starvation):
        #   The ScheduleBuilder will insert CHANGEOVER blocks between SKUs on
        #   every machine that has >1 y=1. Without reserving those minutes,
        #   the LP over-promises production that gets eaten by changeovers
        #   at build-time → actual units < planned → WIP < 0 → STARVATION.
        #
        #   Σ_{s,t} x[s,m,t] ≤ SHIFT_MINS·T  −  co_time[m] · max(0, y_sum[m]−1)
        if co_time_map is not None:
            horizon_mins = self.shift_mins * T
            for mi, mach in enumerate(machines):
                y_sum_m = int(y[:, mi].sum())
                if y_sum_m <= 1:
                    continue  # no CO when ≤1 SKU runs on this machine
                co_t = float(co_time_map.get(str(mach), 60.0))
                co_reserve = co_t * (y_sum_m - 1)
                r = new_row()
                for si in range(S):
                    for t in range(T):
                        r[0, xi(si, mi, t)] = 1.0
                rhs = max(0.0, horizon_mins - co_reserve)
                rows.append(r); bvals.append(rhs)

        # WIP / starvation
        for si, sku in enumerate(sku_list):
            wip0 = float(inv_init.get(sku, 0))
            cum_cure = 0.0
            for t in range(T):
                cum_cure += curing_matrix[si, t]
                need = max(0.0, cum_cure - wip0)
                r = new_row()
                for tau in range(t + 1):
                    for mi, mach in enumerate(machines):
                        ct = ct_map.get(mach, 2.0)
                        if ct > 0:
                            r[0, xi(si, mi, tau)] = -1.0 / ct
                    r[0, sli(si, tau)] = -1.0
                rows.append(r); bvals.append(-need)

        # MIN_CAMPAIGN per active y — collected SEPARATELY so it can be dropped
        # if it makes the LP infeasible. A hard 45-min floor on EVERY active
        # (SKU,machine) pair is unsatisfiable when the GA packs many SKUs onto
        # one machine (45·n + CO·(n−1) > day minutes). Left as a hard constraint
        # that would make the WHOLE LP infeasible → nothing builds → topup masks
        # it. Treated as a best-effort preference instead (see two-pass solve).
        mc_rows, mc_bvals = [], []
        min_camp = float(Config.MIN_CAMPAIGN_MINS)
        for si in range(S):
            for mi in range(M):
                if y[si, mi] >= 0.5:
                    r = new_row()
                    for t in range(T):
                        r[0, xi(si, mi, t)] = -1.0
                    mc_rows.append(r); mc_bvals.append(-min_camp)

        # LOCK
        lock_min = float(Config.LOCKED_MIN_SHIFT0_MINS)
        for m_lock, sku_lock in locked_pairs:
            if m_lock not in m_idx or sku_lock not in sku_idx:
                continue
            si, mi = sku_idx[sku_lock], m_idx[m_lock]
            if y[si, mi] < 0.5:
                continue
            r = new_row()
            r[0, xi(si, mi, 0)] = -1.0
            rows.append(r); bvals.append(-lock_min)

        # End-of-day INV_CAP
        if inv_cap is not None:
            total_inv_init = sum(float(inv_init.get(s, 0)) for s in sku_list)
            cum_cure_T = float(curing_matrix.sum())
            rhs = float(inv_cap) - total_inv_init + cum_cure_T
            r = new_row()
            for si in range(S):
                for mi, mach in enumerate(machines):
                    ct = ct_map.get(mach, 2.0)
                    if ct > 0:
                        for t in range(T):
                            r[0, xi(si, mi, t)] = 1.0 / ct
            rows.append(r); bvals.append(max(rhs, 0.0))

        # PRODUCTION CAP per SKU — limits daily building to avoid gross
        # overproduction.  Two modes:
        #   cap_on_gross=True  (GT):  cap = gross × buf, ignoring WIP.
        #     Used when rolling WIP from DB inventory would otherwise suppress
        #     the cap to zero and leave machines idle.  The starvation constraint
        #     (above) already uses WIP to set the lower bound; this upper bound
        #     just prevents runaway overproduction beyond 1+buf× daily demand.
        #   cap_on_gross=False (S1):  cap = max(0, gross − wip0) × buf.
        #     Classic net-demand cap; safe for carcass whose 1-day shelf life
        #     means surplus inventory should suppress today's build.
        # Floored at the locked minimum so continuity locks never make it
        # infeasible. Σ_{m,t} x[s,m,t]/ct ≤ cap_units[s].
        buf = 1.0 + float(getattr(Config, "OVERBUILD_BUFFER_FRAC", 0.10))
        locked_floor = np.zeros(S)
        for m_lock, sku_lock in locked_pairs:
            if m_lock in m_idx and sku_lock in sku_idx and y[sku_idx[sku_lock], m_idx[m_lock]] >= 0.5:
                ctl = ct_map.get(m_lock, 2.0)
                if ctl > 0:
                    locked_floor[sku_idx[sku_lock]] += lock_min / ctl
        for si, sku in enumerate(sku_list):
            gross = float(curing_matrix[si].sum())
            if cap_on_gross:
                cap_units = max(gross * buf, float(locked_floor[si]))
            else:
                wip0  = float(inv_init.get(sku, 0))
                cap_units = max(max(0.0, gross - wip0) * buf, float(locked_floor[si]))
            r = new_row()
            for mi, mach in enumerate(machines):
                ct = ct_map.get(mach, 2.0)
                if ct > 0:
                    for t in range(T):
                        r[0, xi(si, mi, t)] = 1.0 / ct
            rows.append(r); bvals.append(cap_units)

        bounds = [(0.0, ub[i]) for i in range(n_vars)]

        def _run(extra_rows, extra_b):
            all_rows = rows + extra_rows
            all_b    = bvals + extra_b
            A_ub = sp_vstack(all_rows).tocsr() if all_rows else None
            b_ub = np.array(all_b, dtype=float) if all_b else None
            return linprog(c, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")

        # Pass 1: with the min-campaign floor. Pass 2 (fallback): without it, so
        # an over-packed machine assignment still yields a producing schedule
        # rather than a globally-infeasible LP (which would build nothing).
        res = _run(mc_rows, mc_bvals)
        min_campaign_relaxed = False
        if (not res.success or res.x is None) and mc_rows:
            res = _run([], [])
            min_campaign_relaxed = True
        if not res.success or res.x is None:
            return None, float("inf"), {"status": res.status}

        x_sol = res.x
        units = 0.0
        per_sku_units = np.zeros(S)
        for si in range(S):
            for mi, mach in enumerate(machines):
                ct = ct_map.get(mach, 2.0)
                if ct > 0:
                    for t in range(T):
                        u = x_sol[xi(si, mi, t)] / ct
                        units += u
                        per_sku_units[si] += u
        slack_total = sum(x_sol[sli(si, t)] for si in range(S) for t in range(T))
        # SKUs that actually produced ≥10 units (real campaign, not noise)
        distinct_produced = int((per_sku_units >= 10).sum())
        return x_sol, res.fun, {
            "units": units, "slack": slack_total, "y_sum": int(y.sum()),
            "distinct_produced": distinct_produced,
            "min_campaign_relaxed": min_campaign_relaxed,
        }


# ══════════════════════════════════════════════════════════════════════
# GA OPTIMISER
# ══════════════════════════════════════════════════════════════════════
# Outer optimiser — evolves the binary enablement matrix y[SKU, machine]
# ("which SKUs are active on which press today"). Each GA individual is
# scored by running the LPMinuteSolver on it; fitness = LP objective +
# changeover penalty − diversity bonus. Chromosomes are seeded from
# history_map so random individuals resemble plant practice.
class GeneticOptimiser:
    def __init__(self, lp_solver=None, rng_seed=None):
        self.lp = lp_solver or LPMinuteSolver()
        self.rng = random.Random(rng_seed)

    def optimise(self, sku_list, machines, allow_map, curing_matrix,
                 ct_map, inv_init, locked_pairs, history_map,
                 inv_cap=None, label="GT", co_time_map=None):
        S, M = len(sku_list), len(machines)
        if S == 0 or M == 0:
            return None, None, {}

        m_idx = {m: i for i, m in enumerate(machines)}
        elig = {s: [m for m in allow_map.get(s, ()) if m in m_idx]
                for s in sku_list}

        pop = [self._seed(sku_list, machines, elig, history_map, m_idx)
               for _ in range(GAConfig.POPULATION)]
        self._enforce_locks(pop[0], locked_pairs, sku_list, machines)

        best_y, best_fit, best_x, best_meta = None, math.inf, None, {}
        for gen in range(GAConfig.GENERATIONS):
            scored = []
            for y in pop:
                x, lp_obj, meta = self.lp.solve(
                    y, curing_matrix, ct_map, sku_list, machines,
                    inv_init, locked_pairs, inv_cap=inv_cap,
                    co_time_map=co_time_map,
                )
                fit = self._fitness(lp_obj, y, meta)
                scored.append((fit, y, x, meta))
                if fit < best_fit:
                    best_fit, best_y, best_x, best_meta = fit, y.copy(), x, meta
            scored.sort(key=lambda t: t[0])
            print(f"  [GA-{label}] Gen {gen+1:>2}/{GAConfig.GENERATIONS} | "
                  f"best fit: {scored[0][0]:>14,.1f} | "
                  f"units: {scored[0][3].get('units',0):>7,.0f} | "
                  f"slack: {scored[0][3].get('slack',0):>6,.0f} | "
                  f"distinct: {scored[0][3].get('distinct_produced',0):>3} | "
                  f"y_sum: {scored[0][3].get('y_sum',0):>4}")

            new_pop = [scored[i][1].copy() for i in range(GAConfig.ELITE)]
            while len(new_pop) < GAConfig.POPULATION:
                p1 = self._tournament(scored)
                p2 = self._tournament(scored)
                if self.rng.random() < GAConfig.CROSSOVER_RATE:
                    c1, c2 = self._crossover(p1, p2)
                else:
                    c1, c2 = p1.copy(), p2.copy()
                self._mutate(c1, sku_list, elig, m_idx)
                self._mutate(c2, sku_list, elig, m_idx)
                self._enforce_locks(c1, locked_pairs, sku_list, machines)
                self._enforce_locks(c2, locked_pairs, sku_list, machines)
                new_pop.append(c1)
                if len(new_pop) < GAConfig.POPULATION:
                    new_pop.append(c2)
            pop = new_pop
        return best_y, best_x, best_meta

    def _seed(self, sku_list, machines, elig, history_map, m_idx):
        S, M = len(sku_list), len(machines)
        y = np.zeros((S, M), dtype=np.int8)
        for si, sku in enumerate(sku_list):
            opts = elig.get(sku, [])
            if not opts:
                continue
            weights = [1.0 + math.log1p(history_map.get((m, sku), 0.0))
                       for m in opts]
            lo = min(GAConfig.MIN_MACH_PER_SKU, len(opts))
            hi = min(GAConfig.MAX_MACH_PER_SKU, len(opts))
            k = self.rng.randint(lo, hi) if hi >= lo else len(opts)
            chosen = self._weighted_sample(opts, weights, k)
            for m in chosen:
                y[si, m_idx[m]] = 1
        return y

    def _weighted_sample(self, items, weights, k):
        items, weights = list(items), list(weights)
        out = []
        for _ in range(k):
            if not items: break
            total = sum(weights)
            if total <= 0:
                idx = self.rng.randrange(len(items))
            else:
                r = self.rng.random() * total
                cum = 0.0
                for idx in range(len(items)):
                    cum += weights[idx]
                    if r <= cum: break
            out.append(items[idx]); items.pop(idx); weights.pop(idx)
        return out

    def _enforce_locks(self, y, locked_pairs, sku_list, machines):
        sku_idx = {s: i for i, s in enumerate(sku_list)}
        m_idx   = {m: i for i, m in enumerate(machines)}
        for m_lock, sku_lock in locked_pairs:
            if sku_lock in sku_idx and m_lock in m_idx:
                y[sku_idx[sku_lock], m_idx[m_lock]] = 1

    def _tournament(self, scored):
        cand = self.rng.sample(scored, min(GAConfig.TOURNAMENT, len(scored)))
        return min(cand, key=lambda t: t[0])[1].copy()

    def _crossover(self, p1, p2):
        S = p1.shape[0]
        c1, c2 = p1.copy(), p2.copy()
        for s in range(S):
            if self.rng.random() < 0.5:
                c1[s], c2[s] = p2[s].copy(), p1[s].copy()
        return c1, c2

    def _mutate(self, y, sku_list, elig, m_idx):
        for s_idx, sku in enumerate(sku_list):
            if self.rng.random() > GAConfig.MUTATION_RATE:
                continue
            opts = elig.get(sku, [])
            if not opts: continue
            m = self.rng.choice(opts); mi = m_idx[m]
            y[s_idx, mi] = 1 - y[s_idx, mi]
            if y[s_idx].sum() == 0:
                y[s_idx, m_idx[self.rng.choice(opts)]] = 1

    def _fitness(self, lp_obj, y, meta):
        if lp_obj == math.inf or lp_obj == float("inf"):
            return 1e15
        per_m = y.sum(axis=0)
        cos = float(np.maximum(per_m - 1, 0).sum())
        # Reward SKUs that actually produced ≥10 units, not just y=1.
        # Falls back to y-mask if LP didn't return distinct_produced.
        distinct = meta.get("distinct_produced",
                            int((y.sum(axis=1) > 0).sum()))
        return (lp_obj
                + GAConfig.CO_PENALTY * cos
                - GAConfig.DIVERSITY_BONUS * distinct)


# ══════════════════════════════════════════════════════════════════════
# DEMAND HEURISTIC ASSIGNER  (replaces GeneticOptimiser in bc_lp branch)
# ══════════════════════════════════════════════════════════════════════
# Deterministic, demand-driven machine assignment.
#
# Algorithm (per day):
#   For each SKU sorted by total demand DESC:
#     1. Find eligible machines: allow_map[sku] ∩ machines
#     2. Score by history_count[(machine, sku)] — prefer experienced machines
#     3. Assign top-N machines until demand_minutes ≤ assigned_capacity
#     4. Machines are partially committed — other SKUs share remaining idle time
#
# Returns y[S, M] binary matrix, which is passed directly to LPMinuteSolver.
# ~100 lines vs ~600 lines of GA; deterministic, debuggable, demand-driven.
class DemandHeuristicAssigner:
    """
    Demand-driven heuristic machine assignment — replaces GeneticOptimiser.
    Produces a y[S, M] matrix for LPMinuteSolver.
    """
    MAX_MACH_PER_SKU = 5   # cap matches GA's MAX_MACH_PER_SKU
    MIN_MACH_PER_SKU = 1
    MAX_SKUS_PER_MACHINE = 4  # hard cap to limit CO overhead per machine

    def assign(self, sku_list, machines, allow_map, curing_matrix,
               ct_map, history_map, locked_pairs, co_time_map=None):
        """
        Build y[S, M] assignment matrix.

        Parameters
        ----------
        sku_list      : list[str]  — length S
        machines      : list[str]  — length M
        allow_map     : dict {sku: set(machines)} — eligible machines per SKU
        curing_matrix : ndarray (S, T) — demand per SKU per shift
        ct_map        : dict {machine: cycle_time_min}
        history_map   : dict {(machine, sku): count} — 3-month history score
        locked_pairs  : list[(machine, sku)] — currently-running pairs
        co_time_map   : unused (kept for API compatibility with GA)

        Returns
        -------
        y : ndarray (S, M) int8
        """
        S, M = len(sku_list), len(machines)
        y = np.zeros((S, M), dtype=np.int8)
        if S == 0 or M == 0:
            return y

        m_idx   = {m: i for i, m in enumerate(machines)}
        sku_idx = {s: i for i, s in enumerate(sku_list)}

        # Enforce locks: currently-running machines stay with their SKU.
        for m_lock, sku_lock in locked_pairs:
            if sku_lock in sku_idx and m_lock in m_idx:
                y[sku_idx[sku_lock], m_idx[m_lock]] = 1

        # Demand in machine-minutes per SKU (total across all shifts today).
        T   = curing_matrix.shape[1]
        cap = float(Config.SHIFT_MINS * T)   # full-day capacity per machine

        sku_demand_units = curing_matrix.sum(axis=1)    # (S,) units
        sku_demand_mins  = np.zeros(S, dtype=float)
        for si, sku in enumerate(sku_list):
            elig = [m for m in allow_map.get(sku, []) if m in m_idx]
            if elig:
                avg_ct = sum(ct_map.get(m, 2.0) for m in elig) / len(elig)
            else:
                avg_ct = 2.0
            sku_demand_mins[si] = float(sku_demand_units[si]) * avg_ct

        # Pre-compute per-machine demand eligibility count (how many SKUs can
        # use each machine). Machines with fewer eligible SKUs are more
        # specialised; prefer them to avoid flooding general-purpose machines.
        mach_elig_count = {
            m: sum(1 for si in range(S) if m in allow_map.get(sku_list[si], set()))
            for m in machines
        }

        # Sort SKUs: most-constrained first (fewest eligible GT machines) so
        # restricted SKUs get placed before popular machines fill up; within
        # equal eligibility, highest demand first to prioritise load.
        def _elig_count(si):
            sku = sku_list[si]
            return len([m for m in allow_map.get(sku, []) if m in m_idx])
        order = sorted(range(S), key=lambda si: (_elig_count(si), -sku_demand_mins[si]))

        for si in order:
            sku         = sku_list[si]
            demand_mins = float(sku_demand_mins[si])

            elig = [m for m in allow_map.get(sku, []) if m in m_idx]
            if not elig:
                continue

            # Sort: balance load first, then prefer specialised machines
            # (fewest total eligible SKUs → UNISTAGE before STAGE2-only),
            # then history score, then stable index tiebreak.
            elig.sort(key=lambda m: (
                int(y[:, m_idx[m]].sum()),
                mach_elig_count.get(m, 0),
                -history_map.get((m, sku), 0.0),
                m_idx[m],
            ))

            # Machines needed to cover demand (at least MIN, at most MAX).
            if demand_mins > 0 and cap > 0:
                n_needed = max(self.MIN_MACH_PER_SKU,
                               math.ceil(demand_mins / cap))
            else:
                n_needed = self.MIN_MACH_PER_SKU
            n_needed = min(n_needed, self.MAX_MACH_PER_SKU, len(elig))

            already = int(y[si].sum())   # count locked machines already set

            # Prefer machines under the per-machine SKU cap; fall back to any
            # eligible machine if all are at cap (so no SKU is left unassigned).
            cap_limit = self.MAX_SKUS_PER_MACHINE
            elig_under_cap = [m for m in elig if int(y[:, m_idx[m]].sum()) < cap_limit]
            elig_to_use = elig_under_cap if elig_under_cap else elig

            for m in elig_to_use:
                if already >= n_needed:
                    break
                mi = m_idx[m]
                if y[si, mi] == 1:
                    continue            # lock already set this entry
                y[si, mi] = 1
                already  += 1

        return y


# ══════════════════════════════════════════════════════════════════════
# Allocation extractor
# ══════════════════════════════════════════════════════════════════════
def lp_x_to_alloc(x, sku_list, machines, ct_map):
    S, M, T = len(sku_list), len(machines), Config.SHIFTS_PER_DAY
    rows = []
    def xi(s, m, t): return s * M * T + m * T + t
    for si, sku in enumerate(sku_list):
        for mi, mach in enumerate(machines):
            ct = ct_map.get(mach, 2.0)
            for t in range(T):
                mins = float(x[xi(si, mi, t)])
                if mins < 1e-3 or ct <= 0: continue
                units = mins / ct
                if units < 1.0: continue
                rows.append({
                    "Machine":   str(mach),
                    "SKUCode":   str(sku),
                    "Units":     int(round(units)),
                    "Mins_Used": round(mins, 1),
                    "CT_Min":    round(ct, 3),
                    "ShiftIdx":  int(t),
                })
    return pd.DataFrame(rows, columns=[
        "Machine","SKUCode","Units","Mins_Used","CT_Min","ShiftIdx"
    ])


# ══════════════════════════════════════════════════════════════════════
# HYBRID DAILY SCHEDULER
# ══════════════════════════════════════════════════════════════════════
class HybridDailyScheduler:
    """End-to-end scheduler. Runs one Heuristic + LP pass per plan-day,
    rolling GT/carcass inventory and running-machine state forward between days.

    Machine assignment: DemandHeuristicAssigner (demand-driven, deterministic).
    Minute allocation:  LPMinuteSolver (continuous LP, HiGHS backend).
    """

    def run(self, df_curing, df_gt_inv, df_carcass_inv, df_allow, co_map,
            sku_to_size, df_running, plan_start,
            history_map=None):
        # Pipeline (once per day of horizon):
        #   1. Slice today's curing demand, subtract opening GT inv.
        #   2. Inch-lock Stage1/Unistage presses (coverage-then-balance).
        #   3. Heuristic assignment → LP → today's GT (Stage2+Unistage) alloc.
        #   4. Derive Stage-1 carcass demand from Stage-2 output.
        #   5. Heuristic assignment → LP again → today's Stage-1 (carcass) alloc.
        #   6. Sequence campaigns per press, insert CHANGEOVER rows.
        #   7. TopUp: fill idle tails with pre-build for future demand.
        #   8. Anti-starvation validation (WIP balance per shift).
        #   9. Roll inventories + running-machine state into next day.
        history_map = history_map or {}
        full_days = Config.PLANNING_DAYS
        if full_days <= 0:
            raise ValueError("PLANNING_DAYS must be ≥ 1")

        gt_inv = (df_gt_inv.copy() if df_gt_inv is not None
                  else pd.DataFrame(columns=["SKUCode","GT_Inventory"]))
        s1_inv = (df_carcass_inv.copy() if df_carcass_inv is not None
                  else pd.DataFrame(columns=["SKUCode","Carcass_Inventory"]))
        running = (df_running.copy() if df_running is not None
                   else pd.DataFrame(columns=["Machine","SKUCode"]))

        df_curing = df_curing.copy()
        df_curing["StartTime"] = pd.to_datetime(df_curing["StartTime"])
        df_curing["EndTime"]   = pd.to_datetime(df_curing["EndTime"])

        allow_map = {}
        for _, r in df_allow.iterrows():
            allow_map[str(r["SKUCode"])] = set(map(str, r.get("Machines", [])))

        ct_map = {m: Config.ct_min(m)
                  for m in Config.STAGE1 | Config.STAGE2 | Config.UNISTAGE}

        # Horizon-wide curing demand per SKU (used as topup target so idle
        # machine tails build inventory for future days' curing). Decreases
        # as each day's real production + topup production completes.
        self._gt_remaining = (
            df_curing.groupby("SKUCode")["Qty"].sum().to_dict()
        )
        # Stage1 topup target — approximated as horizon GT demand of SKUs
        # that will be produced on Stage2 downstream (they need carcass).
        # For simplicity, mirror the GT target initially and let stage1_roll
        # drain it as Stage1 actually builds carcass.
        self._s1_remaining = dict(self._gt_remaining)

        # Per-machine representative changeover time (min). LP reserves
        # co_time × (y_sum[m] − 1) minutes out of machine capacity so what
        # it plans survives the ScheduleBuilder inserting CHANGEOVER rows.
        co_time_map = {}
        for m in Config.STAGE1 | Config.STAGE2 | Config.UNISTAGE:
            entry = co_map.get(str(m), {"same": 40, "diff": 60})
            # use "diff" as a conservative worst case — safer vs starvation
            co_time_map[str(m)] = float(entry.get("diff", 60))

        agg = {k: [] for k in (
            "gt_allocation","s1_allocation",
            "shift_schedule","gt_shift_schedule","s1_shift_schedule",
            "starvation_report",
        )}

        for d in range(full_days):
            day_start = plan_start + timedelta(days=d)
            day_end   = day_start + timedelta(days=1)
            print("\n" + "█" * 72)
            print(f"█  [v8 Hybrid] DAY {d+1}/{full_days}  —  {day_start:%Y-%m-%d}")
            print("█" * 72)

            # TODAY's curing: used for starvation validation and inventory roll.
            mask = ((df_curing["StartTime"] >= day_start)
                    & (df_curing["StartTime"] < day_end))
            day_curing = df_curing.loc[mask].copy()
            if day_curing.empty:
                print(f"  [Day {d+1}] No curing — skipping.")
                continue

            # LEAD window: LP/GA targets curing demand BUILD_LEAD_SHIFTS ahead.
            # 3 shifts = 1 full day, so day D builds for day D+1 curing demand.
            lead_delta = timedelta(
                hours=Config.BUILD_LEAD_SHIFTS * Config.HOURS_PER_SHIFT
            )
            build_target_start = day_start + lead_delta
            build_target_end   = day_end   + lead_delta
            mask_lead = ((df_curing["StartTime"] >= build_target_start)
                         & (df_curing["StartTime"] < build_target_end))
            day_curing_lead = df_curing.loc[mask_lead].copy()
            # Last planning day has no lead-window data; fall back to today.
            if day_curing_lead.empty:
                day_curing_lead    = day_curing
                build_target_start = day_start

            # 1-day demand
            prev_days = Config.PLANNING_DAYS
            Config.PLANNING_DAYS = 1
            try:
                deriver = DemandDeriver()
                # LP/GA demand matrix: derived from lead window (next day's curing)
                df_sku_demand, cur_mat, sku_list, _ = (
                    deriver.derive(day_curing_lead, gt_inv, build_target_start)
                )
                # Shift-demand for starvation validator: today's actual curing
                _, _, _, df_shift_demand_day = (
                    deriver.derive(day_curing, gt_inv, day_start)
                )
            finally:
                Config.PLANNING_DAYS = prev_days

            gt_inv_init  = dict(zip(gt_inv["SKUCode"], gt_inv["GT_Inventory"]))
            carc_inv_init = (
                dict(zip(s1_inv["SKUCode"], s1_inv["Carcass_Inventory"]))
                if not s1_inv.empty else {}
            )

            # gt_inv_for_cap feeds the LP starvation constraint (lower bound) so
            # the LP knows how much WIP will remain after today's curing.
            # The PRODUCTION CAP now uses cap_on_gross=True (gross×buf) and
            # ignores wip0 entirely, preventing cap-collapse when DB inventory
            # is large relative to the lead-window demand.
            if Config.BUILD_LEAD_SHIFTS > 0:
                _today_cure_qty = day_curing.groupby("SKUCode")["Qty"].sum().to_dict()
                gt_inv_for_cap = {
                    sku: max(0.0, qty - float(_today_cure_qty.get(sku, 0.0)))
                    for sku, qty in gt_inv_init.items()
                }
            else:
                gt_inv_for_cap = gt_inv_init

            locked = {}
            for _, r in running.iterrows():
                locked[str(r["Machine"])] = str(r["SKUCode"])
            extra = [s for s in set(locked.values()) if s not in sku_list]
            if extra:
                sku_list = sku_list + extra
                cur_mat = np.vstack([cur_mat, np.zeros((len(extra), cur_mat.shape[1]))])

            # Inch-lock for Stage1 + Unistage
            #   - machine_size from currently-running SKU
            #   - SMART assignment for empty restricted machines:
            #       Phase 1 (coverage): every demanded size that *can* be served
            #                            on a restricted machine gets ≥1 machine.
            #       Phase 2 (balance):  remaining machines go to whichever size
            #                            has the largest leftover demand-minutes
            #                            per machine already locked to it.
            #   - per-SKU eligibility shrunk to size-compatible machines
            machine_size = {}
            for m, s in locked.items():
                machine_size[str(m)] = sku_to_size.get(str(s), "")
            restricted = Config.STAGE1 | Config.UNISTAGE
            sku_total_demand = cur_mat.sum(axis=1)

            # ---- per-size demand-minutes (averaged across eligible machines)
            #      and which restricted machines can run that size ----
            size_eligible_machines = defaultdict(set)   # sz → {machines that CAN run sz}
            size_demand_min        = defaultdict(float) # sz → required minutes
            for si, sku in enumerate(sku_list):
                sz  = sku_to_size.get(str(sku))
                dem = float(sku_total_demand[si])
                if not sz or dem <= 0:
                    continue
                elig_for_sku = allow_map.get(str(sku), set()) & restricted
                if not elig_for_sku:
                    continue
                # avg ct across this SKU's eligible restricted machines
                cts = [ct_map.get(m, 2.0) for m in elig_for_sku]
                avg_ct = sum(cts) / len(cts)
                size_demand_min[sz] += dem * avg_ct
                for m in elig_for_sku:
                    size_eligible_machines[sz].add(m)

            cap_per_machine = float(Config.SHIFT_MINS * Config.SHIFTS_PER_DAY)

            # how many machines are already locked to each size (from running)
            machines_per_size = defaultdict(int)
            for s in machine_size.values():
                if s:
                    machines_per_size[s] += 1

            unassigned = [m for m in sorted(restricted) if not machine_size.get(m)]

            # ---- Phase 1: coverage — give each uncovered demanded size ≥1 machine
            sizes_by_constraint = sorted(
                size_eligible_machines.keys(),
                key=lambda sz: (len(size_eligible_machines[sz]),
                                -size_demand_min[sz]),
            )
            for sz in sizes_by_constraint:
                if machines_per_size[sz] >= 1:
                    continue
                cands = [m for m in unassigned if m in size_eligible_machines[sz]]
                if not cands:
                    continue
                # pick the LEAST-flexible candidate (eligible for fewest sizes)
                cands.sort(
                    key=lambda m: sum(1 for sz2 in size_eligible_machines
                                      if m in size_eligible_machines[sz2])
                )
                m = cands[0]
                machine_size[m] = sz
                machines_per_size[sz] += 1
                unassigned.remove(m)

            # ---- Phase 2: balance — fill remaining machines into the size with
            # the highest deficit (demand_min − current capacity)
            while unassigned:
                best, best_def = None, -float("inf")
                for m in unassigned:
                    for sz, elig in size_eligible_machines.items():
                        if m not in elig:
                            continue
                        deficit = (size_demand_min[sz]
                                   - machines_per_size[sz] * cap_per_machine)
                        if deficit > best_def:
                            best_def, best = deficit, (m, sz)
                if best is None:
                    break
                m, sz = best
                machine_size[m] = sz
                machines_per_size[sz] += 1
                unassigned.remove(m)

            # ---- Phase 3: per-SKU rescue ─ for each demanded SKU, ensure
            # at least one of ITS OWN eligible restricted machines is locked
            # to its size. If none is, try to flip a currently-locked-but-
            # reassignable machine from another size. A machine is
            # "reassignable" if:
            #   (i) it is not carrying a currently-running SKU (not in `locked`)
            #   (ii) its current locked size has ≥2 machines (so flipping it
            #        doesn't strand the origin size).
            #   (iii) it is in this SKU's eligible-machine set.
            stranded_skus = []
            for si, sku in enumerate(sku_list):
                dem = float(sku_total_demand[si])
                if dem <= 0:
                    continue
                s_sz = sku_to_size.get(str(sku))
                if not s_sz:
                    continue
                elig_for_sku = allow_map.get(str(sku), set()) & restricted
                if not elig_for_sku:
                    # SKU has no restricted-machine eligibility at all; it
                    # can only be served on Stage2 (or nowhere). Skip.
                    continue
                # Is at least one eligible machine locked to this SKU's size?
                covered = any(machine_size.get(m) == s_sz
                              for m in elig_for_sku)
                if covered:
                    continue
                # Try to rescue — find a reassignable eligible machine.
                rescued = False
                for m in sorted(elig_for_sku):
                    cur_sz = machine_size.get(m, "")
                    if not cur_sz:
                        # Not yet locked (edge case) — just set it.
                        machine_size[m] = s_sz
                        machines_per_size[s_sz] += 1
                        rescued = True
                        break
                    if m in locked:
                        continue  # currently running a real SKU, can't flip
                    if machines_per_size.get(cur_sz, 0) <= 1:
                        continue  # origin size would be stranded
                    # Safe to flip
                    machines_per_size[cur_sz] -= 1
                    machine_size[m] = s_sz
                    machines_per_size[s_sz] = machines_per_size.get(s_sz, 0) + 1
                    rescued = True
                    break
                if not rescued:
                    stranded_skus.append({
                        "sku": sku, "size": s_sz, "demand": int(dem),
                        "eligible_restricted": sorted(elig_for_sku),
                    })

            covered_sizes  = sum(1 for sz in size_eligible_machines
                                 if machines_per_size[sz] > 0)
            total_demanded = len(size_eligible_machines)
            print(
                f"  [Inch-lock] Restricted machines locked: "
                f"{sum(1 for m in restricted if machine_size.get(m))}/{len(restricted)} "
                f"| Sizes covered: {covered_sizes}/{total_demanded}"
            )
            if stranded_skus:
                print(
                    f"  [STRANDED] {len(stranded_skus)} SKU(s) have no "
                    f"size-matching restricted machine after rescue — "
                    f"their curing demand will starve:"
                )
                for info in stranded_skus[:15]:  # cap log length
                    print(
                        f"    • {info['sku']:<26} size={info['size']:<8} "
                        f"demand={info['demand']:>5} "
                        f"eligibles={info['eligible_restricted']}"
                    )
                if len(stranded_skus) > 15:
                    print(f"    ... and {len(stranded_skus) - 15} more")

            def _filter_for_locks(am):
                """Drop restricted machines whose locked size ≠ SKU's size."""
                out = {}
                for sku, machines_set in am.items():
                    s_sz = sku_to_size.get(str(sku))
                    keep = set()
                    for mch in machines_set:
                        if mch in restricted:
                            m_sz = machine_size.get(str(mch))
                            if m_sz and s_sz and m_sz != s_sz:
                                continue
                        keep.add(mch)
                    out[sku] = keep
                return out

            # GT  — Heuristic assignment → LP minute allocation
            gt_machines  = sorted(Config.STAGE2 | Config.UNISTAGE)
            allow_map_gt = _filter_for_locks(allow_map)
            gt_locked    = [(m, s) for m, s in locked.items() if m in gt_machines]

            heuristic_gt = DemandHeuristicAssigner()
            best_y_gt    = heuristic_gt.assign(
                sku_list, gt_machines, allow_map_gt, cur_mat, ct_map,
                history_map or {}, gt_locked, co_time_map=co_time_map,
            )
            lp_gt = LPMinuteSolver()
            best_x_gt, _, meta_gt = lp_gt.solve(
                best_y_gt, cur_mat, ct_map, sku_list, gt_machines,
                gt_inv_for_cap, gt_locked, co_time_map=co_time_map,
                cap_on_gross=True,
            )
            df_gt_alloc = (
                lp_x_to_alloc(best_x_gt, sku_list, gt_machines, ct_map)
                if best_x_gt is not None else pd.DataFrame()
            )
            print(f"  [Heuristic-GT] units={meta_gt.get('units',0):,.0f}  "
                  f"slack={meta_gt.get('slack',0):,.0f}  "
                  f"distinct SKUs="
                  f"{int((best_y_gt.sum(axis=1)>0).sum())}  "
                  f"y_sum={int(best_y_gt.sum())}")

            # Carcass demand from Stage2 only
            if not df_gt_alloc.empty:
                s2_alloc = df_gt_alloc[df_gt_alloc["Machine"].isin(Config.STAGE2)]
            else:
                s2_alloc = pd.DataFrame(columns=["SKUCode","Units"])
            carc_raw = s2_alloc.groupby("SKUCode")["Units"].sum().reset_index()
            carc_raw.columns = ["SKUCode","Raw_Carcass_Demand"]
            carc_raw["Carcass_Inventory"] = carc_raw["SKUCode"].map(carc_inv_init).fillna(0)
            carc_raw["Net_Carcass_Demand"] = (
                carc_raw["Raw_Carcass_Demand"] - carc_raw["Carcass_Inventory"]
            ).clip(lower=0).astype(int)
            df_carc_valid = carc_raw[carc_raw["Net_Carcass_Demand"] > 0].copy()

            # Stage1 GA + LP
            s1_machines = sorted(Config.STAGE1)
            s1_sku_list = list(df_carc_valid["SKUCode"])
            for m, s in locked.items():
                if m in s1_machines and s not in s1_sku_list:
                    s1_sku_list.append(s)

            df_s1_alloc = pd.DataFrame()
            if s1_sku_list:
                T_day = Config.SHIFTS_PER_DAY
                s1_mat = np.zeros((len(s1_sku_list), T_day))
                carc_dem_map = dict(zip(
                    df_carc_valid["SKUCode"], df_carc_valid["Net_Carcass_Demand"]
                ))
                for si, sku in enumerate(s1_sku_list):
                    dem = float(carc_dem_map.get(sku, 0))
                    s1_mat[si, :] = dem / T_day

                s1_locked    = [(m, s) for m, s in locked.items() if m in s1_machines]
                allow_map_s1 = _filter_for_locks(allow_map)

                heuristic_s1 = DemandHeuristicAssigner()
                y_s1 = heuristic_s1.assign(
                    s1_sku_list, s1_machines, allow_map_s1, s1_mat, ct_map,
                    history_map or {}, s1_locked, co_time_map=co_time_map,
                )
                lp_s1 = LPMinuteSolver()
                best_x_s1, _, meta_s1 = lp_s1.solve(
                    y_s1, s1_mat, ct_map, s1_sku_list, s1_machines,
                    carc_inv_init, s1_locked, co_time_map=co_time_map,
                )
                if best_x_s1 is not None:
                    df_s1_alloc = lp_x_to_alloc(
                        best_x_s1, s1_sku_list, s1_machines, ct_map,
                    )

            # Sequence + build
            urgency_map = dict(zip(
                df_sku_demand["SKUCode"], df_sku_demand["First_Curing_Start"]
            ))
            sequencer = CampaignSequencer(
                co_map, sku_to_size, locked, urgency_map, day_start,
            )
            gt_seq = sequencer.sequence(df_gt_alloc) if not df_gt_alloc.empty else {}
            s1_seq = sequencer.sequence(df_s1_alloc) if not df_s1_alloc.empty else {}

            locked_times = {m: day_start for m in locked}
            builder = ScheduleBuilder(day_start, co_map)
            df_gt_sched = builder.build(gt_seq, locked, locked_times,
                                        sku_to_size, "GT BUILD")
            df_s1_sched = builder.build(s1_seq, locked, locked_times,
                                        sku_to_size, "CARCASS")

            # ---- TOPUP idle tail per machine ──────────────────────────
            # ScheduleBuilder stops a machine when its last campaign ends.
            # If that's well before day_end (e.g. 2 am vs 7 am next day),
            # the press sits idle for hours. Use that gap to pre-build only
            # the curing due within TOPUP_LOOKAHEAD_DAYS ahead, net of what
            # opening inventory + today's own build already cover. SKUs that
            # are already covered get a 0 target (no stockpiling); idle
            # capacity then only chases SKUs still short (starving).
            def _future_curing(n_days):
                end = day_end + timedelta(days=n_days)
                return (
                    df_curing.loc[(df_curing["StartTime"] >= day_end)
                                  & (df_curing["StartTime"] < end)]
                    .groupby("SKUCode")["Qty"].sum().to_dict()
                )
            # GT may be pre-built up to 3 days ahead; carcass only 1 day.
            future_dem_gt = _future_curing(Config.TOPUP_LOOKAHEAD_DAYS_GT)
            future_dem_s1 = _future_curing(Config.TOPUP_LOOKAHEAD_DAYS_CARCASS)
            today_cure = day_curing.groupby("SKUCode")["Qty"].sum().to_dict()
            gt_prod_now = (
                df_gt_sched.loc[~df_gt_sched["SKUCode"].isin(
                    ["CHANGEOVER", "MOULD_CLEAN"])]
                .groupby("SKUCode")["Qty"].sum().to_dict()
                if df_gt_sched is not None and not df_gt_sched.empty else {}
            )
            s1_prod_now = (
                df_s1_sched.loc[~df_s1_sched["SKUCode"].isin(
                    ["CHANGEOVER", "MOULD_CLEAN"])]
                .groupby("SKUCode")["Qty"].sum().to_dict()
                if df_s1_sched is not None and not df_s1_sched.empty else {}
            )
            gt_inv_now = (dict(zip(gt_inv["SKUCode"], gt_inv["GT_Inventory"]))
                          if gt_inv is not None and not gt_inv.empty else {})
            s1_inv_now = (dict(zip(s1_inv["SKUCode"], s1_inv["Carcass_Inventory"]))
                          if s1_inv is not None and not s1_inv.empty else {})
            # GT target: 3-day future curing not yet covered by projected
            # end-of-day GT inventory (opening + today's build − today's use).
            gt_topup_target = {
                s: max(0.0, need - (gt_inv_now.get(s, 0.0)
                                    + gt_prod_now.get(s, 0.0)
                                    - today_cure.get(s, 0.0)))
                for s, need in future_dem_gt.items()
            }
            # Carcass target: each future GT needs a carcass, but carcass max
            # age is 1 day → cover only the next 1 day of GT demand not yet
            # held as carcass inventory + today's carcass build.
            s1_topup_target = {
                s: max(0.0, need - (s1_inv_now.get(s, 0.0)
                                    + s1_prod_now.get(s, 0.0)))
                for s, need in future_dem_s1.items()
            }

            df_gt_sched = self._topup_idle_tail(
                df_gt_sched, day_start, day_end, locked, machine_size,
                sku_to_size, allow_map, ct_map, co_map, co_time_map,
                gt_topup_target,
                machine_group=set(Config.STAGE2) | set(Config.UNISTAGE),
                stage_label="GT TOPUP",
            )
            # stage1(carcass) schedule generation

            df_s1_sched = self._topup_idle_tail(
                df_s1_sched, day_start, day_end, locked, machine_size,
                sku_to_size, allow_map, ct_map, co_map, co_time_map,
                s1_topup_target,
                machine_group=set(Config.STAGE1),
                stage_label="CARCASS TOPUP",
            )

            df_all = pd.concat([df_gt_sched, df_s1_sched], ignore_index=True)
            if not df_all.empty:
                df_all["MachineType"] = df_all["Machine"].apply(
                    lambda m: ("STAGE1" if m in Config.STAGE1
                               else "STAGE2" if m in Config.STAGE2
                               else "UNISTAGE" if m in Config.UNISTAGE
                               else "UNKNOWN")
                )
                df_all = df_all.sort_values(["Machine","StartTime"]).reset_index(drop=True)

            df_starv = (
                StarvationValidator().validate(
                    df_gt_sched, df_shift_demand_day, gt_inv,
                ) if not df_shift_demand_day.empty else pd.DataFrame()
            )

            for k, df in (
                ("gt_allocation", df_gt_alloc),
                ("s1_allocation", df_s1_alloc),
                ("shift_schedule", df_all),
                ("gt_shift_schedule", df_gt_sched),
                ("s1_shift_schedule", df_s1_sched),
                ("starvation_report", df_starv),
            ):
                if df is not None and not df.empty:
                    agg[k].append(df)

            # Roll inventory (includes TOPUP production because topup rows
            # are appended to df_gt_sched / df_s1_sched above)
            gt_prod = (
                df_gt_sched.loc[~df_gt_sched["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"]),
                                ["SKUCode","Qty"]]
                .groupby("SKUCode")["Qty"].sum()
                if not df_gt_sched.empty else pd.Series(dtype=float)
            )
            gt_cure = day_curing.groupby("SKUCode")["Qty"].sum()
            gt_inv  = self._roll(gt_inv, "GT_Inventory", gt_prod, gt_cure)

            # Drain horizon-remaining for next day's topup (includes both
            # LP and TOPUP production because gt_prod sums the full schedule)
            for sku, qty in gt_prod.items():
                self._gt_remaining[sku] = max(
                    0.0, self._gt_remaining.get(sku, 0.0) - float(qty)
                )

            s1_prod = (
                df_s1_sched.loc[~df_s1_sched["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"]),
                                ["SKUCode","Qty"]]
                .groupby("SKUCode")["Qty"].sum()
                if not df_s1_sched.empty else pd.Series(dtype=float)
            )
            s2_consumed = (
                df_gt_alloc.loc[df_gt_alloc["Machine"].isin(Config.STAGE2),
                                ["SKUCode","Units"]]
                .groupby("SKUCode")["Units"].sum()
                if not df_gt_alloc.empty else pd.Series(dtype=float)
            )
            s1_inv = self._roll(s1_inv, "Carcass_Inventory", s1_prod, s2_consumed)
            # Hard shelf-life clip: carcass max age is 1 day, so no SKU may
            # carry more carcass into tomorrow than tomorrow's demand. Any
            # surplus (from Stage-1 LP overproduction or uncounted Stage-2
            # topup consumption) is scrapped/aged-out here, guaranteeing
            # end-of-day carcass never exceeds 1-day cover.
            if s1_inv is not None and not s1_inv.empty:
                _scrapped = 0
                _new_carc = []
                for _, _r in s1_inv.iterrows():
                    _have = int(_r["Carcass_Inventory"])
                    _cap  = int(future_dem_s1.get(_r["SKUCode"], 0))
                    if _have > _cap:
                        _scrapped += _have - _cap
                    _new_carc.append(min(_have, _cap))
                s1_inv["Carcass_Inventory"] = _new_carc
                if _scrapped > 0:
                    print(f"  [Carcass shelf-life] Scrapped {_scrapped:,} "
                          f"aged carcass units (>1-day cover).")

            for sku, qty in s1_prod.items():
                self._s1_remaining[sku] = max(
                    0.0, self._s1_remaining.get(sku, 0.0) - float(qty)
                )

            # Only carry forward machine locks from Day 0 (the real-time DB snapshot).
            # From Day 1 onward, all machines are free to be reassigned by the
            # heuristic. Without this, all Unistage machines stay size-locked to
            # their Day-0 end-SKU for the full 30-day horizon, starving SKUs of
            # different sizes from using those machines.
            if d == 0:
                running = self._update_running(df_all, running)
            else:
                running = pd.DataFrame(columns=["Machine", "SKUCode"])

            print(f"  [Day {d+1}] EOD GT inv: {int(gt_inv['GT_Inventory'].sum()):,}  |  "
                  f"Carcass inv: {int(s1_inv['Carcass_Inventory'].sum()):,}")

        # Aggregate
        def cat(lst):
            non = [df for df in lst if df is not None and not df.empty]
            return pd.concat(non, ignore_index=True) if non else pd.DataFrame()

        df_gt_alloc = cat(agg["gt_allocation"])
        df_s1_alloc = cat(agg["s1_allocation"])
        df_all      = cat(agg["shift_schedule"])
        df_gt_sched = cat(agg["gt_shift_schedule"])
        df_s1_sched = cat(agg["s1_shift_schedule"])
        df_stv      = cat(agg["starvation_report"])

        # Reports — full-horizon
        deriver = DemandDeriver()
        df_sku_demand_full, curing_matrix_full, sku_list_full, _ = deriver.derive(
            df_curing, df_gt_inv, plan_start,
        )
        df_summary       = build_summary(df_sku_demand_full, df_all)
        df_util_gt       = build_util(df_gt_sched,
                                      sorted(Config.STAGE2 | Config.UNISTAGE))
        df_util_s1       = build_util(df_s1_sched, sorted(Config.STAGE1))
        df_dw            = build_daywise_report(pd.DataFrame(), df_gt_sched, df_s1_sched)
        df_dcnt          = build_daily_sku_counts(df_gt_sched, df_s1_sched)

        # ── 3-day total summary ──────────────────────────────────────
        self._print_horizon_summary(
            df_all, df_gt_sched, df_s1_sched, df_summary, df_stv,
            df_util_gt, df_util_s1, df_dcnt, df_curing, plan_start, full_days,
        )

        return {
            "gt_allocation":      df_gt_alloc,
            "s1_allocation":      df_s1_alloc,
            "shift_schedule":     df_all,
            "gt_shift_schedule":  df_gt_sched,
            "s1_shift_schedule":  df_s1_sched,
            "demand_summary":     df_summary,
            "starvation_report":  df_stv,
            "gt_utilization":     df_util_gt,
            "s1_utilization":     df_util_s1,
            "daywise_report":     df_dw,
            "daily_sku_counts":   df_dcnt,
            "carcass_demand":     pd.DataFrame(),
            "curing_input":       df_curing,
            "curing_matrix":      pd.DataFrame(
                curing_matrix_full, index=sku_list_full,
                columns=[f"Shift{t}" for t in range(curing_matrix_full.shape[1])],
            ),
            "sku_to_size":        sku_to_size,
        }

    @staticmethod
    def _split_row(start, end, machine, sku, total_units, ct_min, remarks):
        """Split a (start, end) block into per-shift rows (mirrors
        ScheduleBuilder._split_row so topup rows match the existing format)."""
        rows, curr, rem = [], start, total_units
        total_dur = max((end - start).total_seconds() / 60.0, 1e-6)
        while curr < end:
            shift, shift_end = _shift_fn(curr)
            slice_end = min(shift_end, end)
            dur = (slice_end - curr).total_seconds() / 60.0
            if dur <= 0:
                curr = slice_end
                continue
            if total_units == 0 or sku == "CHANGEOVER":
                qty = 0
            elif slice_end == end:
                qty = rem
            else:
                qty = (int(dur / ct_min) if ct_min > 0
                       else int(total_units * dur / total_dur))
                qty = min(qty, rem)
            rows.append({
                "Date": curr.date(), "Shift": shift,
                "Machine": machine, "SKUCode": sku,
                "StartTime": curr, "EndTime": slice_end,
                "Qty": qty, "CT_Min": round(ct_min, 2),
                "Remarks": remarks,
            })
            rem -= qty
            curr = slice_end
        return rows

    def _topup_idle_tail(self, df_sched, day_start, day_end, locked,
                         machine_size, sku_to_size, allow_map, ct_map,
                         co_map, co_time_map, remaining_demand,
                         machine_group, stage_label):
        """
        For each machine in `machine_group`, find the idle window between
        its last campaign end and day_end. If long enough for MIN_CAMPAIGN
        (+ any CO needed), add a pre-build campaign for the eligible SKU
        with the highest leftover horizon demand. Same-size-as-last SKU is
        preferred so no CO is required.
        """
        day_end_ts = pd.to_datetime(day_end)
        machine_group = {str(m) for m in machine_group}

        # Work on a scratch copy so `remaining_demand` (a reference to
        # self._gt_remaining / self._s1_remaining) isn't double-decremented
        # — the rolling step later drains it based on the full day schedule
        # which already includes these topup rows.
        scratch = {s: float(v) for s, v in remaining_demand.items()}

        new_rows = []
        topups_added = 0

        for mach in sorted(machine_group):
            # What's the last activity on this press today?
            if df_sched is None or df_sched.empty:
                last_end = pd.to_datetime(day_start)
                last_sku = locked.get(mach)
            else:
                m_rows = df_sched[df_sched["Machine"] == mach]
                if m_rows.empty:
                    last_end = pd.to_datetime(day_start)
                    last_sku = locked.get(mach)
                else:
                    last_end = pd.to_datetime(m_rows["EndTime"].max())
                    prod = m_rows[~m_rows["SKUCode"].isin(
                        ["CHANGEOVER", "MOULD_CLEAN"]
                    )]
                    last_sku = (prod.sort_values("StartTime")["SKUCode"].iloc[-1]
                                if not prod.empty else locked.get(mach))

            idle_min = (day_end_ts - last_end).total_seconds() / 60.0
            if idle_min < Config.MIN_CAMPAIGN_MINS:
                continue  # not enough time for a real campaign

            last_sku_size = (sku_to_size.get(str(last_sku), "")
                             if last_sku else "")
            m_sz = machine_size.get(str(mach), "")
            is_restr = (mach in Config.STAGE1) or (mach in Config.UNISTAGE)

            # Candidate SKUs: eligible, size-compatible, and STILL SHORT of
            # their near-term cover (scratch > 0). SKUs already covered are
            # skipped — no arbitrary future stockpiling. Prefer a zero-CO
            # (same-size) pick; otherwise the largest shortfall (the most
            # starved SKU). Built units are capped to the shortfall, so a
            # press never builds beyond the lookahead cover.
            cands = []
            for sku, rem_qty in scratch.items():
                if rem_qty <= 0:
                    continue
                if mach not in allow_map.get(str(sku), set()):
                    continue
                s_sz = sku_to_size.get(str(sku), "")
                if is_restr and m_sz and s_sz and m_sz != s_sz:
                    continue
                same_size = (s_sz == last_sku_size and last_sku_size != "")
                cands.append((sku, float(rem_qty), same_size, s_sz))
            if not cands:
                continue
            cands.sort(key=lambda c: (c[2], c[1]), reverse=True)

            ct = ct_map.get(mach, 2.0)
            best_sku = best_sku_size = None
            co_dur = 0.0
            units = 0
            prod_min = 0.0
            for sku, rem_qty, same_size, s_sz in cands:
                if last_sku and last_sku != sku:
                    co_entry = co_map.get(str(mach), {"same": 40, "diff": 60})
                    cand_co = (co_entry["same"]
                               if last_sku_size == s_sz else co_entry["diff"])
                else:
                    cand_co = 0.0
                cand_prod_min = idle_min - cand_co
                if cand_prod_min < Config.MIN_CAMPAIGN_MINS:
                    continue  # CO eats too much of the idle window
                cand_units = int(cand_prod_min / ct) if ct > 0 else 0
                cand_units = min(cand_units, int(rem_qty))  # cap to shortfall
                if cand_units < Config.MIN_CAMPAIGN_UNITS:
                    continue
                best_sku, best_sku_size = sku, s_sz
                co_dur, units = cand_co, cand_units
                prod_min = units * ct
                break

            if best_sku is None:
                continue

            cursor = last_end
            if co_dur > 0:
                co_end = cursor + timedelta(minutes=co_dur)
                new_rows.extend(self._split_row(
                    cursor, co_end, mach, "CHANGEOVER", 0, 0.0,
                    f"C/O → {best_sku} (TOPUP)"
                ))
                cursor = co_end

            prod_end = cursor + timedelta(minutes=prod_min)
            new_rows.extend(self._split_row(
                cursor, prod_end, mach, best_sku, units, ct, stage_label,
            ))

            # Update the scratch copy only — so the next machine this
            # same day doesn't pick the same SKU beyond its remaining.
            # Cross-day accounting happens in the rolling step.
            scratch[best_sku] = max(0.0, scratch.get(best_sku, 0.0) - units)
            topups_added += 1

        if topups_added:
            print(
                f"  [TopUp-{stage_label.split()[0]}] Filled {topups_added} "
                f"idle machine tail(s) with pre-build campaigns."
            )

        if not new_rows:
            return df_sched
        df_topup = pd.DataFrame(new_rows)
        out = (pd.concat([df_sched, df_topup], ignore_index=True)
               if df_sched is not None and not df_sched.empty
               else df_topup)
        return out.sort_values(["Machine", "StartTime"]).reset_index(drop=True)

    @staticmethod
    def _roll(df_inv, col, produced, consumed):
        base = (dict(zip(df_inv["SKUCode"], df_inv[col]))
                if df_inv is not None and not df_inv.empty else {})
        skus = set(base) | set(produced.index) | set(consumed.index)
        rows = []
        for s in skus:
            old = float(base.get(s, 0))
            new = old + float(produced.get(s, 0)) - float(consumed.get(s, 0))
            rows.append({"SKUCode": s, col: int(round(max(0.0, new)))})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["SKUCode", col])

    @staticmethod
    def _update_running(df_sched, prev_running):
        if df_sched is None or df_sched.empty:
            return prev_running
        prod = df_sched[~df_sched["SKUCode"].isin(["CHANGEOVER"])].copy()
        if prod.empty:
            return prev_running
        prod = prod.sort_values(["Machine","StartTime"])
        last = prod.groupby("Machine").tail(1)[["Machine","SKUCode"]]
        return last.reset_index(drop=True)

    @staticmethod
    def _print_horizon_summary(df_all, df_gt_sched, df_s1_sched, df_summary,
                                df_stv, df_util_gt, df_util_s1, df_dcnt,
                                df_curing, plan_start, full_days):
        """Terminal-only consolidated 3-day summary."""
        bar = "═" * 74

        # Per-day curing demand (distinct + qty)
        cd = df_curing.copy()
        cd["StartTime"] = pd.to_datetime(cd["StartTime"])
        cd["DateOnly"]  = cd["StartTime"].dt.date

        prod = (df_all[~df_all["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])]
                if df_all is not None and not df_all.empty
                else pd.DataFrame(columns=["Date","SKUCode","Qty"]))
        if not prod.empty:
            prod = prod.copy()
            prod["DateOnly"] = pd.to_datetime(prod["Date"]).dt.date

        gt_prod = (df_gt_sched[~df_gt_sched["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])]
                   if df_gt_sched is not None and not df_gt_sched.empty
                   else pd.DataFrame(columns=["Date","SKUCode","Qty"]))
        if not gt_prod.empty:
            gt_prod = gt_prod.copy()
            gt_prod["DateOnly"] = pd.to_datetime(gt_prod["Date"]).dt.date

        s1_prod = (df_s1_sched[~df_s1_sched["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])]
                   if df_s1_sched is not None and not df_s1_sched.empty
                   else pd.DataFrame(columns=["Date","SKUCode","Qty"]))
        if not s1_prod.empty:
            s1_prod = s1_prod.copy()
            s1_prod["DateOnly"] = pd.to_datetime(s1_prod["Date"]).dt.date

        co_count = ((df_all["SKUCode"] == "CHANGEOVER").sum()
                    if df_all is not None and not df_all.empty else 0)

        print(f"\n{bar}")
        print(f"  [v8 Hybrid]  HORIZON SUMMARY  —  "
              f"{plan_start:%Y-%m-%d}  →  +{full_days} day(s)")
        print(bar)

        # Per-day breakdown
        header = (f"  {'Date':<12} {'Cure SKU':>10} {'Cure Qty':>10} "
                  f"{'GT SKU':>8} {'GT Qty':>10} {'S1 Qty':>10} "
                  f"{'COs':>6}")
        print(header)
        print("  " + "-" * (len(header) - 2))

        tot_cure_q = tot_gt_q = tot_s1_q = tot_cos = 0
        for d in range(full_days):
            dt = (plan_start + timedelta(days=d)).date()
            cd_day = cd[cd["DateOnly"] == dt]
            gt_day = gt_prod[gt_prod["DateOnly"] == dt] if not gt_prod.empty else gt_prod
            s1_day = s1_prod[s1_prod["DateOnly"] == dt] if not s1_prod.empty else s1_prod
            co_day = (df_all[(df_all["SKUCode"] == "CHANGEOVER") &
                             (pd.to_datetime(df_all["Date"]).dt.date == dt)].shape[0]
                      if df_all is not None and not df_all.empty else 0)

            cure_skus = cd_day["SKUCode"].nunique()
            cure_qty  = int(cd_day["Qty"].sum())
            gt_skus   = gt_day["SKUCode"].nunique() if not gt_day.empty else 0
            gt_qty    = int(gt_day["Qty"].sum())    if not gt_day.empty else 0
            s1_qty    = int(s1_day["Qty"].sum())    if not s1_day.empty else 0

            tot_cure_q += cure_qty; tot_gt_q += gt_qty
            tot_s1_q   += s1_qty;   tot_cos  += co_day
            print(f"  {str(dt):<12} {cure_skus:>10} {cure_qty:>10,} "
                  f"{gt_skus:>8} {gt_qty:>10,} {s1_qty:>10,} {co_day:>6}")

        print("  " + "-" * (len(header) - 2))

        total_cure_skus = cd["SKUCode"].nunique()
        total_gt_skus   = gt_prod["SKUCode"].nunique() if not gt_prod.empty else 0
        total_s1_skus   = s1_prod["SKUCode"].nunique() if not s1_prod.empty else 0
        print(f"  {'TOTAL':<12} {total_cure_skus:>10} {tot_cure_q:>10,} "
              f"{total_gt_skus:>8} {tot_gt_q:>10,} {tot_s1_q:>10,} {tot_cos:>6}")

        # Fulfilment + utilisation + starvation
        td  = int(df_summary["Net_GT_Demand"].sum()) if not df_summary.empty else 0
        tp  = int(df_summary["Planned_GT"].sum())    if not df_summary.empty else 0
        # capped per-SKU coverage (see build_summary KPI note) — never > 100%
        met = (int(np.minimum(df_summary["Planned_GT"], df_summary["Net_GT_Demand"]).sum())
               if not df_summary.empty else 0)
        pct = round(met / td * 100, 1) if td else 0.0
        over = max(tp - met, 0)
        full = (df_summary["Status"] == "FULLY MET").sum() if not df_summary.empty else 0
        part = (df_summary["Status"] == "PARTIAL").sum()   if not df_summary.empty else 0
        unmet= (df_summary["Status"] == "UNMET").sum()     if not df_summary.empty else 0
        stv  = (df_stv["Status"] == "STARVATION").sum()    if not df_stv.empty else 0
        wrn  = (df_stv["Status"] == "WARNING").sum()       if not df_stv.empty else 0
        u_gt = df_util_gt["Utilization_Pct"].mean() if not df_util_gt.empty else 0.0
        u_s1 = df_util_s1["Utilization_Pct"].mean() if not df_util_s1.empty else 0.0

        print(f"\n  Net GT demand        : {td:>10,}")
        print(f"  Demand met (capped)  : {met:>10,}  ({pct}%)")
        print(f"  Total built          : {tp:>10,}  (+{over:,} over-build: topup/rounding)")
        print(f"  SKUs Fully met       : {full:>10}")
        print(f"  SKUs Partial         : {part:>10}")
        print(f"  SKUs Unmet           : {unmet:>10}")
        print(f"  Starvation shifts    : {stv:>10}")
        print(f"  Warning shifts       : {wrn:>10}")
        print(f"  GT avg utilisation   : {u_gt:>9.1f}%")
        print(f"  S1 avg utilisation   : {u_s1:>9.1f}%")
        print(f"  Total changeovers    : {tot_cos:>10}")
        print(bar + "\n")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════
# Top-level entry: reads every input from DB, runs the scheduler for
# Config.PLANNING_DAYS days, writes the output workbook.
def run_from_database_hybrid(plan_start=None, output_path=None):
    if create_engine is None:
        raise ImportError("sqlalchemy not installed.")
    plan_start = plan_start or Config.PLAN_DATE
    if output_path is None:
        output_path = (
            f"BTP_PCR_Building_Hybrid_v8_{plan_start.date()}_"
            f"{Config.PLANNING_DAYS}Days.xlsx"
        )

    # Use the hardened engine (pool_pre_ping / pool_recycle) so a dropped/stale
    # connection reconnects instead of raising "invalid transaction" errors.
    try:
        from cbc_env import make_engine as _mk
        engine = _mk()
    except Exception:  # noqa: BLE001 — fallback to a plain engine
        engine = create_engine(
            f"mysql+pymysql://{Config.DB_USER}:{Config.DB_PASSWORD}"
            f"@{Config.DB_SERVER}/{Config.DB_NAME}",
            pool_pre_ping=True, pool_recycle=280,
        )
    etl = ETL(engine)

    print("\n[Phase 0] ETL from database...")
    df_curing      = etl.load_curing_schedule()
    df_gt_inv      = etl.load_gt_inventory()
    df_carcass_inv = etl.load_carcass_inventory()
    df_allow       = etl.load_machine_allowable()
    co_map         = etl.load_changeover_map()
    sku_to_size    = etl.load_sku_sizes()
    df_running     = etl.load_running_machines()
    history_map    = etl.load_history_map()

    # ── Whole-month build with a one-day lead ────────────────────────────
    # Building day D produces the GTs that curing consumes on day D+1, so we
    # shift the curing plan back LEAD_DAYS AND start building's window LEAD_DAYS
    # earlier. Both must move together: shifting only the curing (not the window)
    # pushed the first curing day before plan_start, silently dropping ~1 day of
    # demand (e.g. 25,810 GTs for June) and collapsing the last building day.
    # With the window aligned, building plans [plan_start-LEAD_DAYS … +DAYS) and
    # covers ALL curing days.
    # Cold start: opening GT/carcass inventory = zero (the day-before-curing
    # build stocks day 1) and no machine is pre-locked mid-run; inventory
    # then rolls forward internally.
    LEAD_DAYS = 1
    df_curing = df_curing.copy()
    df_curing["StartTime"] = df_curing["StartTime"] - pd.Timedelta(days=LEAD_DAYS)
    df_curing["EndTime"]   = df_curing["EndTime"]   - pd.Timedelta(days=LEAD_DAYS)
    build_start    = plan_start - timedelta(days=LEAD_DAYS)
    df_gt_inv      = pd.DataFrame(columns=["SKUCode", "GT_Inventory"])
    df_carcass_inv = pd.DataFrame(columns=["SKUCode", "Carcass_Inventory"])
    df_running     = pd.DataFrame(columns=["Machine", "SKUCode"])

    print(f"\n[v8 Hybrid] Rolling {Config.PLANNING_DAYS}-day horizon as "
          f"{Config.PLANNING_DAYS} × (GA + LP) per day "
          f"(build window starts {build_start:%Y-%m-%d}, one day ahead of curing)")

    scheduler = HybridDailyScheduler()
    results = scheduler.run(
        df_curing, df_gt_inv, df_carcass_inv,
        df_allow, co_map, sku_to_size, df_running,
        build_start,
        history_map=history_map,
    )
    ExcelExporter(output_path).export(results)


if __name__ == "__main__":
    run_from_database_hybrid()

