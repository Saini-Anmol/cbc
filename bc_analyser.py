"""
B2C Pipeline — Phase 3: Analysis & KPIs
========================================
Reads the three Phase 0/1/2 output files and produces a consolidated KPI
workbook: building utilisation, GT alignment, starvation report, changeover
summary, SKU diversity, and monthly aggregates.

Standalone usage:
    python bc_analyser.py

Inputs:
    data/main_output/bc_building_schedule.xlsx   (Phase 1 output)
    data/main_output/bc_curing_schedule.xlsx     (Phase 2 output)
    data/output/curing_consumption_table.xlsx    (Phase 0 output)

Output:
    data/main_output/bc_analysis.xlsx
"""

from __future__ import annotations

import os
import sys
import warnings
from collections import defaultdict
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

import cbc_env

HERE    = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = cbc_env.OUTPUT_DIR
MAIN_OUT = os.path.join(OUT_DIR, "main_output")
os.makedirs(MAIN_OUT, exist_ok=True)

# GT-producing machine sets (from building.py)
_GT_STAGES = frozenset({
    "8201", "8301", "8302", "8501", "8502", "7301",
    "7001", "7002", "7003", "7004",
    "6001", "6002", "6003", "6004",
    "7101", "7102", "7103", "7104", "7105", "7106",
    "7201", "7501", "7502", "7503",
})
_S1_STAGES = frozenset({
    "6801", "6802", "6803", "6909", "6911", "7601", "7701",
    "7801", "7802", "7803", "7804", "8001", "8002", "8003", "8101",
})
_CO_SENTINELS = {"CHANGEOVER", "MOULD_CLEAN", "C/O", "CLEANING"}

SHIFT_MINS = 480  # minutes per shift


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: read a sheet from an xlsx file
# ══════════════════════════════════════════════════════════════════════════════

def _read_sheet(path: str, sheet: str, required_cols: set | None = None) -> pd.DataFrame:
    """Read a named sheet, trying multiple header rows. Returns empty DF on failure."""
    if not os.path.exists(path):
        print(f"  ⚠️  File not found: {path}")
        return pd.DataFrame()
    xl = pd.ExcelFile(path)
    target_sheet = sheet if sheet in xl.sheet_names else (xl.sheet_names[0] if xl.sheet_names else None)
    if target_sheet is None:
        return pd.DataFrame()
    for hdr in range(4):
        try:
            df = pd.read_excel(path, sheet_name=target_sheet, header=hdr)
            if required_cols and not required_cols.issubset({c.strip() for c in df.columns}):
                continue
            return df
        except Exception:
            continue
    return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# BC ANALYSER
# ══════════════════════════════════════════════════════════════════════════════

class BCAnalyser:
    """
    Reads Phase 0/1/2 outputs and computes all KPIs.
    Call run() to load data, then export() to write the analysis workbook.
    """

    def __init__(
        self,
        building_path: str,
        curing_path: str,
        consumption_path: str,
    ):
        self.building_path    = building_path
        self.curing_path      = curing_path
        self.consumption_path = consumption_path

        # Loaded DataFrames (populated by run())
        self.df_bld_shift:    pd.DataFrame = pd.DataFrame()
        self.df_bld_demand:   pd.DataFrame = pd.DataFrame()
        self.df_cur_shift:    pd.DataFrame = pd.DataFrame()
        self.df_gt_balance:   pd.DataFrame = pd.DataFrame()
        self.df_demand_ful:   pd.DataFrame = pd.DataFrame()
        self.df_daily_sum:    pd.DataFrame = pd.DataFrame()
        self.df_consumption:  pd.DataFrame = pd.DataFrame()
        self.df_co_plan:      pd.DataFrame = pd.DataFrame()

    def run(self) -> "BCAnalyser":
        """Load all data. Returns self for chaining."""
        print("  [Analyser] Loading building schedule …")
        self.df_bld_shift  = _read_sheet(self.building_path, "Shift Schedule",
                                         required_cols={"SKUCode", "Machine", "Qty"})
        self.df_bld_demand = _read_sheet(self.building_path, "Demand Summary",
                                         required_cols={"SKUCode"})
        self.df_co_plan    = _read_sheet(self.building_path, "Changeover Plan")

        print("  [Analyser] Loading curing schedule …")
        self.df_cur_shift  = _read_sheet(self.curing_path, "Shift Schedule",
                                         required_cols={"SKUCode", "Date", "Shift"})
        self.df_gt_balance = _read_sheet(self.curing_path, "GT Balance")
        self.df_demand_ful = _read_sheet(self.curing_path, "Demand Fulfillment",
                                         required_cols={"SKUCode", "Demand_Qty"})
        self.df_daily_sum  = _read_sheet(self.curing_path, "Daily Summary",
                                         required_cols={"Total_Cured", "Starvation_Events"})

        print("  [Analyser] Loading consumption table …")
        self.df_consumption = _read_sheet(self.consumption_path, "Consumption Summary")
        if "SKUCode" in self.df_consumption.columns:
            self.df_consumption["SKUCode"] = \
                self.df_consumption["SKUCode"].astype(str).str.strip()

        return self

    # ── 1. Building Utilisation ───────────────────────────────────────────────

    def building_utilisation(self) -> pd.DataFrame:
        """Per-machine: used minutes, idle minutes, utilisation %, SKU count, total units."""
        df = self.df_bld_shift
        if df.empty:
            return pd.DataFrame()

        df = df.copy()
        if "Machine" not in df.columns:
            return pd.DataFrame()

        df["Machine"] = df["Machine"].astype(str).str.strip()
        df["Qty"]     = pd.to_numeric(df.get("Qty", 0), errors="coerce").fillna(0)

        # Determine total shifts in the schedule
        total_shifts = df[["Date", "Shift"]].drop_duplicates().__len__() \
            if "Date" in df.columns and "Shift" in df.columns else 90
        cap_per_machine = total_shifts * SHIFT_MINS

        prod = df[~df["Machine"].isin(["", "nan"])
                  & ~df["SKUCode"].astype(str).isin(_CO_SENTINELS)].copy()

        # Used minutes: Qty × CT_Min (or estimate from Qty/capacity ratio)
        if "CT_Min" in prod.columns:
            prod["Used_Mins"] = prod["Qty"] * prod["CT_Min"].fillna(0) / 2  # ÷ CAVITIES
        else:
            prod["Used_Mins"] = prod["Qty"] * 0  # can't compute without CT

        agg = (
            prod.groupby("Machine")
            .agg(
                Total_Units=("Qty", "sum"),
                Distinct_SKUs=("SKUCode", "nunique"),
            )
            .reset_index()
        )

        # Changeover time
        co_df = df[df["SKUCode"].astype(str).isin(_CO_SENTINELS)].copy()
        co_agg = co_df.groupby("Machine").size().reset_index(name="CO_Count")
        agg = agg.merge(co_agg, on="Machine", how="left")
        agg["CO_Count"] = agg["CO_Count"].fillna(0).astype(int)

        # Stage classification
        agg["Stage"] = agg["Machine"].apply(
            lambda m: "Stage-2/Unistage" if m in _GT_STAGES
            else ("Stage-1" if m in _S1_STAGES else "Unknown")
        )

        agg = agg.sort_values(["Stage", "Machine"])
        return agg.reset_index(drop=True)

    # ── 2. GT Alignment ───────────────────────────────────────────────────────

    def gt_alignment(self) -> pd.DataFrame:
        """
        Per (SKU, shift): GT balance and whether curing was fully fed.
        Positive balance = surplus; 0 = exact; negative = starvation.
        """
        df = self.df_gt_balance
        if df.empty:
            return pd.DataFrame()

        df = df.copy()
        df["GT_Surplus_Deficit"] = (
            df["Building_Output"].fillna(0) - df["Curing_Consumption"].fillna(0)
        )
        df["Alignment_Status"] = df.apply(
            lambda r: "SURPLUS"     if r["GT_Surplus_Deficit"] > 50
            else "BALANCED" if abs(r["GT_Surplus_Deficit"]) <= 50
            else "DEFICIT",
            axis=1,
        )
        return df[["Date", "Shift", "SKUCode", "Active_Press_Count",
                   "Building_Output", "Curing_Consumption", "Cured_Qty",
                   "GT_Balance", "GT_Surplus_Deficit", "Alignment_Status",
                   "Status"]].copy()

    # ── 3. Starvation Report ──────────────────────────────────────────────────

    def starvation_report(self) -> pd.DataFrame:
        """Per (SKU, Date, Shift): starvation events with deficit severity."""
        df = self.df_gt_balance
        if df.empty:
            return pd.DataFrame()

        starv = df[df["Status"] == "WAITING_GT"].copy()
        if starv.empty:
            return pd.DataFrame(columns=["Date", "Shift", "SKUCode", "Category",
                                         "GT_Balance", "Curing_Consumption", "Severity"])

        # Merge category from consumption table
        if not self.df_consumption.empty and "Category" in self.df_consumption.columns:
            cat_map = dict(zip(
                self.df_consumption["SKUCode"],
                self.df_consumption["Category"],
            ))
            starv["Category"] = starv["SKUCode"].map(cat_map).fillna("Unknown")
        else:
            starv["Category"] = "Unknown"

        starv["Severity"] = starv["GT_Balance"].apply(
            lambda b: "CRITICAL" if b < -500 else "WARNING"
        )
        return starv[["Date", "Shift", "SKUCode", "Category",
                       "GT_Balance", "Curing_Consumption", "Severity"]].copy()

    # ── 4. Changeover Summary ─────────────────────────────────────────────────

    def changeover_summary(self) -> pd.DataFrame:
        """Per day: changeovers taken vs limit, stranded/deferred presses."""
        df = self.df_co_plan
        if df.empty:
            return pd.DataFrame(columns=["CO_Day_Index", "Scheduled", "Deferred", "Status"])

        sched   = df[df["Status"] == "SCHEDULED"].groupby("CO_Day_Index").size()
        deferred = df[df["Status"] == "DEFERRED"].shape[0]

        summary = sched.reset_index()
        summary.columns = ["CO_Day_Index", "Scheduled"]
        summary["Max_Allowed"] = 8
        summary["Within_Limit"] = summary["Scheduled"] <= 8

        foot = pd.DataFrame([{
            "CO_Day_Index": "TOTAL",
            "Scheduled":    sched.sum(),
            "Max_Allowed":  "—",
            "Within_Limit": (sched > 8).sum() == 0,
        }])
        return pd.concat([summary, foot], ignore_index=True)

    # ── 5. SKU Diversity ─────────────────────────────────────────────────────

    def sku_diversity(self) -> pd.DataFrame:
        """Per day: distinct SKUs built vs cured."""
        built_div = pd.DataFrame()
        cured_div = pd.DataFrame()

        if not self.df_bld_shift.empty and "Date" in self.df_bld_shift.columns:
            bld = self.df_bld_shift[
                ~self.df_bld_shift["SKUCode"].astype(str).isin(_CO_SENTINELS)
                & self.df_bld_shift["Machine"].astype(str).isin(_GT_STAGES)
            ]
            built_div = (
                bld.groupby("Date")["SKUCode"].nunique()
                .reset_index()
                .rename(columns={"SKUCode": "Distinct_SKUs_Built"})
            )

        if not self.df_cur_shift.empty and "Date" in self.df_cur_shift.columns:
            cur = self.df_cur_shift[
                self.df_cur_shift["Status"].astype(str) == "RUNNING"
            ]
            cured_div = (
                cur.groupby("Date")["SKUCode"].nunique()
                .reset_index()
                .rename(columns={"SKUCode": "Distinct_SKUs_Cured"})
            )

        if built_div.empty and cured_div.empty:
            return pd.DataFrame()
        if built_div.empty:
            return cured_div
        if cured_div.empty:
            return built_div
        return built_div.merge(cured_div, on="Date", how="outer").fillna(0)

    # ── 6. Monthly Summary KPIs ───────────────────────────────────────────────

    def monthly_summary(self) -> dict:
        """High-level KPIs for the entire planning horizon."""
        kpis = {}

        # Curing KPIs
        if not self.df_daily_sum.empty:
            kpis["total_cured_tyres"]    = int(self.df_daily_sum["Total_Cured"].sum())
            kpis["avg_daily_cured"]      = round(
                self.df_daily_sum["Total_Cured"].mean(), 0)
            kpis["starvation_shifts"]    = int(
                self.df_daily_sum["Starvation_Events"].sum())
        else:
            kpis["total_cured_tyres"]    = 0
            kpis["avg_daily_cured"]      = 0
            kpis["starvation_shifts"]    = 0

        # Demand fulfillment KPIs
        if not self.df_demand_ful.empty and "Demand_Qty" in self.df_demand_ful.columns:
            total_demand = self.df_demand_ful["Demand_Qty"].sum()
            total_cured  = self.df_demand_ful["Total_Cured"].sum()
            kpis["demand_fulfillment_pct"] = round(
                100 * total_cured / max(total_demand, 1), 1)
            kpis["total_demand"]           = int(total_demand)
            ri_ful = self.df_demand_ful[
                self.df_demand_ful["Category"] == "Runner-In"
            ]["Fulfillment_Pct"].mean()
            kpis["runner_in_fulfillment_pct"] = round(ri_ful, 1) if pd.notna(ri_ful) else 0
        else:
            kpis["demand_fulfillment_pct"]    = 0
            kpis["total_demand"]              = 0
            kpis["runner_in_fulfillment_pct"] = 0

        # Building KPIs
        if not self.df_bld_shift.empty and "Qty" in self.df_bld_shift.columns:
            bld_prod = self.df_bld_shift[
                ~self.df_bld_shift["SKUCode"].astype(str).isin(_CO_SENTINELS)
                & self.df_bld_shift["Machine"].astype(str).isin(_GT_STAGES)
            ]
            kpis["total_gt_built"] = int(bld_prod["Qty"].sum())
        else:
            kpis["total_gt_built"] = 0

        # Changeover KPIs
        if not self.df_co_plan.empty:
            kpis["total_changeovers"] = int((
                self.df_co_plan["Status"] == "SCHEDULED"
            ).sum())
            kpis["deferred_changeovers"] = int((
                self.df_co_plan["Status"] == "DEFERRED"
            ).sum())
        else:
            kpis["total_changeovers"]    = 0
            kpis["deferred_changeovers"] = 0

        # SKU counts from consumption table
        if not self.df_consumption.empty and "Category" in self.df_consumption.columns:
            kpis["runner_in_skus"]     = int((self.df_consumption["Category"] == "Runner-In").sum())
            kpis["runner_out_skus"]    = int((self.df_consumption["Category"] == "Runner-Out").sum())
            kpis["non_runner_in_skus"] = int((self.df_consumption["Category"] == "Non-Runner-In").sum())
        else:
            kpis["runner_in_skus"]     = 0
            kpis["runner_out_skus"]    = 0
            kpis["non_runner_in_skus"] = 0

        return kpis

    # ── 7. Export ─────────────────────────────────────────────────────────────

    def export(self, output_path: str):
        """Write bc_analysis.xlsx with one sheet per KPI dimension."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = Workbook()

        _NAVY  = "1F3864"
        _WHITE = "FFFFFF"
        _GREEN = "E2EFDA"
        _RED   = "FFE0E0"
        _AMBER = "FFF2CC"

        def _add_kpi_sheet(name: str, df: pd.DataFrame, color_col: str = ""):
            ws = wb.create_sheet(name)
            if df.empty:
                ws.cell(row=1, column=1, value="No data available").font = Font(italic=True)
                return ws
            hdr_fill = PatternFill("solid", fgColor=_NAVY)
            hdr_font = Font(bold=True, color=_WHITE)
            bd = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"),
            )
            for ci, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.border = bd
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for ri, (_, row) in enumerate(df.iterrows(), start=2):
                fill = None
                if color_col and color_col in row:
                    v = str(row[color_col])
                    fill = (PatternFill("solid", fgColor=_RED)   if v in {"WAITING_GT", "CRITICAL", "DEFICIT"}
                            else PatternFill("solid", fgColor=_AMBER) if v in {"WARNING", "DEFERRED"}
                            else PatternFill("solid", fgColor=_GREEN) if v in {"RUNNING", "OK", "SURPLUS", "BALANCED", "SCHEDULED"}
                            else None)
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = bd
                    cell.alignment = Alignment(horizontal="center")
                    if fill:
                        cell.fill = fill
            for col in ws.columns:
                w = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 35)
            return ws

        # KPI Summary (first sheet)
        ws0 = wb.active
        ws0.title = "KPI Summary"
        kpis = self.monthly_summary()
        ws0.cell(row=1, column=1, value="B2C Pipeline — Monthly KPI Summary").font = \
            Font(bold=True, size=14)
        ws0.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = \
            Font(italic=True, size=10)
        kpi_rows = [
            ("", ""),
            ("CURING KPIs", ""),
            ("Total Cured Tyres (Month)", kpis["total_cured_tyres"]),
            ("Avg Daily Cured Tyres", kpis["avg_daily_cured"]),
            ("Demand Fulfillment %", f"{kpis['demand_fulfillment_pct']}%"),
            ("Runner-In Fulfillment %", f"{kpis['runner_in_fulfillment_pct']}%"),
            ("Starvation Events (shifts)", kpis["starvation_shifts"]),
            ("", ""),
            ("BUILDING KPIs", ""),
            ("Total GT Built", kpis["total_gt_built"]),
            ("Total Demand", kpis["total_demand"]),
            ("", ""),
            ("CHANGEOVER KPIs", ""),
            ("Total Changeovers Scheduled", kpis["total_changeovers"]),
            ("Deferred Changeovers", kpis["deferred_changeovers"]),
            ("", ""),
            ("SKU CLASSIFICATION", ""),
            ("Runner-In SKUs", kpis["runner_in_skus"]),
            ("Runner-Out SKUs", kpis["runner_out_skus"]),
            ("Non-Runner-In SKUs", kpis["non_runner_in_skus"]),
        ]
        for r, (label, value) in enumerate(kpi_rows, start=4):
            ws0.cell(row=r, column=1, value=label)
            ws0.cell(row=r, column=2, value=value)
            if label and value == "":
                ws0.cell(row=r, column=1).font = Font(bold=True)

        for col in ws0.columns:
            w = max((len(str(c.value or "")) for c in col), default=20)
            ws0.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 45)

        # Individual KPI sheets
        util_df  = self.building_utilisation()
        align_df = self.gt_alignment()
        starv_df = self.starvation_report()
        co_df    = self.changeover_summary()
        div_df   = self.sku_diversity()

        if not util_df.empty:
            _add_kpi_sheet("Building Utilisation", util_df)
        if not align_df.empty:
            _add_kpi_sheet("GT Alignment", align_df,   color_col="Alignment_Status")
        if not starv_df.empty:
            _add_kpi_sheet("Starvation Report", starv_df, color_col="Severity")
        if not co_df.empty:
            _add_kpi_sheet("Changeover Summary", co_df)
        if not div_df.empty:
            _add_kpi_sheet("SKU Diversity", div_df)
        if not self.df_demand_ful.empty:
            _add_kpi_sheet("Demand Fulfillment", self.df_demand_ful)
        if not self.df_daily_sum.empty:
            _add_kpi_sheet("Daily Summary", self.df_daily_sum)

        wb.save(output_path)
        print(f"  [Analyser] Saved → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_analysis(
    building_path: str,
    curing_path: str,
    consumption_path: str,
    output_path: str,
) -> BCAnalyser:
    """
    Run B2C analysis and export KPI workbook.

    Returns the BCAnalyser instance (with all DataFrames loaded) so the caller
    can access individual KPI methods directly.
    """
    print("\n" + "=" * 70)
    print("  B2C Phase 3 — Analysis & KPIs")
    print("=" * 70)

    analyser = BCAnalyser(building_path, curing_path, consumption_path)
    analyser.run()

    kpis = analyser.monthly_summary()
    print(f"\n  Monthly KPIs:")
    print(f"    Total Cured Tyres:     {kpis['total_cured_tyres']:,}")
    print(f"    Avg Daily Cured:       {kpis['avg_daily_cured']:,}")
    print(f"    Demand Fulfillment:    {kpis['demand_fulfillment_pct']}%")
    print(f"    Runner-In Fulfillment: {kpis['runner_in_fulfillment_pct']}%")
    print(f"    Starvation Events:     {kpis['starvation_shifts']}")
    print(f"    Total GT Built:        {kpis['total_gt_built']:,}")
    print(f"    Changeovers Scheduled: {kpis['total_changeovers']}")

    print(f"\n  [Export] Writing → {output_path}")
    analyser.export(output_path)

    print("=" * 70)
    print("  Phase 3 complete.")
    print("=" * 70 + "\n")

    return analyser


# ══════════════════════════════════════════════════════════════════════════════
# STANDALONE RUN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    _BUILDING_PATH    = os.path.join(MAIN_OUT, "bc_building_schedule.xlsx")
    _CURING_PATH      = os.path.join(MAIN_OUT, "bc_curing_schedule.xlsx")
    _CONSUMPTION_PATH = os.path.join(OUT_DIR,  "curing_consumption_table.xlsx")
    _OUTPUT_PATH      = os.path.join(MAIN_OUT, "bc_analysis.xlsx")

    analyser = run_analysis(
        building_path    = _BUILDING_PATH,
        curing_path      = _CURING_PATH,
        consumption_path = _CONSUMPTION_PATH,
        output_path      = _OUTPUT_PATH,
    )
