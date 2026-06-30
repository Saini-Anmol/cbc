"""
building_unconstrained.py — Max Building Capacity Experiment
============================================================
Generates a building schedule directly from customer demand,
bypassing all curing constraints (no press counts, no CO schedule,
no curing consumption table).

Purpose: measure the theoretical maximum building output when curing
constraints are absent — a pure building-capacity ceiling to compare
against the constrained B2C run.

What's KEPT:
  • Machine eligibility (allowable machines from DB + history union)
  • Inch locking (_UNISTAGE_INCH_POLICY + _MACHINE_HARD_INCH)
  • Demand cap: total build per SKU ≤ Demand_Qty  (hard invariant)
  • Stage-1 → Stage-2 dependency (structural — not a curing rule)
  • LP + Heuristic (HybridDailyScheduler from building.py)
  • April 30 Shift B start (PRE_START_SHIFTS=2 before May 1)
  • 31-day planning horizon

What's REMOVED:
  • curing_consumption_dynamic.py step
  • Consumption table, press counts, curing CO schedule
  • RI / RO / NRI SKU classification
  • DynamicTargetLock, _make_synthetic_curing (press-rate ceilings)
  • Mould-constrained priority boost (requires press count data)
  • 8 COs/day curing cap (irrelevant — no curing CO)

Usage:
    python building_unconstrained.py
    python building_unconstrained.py data/input/demand_may.xlsx

Output: data/output/main_output/bc_building_unconstrained_<date>.xlsx
"""

from __future__ import annotations

import os
import sys
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings("ignore")

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

# ── Import reusable machinery from building.py ────────────────────────────────
from building import (
    Config,
    ETL,
)
from openpyxl.utils import get_column_letter

import cbc_env

HERE     = os.path.dirname(os.path.abspath(__file__))
OUT_DIR  = cbc_env.OUTPUT_DIR
MAIN_OUT = os.path.join(OUT_DIR, "main_output")
os.makedirs(MAIN_OUT, exist_ok=True)

# ── Config overrides ──────────────────────────────────────────────────────────
# ── Unconstrained-mode tuning (different from B2C defaults) ──────────────────
# MIN_CAMPAIGN_MINS: lowered from 120→45 to eliminate idle tails at shift end.
# In B2C mode 120 was needed to stop CO explosion from many NRI press activations
# firing simultaneously. Here there are no curing COs, so shorter campaigns are safe.
# This alone recovers ~15-25k units of idle-tail loss.
Config.MIN_CAMPAIGN_MINS    = 45

# OVERBUILD_BUFFER_FRAC: raised from 0.2→0.4 to give LP more per-shift headroom.
# In B2C the 0.2 cap prevented overbuilding vs curing schedule. Here the only
# ceiling is total demand, so giving LP 40% headroom per shift lets it push harder
# on fast-throughput machines without violating the total demand cap.
Config.OVERBUILD_BUFFER_FRAC = 0.4

# BUILD_LEAD_SHIFTS: set to 0 — no curing schedule to sync with.
# In B2C, =3 was essential to prevent LP cap collapse when prior-day WIP blocked
# same-day building. Here demand is spread evenly across all shifts; setting to 0
# means LP targets today's demand directly, which fully utilises Days 29-31
# (those days had no Day+1 target under BUILD_LEAD_SHIFTS=3 → machines idled).
Config.BUILD_LEAD_SHIFTS    = 0

Config.CURING_PLAN_FILE     = None  # not used

# Inch policies — identical to building_b2c.py
# Source: CLAUDE.md §Inch-Run Study
_UNISTAGE_INCH_POLICY: dict[str, set] = {
    "6001": {"14","15","16","17","18"}, "6002": {"14","15","16","17","18"},
    "6003": {"14","15","16","17","18"}, "6004": {"14","15","16","17","18"},
    "7001": {"14","15","16","17","18"}, "7002": {"14","15","16","17","18"},
    "7003": {"14","15","16","17","18"}, "7004": {"14","15","16","17","18"},
    "7101": {"13","14","15","16"},      "7102": {"13","14","15","16"},
    "7103": {"13","14","15","16"},      "7104": {"13","14","15","16"},
    "7105": {"13","14","15","16"},      "7106": {"13","14","15","16"},
    "7201": {"13","14","15","16"},
    "7501": {"12","13"}, "7502": {"12","13"}, "7503": {"12","13"},
}

_MACHINE_HARD_INCH: dict[str, set] = {
    "7001": {"16"},         "6001": {"14"},         "7002": {"14"},
    "7004": {"14"},         "6002": {"15"},         "7003": {"15"},
    "6003": {"17","18"},    "6004": {"16"},
    "7101": {"15"},         "7102": {"14","15"},    "7103": {"13"},
    "7104": {"14","15"},    "7105": {"13"},         "7106": {"13"},
    "7201": {"16"},
    "7501": {"12"},         "7502": {"13"},         "7503": {"13"},
}


# ══════════════════════════════════════════════════════════════════════════════
# DEMAND LOADER
# ══════════════════════════════════════════════════════════════════════════════

def load_demand(path: str) -> pd.DataFrame:
    """
    Load demand file. Returns [SKUCode, Demand_Qty, Priority_Score].
    Mirrors curing_consumption.ConsumptionETL.load_demand() — handles all
    raw/normalised CSV + XLSX formats used in this project.
    """
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)

    col_map = {
        "skuCode":  "SKUCode", "sku_code": "SKUCode",
        "sapcode":  "SKUCode", "Sapcode":  "SKUCode",
        "requirement":         "Requirement",
        "updated_requirement": "Updated_Requirement",
    }
    df = df.rename(columns={c: col_map[c] for c in df.columns if c in col_map})

    if "SKUCode" not in df.columns:
        raise KeyError(f"Demand file has no SKU column. Found: {df.columns.tolist()}")

    df["SKUCode"] = df["SKUCode"].astype(str).str.strip()

    for candidate in ("Quantity", "Updated_Requirement", "Requirement"):
        if candidate in df.columns:
            qty_col = candidate
            break
    else:
        raise KeyError(f"Demand file has no quantity column. Found: {df.columns.tolist()}")

    if "Priority" in df.columns:
        pri_col = "Priority"
    elif "ConsolidatedPriorityScore" in df.columns:
        pri_col = "ConsolidatedPriorityScore"
    else:
        df["_priority"] = 1.0
        pri_col = "_priority"

    df = (df.groupby("SKUCode")
            .agg(Demand_Qty=(qty_col, "sum"), Priority_Score=(pri_col, "max"))
            .reset_index())
    return df[df["Demand_Qty"] > 0].copy()


# ══════════════════════════════════════════════════════════════════════════════
# DEMAND → df_curing CONVERTER
# ══════════════════════════════════════════════════════════════════════════════

def demand_to_curing_df(
    demand: dict,           # {SKUCode: Demand_Qty}
    plan_start: datetime,
    planning_days: int,
) -> pd.DataFrame:
    """
    Convert SKU→Demand_Qty into the df_curing format that HybridDailyScheduler
    expects: [SKUCode, StartTime, EndTime, Qty].

    Each SKU's total demand is spread EVENLY across all (planning_days × 3)
    shifts. This removes the press-rate ceiling that the consumption table
    imposes — the LP sees the full demand as a reachable target every shift.
    """
    total_shifts = planning_days * Config.SHIFTS_PER_DAY
    sh = Config.SHIFT_START_HOUR
    hrs = [sh, sh + Config.HOURS_PER_SHIFT, sh + Config.HOURS_PER_SHIFT * 2]
    rows = []
    for sku, qty in demand.items():
        if qty <= 0:
            continue
        per_shift = qty / total_shifts
        for day_offset in range(planning_days):
            base = plan_start + timedelta(days=day_offset)
            for h in hrs:
                sh_start = datetime(base.year, base.month, base.day, h % 24)
                if h >= 24:
                    sh_start += timedelta(days=1)
                rows.append({
                    "SKUCode":   sku,
                    "StartTime": sh_start,
                    "EndTime":   sh_start + timedelta(hours=Config.HOURS_PER_SHIFT),
                    "Qty":       per_shift,
                })
    if not rows:
        return pd.DataFrame(columns=["SKUCode", "StartTime", "EndTime", "Qty"])
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# GREEDY PARALLEL CAMPAIGN SCHEDULER
# ══════════════════════════════════════════════════════════════════════════════

class GreedyCampaignScheduler:
    """
    Parallel campaign-to-completion building scheduler.

    For each SKU (sorted by priority then demand), ALL eligible machines are
    dispatched simultaneously.  Each machine contributes to the shared remaining
    demand until demand is fully met, then COs to its next highest-priority
    eligible SKU.  With per-machine hard-inch locking every CO is a same-size
    CO (20 min on VMI), so CO overhead collapses from ~10-20% to <1% and
    machines spend nearly all their time in production.

    Machine dispatch order: VMI first (spare capacity + cheapest CO), then
    UNI_NARROW, then BJ, then Stage-2.
    """

    _ORDER = [
        "6001","6002","6003","6004",          # VMIMAXX
        "7001","7002","7003","7004",          # VMIMAXX
        "7501","7502","7503",                  # UNI_NARROW
        "7101","7102","7103","7104",           # BJ
        "7105","7106","7201",                  # BJ
        "7301","8201","8301","8302","8501","8502",  # STAGE2
    ]

    def run(
        self,
        allow_map_dict: dict,
        demand_dict:    dict,
        priority_map:   dict,
        sku_to_size:    dict,
        co_map_raw:     dict,
        total_mins:     float,
    ) -> tuple:
        """
        Returns
        -------
        campaigns        list[dict]  — one entry per CO or production campaign
        demand_remaining {sku: float} — unbuilt qty per SKU after scheduling
        machine_stats    {machine: dict} — utilization breakdown per machine
        """
        demand_remaining  = {s: float(q) for s, q in demand_dict.items() if q > 0}
        machine_avail_at  = {m: 0.0 for m in self._ORDER}
        machine_last_inch = {m: None for m in self._ORDER}
        machine_prod_time = {m: 0.0 for m in self._ORDER}
        machine_co_mins   = {m: 0.0 for m in self._ORDER}
        campaigns: list   = []

        # Stable dispatch index: VMI first, then UNI_NARROW, then BJ, then STAGE2.
        # Used as tiebreaker when multiple machines have the same avail_at (e.g. t=0).
        # Without this, set iteration order is random → non-deterministic results.
        _order_idx = {m: i for i, m in enumerate(self._ORDER)}

        # Highest priority first; largest demand as tiebreaker (critical → scheduled first)
        skus_sorted = sorted(
            [s for s in demand_dict if demand_dict[s] > Config.MIN_CAMPAIGN_UNITS],
            key=lambda s: (-priority_map.get(s, 1.0), -demand_dict.get(s, 0)),
        )

        for sku in skus_sorted:
            sku_inch = str(sku_to_size.get(sku, "")).strip().replace('"', "")

            # Dispatch eligible machines: earliest-available first.
            # Tiebreak by _ORDER position (VMI before BJ before Stage-2) so that
            # when machines are all free (t=0), VMI machines are dispatched first.
            eligible = sorted(
                [m for m in allow_map_dict.get(sku, set()) if m in machine_avail_at],
                key=lambda m: (machine_avail_at[m], _order_idx.get(m, 999)),
            )

            for machine in eligible:
                if demand_remaining.get(sku, 0) <= Config.MIN_CAMPAIGN_UNITS:
                    break  # demand fully served by earlier machines in this loop

                t_free    = machine_avail_at[machine]
                prev_inch = machine_last_inch[machine]

                # CO cost — with hard-inch locking, almost always "same" (same mould)
                if prev_inch is not None:
                    co_key  = "same" if prev_inch == sku_inch else "diff"
                    co_time = float(co_map_raw.get(machine, {}).get(co_key, 45))
                else:
                    co_time = 0.0  # first campaign: no CO

                t_prod     = t_free + co_time
                time_avail = total_mins - t_prod

                if time_avail < Config.MIN_CAMPAIGN_MINS:
                    continue  # not enough room for a meaningful run

                ct        = Config.ct_min(machine)          # min per unit
                qty_need  = demand_remaining[sku]
                camp_time = min(qty_need * ct, time_avail)  # min needed vs time left
                qty_built = camp_time / ct

                if qty_built < Config.MIN_CAMPAIGN_UNITS:
                    continue

                if co_time > 0:
                    campaigns.append({
                        "Machine": machine, "SKUCode": "CHANGEOVER",
                        "Start_Min": t_free,  "End_Min": t_prod,
                        "Duration_Min": co_time, "Qty": 0.0,
                        "From_Inch": prev_inch, "To_Inch": sku_inch,
                    })
                    machine_co_mins[machine] += co_time

                campaigns.append({
                    "Machine": machine, "SKUCode": sku,
                    "Start_Min": t_prod, "End_Min": t_prod + camp_time,
                    "Duration_Min": camp_time, "Qty": qty_built,
                    "Inch": sku_inch,
                })
                machine_prod_time[machine] += camp_time

                demand_remaining[sku]     = max(0.0, demand_remaining[sku] - qty_built)
                machine_avail_at[machine]  = t_prod + camp_time
                machine_last_inch[machine] = sku_inch

        machine_stats = {
            m: {
                "prod_time":       machine_prod_time[m],
                "co_time":         machine_co_mins[m],
                "idle_time":       max(0.0, total_mins - machine_prod_time[m] - machine_co_mins[m]),
                "utilization":     (machine_prod_time[m] + machine_co_mins[m]) / total_mins,
                "prod_utilization": machine_prod_time[m] / total_mins,
            }
            for m in self._ORDER
        }
        return campaigns, demand_remaining, machine_stats


# ══════════════════════════════════════════════════════════════════════════════
# GREEDY OUTPUT WRITER
# ══════════════════════════════════════════════════════════════════════════════

def _write_greedy_xlsx(
    output_path:   str,
    campaigns:     list,
    demand_dict:   dict,
    df_demand:     pd.DataFrame,
    gt_inv_dict:   dict,
    allow_map:     dict,
    machine_stats: dict,
    total_mins:    float,
) -> dict:
    """
    Write greedy campaign output to Excel.

    Sheets
    ------
    KPI Summary         — headline numbers
    Demand Fulfillment  — per-SKU coloured table (same format as LP run)
    Machine Utilization — per-machine prod/CO/idle breakdown
    Campaign Schedule   — full chronological campaign list per machine
    """
    from openpyxl import Workbook as _WB

    # ── Derive prod_by_sku ────────────────────────────────────────────────────
    STAGE1_SET = {str(m) for m in Config.STAGE1}
    prod_by_sku: dict = {}
    co_count = 0
    for c in campaigns:
        if c["SKUCode"] == "CHANGEOVER":
            co_count += 1
        else:
            if c["Machine"] not in STAGE1_SET:   # only count GT machines
                prod_by_sku[c["SKUCode"]] = (
                    prod_by_sku.get(c["SKUCode"], 0.0) + c["Qty"]
                )

    total_demand = float(df_demand["Demand_Qty"].sum())
    total_built  = sum(prod_by_sku.values())
    total_inv_kpi = sum(gt_inv_dict.get(str(r["SKUCode"]), 0.0) for _, r in df_demand.iterrows())
    n_full  = sum(1 for _, r in df_demand.iterrows()
                  if (prod_by_sku.get(str(r["SKUCode"]), 0)
                      + gt_inv_dict.get(str(r["SKUCode"]), 0.0)) >= float(r["Demand_Qty"]) * 0.95)
    n_part  = sum(1 for _, r in df_demand.iterrows()
                  if 0 < (prod_by_sku.get(str(r["SKUCode"]), 0)
                          + gt_inv_dict.get(str(r["SKUCode"]), 0.0)) < float(r["Demand_Qty"]) * 0.95)
    n_unmet = len(df_demand) - n_full - n_part

    wb = _WB()

    # ── KPI Summary ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "KPI Summary"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")

    kpi_rows = [
        ("Metric",                       "Value"),
        ("Total Customer Demand",        f"{total_demand:,.0f}"),
        ("Total GT Built (New)",         f"{total_built:,.0f}"),
        ("Opening GT Inventory",         f"{total_inv_kpi:,.0f}"),
        ("Coverage % (Built+Inventory)", f"{100*(total_built+total_inv_kpi)/total_demand:.2f}%"),
        ("SKUs Fully Met (≥95%)",        n_full),
        ("SKUs Partial",                 n_part),
        ("SKUs Unmet (0 built)",         n_unmet),
        ("Total COs Scheduled",          co_count),
        ("Scheduler",                    "GreedyCampaignScheduler"),
    ]
    for ri, (label, val) in enumerate(kpi_rows, 1):
        ws.cell(ri, 1, label).font = hdr_font if ri == 1 else Font(bold=(ri in (1, 4)))
        ws.cell(ri, 1).fill = hdr_fill if ri == 1 else PatternFill()
        ws.cell(ri, 2, val)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    # ── Demand Fulfillment ────────────────────────────────────────────────────
    ws_df = wb.create_sheet("Demand Fulfillment (UC)")
    df_cols = ["SKUCode","Priority","Demand","GT_Inventory","Net_Demand",
               "Built_GT","Gap","Fulfillment_%","Status","Eligible_Machines"]
    for ci, col in enumerate(df_cols, 1):
        c = ws_df.cell(1, ci, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    status_fill = {
        "FULLY MET": PatternFill("solid", fgColor=_GREEN),
        "PARTIAL":   PatternFill("solid", fgColor=_AMBER),
        "UNMET":     PatternFill("solid", fgColor=_RED),
    }
    ri = 2
    totals_built = 0
    totals_inv   = 0
    n_f = n_p = n_u = 0
    for _, dr in df_demand.sort_values("Priority_Score", ascending=False).iterrows():
        sku  = str(dr["SKUCode"])
        dem  = int(dr["Demand_Qty"])
        pri  = round(float(dr.get("Priority_Score", 1.0) or 1.0), 4)
        inv  = int(gt_inv_dict.get(sku, 0))
        net  = max(0, dem - inv)
        plan = int(round(prod_by_sku.get(sku, 0)))
        effective = plan + inv                              # built + existing inventory
        gap  = max(0, dem - effective)                     # real unmet gap after inventory
        pct  = round(effective / dem * 100, 1) if dem > 0 else 0.0
        elig = len(allow_map.get(sku, set()))
        totals_built += plan
        totals_inv   += inv
        status = "FULLY MET" if plan >= net else ("PARTIAL" if plan > 0 else "UNMET")
        if status == "FULLY MET": n_f += 1
        elif status == "PARTIAL": n_p += 1
        else:                     n_u += 1
        row_vals = [sku, pri, dem, inv, net, plan, gap, f"{pct}%", status, elig]
        fill = status_fill.get(status, PatternFill())
        for ci, val in enumerate(row_vals, 1):
            c = ws_df.cell(ri, ci, val)
            c.fill = fill; c.alignment = Alignment(horizontal="center")
        ri += 1

    # Footer
    ri += 1
    ws_df.cell(ri,   1, "KPI SUMMARY").font = Font(bold=True)
    ws_df.cell(ri+1, 1, "Total Demand");    ws_df.cell(ri+1, 2, int(total_demand))
    ws_df.cell(ri+2, 1, "Total Built (New GT)"); ws_df.cell(ri+2, 2, totals_built)
    ws_df.cell(ri+3, 1, "Coverage % (Built+Inv)").font = Font(bold=True)
    ws_df.cell(ri+3, 2,
        f"{100*(totals_built+totals_inv)/total_demand:.2f}%").font = Font(bold=True)
    ws_df.cell(ri+4, 1, "Fully Met (≥95%)"); ws_df.cell(ri+4, 2, n_f)
    ws_df.cell(ri+5, 1, "Partial");          ws_df.cell(ri+5, 2, n_p)
    ws_df.cell(ri+6, 1, "Unmet (0 built)");  ws_df.cell(ri+6, 2, n_u)
    ws_df.cell(ri+8, 1,
        "Scheduler: GreedyCampaignScheduler — campaign-to-completion, no daily LP"
    ).font = Font(italic=True)
    for col in "ABCDEFGHIJ":
        ws_df.column_dimensions[col].width = 18

    # ── Machine Utilization ───────────────────────────────────────────────────
    ws_util = wb.create_sheet("Machine Utilization")
    util_cols = ["Machine","Group","Prod_Min","CO_Min","Idle_Min",
                 "Utilization_%","Prod_Util_%","Total_Mins"]
    for ci, col in enumerate(util_cols, 1):
        c = ws_util.cell(1, ci, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    def _mgrp(m):
        if m in {"6001","6002","6003","6004","7001","7002","7003","7004"}: return "VMIMAXX"
        if m in {"7101","7102","7103","7104","7105","7106","7201"}:        return "BJ"
        if m in {"7501","7502","7503"}:                                    return "UNI_NARROW"
        if m in {"7301","8201","8301","8302","8501","8502"}:               return "STAGE2"
        return "STAGE1"

    ri = 2
    for m, s in machine_stats.items():
        row = [m, _mgrp(m),
               round(s["prod_time"],1), round(s["co_time"],1), round(s["idle_time"],1),
               f"{100*s['utilization']:.1f}%", f"{100*s['prod_utilization']:.1f}%",
               round(total_mins,0)]
        for ci, val in enumerate(row, 1):
            ws_util.cell(ri, ci, val).alignment = Alignment(horizontal="center")
        ri += 1
    for col in "ABCDEFGH":
        ws_util.column_dimensions[col].width = 16

    # ── Campaign Schedule ─────────────────────────────────────────────────────
    ws_camp = wb.create_sheet("Campaign Schedule")
    camp_cols = ["Machine","SKUCode","Type","Start_Min","End_Min",
                 "Duration_Min","Qty","Inch"]
    for ci, col in enumerate(camp_cols, 1):
        c = ws_camp.cell(1, ci, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    co_fill   = PatternFill("solid", fgColor="D9D9D9")
    prod_fill = PatternFill("solid", fgColor="E2EFDA")
    for ri, c in enumerate(campaigns, 2):
        is_co = c["SKUCode"] == "CHANGEOVER"
        fill  = co_fill if is_co else prod_fill
        row = [
            c.get("Machine",""), c.get("SKUCode",""),
            "CO" if is_co else "PROD",
            round(c.get("Start_Min",0), 1), round(c.get("End_Min",0), 1),
            round(c.get("Duration_Min",0), 1), round(c.get("Qty",0), 1),
            c.get("Inch", c.get("To_Inch","")) if is_co else c.get("Inch",""),
        ]
        for ci, val in enumerate(row, 1):
            cell = ws_camp.cell(ri, ci, val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
    for col in "ABCDEFGH":
        ws_camp.column_dimensions[col].width = 16

    wb.save(output_path)
    print(f"  [Export] Saved → {output_path}")
    print(f"  [Export] Demand Fulfillment (UC) sheet  "
          f"[{n_f} full / {n_p} partial / {n_u} unmet]")
    return prod_by_sku


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_unconstrained(
    demand_path: str | None = None,
    output_path: str | None = None,
    plan_start: datetime | None = None,
    planning_days: int | None = None,
    engine=None,
) -> dict:
    """
    Run the unconstrained building scheduler.

    Demand is the only ceiling — no curing press counts, no CO schedule,
    no press-rate limits. Measures the theoretical max building capacity.
    """
    from cbc_env import make_engine as _mk

    if plan_start is None:
        plan_start = datetime(2026, 5, 1, 7, 0, 0)
    if demand_path is None:
        demand_path = os.path.join(cbc_env.INPUT_DIR, "demand_may.xlsx")
    if engine is None:
        engine = _mk()
    if planning_days is not None:
        Config.PLANNING_DAYS = planning_days
    else:
        Config.PLANNING_DAYS = 31
    if output_path is None:
        output_path = os.path.join(
            MAIN_OUT, f"bc_building_unconstrained_{plan_start.date()}.xlsx"
        )

    print("\n" + "=" * 70)
    print("  UNCONSTRAINED BUILDING — Max Capacity Experiment")
    print("  (Demand cap only — no curing press / CO constraints)")
    print("=" * 70)

    # ── Building CT from DB ───────────────────────────────────────────────────
    try:
        df_bct = pd.read_sql(
            "SELECT `SAP Machine Code` AS machine, `Cycle Time (Minutes)` AS ct_min "
            "FROM Master_Building_Machine_Design_cycleTime",
            engine,
        )
        db_ct = {
            str(r["machine"]).strip(): round(float(r["ct_min"]) * 60, 4)
            for _, r in df_bct.iterrows()
            if r["machine"] is not None
        }
        Config._CT_SEC.update(db_ct)
        print(f"  [CT] Loaded {len(db_ct)} machine cycle times from DB")
    except Exception as exc:
        print(f"  [CT] DB load failed ({exc}); using hardcoded fallback")

    # ── Build start: 2 shifts before plan_start ───────────────────────────────
    PRE_START_SHIFTS = 2
    build_start = plan_start - timedelta(hours=Config.HOURS_PER_SHIFT * PRE_START_SHIFTS)
    print(f"  [Horizon] plan_start={plan_start}  build_start={build_start}"
          f"  ({PRE_START_SHIFTS} shifts early)  days={Config.PLANNING_DAYS}")

    # ── ETL ───────────────────────────────────────────────────────────────────
    etl = ETL(engine)

    print("\n  [ETL] Loading demand …")
    df_demand = load_demand(demand_path)
    total_demand = float(df_demand["Demand_Qty"].sum())
    print(f"        {len(df_demand)} SKUs  |  {total_demand:,.0f} total units")

    print("  [ETL] Loading GT inventory …")
    df_gt_inv = pd.read_sql(
        f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory "
        f"FROM {Config.DB_NAME}.gt_inventory_manual",
        engine,
    )
    print(f"        {len(df_gt_inv)} SKUs with opening inventory")

    print("  [ETL] Loading carcass inventory (zeroing for cold start) …")
    df_carcass_inv = etl.load_carcass_inventory()
    if df_carcass_inv is not None and not df_carcass_inv.empty:
        df_carcass_inv = df_carcass_inv.copy()
        df_carcass_inv["Carcass_Inventory"] = 0

    print("  [ETL] Loading allowable machines …")
    df_allow = etl.load_machine_allowable()

    print("  [ETL] Loading changeover times …")
    co_map = etl.load_changeover_map()

    print("  [ETL] Loading SKU sizes …")
    sku_to_size = etl.load_sku_sizes()

    print("  [ETL] Loading running building machines …")
    df_running = etl.load_running_machines()

    print("  [ETL] Loading history map (machine discovery) …")
    history_map = etl.load_history_map()

    # ── Merge history → allow_map (machine discovery, NOT sort bias) ──────────
    print("  [Allow] Merging historical machine-SKU pairs …")
    hist_by_sku: dict = {}
    for (machine, sku), count in history_map.items():
        if count > 0:
            hist_by_sku.setdefault(sku, set()).add(machine)

    allow_sku_idx = {str(r["SKUCode"]): i for i, r in df_allow.iterrows()}
    extra_pairs, new_hist_rows = 0, []
    for sku, hist_machs in hist_by_sku.items():
        if sku in allow_sku_idx:
            idx = allow_sku_idx[sku]
            cur_set = set(df_allow.at[idx, "Machines"] or [])
            added = hist_machs - cur_set
            if added:
                df_allow.at[idx, "Machines"] = list(cur_set | added)
                extra_pairs += len(added)
        else:
            new_hist_rows.append({"SKUCode": sku, "Machines": list(hist_machs)})
            extra_pairs += len(hist_machs)
    if new_hist_rows:
        df_allow = pd.concat(
            [df_allow, pd.DataFrame(new_hist_rows)], ignore_index=True
        )
    print(f"  [Allow] +{extra_pairs} pairs from history "
          f"({len(new_hist_rows)} new SKUs via history)")

    # ── Inch-group policy filter ──────────────────────────────────────────────
    removed_group = 0
    for idx, row in df_allow.iterrows():
        sku       = str(row["SKUCode"])
        mach_list = list(row.get("Machines", []) or [])
        sku_inch  = str(sku_to_size.get(sku, "")).strip().replace('"', "")
        if not sku_inch:
            continue
        filtered = [
            m for m in mach_list
            if str(m) not in _UNISTAGE_INCH_POLICY
            or sku_inch in _UNISTAGE_INCH_POLICY[str(m)]
        ]
        removed_group += len(mach_list) - len(filtered)
        df_allow.at[idx, "Machines"] = filtered
    print(f"  [Inch] Removed {removed_group} pairs violating group inch policies")

    # ── Per-machine hard inch lock ────────────────────────────────────────────
    removed_hard = 0
    for idx, row in df_allow.iterrows():
        sku       = str(row["SKUCode"])
        mach_list = list(row.get("Machines", []) or [])
        sku_inch  = str(sku_to_size.get(sku, "")).strip().replace('"', "")
        if not sku_inch:
            continue
        filtered = [
            m for m in mach_list
            if str(m) not in _MACHINE_HARD_INCH
            or sku_inch in _MACHINE_HARD_INCH[str(m)]
        ]
        removed_hard += len(mach_list) - len(filtered)
        df_allow.at[idx, "Machines"] = filtered
    print(f"  [Inch] Removed {removed_hard} pairs via per-machine hard inch lock")

    # ── VMI overflow for BJ-only Unistage SKUs ───────────────────────────────
    # BJ machines (7101-7106, 7201) are oversubscribed: 249k demand vs ~191k
    # physical capacity. VMIMAXX machines (6001-6004, 7001-7004) have 20-45%
    # spare capacity. For pure BJ-Unistage SKUs (no Stage-1/Stage-2 path)
    # whose inch overlaps VMI hard-inch locks, add the compatible VMI machine
    # so the LP/heuristic can spill overflow there.
    _BJ_SET  = {"7101","7102","7103","7104","7105","7106","7201"}
    _S1_SET  = {str(m) for m in Config.STAGE1}
    _S2_SET  = {str(m) for m in Config.STAGE2}
    _VMI_INCH_OVERFLOW: dict[str, set] = {
        "14": {"6001", "7002", "7004"},
        "15": {"6002", "7003"},
        "16": {"7001", "6004"},
        "17": {"6003"},
        "18": {"6003"},
    }
    vmi_overflow_added = 0
    for idx, row in df_allow.iterrows():
        sku      = str(row["SKUCode"])
        mach_set = set(map(str, row.get("Machines", []) or []))
        sku_inch = str(sku_to_size.get(sku, "")).strip().replace('"', "")
        if not sku_inch:
            continue
        # Must have at least one BJ machine to qualify
        if not (mach_set & _BJ_SET):
            continue
        # Skip two-stage SKUs (they use Stage-1 or Stage-2 machines)
        if mach_set & _S1_SET or mach_set & _S2_SET:
            continue
        # Only add VMI machines not already in the set
        candidates = _VMI_INCH_OVERFLOW.get(sku_inch, set())
        new_vmi = candidates - mach_set
        if new_vmi:
            df_allow.at[idx, "Machines"] = list(mach_set | new_vmi)
            vmi_overflow_added += len(new_vmi)
    print(f"  [VMI Overflow] +{vmi_overflow_added} VMI machine-SKU pairs added "
          f"for BJ-only Unistage SKUs")

    # ── Diagnostic: demand SKUs with zero eligible building machines ──────────
    allow_map_lookup = {
        str(r["SKUCode"]): set(map(str, r.get("Machines", []) or []))
        for _, r in df_allow.iterrows()
    }
    zero_machine_skus = [
        str(r["SKUCode"]) for _, r in df_demand.iterrows()
        if not allow_map_lookup.get(str(r["SKUCode"]))
    ]
    if zero_machine_skus:
        print(f"  [WARN] {len(zero_machine_skus)} demand SKUs have 0 eligible "
              f"building machines after inch filter — will produce 0 GT:")
        for s in zero_machine_skus[:10]:
            print(f"    {s}  inch={sku_to_size.get(s, '?')}")
        if len(zero_machine_skus) > 10:
            print(f"    … and {len(zero_machine_skus)-10} more")
    else:
        print("  [OK] All demand SKUs have ≥1 eligible building machine")

    # ── Build demand dict, priority map, demand cap ───────────────────────────
    demand_dict  = {
        str(r["SKUCode"]): float(r["Demand_Qty"])
        for _, r in df_demand.iterrows()
    }
    priority_map = {
        str(r["SKUCode"]): float(r.get("Priority_Score", 1.0) or 1.0)
        for _, r in df_demand.iterrows()
    }
    # ── GT inventory dict ─────────────────────────────────────────────────────
    gt_inv_dict = dict(zip(
        df_gt_inv["SKUCode"].astype(str).str.strip(),
        df_gt_inv["GT_Inventory"].astype(float),
    )) if df_gt_inv is not None and not df_gt_inv.empty else {}

    # Net demand = Demand_Qty − GT_inventory already on hand (cap build at what's needed)
    net_demand_dict = {
        s: max(0.0, d - gt_inv_dict.get(s, 0.0))
        for s, d in demand_dict.items()
    }
    inv_skus = sum(1 for s, d in demand_dict.items() if gt_inv_dict.get(s, 0.0) > 0)
    if inv_skus:
        print(f"  [Greedy] {inv_skus} SKUs have existing GT inventory → net demand reduced")

    # ── Total machine-minutes available (including 2 pre-start shifts) ────────
    # PRE_START_SHIFTS=2 adds 2×480=960 extra min per machine before Day 1.
    total_mins = (Config.PLANNING_DAYS * Config.SHIFTS_PER_DAY + PRE_START_SHIFTS) \
                 * Config.SHIFT_MINS
    print(f"\n  [Greedy] Total machine-minutes per machine: {total_mins:,.0f} "
          f"({Config.PLANNING_DAYS} days × 3 shifts + {PRE_START_SHIFTS} pre-start shifts)")

    # ── Run GreedyCampaignScheduler ───────────────────────────────────────────
    print("  [Greedy] Launching GreedyCampaignScheduler …")
    gsched = GreedyCampaignScheduler()
    campaigns, demand_remaining, machine_stats = gsched.run(
        allow_map_dict = allow_map_lookup,
        demand_dict    = net_demand_dict,
        priority_map   = priority_map,
        sku_to_size    = sku_to_size,
        co_map_raw     = co_map,
        total_mins     = total_mins,
    )

    # ── Print greedy summary ──────────────────────────────────────────────────
    prod_campaigns = [c for c in campaigns if c["SKUCode"] != "CHANGEOVER"]
    co_campaigns   = [c for c in campaigns if c["SKUCode"] == "CHANGEOVER"]
    print(f"  [Greedy] Done — {len(prod_campaigns)} production campaigns, "
          f"{len(co_campaigns)} COs across {len(machine_stats)} machines")

    unmet_skus = [s for s in demand_dict if demand_remaining.get(s, 0) > 1]
    if unmet_skus:
        print(f"  [Greedy] {len(unmet_skus)} SKUs with remaining unbuilt demand:")
        for s in sorted(unmet_skus, key=lambda x: -demand_remaining.get(x, 0))[:10]:
            print(f"    {s}  unbuilt={demand_remaining[s]:,.0f}  "
                  f"machines={list(allow_map_lookup.get(s, set()))}")

    # ── Write Excel output ────────────────────────────────────────────────────
    print(f"\n  [Export] Writing → {output_path}")
    prod_by_sku = _write_greedy_xlsx(
        output_path   = output_path,
        campaigns     = campaigns,
        demand_dict   = demand_dict,
        df_demand     = df_demand,
        gt_inv_dict   = gt_inv_dict,
        allow_map     = allow_map_lookup,
        machine_stats = machine_stats,
        total_mins    = total_mins,
    )

    # ── KPI summary ───────────────────────────────────────────────────────────
    gt_built  = float(sum(prod_by_sku.values()))
    total_inv = sum(gt_inv_dict.get(str(r["SKUCode"]), 0.0) for _, r in df_demand.iterrows())
    n_full   = sum(1 for _, r in df_demand.iterrows()
                   if (prod_by_sku.get(str(r["SKUCode"]), 0)
                       + gt_inv_dict.get(str(r["SKUCode"]), 0.0)) >= float(r["Demand_Qty"]) * 0.95)
    n_part   = sum(1 for _, r in df_demand.iterrows()
                   if 0 < (prod_by_sku.get(str(r["SKUCode"]), 0)
                           + gt_inv_dict.get(str(r["SKUCode"]), 0.0)) < float(r["Demand_Qty"]) * 0.95)
    n_unmet  = len(df_demand) - n_full - n_part
    print("\n" + "=" * 70)
    print("  UNCONSTRAINED BUILDING — KPI SUMMARY (Greedy Campaign Scheduler)")
    print("=" * 70)
    print(f"  Total Customer Demand  : {total_demand:>12,.0f} units")
    print(f"  Total GT Built (New)   : {gt_built:>12,.0f} units")
    print(f"  Opening GT Inventory   : {total_inv:>12,.0f} units")
    print(f"  Coverage (Built+Inv)   : {100*(gt_built+total_inv)/total_demand:>11.2f}%")
    print(f"  SKUs Fully Met (≥95%)  : {n_full}")
    print(f"  SKUs Partial           : {n_part}")
    print(f"  SKUs Unmet             : {n_unmet}")
    print(f"  Total COs              : {len(co_campaigns)}")
    print(f"  Output                 : {output_path}")
    print("=" * 70)

    return {
        "campaigns":      campaigns,
        "machine_stats":  machine_stats,
        "prod_by_sku":    prod_by_sku,
        "demand_dict":    demand_dict,
        "output_path":    output_path,
    }


# ══════════════════════════════════════════════════════════════════════════════
# DEMAND FULFILLMENT SHEET APPENDER
# ══════════════════════════════════════════════════════════════════════════════

_GREEN = "C6EFCE"
_AMBER = "FFEB9C"
_RED   = "FFC7CE"
_GREY  = "F2F2F2"


def _append_demand_fulfillment(
    output_path: str,
    df_demand: pd.DataFrame,
    prod_by_sku: dict,
    gt_inv_dict: dict,
    allow_map: dict,
) -> None:
    """
    Append a 'Demand Fulfillment (Unconstrained)' sheet to the Excel output
    showing per-SKU demand vs actual build vs gap, coloured by status.
    """
    try:
        wb = load_workbook(output_path)
        sheet_name = "Demand Fulfillment (UC)"

        rows = []
        for _, dr in df_demand.sort_values("Priority_Score", ascending=False).iterrows():
            sku  = str(dr["SKUCode"])
            dem  = int(dr["Demand_Qty"])
            pri  = round(float(dr.get("Priority_Score", 1.0) or 1.0), 4)
            inv  = int(gt_inv_dict.get(sku, 0))
            net  = max(0, dem - inv)
            plan = int(prod_by_sku.get(sku, 0))
            gap  = max(0, dem - plan)
            pct  = round(plan / dem * 100, 1) if dem > 0 else 0.0
            elig = len(allow_map.get(sku, set()))
            status = (
                "FULLY MET" if plan >= net
                else "PARTIAL" if plan > 0 else "UNMET"
            )
            rows.append({
                "SKUCode": sku, "Priority": pri, "Demand": dem,
                "GT_Inventory": inv, "Net_Demand": net,
                "Built_GT": plan, "Gap": gap,
                "Fulfillment_%": f"{pct}%", "Status": status,
                "Eligible_Machines": elig,
            })

        df_out = pd.DataFrame(rows)

        ws = wb.create_sheet(sheet_name)
        cols = list(df_out.columns)
        status_colors = {
            "FULLY MET": _GREEN, "PARTIAL": _AMBER, "UNMET": _RED,
        }

        # Header
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for ri, row_data in enumerate(df_out.itertuples(index=False), start=2):
            status = row_data.Status
            fill_hex = status_colors.get(status, _GREY)
            fill = PatternFill("solid", fgColor=fill_hex)
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

        # KPI footer
        total_dem   = int(df_demand["Demand_Qty"].sum())
        total_built = sum(int(prod_by_sku.get(str(r["SKUCode"]), 0))
                          for _, r in df_demand.iterrows())
        kpi_pct = round(100 * total_built / total_dem, 1) if total_dem else 0.0
        n_full  = sum(1 for r in rows if r["Status"] == "FULLY MET")
        n_part  = sum(1 for r in rows if r["Status"] == "PARTIAL")
        n_unmet = sum(1 for r in rows if r["Status"] == "UNMET")

        footer = len(df_out) + 3
        ws.cell(footer,   1, "KPI SUMMARY").font = Font(bold=True)
        ws.cell(footer+1, 1, "Total Demand");   ws.cell(footer+1, 2, total_dem)
        ws.cell(footer+2, 1, "Total Built");    ws.cell(footer+2, 2, total_built)
        _kpi = ws.cell(footer+3, 1, "Coverage %"); _kpi.font = Font(bold=True)
        _kpv = ws.cell(footer+3, 2, f"{kpi_pct}%"); _kpv.font = Font(bold=True)
        ws.cell(footer+4, 1, "Fully Met (≥95%)"); ws.cell(footer+4, 2, n_full)
        ws.cell(footer+5, 1, "Partial");          ws.cell(footer+5, 2, n_part)
        ws.cell(footer+6, 1, "Unmet (0 built)");  ws.cell(footer+6, 2, n_unmet)
        ws.cell(footer+8, 1,
                "Note: no curing press / CO constraints — demand cap only").font = Font(italic=True)

        wb.save(output_path)
        print(f"  [Export] Demand Fulfillment (UC) sheet appended  "
              f"[{n_full} full / {n_part} partial / {n_unmet} unmet]")
    except Exception as exc:
        print(f"  [WARN] Could not append Demand Fulfillment sheet: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    _demand = sys.argv[1] if len(sys.argv) > 1 else None
    run_unconstrained(demand_path=_demand)
