# Just confirm that new table in db- building allowable machines is got updated succesfully and now for more SKU will be eligible on the VMI machines in allowable matrix. 
# Now, I have updated the building allowable machines data from the plant, Now, i think we should get the better KPIs in the building side so accordingly curing side KPIs should also get better automatically. But, now i am getting the output KPIs around- GT built= 609507 and curing 610955, from the same code we are able to generate GT of around 615535 and curing qty around- 616626. Why our KPIs got decreased now. So, where we are making mistake now. We only have to work on this- update in the inch size locking to machine group/machine? Anything else needed or some issues in the logic as well. Now, I have added 2 more SKU in the demand file- 1. "1D25212812074FXC10	28089	0.040417398	and 1325215813079TTMX0	12521	0.018016527	[sku, qty, prioirty_score]. And, after adding more than 40k qty we are getting the KPIs as building- 615,533 and 617,359. Now, why we laggin in KPIs now? Find the Root cause for the same. 


"""
b2c_pipeline.py — End-to-end B2C scheduling pipeline.

Two modes:
  LEGACY (run_pipeline):
    Step 1: Curing Consumption (dynamic) — 31-day CO schedule
    Step 2: Building Scheduler (B2C)     — 31-day LP building plan
    Step 3: Curing Schedule (B2C)        — shift-wise curing from GT output

  ROLLING (run_rolling_pipeline):   <-- NEW DEFAULT
    Pre-compute: CO schedule + master data (once)
    For each Day D (1..31):
      1. Compute curing demand from actual press_state
      2. Greedy building assignment (projected GT — Option B)
      3. Per-shift: add building to GT inventory FIRST, then cure min(capacity, gt_available)
      4. GT shelf-life writeoff at end of day
      5. Apply CO transitions

    Output: SAME Excel files and sheet names as the legacy pipeline:
      Building → bc_building_schedule_{date}.xlsx
        Sheets: Shift Schedule | Changeover Plan | SKU Classification |
                Daily GT & Carcass | Demand Fulfillment (B2C)
      Curing  → bc_curing_b2c.xlsx
        Sheets: Demand Fulfillment | Machine Utilization | Shift Schedule |
                Mould Tracker | Machine Schedule | Daily Cured tyres | GT Gap Diagnostic

All parameters are read from bc_config.py — edit there, not here.

Usage:
    python b2c_pipeline.py                           # rolling pipeline (new default)
    python b2c_pipeline.py data/input/demand_may.xlsx
    python b2c_pipeline.py --legacy                  # run old 31-day LP pipeline
"""

import os
import math
import sys
import tempfile
import hashlib
from collections import defaultdict, Counter
from datetime import datetime, timedelta

# VERBOSE terminal logging. Default OFF = quiet (only the setup lines, warnings, the final
# KPI summary, and output paths print). VERBOSE=1 restores the per-day / per-machine trace
# (Pool moves, deferrals, pre-positions, per-day built/cured, cold-start / Runner-Out lines).
_VERBOSE = (os.environ.get("VERBOSE", "0") != "0")

import pandas as pd

import bc_config
from curing_consumption_dynamic import (
    run_dynamic_consumption, ConsumptionConfig, COScheduler, _SURPLUS_RELEASE_ENABLED,
    press_efficiency as _press_efficiency,
)
from building_b2c import run_from_database_b2c
from curing_b2c import run_curing_b2c
from connection import ConsumptionETL

# ── All params from bc_config (single source of truth) ────────────────────────
from bc_config import (
    PLAN_START,
    PLANNING_DAYS,
    DEMAND_FILE,
    GT_SHELF_LIFE_DAYS,
    MAX_ENDOFDAY_GT_INVENTORY,
    MOULD_CLEAN_CYCLES,
    MOULD_CLEAN_MINS,
    CURING_CO_CHANGEOVER_MINS,
    CURING_PRESS_COUNT,
    MAX_CHANGEOVERS_PER_DAY,
    VMI_JIT_MARGIN,
    VMI_MAX_DIFF_CO_PER_DAY,
    RT_SAME_INCH,
    MIN_CAMPAIGN_MINS,
    MIN_CAMPAIGN_UNITS,
    MIN_INCH_DWELL_DAYS,
    MAX_BUILDING_SKUS_PER_DAY,
    INCH_PLUS3_CO_MINS,
    INCH_PLUS3_MIN_DAYS_LEFT,
    DIFF_CO_MIN_DWELL_DAYS,
    DIFF_CO_MIN_TARGET_UNITS,
    BUILD_LEAD_SHIFTS,
    MAX_BUILDING_COS_PER_MACHINE_PER_SHIFT,
    GT_BUFFER_SHIFTS,
    GT_BUFFER_SHIFTS_VMI,
    GT_BUFFER_SHIFTS_OTHER,
    BUILDING_MACHINE_NAMES,
    BUILDING_CO_SAME_SIZE,
    BUILDING_CO_DIFF_SIZE,
    BJ_SAME_SIZE_CO_EXCEPTIONS,
    SHIFT_MINS,
    SHIFT_STARTS,
    POOL_SIZE,
    STARVATION_BUFFER_MINS,
    CO_CLASS_B_THRESHOLD,
    DYNAMIC_CC_OUTPUT  as CC_OUTPUT,
    BUILDING_OUTPUT    as BUILD_OUTPUT,
    CURING_B2C_OUTPUT  as CURING_OUTPUT,
)
import bc_config as _bc_cfg

# Optional env override for the daily curing-CO cap — lets us sweep it (e.g. 8-13)
# without editing bc_config. Unset ⇒ the committed bc_config value.
if os.environ.get("MAX_CO"):
    MAX_CHANGEOVERS_PER_DAY = int(os.environ["MAX_CO"])

# Optional env override for the max distinct SKUs a building machine may build per day
# (committed default 4 in bc_config). Lets us A/B a tighter cap (e.g. 3) without editing
# bc_config. Unset ⇒ the committed value. Gated by _BLD_SKU_CAP_ENABLED as usual.
if os.environ.get("BLD_SKU_MAX"):
    MAX_BUILDING_SKUS_PER_DAY = int(os.environ["BLD_SKU_MAX"])

# Optional OUT_TAG — suffix all output files so parallel runs don't clobber each
# other (used for concurrent CO-cap sweeps). Unset ⇒ the normal output paths.
if os.environ.get("OUT_TAG"):
    _ot = os.environ["OUT_TAG"]
    BUILD_OUTPUT  = BUILD_OUTPUT.replace(".xlsx", f"_{_ot}.xlsx")
    CURING_OUTPUT = CURING_OUTPUT.replace(".xlsx", f"_{_ot}.xlsx")
    CC_OUTPUT     = CC_OUTPUT.replace(".xlsx", f"_{_ot}.xlsx")

# ── Machine group map ─────────────────────────────────────────────────────────
# NOTE: these values are INTERNAL LOGIC KEYS — they are compared against string
# literals throughout (_S1_MACHINES, _buf_of, _co_cost, _pri, captive-max, the
# Stage-1 carcass step) AND used as dict keys into bc_config's
# BUILDING_CO_SAME_SIZE / BUILDING_CO_DIFF_SIZE. Do NOT rename them.
# For plant-friendly names in the output sheets, edit _MACHINE_GROUP_DISPLAY below.
_MACHINE_GROUP: dict[str, str] = {}
for _m in ("6001","6002","6003","6004","7001","7002","7003","7004"):
    _MACHINE_GROUP[_m] = "VMI"
for _m in ("7101","7102","7103","7104","7105","7106","7201"):
    _MACHINE_GROUP[_m] = "BJ"
for _m in ("7501","7502","7503"):
    _MACHINE_GROUP[_m] = "UNISTAGE"
for _m in ("8201","8301","8302","8501","8502","7301"):
    _MACHINE_GROUP[_m] = "STAGE2"
for _m in ("6801","6802","6803","6909","6911","7601","7701",
           "7801","7802","7803","7804","8001","8002","8003","8101"):  # Stage-1 (15; 6801 is Stage-1 carcass, not GT)
    _MACHINE_GROUP[_m] = "STAGE1"
for _m in ("ps2","ps3","ps4"):   # NEW GT machines (independent GT, like Unistage; CO 25/45 via "PS" key)
    _MACHINE_GROUP[_m] = "PS"

# Plant-facing display labels — used ONLY for the Machine_Group column in the
# building Shift Schedule output sheet. Never used in scheduling logic, so these
# are safe to rename freely.
_MACHINE_GROUP_DISPLAY: dict[str, str] = {
    "VMI":      "VMIMAXX GROUP",
    "BJ":       "BJ GROUP",
    "UNISTAGE": "UNISTAGE GROUP",
    "STAGE2":   "TBM STAGE2",
    "STAGE1":   "TBM STAGE1",
}


def _group_label(machine: str) -> str:
    """Plant-facing group name for output sheets (falls back to the internal key)."""
    g = _MACHINE_GROUP.get(str(machine), "")
    return _MACHINE_GROUP_DISPLAY.get(g, g)


_S1_MACHINES = frozenset(m for m, g in _MACHINE_GROUP.items() if g == "STAGE1")

# ── SOFT "one SKU → one machine group" constraint (env SAME_GROUP, default OFF) ──
# Reduces UNNECESSARY cross-group shuffling of a GT SKU on plan Day 3+ (after the 2-day
# plant replay). A FINER group map than _MACHINE_GROUP: it SPLITS the VMI group into
# VMI-Maxx (7001-7004) and VMI-Exium (6001-6004) because those are physically distinct
# machine families that the plant does not treat as interchangeable. Used ONLY by this
# lever — _MACHINE_GROUP (which drives CO-time keys, Stage-1 detection, etc.) is untouched.
_SG_MACHINE_GROUP: dict[str, str] = {}
for _m in ("ps2", "ps3", "ps4"):
    _SG_MACHINE_GROUP[_m] = "PS"
for _m in ("7001", "7002", "7003", "7004"):
    _SG_MACHINE_GROUP[_m] = "MAXX"
for _m in ("6001", "6002", "6003", "6004"):
    _SG_MACHINE_GROUP[_m] = "EXIUM"
for _m in ("7101", "7102", "7103", "7104", "7105", "7106", "7201"):
    _SG_MACHINE_GROUP[_m] = "BJ"
for _m in ("7501", "7502", "7503"):
    _SG_MACHINE_GROUP[_m] = "US"
for _m in ("8201", "8301", "8302", "8501", "8502", "7301"):
    _SG_MACHINE_GROUP[_m] = "STAGE2"
for _m in ("6801", "6802", "6803", "6909", "6911", "7601", "7701",
           "7801", "7802", "7803", "7804", "8001", "8002", "8003", "8101"):
    _SG_MACHINE_GROUP[_m] = "STAGE1"


def _sku_group_of(machine: str) -> str:
    """Finer 7-group label (Maxx/Exium split) for the same-group soft lever."""
    return _SG_MACHINE_GROUP.get(str(machine), "")


# The GT groups a "home group" may be pinned to (Stage-1 carcass is inherently its own
# group and is never a home for a GT SKU).
_SG_GT_GROUPS = frozenset(("PS", "MAXX", "EXIUM", "BJ", "US", "STAGE2"))

# Toggle (env SAME_GROUP=1/0, or bc_config.SAME_GROUP_SOFT_A). DEFAULT ON (SOFT-A adopted
# 2026-09): SAME_GROUP=0 fully reverts to the bit-for-bit OFF baseline.
_SAME_GROUP_SOFT = os.environ.get(
    "SAME_GROUP",
    "1" if getattr(_bc_cfg, "SAME_GROUP_SOFT_A", True) else "0") != "0"
# Penalty WEIGHT (rank levels a cross-group pair is demoted). Default 1 = one soft rank.
# 0 → inert even when the toggle is on (home == cross == 0).
_SAME_GROUP_PEN = int(os.environ.get(
    "SAME_GROUP_PEN", str(getattr(_bc_cfg, "SAME_GROUP_PEN", 1))))
# Placement strength (env SG_STRONG, default 0 = weak). WEAK: the home-group term sits
# AFTER the continuous per-machine demand-ratio `primary`, so it only breaks exact ties
# (rarely fires → near-neutral). STRONG: it sits AFTER the RI-vs-NRI tier but BEFORE
# `primary`, so among same-inch, same-RI-status candidates a home pairing is preferred over
# a cross one — this actually bites, at the risk of deferring a slightly-more-urgent cross
# SKU behind a home one (still bounded: it never crosses the inch band or the RI/NRI tier,
# and a home machine with no free capacity is simply absent from the pool → cross is taken).
_SG_STRONG = os.environ.get("SG_STRONG", "0") != "0"

# EXCEPTION list (bc_config.SAME_GROUP_SOFT_A_EXEMPT_SKUS, env SG_EXEMPT_SKUS override):
# SKUs that are EXEMPT from the whole SAME_GROUP rule — they may be built on ANY allowable
# machine across groups (incl. simultaneously in different groups the same shift). Every
# SAME_GROUP gate/penalty short-circuits to neutral for a listed SKU (home-pen 0, move-pen 0,
# no HARD same-shift/purity drop); all other constraints (demand cap, inch-lock, mould, CT)
# still apply. Empty → no exemption (bit-for-bit).
_SG_EXEMPT_SKUS: set = set(
    s.strip() for s in os.environ.get(
        "SG_EXEMPT_SKUS",
        ",".join(getattr(_bc_cfg, "SAME_GROUP_SOFT_A_EXEMPT_SKUS", []) or [])
    ).split(",") if s.strip())
if _SG_EXEMPT_SKUS:
    print(f"  [SAME_GROUP] {len(_SG_EXEMPT_SKUS)} SKU(s) EXEMPT from the group rule "
          f"(any allowable machine): {', '.join(sorted(_SG_EXEMPT_SKUS))}")

# ── DELIBERATE + STABLE per-SKU group allocation (extends SAME_GROUP) ───────────────
# When SAME_GROUP is ON, this sub-mode (SG_DELIBERATE, default ON) REPLACES the old soft
# _group_pen tiebreak with a DELIBERATE, STABLE group assignment:
#   1. each GT SKU gets a deliberate TARGET group SET (_best_group / _compute_grp_targets):
#      the single group whose monthly capacity completes the SKU's remaining demand; if none
#      suffices, the MINIMAL completing SET (high-demand SKUs like LSTL0/SUNE1).
#   2. a pair (machine, sku) whose finer group is IN the SKU's target set costs 0; a pair
#      OUTSIDE it is a cross-group MOVE, admitted (penalty waived) only when the SKU is
#      genuinely starving in its target AND a per-SKU cooldown has elapsed (hysteresis) —
#      else it is deprioritized by _HYST_BIG so the pair loses (stable, no per-shift churn).
#      An admitted move ADDS the new group to the SKU's target set (permanently sanctioned →
#      no ping-pong) and stamps the move day.
#   3. a HARD guard forbids the same SKU being produced in >1 group in the SAME shift.
# SG_DELIBERATE=0 → the old soft-tiebreak SAME_GROUP behaviour (still gated by SAME_GROUP).
# SAME_GROUP=0 → the whole feature is inert (bit-for-bit).
_SG_DELIB = _SAME_GROUP_SOFT and (os.environ.get("SG_DELIBERATE", "1") != "0")
# DELIBERATE MOVE-GATE master (env SG_MOVE_ADMIT, default OFF). When OFF, a SKU NEVER changes
# its frozen target group set at runtime → maximum stability (distinct-group = frozen set size,
# ≈1 except the minimal-SET mega-SKUs) at the accepted coverage cost (priority #7). When ON, a
# cross-group MOVE may be admitted by the hysteresis below (dead-band + structural-gap + cooldown)
# — a relief valve for SLACK months where recovering coverage is worth an occasional 2nd group.
_SG_MOVE_ADMIT = os.environ.get("SG_MOVE_ADMIT", "0") != "0"
# Dead-band: a cross-group move is admitted only if the SKU's target group is failing to feed
# it — projected GT below (1 + band)·this-shift-draw. Larger band → moves fire more readily.
_SG_MOVE_BAND = float(os.environ.get("SG_MOVE_BAND", str(getattr(_bc_cfg, "SG_MOVE_BAND", 0.15))))
# Cooldown: min days between successive group MOVES for one SKU (anti-ping-pong).
_SG_MOVE_COOLDOWN_DAYS = int(os.environ.get(
    "SG_MOVE_COOLDOWN_DAYS", str(getattr(_bc_cfg, "SG_MOVE_COOLDOWN_DAYS", 3))))
# STRUCTURAL move trigger: a deliberate cross-group move is admitted only when the SKU's
# TARGET group is falling structurally behind — its cumulative monthly gap (demand_remaining
# − projected GT) exceeds this many shifts of its current draw (i.e. the group cannot keep
# up), NOT merely momentary per-shift starvation. Larger → moves fire less readily. 9 = ~3
# working days behind. This is what keeps distinct-group ≈ 1 (moves are rare + deliberate).
_SG_MOVE_GAP_SHIFTS = float(os.environ.get(
    "SG_MOVE_GAP_SHIFTS", str(getattr(_bc_cfg, "SG_MOVE_GAP_SHIFTS", 9.0))))
# Group-capacity DERATE for the _best_group completability test: a single group is judged able
# to complete a SKU only if its DERATED monthly building capacity ≥ remaining demand. Raw
# floor(shift/ct)·shifts overstates deliverable output (COs, contention with the group's other
# SKUs, curing-draw limits), so a mega-SKU that no group can really finish alone would wrongly
# read as single-group-completable. 0.55 cleanly separates the two 65-69k mega-SKUs (→ minimal
# SET) from the ≤45k rest (→ single group). 1.0 = raw (no derate).
_SG_GRP_CAP_DERATE = float(os.environ.get(
    "SG_GRP_CAP_DERATE", str(getattr(_bc_cfg, "SG_GRP_CAP_DERATE", 0.55))))
# INCH-AWARE stable target sizing (env SG_INCH_AWARE, default ON). "Prefer 1 group, add a stable
# 2nd group only when genuinely needed." The plain completability test in _compute_grp_targets
# sizes a group's capacity against ONE SKU's demand — it ignores INCH CONTENTION (all SKUs on an
# inch share that group's inch-locked machines), so on an oversubscribed inch (13"/15") a group
# looks "sufficient", gets locked, then saturates while sibling groups idle (measured −26.6k on
# Sept). Inch-aware sizing instead credits each SKU its DEMAND-PROPORTIONAL SHARE of a group's
# capacity (share = cap[g]·min(1, demand_s / Σdemand of inch SKUs competing for g)); a SKU on an
# oversubscribed inch then needs ≥2 groups → gets a STABLE 2-group target UP-FRONT (frozen at day
# 3, no ping-pong, no runtime move-gate). Non-saturated inches still resolve to 1 group.
_SG_INCH_AWARE_TARGETS = os.environ.get("SG_INCH_AWARE", "1") != "0"
# HARD group purity: a non-admitted cross-group pair is DROPPED from the candidate pool (the
# machine idles rather than build a foreign SKU) — the faithful realization of "one group at
# a time" (priority #7: stable groups even at a coverage cost). DEFAULT OFF (SOFT-A adopted
# 2026-09): keeps only the _HYST_BIG soft penalty (foreign builds allowed when a machine has
# no in-target deficit SKU → some oversubscription spill, higher coverage). SG_HARD=1 restores
# the hard-drop behavior for A/B.
_SG_HARD = os.environ.get("SG_HARD", "0") != "0"
# SAME-SHIFT purity guard (env SG_SAMESHIFT_HARD, DEFAULT True = current hard behavior).
# The HARD no-two-groups-same-shift guard forbids a SKU being produced in >1 finer group in
# the SAME shift (invariant #3, enforced by the `_shift_grp` skip in Phase-B/C). When relaxed
# (=0), an IDLE machine may build a foreign-group SKU even when that SKU is already building in
# its home group this shift — the "Soft-B" idle-avoidance path (recovers coverage at the cost
# of allowing a SKU in 2 groups within one shift). Home group is STILL preferred via the
# _sg_move_pen scoring penalty (a cross build only wins when it avoids an idle machine, never
# gratuitously). Downstream of PM/MTC/demand-cap/inch/allowable filters; day>2 only (rides
# _sg_delib, which is False for the Days-1-2 plant replay). SAME_GROUP=0 → inert (bit-for-bit).
_SG_SAMESHIFT_HARD = os.environ.get("SG_SAMESHIFT_HARD", "1") != "0"

# ── Machine-level CONTINUATION stickiness (the CORE of the same-group feature) ──────
# Eliminates avoidable building changeovers caused by SKU MIGRATION across machines:
# a machine STAYS on its current SKU across shifts whenever that SKU still has curable
# in-demand work (demand left AND a live curing draw) — rather than releasing the machine
# (once its thin dynamic buffer is momentarily full) so Phase B gratuitously reshuffles the
# SKU to another machine (a CO) while this one idles or CO's elsewhere. Continuation is
# extended to a shelf-safe, demand-capped, curable-capped depth (never overbuilds past the
# 3-day GT shelf or the demand cap), so keeping the machine on its own SKU is pure upside.
# Enabled together with the group-home soft rule by SAME_GROUP (env STICKY_MACHINE overrides).
_STICKY_MACHINE = os.environ.get(
    "STICKY_MACHINE",
    "1" if (_SAME_GROUP_SOFT or getattr(_bc_cfg, "STICKY_MACHINE", False)) else "0") != "0"
# Continuation depth in shifts (how far a machine keeps feeding its own SKU ahead of draw).
# Sept sweep: cured cost RISES with depth (2→-14.7k, 4→-9.6k, 6→-14.9k, 9→-22.1k) while the
# building-CO cut also rises (4→-463, 9→-524) — a genuine trade-off (deeper continuation
# front-loads GT that partly expires under the 3-day shelf). Committed opt-in default = 4
# (least-bad on the sweep); STICKY_CONT_SHIFTS=0 disables the continuation extension entirely
# (leaves only the weak pickup+group soft tiebreak, which is cured-neutral: Sept -73).
_STICKY_CONT_SHIFTS = int(os.environ.get(
    "STICKY_CONT_SHIFTS", str(getattr(_bc_cfg, "STICKY_CONT_SHIFTS", 4))))
# Phase-B pickup damper: penalize a machine CO'ing to a SKU ANOTHER machine is already
# validly building (a migration/duplication) unless the deficit genuinely needs a 2nd
# machine. Soft rank term; 0 → off. Default ON with the feature.
_STICKY_PICKUP_PEN = int(os.environ.get(
    "STICKY_PICKUP_PEN", str(getattr(_bc_cfg, "STICKY_PICKUP_PEN", 1))))

# ── NARROW day-2→day-3 BOUNDARY continuation (STICKY_HANDOFF, default OFF) ──────────
# A one-time, cheap variant of the full-month continuation above: at the FIRST assigned
# day (day 3, Shift A — the shift right after the 2-day plant replay ends) each GT machine
# CONTINUES the SKU it carried at the END of day 2, instead of letting Phase-B / global
# reassignment swap it off at the boundary (the ~9 avoidable day2→day3 building COs, e.g.
# 7105 TUNE6→SUNE1 abandoning TUNE6 at 20%). Applies ONLY when the carried SKU still has
# curable in-demand work for that machine (demand left AND a live curing draw); a genuinely
# demand-done / undrawn / would-starve carry is still free to change. INDEPENDENT of
# _STICKY_MACHINE / _SAME_GROUP (its own toggle). Because it fires on ONE shift only, it
# avoids the boundary shuffle at ~zero coverage cost (unlike the per-shift CONT lever).
_STICKY_HANDOFF = os.environ.get(
    "STICKY_HANDOFF",
    "1" if getattr(_bc_cfg, "STICKY_HANDOFF", False) else "0") != "0"
# Boundary continuation depth (shifts of draw the machine may bank at the handoff). Kept
# THIN (default 2 ≈ the VMI flat buffer) so the one-time continuation does not front-load.
_STICKY_HANDOFF_SHIFTS = int(os.environ.get(
    "STICKY_HANDOFF_SHIFTS", str(getattr(_bc_cfg, "STICKY_HANDOFF_SHIFTS", 2))))
# HARD lock sub-mode (env STICKY_HANDOFF_LOCK): also exclude the boundary machine from
# Phase B/C this shift (guarantees no swap even when the carried SKU's GT is already banked
# past the thin continuation depth, but idles residual capacity → costs coverage on a
# saturated month). Default OFF: continuation-room-only keeps the carried SKU as the day-3
# PRIMARY (Phase A builds it first) while leaving residual capacity to Phase B/C.
_STICKY_HANDOFF_LOCK = os.environ.get(
    "STICKY_HANDOFF_LOCK",
    "1" if getattr(_bc_cfg, "STICKY_HANDOFF_LOCK", False) else "0") != "0"

# ── PLANT_SET_LOCK: each GT machine committed to its plant Days-1-2 SKU SET (day 3+) ──────
# From day 3 on, a GT-producing machine (VMI/BJ/Unistage/PS + Stage-2) may build ONLY the
# SKUs the plant ran on it in Days 1-2 (its 1-3 SKUs) — fed first, rotating among them, never
# abandoned — until every one of those SKUs is demand-complete, after which the machine is
# RELEASED to its full allowable matrix. Spare/idle capacity (Phase C) still serves other
# SKUs. Stage-1 is EXCLUDED (its carcass auto-follows the Stage-2 SKUs). Days 1-2 are the
# plant replay (assigner bypassed). PLANT_SET_LOCK=0 → empty set → identity, bit-for-bit OFF.
_PLANT_SET_LOCK = os.environ.get(
    "PLANT_SET_LOCK",
    "1" if getattr(_bc_cfg, "PLANT_SET_LOCK", False) else "0") != "0"
# (mid-month activation boundary _PLANT_SET_LOCK_FROM is defined below, after _PLANT_2DAY_DAYS)
# STARVING-PRESS CO BYPASS: allow a SAME-INCH building changeover past the 30%-of-remaining CO-cost
# guard when the target SKU's curing presses are RUNNING but STARVED (live draw, no GT on hand or
# built this shift, demand remaining). Otherwise a machine with a half-used shift idles its residual
# minutes rather than pay a cheap same-inch CO to feed its own empty presses (e.g. 7503→TUHL0-73).
# Same-inch only + demand/curable-capped → no waste-GT risk. Env CO_STARVE_BYPASS=0 reverts.
_CO_STARVE_BYPASS = os.environ.get(
    "CO_STARVE_BYPASS",
    "1" if getattr(_bc_cfg, "CO_STARVE_BYPASS", True) else "0") != "0"
# PLANT-LOCK STARVED-FEED (Fix C): let a plant-locked machine build a SAME-INCH, live-draw, STARVED
# non-plant SKU (presses RUNNING, 0 GT on hand + none built this shift, demand left) even when its
# plant SKUs still show a nominal deficit — because those plant SKUs are draw-capped and a real
# press is waiting empty. Same-inch keeps the CO cheap; bounded to genuinely starved targets → no
# waste-GT. Env PLANT_STARVE_FEED=0 reverts.
_PLANT_STARVE_FEED = os.environ.get(
    "PLANT_STARVE_FEED",
    "1" if getattr(_bc_cfg, "PLANT_STARVE_FEED", False) else "0") != "0"   # default OFF: measured −4,868 (churn), kept for the record
# SPARE-RELEASE threshold (units): a locked machine may ALSO build other curing-drawn SKUs in a
# shift once its plant-set SKUs can no longer use a meaningful chunk of its time — i.e. their best
# residual draw-deficit is below this many units (a machine can't fill a real campaign with them
# because other machines already feed them). The plant SKU is NEVER removed (stays the committed
# primary, rebuilt whenever it regains real draw); only the machine's IDLE time goes to others.
# Fixes the single-heavily-shared-never-completing-SKU strand (e.g. 7106 → SUNE1). 0 → old strict
# gate (spare only when plant deficit is exactly 0). Env PLANT_SPARE_MIN_UNITS overrides.
_PLANT_SPARE_MIN_UNITS = float(os.environ.get(
    "PLANT_SPARE_MIN_UNITS", str(getattr(_bc_cfg, "PLANT_SPARE_MIN_UNITS",
                                         getattr(_bc_cfg, "MIN_CAMPAIGN_UNITS", 40)))))

# ── Idle-recoverability diagnostic (env IDLE_DIAG=1, read-only, plan-neutral) ──
# For each (machine, shift) with meaningful idle time, decide whether a REACHABLE
# SKU (allowable + in ±2 band + dwell-OK) with a live press draw and a GT deficit
# went unbuilt this shift. Splits the momentary curing shortfall into:
#   - recoverable_units : a reachable idle machine could have pre-built the missing GT
#                         → allocation/timing-fixable WITHOUT relaxing the plant rules
#   - ceiling_units     : no reachable machine existed → true curing/press/mould ceiling
# The accumulator is a plain dict so it survives across the day loop; printed at run end.
_IDLE_DIAG_ON = os.environ.get("IDLE_DIAG", "0") != "0"
_IDLE_DIAG = {
    "rec_units": 0.0, "ceil_units": 0.0,          # per-shift press shortfall, by reachability
    "rec_shifts": 0, "ceil_shifts": 0,            # #(machine,shift) idle buckets touched
    "gt_idle_min": 0.0, "s1_idle_min": 0.0,       # idle minutes, GT vs Stage-1
    "rec_by_inch": {}, "ceil_by_inch": {},        # shortfall units by inch
}

# Round-trip buffer sizing: when a machine alternates between its current SKU
# and another live, unfulfilled SKU, the buffer left behind for the current
# SKU must survive CO(cur->partner) + partner's own dwell time + CO(partner->cur),
# not just a flat GT_BUFFER_SHIFTS multiplier. Skipped entirely (falls back to
# the flat buffer) when the machine has only one eligible SKU, when no other
# eligible SKU has unmet demand, or when no other eligible SKU currently has a
# real curing-driven deficit — see _assign_building_shift.
_ROUND_TRIP_BUFFER_ENABLED = True

# Round-trip partner SAME-INCH preference (env RT_SAME_INCH). The round-trip buffer sizes itself to
# a rotation partner; today it picks the HIGHEST-DEFICIT partner, which may be a different inch → an
# expensive 120-min diff round-trip that oversizes the buffer (front-loading). ON prefers a SAME-INCH
# partner (a ~20-min same_size_CO, no building diff-CO, consistent with Phase-B same-inch-first) →
# smaller, more accurate buffer. Measured: big win on low-demand months (May +15k, starv down, VMI
# 11→8) but a HARD same-inch restriction HURTS high-demand July (−14k) — July's imbalanced demand
# genuinely needs cross-inch rotations. So it is DEFICIT-AWARE: prefer the same-inch partner only when
# its deficit ≥ RT_SAME_INCH_FRAC × the best (any-inch) partner deficit; else fall back to the biggest
# deficit even if off-inch (July keeps its flexibility). OFF → deficit-first partner, bit-for-bit.
# DEFAULT OFF. Month-dependent: hard same-inch (FRAC=0) is +14,964 on the low-demand month (May,
# VMI 11→8 = plant, starvation down) but −14k on high-demand July (needs cross-inch rotations); the
# deficit-aware FRAC>0 gate cannot keep May's gain without July's loss. So it is a LOW-DEMAND-MONTH
# lever (set RT_SAME_INCH=1 for May-like months, leave OFF for July). FRAC default 0 = hard restrict.
_RT_SAME_INCH = os.environ.get("RT_SAME_INCH", "1" if RT_SAME_INCH else "0") != "0"  # bc_config per-month
_RT_SAME_INCH_FRAC = float(os.environ.get("RT_SAME_INCH_FRAC", "0"))

# T2 — departure-gated round-trip buffer (env RT_IMMINENT, default OFF). Today the wide round-trip
# cushion (eff_buf) is held EVERY shift the SKU is current → front-loads even machines that never
# rotate. ON: Phase A builds only the FLAT buffer; the eff_buf cushion becomes a GATE — a machine may
# CO away from its current SKU only once that SKU holds eff_buf (else it tops the current SKU up toward
# eff_buf and DEFERS the rotation to a later shift). So the wide cushion is paid only by machines that
# actually rotate → less front-load. OFF → Phase A builds to eff_buf, no gate (bit-for-bit).
_RT_IMMINENT = os.environ.get("RT_IMMINENT", "0") != "0"

# T1 — SKU-B-aware round-trip sizing (env RT_PARTNER_RT, default OFF). Today A's away-cushion assumes
# the partner B builds only to its FLAT buffer. ON: the partner-dwell term uses B's OWN round-trip
# buffer (one level deep — B rotates back to A with A's dwell forced flat, no recursion), so A's
# cushion reflects B needing a full campaign. Widens eff_buf (opposes T2). OFF → flat, bit-for-bit.
_RT_PARTNER_RT = os.environ.get("RT_PARTNER_RT", "0") != "0"

# Secondary-SKU priority ordering for the live global-assign Phase-B (machine,SKU)
# pairing key (env BLD_SEC_ORDER, default "baseline" = committed key bit-for-bit).
# Same-size (inch_penalty) ALWAYS stays first; only the middle factors are permuted;
# tail is always (cost, m, sku) for a deterministic total order. Factors:
#   SAME  = inch_penalty (0 same-size CO, 1 diff-size)           — fixed first
#   URG   = 0 if _urgency_score>0 (demand uncoverable in horizon) else 1  (Class-A first)
#   STARV = 0 if a press draws this SKU now with GT < 1 shift of draw else 1 (about-to-starve first)
#   DEFC  = -_defc(sku, buf)  (larger this-shift GT deficit first)
# Variants: baseline | UD | USD | SUD | SD | DSU | INS_S (promote STARV into the
# committed tier/primary/constraint key without dropping the load-balancing terms).
_BLD_SEC_ORDER = os.environ.get("BLD_SEC_ORDER", "baseline")

# Idle-machine monthly-gap fill (env IDLE_GAP_FILL, default OFF). Phase-C forward-buffer
# already targets min(demand_remaining, draw × 9-shift shelf), but the starvation-risk gate
# only fires when a SKU is about to run dry THIS shift — so an idle eligible machine sits idle
# next to a SKU whose CUMULATIVE monthly demand is unmet but is kept just-fed by others (the
# 7502 case). This relaxes that gate for BUILDING-LIMITED SKUs only (_urgency_score==0 → curing
# can still cover the demand in the horizon → extra GT will be cured), so idle machines pre-build
# toward the shelf-capped monthly gap. Fully relaxing the gate for ALL SKUs was measured worse
# (IDLE_UNMET_KEEP_GATE=0: May −6,180/July −12,698 — GT-cap clog); the building-limited restriction
# is the discriminator. Target/shelf/7k-cap/inch-rules/CO-caps/ranking unchanged. OFF → bit-for-bit.
_IDLE_GAP_FILL = os.environ.get("IDLE_GAP_FILL", "0") != "0"

# Building-side RI-first-then-ratio: NRI (press_count<=0) candidates ranked by
# static demand[sku]/machine_total_demand[machine] instead of raw deficit; RI
# candidates (any live press) keep raw-deficit ranking unchanged and always
# sort ahead of NRI (tier 0 vs tier 1) — see _priority_tier.
_BUILDING_RATIO_ENABLED = True

# ── EXPERIMENT: RI-by-ratio (two approaches, both default OFF = bit-for-bit) ────
# User request: assign RI SKUs by the max-ratio formula instead of raw deficit.
# Approach 1 (_RI_RATIO_ENABLED): ranking-only. In _priority_tier, RI candidates
#   (live press) are ranked by static demand[sku]/machine_total_demand[machine]
#   (same formula as NRI) instead of raw curing-deficit. RI still outranks NRI
#   (tier 0 vs 1); only the WITHIN-RI order changes. Machine order + Campaign
#   loop unchanged. NOTE: drops the live deficit (starvation-urgency) signal for
#   RI, so this may raise starvation — that is what the experiment measures.
# Approach 2 (_RI_RATIO_GLOBAL): drop the fixed VMI->BJ->US->Stage2 machine order
#   and instead process machines in DESCENDING order of the best (highest) ratio
#   deficit-SKU each can serve, so the ratio — not the group order — drives which
#   machine-SKU combination is claimed first. Stage-1 stays last. Implies the
#   ratio ranking (auto-enables Approach 1's _priority_tier branch for RI).
_RI_RATIO_ENABLED = os.environ.get("RI_RATIO") == "1"
_RI_RATIO_GLOBAL  = os.environ.get("RI_RATIO_GLOBAL") == "1"

# EXPERIMENT: seed each building machine's starting SKU from the plant's ACTUAL
# running-machine snapshot (data/running_prod/building_running_machines_39_near7AM.xlsx)
# instead of the DB loader (which returns empty here -> machines start blank and
# the scheduler derives its own machine->SKU). Off = current behaviour (blank/DB).
_SEED_FROM_PLANT_RUNNING = True
# Start all building machines FREE (no Day-0 running SKU) → the very first shift seeds
# each machine as a "start" campaign (0 CO), so there are NO initial same/diff-size
# building COs; the scheduler's own logic decides the first SKU per machine. Skips
# both the DB running-machines load and the plant-file seed. Env BLD_START_FREE=1
# force-enables; default OFF keeps the seeded-from-running behaviour bit-for-bit.
_BLD_START_FREE = os.environ.get(
    "BLD_START_FREE",
    "1" if getattr(_bc_cfg, "BLD_START_FREE_ENABLED", False) else "0"
) != "0"
# Per-month building running-machine snapshot: env PLANT_RUNNING_FILE overrides (June/July use
# their own near-8AM snapshots built from the stage1+stage2 running data). Default = May.
_PLANT_RUNNING_FILE = os.environ.get(
    "PLANT_RUNNING_FILE",
    "data/running_prod/building_running_machines_july_near8AM.xlsx")

# ── Day-1 building seed from ACTUAL plant production (see run_rolling_pipeline seed block) ──
# Each building machine starts Day-1 on the SKU it was ACTUALLY building in the latest plant
# production; the seeded SKU's inch becomes the machine's DOMINANT/anchor inch. TAKES PRECEDENCE
# over BLD_START_FREE for seeded machines. Default ON (from bc_config); BLD_ACTUAL_SEED=0 reverts
# bit-for-bit. Only the full-month / Run-1 Day-1 is seeded (mid-month initial_state wins).
_BLD_ACTUAL_SEED = os.environ.get(
    "BLD_ACTUAL_SEED",
    "1" if getattr(_bc_cfg, "BLD_ACTUAL_SEED_ENABLED", False) else "0") != "0"
_BLD_ACTUAL_SEED_FILE = os.environ.get(
    "BLD_ACTUAL_SEED_FILE", getattr(_bc_cfg, "BLD_ACTUAL_SEED_FILE", ""))
# Populated per-run in run_rolling_pipeline: seeded machines + each machine's seed inch.
# Read by _assign_building_shift (dom override + Day-1 flex_reclaim guard). Empty ⇒ feature OFF.
_BLD_SEED_MACHINES: set = set()
_BLD_SEED_INCH: dict[str, str] = {}
_BLD_SEED_SKU: dict[str, str] = {}       # BLD_SEED_STICKY: machine -> its Day-1 seed SKU

# ── BLD_SEED_STICKY: give a seeded machine a PRIORITY (reserved) claim on its seed SKU ──────
# so it builds a CONSISTENT day-to-day amount of that SKU (like a captive machine) instead of
# absorbing the residual of a shared draw and swinging. In Phase A the sticky-seed machine is
# processed FIRST (before captive-max peers), so it claims its seed SKU's DRAW-BOUNDED deficit
# up front and the other eligible machines fill only the remainder. Draw-bounded (via _defc) →
# it never builds more than the curing draw can consume (no extra expiry) and never exceeds the
# demand cap. Only for the first _BLD_SEED_STICKY_DAYS days and only while the machine is still
# on its seed SKU. Default ON whenever BLD_ACTUAL_SEED is on; BLD_SEED_STICKY=0 reverts bit-for-bit.
_BLD_SEED_STICKY = _BLD_ACTUAL_SEED and os.environ.get(
    "BLD_SEED_STICKY",
    "1" if getattr(_bc_cfg, "BLD_SEED_STICKY_ENABLED", True) else "0") != "0"
_BLD_SEED_STICKY_DAYS = int(os.environ.get(
    "BLD_SEED_STICKY_DAYS", str(getattr(_bc_cfg, "BLD_SEED_STICKY_DAYS", 10))))
# BLD_SEED_PIN (default OFF): the AGGRESSIVE variant — re-anchor the sticky-seed machine onto its
# seed SKU EVERY shift (not just when it happens to be current), so it fills its whole shift with
# the seed SKU and its per-day output goes flat (like a captive machine). MEASURED: it flattens
# bj5/7102 (day1-10 stdev 225→141, cv 34.8%→21.2%) but costs ~2.7k cured on September because a
# flexible machine like 7102 is the SOLE builder for scarce 2-mould-pair SKUs (LSTL0-10, QSTL0) —
# pinning it off them, plus the global rebalance, loses more than it gains. Left OFF (violates
# cured >= baseline); flip on only when steady seed output is worth a KPI hit. Requires STICKY on.
_BLD_SEED_PIN = _BLD_SEED_STICKY and os.environ.get(
    "BLD_SEED_PIN",
    "1" if getattr(_bc_cfg, "BLD_SEED_PIN_ENABLED", False) else "0") != "0"

# ── BLD_SEED_PIN_D1A: HARD Day-1 Shift-A seed pin ──────────────────────────────
# On Day 1 Shift A ONLY, every seeded GT (non-Stage-1) machine builds EXACTLY its seed SKU
# (100% alignment with the plant Day-0 snapshot). Unlike _BLD_SEED_PIN (which only re-anchors
# a machine to its seed when SAME-INCH + curable-deficit>0 and applies to the whole sticky
# window), this FORCES cur = seed UNCONDITIONALLY on D1-A, builds it bounded ONLY by the demand
# cap + curable/GT cap (so a seed SKU with 0 live draw still gets a modest build — the plant was
# building it), and LOCKS the machine out of Phase B / global-assign / forward-buffer so nothing
# reassigns its D1-A production. Stage-1 (carcass) alignment is handled separately in _gate_build
# (best-effort: a Stage-1 machine can only align if its seed SKU's Stage-2 GT is built that shift).
# Default ON whenever BLD_ACTUAL_SEED is on; BLD_SEED_PIN_D1A=0 reverts bit-for-bit.
_BLD_SEED_PIN_D1A = _BLD_ACTUAL_SEED and os.environ.get(
    "BLD_SEED_PIN_D1A",
    "1" if getattr(_bc_cfg, "BLD_SEED_PIN_D1A_ENABLED", True) else "0") != "0"


# ── 2-DAY PLANT PLAYBACK (building) ───────────────────────────────────────────────
# For plan Days 1 and 2 (ALL shifts A/B/C), building EXACTLY replays the plant's actual
# day-0 snapshot (bc_config.PLANT_2DAY_SCHEDULE_FILE): the exact SKU(s) per (machine, shift)
# for all 41 machines, at qty = min(plant_qty, shift_capacity) — build the plant qty unless
# the shift physically cannot hold it (capacity < plant qty), then build the max that fits;
# NEVER overbuild past the plant qty, NEVER carry the shortfall into the next shift, demand
# cap IGNORED (follow the plant even beyond monthly demand). Curing runs normally (derives
# from the forced GT). From Day 3 the NORMAL pipeline takes over; state (GT inv / press /
# machine / inch / demand) carries forward in the single continuous run. Stage-1 machines
# emit CARCASS rows; GT machines add to gt_inventory. Default ON when the file exists;
# PLANT_2DAY_REPLAY=0 reverts bit-for-bit. Supersedes the Day-1 Shift-A seed pin on days 1-2.
_PLANT_2DAY_FILE = str(getattr(_bc_cfg, "PLANT_2DAY_SCHEDULE_FILE", "") or "")
# MID-MONTH GUARD: the schedule file's `Day` column is a PLAN-DAY index (1,2) describing the
# month's first two days. On a mid-month start, plan-day 1 is NOT the 1st, so replaying it would
# stamp the 1st-2nd's plant schedule onto the start date. Only replay for a 1st-of-month start.
_PLANT_2DAY_REPLAY = (
    os.environ.get("PLANT_2DAY_REPLAY", "1") != "0"
    and bool(_PLANT_2DAY_FILE) and os.path.exists(_PLANT_2DAY_FILE)
    and int(getattr(getattr(_bc_cfg, "PLAN_START", None), "day", 1) or 1) == 1)
_PLANT_2DAY_DAYS = 2                       # replay covers plan days 1..2 (all shifts)
# PLANT_SET_LOCK mid-month activation boundary. The gate is `day > _PLANT_SET_LOCK_FROM`, where
# `day` is the PLAN-day index. For a 1st-of-month start, plan days 1-2 ARE the plant replay, so
# the lock correctly starts at plan day 3 (= _PLANT_2DAY_DAYS) — unchanged, bit-for-bit. For a
# MID-MONTH start the plant days already physically happened (production deducted from demand,
# replay off), so every PLANNED day is "after" them and the lock must govern from plan day 1 —
# otherwise the first two planned days run unlocked and then snap onto the plant set, churning
# machines. Env PSL_MIDMONTH_FROM_DAY1=0 reverts to the verbatim bc_lp boundary (for A/B).
_PSL_FROM_DAY1 = os.environ.get("PSL_MIDMONTH_FROM_DAY1", "1") != "0"
_PLANT_SET_LOCK_FROM = (
    0 if (_PSL_FROM_DAY1
          and int(getattr(getattr(_bc_cfg, "PLAN_START", None), "day", 1) or 1) != 1)
    else _PLANT_2DAY_DAYS)
_PLANT_2DAY_BY_DS: dict = None             # {(day:int, shift:str): [(machine, sku, qty), ...]} — lazy


def _load_plant_2day_schedule() -> dict:
    """Load the plant 2-day building snapshot → {(day, shift): [(machine, sku, qty), ...]}.

    Row ORDER within a (day, shift) is preserved so multi-SKU machine campaigns replay in the
    plant's sequence (with a building CO emitted between them). Machine ids are kept as strings
    exactly as in the schedule (e.g. '7502', 'ps2'). Cached at module scope."""
    global _PLANT_2DAY_BY_DS
    if _PLANT_2DAY_BY_DS is not None:
        return _PLANT_2DAY_BY_DS
    out: dict = {}
    try:
        _df = pd.read_excel(_PLANT_2DAY_FILE)
    except Exception as _e:                                # noqa: BLE001
        print(f"  [Plant2Day] schedule load FAILED ({_e}) → replay INERT")
        _PLANT_2DAY_BY_DS = out
        return out
    for _, _r in _df.iterrows():
        try:
            _day = int(_r["Day"]); _sh = str(_r["Shift"]).strip().upper()
            _m = str(_r["Machine"]).strip(); _s = str(_r["SKUCode"]).strip()
            _q = float(_r["Qty"])
        except Exception:                                  # noqa: BLE001
            continue
        if not _m or not _s or _q <= 0:
            continue
        out.setdefault((_day, _sh), []).append((_m, _s, _q))
    _PLANT_2DAY_BY_DS = out
    _nrows = sum(len(v) for v in out.values())
    _nmach = len({m for v in out.values() for (m, _, _) in v})
    print(f"  [Plant2Day] loaded {_nrows} plant rows, {_nmach} machines, days 1-{_PLANT_2DAY_DAYS} "
          f"(replay {'ON' if _PLANT_2DAY_REPLAY else 'OFF'})")
    return out


_MACHINE_PLANT_SET: dict = None


def _get_machine_plant_set() -> dict:
    """PLANT_SET_LOCK: {GT-machine: set(SKUCodes it built in the plant Days-1-2 schedule)}.

    Stage-1 machines are EXCLUDED (their carcass auto-follows the Stage-2 SKUs). Returns an
    EMPTY dict when the lock is OFF or the plant schedule is absent → the assigner's plant-set
    gate is then identity (bit-for-bit OFF). Cached at module scope; iterated deterministically."""
    global _MACHINE_PLANT_SET
    if _MACHINE_PLANT_SET is not None:
        return _MACHINE_PLANT_SET
    out: dict = {}
    if _PLANT_SET_LOCK:
        for (_d, _sh), _rows in sorted(_load_plant_2day_schedule().items()):
            for (_m, _s, _q) in _rows:
                _mk = str(_m)
                if _mk in _S1_MACHINES:
                    continue                       # Stage-1 auto-follows Stage-2
                out.setdefault(_mk, set()).add(str(_s))
        print(f"  [PlantSetLock] ON — {len(out)} GT machines pinned to their Days-1-2 plant SKU set "
              f"(sizes: {sorted({len(v) for v in out.values()})})")
    _MACHINE_PLANT_SET = out
    return out


def _plant_2day_gt_plan(day: int, shift: str, machine_current_sku: dict,
                         machine_down_mins: dict | None = None) -> dict:
    """Build the forced GT-machine building plan for (day, shift) from the plant snapshot.

    Returns {machine: [(sku, qty_int, co_type)]} for NON-Stage-1 (GT) machines only —
    Stage-1 carcass is injected post-plan (see the carcass-injection block). qty =
    min(plant_qty, floor(SHIFT_MINS*60 / _bld_ct_sec(machine, sku))). co_type is 'start'
    when the SKU continues the machine's carryover SKU (no CO), else a CO marker so the
    emission loop charges + labels a building CO between in-shift SKU changes.

    PM/MTC takes priority over the plant seed (`machine_down_mins`, optional — the SAME
    `_mdown` dict the normal path builds via `_pm_maint_free`/`_best_maint_free` just above
    the replay call site, so this reuses the one existing maintenance-free calculation rather
    than recomputing it). For a machine with NO maintenance this shift the row is capped
    exactly as before (full-shift capacity, byte-identical). For a machine WITH maintenance
    this shift, `_mdown[machine]` already equals `SHIFT_MINS − (largest maintenance-free
    sub-interval)`, so `SHIFT_MINS − _mdown[machine]` reconstructs that free-interval length —
    the qty is capped to what fits inside it (`floor(free_mins*60 / _bld_ct_sec(machine,sku))`),
    shared cumulatively across the machine's rows this shift (in plant order) so a multi-SKU
    shift's TOTAL production also stays inside the free time. The emission loop already starts
    the row's wall-clock at that same free interval's start (`_mprod_start`, unchanged) — so
    capping the qty here is what stops the row from running past the interval INTO the window.
    Any plant qty beyond what fits is simply dropped (partial build, never carried to the next
    shift/day, never overbuilt) — never a byte-for-byte match here since maintenance won."""
    sched = _load_plant_2day_schedule()
    _mdown = machine_down_mins or {}
    _used_mins: dict = defaultdict(float)   # cumulative maintenance-budget production minutes charged this shift, per machine
    plan: dict = {}
    for (_m, _s, _q) in sched.get((day, shift), ()):
        if _m in _S1_MACHINES:                             # Stage-1 → carcass, handled separately
            continue
        _ct = _bld_ct_sec(_m, _s)
        _dn = float(_mdown.get(_m, 0.0))
        if _dn > 0:
            # Maintenance intrudes on this machine's shift — cap to the maintenance-free
            # budget (shared across this machine's rows this shift), never the full shift.
            _avail = max(0.0, float(SHIFT_MINS) - _dn)
            _left  = max(0.0, _avail - _used_mins[_m])
            _cap   = int(_left * 60.0 / _ct) if _ct > 0 else 0
        else:
            _cap = int(SHIFT_MINS * 60.0 / _ct) if _ct > 0 else 0
        _qi = int(min(round(_q), _cap))                    # min(plant qty, capacity); no overbuild
        if _qi <= 0:
            continue
        _prev = plan[_m][-1][0] if _m in plan else machine_current_sku.get(_m, "")
        _co = "start" if _s == _prev else "co"             # CO between different in-shift SKUs
        plan.setdefault(_m, []).append((_s, _qi, _co))
        if _dn > 0:
            _used_mins[_m] += _qi * _ct / 60.0
    return plan


def _load_actual_seed(path: str) -> dict[str, str]:
    """Read the Day-1 building seed file → {Machine(str): SKUCode(str)}.

    The file is a simple 2-column sheet (`Machine`, `SKUCode`) — one row per building
    machine and the SKU it is currently building; edit it to change a Day-1 assignment.
    Every row is a seed. (Back-compat: if the file still carries a legacy `Seed_Action`
    column, only its `SEED` rows are used and `FALLBACK` rows are skipped.)"""
    seed: dict[str, str] = {}
    try:
        _df = pd.read_excel(path)
    except Exception as _e:
        print(f"  [Rolling] BLD_ACTUAL_SEED: seed file load FAILED ({_e}) → no seed")
        return seed
    # accept flexible headers: Machine / "Machine ID"  and  SKUCode / "SKU Code" / SKU_Code
    _df = _df.rename(columns={"Machine ID": "Machine", "MachineID": "Machine",
                              "SKU Code": "SKUCode", "SKU_Code": "SKUCode"})
    _has_action = "Seed_Action" in _df.columns
    for _, r in _df.iterrows():
        if _has_action and str(r.get("Seed_Action", "")).strip().upper() != "SEED":
            continue
        _m   = str(r.get("Machine", "")).strip()
        _sku = str(r.get("SKUCode", "")).strip()
        if not _m or _m.lower() == "nan" or not _sku or _sku.lower() == "nan":
            continue
        seed[_m] = _sku
    return seed


_MIDMONTH_SET: dict = {}          # {machine: set(SKUs it built in the N days before PLAN_START)}
_MIDMONTH_LAST: dict = {}         # {machine: the LAST SKU it was building before PLAN_START}
# DEDICATED MACHINES (bc_config.MACHINE_DEDICATED_SKU): {machine: the ONLY SKU it may build}.
_DEDICATED: dict = {str(k): str(v) for k, v in
                    (getattr(_bc_cfg, "MACHINE_DEDICATED_SKU", {}) or {}).items()}
_DEDICATED_SKUS: set = set(_DEDICATED.values())
_DEDICATED_FIRST = bool(getattr(_bc_cfg, "DEDICATED_SKU_FIRST", True)) and bool(_DEDICATED)
# EARLY FULL-LOAD (bc_config.EARLY_FULL_LOAD_*): SKUs certified on these machines get first call
# during plan days 1..N so the machines run flat out early. Resolved lazily (needs the CT map).
_EARLY_MACHINES: list = [str(x) for x in (getattr(_bc_cfg, "EARLY_FULL_LOAD_MACHINES", []) or [])]
_EARLY_DAYS: int = int(getattr(_bc_cfg, "EARLY_FULL_LOAD_DAYS", 0) or 0)
_EARLY_SKUS: set = set()


def _derive_midmonth_sets(plan_start, path: str = None, days: int = None) -> dict:
    """MID-MONTH CARRY-IN as a rolling N-day SKU SET per building machine.

    Reads a FULL-MONTH baseline plan (built with NO production deduction) and returns
    {machine: set(SKUCode)} for the `days` calendar days immediately BEFORE plan_start.
    The machine then starts on whichever SKU in its own set has the best live need —
    it demonstrably ran all of them recently, so no changeover is charged.

    Returns {} on any failure so the caller falls back to the previous behaviour.
    """
    # SOURCE = the plant's own Days-1-2 schedule (real plant data) when configured.
    if str(getattr(_bc_cfg, "MIDMONTH_SET_SOURCE", "baseline")).lower() == "plant_2day":
        _pp = getattr(_bc_cfg, "PLANT_2DAY_SCHEDULE_FILE", "") or ""
        if _pp and os.path.exists(_pp):
            try:
                _pd_ = pd.read_excel(_pp)
                _pd_["_M"] = _pd_["Machine"].astype(str).str.strip()
                _pd_["_S"] = _pd_["SKUCode"].astype(str).str.strip()
                _out = {m: {x for x in g["_S"] if len(x) == 18}
                        for m, g in _pd_.groupby("_M")}
                _out = {m: v for m, v in _out.items() if v}
                _MIDMONTH_LAST.clear()
                for _m, _g in _pd_.groupby("_M"):          # last plant row = its latest SKU
                    _r = _g.sort_values(["Day", "Shift"], kind="stable")
                    if len(_r):
                        _MIDMONTH_LAST[str(_m)] = str(_r.iloc[-1]["_S"])
                for _om, _os in (getattr(_bc_cfg, "MIDMONTH_SET_OVERRIDE", {}) or {}).items():
                    _out[str(_om)] = {str(x) for x in _os}
                    _MIDMONTH_LAST[str(_om)] = str(list(_os)[0])
                    print(f"  [midmonth-override] {_om} -> {sorted(_out[str(_om)])}")
                return _out
            except Exception as _e:
                print(f"  [midmonth-set] plant_2day read failed ({_e}); falling back")
    path = path or getattr(_bc_cfg, "MIDMONTH_BASELINE_PLAN", "") or ""
    days = int(days or getattr(_bc_cfg, "MIDMONTH_SET_DAYS", 3) or 3)
    if not path or not os.path.exists(path):
        return {}
    try:
        d = pd.read_excel(path, sheet_name="Shift Schedule", header=2)
        d.columns = [str(c).strip() for c in d.columns]
        _mode = str(getattr(_bc_cfg, "MIDMONTH_SET_MODE", "window") or "window").lower()
        if _mode == "start_day":
            # the baseline plan's rows ON the mid-month start date itself — the exact machine
            # state the 26-day plan begins from, including every SKU a machine ran that day.
            win = {plan_start.strftime("%Y-%m-%d")}
        else:
            win = {(plan_start - timedelta(days=k)).strftime("%Y-%m-%d") for k in range(1, days + 1)}
        d["_D"] = d["Date"].astype(str).str[:10]
        d["_S"] = d["SKUCode"].astype(str).str.strip()
        d["_M"] = d["Machine"].astype(str).str.strip()
        d["_C"] = d["CO_Type"].astype(str).str.upper()
        d["_Q"] = pd.to_numeric(d["Qty"], errors="coerce").fillna(0.0)
        _sent = {"CHANGEOVER", "MOULD_CLEAN", "PM", "MTC", "NAN", "—", "",
                 "EXPIRED_GT", "EXPIRED_CARCASS"}
        d = d[(d["_D"].isin(win)) & (d["_Q"] > 0)
              & (~d["_S"].str.upper().isin(_sent))
              & (~d["_C"].isin(["PM", "MTC", "EXPIRED_GT", "EXPIRED_CARCASS"]))]
        d = d.reset_index()
        d["_so"] = d["Shift"].map({"A": 0, "B": 1, "C": 2}).fillna(0)
        out: dict = {}
        _MIDMONTH_LAST.clear()
        for _m, _g in d.groupby("_M"):
            _sk = {x for x in _g["_S"] if len(x) == 18}
            if not _sk:
                continue
            out[str(_m)] = _sk
            # the SKU physically on the machine at PLAN_START 07:00
            _MIDMONTH_LAST[str(_m)] = str(
                _g.sort_values(["_D", "_so", "index"], kind="stable").iloc[-1]["_S"])
        return out
    except Exception as exc:                       # never block a run on the baseline file
        print(f"  [midmonth-set] could not read {os.path.basename(path)}: {exc}")
        return {}


def _derive_seed_from_plant_2day(path: str = None) -> dict[str, str]:
    """MID-MONTH carry-in: derive {Machine: SKUCode} = each building machine's SKU at the
    END of the plant's last replayed day, from PLANT_2DAY_SCHEDULE_FILE.

    Why this exists: on a mid-month start the plant has already run days 1..K and the plan
    begins at day K+1, which must CONTINUE what each machine was building — not start free
    and not replay days 1..K (their production is already deducted from demand). The
    static day-1 seed file is the WRONG state for that (it is the state before those days
    ran). This derives the correct carry-in straight from the plant schedule, so no
    hand-made file is needed and it stays correct when the plant file is refreshed.

    Per machine: take its LAST row on the MAX `Day` present (shift order A<B<C, and the
    last row within that shift), i.e. the SKU it ends the plant window on.
    Returns {} if the file is absent/unreadable (caller falls back to the seed file).
    """
    path = path or _PLANT_2DAY_FILE
    if not path or not os.path.exists(path):
        return {}
    try:
        _d = pd.read_excel(path)
    except Exception as _e:
        print(f"  [Rolling] PLANT_2DAY carry-in: load FAILED ({_e}) → falling back to seed file")
        return {}
    _cols = {str(c).strip().lower(): c for c in _d.columns}
    _mc, _sc = _cols.get("machine"), _cols.get("skucode") or _cols.get("sku code")
    _dc, _sh = _cols.get("day"), _cols.get("shift")
    if not all((_mc, _sc, _dc, _sh)):
        print(f"  [Rolling] PLANT_2DAY carry-in: unexpected columns {list(_d.columns)} → skipped")
        return {}
    _d = _d.copy()
    _d["_ord"] = _d[_sh].astype(str).str.strip().str.upper().map({"A": 0, "B": 1, "C": 2}).fillna(0)
    _last_day = pd.to_numeric(_d[_dc], errors="coerce").max()
    _d = _d[pd.to_numeric(_d[_dc], errors="coerce") == _last_day]
    _d = _d.reset_index().sort_values(["_ord", "index"])          # stable: shift, then row order
    out: dict[str, str] = {}
    for _, r in _d.iterrows():                                     # later rows overwrite → last wins
        _m, _s = str(r[_mc]).strip(), str(r[_sc]).strip()
        if _m and _m.lower() != "nan" and _s and _s.lower() != "nan":
            out[_m] = _s
    print(f"  [Rolling] PLANT_2DAY carry-in: derived {len(out)} machine→SKU from end of plant "
          f"day {int(_last_day)} ({os.path.basename(path)})")
    return out

# EXPERIMENT: captive-first ordering. A captive building machine (eligible for
# exactly ONE SKU, e.g. 7301 -> only LSTL0) sits idle whenever flexible machines
# processed earlier in the VMI->BJ->US->Stage2 order drain that SKU's deficit
# first. Promoting captives to the FRONT of the machine loop lets them claim
# their sole SKU before flexible machines, maxing captive utilisation and freeing
# the flexible (oversubscribed) machines for SKUs only they can serve. Stage-1
# captives are NOT promoted (carcass, always last). Off = current behaviour.
# _CAPTIVE_FIRST_ENABLED = os.environ.get("CAPTIVE_FIRST") == "1"
_CAPTIVE_FIRST_ENABLED = True

# EXPERIMENT: global machine-SKU scoring assignment. Replaces the sequential
# VMI->BJ->US->Stage2 per-machine greedy with ONE global pass: after each machine
# continues its current SKU (Phase A, no CO), all remaining (machine,SKU) pairs are
# scored together and assigned best-first (Phase B). The score includes
# constraint = min(flex_machine, flex_sku), so a captive machine (e.g. 7301) OR a
# sole-supplier SKU wins WITHOUT any hardcoded rule — symmetric on both sides,
# unlike the machine-only _SCARCITY_ORDER that regressed. When on, it early-returns
# before the captive-first re-sort (so it fully supersedes it). Off = per-machine greedy.
_GLOBAL_ASSIGN_ENABLED  = True
# constraint placement in the pair sort key: "above" (constraint before SKU tier),
# "below" (after tier), or "captive" (only min(flex)<=1 boosted). Sensitivity knob.
_GLOBAL_CONSTRAINT_MODE = "below"

# EXPERIMENT (global branch only): force a CAPTIVE machine (exactly 1 eligible SKU,
# e.g. 7301 -> LSTL0) to build its sole SKU at FULL shift capacity in Phase A —
# capped ONLY by the demand cap, not by the curing-demand buffer — so it never
# idles while its SKU still has unmet demand. Remaining demand is filled by the
# normal global logic on other machines. Off = buffer-throttled (current global).
# Flip this line True/False to turn captive-max on/off (env CAPTIVE_MAX also works):
# _CAPTIVE_MAX_ENABLED = os.environ.get("CAPTIVE_MAX") == "1"
_CAPTIVE_MAX_ENABLED = True

# BLD_CURABLE_CAP (default ON) — no-waste-GT fix (invariant #4). A building machine
# must not build a SKU faster than the SKU's curing side can CONSUME it, or the excess
# GT ages out (3-day shelf) as expired waste. This bites hardest on single-source / PS
# machines that are captive to ONE SKU with only 1-2 eligible curing presses: captive-max
# builds toward the full DEMAND cap, ignoring that (say) 1 press can only drain a fraction
# of that GT before it ages out — the surplus expires (e.g. ps2 → TUXPE: built 3,417,
# cured 1,548, ~1,900 expired). The fix bounds the captive-max / sticky build so the SKU's
# GT-ON-HAND can never exceed its curable STOCK cap = (#eligible curing presses) ×
# cure_rate/shift × GT_SHELF_LIFE_SHIFTS — the most its presses can drain over the 3-day
# shelf. It never UNDER-builds a genuinely curable SKU: a machine feeding many live presses
# has a large cap and rebuilds as curing drains, so the 3-day forward buffer is preserved;
# only GT that WOULD have expired is withheld. BLD_CURABLE_CAP=0 reverts bit-for-bit.
_BLD_CURABLE_CAP = os.environ.get(
    "BLD_CURABLE_CAP",
    "1" if getattr(_bc_cfg, "BLD_CURABLE_CAP", True) else "0") != "0"

# ps2-ONLY batch exception (BLD_BATCH). The curable cap correctly bounds a single-press SKU's
# GT-on-hand (no expiry), but it then dribbles a tiny top-up EVERY shift (e.g. 86/shift = ~14%
# of a 480-min shift) → an under-utilized Shift Schedule. For the listed exception machines
# (ps2 only), CONSOLIDATE: skip a shift unless a chunk of >= _BLD_BATCH_MIN units can be built,
# so the machine builds fuller shifts less often (the press drains between). The curable cap
# still bounds the batch → NO expiry returns; total build/cured is preserved. Applies ONLY to
# the exception machines; every other machine is unchanged. BLD_BATCH=0 disables.
_BLD_BATCH_ENABLED = os.environ.get("BLD_BATCH", "0") != "0"
_BLD_BATCH_MACHINES = set(str(x) for x in getattr(_bc_cfg, "BLD_BATCH_MACHINES", ["ps2"]))
_BLD_BATCH_MIN = int(os.environ.get("BLD_BATCH_MIN", str(getattr(_bc_cfg, "BLD_BATCH_MIN", 300))))

# EXPERIMENT: End-of-day 12k total GT-inventory cap (hard plant constraint).
# Plant can hold at most MAX_ENDOFDAY_GT_INVENTORY units of GT overnight (summed
# over all SKUs). Enforced PROACTIVELY in the build-room calc (never build past the
# ceiling) — see _defc / captive-max _room. An end-of-day audit column records the
# actual total so we can confirm it never violates. Off = no cap (current behaviour).
# env GT_CAP=0 forces OFF for the bit-for-bit baseline check; default ON.
_ENDOFDAY_GT_CAP_ENABLED = True

# EXPERIMENT: Forward-buffer level-loading. Use idle building capacity (weeks 4-5 run
# at 51-56% util while presses starve) to pre-build a SHELF-LIFE-SAFE forward buffer
# for SKUs that WILL be cured in the next 3 days. Phase C (slack-fill) runs after
# Phase A/B: for a machine that would otherwise idle, build MORE of its most
# starvation-prone eligible SKU (must have a LIVE cure-draw shift_cure_demand>0 — a
# press is actively pulling it — and demand remaining), up to min(demand_remaining,
# 3-day cure-draw) and bounded by the 12k cap. Builds only REQUIRED GT, never random
# SKUs. Auto-targets building-limited SKUs, auto-skips press-limited ones. Off = idle.
# env FWD_BUF=0 forces OFF; default ON.
_FORWARD_BUFFER_ENABLED = True

# Starvation-risk gate for the forward buffer: only pre-build a SKU whose on-hand GT
# is BELOW this many shifts of its live cure-draw (i.e. it is about to starve). This
# stops early-month front-loading (weeks 1-3 machines have slack but presses are NOT
# starving — building keeps up), so the forward buffer fires mainly where presses
# actually run dry (weeks 4-5), lifting the tail without pulling demand forward.
# Higher = fires more aggressively (more front-load); lower = only near-starvation.
# 0 ⇒ gate off (fill whenever slack+demand exist, the pure front-loading behaviour).
_FWD_RISK_SHIFTS = True

# Shelf-life expressed in shifts (3 days x 3 shifts/day = 9). Forward-buffer never
# pre-builds a SKU beyond this many shifts of its live cure-draw (so it cannot age
# out to writeoff within the shelf window).
GT_SHELF_LIFE_SHIFTS = GT_SHELF_LIFE_DAYS * 3

# Build-to-draw PACING = SMALL-BUFFER ROTATION (env PACING, default OFF). Models the plant's
# real building method: a machine runs a SKU only until it has built a SMALL forward buffer
# (PACING_BUFFER_SHIFTS = 1-2 shifts of that SKU's draw, NOT the full 9-shift GT_SHELF_LIFE
# bank), then ROTATES (CO's) to the next in-demand SKU curing is drawing and builds a small
# buffer there too — spreading building across MANY distinct SKUs/day (plant ~48-50 vs our
# ~40) instead of concentrating on a few + front-loading. Two coupled throttles, ALL gated by
# _PACING_ENABLED and applied ONLY from day 3+ (days 1-2 are the plant replay, untouched):
#   (a) per-SKU forward buffer shrunk 9 -> PACING_BUFFER_SHIFTS (thin buffers don't age out);
#   (b) a FLAT daily-build target caps the day's TOTAL forward-added GT so daily GT built
#       tracks a flat pace (no banking once the day has hit pace) — a SKU about to starve
#       this shift (on-hand < its draw) still gets fed (no new starvation);
#   (c) Phase-C ranking WIDENS the active SKU set (prefer a SKU not yet built today, then
#       nearest-to-starve) instead of IUkeep's biggest-gap concentration.
# Target modes (PACING_TARGET_MODE): "adaptive" (default) = min(remaining_curable_demand /
# remaining_working_days, that day's curing capacity ~ draw/shift x 3) recomputed daily; or
# "fixed" = PACING_DAILY_TARGET (~22000) for A/B. Trades the forward-buffer's early-cure
# accelerator for flatness + wider SKU spread + less expiry. OFF => forward-buffer unchanged
# (9-shift bank, IUkeep ranking, no daily cap), bit-for-bit.
_PACING_ENABLED       = os.environ.get("PACING", "0") != "0"
# per-SKU thin buffer (shifts of draw); PACING_SHIFTS kept as a back-compat alias.
PACING_BUFFER_SHIFTS  = int(os.environ.get("PACING_BUFFER_SHIFTS",
                                           os.environ.get("PACING_SHIFTS", "2")))
PACING_SHIFTS         = PACING_BUFFER_SHIFTS
PACING_TARGET_MODE    = os.environ.get("PACING_TARGET_MODE", "adaptive").lower()
PACING_DAILY_TARGET   = float(os.environ.get("PACING_DAILY_TARGET", "22000"))
# Plant builds only ~2k/day AHEAD of curing (the thin overnight buffer it maintains, NOT the
# 5-10k our forward-buffer banks). The adaptive daily cap = curing capacity (draw/shift x 3)
# + this thin headroom, so building tracks cure + ~2k. env PACING_OVERBUILD.
PACING_OVERBUILD      = float(os.environ.get("PACING_OVERBUILD", "2000"))

# IDLE-MACHINE → HIGHEST-UNMET-DEMAND targeting (ADOPTED, default ON). The forward
# buffer above (Phase C) fills idle building capacity toward the NEAREST-TO-STARVE SKU.
# This lever instead points idle machines at the BIGGEST unmet-demand gaps: it re-ranks
# the Phase-C candidates by remaining demand (largest gap first). It stays on a CURABLE
# PATH — the draw>0 gate is kept, so a candidate is only ever a SKU a press is actively
# drawing OR CO'ing to today (the Shift-A pre-build injection makes today's CO targets
# draw>0), and it is bounded by the shelf-safe target, the 7k end-of-day GT cap, and the
# hard demand cap → no waste GT (invariant #4 holds).
# MEASURED (cap=12, mould-audit PASS, deterministic): May 684,910 (+2,650), June 632,168
# (-1,870), July 694,161 (+12,732) → net +13,512, biggest gain on the weak month (July
# 87.5%→89.1%). Also LOWERS writeoff (1,864→1,845) and starvation (1,203→1,063).
# Pinned ON (like the other adopted toggles). Set this to False to reproduce the
# pre-adoption forward buffer bit-for-bit (verified May 682,260 / June 634,038 / July 681,429).
_IDLE_UNMET_ENABLED = True
# Sub-mode (default ON = the shipped variant): KEEP the starvation-risk throttle (only
# near-starving SKUs are pre-built), just re-ranked by biggest-gap first. Isolates "aim
# at the biggest gaps" from "build ahead of need". IDLE_UNMET_KEEP_GATE=0 relaxes the
# throttle too (pure front-loading) — MEASURED WORSE (May -6,180, July -12,698), left OFF.
_IDLE_UNMET_KEEP_GATE = True

# BUSINESS RULE: curing-press mould clean. After every MOULD_CLEAN_CYCLES cycles
# (= 6,000 tyres) a press takes an 8h (MOULD_CLEAN_MINS = 480 = 1 shift) mould clean
# during which it produces nothing; mould life then resets. A curing CO also resets
# mould life (the CO already includes a clean). env MOULD_CLEAN=0 disables → the
# pre-mould-clean 690,180 baseline reproduces bit-for-bit.
_MOULD_CLEAN_ENABLED = True

# Mould life v2: seed each press's OPENING mould life from the real DB remaining life
# (min over its 2 moulds, so both clean together) instead of a flat 3,000. env
# MOULD_LIFE_DB=0 → v1 (everyone opens fresh at 3,000) bit-for-bit. Only the FIRST
# clean's timing changes; a clean/CO still resets to 3,000 (per-press, unchanged).
_MOULD_LIFE_FROM_DB = True

# ── MOULD→SKU AVAILABILITY GATE (client hard rule) ────────────────────────────
# A press can only run/CO to an SKU if it physically has 2 eligible moulds mounted
# (Master_Mapping_Mould_SKU). Inventory = 1,284 physical moulds, one copy each; a
# mould serves one press at a time (contention). Mounting a spare is free; mould
# movement rides the existing 480-min press CO (no extra time). Mould life stays
# per-press 3,000 for v1 (real per-mould life is v2).
# Default ON — this makes the plan physically real (a moderate KPI drop is the new
# baseline). MOULD_GATE=0 reproduces the current mould-blind engine bit-for-bit.
_MOULD_GATE_ENABLED = True

# Phase 2 mould optimisation (raise the Phase-1 contention baseline). Only meaningful
# when the gate is ON. Two levers, both toggle-gated by MOULD_OPT:
#   (a) scarce-first ordering — when several presses CO the same day, allocate moulds
#       to the SCARCEST new-SKU first (fewest eligible moulds), so a 2-mould SKU is
#       not blocked by a 6-mould SKU grabbing a shared mould first.
#   (b) retarget-on-block — a planned CO whose mould claim fails does NOT just idle on
#       its (usually demand-done) old SKU; it retargets to the most-needy eligible SKU
#       that still HAS 2 free moulds, recovering the wasted CO slot.
# MOULD_OPT=0 → pure Phase-1 gate (scheduling identical to the locked mould baseline).
_MOULD_OPT_ENABLED = True

# Phase 3 — Unified CO scorer. Replaces the three ad-hoc changeover paths (static
# planned COs, per-press retarget-on-block, mid-shift dynamic CO) with ONE scoring
# function that ranks {execute planned, pull-forward tomorrow's planned, dynamic,
# retarget, idle} per press and solves them as a GLOBAL greedy over the shared
# resources (mould pool + a quantitative building-feed estimate + daily CO cap).
# Default ON — measured ≥ Phase-2 on all 3 months (May +3,780, June +2,261, July +2,987;
# all exact-mould-audit PASS, deterministic, demand cap holds). CO_SCORER=0 → Phase-2 path.
_CO_SCORER_ENABLED = True
# Sub-flag (measured rollout): False = ADDITIVE (planned COs always kept; scorer only
# fills idle presses + pulls forward tomorrow's planned COs) — the shipped mode.
# True = FULL RE-OPT (planned COs may be cancelled/replaced) — MEASURED WORSE (utility
# picker churns and drops good planned COs, e.g. May −40k), left OFF.
_SCORER_FULL_REOPT = os.environ.get("SCORER_FULL", "0") != "0"
# ── P1: reactive building-supply CO gate — REJECTED (env REACTIVE_CO, default OFF) ──
# When ON, every reactive curing CO must pass a HARD building-supply test (_supply_ok):
# the target SKU must have curable GT already banked (≥1 shift of draw) OR a building
# machine that can be reserved to feed it this shift (same-inch/flex/Stage-1, via
# _bld_capacity). Intended to stop a press changing over toward a SKU whose GT never
# arrives. A press whose demand is done but whose only reachable targets all fail the
# supply test does NOT idle — a FORCED CO fires on the plain best target (bypasses ONLY
# the supply test; the mould + allowable gates ALWAYS still apply), counted in
# co_scorer_stats["forced"].
#
# MEASURED + REJECTED (2026-08, cap=12, deterministic 2-hashseed, feasibility no NEW
# violations, OFF bit-for-bit): net −904 cured over 3 months — July +924 / June 0 /
# Aug −1,828. forced=0 and build_blocked≈0 on ALL months → the target scenario (a CO
# toward absent GT) essentially never occurs; the existing soft build-feed veto
# (_co_utility min-with-feed + the _bld_capacity check in _commit/_best_alt) already
# prevents it. The hard gate therefore only REORDERS targets (best supply-feasible over
# best-overall): helps the building-limited month (July) but churns the well-supplied
# month more (Aug). Same shape as the rejected Lever B. Kept OFF, code retained for the
# record. OFF (default) = bit-for-bit identical to today.
_REACTIVE_CO = os.environ.get("REACTIVE_CO", "0") != "0"
# ── Same-press-return guard (cosmetic monotonicity fix, env PRESS_RETURN_BLOCK, default OFF) ──
# The COScheduler plan is monotone-clean, but the planned-CO RETARGET-on-mould-block (_solve_day_cos
# additive branch → _pick_retarget) can re-choose a blocked planned CO's target to a SKU the press
# JUST LEFT (it still holds that SKU's moulds), rendering a 1-day excursion + return in the output
# ("Planned" CO row). This guard makes the retarget PREFER a target the press has not left; it falls
# back to the unguarded pick when no alternative exists (never strands → bit-parity in that case).
# Target-side only: never touches the mould gate, CO-cap, n-1 donor guard, or demand cap. Exempts
# delivery-priority SKUs (deadline > KPI). Same env as the (inert) Phase-0 guard so ONE flag drives both.
# ADOPTED (default ON): removes the same-press-return display artifact (June 5→0 / July 7→1 / Aug 13→2,
# net cured +143 across 3 months, feasibility-clean). PRESS_RETURN_BLOCK=0 reverts bit-for-bit.
_PRESS_RETURN_BLOCK = os.environ.get("PRESS_RETURN_BLOCK", "1") != "0"

# Phase 4 — GLOBAL MOULD OPTIMISER (experiment, default OFF). The gate + scorer above
# allocate moulds per-press greedily (each CO grabs the first 2 free/own eligible
# moulds). That leaves scarce moulds (15"/13" tooling) "stuck" on presses that no longer
# need them, so eligible presses can't serve the biggest-gap SKUs. This step runs once
# per day (after the day's COs are fixed, before they drive the sim) and, ranked by the
# most-under-served scarce SKU first:
#   (1) DIRECT ADD — put a sacrificeable eligible press onto an under-served scarce SKU
#       when it can already mount 2 moulds; and
#   (2) LIBERATION — proactively CO a sacrificeable HOLDER of that SKU's mould to a needy
#       retarget SKU chosen to RELEASE the scarce mould, so a future day can mount it.
# Both respect the daily CO cap. Two aggressiveness MODES (env MOULD_OPT_MODE):
#   "ro_only"    — only sacrifice / evict presses whose current SKU demand is DONE
#                  (Runner-Out). Those COs are near-free. Lower risk.
#   "full_evict" — additionally evict a RUNNING press when the target SKU's amortised
#                  marginal value strictly exceeds the current SKU's. Bigger lever,
#                  more churn, more CO-cap hungry.
# MOULD_GLOBAL_OPT=0 (default) reproduces the current engine bit-for-bit.
_MOULD_GLOBAL_OPT_ENABLED = os.environ.get("MOULD_GLOBAL_OPT", "0") != "0"
_MOULD_GLOBAL_OPT_MODE    = os.environ.get("MOULD_OPT_MODE", "ro_only").strip().lower()

# BUSINESS RULE: spread planned curing COs across shifts A/B/C.
# Planned COs were hardcoded to Shift A (all 147), so 97% of changeover downtime
# landed in Shift A and its curing output sat ~6.6k below Shift B — an artifact, not
# plant physics. Real plants change over across all three shifts, firing a CO as soon
# as a press finishes its current SKU (subject to the daily cap). With this ON, each
# planned CO is placed in the shift where its press is projected to exhaust its old
# SKU — Shift A if it is ALREADY finished, so a free press never waits — falling back
# to Shift A when the SKU will not finish today (the static scheduler booked that CO
# for a reason: a preemptive Class-A move). Flip to False to disable (all COs → Shift A).
_CO_SHIFT_SPREAD_ENABLED = True

# Reversed machine processing order: Stage2 -> Unistage -> BJ -> VMI instead of
# today's VMI -> BJ -> Unistage -> Stage2. Tests whether scarce/inch-locked
# groups should claim their deficit signal before flexible VMI machines mop up
# residual demand last.
_MACHINE_ORDER_REVERSED = False

# Scarcity-first machine ordering: process machines with the FEWEST eligible
# SKUs first, so captive/specialized machines (e.g. 7301, eligible only for
# LSTL0) claim their specific SKU's deficit before flexible machines that
# share that SKU but have many other options poach it. Complements the ratio
# formula (which ranks WHICH SKU a machine picks); this fixes WHICH MACHINE
# claims a shared SKU first. Fixes the 7301-at-7%-util problem generally.
_SCARCITY_ORDER_ENABLED = False

# Starvation-feed: when a machine would otherwise IDLE and an eligible SKU has
# starving presses (curing demand this shift but zero GT), let the machine take
# the CO to feed it as a LAST RESORT. Inch preference is FULLY preserved — the
# starving SKU stays low priority (same/diff bucketing + inch_penalty unchanged),
# so a machine only feeds an off-inch starving SKU when it has no same-inch work
# left. Only the 30%-CO-cost guard is relaxed for starving SKUs (generalizing
# the existing curing-CO-target bypass to ANY starving SKU); MIN_CAMPAIGN
# feasibility and the demand cap stay enforced, so no overbuild, no forced
# infeasible CO. General case — all machines, all SKUs.
_STARVATION_FEED_ENABLED = False

# Reactive "instant CO" mechanism (dynamic_co_tracker): when a curing press
# finishes its SKU's demand mid-month, it is immediately changed over to a new
# needed SKU. Default True = current behaviour ("dynamic consumption": upfront
# pre-planned Phase-0 schedule PLUS reactive COs). Set False for a "static
# consumption" run where curing follows ONLY the pre-planned schedule.
_DYNAMIC_CO_TRACKER_ENABLED = True

# Dynamic per-day curing CO planner: replaces the static, upfront 31-day CO
# schedule (COScheduler's proxy-simulated demand drain) with a fresh decision
# made once per day, before Shift A, using REAL live state (demand_remaining,
# press_state, gt_inventory) — same real-state discipline as the existing
# reactive dynamic_co_tracker mechanism, extended to the whole press fleet.
# COScheduler/co_by_day stays completely untouched when this is off.
# ABANDONED — a purely reactive planner (no lookahead) regressed severely
# (670k->~594k GT, 96.7%->~86% coverage) across three attempts; root cause:
# it can only redistribute capacity from presses already idle today, never
# proactively reassign a press ahead of a deadline the way a whole-horizon
# simulation can. Code kept in place, toggle permanently off. Superseded by
# _ROLLING_HORIZON_CO_ENABLED below.
_DYNAMIC_CO_PLANNER_ENABLED = (os.environ.get("DYNCO", "0") == "1")  # env-gated for A/B; default OFF = bit-for-bit

# Nested sub-toggle (only meaningful when _DYNAMIC_CO_PLANNER_ENABLED=True):
# adds sku_campaign_tier (building's primary/secondary/tertiary campaign
# position for each SKU) as a tiebreak in CO target ranking — a SKU with
# committed primary-campaign building capacity is a safer CO target than one
# only produced as an opportunistic secondary/tertiary blip. Day-granularity,
# tiebreak-only role (never a primary gating signal) — avoids the live
# per-shift-signal thrashing failure mode seen earlier this session.
_CAMPAIGN_TIER_TIEBREAK_ENABLED = False

# Rolling-Horizon COScheduler: reuses the existing, proven COScheduler.schedule()
# (the artifact behind the 670,431/670,649/96.7% baseline) as a receding-horizon
# planner — called once per simulated day, seeded from live rolling-pipeline
# state (press_state/press_count/demand_remaining), with a SHRINKING remaining
# horizon (Day 1 plans 31 days ahead, Day 10 plans 22, ...). Only that call's
# relative-day-1 events are kept; the rest of its lookahead is discarded and
# recomputed fresh tomorrow. The scheduling ALGORITHM is not rewritten — only
# its inputs change per call. See plan file for the full design rationale.
_ROLLING_HORIZON_CO_ENABLED = False

# Ratio-based curing CO allocation: no proactive CO at all — a press keeps
# producing its current SKU until that SKU's demand is FULLY exhausted (no
# urgency/Class A/B/horizon-threshold early CO). When a press frees up, pick
# the highest-static-ratio compatible SKU that still needs more press
# capacity to finish within the remaining horizon (a "presses needed" tracker
# skips SKUs that already have enough). Building scheduler untouched. Reuses
# the existing reactive trigger site (press's demand hits 0, mid-shift) —
# only the SELECTION function changes; co_by_day is emptied so nothing
# proactive from the static schedule fires under this toggle.
_RATIO_CO_ALLOCATION_ENABLED = False

# Nested sub-toggle (only meaningful when _RATIO_CO_ALLOCATION_ENABLED=True):
# enriches _select_ratio_co_target's ranking with COScheduler's other four
# levels (Class A/B, after_days, cycle time, target-side scarcity) on top of
# ratio, testing whether a richer ranking recovers ground among whatever
# candidates are ALREADY eligible right now. Does NOT touch eligibility
# timing — a press still only frees up once its current SKU is fully done,
# exactly as specified; this only changes which SKU wins once it's free.
_RATIO_CO_RICH_RANKING_ENABLED = False

# Early-CO (dynamic consumption via surplus-press reassignment): the one lever
# every prior dynamic attempt left unpulled. All of them waited for a press's
# demand to hit ZERO before reassigning, so COs clustered late and regressed.
# The static approach wins because its proxy sim frees presses EARLY. This adds
# that: a press may CO away from its current SKU BEFORE its demand is complete,
# IFF that SKU is already on track to finish with its REMAINING presses (n-1)
# within the horizon — i.e. this press is surplus to it. The next SKU is chosen
# by the same max-ratio formula as building (via _select_ratio_co_target), and
# only fires if a genuinely under-served target exists. Builds ON TOP of the
# current hybrid (static co_by_day base + reactive layer) — it can only ADD COs,
# so it has a real shot at beating baseline rather than replacing what works.
_EARLY_CO_ENABLED = False

# ── Part B: pure-reactive curing COs (env REACTIVE_ONLY, default OFF = current hybrid) ──
# Master toggle for Part B. When ON: the whole-horizon planned schedule (COScheduler +
# co_by_day) and the curing_consumption_*.xlsx workbook are DROPPED; a single reactive CO
# arbiter (_reactive_co, run once per shift AFTER building assignment) makes every
# changeover decision. The user accepted the documented risk that pure-reactive planning
# regresses coverage (three prior attempts 670k->~594k GT, 96.7%->~86% — see the
# _DYNAMIC_CO_PLANNER_ENABLED note above); the deliverable is the clean single-rule
# architecture, measured honestly. OFF (default) = bit-for-bit the current hybrid.
#
# MEASURED (2026-08, deterministic 2-seed, OFF bit-for-bit):
#   B-1 (planned COs removed, existing reactive layer): Jun 604,774 / Jul 634,189 /
#       Aug 617,384 = −128k cured vs hybrid (~−6pp; 245 mould-blocked COs on July).
#   B-2 (single once-per-shift _reactive_co arbiter + machine-swap depth-1 + B-3 surplus):
#       Jun 525,040 / Jul 579,936 / Aug 560,723 = −319k cured vs hybrid (~−15pp), WORSE
#       than B-1. Clean architecture (one rule, 0 mould-blocks) but the once-per-shift model
#       makes every CO a FULL-shift press idle and the arbiter fires the 12/day cap every
#       day → curing capacity collapses. Confirms the documented pure-reactive regression.
#   B-3 surplus release is the over-firing culprit — RCO_SURPLUS=0 (arbiter COs only
#       TRULY-idle presses) recovers massively: Jun 588,581 / Jul 658,949 (90.5%, only
#       −1.9pp!) / Aug 634,523 = −103k vs hybrid (vs −319k with surplus ON). So surplus
#       should stay OFF. The residual gap (June/Aug) is the remaining full-shift-CO cost.
# Kept OFF, code retained.
#
# ADOPTED reactive config = the B-1 MID-SHIFT base (RCO_ARBITER=0, the default) upgraded, in
# order of impact:
#   • RETARGET-ON-BLOCK — a mould-blocked press picks the neediest allowable SKU it CAN mount
#     instead of idling (July mould-blocked 394->19). The single biggest lever.
#   • FEED-GUARD RELAX under reactive — trust REAL starvation over the optimistic buildable_rate
#     estimate: a press starved RCO_STARV_SHIFTS shifts leaves for a GT-on-hand SKU (starvation
#     July 3,566 -> 2,678, below the hybrid). RCO_STARV_SHIFTS=4 (churn-tuned; 2 over-fires).
#   • Supply-gate (point 3, _supply_ok) + depth-1 machine-swap + STRICT allowable gate (R3C:
#     curing-allowable set only, no building-map fallback) + 8-shift->N starvation CO.
#   • PRE-POSITIONING (RCO_PREPOS, point 2) — light foresight from LIVE state: an under-served
#     buildable SKU (running < presses_needed) pulls a SURPLUS press (n-1 safe) via a proactive
#     Day CO added to today_cos so co_press_map PRE-BUILDS the target's GT (building couples).
#     Need-gated both sides + rate-limited (RCO_PREPOS_MAX=4/day). Net +11,546 (Aug +11k, Jul +2.9k).
# MEASURED (deterministic 2-seed, OFF bit-for-bit, R3C/R17/R10 PASS): Jun 649,900 (+4,041, BEATS
# hybrid) / Jul 670,131 (−2,455) / Aug 654,206 (−11,955) = 1,974,237 = ~−0.5pp vs hybrid.
# Progression: plain B-1 −128k -> +supply/swap/retarget −56k -> +feed-guard-relax −22k -> +prepos −10k.
# CAVEAT: the reactive CO churn leaves R5 (Stage-2 GT ≤ carcass) FAIL=3 on Aug (hybrid is R5=0) —
# a small carcass-timing residual to resolve before any adoption. Aug stays the structural laggard.
_REACTIVE_ONLY = (os.environ.get("REACTIVE_ONLY", "0") != "0")
# B-3 sub-toggle: proactive SURPLUS-press release inside _reactive_co (a press whose SKU has
# more presses than needed to finish in the remaining horizon COs early, with n-1 protection
# + 3-shift hysteresis). RCO_SURPLUS=0 → the arbiter only COs TRULY-idle presses (demand
# fully done) — tests whether stopping the every-shift over-firing recovers coverage. Only
# meaningful when _REACTIVE_ONLY.
_RCO_SURPLUS = (os.environ.get("RCO_SURPLUS", "1") != "0")
# Reactive ENGINE selector (only under _REACTIVE_ONLY). Default OFF = the B-1 MID-SHIFT base
# (a press cures until its demand hits 0, THEN COs mid-shift — best timing), now upgraded
# with the supply-gate + depth-1 machine-swap + the 8-shift starvation CO (_CURING_ADAPT_CO).
# RCO_ARBITER=1 = the once-per-shift _reactive_co arbiter (cleaner rule, full-shift CO cost).
_RCO_ARBITER = (os.environ.get("RCO_ARBITER", "0") != "0")
# Point 1 (better than the 8-shift wait): under _REACTIVE_ONLY the sustained-starvation
# switch fires after only _RCO_STARV_SHIFTS consecutive 0-GT shifts (default 2), GATED by
# the feed guard (only switches a SKU building genuinely CAN'T feed — buildable < curing
# draw). Supply-aware + fast, instead of idling 8 shifts. env RCO_STARV_SHIFTS.
_RCO_STARV_SHIFTS = int(os.environ.get("RCO_STARV_SHIFTS", "4"))
# Point 2 (default ON under REACTIVE_ONLY): light forward-looking PRE-POSITIONING. Each day,
# from LIVE state, a SKU under-served for the remaining horizon (running presses <
# presses_needed) pulls a SURPLUS press (its SKU over-served, n-1 safe) via a proactive Day
# CO added to today_cos → co_press_map pre-builds the target's GT (building couples to the
# move). Need-gated on BOTH sides + buildable-checked → foresight without over-firing. env RCO_PREPOS.
_RCO_PREPOS = (os.environ.get("RCO_PREPOS", "1") != "0")
_RCO_PREPOS_MAX = int(os.environ.get("RCO_PREPOS_MAX", "4"))   # max pre-position COs/day (rate limit)
# ── HYBRID planned-CO fixes (env, default OFF = bit-for-bit hybrid) ─────────────
# MEASURED (2026-08, hybrid, restrict=ON, corrected CT, deterministic, OFF bit-for-bit):
# ADOPTED best config = HYBRID_CO_DEFER=1 + PERSKU_FEED_V2=1 (item 2 + item 1) = +29,836
# over 3 months (Jun 650,448 / Jul 685,342 / Aug 648,687 vs baseline 646,275/672,157/636,209),
# R10/R8C/R17/R3C PASS. The four fixes OVERLAP (all cut wasteful COs) → non-additive: item 2
# defer is the biggest (Jul +14,137 alone), item 1 V2 lifts Jun/Aug; item 3 (+4,845 Jul alone)
# and item 4 don't stack on top of 1+2 (ALL-on under-prunes COs). Item 3 is a correctness fix
# (stale-CO removal) — keep OFF for max KPI, ON if you want strict staleness cleanup.
# Item 3: a dynamic CO changes the press's SKU, so ALL its future planned COs (booked against
# the OLD sku) are stale. ON wipes them for ANY dynamic CO (not just a starvation switch).
_HYBRID_CO_CANCEL = (os.environ.get("HYBRID_CO_CANCEL", "0") != "0")
# Item 2 (ADOPT): defer a planned CO (to the next working day) instead of preempting it in
# Shift A when the press's old SKU still has FULFILLABLE demand (live _supply_ok) and the press
# is NOT surplus. The single biggest lever (Jul +14,137). Pair with PERSKU_FEED_V2.
_HYBRID_CO_DEFER = True

# ── Holiday fix #1: NO new curing CO fires on a plant holiday (decision: no setup-crew
# starts a NEW changeover on the idle day; only in-flight COs/cleans finish). Every CO the
# plan placed on a holiday is DEFERRED to the next WORKING day that still has CO budget; if
# none remain (month-end) it is dropped. Makes planned/dynamic/scorer COs match the reactive
# guard (_reactive_co already returns on a holiday). INERT when PLANT_HOLIDAYS is empty →
# no-holiday runs are bit-for-bit identical. HOLIDAY_CO_DEFER=0 reproduces the old behavior
# (planned COs fire on the holiday). ──
_HOLIDAY_CO_DEFER = (os.environ.get("HOLIDAY_CO_DEFER", "1") != "0")

# ── Holiday fix #2/#3: make BUILDING holiday-aware. Both default OFF (env-gated) + inert when
# PLANT_HOLIDAYS is empty → double-layer bit-for-bit parity.
#  #2 NO-PERISH (avoid waste): don't pre-build perishable stock that will just age out over an
#     upcoming holiday — cap the carcass PASS-2 lead + the GT forward-buffer window to the WORKING
#     shifts actually reachable before the holiday. Only ever SHRINKS a build target → no overbuild.
#  #3 BRIDGE: pre-build EXTRA GT before a holiday so presses run full-rate on the first post-holiday
#     shift. MEASURED NO-OP (July, all 3 holiday scenarios byte-identical with/without it): the
#     existing 9-shift forward-buffer already pre-builds enough to bridge a 1-2 day holiday, and a
#     ≥3-day holiday is shelf-blocked — so no gap is left for a separate bridge lever. Kept OFF +
#     documented (like FIXED_ESCAPE / global-mould-opt). #2 is ADOPTED ON: cuts pre-holiday carcass
#     writeoff ~0.9-1.4k, cured-neutral, feasibility-clean, and no-holiday runs stay bit-for-bit.
_HOLIDAY_NO_PERISH = (os.environ.get("HOLIDAY_NO_PERISH_PREBUILD", "1") != "0")
_HOLIDAY_BRIDGE    = (os.environ.get("HOLIDAY_BRIDGE_BUILD",       "0") != "0")

# ── Holiday fix #7: pre-holiday Shift-C midnight cap. A plant holiday idles day D's own shifts,
# but day D-1's Shift C (23:00→07:00) is a working-day shift whose 00:00→07:00 tail lands ON the
# holiday — building would otherwise produce + start NEW COs after midnight. Cap that pre-holiday
# Shift-C to the 23:00→00:00 window (=MIN_CAMPAIGN_MINS 60): the current SKU continues for ≤60 min,
# ZERO new COs fire (60 - co_cost < 60 always skips), and nothing renders past 00:00. In-flight
# curing COs (co_carry drain on the holiday) are untouched → they still complete. Default ON but
# INERT without holidays (the (day+1) in _holiday_days guard is empty) → no-holiday bit-for-bit.
_HOLIDAY_SHIFTC_CAP      = (os.environ.get("HOLIDAY_SHIFTC_CAP", "1") != "0")
_HOLIDAY_SHIFTC_CAP_MINS = int(os.environ.get("HOLIDAY_SHIFTC_CAP_MINS", "60"))  # 23:00→00:00 pre-midnight window

# ── Shift-level Minimum Production Quantity (MPQ) — per (machine/press × SKU × shift) ──
# A production block below the floor is not emitted: batched into a later same-machine/press+SKU
# shift where possible, else DROPPED (unmet, never over-produced). 0 = disabled for that stage.
# Env BUILDING_MPQ / CURING_MPQ override bc_config. Additional to MIN_CAMPAIGN_* (independent).
_BUILDING_MPQ = int(os.environ.get("BUILDING_MPQ", getattr(_bc_cfg, "BUILDING_MPQ", 20)))
_CURING_MPQ   = int(os.environ.get("CURING_MPQ",   getattr(_bc_cfg, "CURING_MPQ",   0)))
_CARCASS_MPQ  = int(os.environ.get("CARCASS_MPQ",  getattr(_bc_cfg, "CARCASS_MPQ",  0)))

# ── Carcass row CONSOLIDATION (representation fix, KPI/feasibility-NEUTRAL) ──────────
# The Stage-1 carcass renderer emits many tiny carcass rows (a machine dribbles a few
# units of a SKU across shifts A/B/C, and a SKU's daily carcass is split over machines).
# This lever REPACKS the emitted carcass builds WITHIN each calendar day: (1) merge each
# machine's own cross-shift fragments for a SKU into one fuller block laid earliest-first,
# and (2) fold a machine's sub-threshold (< CARCASS_CONSOLIDATE_MIN) daily carcass for a
# SKU onto a SIBLING machine already building that same (date, SKU) when that sibling has
# spare day-capacity. It is provably KPI/feasibility-neutral: per-(date, SKU) carcass total
# is preserved EXACTLY (integers, no drift), movement stays WITHIN one calendar day (carcass
# aging + R5/R9C are DAY-granular), no cross-machine move ever targets a machine not already
# building that (date, SKU) (Stage-1 eligibility / inch-lock preserved), and no machine's
# per-day carcass minutes exceed its 3 shifts (R11B is shift-clipped). Carcass ≠ GT (not in
# gt_inventory) so cured/coverage/demand-cap are untouched by construction. Default ON;
# CARCASS_CONSOLIDATE=0 reverts bit-for-bit.
_CARCASS_CONSOLIDATE = os.environ.get("CARCASS_CONSOLIDATE",
                        "1" if bool(getattr(_bc_cfg, "CARCASS_CONSOLIDATE_ENABLED", True)) else "0") != "0"
_CARCASS_CONSOLIDATE_MIN = int(os.environ.get("CARCASS_CONSOLIDATE_MIN",
                              getattr(_bc_cfg, "CARCASS_CONSOLIDATE_MIN", 40)))

# ── Carcass FINAL shift-cap enforcement (residual cumulative-rounding fix) ───────────
# After every fold/split, a residual cumulative-rounding artifact can leave a carcass
# row's Qty 1-2 units ABOVE the strict floor(shift_free_min*60/ct) capacity of its OWN
# (machine, date, shift) — the boundary-split's LAST segment of a multi-shift block is
# assigned `qty_floor - already_assigned`, inheriting the flooring loss of every earlier
# segment. This lever re-caps every carcass row at its strict physical floor and moves
# any excess to a SAME (machine, date, SKU) sibling row on ANOTHER shift of the SAME
# calendar day that has spare room (carcass 1-day shelf); a residual with no same-day
# home is dropped from the display (never touches gt_inventory/cured/built) and
# reported. Display-only, KPI-neutral, deterministic. Default ON; CARCASS_SHIFT_CAP=0
# reverts bit-for-bit (pre-existing rounding artifact returns).
_CARCASS_SHIFT_CAP_ENFORCE = os.environ.get("CARCASS_SHIFT_CAP", "1") != "0"

# ── Shift-contained building CO (client hard rule) — a building CO must START and FINISH
# within one shift (07:00-15:00 / 15:00-23:00 / 23:00-07:00); it may never cross a boundary.
# ON (default): a CO that would cross is deferred to the next shift's start. OFF reverts
# bit-for-bit (COs may cross, split for display). Env CO_SHIFT_CONTAINED=0 disables.
_CO_SHIFT_CONTAINED = os.environ.get("CO_SHIFT_CONTAINED", "1") != "0"

# ── Building machine CT (seconds/unit) ────────────────────────────────────────
_BLD_CT_SEC: dict[str, float] = {
    "7001":51.6,  "7002":52.6,  "7003":56.0,  "7004":53.0,
    "6001":53.0,  "6002":52.0,  "6003":73.8,  "6004":60.0,
    "7101":83.0, "7102":86.0, "7103":60.0,  "7104":87.0,
    "7105":60.0, "7106":60,  "7201":70.0,
    "7501":90.0, "7502":90.0, "7503":90.0,
    "8201":62.0,  "8301":60.0,  "8302":60.0,
    "8501":70.0, "8502":70.0, "7301":70.0,
    "6802":146,   "6803":146,
    "6909":157,   "6911":115,   "7601":186,
    "7701":163,   "7801":135,   "7802":135,
    "7803":135,   "7804":135,   "8001":113,
    "8002":113,   "8003":113,   "8101":230,
    "ps2":48.0,   "ps3":48.0,   "ps4":48.0,   # NEW plant GT machines: ps2 dom 13", ps3 15", ps4 16" (CT 48 fallback; file CT used when present)
}

# ── Per-(SKU × machine) building CT (sec/unit) — LIVE, toggle-gated ───────────
# Loaded from bc_config.BLD_CT_FILE (data/input/Cycle_time_Building.xlsx; .csv also read). Gives a
# distinct CT for a machine depending on WHICH SKU it builds (e.g. VMI builds a
# small 12"-14" tyre in ~43s but a large 15" in 51-74s). The file is the source
# of truth for CT ONLY — allowability stays from the DB allowable matrix, and any
# (SKU, machine) pair absent from the file falls back to the per-machine fixed
# _BLD_CT_SEC above. Env BLD_CT_FILE=0 (or bc_config.BLD_CT_FILE_ENABLED=False)
# disables the lookup → every _bld_ct_sec()/_bld_qty_per_shift() call reproduces
# the fixed-per-machine plan bit-for-bit.
_BLD_CT_FILE_ENABLED = (
    getattr(_bc_cfg, "BLD_CT_FILE_ENABLED", False)
    and os.environ.get("BLD_CT_FILE", "1") != "0"
)
_BLD_CT_SKU_MACH: dict[tuple[str, str], float] = {}

def _load_bld_ct_file() -> None:
    """Populate _BLD_CT_SKU_MACH from the per-(SKU,machine) CT CSV. Silent no-op if
    the toggle is off or the file is missing/unreadable (falls back to _BLD_CT_SEC)."""
    if not _BLD_CT_FILE_ENABLED:
        return
    path = getattr(_bc_cfg, "BLD_CT_FILE", "")
    if not path or not os.path.exists(path):
        print(f"  [BLD_CT] file not found ({path}); using fixed per-machine CT")
        return
    try:
        if str(path).lower().endswith((".xlsx", ".xls")):
            _df = pd.read_excel(path, dtype=str)
        else:
            _df = pd.read_csv(path, dtype=str)
    except Exception as _e:
        print(f"  [BLD_CT] read failed ({_e}); using fixed per-machine CT")
        return
    # Rename new-machine CT columns to internal ids (file uses 6403/6404 for ps3/ps4). Match by
    # str(col) because some headers are int (6403) and some str — a plain str-key rename misses them.
    _ctmap = getattr(_bc_cfg, "BLD_CT_COL_MAP", {})
    if _ctmap:
        _df = _df.rename(columns={_c: _ctmap[str(_c)] for _c in _df.columns if str(_c) in _ctmap})
    # accept both the spaced (Yuvraj) and underscored (Consolidated_Time_Matrix) SKU-code header
    _df = _df.rename(columns={"SKU_Code": "SKU Code", "SKU_Name": "SKU Name"})
    _mach_cols = [c for c in _df.columns if str(c) in _BLD_CT_SEC]
    _n = 0
    for _, _row in _df.iterrows():
        _sku = str(_row.get("SKU Code", "")).strip()
        if not _sku:
            continue
        for _m in _mach_cols:
            _v = _row.get(_m)
            if pd.notna(_v) and str(_v).strip() != "":
                try:
                    _BLD_CT_SKU_MACH[(_sku, str(_m))] = float(_v)
                    _n += 1
                except (TypeError, ValueError):
                    pass
    print(f"  [BLD_CT] loaded {_n} per-(SKU,machine) CTs "
          f"for {len({k[0] for k in _BLD_CT_SKU_MACH})} SKUs from {os.path.basename(path)}")

_load_bld_ct_file()

def _bld_ct_sec(machine, sku=None) -> float:
    """Building cycle time (sec/unit) for `machine`, optionally specialised to `sku`.
    Per-(SKU,machine) value when the CT file is loaded AND the pair is present; else
    the per-machine fixed _BLD_CT_SEC; else 120.0. When the toggle is off, `sku` is
    ignored → identical to the historical _BLD_CT_SEC.get(machine) behaviour."""
    m = str(machine)
    if sku is not None and _BLD_CT_SKU_MACH:
        _v = _BLD_CT_SKU_MACH.get((str(sku), m))
        if _v is not None:
            return _v
    return _BLD_CT_SEC.get(m, 120.0)

# ── PM / MTC maintenance downtime (bc_config.PM_MTC_ENABLED, default OFF) ────────────
# A machine (building OR curing press) is DOWN during its window (no prod/CO); the overlapped
# shift's minutes are reduced minute-precisely. Stage detected by _MACHINE_GROUP membership.
_PM_MTC_ENABLED = bool(getattr(_bc_cfg, "PM_MTC_ENABLED", False))
_BLD_DOWN: dict[str, list] = {}    # building machine -> [(start_dt, end_dt)]
_CUR_DOWN: dict[str, list] = {}    # curing press    -> [(start_dt, end_dt)]

def _load_pm_mtc() -> None:
    if not _PM_MTC_ENABLED:
        return
    _path = getattr(_bc_cfg, "PM_MTC_FILE", "")
    if not _path or not os.path.exists(_path):
        print(f"  [PM_MTC] file not found ({_path}); no downtime applied")
        return
    from datetime import datetime as _dt
    def _parse(_t):
        try:
            return _dt.strptime(str(_t).strip(), "%d/%m/%Y:%I:%M %p")
        except (TypeError, ValueError):
            return None
    _wb = pd.ExcelFile(_path)
    for _sh in _wb.sheet_names:
        # Route STRICTLY by sheet name: 'building' sheet -> building machines,
        # 'curing' sheet -> curing presses (user-confirmed; sheets renamed 2026-08-29).
        _sl = str(_sh).strip().lower()
        if _sl in ("building", "mtc"):
            _target = _BLD_DOWN
        elif _sl in ("curing", "pm"):
            _target = _CUR_DOWN
        else:
            continue
        _df = _wb.parse(_sh)
        _df.columns = [str(c).strip() for c in _df.columns]
        if "Machine ID" not in _df.columns:
            continue
        for _, _r in _df.iterrows():
            _m = str(_r.get("Machine ID", "")).strip().split(".")[0]
            if _m == "6801":
                continue                              # plant-retired: no production, no PM/MTC
            _s, _e = _parse(_r.get("Scheduled Start Time")), _parse(_r.get("Scheduled End Time"))
            if not _m or _m.lower() == "nan" or _s is None or _e is None or _e <= _s:
                continue
            # Maintenance TYPE is the per-row 'Downtime Reason' column (PM / MTC) — NOT the
            # sheet. Stored with each window so the schedule sheets can label each row.
            _reason = str(_r.get("Downtime Reason", "")).strip().upper()
            _reason = _reason if _reason in ("PM", "MTC") else "PM"
            _target.setdefault(_m, []).append((_s, _e, _reason))
    print(f"  [PM_MTC] {sum(len(v) for v in _BLD_DOWN.values())} building + "
          f"{sum(len(v) for v in _CUR_DOWN.values())} curing downtime windows loaded")

_load_pm_mtc()

# PM/MTC NO-OVERLAP building placement: the existing building handling only REDUCES the shift's
# available minutes (so the produced qty is right) but still emitted a single StartTime→EndTime run
# that SPANNED a maintenance window. The fix lives in `_split_rows_at_shift_boundaries` (building,
# not even_qty): a machine's production skips OVER any window it hits (the window is an idle gap; the
# same production minutes resume after it, extending wall-clock — NO quantity is dropped, so the
# sheet still reconciles to the built/cured KPI and mould feasibility). A CO (indivisible) is moved
# to start just after any window it would collide with. Toggle `_PM_MTC_NO_OVERLAP` (defined below).

def _down_mins(windows, sh_start, sh_end) -> float:
    """Minutes of [sh_start, sh_end] covered by any maintenance window (minute-precise overlap)."""
    if not windows:
        return 0.0
    _tot = 0.0
    for _w in windows:
        _s, _e = _w[0], _w[1]
        _lo = max(_s, sh_start); _hi = min(_e, sh_end)
        if _hi > _lo:
            _tot += (_hi - _lo).total_seconds() / 60.0
    return _tot

# ── PM/MTC time-placement: emit activity only in the shift minutes NOT under a
# maintenance window (env PM_MTC_NO_OVERLAP, default ON whenever PM_MTC is on). A
# press's (already capacity-reduced) curing is laid into the shift's FREE
# sub-intervals so an emitted StartTime→EndTime never overlaps a window. OFF (or
# PM_MTC off) = the naive shift-start cursor (bit-for-bit baseline).
_PM_MTC_NO_OVERLAP = (_PM_MTC_ENABLED
                      and os.environ.get("PM_MTC_NO_OVERLAP", "1") != "0")

def _free_intervals(windows, sh_start, sh_end):
    """Free (non-maintenance) sub-intervals of [sh_start, sh_end], as a list of
    (start_dt, end_dt) sorted chronologically. Subtracts every maintenance window
    overlapping the shift. Returns [(sh_start, sh_end)] when there is no downtime."""
    if not windows:
        return [(sh_start, sh_end)]
    # Clip + merge the maintenance windows that touch this shift.
    _busy = []
    for _w in windows:
        _lo = max(_w[0], sh_start); _hi = min(_w[1], sh_end)
        if _hi > _lo:
            _busy.append((_lo, _hi))
    if not _busy:
        return [(sh_start, sh_end)]
    _busy.sort()
    _merged = [list(_busy[0])]
    for _lo, _hi in _busy[1:]:
        if _lo <= _merged[-1][1]:
            _merged[-1][1] = max(_merged[-1][1], _hi)
        else:
            _merged.append([_lo, _hi])
    # Complement of the merged busy windows within the shift.
    _free = []
    _cur = sh_start
    for _lo, _hi in _merged:
        if _lo > _cur:
            _free.append((_cur, _lo))
        _cur = max(_cur, _hi)
    if _cur < sh_end:
        _free.append((_cur, sh_end))
    return _free

def _post_maint_free(windows, sh_start, sh_end):
    """(post_free_mins, post_start_dt) for a press's shift: production resumes strictly AFTER
    the LAST maintenance window that overlaps this shift. post_start = the latest window-end
    within the shift; post_free = the free minutes from there to shift end. The pre-maintenance
    gap (before the last window) is NOT counted — a deferred remainder carries to a later shift,
    not backwards into that gap (per plant rule). No overlap → (full shift, sh_start)."""
    _wend = sh_start
    for _w in (windows or []):
        _lo = max(_w[0], sh_start); _hi = min(_w[1], sh_end)
        if _hi > _lo and _hi > _wend:
            _wend = _hi
    _wend = min(_wend, sh_end)
    return max(0.0, (sh_end - _wend).total_seconds() / 60.0), _wend

# Sub-toggle of PM_MTC_NO_OVERLAP (default ON when no-overlap is on): also USE the
# free minutes that fall BEFORE a maintenance window inside the same shift, not only
# the post-window tail. A window like 10:00–18:00 leaves 07:00–10:00 free in Shift A;
# the machine can legitimately build there and go down at 10:00 — that block does NOT
# overlap the window. The old `_post_maint_free` discarded that pre-window gap (it only
# counted time after the LAST window), permanently losing it for a fully-loaded machine
# (the "deferred remainder" it assumed would rebuild later never can when the machine is
# at capacity). `_best_maint_free` returns the LARGEST free sub-interval of the shift
# (its length + start) so production fits in one contiguous, window-free block. Set
# PM_MTC_PREWINDOW=0 to revert to post-window-only (the prior no-overlap behaviour).
_PM_MTC_PREWINDOW = (_PM_MTC_ENABLED
                     and os.environ.get("PM_MTC_PREWINDOW", "1") != "0")

def _best_maint_free(windows, sh_start, sh_end):
    """(free_mins, start_dt) = the LARGEST maintenance-free sub-interval of [sh_start, sh_end].
    Unlike `_post_maint_free` (post-last-window only) this also considers the gap BEFORE a
    window, so a machine keeps the pre-window minutes it can genuinely use without ever
    overlapping a window. A single contiguous campaign is emitted from `start_dt` for at most
    `free_mins`, so it stays inside this one free interval. No window → (full shift, sh_start)."""
    _free = _free_intervals(windows, sh_start, sh_end)
    _best_len, _best_start = 0.0, sh_start
    for _lo, _hi in _free:
        _len = (_hi - _lo).total_seconds() / 60.0
        if _len > _best_len:
            _best_len, _best_start = _len, _lo
    return _best_len, _best_start

def _pm_maint_free(windows, sh_start, sh_end):
    """No-overlap free-time picker: the pre-window-aware `_best_maint_free` when
    `_PM_MTC_PREWINDOW` is on (the fix), else the post-window-only `_post_maint_free`."""
    return (_best_maint_free(windows, sh_start, sh_end) if _PM_MTC_PREWINDOW
            else _post_maint_free(windows, sh_start, sh_end))

def _pm_shift_of_bounds(dt):
    """(shift_start, shift_end) datetimes for the plant shift containing dt (07/15/23)."""
    _h = dt.hour
    _d0 = dt.replace(minute=0, second=0, microsecond=0)
    if 7 <= _h < 15:
        _s = _d0.replace(hour=7)
    elif 15 <= _h < 23:
        _s = _d0.replace(hour=15)
    elif _h >= 23:
        _s = _d0.replace(hour=23)
    else:                                  # 00:00–06:59 → shift C started 23:00 the prior day
        _s = (_d0 - timedelta(days=1)).replace(hour=23)
    return _s, _s + timedelta(minutes=float(SHIFT_MINS))

def _pm_relocate_curing_rows(rows):
    """POST-PROCESS (env PM_MTC_NO_OVERLAP, ON with PM_MTC): after rows are shift-split, move
    every CURING (Qty>0) row whose [StartTime,EndTime] overlaps this press's maintenance window
    into the FREE (non-maintenance) sub-interval(s) of its OWN shift — quantity unchanged. Runs
    AFTER the sim + shift-boundary split, so it touches NO plan state and never feeds the
    per-press continuous-cursor / even-Qty machinery (no cascade). A row that already clears its
    windows is left byte-identical. A row that fits one free interval stays a SINGLE row (the R17
    bipartite mould audit counts one (press,SKU) per Qty>0 row → unaffected); only a row longer
    than every single free interval (a window bisecting the shift with production exceeding either
    side) is SPLIT across intervals — does not occur in the current data. OFF → not called."""
    if not (_PM_MTC_NO_OVERLAP and _CUR_DOWN):
        return rows
    def _p(s):
        try:
            return datetime.strptime(str(s), "%Y-%m-%d %H:%M")
        except (TypeError, ValueError):
            return None
    _out = []
    _moved = 0
    for r in rows:
        try:
            _q = float(r.get("Qty", 0) or 0)
        except (TypeError, ValueError):
            _q = 0.0
        _wins = _CUR_DOWN.get(str(r.get("Machine", "")))
        _st = _p(r.get("StartTime")); _en = _p(r.get("EndTime"))
        if _q <= 0 or not _wins or _st is None or _en is None or _en <= _st:
            _out.append(r)                                  # not a curing row / no window data → unchanged
            continue
        _sh_s, _sh_e = _pm_shift_of_bounds(_st)
        if _down_mins(_wins, _sh_s, _sh_e) <= 0:            # no maintenance in this shift → unchanged
            _out.append(r)
            continue
        _post_free, _post_start = _pm_maint_free(_wins, _sh_s, _sh_e)
        _dur = (_en - _st).total_seconds() / 60.0
        if _post_free >= _dur - 1e-6 and _post_free > 0:
            # place the whole (sim-capped) run inside the largest window-free block → single
            # row, no split, no overlap (the block start may be pre- or post-window).
            _nr = dict(r)
            _nr["StartTime"] = _fmt_dt(_post_start)
            _nr["EndTime"]   = _fmt_dt(_post_start + timedelta(minutes=_dur))
            _out.append(_nr)
            _moved += 1
        elif _post_free > 0:
            # defensive: run longer than the free block (should not occur — the sim caps curing
            # to the free minutes). Clamp the row to END of the free block so it never re-enters
            # the window.
            _nr = dict(r)
            _nr["StartTime"] = _fmt_dt(_post_start)
            _nr["EndTime"]   = _fmt_dt(_post_start + timedelta(minutes=_post_free))
            _out.append(_nr)
            _moved += 1
        else:
            _out.append(r)                                  # no post-maintenance time (window to shift end) → leave
    if _moved:
        print(f"  [PM_MTC] no-overlap: relocated {_moved} curing row(s) out of maintenance windows")
    return _out

def _pm_relocate_carcass_rows(rows):
    """POST-PROCESS (env PM_MTC_NO_OVERLAP, ON with PM_MTC): after carcass rows are shift-split
    AND shift-capacity-capped, move every Stage-1 CARCASS (CO_Type=='carcass', Qty>0) row whose
    [StartTime,EndTime] straddles its machine's PM/MTC window into the shift's FREE
    (non-maintenance) sub-interval(s) — quantity unchanged. Mirrors `_pm_relocate_curing_rows`
    for the building carcass rows (GT building already clips at production time via `_mdown`;
    carcass rows are rendered post-plan and only had their QTY reduced for maintenance, never
    their time span, so the emitted StartTime→EndTime could still bleed into a window). The Qty
    was already capped to the free minutes upstream (`_avail_for_sku`/`_enforce_carcass_shift_cap`
    both subtract PM), so the run always fits the total free time; if it exceeds the LARGEST single
    free block it is SPLIT across free intervals (Qty apportioned by placed minutes, integer,
    remainder to the last slice). Display-only / KPI-neutral (carcass not in gt_inventory; per-
    (machine,date,SKU,shift) Qty total preserved EXACTLY). A row that already clears its windows is
    left byte-identical. OFF → not called."""
    if not (_PM_MTC_NO_OVERLAP and _BLD_DOWN):
        return rows
    def _p(s):
        try:
            return datetime.strptime(str(s), "%Y-%m-%d %H:%M")
        except (TypeError, ValueError):
            return None
    _out = []
    _moved = 0
    for r in rows:
        if r.get("CO_Type") != "carcass" or str(r.get("SKUCode")) == "CHANGEOVER":
            _out.append(r); continue
        try:
            _q = int(round(float(r.get("Qty", 0) or 0)))
        except (TypeError, ValueError):
            _q = 0
        _wins = _BLD_DOWN.get(str(r.get("Machine", "")))
        _st = _p(r.get("StartTime")); _en = _p(r.get("EndTime"))
        if _q <= 0 or not _wins or _st is None or _en is None or _en <= _st:
            _out.append(r); continue
        _sh_s, _sh_e = _pm_shift_of_bounds(_st)
        if _down_mins(_wins, _sh_s, _sh_e) <= 0 or _down_mins(_wins, _st, _en) <= 1e-6:
            _out.append(r); continue                       # no maintenance / no actual straddle
        _ct = _bld_ct_sec(str(r.get("Machine", "")), str(r.get("SKUCode")))
        _free = _free_intervals(_wins, _sh_s, _sh_e)
        _total_free = sum((b - a).total_seconds() / 60.0 for a, b in _free)
        _dur = (_en - _st).total_seconds() / 60.0
        if _total_free <= 0 or _ct <= 0:
            _out.append(r); continue
        _rem = min(_dur, _total_free)
        _placed = []                                       # (start_dt, mins) chronological
        for a, b in _free:
            if _rem <= 1e-6:
                break
            _take = min((b - a).total_seconds() / 60.0, _rem)
            if _take > 1e-6:
                _placed.append((a, _take)); _rem -= _take
        if len(_placed) == 1:
            _a, _m = _placed[0]
            _nr = dict(r)
            _nr["StartTime"] = _fmt_dt(_a)
            _nr["EndTime"]   = _fmt_dt(_a + timedelta(minutes=min(_m, _q * _ct / 60.0)))
            _out.append(_nr); _moved += 1
        else:
            _sum_m = sum(m for _, m in _placed) or 1.0
            _assigned = 0
            for _i, (_a, _m) in enumerate(_placed):
                _qi = int(_q * _m / _sum_m) if _i < len(_placed) - 1 else _q - _assigned
                _assigned += _qi
                if _qi <= 0:
                    continue
                _nr = dict(r)
                _nr["Qty"]       = _qi
                _nr["StartTime"] = _fmt_dt(_a)
                _nr["EndTime"]   = _fmt_dt(_a + timedelta(minutes=min(_m, _qi * _ct / 60.0)))
                _out.append(_nr)
            _moved += 1
    if _moved:
        print(f"  [PM_MTC] no-overlap: relocated {_moved} carcass row(s) out of maintenance windows")
    return _out

def _enforce_carcass_min_qty(rows):
    """HARD min-carcass floor on the DISPLAY rows (env CARCASS_MIN_ENFORCE, MIN_CARCASS_QTY).
    The gate already enforces ≥MIN on the BUILT carcass (and clamps GT/curing to it), but the
    post-plan carcass RENDERER + shift-cap redistribution can re-slice a machine's day total into
    a sub-MIN shift fragment (the 'over-production shifting' leftover). This folds every sub-MIN
    carcass slice into the largest same-(machine, date, SKU) SIBLING that has spare shift capacity
    — so the per-(machine, date, SKU) carcass TOTAL is preserved EXACTLY (GT ≤ carcass / R5 sync
    untouched; carcass not in gt_inventory) and no carcass row < MIN is emitted. A sub-MIN slice
    with no capacity-safe sibling that day is DROPPED (its units are genuinely unplaceable ≥MIN);
    this is rare and display-only. Runs after _enforce_carcass_shift_cap. OFF → bit-for-bit."""
    if not (_CARCASS_MIN_ENFORCE and _CARCASS_MIN_QTY > 0):
        return rows
    from collections import defaultdict as _dd
    def _is_carc(r):
        return r.get("CO_Type") == "carcass" and str(r.get("SKUCode")) != "CHANGEOVER"
    def _pdt(s):
        try:
            return datetime.strptime(str(s), "%Y-%m-%d %H:%M")
        except (TypeError, ValueError):
            return None
    _grp = _dd(list)                                   # (machine, date, sku) -> [rows]
    for r in rows:
        if _is_carc(r):
            _grp[(str(r["Machine"]), str(r["Date"])[:10], str(r["SKUCode"]))].append(r)
    _drop = set(); _moved = 0; _dropped_units = 0
    for (_m, _dt0, _sk), _rs in _grp.items():
        _ct = _bld_ct_sec(_m, _sk)
        for _r in _rs:
            _q = int(round(_r.get("Qty", 0) or 0))
            if not (0 < _q < _CARCASS_MIN_QTY) or id(_r) in _drop:
                continue
            _recv = None
            for _cand in sorted((x for x in _rs if x is not _r and id(x) not in _drop),
                                 key=lambda x: -int(round(x.get("Qty", 0) or 0))):
                _cs, _ce = _pdt(_cand.get("StartTime")), _pdt(_cand.get("EndTime"))
                _span = (_ce - _cs).total_seconds() / 60.0 if (_cs and _ce) else 0.0
                _add = (_q * _ct / 60.0) if _ct > 0 else 0.0
                if _ct <= 0 or _span + _add <= SHIFT_MINS + 1e-6:
                    _recv = _cand; break
            if _recv is None:                          # nothing can absorb it ≥MIN → drop (rare)
                _drop.add(id(_r)); _dropped_units += _q
                continue
            _recv["Qty"] = int(round(_recv.get("Qty", 0) or 0)) + _q
            _cs = _pdt(_recv.get("StartTime"))
            if _cs and _ct > 0:
                _recv["EndTime"] = _fmt_dt(_cs + timedelta(minutes=int(round(_recv["Qty"])) * _ct / 60.0))
            _drop.add(id(_r)); _moved += 1
    if _drop:
        rows = [r for r in rows if id(r) not in _drop]
    if _moved or _dropped_units:
        print(f"  [Stage-1 carcass] min-qty ≥{_CARCASS_MIN_QTY}: folded {_moved} sub-{_CARCASS_MIN_QTY} "
              f"slice(s) into same-machine-day siblings"
              + (f"; dropped {_dropped_units} unplaceable unit(s) (display-only)" if _dropped_units else ""))
    return rows

def _shift_of(_dt) -> str:
    """Shift letter from a datetime's hour: A 07-15, B 15-23, C 23-07."""
    _h = _dt.hour
    return "A" if 7 <= _h < 15 else ("B" if 15 <= _h < 23 else "C")

def _pm_mtc_display_rows(plan_start, planning_days, stage):
    """DISPLAY-ONLY maintenance rows for a Shift Schedule sheet — one row per (equipment,
    window) overlapping the plan horizon, from BOTH sources (PM=curing presses +
    MTC=building machines). BOTH sheets show BOTH types, each labelled: building via
    CO_Type ('PM'/'MTC'), curing via Remarks ('PM Schedule'/'MTC Schedule'). Never fed
    to prod_rows / KPIs / feasibility. stage: 'building' or 'curing' (row schema)."""
    from datetime import timedelta as _td
    if not _PM_MTC_ENABLED:
        return []
    _h_start = plan_start
    _h_end   = plan_start + _td(days=planning_days)
    _rows = []
    for _down_dict in (_BLD_DOWN, _CUR_DOWN):
        for _id, _wins in (_down_dict or {}).items():
            # Route by EQUIPMENT TYPE: building machines → building sheet (label in CO_Type),
            # curing presses → curing sheet (label in Remarks). The PM/MTC TYPE comes from
            # each window's 'Downtime Reason' — so BOTH sheets carry BOTH PM and MTC rows.
            _is_bld = _id in _MACHINE_GROUP
            if (stage == "building") != _is_bld:      # building sheet ⇔ building machine only
                continue
            for _w in _wins:
                _s, _e = _w[0], _w[1]
                _reason = _w[2] if len(_w) > 2 else "PM"   # "PM" or "MTC"
                if _e <= _h_start or _s >= _h_end:  # outside plan horizon (e.g. Sept windows on a July run)
                    continue
                _mins = round((_e - _s).total_seconds() / 60.0)
                _st = _s.strftime("%Y-%m-%d %H:%M"); _en = _e.strftime("%Y-%m-%d %H:%M")
                if stage == "building":
                    _rows.append({"Machine": _id, "Date": _s.strftime("%Y-%m-%d"),
                                  "Shift": _shift_of(_s), "SKUCode": "—", "Qty": 0,
                                  "CO_Mins": _mins, "StartTime": _st, "EndTime": _en,
                                  "Machine_Group": _MACHINE_GROUP.get(_id, ""), "CO_Type": _reason,
                                  "Remarks": f"{_reason} Schedule"})
                else:
                    _rows.append({"Date": _s.strftime("%Y-%m-%d"), "Shift": _shift_of(_s),
                                  "Machine": _id, "SKUCode": "—", "StartTime": _st,
                                  "EndTime": _en, "Qty": 0, "CO_Mins": _mins,
                                  "Mould_Clean_Mins": 0, "CycleTime_min": "", "GT_Inventory": "",
                                  "CO_Type": _reason, "Remarks": f"{_reason} Schedule",
                                  "_status": "PM_MTC"})
    return _rows

DEFAULT_CURING_CT = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
CURING_CAVITIES   = 2

# ── Dominant inch per building machine ────────────────────────────────────────
# Used to prioritise same-dominant-inch SKUs when choosing campaigns/COs.
# When a machine's dominant-inch SKUs have no deficit, it can build other inches
# per its allowable list — but dominant inch is always tried first.
_MACHINE_DOMINANT_INCH: dict[str, str] = {
    "6001": "14", "6002": "15", "6003": "17", "6004": "16",
    "7001": "16", "7002": "14", "7003": "15", "7004": "14",
    "7101": "15", "7102": "15", "7103": "13", "7104": "15",
    "7105": "13", "7106": "13", "7201": "16",
    "7501": "12", "7502": "13", "7503": "13",
}

# ── Data-driven multi-inch dominant ranking (DOMINANT_INCH_FILE) — Phase 5 ─────
# Replaces the single hardcoded dominant inch above with an ORDERED multi-inch band
# per machine, extracted from the plant's last-N-day running data
# (data/analysis_aug/machine_inch_dominant_aug.xlsx). ranked[0] = dominant inch →
# overrides the scalar _MACHINE_DOMINANT_INCH so every existing consumer keeps
# working but on real-data dominance; the full ordered list is exposed as
# _MACHINE_DOMINANT_INCH_RANKED for the inch band / graded inch penalty in later
# phases. OFF (default) or a missing file → 1-element lists from the hardcoded map
# and the scalar map is untouched → bit-for-bit baseline. Env DOMINANT_INCH_FILE=0
# also disables.
_DOMINANT_INCH_FILE_ENABLED = os.environ.get(
    "DOMINANT_INCH_FILE",
    "1" if getattr(_bc_cfg, "DOMINANT_INCH_FILE_ENABLED", False) else "0"
) != "0"
_MACHINE_DOMINANT_INCH_RANKED: dict[str, list[str]] = {
    _m: [_v] for _m, _v in _MACHINE_DOMINANT_INCH.items()
}

def _load_dominant_inch_file() -> None:
    """Populate _MACHINE_DOMINANT_INCH_RANKED (ordered, dominant-first) and override the
    scalar _MACHINE_DOMINANT_INCH from the plant dominant-inch xlsx. No-op if the toggle
    is off or the file is missing/unreadable (keeps the hardcoded map bit-for-bit)."""
    if not _DOMINANT_INCH_FILE_ENABLED:
        return
    path = getattr(_bc_cfg, "DOMINANT_INCH_FILE", "")
    if not path or not os.path.exists(path):
        print(f"  [DOM_INCH] file not found ({path}); using hardcoded dominant inch")
        return
    try:
        _df = pd.read_excel(path, sheet_name="Dominant_Inch", dtype=str)
    except Exception as _e:
        print(f"  [DOM_INCH] read failed ({_e}); using hardcoded dominant inch")
        return
    _n = 0
    for _, _row in _df.iterrows():
        _m = str(_row.get("Machine", "")).strip()
        _ranked = str(_row.get("Ranked_Inches", "") or "").strip()
        if not _m or not _ranked:
            continue
        _lst = [x.strip() for x in _ranked.split(",") if x.strip()]
        if not _lst:
            continue
        _MACHINE_DOMINANT_INCH_RANKED[_m] = _lst
        _MACHINE_DOMINANT_INCH[_m] = _lst[0]      # scalar dominant = data-driven top inch
        _n += 1
    print(f"  [DOM_INCH] loaded ranked inch bands for {_n} machines "
          f"from {os.path.basename(path)}")

_load_dominant_inch_file()

# ── Historical inch-LOCK (INCH_HIST_LOCK) ─────────────────────────────────────
# Per-machine ALLOWED-INCH sets from the 4-month plant report replace the anchor±2
# band. FIXED machines (single historical inch) do zero diff-size CO; FLEXIBLE
# machines may only build/CO among their ranked historical inches (±2 discontinued).
# OFF (env INCH_HIST_LOCK=0) or a missing file → current ±2 behaviour bit-for-bit.
_INCH_HIST_LOCK_ENABLED = bool(getattr(_bc_cfg, "INCH_HIST_LOCK_ENABLED", False))
_INCH_HIST_LOCK_STAGE1  = bool(getattr(_bc_cfg, "INCH_HIST_LOCK_STAGE1", False))
_MACHINE_ALLOWED_INCHES: dict[str, list[str]] = {}   # ranked, dominant-first (per machine)
_MACHINE_ALLOWED_INCH_SET: dict[str, set] = {}       # same as a set, for membership tests
# #1 BJ never takes a +3/-3 inch jump (plant rule). ADOPTED default ON; BJ_NO_PLUS3=0 reverts.
_BJ_MACHINES = {"7101", "7102", "7103", "7104", "7105", "7106", "7201"}
_BJ_NO_PLUS3 = os.environ.get("BJ_NO_PLUS3", "1") != "0"
# #2 non-BJ/non-Stage-2: a DIRECT +3/-3 in one CO is priced at the 8h INCH_PLUS3 cost, so the scorer
# picks the cheaper two-hop (15→16→18) when productive and pays 8h only when it isn't. ADOPTED default
# ON; VMI_TWO_HOP=0 reverts (direct +3 at the normal diff cost, bit-for-bit).
_VMI_TWO_HOP = os.environ.get("VMI_TWO_HOP", "1") != "0"
# #5 revert-dwell: a machine may revert to a left inch only after building its current inch this many
# days (blocks RAPID Size1→Size2→Size1 flip-flops; 0 = off = bit-for-bit). Hard-ban measured −25k.
_REVERT_DWELL_DAYS = int(os.environ.get("REVERT_DWELL_DAYS", "0"))
# SKU no-revert (HARD, max-productivity rule): once a machine LEAVES a SKU it may never build that
# SKU again (no round-trips). Inch reverts still allowed (dwell-gated) to a DIFFERENT SKU. Default OFF.
_SKU_NO_REVERT = os.environ.get("SKU_NO_REVERT", "0") != "0"

def _load_inch_hist_lock() -> None:
    """Build _MACHINE_ALLOWED_INCHES/_SET from the 4-month plant building report's
    Inch_Counts_Matrix: keep every inch that is >= INCH_HIST_LOCK_MIN_SHARE of the
    machine's records, ranked by count, capped at INCH_HIST_LOCK_MAX_INCHES (always
    at least the dominant). Also overrides the dominant-inch maps (dominant = the
    top historical inch) so anchor-seeding + inch_penalty use the plant history.
    No-op if the toggle is off or the file is missing (keeps ±2 bit-for-bit)."""
    if not _INCH_HIST_LOCK_ENABLED:
        return
    # FAST-ITERATION HOOK: env INCH_LOCK_JSON → load per-machine allowed-inch sets from a
    # JSON file {machine: [inches]} (overrides bc_config). For tuning the Sept inch-lock.
    _lj = os.environ.get("INCH_LOCK_JSON", "")
    if _lj and os.path.exists(_lj):
        import json as _json
        _sets = _json.load(open(_lj))
        _n = 0
        for _m, _al in _sets.items():
            _m = str(_m).strip(); _keep = [str(i).strip() for i in _al if str(i).strip()]
            if not _m or not _keep:
                continue
            _MACHINE_ALLOWED_INCHES[_m] = list(_keep); _MACHINE_ALLOWED_INCH_SET[_m] = set(_keep)
            _MACHINE_DOMINANT_INCH[_m] = _keep[0]; _MACHINE_DOMINANT_INCH_RANKED[_m] = list(_keep)
            _n += 1
        print(f"  [INCH_HIST_LOCK] loaded {_n} machines from INCH_LOCK_JSON={_lj}")
        return
    # Preferred source: the hardcoded, already-final sets in bc_config. When present,
    # these are the FINAL ranked allowed-inch lists (2%/max-3/BJ-no-+3 already applied),
    # so build the four maps DIRECTLY and SKIP the Excel read + per-2%/rank/BJ logic.
    _cfg_sets = getattr(_bc_cfg, "INCH_HIST_LOCK_SETS", None)
    if isinstance(_cfg_sets, dict) and _cfg_sets:
        _n = 0
        for _m, _al in _cfg_sets.items():
            _m = str(_m).strip()
            _keep = [str(i).strip() for i in _al if str(i).strip()]
            if not _m or not _keep:
                continue
            _MACHINE_ALLOWED_INCHES[_m] = list(_keep)
            _MACHINE_ALLOWED_INCH_SET[_m] = set(_keep)
            _MACHINE_DOMINANT_INCH[_m] = _keep[0]
            _MACHINE_DOMINANT_INCH_RANKED[_m] = list(_keep)
            _n += 1
        _nf = sum(1 for v in _MACHINE_ALLOWED_INCHES.values() if len(v) == 1)
        print(f"  [INCH_HIST_LOCK] loaded allowed-inch sets for {_n} machines from "
              f"bc_config.INCH_HIST_LOCK_SETS ({_nf} FIXED single-inch, "
              f"{_n - _nf} FLEXIBLE); ±2 band discontinued")
        return
    path = getattr(_bc_cfg, "INCH_HIST_LOCK_FILE", "")
    if not path or not os.path.exists(path):
        print(f"  [INCH_HIST_LOCK] file not found ({path}); ±2 band kept")
        return
    try:
        _df = pd.read_excel(path, sheet_name="Inch_Counts_Matrix")
    except Exception as _e:
        print(f"  [INCH_HIST_LOCK] read failed ({_e}); ±2 band kept")
        return
    _min_share = float(getattr(_bc_cfg, "INCH_HIST_LOCK_MIN_SHARE", 0.02))
    _max_inch  = int(getattr(_bc_cfg, "INCH_HIST_LOCK_MAX_INCHES", 3))
    _inch_cols = [c for c in _df.columns
                  if str(c).replace('"', '').strip().isdigit()]
    _n = 0
    for _, _row in _df.iterrows():
        _m = str(_row.get("Machine", "")).strip()
        if not _m:
            continue
        _counts = {str(c).replace('"', '').strip(): float(_row[c] or 0)
                   for c in _inch_cols}
        _tot = sum(_counts.values())
        if _tot <= 0:
            continue
        _ranked = sorted((i for i, c in _counts.items() if c > 0),
                         key=lambda i: (-_counts[i], int(i)))
        _keep = [i for i in _ranked if _counts[i] / _tot >= _min_share][:_max_inch]
        if not _keep:
            _keep = _ranked[:1]
        # #1 BJ_NO_PLUS3: BJ machines may NEVER take a +3/-3 inch jump (plant rule). Drop any
        # allowed inch >2 away from the dominant, so a BJ machine's span is ≤2 (e.g. 7103
        # {13,15,16}→{13,15}, 7201 {13,16}→{16}). Env BJ_NO_PLUS3=0 → keep historical set.
        if _BJ_NO_PLUS3 and str(_m) in _BJ_MACHINES and len(_keep) > 1:
            _dom_n = int(_keep[0])
            _keep = [i for i in _keep if abs(int(i) - _dom_n) <= 2]
        _MACHINE_ALLOWED_INCHES[_m] = _keep
        _MACHINE_ALLOWED_INCH_SET[_m] = set(_keep)
        _MACHINE_DOMINANT_INCH[_m] = _keep[0]
        _MACHINE_DOMINANT_INCH_RANKED[_m] = list(_keep)
        _n += 1
    _nf = sum(1 for v in _MACHINE_ALLOWED_INCHES.values() if len(v) == 1)
    print(f"  [INCH_HIST_LOCK] loaded allowed-inch sets for {_n} machines from "
          f"{os.path.basename(path)} ({_nf} FIXED single-inch, {_n - _nf} FLEXIBLE); "
          f"±2 band discontinued")

_load_inch_hist_lock()

# ── 18-inch EXCEPTION (env INCH18_EXC, default OFF = bit-for-bit) ─────────────────────────
# 18" SKUs are DB-allowable on all 4 VMI-Maxx machines (7001-7004) but the historical inch-lock
# pins 18" to 7001 ONLY (7001 set = {15,16,18}). 7001 is 91% busy on 15"/16" until ~day 22, so the
# Day-0 presses committed to 18" starve ~22 days before 7001 can build 18" GT (build↔cure misalign,
# ~200 phantom starvation events). Fix by rebalancing the VMI-Maxx inch-locks so 18" lives on the
# under-loaded 7003 and 7001 is freed for 15":
#   7001 {15,16,18} → {15}   (15"-dedicated; its 16" work moves to 6004)
#   7003 {15}       → {15,18} (FLEXIBLE; builds 18" FIRST when a press is drawing it — see §2 at _key)
#   6004 {16}       → {16}    (unchanged; absorbs 7001's 16")
# Building auto-follows curing draw (no-waste coupling) + Day-0 presses already draw 18" from day 1,
# so relaxing 7003 + prioritizing 18"-when-drawn makes 18" GT arrive in time — the misalignment
# self-heals; no Phase-0 change needed. DB-allowable only (no invented pairs). All edits ORDERING/gate.
_INCH18_EXC = os.environ.get("INCH18_EXC", "0") != "0"
# INCH18_DEFER (the "superior" coordinated rule): same inch-lock rebalance, BUT 7003 builds 15" ONLY
# until a computed switch_day (option b: when the OTHER 15" machines can cover the remaining 15"),
# then takes a direct 8h CO 15"→18" (never 16"), builds 18" the rest of the month, one-way (no revert).
# Phase-0 curing defers 18" press acquisition to switch_day + frees the Day-0 18" press(es) early.
_INCH18_DEFER = os.environ.get("INCH18_DEFER", "1") != "0"    # ADOPTED (default ON): VMI realloc —
#   7001→15, 6004→16, 7003→16"(d1-20)→18"(d21+), 7002→14"→16" one-way (no churn). INCH18_DEFER=0 reverts.
_INCH18_MACHINE = os.environ.get("INCH18_MACHINE", "7003")     # the machine 18" moves to
# 7501 one-way 12"→13" (env INCH_7501_FLEX, default ON): 7501 builds 12" until all 12" demand is
# done, then takes ONE diff-CO to 13" and stays (no revert). =0 keeps 7501 locked to its config set.
_INCH_7501_FLEX = os.environ.get("INCH_7501_FLEX", "1") != "0"
if _INCH18_EXC or _INCH18_DEFER:
    # Optimal VMI allocation (June+July analysis): 7001→15, 6004→16, 7003→16(early)+18(from day21),
    # 7002→14(early)+16(one-way, kills its 14↔16 churn). Others keep historical (6001/7004→14, 6002→15,
    # 6003→17). 16→18 and 14→16 are +2 jumps (normal diff-CO, not the 8h +3).
    # NOTE: bc_config.INCH_HIST_LOCK_SETS now ALSO carries these adopted sets (7001→[15], 7003→[16,18],
    # 7002→[14,16], 6004→[16]) so the config file reflects reality — this loop is idempotent belt-and-
    # suspenders (re-asserts them when INCH18_DEFER on). The per-day timing (7003 16→18@d21, 7002 one-way)
    # lives in the day loop; INCH18_DEFER=0 keeps the config sets but drops the timing.
    # 7003 carries 18" AND 17": its plant Days-1-2 set is three 17" SKUs (HRHT0/HRHP0/HURL0) and
    # 1325223517104HRHT0 (7,000 units) has NO other allowable machine at all — pinning 7003 to 18"
    # alone left the plant-set intersection EMPTY, so _plant_set_done could never clear and those
    # units were permanently unbuildable. 18" stays FIRST (anchor) so the VMI realloc intent holds.
    _inch18_override = {"7001": ["15"], _INCH18_MACHINE: ["18", "17"], "6004": ["16"], "7002": ["14", "16"]}
    for _m, _al in _inch18_override.items():
        _MACHINE_ALLOWED_INCHES[_m] = list(_al)
        _MACHINE_ALLOWED_INCH_SET[_m] = set(_al)
        _MACHINE_DOMINANT_INCH[_m] = _al[0]                    # anchor = first (15" 7001, 16" 7003, 14" 7002)
        _MACHINE_DOMINANT_INCH_RANKED[_m] = list(_al)
    print("  [INCH18_DEFER] VMI realloc — " + ", ".join(
        f"{_m}→{{{','.join(_al)}}}" for _m, _al in _inch18_override.items())
        + " (DB-allowable, no invented pairs)")

# ── CT_INCH_TRUST (env CT_INCH_TRUST, default ON; =0 reverts bit-for-bit) ─────────────────
# The CT matrix (bc_config.BLD_CT_FILE) and the historical inch-lock disagree: the CT file
# CERTIFIES (machine, SKU) pairs whose inch the lock then forbids, so the pair is promised in
# the matrix and silently unusable in the plan. Two measured casualties:
#   8302 is CT-certified for 1325213615099MSXT0 (15") but inch-locked to {12,13} -> the SKU's
#        only other Stage-2 (8301) is 123% oversubscribed, so 6911 (its SOLE Stage-1) idles 100%.
#   8501 is CT-certified for 14" SKUs but inch-locked to {15,12,13} -> 4,068 units of 14" work
#        are structurally dead and 7803 (sole Stage-1 for 4 of its 5 SKUs) sits at 19%.
# Plant decision: the CT file is authoritative for capability. Widen each machine's allowed-inch
# set to the inches its OWN CT-certified SKUs need. Only ever ADDS inches, never removes.
# The explicit INCH18_DEFER realloc machines are EXEMPT — those sets are a deliberate plant call.
_CT_INCH_TRUST = os.environ.get("CT_INCH_TRUST", "1") != "0"
if _CT_INCH_TRUST and _MACHINE_ALLOWED_INCH_SET:
    try:
        _exempt = set(_inch18_override) if (_INCH18_EXC or _INCH18_DEFER) else set()
        _ctdf = pd.read_excel(getattr(_bc_cfg, "BLD_CT_FILE", ""), sheet_name=0)
        _ren = {"6402": "ps2", "6403": "ps3", "6404": "ps4"}
        _skc = next(c for c in _ctdf.columns if str(c).strip().lower() in ("sku code", "skucode"))
        _ctdf["_sk"] = _ctdf[_skc].astype(str).str.strip()
        _added = []
        for _c in _ctdf.columns:
            _m = _ren.get(str(_c).strip(), str(_c).strip())
            if _m in ("6801", "6401") or _m not in _MACHINE_ALLOWED_INCH_SET or _m in _exempt:
                continue
            _v = pd.to_numeric(_ctdf[_c], errors="coerce")
            _need = {x[8:10] for x in _ctdf.loc[_v.notna(), "_sk"] if len(x) == 18}
            _new = _need - set(_MACHINE_ALLOWED_INCH_SET[_m])
            if _new:
                _MACHINE_ALLOWED_INCH_SET[_m] = set(_MACHINE_ALLOWED_INCH_SET[_m]) | _new
                _MACHINE_ALLOWED_INCHES[_m] = list(_MACHINE_ALLOWED_INCHES.get(_m, [])) + sorted(_new)
                _MACHINE_DOMINANT_INCH_RANKED[_m] = list(_MACHINE_ALLOWED_INCHES[_m])
                _added.append(f"{_m}+{'/'.join(sorted(_new))}")
        if _added:
            print(f"  [CT_INCH_TRUST] CT-certified inches admitted on {len(_added)} machine(s): "
                  f"{', '.join(_added)}")
    except Exception as _exc:
        print(f"  [CT_INCH_TRUST] skipped ({_exc})")

# Flexible machines under the historical lock (allowed-inch set >= 2) — the only
# machines that can redirect between inches; fixed machines are single-inch.
_FLEX_MACHS_HIST: frozenset = frozenset(
    m for m, v in _MACHINE_ALLOWED_INCHES.items() if len(v) >= 2)
# Fixed machines under the lock (single historical inch) — candidates for the escape.
_FIXED_MACHS_HIST: frozenset = frozenset(
    m for m, v in _MACHINE_ALLOWED_INCHES.items() if len(v) == 1)
# Lever B (FIXED_ESCAPE, default OFF): a fixed machine may take <= FIXED_ESCAPE_MAX_COS
# diff-size CO, only after its own inch demand is done, to a DB-allowable scarce inch.
_FIXED_ESCAPE_ENABLED = bool(getattr(_bc_cfg, "FIXED_ESCAPE_ENABLED", False))
_FIXED_ESCAPE_MAX_COS = int(getattr(_bc_cfg, "FIXED_ESCAPE_MAX_COS", 1))

# ── ONE-WAY inch model (Step-1 building-unlock, env ONEWAY_INCH, default OFF = hist-lock bit-for-bit) ──
# Replaces the FIXED historical allowed-inch SET with: start at the historical DOMINANT inch, then take
# a diff-size CO to ANY DB-allowable inch — but (a) ONE-WAY: never return to an inch the machine has
# already left (machine_used_inches), and (b) at most a ±2 jump per CO (so NO direct +3/-3 for ANY
# machine — a +3 must go via a two-hop through the intermediate inch; this satisfies "BJ no +3 ever"
# and "VMI no direct +3"). STAGE-2 is EXEMPT (free diff-COs, may revisit, max util). Unlocks the idle
# capacity stranded on met inches (the +6,392 re-lock finding) while the one-way + ≤2 discipline
# prevents the inch-wander that regressed the unconstrained unlock. ONEWAY_INCH=0 → current lock.
_ONEWAY_INCH_ENABLED = os.environ.get("ONEWAY_INCH", "0") != "0"
_ONEWAY_MAX_JUMP = int(os.environ.get("ONEWAY_MAX_JUMP", "2"))   # max inch distance per single diff-CO
# ── GENERAL one-way inch rule (env ONEWAY_INCH_GENERAL, default ON) ─────────────────────────────
# Every building machine starts on its DOMINANT/historical inch and builds it until that inch's
# servable demand is DONE; then it takes exactly ONE diff-CO to the NEEDIEST OTHER inch it is
# CT-allowable for (from latest_building_CT.xlsx) and STAYS there (one-way, no revert). Subsumes the
# per-machine 7003/7501/7002 switches. Enforced purely by flipping machine_locked_inches (the
# _inch_gate choke point); eligibility comes from the allowable matrix. =0 → hist-lock (fixed) sets.
_ONEWAY_INCH_GENERAL = os.environ.get("ONEWAY_INCH_GENERAL", "0") != "0"   # default OFF — MEASURED WORSE (July −33k): forcing dominant-only+1 switch beats the existing flexible hist-lock only when inches finish early, but they are curing-limited
_ONEWAY_GEN_DONE_EPS = float(os.environ.get("ONEWAY_GEN_DONE_EPS", "1.0"))  # inch "done" when remaining ≤ this
# KEEP_DOMINANT: a machine may ALWAYS return to its historical dominant inch (bread-and-butter);
# the no-revert rule applies only to SECONDARY excursions. Pure one-way (=0) strands machines off
# their main inch and regressed −36k..−58k/month, so this defaults ON.
_ONEWAY_KEEP_DOMINANT = os.environ.get("ONEWAY_KEEP_DOMINANT", "1") != "0"

# Delivery-date / priority-flag committed-delivery SKUs (DELIVERY_PRIORITY). Master
# toggle mirrors bc_config; the per-run active flag is DELIVERY_PRIORITY_ENABLED AND
# a non-empty priority_deadline_map (built from the demand file). When inactive every
# priority insertion collapses to identity → bit-for-bit baseline. See the feature
# block in bc_config.py and _build_priority_deadline_map below. Env DELIVERY_PRIORITY=0
# forces OFF everywhere.
# _DELIVERY_PRIORITY_ENABLED is now DERIVED (merged): the feature is enabled iff EITHER mode
# toggle is on. Defined just below, after both are read. (env DELIVERY_PRIORITY=0 forces OFF.)
# PRIORITY-FLAG month-end rule-relaxation (bc_config.PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED,
# default OFF): governs FLAG-only (UNDATED) SKUs → commit to MONTH-END and relax soft rules ONLY
# WHEN AT RISK of missing month-end. Also the commitment scope for undated flags (old
# old undated→month-end commitment semantics, renamed).
_PRIORITY_FLAG_MONTHEND_RELAX = bool(
    getattr(_bc_cfg, "PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED", False))
# Building-side committed-delivery boost sub-toggle (bisection A/B; default ON). DP_BLD=0
# drops the _bld_prio building boost + Phase-C risk-gate bypass, leaving only the Phase-0
# CO levers. Used only when priority_deadline_map is non-empty.
_DP_BLD = os.environ.get("DP_BLD", "1") != "0"
# DELIVERY-DATE rule-relaxation (bc_config.DELIVERY_DATE_ALL_SOFT_RULES_RELAXED, env DELIVERY_RELAX,
# default ON): governs DATED SKUs → relax ALL soft building rules on their ALLOWABLE machines
# (inch-lock, dwell, SAME_GROUP, CO-cost + CO-per-shift-cap guards, 4-SKU/day cap, min-campaign,
# Phase-C risk gate) + machine preemption ONLY WHEN AT RISK of missing the date (behind the linear
# pace). Keeps allowable/tooling + demand cap + mould feasibility. See _delivery_relax().
_DELIVERY_DATE_RELAX = (os.environ.get(
    "DELIVERY_RELAX", "1" if bool(getattr(_bc_cfg, "DELIVERY_DATE_ALL_SOFT_RULES_RELAXED", True)) else "0") != "0")
# The two relaxation MODES are mutually exclusive (re-checked here, so a cloud CLOUD_CONFIG that
# sets both — applied before this import — is caught too, not only the bc_config import-time check).
if _DELIVERY_DATE_RELAX and _PRIORITY_FLAG_MONTHEND_RELAX:
    raise ValueError(
        "DELIVERY_DATE_ALL_SOFT_RULES_RELAXED and PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED "
        "are MUTUALLY EXCLUSIVE — enable only ONE (check bc_config / main.CLOUD_CONFIG / env DELIVERY_RELAX).")
_ANY_DELIVERY_RELAX = _DELIVERY_DATE_RELAX or _PRIORITY_FLAG_MONTHEND_RELAX
# Merged master enable: feature ON iff a mode toggle is ON (env DELIVERY_PRIORITY=0 forces OFF).
_DELIVERY_PRIORITY_ENABLED = _ANY_DELIVERY_RELAX and (os.environ.get("DELIVERY_PRIORITY", "1") != "0")
# Committed SKUs that carry a real Delivery Date (vs flag-only month-end). Populated once per run
# by run_rolling_pipeline (mutated in place, so _assign_building_shift's _delivery_relax reads the
# current set without threading it through every call site). DATE-mode vs MONTHEND-mode gating.
_PRIO_DATED_SKUS: set = set()
# Continuous cycle/unit carry across shifts (PLANT RULE, permanent ON) — the fractional cure
# cycle / build unit is NOT reset at a shift boundary; it carries to the next shift while the
# press/machine keeps running the same SKU, so a press cures the continuous cumulative
# floor(Σtime/ct)·2 (e.g. LSTL0 ct 30.3 → 32/shift avg, the 30,32,32,… pattern), not per-shift
# floor(480/ct)·2 = 30. A cycle that straddles a shift boundary is emitted as a small completion
# row + the shift's main row; those are CONSOLIDATED per (press,date,shift,SKU) below so each
# shift shows ONE row, and the per-shift feasibility rules (R17/R11C) validate the press's REAL
# production (one press = 2 mould cavities/shift, one straddling cycle allowed), not per display row.
_BLD_CYCLE_CARRY  = os.environ.get("BLD_CYCLE_CARRY", "1") != "0"
_CURE_CYCLE_CARRY = os.environ.get("CURE_CYCLE_CARRY", "1") != "0"
# CONTINUOUS BUILD CARRY (PLANT RULE — permanent): per-machine fractional UNITS carried from the
# prior shift's flat-out continuation of the SAME SKU. Building is continuous, so a machine that
# ran time-saturated on one SKU does NOT reset its fractional unit-time each shift — the leftover
# finishes early next shift. {machine: (sku, fractional_units)}; carry-in used only if the machine
# continues that SKU; carry-out set only when the continuation was TIME-bound (spent the whole
# shift). Reset per run by run_rolling_pipeline. Analogous to the curing _cure_carry_min.
_BLD_CARRY_UNITS: dict = {}
# Lever A (FLEX_SCARCE_INCH, ADOPTED — default ON): among a FLEXIBLE machine's allowed
# inches, prefer the SCARCEST (biggest live curing-draw shortfall) over same-inch
# stickiness, so flex capacity feeds about-to-starve scarce inches (15"/13") the fixed
# machines can't reach. Fixed machines unaffected. Measured under the historical lock,
# 3 months, deterministic, mould-audit PASS, demand-cap 0-over: June +570, July +1,860,
# August +1,374 = +3,804 net cured, starvation down/flat every month, no regression.
# Env FLEX_SCARCE_INCH=0 → dominant/same-inch ranking, bit-for-bit.
_FLEX_SCARCE_INCH = (os.environ.get("FLEX_SCARCE_INCH", "1") != "0")

# #2 Lever A → marginal coverage-value + hysteresis (FLEX_MCV, default OFF = Lever A bit-for-bit).
# Rank a flexible machine's allowed inches by VALUE = w_now·(this-shift draw shortfall) +
# w_mon·(cumulative monthly unmet units) — blends immediate starvation with chronic behind-ness.
# HYSTERESIS: a flex machine leaves its CURRENT inch only if another inch's value exceeds it by
# Δ_switch = HYS_BAND·val_scale + CO_LAMBDA·(diff-CO forgone production) — a symmetric dead-band
# that forbids A→B→A oscillation on near-ties while still reacting to a meaningful gain; plus a
# COOLDOWN (no diff-inch switch within N days of the last one). See _key.
# Defaults = the adopted "config5" (hysteresis-only): the monthly-gap blend (W_MON>0) regressed
# on the sweep, so it is OFF; a light dead-band (band 0.1, λ 0.5) + 1-day cooldown is the winner
# (June +272 / July −780 / Aug +1,330 = net +822 vs plain Lever A). FLEX_MCV default OFF here;
# flipped ON at the merge step.
_FLEX_MCV_ENABLED   = (os.environ.get("FLEX_MCV", "1") != "0")   # ADOPTED (config5, net +822)
_FLEX_MCV_W_NOW     = float(os.environ.get("FLEX_MCV_W_NOW", "1.0"))
_FLEX_MCV_W_MON     = float(os.environ.get("FLEX_MCV_W_MON", "0.0"))
_FLEX_MCV_HYS_BAND  = float(os.environ.get("FLEX_MCV_HYS_BAND", "0.1"))
_FLEX_MCV_CO_LAMBDA = float(os.environ.get("FLEX_MCV_CO_LAMBDA", "0.5"))
_FLEX_MCV_COOLDOWN  = int(os.environ.get("FLEX_MCV_COOLDOWN", "1"))
_HYST_BIG           = 10_000   # rank bump that deprioritizes a marginal/cooldown-blocked switch

# ── Dynamic GT buffer (Phase 1, DYN_BUFFER) ───────────────────────────────────
# Replaces the flat per-group GT buffer with a per-SKU, per-shift horizon H_s
# (see bc_config DYN_BUFFER block). Default OFF → the flat buffer, bit-for-bit.
# Env DYN_BUFFER=1 force-enables for A/B; DYN_BUFFER=0 disables.
_DYN_BUFFER_ENABLED = os.environ.get(
    "DYN_BUFFER",
    "1" if getattr(_bc_cfg, "DYN_BUFFER_ENABLED", False) else "0"
) != "0"
# Adaptive curing CO on sustained starvation (unified delay+switch, default OFF).
# See bc_config CURING_ADAPT_CO block. Env CURING_ADAPT_CO=1 force-enables;
# CURING_STARV_SWITCH_SHIFTS overrides the consecutive-zero-GT threshold N.
_CURING_ADAPT_CO = os.environ.get(
    "CURING_ADAPT_CO",
    "1" if getattr(_bc_cfg, "CURING_ADAPT_CO_ENABLED", False) else "0"
) != "0"
# B-1 point 2: under pure-reactive there is no planned base, so the 8-shift starvation
# CO (a press starved N=_CURING_STARV_SWITCH_SHIFTS shifts switches to a feedable SKU) is
# wanted. Force it ON under _REACTIVE_ONLY (env CURING_ADAPT_CO=0 still overrides to off).
if _REACTIVE_ONLY and "CURING_ADAPT_CO" not in os.environ:
    _CURING_ADAPT_CO = True
_CURING_STARV_SWITCH_SHIFTS = int(os.environ.get(
    "CURING_STARV_SWITCH_SHIFTS",
    str(getattr(_bc_cfg, "CURING_STARV_SWITCH_SHIFTS", 8))))
# Refinement: only SWITCH a press off a SKU that is GENUINELY building-limited
# (building can't sustain its curing draw). Without this, a press transiently
# starved because the dynamic buffer is front-loading OTHER SKUs gets switched off
# a recoverable SKU (the buffer↔CO conflict). Default OFF preserves the standalone
# N=8 result; ON is the fix to test alongside the dynamic buffer.
_CURING_ADAPT_FEED_GUARD = os.environ.get(
    "CURING_ADAPT_FEED_GUARD",
    "1" if getattr(_bc_cfg, "CURING_ADAPT_FEED_GUARD_ENABLED", False) else "0"
) != "0"
# Point 1: the supply-aware switch REQUIRES the feed guard (else the fast 2-shift trigger
# would fire on transient dips). Force it on under _REACTIVE_ONLY unless explicitly disabled.
if _REACTIVE_ONLY and "CURING_ADAPT_FEED_GUARD" not in os.environ:
    _CURING_ADAPT_FEED_GUARD = True
DYN_BUF_FLOOR_VMI   = int(os.environ.get("DYN_BUF_FLOOR_VMI",
                          str(getattr(_bc_cfg, "DYN_BUF_FLOOR_VMI", 2))))
DYN_BUF_FLOOR_OTHER = int(os.environ.get("DYN_BUF_FLOOR_OTHER",
                          str(getattr(_bc_cfg, "DYN_BUF_FLOOR_OTHER", 1))))
DYN_BUF_ALPHA       = float(os.environ.get("DYN_BUF_ALPHA",
                            str(getattr(_bc_cfg, "DYN_BUF_ALPHA", 0.5))))
DYN_BUF_BETA        = float(os.environ.get("DYN_BUF_BETA",
                            str(getattr(_bc_cfg, "DYN_BUF_BETA", 1.0))))
DYN_BUF_CURE_CREDIT = float(os.environ.get("DYN_BUF_CURE_CREDIT",
                            str(getattr(_bc_cfg, "DYN_BUF_CURE_CREDIT", 1.0))))
# Global scored assignment (Phase 2, GLOBAL_SCORE_V2). See bc_config block. Default
# OFF → committed _key ranking bit-for-bit. Env GLOBAL_SCORE_V2=1 force-enables.
_GLOBAL_SCORE_V2 = os.environ.get(
    "GLOBAL_SCORE_V2",
    "1" if getattr(_bc_cfg, "GLOBAL_SCORE_V2", False) else "0") != "0"
GS_W_DEF    = float(os.environ.get("GS_W_DEF",    str(getattr(_bc_cfg, "GS_W_DEF",    1.0))))
GS_W_STARV  = float(os.environ.get("GS_W_STARV",  str(getattr(_bc_cfg, "GS_W_STARV",  1.0))))
GS_W_GAP    = float(os.environ.get("GS_W_GAP",    str(getattr(_bc_cfg, "GS_W_GAP",    0.5))))
GS_W_SCARCE = float(os.environ.get("GS_W_SCARCE", str(getattr(_bc_cfg, "GS_W_SCARCE", 1.0))))
GS_W_CO     = float(os.environ.get("GS_W_CO",     str(getattr(_bc_cfg, "GS_W_CO",     0.75))))
GS_W_INCH   = float(os.environ.get("GS_W_INCH",   str(getattr(_bc_cfg, "GS_W_INCH",   0.5))))
GS_W_OVER   = float(os.environ.get("GS_W_OVER",   str(getattr(_bc_cfg, "GS_W_OVER",   1.0))))
GS_INCH_OFFBAND = int(os.environ.get("GS_INCH_OFFBAND", str(getattr(_bc_cfg, "GS_INCH_OFFBAND", 5))))

# Machines removed from _HARD (soft-locked): serve non-dominant inch ONLY when primary demand done.
# All other machines (BJ etc.) were never hard-locked; their Campaign 2+ inch freedom is unchanged.
_SOFT_LOCK_MACHINES: frozenset[str] = frozenset({"7001", "7003"})

# ── Inch-flexibility (_INCH_FLEX_ENABLED) ──────────────────────────────────────
# Generalizes the 7001/7003 soft-lock to the remaining HARD-locked machines so
# they don't sit idle once their own inch is met — plant data proves these run
# 4-7 inches. When off (empty flex set) the scheduler runs bit-for-bit baseline.
# UNI_NARROW (7501-7503) is a separate sub-toggle: CLAUDE.md says they can't
# physically run 14"+, so it's independently back-out-able.
_INCH_FLEX_ENABLED            = True   # master
_INCH_FLEX_INCLUDE_UNI_NARROW = False   # add 7501-7503 (trust DB allowable)
# Off-inch candidate ordering (tiebreak WITHIN the off-inch bucket only — never
# overrides inch preference). "starving_first" = feed urgent starving presses
# first; "demand_first" = largest remaining demand first (best amortization of
# the expensive diff-inch CO). Tested both ways — see plan verification.
_INCH_FLEX_OFFINCH_ORDER      = "starving_first"
_INCH_FLEX_EXTRA_COS          = int(os.environ.get("INCH_FLEX_EXTRA_COS", "2"))   # env raises off-inch CO budget for flex machines

# ── CLIENT INCH RULES (hard plant rules on building inch movement) ────────────
# Rule 1 — one-way inch movement. A machine may take a diff_size_CO only when the
#          demand it can still serve at its CURRENT inch is finished, and it may
#          NEVER return to an inch it has already left (14->15->14 illegal;
#          14->15->13 legal, 13 was never used).
# Rule 2 — +/-2 band. The machine's inch must stay within anchor +/- 2 for the
#          whole month (anchor 14" => 12".."16"; 14"->17" illegal).
#
# Anchor = the inch of the machine's FIRST assignment. There is no Day-0 building
# state to anchor on (machine_current_sku starts empty; TBMStage1/2_ProductionEventData
# are both empty), so the first SKU the scheduler assigns fixes the band.
#
# These are RESTRICTIONS: they cut expensive diff-size COs (freeing production
# minutes) but remove flexibility (some SKUs become unreachable). Net KPI effect
# is measured, not assumed. OFF reproduces the previous behaviour bit-for-bit.
# DEFAULT OFF — the inch rules are a SEPARATE, parked, not-yet-approved plan. They
# must not be live for the mould baseline (they dropped July ~90%→79%). Turn on
# explicitly with INCH_RULES=1 only when working the inch plan.
_INCH_RULES_ENABLED      = True
_INCH_DBG                = [0, 0, 0]   # INCH_DEBUG: [deficit-done, dwell-pass, dwell-BLOCK]
# Plant rule: max MAX_BUILDING_SKUS_PER_DAY (4) distinct SKUs per building machine per day
# (carryover counts as #1; both same/diff-size COs count). Default ON; BLD_SKU_CAP=0 off.
_BLD_SKU_CAP_ENABLED     = True
# EXPERIMENT: one-time +3/-3 inch escape per machine per month, at an 8h building CO
# (INCH_PLUS3_CO_MINS). Default OFF; INCH_PLUS3=1 to test. Only fires for a stranded
# machine (no ±2 in-band work left) with real +3/-3 demand and enough days to amortise.
_INCH_PLUS3_ENABLED      = os.environ.get("INCH_PLUS3", "0") != "0"
_P3DBG                   = os.environ.get("INCH_PLUS3_DEBUG") is not None
_PLUS3_DBG               = [0, 0, 0, 0]   # [has-room, stranded, dwell-ok, has-±3-SKU]
_INCH_BAND_WIDTH         = int(os.environ.get("INCH_BAND", "2"))   # Rule 2: anchor +/- N
# Stepwise inch-DRIFT (Phase-0a): bounded ±1-step relaxation of the hist-lock for a stranded
# idle machine, DB-certified only, cumulative cap _INCH_STEP_MAX, no direct ±3. Default OFF =
# today bit-for-bit. See bc_config.INCH_STEP_DRIFT_ENABLED / the drift block after Phase-B2.
_INCH_STEP_DRIFT         = (os.environ.get("INCH_STEP_DRIFT",
                            "1" if bool(getattr(_bc_cfg, "INCH_STEP_DRIFT_ENABLED", False)) else "0") != "0")
_INCH_STEP_MAX           = int(os.environ.get("INCH_STEP_MAX",
                               str(getattr(_bc_cfg, "INCH_STEP_MAX", 2))))
_STEP_DBG                = [0, 0, 0]   # [idle+room, stranded, drifted]
# Lookahead buffer (Phase-1a): size _dyn_H + the forward-buffer risk gate to the ANTICIPATED peak
# draw (running + incoming-CO presses today) instead of only the current shift's draw. Default OFF
# = today bit-for-bit. See bc_config.LOOKAHEAD_BUF_ENABLED.
_LOOKAHEAD_BUF           = (os.environ.get("LOOKAHEAD_BUF",
                            "1" if bool(getattr(_bc_cfg, "LOOKAHEAD_BUF_ENABLED", False)) else "0") != "0")
# Variant A (True): the +/-2 band REPLACES the _HARD dominant-inch locks.
# Variant B (False): keep _HARD as well, so the machine is bound by the
# intersection (most restrictive). Chosen by measurement — see plan.
_INCH_BAND_REPLACES_HARD = os.environ.get("INCH_BAND_REPLACES_HARD", "1") != "0"
# Keep the opportunistic forward buffer (Phase C) on the machine's CURRENT inch:
# under one-way movement an inch change is irreversible, so it should be spent on
# real demand (Phase B), not on speculative pre-building.
# DEFAULT OFF (Lever C): with the 5-day dwell + revisits an inch change is NO LONGER a
# one-time door, so the idle-machine forward-buffer may pre-build any IN-BAND inch (still
# gated by the ±2 band + 5-day leave gate + starvation gate). Current-inch is preferred
# first via the Phase-C sort key. INCH_PHASEC_SAME=1 restores the old same-inch-only lock.
_INCH_RULES_PHASE_C_SAME_INCH = False
# Rule 1a (never re-use an inch the machine has left) as its own sub-toggle, so
# the cost of the one-way rule can be measured separately from the +/-2 band.
_INCH_NO_REVISIT = os.environ.get("INCH_NO_REVISIT", "1") != "0"
# Treat a sub-campaign leftover deficit as "inch finished" so a pinned machine may leave
# early instead of idling to day 5 (Lever B). DEFAULT ON. The earlier "made it worse"
# result (May -13,197 / June -8,550 / July -28,304) was measured under PERMANENT one-way
# movement, where an easier exit burned the machine's limited inches forever. That no
# longer applies: with the 5-day dwell + revisits allowed, leaving is not permanent, so a
# nearly-exhausted inch should be releasable. INCH_GATE_THRESH=0 restores the strict gate.
_INCH_GATE_CAMPAIGN_THRESHOLD = True

# ── STRICT INCH RULES (experiments, default OFF) ─────────────────────────────
# Rule 1 — STRICT inch dwell (env INCH_STRICT). Redefines "inch demand done" in
# _inch_demand_done from the MOMENTARY buffer-filled deficit to FULL remaining-demand
# exhaustion: a machine may only leave its inch early when every same-inch SKU it can
# serve has demand_remaining - projected_gt <= 0 (and the Lever-B sub-campaign shortcut
# is disabled). The 5-day dwell early-exit and the +/-2 band are unchanged. Removes the
# temporary off-inch excursions (7002's 14->15->16 hop) — a machine stays single-inch
# through each dwell window, building the SAME inch ahead (or idling) when buffer-full.
# INCH_STRICT=0 reproduces the pre-adoption lenient behaviour bit-for-bit.
# ADOPTED default ON — this is the plant rule ("a machine changes inch only once per 5 days
# OR when its current inch's demand is complete"). It removes the JIT inch-hopping churn
# (diff-size COs fall from ~293-354 toward the rule's natural level). Paired with the anchor
# allocation below to recover the 13"/15" coverage. NEW baseline KPIs (cap=12): see §KPIs.
_INCH_STRICT = True

# Cooldown variant of the leave rule (env INCH_COOLDOWN, default OFF). Instead of "dwell ≥5 days
# on the CURRENT inch before leaving" (which measures the clock from arrival and can strand a
# machine that arrived recently), the machine may change inch whenever its LAST diff-size CO was
# ≥ INCH_COOLDOWN_DAYS ago — i.e. "one diff-size CO per machine per 5 days" — with the SAME
# demand-complete immediate-switch override. Because the frequency cap forbids a second diff-CO
# within the window, a machine also cannot RETURN to a size it left for < that many days (re-entry
# is naturally allowed only after the cooldown). The ±2 band and 4-SKU/day cap are unchanged.
# INCH_COOLDOWN=0 → the strict dwell-from-arrival rule above (bit-for-bit).
_INCH_COOLDOWN_RULE = os.environ.get("INCH_COOLDOWN", "0") != "0"
_INCH_COOLDOWN_DAYS = int(os.environ.get("INCH_COOLDOWN_DAYS", str(MIN_INCH_DWELL_DAYS)))

# ── Part 1: single-inch-majority locked inch-set (env LOCK_INCH_SET, default OFF) ──
# The plant runs MOST machines on one inch (page 4 of the report); our ±2 band lets a machine
# wander 3-4 inches (page 5). This assigns each GT machine (Unistage + Stage-2) a small locked
# inch-SET via a MATHEMATICAL minimum-assignment covering optimisation: minimise the total
# (machine,inch) assignments subject to per-inch capacity covering demand — so most machines
# get exactly ONE inch and only where an inch's demand cannot be covered by whole machines does
# a machine take a 2nd/3rd inch. The split EMERGES from the demand (no preset count), and it is
# HYBRID-seeded from the plant building-running-machine data: a running machine whose real size
# has demand this month is PINNED to that size (no wasted Day-1 CO); idle / demand-mismatched
# machines are free and routed to residual demand. At runtime a machine may only build/CO to an
# inch in its locked set → single-inch machines do ZERO diff-size COs by construction.
# LOCK_INCH_SET=0 → current ±2-band behaviour (bit-for-bit). Pairs with PLANT_SEED for the
# running-data pins; without running data (June/July) it degrades to pure demand-driven sets.
_LOCK_INCH_SET = os.environ.get("LOCK_INCH_SET", "0") != "0"
# demand floor (fraction of one machine's monthly capacity) for pinning a running machine to its
# real size — below this the size is "no real demand" and the machine becomes free.
_LOCK_PIN_DEMAND_FRAC = float(os.environ.get("LOCK_PIN_FRAC", "0.34"))  # ~5 days of a machine

# ── Part 2: JIT inch-switch rule (env JIT_INCH, default OFF) ──────────────────────
# The plant switches size JIT (median dwell ~7 h) — 95% of its size changes are excursions our
# 5-day dwell blocks. This DROPS the dwell/cooldown and instead lets a machine change inch the
# moment a curing press needs it, controlled by two demand-adaptive limiters so it does NOT
# bounce to 293 diff-COs: (1) an URGENCY MARGIN — a diff-inch switch fires only when the target
# inch's aggregate curing-draw deficit exceeds the CURRENT inch's residual deficit by
# JIT_URGENCY_MARGIN (so a machine won't abandon a size that still needs it for a marginally
# worse one — kills A→B→A thrash, self-adapts to any month's mix); and (2) a per-machine per-day
# diff-CO BUDGET (hard backstop). Amortization: the target inch must have ≥ DIFF_CO_MIN_TARGET_UNITS
# of sustained remaining demand to pay back the 88-180 min CO. JIT_INCH=0 → current dwell rule
# (bit-for-bit). The plant-like single-inch CONCENTRATION emerges because JIT excursions are
# short (no 5-day campaign), so a machine's dominant inch stays dominant.
_JIT_INCH = os.environ.get("JIT_INCH", "1") != "0"   # default ON (bit-for-bit); =0 → dwell/strict-leave gate
_JIT_URGENCY_MARGIN = int(os.environ.get("JIT_URGENCY_MARGIN", "150"))  # units; hysteresis
_MAX_DIFF_CO_PER_MACHINE_PER_DAY = int(os.environ.get("MAX_DIFF_CO_PER_DAY", "2"))

# Part 1: curing CO same-inch alignment (env CURING_INCH_ALIGN, default OFF). When on, b2c_pipeline
# computes sku_inch BEFORE the Phase-0 curing scheduler and passes it in, so a press changing over
# prefers a same-inch target — keeping each press on one inch so building feeds it without
# different-size COs (mirrors the toggle of the same name in curing_consumption_dynamic.py).
_CURING_INCH_ALIGN = os.environ.get("CURING_INCH_ALIGN", "1") != "0"  # ADOPTED co-plan config (July): ON

# Part 2: per-group building inch policy (env GROUP_INCH_POLICY, default OFF). Matches the plant's
# per-group changeover pattern (report p.4): BJ + UNISTAGE(US) machines are HARD single-inch
# (0 different-size CO) — locked to their hybrid/real anchor inch via machine_locked_inches;
# VMI is left on JIT but TIGHTER (see VMI_JIT_MARGIN below) so it does few diff-COs (the plant's
# ~3 ±3 jumps); Stage-2 stays fully flexible (JIT). Paired with CURING_INCH_ALIGN so the locked
# BJ/US presses are drawn on-inch and never starve.
_GROUP_INCH_POLICY = os.environ.get("GROUP_INCH_POLICY", "1") != "0"  # ADOPTED co-plan config (July): ON
# VMI-specific JIT tightness (only applied under GROUP_INCH_POLICY): a higher urgency margin and
# lower per-day diff-CO budget than the global JIT, so VMI churns less. Default = global values.
# ADOPTED July co-plan config: 250 / 1 (the "Balanced" operating point → 680,106 cured, VMI 27 diff-COs).
# For May/June use 300 / 1 (env VMI_JIT_MARGIN=300).
_VMI_JIT_MARGIN = int(os.environ.get("VMI_JIT_MARGIN", str(VMI_JIT_MARGIN)))              # bc_config single source
_VMI_MAX_DIFF_CO_PER_DAY = int(os.environ.get("VMI_MAX_DIFF_CO_PER_DAY", str(VMI_MAX_DIFF_CO_PER_DAY)))

# ── Part A: dynamic hybrid initial allocation (env INIT_HYBRID, default OFF) ───────
# Seeds each GT machine's ANCHOR inch (which fixes its ±2 band) from the plant building-running
# snapshot, but DEMAND-ADAPTIVELY: a running machine whose real size has demand THIS MONTH is
# pinned to it (no wasted Day-1 CO); a machine whose real size lacks demand (or is idle) is
# reassigned to the neediest reachable inch by the demand-weighted greedy. Fully dynamic per
# demand file — the same May running data feeds June/July, where machines whose May size does not
# fit the month's demand are re-anchored (this also verifies the dynamic re-allocation). Sets only
# the anchor (soft), NOT a hard lock, so JIT can still flex within ±2. INIT_HYBRID=0 → current
# (raw running-seed or INCH_ANCHOR_OPT).
_INIT_ALLOC_HYBRID = True

# Rule 2 — Stage-1 single inch/month (env S1_SINGLE_INCH). Each Stage-1 carcass machine is
# locked to ONE inch for the whole month (zero different-size CO). The inch is pre-assigned
# day-0 by the demand-optimal solver (Stage-2 carcass demand per inch), and Step-3b
# eligibility is tightened to an EXACT inch match. ADOPTED default ON — measured KPI-neutral
# (Stage-1 is carcass/utilization only; 0 cured-tyre cost, 0 rule violations, mould-audit
# PASS all 3 months). S1_SINGLE_INCH=0 restores the band-only behaviour.
_STAGE1_SINGLE_INCH = True

# Correct Stage-1 carcass schedule (env STAGE1_CARCASS_PASS, default ON). Replaces the
# tracking-only Step-3b carcass rows — which UNDERCOUNT because they log one Stage-1 machine
# per SKU per shift (July: 145k recorded vs 180k Stage-2 GT) — with a full 1:1 carcass
# allocation computed by an exact time-windowed max-flow AFTER the plan is built: carcass a
# machine makes in shift τ may feed Stage-2 in shifts τ..τ+STAGE1_CARCASS_LEAD (a 1-2 shift
# pre-build, ≤1-day aging). Emits a FEASIBILITY flag if any shift/inch cannot be fed (e.g. an
# inch with Stage-2 demand but no eligible Stage-1 machine) — so an infeasible plan is caught
# instead of silently assumed. Carcass does NOT gate GT, so this is correct utilization/qty/
# time accounting only — ZERO effect on cured/starvation. STAGE1_CARCASS_PASS=0 restores the
# old tracking rows. STAGE1_CARCASS_LEAD (default 2) = max shifts of carcass pre-build.
_STAGE1_CARCASS_PASS = os.environ.get("STAGE1_CARCASS_PASS", "1") != "0"
_STAGE1_CARCASS_LEAD = int(os.environ.get("STAGE1_CARCASS_LEAD", "2"))
# #1 Carcass inventory-first: consume opening carcass before Stage-1 builds new (KPI-neutral).
_CARCASS_INV_ENABLED = bool(getattr(_bc_cfg, "CARCASS_INV_ENABLED", False))
# ── Stage-2 carcass GATE (hard constraint) ───────────────────────────────────
# When ON, Stage-2 GT is CAPPED each shift by the feasible same-shift Stage-1
# carcass supply (its eligible, inch-matched Stage-1 machines' capacity, honoring
# one-machine-per-SKU contention): Stage-2 can NEVER build GT that carcass cannot
# back — it WAITS for carcass instead. This enforces invariant #3 as a scheduling
# constraint (Stage-1 built BEFORE Stage-2 committed, carcass rows emitted here),
# not just the post-hoc report. It only ever REDUCES Stage-2 GT (demand cap +
# no-waste-GT stay safe) and can lower cured — the honest, physically-realizable
# plan. OFF (default) reproduces prior behaviour bit-for-bit (Step-3b post-hoc
# tracking runs instead). Env STAGE2_CARCASS_GATE=1 enables. See the gate block in
# the shift loop + bc_config for the committed default.
_STAGE2_CARCASS_GATE = (os.environ.get("STAGE2_CARCASS_GATE",
                        "1" if bool(getattr(_bc_cfg, "STAGE2_CARCASS_GATE_ENABLED", False)) else "0") != "0")
# Carcass PRE-BUILD (aging window): Stage-1 uses residual capacity to bank carcass
# AHEAD (within the ≤1-day shelf) so a Stage-2 burst is backed by carcass built in
# the prior 1-2 shifts — honoring "within 1 day aging". Without it the gate is
# same-shift-only and over-clamps (Stage-2 waits even when yesterday's carcass could
# feed it). ON by default when the gate is on. Env STAGE2_CARCASS_PREBUILD=0 → strict
# same-shift gate (measure only).
_STAGE2_CARCASS_PREBUILD = os.environ.get("STAGE2_CARCASS_PREBUILD", "1") != "0"

# ── HARD min-carcass-build rule (bc_config.MIN_CARCASS_QTY, default 10) ────────────
# No carcass may be BUILT at < MIN per (Stage-1 machine, SKU, shift). When a shift's
# over-production leaves a sub-MIN carcass fragment for a machine, that fragment is NOT
# built (dropped from the bank + the display log); the Stage-2 GT is then made only to the
# carcass that IS available (the gate's _take clamp reduces GT to the reduced bank), and
# curing derives from the reduced GT — so carcass / GT / curing stay in sync (R5 GT≤carcass
# holds). Enforced at the Stage-2 carcass gate, right before the GT clamp. Requires the gate.
# CARCASS_MIN_ENFORCE=0 (or MIN_CARCASS_QTY≤0) → bit-for-bit baseline.
_CARCASS_MIN_ENFORCE = (os.environ.get("CARCASS_MIN_ENFORCE", "1") != "0")
_CARCASS_MIN_QTY = int(os.environ.get(
    "CARCASS_MIN_QTY", str(getattr(_bc_cfg, "MIN_CARCASS_QTY", 10))))

# ── Stage-1 building CHANGEOVER time (STAGE1_CO) ──────────────────────────────
# Charge real building CO on the 15 Stage-1 carcass machines — same_size_CO = 60,
# diff_size_CO = 180 (flat, all 15; already in BUILDING_CO_SAME/DIFF_SIZE via
# _co_cost) — modelled like the 24 GT machines: NO production during the CO block,
# BINDING on capacity. Charged in two places, both guarded by this flag:
#   Site 1 (KPI-binding): the carcass GATE's _gate_build reserves CO minutes when a
#     Stage-1 machine switches carcass SKU (fewer units that shift → Stage-2 clamps).
#   Site 2 (accounting):  the post-plan _stage1_carcass_schedule lays CO blocks into
#     the carcass rows (CO_Mins + honest Stage-1 occupancy).
# Requires _STAGE2_CARCASS_GATE (the binding site lives inside the gate). Default OFF
# → both sites reproduce the current plan bit-for-bit. Env STAGE1_CO=1 enables.
_STAGE1_CO = (os.environ.get("STAGE1_CO",
              "1" if bool(getattr(_bc_cfg, "STAGE1_CO_ENABLED", False)) else "0") != "0")
# End-of-day carcass buffer cap (units held overnight, all SKUs) — analogous to the 8k GT
# cap. Bounds the gate's carcass pre-build so the buffer carried between shifts stays small.
# Applied only when _STAGE1_CO (carcass-realism model); OFF path untouched.
_MAX_EOD_CARCASS = int(os.environ.get("CARCASS_EOD_CAP",
                       str(getattr(_bc_cfg, "MAX_ENDOFDAY_CARCASS_INVENTORY", 1200))))
# Carcass build-to-consumption: pre-build to the Stage-2 BUILD rate (the true carcass consumer)
# instead of max(Stage-2 rate, curing draw) → no aging-out over-build; also drives the carcass-row
# builder to a time-windowed FIFO match (rows track GT consumption, aged-out carcass hidden).
# Default ON; env CARCASS_NO_OVERBUILD=0 reverts to over-build + front-loaded rows bit-for-bit.
_CARCASS_NO_OVERBUILD = (os.environ.get("CARCASS_NO_OVERBUILD",
                         "1" if bool(getattr(_bc_cfg, "CARCASS_NO_OVERBUILD_ENABLED", False)) else "0") != "0")
# Balanced (demand-proportional) Stage-1 inch allocation — spreads the 15 carcass machines
# across inches ∝ carcass demand. MEASURED a no-op on the real data (carcass eligibility is
# inch-lock-filtered, so a machine's Stage-2-carcass inch set is near-forced) → default OFF
# (env S1_BALANCED_INCH=1 to experiment). The real Stage-1 lever is the gate's per-shift
# machine selection, not this static inch assignment.
_S1_BALANCED_INCH = (os.environ.get("S1_BALANCED_INCH", "0") != "0")
# Stage-1 one-way inch flexibility (client rule): a Stage-1 machine may take a DIFFERENT-size
# CO (180 min) to another inch ONLY once its current inch's Stage-2 carcass demand is complete,
# then it continues on the new inch and NEVER reverts. This mobilises the ~48% idle Stage-1
# capacity (a machine on a small/near-done inch moves to a still-demanded one) → fewer Stage-2
# clamps. Default tied to STAGE1_CO; env S1_INCH_FLEX overrides.
_S1_INCH_FLEX = (os.environ.get("S1_INCH_FLEX", "1" if _STAGE1_CO else "0") != "0")
# Stage-1 diff-size CO (separate rule): let an idle/surplus Stage-1 machine take a bounded number of
# diff-size COs toward a TIGHT (not just fully-uncoverable) inch — activates idle Stage-1 to relieve
# local carcass clamps. Bounded (per-machine monthly cap) so COs stay "optimal, not too much"; revisit
# allowed (building may revert). Default OFF = current S1_INCH_FLEX (uncoverable-only, one-way) bit-for-bit.
_S1_DIFF_CO = os.environ.get("S1_DIFF_CO", "0") != "0"
_S1_MAX_DIFF_CO = int(os.environ.get("S1_MAX_DIFF_CO", "3"))      # diff-COs a Stage-1 machine may take/month
_S1_TIGHT_MARGIN = float(os.environ.get("S1_TIGHT_MARGIN", "1.15"))  # fire when coverage < demand×margin
# ── Part A: Stage-1 carcass hardening — A1+A2 REJECTED (env CARCASS_V2, default OFF) ──
# Ported from optimizer/carcass_sched.py. A1 = most-constrained-SKU-first carcass ordering
# (SKUs with the FEWEST eligible Stage-1 machines claim their machines first, before
# flexible SKUs can steal them); A2 = cheapest-CO-first machine ranking in _gate_build
# (continuation 0 / same-inch 60 / diff-inch 180 min — real CO minutes, not a binary flag).
# A3 (successor-aware capacity) is a NO-OP in the forward-only greedy (no backward spill →
# the successor's changeover can never be disturbed), so _gate_build's existing minute
# budget already covers it.
#
# MEASURED + REJECTED (2026-08, cap=12, OFF bit-for-bit, isolated via CV2_A1/CV2_A2):
#   • A2 is a COMPLETE NO-OP on the real data — A2-only == OFF baseline byte-for-byte
#     (the CO-minute ranking selects the same machines as the existing CO-free-first rank).
#   • A1 drives the whole effect and is NET-NEGATIVE: June +2,156 / July +828 / Aug −9,835
#     = net −6,851. Most-constrained-first over-serves low-demand scarce SKUs early and
#     starves the bulk SKUs on the well-supplied month (Aug built −10,345 → fewer Stage-2
#     GT). Helps the constrained months, craters Aug — same shape as the rejected P1 gate.
# Kept OFF, code retained for the record. A4 (global FIFO reconcile — makes R5/R9C pass by
# construction but drops coverage) is a SEPARATE, coverage-dropping decision, not bundled here.
# Effective only when the carcass model is on (_STAGE2_CARCASS_GATE + _STAGE1_CO, both ON
# by default). Default OFF; CARCASS_V2=1 enables A1+A2 (CV2_A1/CV2_A2 isolate each).
_CARCASS_V2 = (os.environ.get("CARCASS_V2", "0") != "0")
# Split sub-flags to isolate A1 vs A2 (both default to CARCASS_V2's state).
_CV2_A1 = _CARCASS_V2 and (os.environ.get("CV2_A1", "1") != "0")   # most-constrained-SKU-first
_CV2_A2 = _CARCASS_V2 and (os.environ.get("CV2_A2", "1") != "0")   # cheapest-CO-first machine rank
# A4 = global per-SKU carcass->Stage-2 FIFO reconcile (1 calendar-day aging): cap Stage-2 GT
# carcass can't back + cascade to cured. MEASURED a NO-OP (2026-08, June/July/Aug bit-for-bit):
# the greedy's in-loop per-shift clamp (_take=min(desired,bank), ~line 7745) ALREADY guarantees
# R5 (Stage-2 <= carcass) cumulatively, so the global FIFO finds NO shortfall to cap → 0 units
# reduced, coverage unchanged. It does NOT fix R9C=9 either — that residual is a RENDERER
# aging-granularity artifact (3-shift window vs calendar-day boundary in _stage1_carcass_rows_co),
# which a Stage-2-reduction cannot touch (an earlier version that also dropped carcass from
# prod_log to chase R9C broke R5 by desyncing from the renderer — reverted). Kept OFF, safe/inert.
# CV2_A4=1 turns it on alone; default = CARCASS_V2 state.
_CV2_A4 = (os.environ.get("CV2_A4", "1" if _CARCASS_V2 else "0") != "0")

# ── Stage-2 campaign consolidation (S2_CAMPAIGN) ──────────────────────────────
# Reduce the churn of the 6 Stage-2 GT machines {8201,8301,8302,8501,8502,7301}
# (217 building COs in July under STAGE1_CO=1, e.g. 8501 = 15 SKUs / 71 COs) by
# forcing LONGER Stage-2 campaigns. Fewer, longer Stage-2 campaigns -> smoother
# per-shift carcass demand -> the FIXED 2000-unit overnight carcass buffer absorbs
# it -> the cured KPI the spiky carcass demand costs is recovered WITHOUT raising the
# cap. STAGE2 GT machines only. The ADOPTED knob is S2_MIN_CAMPAIGN_MINS; the other
# two were measured WORSE on July and default to no-ops:
#   • S2_MIN_CAMPAIGN_MINS (ADOPTED, default 185): a Stage-2 machine may CO to a NEW
#     sku only if the resulting campaign runs >= this long. If every candidate is too
#     short the machine IDLES rather than doing a short churn switch. Counter-intuitively
#     this RAISES Stage-2 GT (freed CO time -> longer productive runs) so long as the
#     threshold stays on the measured stable step: July STAGE1_CO=1, min in [180,190]
#     all give 665,599 (+1,899, Stage-2 COs 217->142); min=200 falls off a cliff (633k),
#     so keep 185 (mid-step). Same-sku continuation (Phase A) is NEVER blocked.
#   • S2_SKU_CAP (default 4 = plant-wide cap = NO-OP): a tighter distinct-SKU-per-day
#     cap. Measured WORSE on July (constrains the SKU mix -> less GT) -> left a no-op.
#   • S2_MAX_CO_PER_DAY (default 0 = disabled): a blunt per-day building-CO budget per
#     Stage-2 machine. Measured WORSE than the min-campaign (budget<=2 idles/collapses).
# All three are STAGE2-only and no-ops when the toggle is OFF, so with the toggle OFF the
# STAGE1_CO=1 plan is reproduced bit-for-bit (verified: 663,700, 217 Stage-2 COs).
_S2_CAMPAIGN = (os.environ.get("S2_CAMPAIGN",
                "1" if bool(getattr(_bc_cfg, "S2_CAMPAIGN_ENABLED", False)) else "0") != "0")
_S2_SKU_CAP = int(os.environ.get("S2_SKU_CAP",
                  str(getattr(_bc_cfg, "S2_SKU_CAP", 4))))
_S2_MIN_CAMPAIGN_MINS = int(os.environ.get("S2_MIN_CAMPAIGN_MINS",
                            str(getattr(_bc_cfg, "S2_MIN_CAMPAIGN_MINS", 185))))
# Optional 3rd knob: a hard per-day building-CO budget on each Stage-2 machine
# (0 = disabled). Once a Stage-2 machine spends its daily COs it can only continue
# its current SKU. A blunter alternative to the min-campaign gate; measured worse
# than the min-campaign on July → left disabled by default (env S2_MAX_CO_PER_DAY).
_S2_MAX_CO_PER_DAY = int(os.environ.get("S2_MAX_CO_PER_DAY",
                         str(getattr(_bc_cfg, "S2_MAX_CO_PER_DAY", 0))))

# Concentration allocation (CONC_ALLOC): defer redundant machines onto already-paced SKUs so
# each SKU runs on ~ceil(draw/rate) machines (fewer machines, longer campaigns, fewer building
# COs). Deferral-only + starvation-override; OFF = current selection bit-for-bit. See bc_config.
_CONCENTRATION = (os.environ.get("CONC_ALLOC",
                  "1" if bool(getattr(_bc_cfg, "CONCENTRATION_ENABLED", False)) else "0") != "0")
_CONC_STARV_SHIFTS = float(os.environ.get("CONC_STARV_SHIFTS",
                           str(getattr(_bc_cfg, "CONC_STARV_SHIFTS", 1.0))))

# IDLE_PRESS_ACTIVATE (env IDLE_PRESS_ACT, default ON): press roster = UNION of the Day-0
# running-moulds snapshot and the allowable matrix (Master_Curing_Allowable_Machines_source,
# cetl.load_allowable_press_ids()). Any allowable press NOT present in the Day-0 snapshot (idle /
# mid-CO / clean at 07:00 on Day 1) is brought online via a cold-start curing CO (nothing -> SKU)
# in Day-1 Shift A, then produces from Day-1 Shift B. Target = neediest allowable SKU with 2 free
# moulds (reuses _pick_retarget). Every Day-0 snapshot press is used as-is (running ⊆ allowable).
# OFF = no cold-start (snapshot presses only), bit-for-bit.
_IDLE_PRESS_ACTIVATE = (os.environ.get("IDLE_PRESS_ACT",
                        "1" if bool(getattr(_bc_cfg, "IDLE_PRESS_ACTIVATE_ENABLED", False)) else "0") != "0")

# FULL_PRESS_ROSTER (env FULL_PRESS_ROSTER, DEFAULT OFF): explicit named gate that unions EVERY
# allowable-matrix press absent from the Day-0 running-moulds snapshot into the sim as a cold-start
# idle press (same mechanism as IDLE_PRESS_ACTIVATE — Day-1 cold-start, fresh mould life, acquires a
# mould-feasible in-demand allowable SKU). For 2026-09 this set is exactly the 8 presses in the
# allowable roster (185) but missing from the snapshot (177): 14802, 14809, 15204, 24809, 24819,
# 4410, 85209, 9803. NOTE: on the ADOPTED config IDLE_PRESS_ACTIVATE is already ON and already
# cold-starts precisely this set, so this toggle's union is IDEMPOTENT there (ON==OFF, bit-for-bit) —
# it exists as an independent lever for runs where IDLE_PRESS_ACTIVATE is off (IDLE_PRESS_ACT=0) so
# the full 185-press roster can still be forced. OFF = no extra presses beyond IDLE_PRESS_ACTIVATE's,
# bit-for-bit.
_FULL_PRESS_ROSTER = (os.environ.get("FULL_PRESS_ROSTER",
                      "1" if bool(getattr(_bc_cfg, "FULL_PRESS_ROSTER_ENABLED", False)) else "0") != "0")

# SUPPLY-AWARE cold-start activation (bc_config.IDLE_PRESS_SUPPLY_AWARE, default ON): a roster
# press absent from the Day-0 snapshot is cold-started ONLY when Building can realistically feed
# its target SKU — otherwise it is DEFERRED and retried on a LATER working day when supply exists
# (activation now runs every working day, not just Day 1). The gate is PER-SKU MARGINAL, not
# inch-level: a prior INCH-level gate (skip when the target inch's committed curing draw exceeds
# `_building_inch_capacity`) over-skipped every inch (committed draw always exceeds building
# capacity — the starvation story) and measured net −20,685, so it was rejected. The per-SKU
# marginal test instead asks "can building supply THIS extra press of THIS SKU" — the target must
# (a) have eligible building machines and (b) not be curing-over-supplied vs building's per-SKU
# feed. Deferral (not permanent skip) is the key difference. IDLE_PRESS_SUPPLY_AWARE=0 reverts to
# the old blind Day-1 activate-all (bit-for-bit).
_IDLE_PRESS_SUPPLY_AWARE = (os.environ.get("IDLE_PRESS_SUPPLY_AWARE",
                            "1" if bool(getattr(_bc_cfg, "IDLE_PRESS_SUPPLY_AWARE", True)) else "0") != "0")
_IDLE_PRESS_SUPPLY_MARGIN = float(os.environ.get("IDLE_PRESS_SUPPLY_MARGIN", "1.0"))

# MONOTONICITY experiment — building-draw per-inch cap (env BLD_DRAW_CAP, default OFF). REJECTED.
# Idea: clamp each inch's aggregate curing-draw signal (shift_cure_demand) passed to the building
# assigner to that inch's building GT/shift capacity, so extra presses on a building-SATURATED inch
# cannot inflate the signal and pull machines off productive inches. MEASURED WORSE: building output
# tracks the draw MAGNITUDE (it produces proportional to the signal), so scaling the signal DOWN
# under-produces the high-demand inches — July WITH-852 679,499→676,126 and even the WITHOUT-852
# baseline 686,201→685,292 (−909). The gap WIDENED (−6,702→−9,166). The non-monotonicity is a
# machine↔inch RE-RANKING/reshuffle (propagates through Stage-1→Stage-2), NOT signal magnitude, so a
# draw-reduction cap is the wrong lever. Kept OFF for the record.
_BLD_DRAW_CAP = os.environ.get("BLD_DRAW_CAP", "0") != "0"

# MONOTONICITY FIX — mould-contention-aware activation gate (env MOULD_CONTENTION_GATE, default
# OFF = bit-for-bit). Root cause of "adding eligible curing presses lowers cured": scarce moulds
# are SHARED across several demanded SKUs (e.g. the 13" QXPC0 pool = 11 moulds shared by 4 SKUs).
# When extra eligible presses (the 852xx additions) mount those shared moulds, they DISPLACE a
# productive press that needed them → the displaced SKU cures less (measured: 1D25215Z13008QXPC0
# eligible presses 7→9 but cured 8,027→5,747). Building then builds less of it (downstream). The
# gate: a mount may take a mould only if it leaves every OTHER mould-sharing DEMANDED SKU at least
# 2 usable moulds (free or owned by a press already running it). If the only available moulds would
# strip a sharing SKU's last pair, the mount is refused → the extra press stays idle instead of
# displacing a productive one → adding presses can never REDUCE cured (monotonic).
_MOULD_CONTENTION_GATE = os.environ.get("MOULD_CONTENTION_GATE", "0") != "0"

# Demand-optimal machine->inch anchor pre-solve for GT machines (env INCH_ANCHOR_OPT).
# Seeds machine_anchor_inch/inch_now/inch_since before the day loop from the shared solver
# instead of letting the day-1 starvation-driven greedy pick each anchor. Ignores running-
# machine state (decision). ADOPTED default ON alongside the strict rule — the demand-weighted
# allocation dedicates machines to the high-demand 13"/15" so those inches don't starve under
# the rule (measured: recovers July +8.4k vs emergent anchors under strict; net +12k/3-months).
# INCH_ANCHOR_OPT=0 restores emergent day-1 anchoring.
_INCH_ANCHOR_OPT = False

# Diff-size-CO amortization gate (env DIFF_CO_GATE, default OFF). A machine may do a
# DIFFERENT-inch CO only if (a) ≥ DIFF_CO_MIN_DWELL_DAYS since its last diff-size CO and
# (b) the target inch has ≥ DIFF_CO_MIN_TARGET_UNITS sustained servable demand for it.
# Blunt per-machine diff-CO frequency cap (env DIFF_CO_GATE, default OFF — SUPERSEDED).
# Measured to only trade KPI for CO reduction (the diff-COs are productive coverage), so it
# is NOT the right lever. The plant rule is instead "diff-CO once per 5 days OR when the inch's
# demand is complete" (= _INCH_STRICT), and the KPI under that rule is an ALLOCATION problem
# (13"/15" under-served) addressed by the anchor allocation. Kept off, for the record.
_DIFF_CO_GATE = os.environ.get("DIFF_CO_GATE", "0") != "0"
# Env overrides for the two gate thresholds (sweep without editing bc_config).
if os.environ.get("DIFF_CO_TARGET"):
    DIFF_CO_MIN_TARGET_UNITS = int(os.environ["DIFF_CO_TARGET"])
if os.environ.get("DIFF_CO_DWELL"):
    DIFF_CO_MIN_DWELL_DAYS = int(os.environ["DIFF_CO_DWELL"])
# Anchor solver: GREEDY by default (env ANCHOR_EXACT=1 → exact MILP). MEASURED: the exact
# MILP maximises a DEGENERATE static coverage objective (once an inch is covered, extra
# machines add nothing → excess machines placed arbitrarily), so it LOSES to the greedy's
# concentrate-on-demand heuristic under strict (3-mo net: greedy 1,954,791 vs exact
# 1,935,916 vs emergent 1,942,695). Greedy kept as default; exact retained, off, for record.
_ANCHOR_EXACT = os.environ.get("ANCHOR_EXACT", "0") != "0"
# Lever 1 — draw-matched anchor allocation (env ANCHOR_PHASED, default OFF pending A/B).
# The default coverage-maximising greedy strands the last (broadest) machines on tiny extreme
# inches (17"/18", over-served 3-60x) while the scarce high-demand inches (15"/13"/16") sit at
# cap/target ~1.11 — no slack, so any CO/mould-clean downtime momentarily starves their presses
# (measured: ~half the strict-rule KPI cost is exactly this reachable starvation). The phased
# solver WATER-FILLS by cap/target ratio instead: it equalises provisioning across DEMANDED
# inches, giving 15"/13"/16" a flex buffer (~1.4-1.6x) funded by the extremes — whose small
# demand is then covered by a band-neighbour (±2), needing no dedicated all-month anchor.
_INCH_ANCHOR_PHASED = os.environ.get("ANCHOR_PHASED", "0") != "0"
# Diagnostic: curing-aware ceiling on the anchor target (env ANCHOR_CURE_CAP, default ON).
# =0 → no curing cap. MEASURED not binding on the tested months (moulds abundant); kept for A/B.
_ANCHOR_CURE_CAP = os.environ.get("ANCHOR_CURE_CAP", "1") != "0"
# counter for the pre-solve: [gt machines assigned, stage1 machines assigned]
_INCH_OPT_DBG = [0, 0]

# Machine groups for the flex brute-force (plant multi-inch ranking: VMI 4.6 >
# BJ 2.4 ~ UNI 2.3 > IRM 1.5 inches/machine).
_FLEX_GROUPS: dict[str, frozenset[str]] = {
    "VMIMAXX": frozenset({"6001", "6002", "6003", "6004", "7002", "7004"}),
    "VMISOFT": frozenset({"7001", "7003"}),
    "BJ":      frozenset({"7101", "7102", "7103", "7104", "7105", "7106", "7201"}),
    "UNI":     frozenset({"7501", "7502", "7503"}),
    "STAGE2":  frozenset({"8201", "8301", "8302", "8501", "8502", "7301"}),
}
_INCH_FLEX_VMIMAXX = _FLEX_GROUPS["VMIMAXX"]
_INCH_FLEX_UNI     = _FLEX_GROUPS["UNI"]

# Brute-force override: env FLEX_SET="VMIMAXX,BJ" (group names) selects the flex
# set for a run without editing code. Falls back to the toggle-based default.
def _resolve_flex_machines() -> frozenset:
    _env = os.environ.get("FLEX_SET")
    if _env is not None:
        out: set = set()
        for tok in _env.split(","):
            tok = tok.strip()
            if tok in _FLEX_GROUPS:
                out |= set(_FLEX_GROUPS[tok])
            elif tok:
                out.add(tok)
        return frozenset(out)
    if not _INCH_FLEX_ENABLED:
        return frozenset()
    # Brute-force winner (plant-guided): the multi-inch groups VMI (VMIMAXX +
    # VMISOFT) + BJ = 670,744 cured / 96.7% / 1,845 starvation / 627 write-off,
    # best on every metric. STAGE2/IRM is catastrophic in the flex set (-57k,
    # carcass-dependency disruption) and UNI hurts (tight 13") — both excluded.
    _sel = _FLEX_GROUPS["VMIMAXX"] | _FLEX_GROUPS["VMISOFT"] | _FLEX_GROUPS["BJ"]
    if _INCH_FLEX_INCLUDE_UNI_NARROW:
        _sel = _sel | _FLEX_GROUPS["UNI"]
    return _sel

_INCH_FLEX_MACHINES: frozenset[str] = _resolve_flex_machines()


def _curing_inch_ceiling(inch_skus: dict, sku_moulds: dict, cure_ct_map: dict,
                         planning_days: int, n_presses: int) -> dict:
    """Per-inch CURING ceiling: how much of an inch can be cured over the horizon, bounded
    by mould-feasible presses. Each running press needs 2 eligible moulds, so an inch's max
    simultaneous presses = min(n_presses, |eligible moulds for its SKUs| // 2) — this is what
    makes 15″/13″ (few moulds) a low ceiling, the real July bottleneck. Returns {inch: units}.
    Empty sku_moulds (mould gate off) → {} (no cap). A ceiling proxy (shared moulds over-count)."""
    if not sku_moulds:
        return {}
    out: dict = {}
    for i, skus in inch_skus.items():
        moulds: set = set()
        cts: list = []
        for s in skus:
            moulds |= set(sku_moulds.get(s, ()) or ())
            cts.append(cure_ct_map.get(s, DEFAULT_CURING_CT))
        presses = min(n_presses, len(moulds) // 2)
        avg_ct = (sum(cts) / len(cts)) if cts else DEFAULT_CURING_CT
        out[i] = presses * _cure_qty_per_shift(avg_ct) * 3 * planning_days
    return out


def _greedy_inch_assignment(machines, elig_inches: dict, cap: dict, inch_demand: dict) -> dict:
    """Deterministic greedy generalized assignment (the MILP fallback). Each machine → its
    highest-remaining-target eligible inch, most-constrained-first; its capacity is subtracted
    from that inch's remaining target so later machines flow to still-under-served inches."""
    remaining = {i: float(d) for i, d in inch_demand.items()}
    result: dict = {}
    order = sorted(machines,
                   key=lambda m: (len(elig_inches.get(m, ()) or ()), -cap.get(m, 0.0), str(m)))
    for m in order:
        opts = sorted(i for i in (elig_inches.get(m, ()) or ()) if i)
        if not opts:
            continue
        best = max(opts, key=lambda i: (remaining.get(i, 0.0), inch_demand.get(i, 0.0)))
        result[m] = best
        remaining[best] = remaining.get(best, 0.0) - cap.get(m, 0.0)
    return result


def _balanced_inch_assignment(machines, elig_inches: dict, cap: dict, inch_demand: dict) -> dict:
    """Stage-1 carcass allocation that spreads machines PROPORTIONAL to per-inch demand,
    instead of the coverage-greedy's "one big machine covers a small inch, strand the next
    on a dead inch". Each machine (most-constrained first, then most capacity) takes the
    eligible DEMANDED inch with the lowest capacity/demand provisioning so far — so every
    machine flows to where per-shift carcass is scarcest, and no machine is parked on a
    near-zero-demand inch while a high-demand inch stays thin. Deterministic; a machine with
    no demanded eligible inch falls back to its lowest eligible inch (never dropped)."""
    dem = {str(i): float(d) for i, d in inch_demand.items() if d and float(d) > 0}
    used: dict = defaultdict(float)                       # capacity already homed per inch
    result: dict = {}
    order = sorted(machines,
                   key=lambda m: (len(elig_inches.get(m, ()) or ()), -cap.get(m, 0.0), str(m)))
    for m in order:
        cm = float(cap.get(m, 0.0))
        opts = sorted(str(i) for i in (elig_inches.get(m, ()) or ()) if str(i) in dem)
        if not opts:                                     # no demanded inch → keep eligible, lowest
            allo = sorted(str(i) for i in (elig_inches.get(m, ()) or ()) if i)
            if allo:
                result[m] = allo[0]
            continue
        # pick the inch that most needs another machine = highest demand/(homed cap + this cap)
        best = max(opts, key=lambda i: (dem[i] / (used[i] + cm + 1.0), dem[i], i))
        result[m] = best
        used[best] += cm
    return result


def _phased_inch_assignment(machines, elig_inches: dict, cap: dict, target: dict) -> dict:
    """Lever 1 — draw-matched anchor allocation (water-filling by cap/target ratio).

    Instead of maximising absolute coverage (which dumps the last broad machines on tiny
    extreme inches they over-serve while the scarce high-demand inches keep zero slack), assign
    each machine to the eligible DEMANDED inch whose cap/target ratio stays LOWEST after adding
    it — equalising provisioning so 15"/13"/16" accumulate a flex buffer against CO/clean
    downtime. A tiny inch (e.g. 18" = 840) has such a high post-add ratio that no machine ever
    homes there; its small demand is served by a ±2 band-neighbour. Deterministic. Any machine
    with no in-target eligible inch falls back to its lowest eligible inch (never dropped)."""
    tgt = {i: max(1.0, float(v)) for i, v in target.items()}
    # Start from the demand-priority base greedy (a good allocation), then BOUNDED-rebalance:
    # move a machine off a grossly over-provisioned inch (ratio > HI) onto THIS MONTH's most
    # under-provisioned inch (lowest ratio, < LO) that the machine is ELIGIBLE to reach. This is
    # month-dynamic (it reads each month's ratios) and reachability-safe: if the starved inch is
    # unreachable by any wasteful machine (e.g. 13" in June/July — the 17"/18" machines can't build
    # it), no move happens and the base allocation is preserved (no regression). Full equalisation
    # was too aggressive; this only ever redistributes genuine waste toward genuine scarcity.
    result = dict(_greedy_inch_assignment(machines, elig_inches, cap, tgt))
    HI    = float(os.environ.get("ANCHOR_HI", "1.8"))      # donor: over-provisioned above this
    LO    = float(os.environ.get("ANCHOR_LO", "1.2"))      # recipient: only feed GENUINELY starved inches (<20% buffer)
    SAFE  = float(os.environ.get("ANCHOR_SAFE", "1.3"))    # a real-demand donor must stay this provisioned after donating
    SMALL = float(os.environ.get("ANCHOR_SMALL", "3000"))  # demand below this = band-coverable (may strip to 0)

    def _assigned():
        a = {i: 0.0 for i in tgt}
        for _m, _i in result.items():
            if _i in a:
                a[_i] += cap.get(_m, 0.0)
        return a

    # Snapshot THIS MONTH's genuinely-starved inches (ratio < LO = under a 20% buffer). Only these
    # trigger a rebalance; an inch merely "not lavish" (e.g. June 15" at 1.30) is left alone. This
    # is what makes the lever a no-op on capacity-tight months whose scarce inch has no reachable
    # surplus to draw from (June/July 13") — nothing regresses — while May's 15", starved next to
    # a grossly wasteful 17"/18", gets drained into.
    a0 = _assigned()
    starved = sorted((i for i in tgt if tgt[i] > 1.0 and a0[i] / tgt[i] < LO),
                     key=lambda i: (a0[i] / tgt[i], int(i) if str(i).isdigit() else 99))
    for R in starved:
        for _ in range(len(machines)):                # bounded per-recipient drain (deterministic)
            a = _assigned()
            rR = a[R] / tgt[R]
            if rR >= HI:                              # recipient now well-buffered → done
                break
            # Donor = machine on a strictly richer, over-provisioned inch (>HI), eligible for R,
            # that stays healthy after donating (new ratio ≥ SAFE) or whose demand is tiny enough
            # for a ±2 band-neighbour to absorb (tgt < SMALL). Pull the most-wasteful first.
            cands = []
            for _m, D in result.items():
                if D == R or R not in (elig_inches.get(_m, ()) or ()):
                    continue
                rD = a[D] / tgt[D]
                if rD <= rR:                          # donor must be richer than the starved inch
                    continue
                newD = (a[D] - cap.get(_m, 0.0)) / tgt[D]
                # Donate only genuine SURPLUS: the donor stays healthy (≥ SAFE) or its demand is
                # tiny enough for a ±2 band-neighbour (tgt < SMALL). This is the guard that keeps
                # capacity-tight months (June/July) untouched — no inch there survives donating.
                if newD >= SAFE or tgt[D] < SMALL:
                    cands.append((-rD, newD, str(_m), _m))
            if not cands:
                break
            cands.sort()
            result[cands[0][3]] = R
    return result


def _optimal_inch_assignment(machines, elig_inches: dict, cap: dict,
                             inch_demand: dict, target: dict | None = None) -> dict:
    """EXACT curing-aware static machine→inch assignment (MILP), greedy fallback.

    Assigns each machine ONE anchor inch to MAXIMISE Σ_i min(target[i], supply[i]), where
    supply[i] = Σ cap[m] over machines anchored at i and target[i] = min(demand, curing
    ceiling) — so machines are never sent to an inch curing cannot absorb. `target` defaults
    to inch_demand (pure demand-optimal). Deterministic; falls back to the greedy on any
    solver failure or partial solution. Returns {machine: inch}.
    """
    tgt = {i: float(v) for i, v in (target if target is not None else inch_demand).items()}
    if _INCH_ANCHOR_PHASED:                  # Lever 1 — draw-matched water-filling
        return _phased_inch_assignment(machines, elig_inches, cap, tgt)
    if not _ANCHOR_EXACT:                    # greedy is the shipped default (beats the MILP)
        return _greedy_inch_assignment(machines, elig_inches, cap, tgt)
    try:
        import numpy as np
        from scipy.optimize import milp, LinearConstraint, Bounds
        ms = sorted(str(m) for m in machines)
        pairs = [(m, i) for m in ms
                 for i in sorted({str(x) for x in (elig_inches.get(m, ()) or ()) if x})]
        if not pairs:
            return {}
        inches = sorted({i for _, i in pairs} | set(tgt.keys()))
        nx, ni = len(pairs), len(inches)
        ipos = {i: k for k, i in enumerate(inches)}
        nvar = nx + ni                                   # x[m,i] binaries + covered[i]
        c = np.zeros(nvar)
        for i in inches:
            c[nx + ipos[i]] = -1.0                       # maximise Σ covered
        A_eq = np.zeros((len(ms), nvar))                 # each machine exactly one anchor
        mrow = {m: r for r, m in enumerate(ms)}
        for k, (pm, _pi) in enumerate(pairs):
            A_eq[mrow[pm], k] = 1.0
        A_ub = np.zeros((ni, nvar))                      # covered[i] − Σ cap·x[m,i] ≤ 0
        for i in inches:
            A_ub[ipos[i], nx + ipos[i]] = 1.0
        for k, (pm, pi) in enumerate(pairs):
            A_ub[ipos[pi], k] = -float(cap.get(pm, 0.0))
        lb = np.zeros(nvar)
        ub = np.ones(nvar)
        for i in inches:
            ub[nx + ipos[i]] = max(0.0, tgt.get(i, 0.0))
        integ = np.zeros(nvar)
        integ[:nx] = 1
        res = milp(c, constraints=[LinearConstraint(A_eq, 1.0, 1.0),
                                   LinearConstraint(A_ub, -np.inf, 0.0)],
                   bounds=Bounds(lb, ub), integrality=integ)
        if not getattr(res, "success", False) or res.x is None:
            raise RuntimeError("milp failed")
        out = {pm: pi for k, (pm, pi) in enumerate(pairs) if res.x[k] > 0.5}
        miss = [m for m in ms if m not in out]
        if miss:                                          # solver quirk → greedy for the rest
            out.update(_greedy_inch_assignment(miss, elig_inches, cap, tgt))
        return out
    except Exception:
        return _greedy_inch_assignment(machines, elig_inches, cap, tgt)


def _min_assignment_inch_sets(machines, elig_inches: dict, cap: dict,
                              target: dict, forced_inch: dict) -> dict:
    """Part 1 — mathematical minimum-assignment covering: give each machine a small inch-SET
    (mostly ONE) so per-inch capacity covers `target`, minimising the number of multi-inch
    machines. Deterministic greedy that provably yields single-inch for every machine not needed
    to cover a fractional inch remainder:

      1. PIN forced machines (running machines whose real size has demand) to that inch.
      2. Give every remaining machine its single most under-covered eligible inch.
      3. Only if an inch is still short, add it as a 2nd inch to an ADJACENT-reachable machine
         (|existing inch − needed inch| ≤ 2) with the most spare capacity — repeat until covered.

    Returns {machine: set(inches)}. `forced_inch[m]` (optional) pins m's real running size."""
    def _num(i):
        try: return int(i)
        except Exception: return 99
    residual = {i: float(t) for i, t in target.items()}
    locked = {str(m): set() for m in machines}
    # 1. pin running machines whose real size has demand
    for m in machines:
        m = str(m); fi = forced_inch.get(m)
        if fi and fi in (elig_inches.get(m, ()) or ()):
            locked[m].add(fi)
            residual[fi] = residual.get(fi, 0.0) - cap.get(m, 0.0)
    # 2. each still-unassigned machine → its most under-covered eligible inch (most-constrained first)
    order = sorted((str(m) for m in machines),
                   key=lambda m: (len(elig_inches.get(m, ()) or ()), -cap.get(m, 0.0), m))
    for m in order:
        if locked[m]:
            continue
        opts = [i for i in (elig_inches.get(m, ()) or ()) if i in residual]
        if not opts:
            _any = sorted((i for i in (elig_inches.get(m, ()) or ()) if i), key=_num)
            if _any:
                locked[m].add(_any[0])
            continue
        best = max(opts, key=lambda i: (residual.get(i, 0.0), target.get(i, 0.0), -_num(i)))
        locked[m].add(best)
        residual[best] -= cap.get(m, 0.0)
    # 3. cover any still-short inch by adding a 2nd/3rd inch to an adjacent-reachable machine
    guard = 0
    while guard < len(locked) * 3:
        guard += 1
        under = [i for i, r in residual.items() if r > 1.0]
        if not under:
            break
        i = max(under, key=lambda z: residual[z])
        cands = [m for m in locked
                 if i in (elig_inches.get(m, ()) or ()) and i not in locked[m]
                 and any(abs(_num(a) - _num(i)) <= _INCH_BAND_WIDTH for a in locked[m])]
        if not cands:
            break                                  # structurally uncoverable (not a lock artefact)
        m = max(cands, key=lambda z: (cap.get(z, 0.0), z))
        locked[m].add(i)
        residual[i] -= cap.get(m, 0.0)
    return {m: s for m, s in locked.items() if s}


# ══════════════════════════════════════════════════════════════════════════════
# ROLLING PIPELINE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _bld_qty_per_shift(machine: str, sku=None) -> int:
    ct_min = _bld_ct_sec(machine, sku) / 60.0
    return int(SHIFT_MINS / ct_min)


def _compute_buildable_rate(engine, demand_path: str) -> dict:
    """Per-SKU sustainable building GT/day for the surplus-release 5b guard.

    For each eligible building machine, its GT/day (_bld_qty_per_shift * 3) is
    apportioned across the SKUs it can build by demand share, then summed per SKU.
    A building-oversubscribed SKU (e.g. BJ 15") gets a rate below its cure demand,
    so the guard blocks moving more curing presses onto it (which would starve).
    Stage-1 machines are excluded (carcass, not GT).
    """
    import ast
    from connection import B2C_ETL as _BETL
    ddf  = pd.read_excel(demand_path)
    scol = next((c for c in ddf.columns if "SKU" in str(c)), ddf.columns[0])
    qcol = next((c for c in ddf.columns
                 if any(x in str(c) for x in ("Requirement", "Demand", "Qty", "Quantity"))),
                ddf.columns[1])
    # Sum duplicate SKU rows (see Section E) so buildable-rate isn't understated.
    _dq = ddf[[scol, qcol]].copy()
    _dq[scol] = _dq[scol].astype(str).str.strip()
    _dq[qcol] = pd.to_numeric(_dq[qcol], errors="coerce")
    _dq = _dq.dropna(subset=[qcol])
    demand_dict = _dq.groupby(scol)[qcol].sum().to_dict()
    df_allow = _BETL(engine).load_machine_allowable()
    machine_skus: dict[str, set] = {}
    for _, r in df_allow.iterrows():
        sku = str(r["SKUCode"]).strip()
        ms  = r.get("Machines", [])
        if isinstance(ms, str):
            try:
                ms = ast.literal_eval(ms)
            except Exception:
                ms = []
        for m in ms:
            m = str(m).strip()
            if _MACHINE_GROUP.get(m, "") == "STAGE1":
                continue
            machine_skus.setdefault(m, set()).add(sku)
    # Building is DYNAMIC: a machine can pour its FULL throughput into one SKU when
    # its other SKUs don't need production that shift. So a SKU's buildable rate is
    # the FULL summed throughput of its eligible machines — NOT apportioned by demand
    # share (which underestimated by median ~44x and blocked almost every valid CO).
    # The guard then blocks only genuinely building-oversubscribed SKUs (full
    # capacity < what the added presses would consume).
    buildable: dict[str, float] = {}
    for m, skus in machine_skus.items():
        m_day = _bld_qty_per_shift(m) * 3                      # GT/day this machine
        for s in skus:
            buildable[s] = buildable.get(s, 0.0) + m_day       # full throughput, not apportioned
    return buildable


def _co_plan_supply(engine, demand_path: str, sku_inch: dict, planning_days: int = 31) -> dict:
    """Co-planning pre-solve (Part A+B). Computes the demand-optimal single-inch allocation for the
    GT machines, then derives:
      - bjus_lock[machine]        : the single locked inch for each BJ/UNISTAGE(US) machine;
      - building_inch_capacity[i] : GT/day the building side supplies inch i under that allocation
                                    (= Σ GT/day of GT machines anchored to i) — used by the curing
                                    scheduler to migrate the press draw toward what building supplies;
      - buildable[sku]            : LOCK-AWARE per-SKU building GT/day (a locked BJ/US machine
                                    contributes ONLY to its locked-inch SKUs) for the 5b guard.
    Reuses _greedy_inch_assignment + _bld_qty_per_shift. Independent of curing state (uses demand +
    allowable + sku_inch only), so it can run BEFORE the Phase-0 curing scheduler."""
    import ast
    from connection import B2C_ETL as _BETL
    ddf  = pd.read_excel(demand_path)
    scol = next((c for c in ddf.columns if "SKU" in str(c)), ddf.columns[0])
    qcol = next((c for c in ddf.columns
                 if any(x in str(c) for x in ("Requirement", "Demand", "Qty", "Quantity"))),
                ddf.columns[1])
    _dq = ddf[[scol, qcol]].copy()
    _dq[scol] = _dq[scol].astype(str).str.strip()
    _dq[qcol] = pd.to_numeric(_dq[qcol], errors="coerce")
    _dq = _dq.dropna(subset=[qcol])
    demand = _dq.groupby(scol)[qcol].sum().to_dict()
    df_allow = _BETL(engine).load_machine_allowable()
    machine_skus: dict[str, set] = {}
    for _, r in df_allow.iterrows():
        sku = str(r["SKUCode"]).strip()
        ms  = r.get("Machines", [])
        if isinstance(ms, str):
            try: ms = ast.literal_eval(ms)
            except Exception: ms = []
        for m in ms:
            m = str(m).strip()
            if _MACHINE_GROUP.get(m, "") == "STAGE1":
                continue
            machine_skus.setdefault(m, set()).add(sku)
    gt = sorted(machine_skus)
    inch_dem: dict[str, float] = defaultdict(float)
    for s in set().union(*machine_skus.values()) if machine_skus else set():
        i = sku_inch.get(str(s), "")
        if i:
            inch_dem[i] += demand.get(s, 0.0)
    elig = {m: {sku_inch.get(str(s), "") for s in machine_skus[m] if sku_inch.get(str(s), "")}
            for m in gt}
    # anchor uses MONTHLY cap (to match monthly demand); building_inch_capacity is GT/DAY.
    cap_month = {m: float(_bld_qty_per_shift(m) * 3 * planning_days) for m in gt}
    cap_day   = {m: float(_bld_qty_per_shift(m) * 3) for m in gt}
    anchor = _greedy_inch_assignment(gt, elig, cap_month, dict(inch_dem))
    inch_cap: dict[str, float] = defaultdict(float)
    for m, i in anchor.items():
        inch_cap[i] += cap_day.get(m, 0.0)
    # §3 18-inch exception: the single-anchor grid gives each machine ONE inch, so the flexible
    # 18"-machine (7003) contributes only to 15" — leaving building_inch_capacity["18"]=0, which
    # makes the SIZE_BAL/_over_cap supply gates SUPPRESS all 18" building/draw (18" collapses to 0).
    # 7003 genuinely flexes to 18" when drawn, so credit 18" with its day-capacity (kept in 15" too:
    # 18" is small + draw-timed, so the slight double-count only stops the false 0-cap suppression).
    if _INCH18_EXC and _INCH18_MACHINE in cap_day:
        inch_cap["18"] += cap_day[_INCH18_MACHINE]
    # Only US (UNISTAGE) is HARD-locked (max 1 Day-1 setup CO). BJ is left flexible (soft, tight
    # JIT → 2-3 COs/month) so it keeps its ±2 band and the KPI recovers.
    bjus_lock = {m: anchor[m] for m in gt
                 if _MACHINE_GROUP.get(m, "") == "UNISTAGE" and m in anchor}
    buildable: dict[str, float] = {}
    sku_machines: dict[str, list] = {}      # PERSKU_FEED: SKU -> lock-aware eligible GT machines
    machine_la_skus: dict[str, set] = {}    # PERSKU_FEED: machine -> lock-aware SKUs it builds (contention)
    machine_gtday: dict[str, float] = {}    # PERSKU_FEED: machine -> GT/day
    for m, skus in machine_skus.items():
        _locked = bjus_lock.get(m)                                  # only BJ/US are locked
        m_day = _bld_qty_per_shift(m) * 3
        machine_gtday[m] = m_day
        for s in skus:
            if _locked is not None and sku_inch.get(str(s), "") != _locked:
                continue                                            # locked machine can't build off-inch
            buildable[s] = buildable.get(s, 0.0) + m_day
            sku_machines.setdefault(str(s), []).append(m)
            machine_la_skus.setdefault(m, set()).add(str(s))
    # PRESS_STABLE_SINGLESRC: lock-BLIND SKU -> ALL eligible GT machines (ignores the BJ/US anchor
    # lock, unlike sku_machines above). A sole-builder SKU (e.g. 12" 1325114812074TUHL0 → only 7501)
    # is dropped from the lock-aware map when its one machine is anchored to a different inch, so the
    # single-source detector needs this raw view of who can PHYSICALLY build the SKU.
    raw_sku_machines: dict[str, list] = {}
    for m, skus in machine_skus.items():
        for s in skus:
            raw_sku_machines.setdefault(str(s), []).append(m)
    # SKU-SPECIFIC buildable GT/day for (near-)single-source SKUs (<=2 machines). Uses the
    # per-(SKU,machine) building CT (_bld_qty_per_shift(m, s)), NOT the generic machine GT/day —
    # e.g. 12" 1325114812074TUHL0 on 7501 builds 687/day (its CT), not 7501's generic 960/day. This
    # is what PRESS_STABLE_SINGLESRC pins the curing draw-drain to so the planner tracks real output.
    singlesrc_feed: dict[str, float] = {}
    for s, ms in raw_sku_machines.items():
        if len(ms) <= 2:
            singlesrc_feed[s] = float(sum(_bld_qty_per_shift(m, s) * 3 for m in ms))
    return {"bjus_lock": bjus_lock, "building_inch_capacity": dict(inch_cap), "buildable": buildable,
            "inch_dem": dict(inch_dem),      # per-inch total demand (INCH18_DEFER switch_day calc)
            "feed_ctx": {"sku_machines": sku_machines, "machine_skus": machine_la_skus,
                         "machine_gtday": machine_gtday, "raw_sku_machines": raw_sku_machines,
                         "singlesrc_feed": singlesrc_feed}}


def _shift_start_dt(date_str: str, shift: str) -> "datetime":
    """Wall-clock datetime at which (date_str, shift) begins.

    Uses SHIFT_STARTS from bc_config ("07:00"/"15:00"/"23:00"). Shift C starts
    at 23:00 on the plan date and runs into 07:00 the next calendar day — the
    date_str is still the shift's own calendar day, matching how every other
    sheet keys shifts. Downstream schedulers consuming StartTime/EndTime get a
    monotonic per-machine timeline because within a shift the cursor only ever
    advances (production + changeover minutes), and shift boundaries never
    overlap for a given machine.
    """
    d  = datetime.strptime(date_str, "%Y-%m-%d")
    hh, mm = SHIFT_STARTS.get(shift, "07:00").split(":")
    return d.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)


def _fmt_dt(dt: "datetime") -> str:
    """Format a building-schedule row timestamp for the output sheet."""
    return dt.strftime("%Y-%m-%d %H:%M")


def _mould_in_use_rows(cure_shift_rows: list, mould_info: dict,
                       demand_skus, planning_days: int, plan_start: datetime) -> list:
    """MouldInUse sheet — a FIXED daily grid: PLANNING_DAYS × #demand-SKUs rows.

    Exactly one row per (calendar day, demand SKU). For each such (day, SKU):
        Mould in USE = that SKU's moulds OCCUPIED that day = moulds mounted on presses
                       COMMITTED to the SKU (holding its moulds). A press occupies the
                       moulds from CO-IN to the SKU until it CO's to another SKU — so it
                       is counted on EVERY held day, including dry / GT-starved / idle
                       days when it produced 0 (building under-fed it). The count only
                       drops when a press actually CO's AWAY — NOT when a press merely
                       runs dry. (Previously this was moulds on presses PRODUCING that
                       shift, so a starved day falsely read 0 — a representation bug.)
        Total Eligible Moulds = size of the SKU's eligible mould pool
                       (Master_Mapping_Mould_SKU) — CONSTANT per SKU, 0 if the SKU
                       has no eligible moulds. So Mould in USE ≤ Total Eligible always.
    A day on which no press holds the SKU → Mould in USE = 0 (row still present). The
    SKU universe is EVERY SKU in the demand file, so the sheet is a complete days×SKUs
    grid. Column names are unchanged. Empty if no mould_info."""
    if not mould_info:
        return []
    eligible = {str(s): set(ms) for s, ms in (mould_info.get("sku_moulds") or {}).items()}

    # OCCUPIED moulds per SKU per day. A press's 2 moulds are OCCUPIED by the SKU it is COMMITTED to
    # (holds its moulds for) — from the moment it CO's IN to that SKU until it CO's to another SKU.
    # This counts the press EVERY day it holds the moulds, including dry / GT-starved / idle days when
    # it produced 0 (building under-fed it that day). So a SKU whose presses produce 0 GT on a day no
    # longer falsely drops to 0 moulds — the count only falls when a press actually CO's AWAY. The
    # press's held SKU per day = the SKU of its LAST shift-row that day (a CHANGEOVER → the NEW target
    # it is mounting); days with no row are forward-filled from the last held SKU (moulds stay mounted
    # on the press until the next CO). Value column stays "Mould in USE" (now = occupied, not producing).
    day_dates = [(plan_start + timedelta(days=i)).strftime("%Y-%m-%d")
                 for i in range(planning_days)]
    _didx = {d: i for i, d in enumerate(day_dates)}
    _press_day_sku: dict = defaultdict(dict)      # press -> {day_index: end-of-day held SKU}
    for r in cure_shift_rows:                     # rows are chronological → last write = end of day
        _di = _didx.get(str(r.get("Date")))
        if _di is None:
            continue
        _st = str(r.get("_status", "RUNNING"))
        sku = str(r.get("SKUCode", "")).strip()
        if _st == "CHANGEOVER":                   # the press is mounting the NEW SKU's moulds
            _rem = str(r.get("Remarks", ""))
            sku = _rem.split("→")[-1].strip() if "→" in _rem else sku
        elif _st not in ("RUNNING", "MOULD_CLEAN"):
            continue
        if not sku or sku in ("CHANGEOVER", "MOULD_CLEAN", ""):
            continue
        _press_day_sku[str(r.get("Machine"))][_di] = sku

    # Forward-fill each press's held SKU across all days (moulds stay mounted until the next change),
    # counting distinct COMMITTED PRESSES per (day, SKU).
    inuse_day: dict = defaultdict(lambda: defaultdict(set))   # day_index -> sku -> set(presses)
    for _press, _dmap in _press_day_sku.items():
        _held = None
        for _di in range(len(day_dates)):
            if _di in _dmap:
                _held = _dmap[_di]
            if _held:
                inuse_day[_di][_held].add(_press)

    # A press occupies exactly 2 moulds (2 cavities), so Mould in USE = 2 × committed presses — always
    # EVEN. (Counting the union of mould IDs went ODD when a mould was re-mounted / moved between
    # presses over the month and got deduplicated, e.g. STMX0 = 7 for 3 presses instead of 6.)
    skus = sorted({str(s).strip() for s in demand_skus if str(s).strip()})
    rows = []
    for _di, d in enumerate(day_dates):           # date-major (day 1 all SKUs, then day 2…)
        _occ = inuse_day.get(_di, {})
        for sku in skus:
            rows.append({"Date": d, "SKU Code": sku,
                         "Mould in USE": 2 * len(_occ.get(sku, set())),
                         "Total Eligible Moulds": len(eligible.get(sku, set()))})
    return rows


def _sku_data_skip_reasons(sku, sku_machine_map=None, cure_ct_map=None,
                           curing_allowable=None, sku_moulds=None,
                           bld_matrix_skus=None, cur_master_skus=None) -> str:
    """Skip_Reason for one SKU's Demand-Fulfillment row.

    RULE (user-specified): key ONLY on ALLOWABLE-MATRIX membership.
      • present in the allowable matrix  → write NOTHING (blank), even if the SKU is
        inch-locked off its inch or press-starved this snapshot (those are scheduling
        outcomes, NOT data gaps — verified: QSTL0/LSTL0 are in the masters and cure/build).
      • absent from the allowable matrix → write "missing from <building|curing> allowable matrix".
    `bld_matrix_skus` = set of SKUs present in the RAW building allowable matrix (lock-blind);
    `cur_master_skus` = set of SKUs present in the curing allowable master. When these sets are
    supplied they are authoritative. (Legacy fallback below runs only if a set is None.)"""
    s = str(sku)
    reasons = []
    # ── building: raw allowable-matrix membership only ──
    if bld_matrix_skus is not None:
        if s not in bld_matrix_skus:
            reasons.append("missing from building allowable matrix")
    elif sku_machine_map is not None and not sku_machine_map.get(s):
        reasons.append("missing from building allowable matrix")     # legacy fallback
    # ── curing: allowable-master membership only ──
    if cur_master_skus is not None:
        if s not in cur_master_skus:
            reasons.append("missing from curing allowable matrix")
    elif cur_master_skus is None and curing_allowable is not None and not curing_allowable.get(s):
        reasons.append("missing from curing allowable matrix")       # legacy fallback
    return "; ".join(reasons)


def _stage1_carcass_schedule(bld_shift_rows: list, s1_sku_to_machines: dict,
                             planning_days: int, lead: int = 2,
                             opening_carcass: dict | None = None,
                             shelf_days: int = 1) -> tuple:
    """Post-plan Stage-1 carcass scheduler — replaces the tracking-only Step-3b rows.

    Allocates the FULL 1:1 carcass for every Stage-2 GT unit to eligible Stage-1 machines,
    respecting each machine's per-shift capacity, via an EXACT time-windowed max-flow: carcass
    built by machine m in shift τ can feed Stage-2 built in shifts τ..τ+lead (a 1-2 shift
    pre-build, ≤1-day aging). Stage-1 does NOT gate GT (this never touches gt_inventory/cured),
    so it is correct utilization/qty/time accounting plus a feasibility check.

    Returns (carcass_rows, report). This is the STAGE1_CO=OFF path; the CO-charged carcass rows
    (STAGE1_CO ON) are built from the gate's own production by _stage1_carcass_rows_co.
    """
    import numpy as np
    from datetime import timedelta
    from scipy.sparse import csr_matrix
    from scipy.sparse.csgraph import maximum_flow
    _SORD = {"A": 0, "B": 1, "C": 2}

    # Stage-2 production per (date, shift) -> {sku: qty}
    s2: dict = defaultdict(lambda: defaultdict(float))
    for r in bld_shift_rows:
        if r.get("Machine_Group") == "TBM STAGE2":
            sku = r.get("SKUCode"); q = r.get("Qty", 0) or 0
            if sku and sku != "CHANGEOVER" and q > 0:
                s2[(r["Date"], r["Shift"])][str(sku)] += q
    if not s2:
        return [], {"demand": 0, "supplied": 0, "unmet": 0,
                    "no_elig_skus": [], "no_elig_units": 0}

    shifts = sorted(s2.keys(), key=lambda k: (k[0], _SORD.get(k[1], 0)))
    # #1 Carcass inventory-first: consume opening carcass FIRST (same SKU code, within the
    # shelf window from Day-0), so the plant's on-hand carcass is not wasted and Stage-1 builds
    # less. POST-HOC → KPI-neutral; only reduces Stage-1 output + the INFEASIBLE count.
    opening_used = 0.0
    if opening_carcass:
        from datetime import datetime as _dtmod
        def _dt_date(_s):
            for _f in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return _dtmod.strptime(str(_s)[:10], _f)
                except Exception:
                    pass
            return None
        _start = _dt_date(shifts[0][0])
        _pool = {str(_s): float(_q) for _s, _q in opening_carcass.items() if _q > 0}
        for k in shifts:                                  # chronological (sorted)
            _kd = _dt_date(k[0])
            if _start is not None and _kd is not None and (_kd - _start).days >= shelf_days:
                break                                     # past shelf window → all later too
            for sku in list(s2[k].keys()):
                avail = _pool.get(sku, 0.0)
                if avail <= 0:
                    continue
                take = min(avail, s2[k][sku])
                if take > 0:
                    s2[k][sku] -= take
                    _pool[sku] = avail - take
                    opening_used += take
                    if s2[k][sku] <= 0:
                        del s2[k][sku]
    tidx = {k: i for i, k in enumerate(shifts)}
    T = len(shifts)
    S1 = sorted(_S1_MACHINES)
    # BUGFIX (per-machine carcass CAP): use the machine's ACTUAL carcass-SKU cycle time, not the
    # machine-default CT. 7701 default CT=163s → CAP 176/shift, but its real carcass SKU
    # (1225170015012LSTL0) runs at 261.8s → only ~110/shift; the default over-assigned it to 133%
    # util (infeasible). Cap by the SLOWEST eligible carcass SKU so no machine can exceed its true
    # minute capacity; the excess redistributes to idle Stage-1 machines (Stage-1 is over-provisioned).
    _m_carc_skus: dict = {}
    for _sk, _ms in (s1_sku_to_machines or {}).items():
        for _mm in _ms:
            _m_carc_skus.setdefault(str(_mm), set()).add(_sk)
    def _s1_cap(m: str) -> int:
        _sks = _m_carc_skus.get(str(m))
        if not _sks:
            return int(round(_bld_qty_per_shift(m)))
        return max(1, min(_bld_qty_per_shift(m, _s) for _s in _sks))   # slowest SKU → feasible
    CAP = {m: _s1_cap(m) for m in S1}                            # carcass units/shift (per-SKU-CT)

    # ABSOLUTE calendar-shift index per produced shift. `shifts` excludes holidays (no Stage-2
    # is built on a holiday), so consecutive entries can straddle a holiday gap. Using the
    # calendar distance for the feed window means carcass built just before a holiday CANNOT
    # feed Stage-2 after it (it would exceed the 1-day shelf) — physical holiday-gap handling.
    # No holidays ⇒ cal[i+1]-cal[i] == 1 ⇒ window identical to the contiguous one (parity).
    from datetime import datetime as _dtc
    def _cal_date(_s):
        for _f in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return _dtc.strptime(str(_s)[:10], _f)
            except Exception:
                pass
        return None
    _c0 = _cal_date(shifts[0][0])
    cal = []
    for _k in shifts:
        _cd = _cal_date(_k[0])
        cal.append(((_cd - _c0).days * 3 + _SORD.get(_k[1], 0)) if (_cd and _c0) else 0)

    demlist = [(tidx[k], sku) for k in shifts for sku in s2[k]]
    NMS = len(S1) * T
    def _ms(mi, t): return 1 + mi * T + t
    dbase = 1 + NMS
    dnode = {d: dbase + j for j, d in enumerate(demlist)}
    SINK = dbase + len(demlist)
    BIG = 10 ** 7

    ri = []; ci = []; da = []
    for mi, m in enumerate(S1):                                   # source -> machine-shift
        for t in range(T):
            ri.append(0); ci.append(_ms(mi, t)); da.append(CAP[m])
    no_elig: dict = defaultdict(float)
    for (ti, sku) in demlist:
        dn = dnode[(ti, sku)]
        ri.append(dn); ci.append(SINK); da.append(int(round(s2[shifts[ti]][sku])))
        me = [m for m in (s1_sku_to_machines.get(sku) or ()) if m in _S1_MACHINES]
        if not me:
            no_elig[sku] += s2[shifts[ti]][sku]
        for m in me:                                             # machine-shift -> demand (aging window)
            mi = S1.index(m)
            for tau in range(max(0, ti - lead), ti + 1):
                ri.append(_ms(mi, tau)); ci.append(dn); da.append(BIG)

    g = csr_matrix((np.array(da), (np.array(ri), np.array(ci))), shape=(SINK + 1, SINK + 1))
    res = maximum_flow(g, 0, SINK)
    F = res.flow.tocoo()
    total_dem = sum(s2[k][s] for k in s2 for s in s2[k])

    # per (machine, production-shift τ, sku) from the machine-shift -> demand flows
    prod: dict = defaultdict(float)
    for u, v, f in zip(F.row.tolist(), F.col.tolist(), F.data.tolist()):
        if f <= 0 or not (1 <= u <= NMS) or not (dbase <= v < SINK):
            continue
        mi = (u - 1) // T; tau = (u - 1) % T
        _ti, sku = demlist[v - dbase]
        prod[(mi, tau, sku)] += f

    # build carcass rows, stacked in time within each (machine, shift)
    rows: list = []
    bymt: dict = defaultdict(list)
    for (mi, tau, sku), q in prod.items():
        if q > 0:
            bymt[(mi, tau)].append((sku, q))
    for (mi, tau), items in bymt.items():
        m = S1[mi]; dstr, sh = shifts[tau]
        cursor = _shift_start_dt(dstr, sh)
        # PM/MTC: a Stage-1 carcass machine's carcass is 1-day-shelf and aged DAY-granular, so a
        # maintenance window only needs the emitted times kept OUT of the window (done by the
        # window-skip in `_split_rows_at_shift_boundaries`, which splits a spanning carcass run
        # around the window in the SAME shift/day → aging unchanged). Carcass qty is NOT
        # maintenance-reduced here (it is sized to feed Stage-2), so forcing it strictly
        # post-window would overflow into later days and strand the Stage-2 GT it backs; the
        # same-day split is the correct, aging-safe handling for carcass.
        for sku, q in sorted(items):
            # MPQ (shift-level floor) for Stage-1 carcass: don't emit a sub-floor carcass block.
            # This is a post-plan display render (carcass ≠ GT, not in gt_inventory), so a skip
            # does NOT change cured output — it only removes tiny carcass rows from the schedule.
            if _CARCASS_MPQ > 0 and 0 < int(round(q)) < _CARCASS_MPQ:
                continue
            ct = _bld_ct_sec(m, sku)   # per-SKU CT (Stage-1 CT is constant per machine)
            _st = cursor
            cursor = cursor + timedelta(minutes=round(q) * ct / 60.0)
            rows.append({
                "Machine": m, "Date": dstr, "Shift": sh, "SKUCode": sku,
                "Qty": int(round(q)), "CO_Mins": 0,
                "StartTime": _fmt_dt(_st), "EndTime": _fmt_dt(cursor),
                "Machine_Group": _group_label(m), "CO_Type": "carcass",
            })
    report = {
        # demand = ORIGINAL Stage-2 carcass demand (before opening-carcass consumption);
        # supplied = built by Stage-1 in-window + covered by opening carcass.
        "demand": round(total_dem + opening_used),
        "supplied": round(res.flow_value + opening_used),
        "unmet": round(total_dem - res.flow_value),   # Stage-1 could not build in-window
        "opening_used": round(opening_used),
        "no_elig_skus": sorted(no_elig.keys()),
        "no_elig_units": round(sum(no_elig.values())),
    }
    return rows, report


def _fifo_reconcile_greedy(opening_carcass: dict, bld_shift_rows: list,
                           s2_consume: dict, prod_log: list, sord: dict):
    """A4 (port of optimizer/carcass_sched.py _fifo_reconcile): replay the plant
    validator's GLOBAL per-SKU carcass->Stage-2 FIFO with 1 CALENDAR-day aging (R9C/R5).
    Where carcass cannot back a Stage-2 GT unit at a shift, REDUCE that Stage-2 GT row;
    DROP carcass unconsumed past 1 calendar day. MUTATES Stage-2 rows in bld_shift_rows
    (Qty) and prod_log entries (qty) in place; ONLY reduces (never invents). Deterministic
    (all iteration over sorted keys). Returns (reduced_by_sku, reduced_by_day)."""
    from collections import deque
    build_at: dict = defaultdict(float)      # (sku,g) -> carcass units built
    keys_at:  dict = defaultdict(list)       # (sku,g) -> backing prod_log entries
    for e in prod_log:
        g = (int(e["day"]) - 1) * 3 + sord.get(e["shift"], 0)
        build_at[(e["sku"], g)] += float(e.get("qty", 0.0))
        keys_at[(e["sku"], g)].append(e)
    s2_rows: dict = defaultdict(list)        # (sku,g) -> Stage-2 GT rows
    for r in bld_shift_rows:
        if (r.get("Machine_Group") == "TBM STAGE2" and str(r.get("SKUCode")) != "CHANGEOVER"
                and (r.get("Qty", 0) or 0) > 0):
            # NOTE (mid-month latent): this uses day-of-month; prod_log above uses loop-day. Only
            # consistent for a 1st-of-month start. _CV2_A4 (the only caller) is default OFF, so it
            # doesn't hit the mid-month path today; if enabled mid-month, thread plan_start in and
            # make this plan-relative (see _planday in run_rolling_pipeline).
            _day = int(str(r["Date"]).split("-")[-1])
            s2_rows[(str(r["SKUCode"]), (_day - 1) * 3 + sord.get(r.get("Shift"), 0))].append(r)
    demand: dict = defaultdict(float)        # (sku,g) -> Stage-2 GT consumption
    for _s, lst in s2_consume.items():
        for (_day, _so, _q) in lst:
            demand[(_s, (_day - 1) * 3 + _so)] += float(_q)

    reduced_sku: dict = defaultdict(float)
    reduced_day: dict = defaultdict(float)

    def _reduce_s2(_s, _g, _amt):
        _cut = 0.0
        for _r in sorted(s2_rows.get((_s, _g), []),
                         key=lambda x: str(x.get("Machine", "")), reverse=True):
            if _amt <= 1e-9:
                break
            _q = float(_r.get("Qty", 0.0))
            _t = min(_q, _amt)
            _r["Qty"] = int(round(_q - _t))
            _amt -= _t; _cut += _t
            if _t > 0:
                reduced_day[str(_r["Date"])] += _t
        return _cut

    skus = ({s for (s, _g) in build_at} | {s for (s, _g) in demand}
            | set(opening_carcass or {}))
    for _s in sorted(skus):
        _gs = [g for (s2, g) in build_at if s2 == _s] + [g for (s2, g) in demand if s2 == _s]
        if not _gs:
            continue
        _dmin, _dmax = min(_gs) // 3, max(_gs) // 3
        q = deque()
        _op = float((opening_carcass or {}).get(_s, 0.0))
        if _op > 0:
            q.append([-1, _op, None])            # opening carcass: never ages out
        for _dn in range(_dmin, _dmax + 1):
            for _sh in range(3):
                _g = _dn * 3 + _sh
                _add = build_at.get((_s, _g), 0.0)
                if _add > 0:
                    q.append([_dn, _add, _g])
                _need = demand.get((_s, _g), 0.0)
                while _need > 1e-9 and q:
                    _use = min(q[0][1], _need)
                    q[0][1] -= _use; _need -= _use
                    if q[0][1] <= 1e-9:
                        q.popleft()
                if _need > 1e-9:                 # carcass short here -> cap Stage-2 GT
                    _c = _reduce_s2(_s, _g, _need)
                    reduced_sku[_s] += _c
            # >1-calendar-day carcass can't back FUTURE demand: drop it from the deque so
            # the next shift's shortfall test is correct. The RENDERER (_stage1_carcass_rows_co)
            # already drops aged carcass from the emitted rows — do NOT mutate prod_log here.
            while q and q[0][0] >= 0 and (_dn - q[0][0]) > 1:
                q.popleft()
    return dict(reduced_sku), dict(reduced_day)


def _stage1_carcass_rows_co(prod_log: list, s2_gt_per_sku: dict, sku_inch: dict,
                            opening_carcass: dict | None = None,
                            s2_gt_consume: dict | None = None,
                            aging_shifts: int = 3) -> tuple:
    """STAGE1_CO Site 2: build carcass rows + building-CO events from the Stage-2 gate's
    OWN production log, CAPPED per SKU at what Stage-2 actually consumes (Stage-2 GT minus
    opening carcass). CO is recomputed on the resulting consolidated sequence (60 same-inch
    / 180 diff, no production during the CO); carcass total == Stage-2 consumption.

    Row-attribution has two modes:
      • _CARCASS_NO_OVERBUILD ON (default, needs `s2_gt_consume` = per-SKU list of
        (day, shift_ord, qty) Stage-2 consumption): a TIME-WINDOWED FIFO MATCH — each
        Stage-2 consumption unit is matched to carcass BUILT within its `aging_shifts` window
        (built same shift or up to aging−1 shifts earlier), oldest build first. Matched builds
        are KEPT at their REAL build (day,shift,machine); carcass that no in-window consumer
        ever pulls (true aged-out over-production) is DROPPED. Result: carcass rows track GT
        consumption day-by-day and no aged-out carcass is shown.
      • OFF: the legacy chronological emit — the gate's builds earliest-first up to the scalar
        per-SKU target, dropping the tail (front-loads rows; kept for bit-for-bit parity).
    Returns (rows, report, co_events)."""
    from datetime import timedelta
    _SORD = {"A": 0, "B": 1, "C": 2}
    _si = sku_inch or {}

    def _s1_row_cap(machine: str, sku: str) -> int:
        """Max units a SINGLE (machine, shift) row can physically hold: floor(SHIFT_MINS*60/ct).
        Bounds the exact-T top-up below so it can never manufacture a row implying more than
        one shift's production (which would spill into fabricated future-day rows)."""
        _ct = _bld_ct_sec(machine, sku)
        return int(SHIFT_MINS * 60.0 / _ct) if _ct > 0 else 0

    _open = {str(k): float(v) for k, v in (opening_carcass or {}).items() if v and float(v) > 0}
    # HARD RULE (business): every Stage-2 GT unit is backed 1:1 by carcass. Per SKU the
    # carcass shown = EXACTLY Stage-2 GT − opening carcass consumed (integer-exact, zero
    # gap). The Stage-2 carcass GATE already enforces feasibility (Stage-2 GT is clamped to
    # available carcass, so the gate log holds ≥ this target); here we just render it exactly.
    gt_int = {str(s): int(round(float(q))) for s, q in s2_gt_per_sku.items()}
    open_used_sku = {s: min(int(_open.get(s, 0.0)), q) for s, q in gt_int.items()}
    total_gt = float(sum(gt_int.values()))
    opening_used = float(sum(open_used_sku.values()))
    tgt = {s: gt_int[s] - open_used_sku[s] for s in gt_int}          # Stage-1's integer share
    by_sku: dict = defaultdict(list)
    for i, e in enumerate(prod_log):
        by_sku[str(e["sku"])].append((int(e["day"]), _SORD.get(e["shift"], 0), i, e))
    kept: list = []
    if _CARCASS_NO_OVERBUILD and s2_gt_consume is not None:
        # ── Time-windowed FIFO match: attribute each Stage-2 consumption unit to carcass
        # built within the 1-day aging window; keep those builds at their real (day,shift,
        # machine); drop carcass no in-window consumer pulls (true aged-out over-production).
        _AG = max(1, int(aging_shifts))
        for s in by_sku:
            T = tgt.get(s, 0)
            if T <= 0:
                continue
            # supply = gate carcass builds for s, oldest first (gidx then machine = deterministic)
            sup = [[(int(e["day"]) - 1) * 3 + _so, int(e["day"]), e["date"], e["shift"],
                    str(e["machine"]), float(e["qty"]), 0.0]            # [.., qty, matched]
                   for (_d, _so, _i, e) in by_sku[s]]
            sup.sort(key=lambda r: (r[0], r[4]))
            # demand = Stage-2 consumption per (day,shift), earliest first; drop the opening-fed
            # head (opening carcass has no Stage-1 build row) so residual demand sums to T.
            dem = sorted(((int(_dd) - 1) * 3 + int(_so), float(_q))
                         for (_dd, _so, _q) in s2_gt_consume.get(s, ()))
            _skip = float(open_used_sku.get(s, 0))
            for cidx, dq in dem:
                if _skip > 0:
                    _t = min(_skip, dq); _skip -= _t; dq -= _t
                need = dq
                if need <= 1e-9:
                    continue
                for r in sup:                            # FIFO within [cidx-(AG-1), cidx]
                    if need <= 1e-9:
                        break
                    if r[0] > cidx or r[0] < cidx - (_AG - 1):
                        continue
                    _av = r[5] - r[6]
                    if _av <= 1e-9:
                        continue
                    _take = min(need, _av); r[6] += _take; need -= _take
            # emit matched supply at its real build shift, integer-rounded to hit T exactly —
            # BUT never inflate a row past its own (machine, shift) physical build capacity
            # (floor(SHIFT_MINS*60/ct)): the target T (Stage-2 GT need) can exceed what the
            # sole eligible Stage-1 machine can physically supply within the month (e.g. under
            # the 2-day plant replay, which force-builds Stage-2 GT past the normal carcass
            # gate's clamp — see _PLANT_2DAY_REPLAY). Forcing the full T onto one row used to
            # manufacture a physically-impossible multi-day row (Util_Pct > 100%, calendar
            # rows past month-end). Any genuine shortfall now stays unmet (report["unmet"]).
            _emit = 0; _cum = 0.0
            _matched = [r for r in sup if r[6] > 1e-9]
            for _k, r in enumerate(_matched):
                if _emit >= T:
                    break
                _cum += r[6]
                qi = int(round(_cum)) - _emit
                if _k == len(_matched) - 1 and _emit + qi < T:
                    qi = min(T - _emit, _s1_row_cap(r[4], s))
                if qi <= 0:
                    continue
                _emit += qi
                kept.append((r[1], r[0] % 3, r[2], r[3], r[4], s, qi))
            if _emit < T and kept and kept[-1][5] == s:  # rounding-only safeguard: hit T exactly
                rr = kept[-1]
                _topup = min(T - _emit, max(0, _s1_row_cap(rr[4], s) - rr[6]))
                if _topup > 0:
                    kept[-1] = (rr[0], rr[1], rr[2], rr[3], rr[4], rr[5], rr[6] + _topup)
    else:
        # Legacy: emit each SKU's gate production chronologically up to its integer target,
        # dropping the tail; cumulative rounding + a final top-up make the per-SKU sum EXACT.
        for s, entries in by_sku.items():
            T = tgt.get(s, 0)
            if T <= 0:
                continue
            ents = sorted(entries)
            cum = 0.0; emitted = 0
            for _k, (_d, _so, _i, e) in enumerate(ents):
                if emitted >= T:
                    break
                take = min(float(e["qty"]), T - cum)
                cum += take
                qi = int(round(cum)) - emitted
                if _k == len(ents) - 1 and emitted + qi < T:
                    # cap the forced last-chunk top-up at this row's own one-shift physical
                    # capacity — never manufacture a row implying more than one shift's
                    # production (see _s1_row_cap docstring). Genuine shortfall stays unmet.
                    qi = min(T - emitted, _s1_row_cap(str(e["machine"]), s))
                if qi <= 0:
                    continue
                emitted += qi
                kept.append((int(e["day"]), _SORD.get(e["shift"], 0), e["date"], e["shift"],
                             str(e["machine"]), s, qi))
            if emitted < T and kept and kept[-1][5] == s:  # rounding-only safeguard
                r = kept[-1]
                _topup = min(T - emitted, max(0, _s1_row_cap(r[4], s) - r[6]))
                if _topup > 0:
                    kept[-1] = (r[0], r[1], r[2], r[3], r[4], r[5], r[6] + _topup)
    # Lay out per machine, chronological; recompute CO on the consolidated sequence.
    bym: dict = defaultdict(list)
    for row in kept:
        bym[row[4]].append(row)
    rows: list = []; co_events: list = []
    produced = 0.0; co_cnt = 0; co_min_tot = 0.0
    for m in sorted(bym):
        cur = ""
        byds: dict = defaultdict(dict)     # (day,so,date,shift) -> {sku: qty}  (merge same-sku)
        for (day, so, date, shift, _m, s, q) in bym[m]:
            d = byds[(day, so, date, shift)]
            d[s] = d.get(s, 0.0) + q
        _shift_keys = sorted(byds)
        _defer: dict = {}     # {sku: qty} carried to the next shift when a CO would cross a boundary
        for _ski, (day, so, date, shift) in enumerate(_shift_keys):
            cursor = _shift_start_dt(date, shift)
            _shift_end = cursor + timedelta(minutes=SHIFT_MINS)
            _has_next = _ski < len(_shift_keys) - 1
            # SKUs carried from a previous shift's deferred CO are produced FIRST in this shift.
            _bucket = dict(byds[(day, so, date, shift)])
            for _ds, _dq in _defer.items():
                _bucket[_ds] = _bucket.get(_ds, 0.0) + _dq
            _defer = {}
            # continuing SKU (no CO) first, then by sku for determinism
            order = sorted(_bucket.items(),
                           key=lambda t: (0 if t[0] == cur else 1, t[0]))
            for _oi, (s, q) in enumerate(order):
                q = int(round(q))
                if q <= 0:
                    continue
                # MPQ (shift-level floor) for Stage-1 carcass: skip a sub-floor carcass block
                # (do NOT CO or emit it). Carcass is a post-plan display render (≠ GT, not in
                # gt_inventory), so this does NOT change cured output — it only removes tiny
                # carcass rows from the schedule. Set CARCASS_MPQ=0 to disable.
                if _CARCASS_MPQ > 0 and q < _CARCASS_MPQ:
                    continue
                if cur not in ("", s):
                    comin = int(_co_cost(m, _si.get(cur, ""), _si.get(s, "")))
                    # SHIFT-CONTAINMENT (client hard rule): a building CO must START and FINISH
                    # within one shift. If this carcass CO would cross the shift boundary, DEFER it
                    # (and every not-yet-produced SKU this shift, incl. this one) to the NEXT shift's
                    # start — the rest of this shift is left idle and the CO runs cleanly at the
                    # next boundary. Total carcass/SKU is preserved (re-timed within the ≤2-shift
                    # pre-build lead, so it still precedes its Stage-2 draw). A deferred CO always
                    # fits (it starts at the boundary; comin ≤ 180 < 480). On the LAST shift there
                    # is no next shift to cross into, so the CO is kept.
                    if _CO_SHIFT_CONTAINED and _has_next and cursor + timedelta(minutes=comin) > _shift_end:
                        for _rs, _rq in order[_oi:]:
                            _defer[_rs] = _defer.get(_rs, 0.0) + int(round(_rq))
                        break
                    _cot = ("same_size_CO" if _si.get(cur, "") == _si.get(s, "")
                            else "diff_size_CO")
                    _co_start = cursor
                    cursor = cursor + timedelta(minutes=comin)
                    # Separate CHANGEOVER row in the Shift Schedule — SAME shape as the GT /
                    # Stage-2 machines (SKUCode=CHANGEOVER, Qty=0, CO_Mins=comin, no production
                    # during it). Makes Stage-1 COs show identically to every other machine.
                    rows.append({
                        "Machine": m, "Date": date, "Shift": shift, "SKUCode": "CHANGEOVER",
                        "Qty": 0, "CO_Mins": comin,
                        "StartTime": _fmt_dt(_co_start), "EndTime": _fmt_dt(cursor),
                        "Machine_Group": _group_label(m), "CO_Type": _cot,
                    })
                    co_events.append({
                        "Machine": m, "Date": date, "Shift": shift, "Day": day,
                        "CO_Day_Index": day, "From_SKU": cur, "Target_SKU": s,
                        "CO_Type": _cot, "CO_Cost_Mins": comin,
                        "Status": f"Stage-1 carcass CO ({_cot})",
                    })
                    co_cnt += 1; co_min_tot += comin
                ct = _bld_ct_sec(m, s)
                _st = cursor
                cursor = cursor + timedelta(minutes=q * ct / 60.0)
                rows.append({
                    "Machine": m, "Date": date, "Shift": shift, "SKUCode": s,
                    "Qty": q, "CO_Mins": 0,     # the CO time lives in the CHANGEOVER row above
                    "StartTime": _fmt_dt(_st), "EndTime": _fmt_dt(cursor),
                    "Machine_Group": _group_label(m), "CO_Type": "carcass",
                })
                cur = s; produced += q
    supplied = produced + opening_used
    report = {
        "demand": round(total_gt), "supplied": round(supplied),
        "unmet": round(max(0.0, total_gt - supplied)), "produced": round(produced),
        "opening_used": round(opening_used), "co_count": co_cnt,
        "co_mins": round(co_min_tot), "no_elig_skus": [], "no_elig_units": 0,
    }
    return rows, report, co_events


def _consolidate_carcass_rows(carc_rows: list, sku_inch: dict) -> tuple:
    """Carcass row CONSOLIDATION (representation fix; KPI/feasibility-NEUTRAL).

    Input = the ALREADY-SHIFT-SPLIT Stage-1 carcass rows (CO_Type "carcass" + the carcass
    CHANGEOVER rows) — so every row's Date/Shift is the FINAL, wall-clock-correct one. This
    rebuilds them FULLER and FEWER by moving carcass ONLY within the SAME calendar day, which
    keeps the per-(date, SKU) carcass total EXACTLY (integers): R5 / R9C (both DAY-granular),
    every GT/cured/demand-cap figure, and the carcass mass-balance are untouched. Two moves:
      (1) merge each machine's cross-shift dribbles for a SKU into one contiguous block, and
      (2) FOLD a machine's sub-threshold (< _CARCASS_CONSOLIDATE_MIN) daily carcass for a SKU
          onto a SIBLING machine ALREADY building that same (date, SKU) with spare day-capacity
          — never a machine outside that (date, SKU)'s set, so Stage-1 eligibility / inch-lock
          is preserved.
    Each (machine, date) is re-laid CONTIGUOUSLY from 07:00: its production is fixed and its
    consolidated CO count is <= what the split rows already carried on that date, so the run
    stays inside [07:00, +1 day] and no carcass spills into the next day (the caller re-splits
    at shift boundaries, re-deriving Date/Shift from wall clock and apportioning Qty by shift —
    all WITHIN the day). Returns (rows, co_events). CARCASS_CONSOLIDATE=0 → caller skips this.
    """
    from datetime import datetime as _dtp
    _si = sku_inch or {}
    _INV = {0: "A", 1: "B", 2: "C"}
    _SORD = {"A": 0, "B": 1, "C": 2}

    # non-carcass rows (defensive: input should be carcass-only) pass through untouched
    def _is_carc(_r):
        return _r.get("CO_Type") == "carcass"
    def _is_carc_co(_r):
        return (str(_r.get("SKUCode")) == "CHANGEOVER"
                and _r.get("CO_Type") in ("same_size_CO", "diff_size_CO"))
    passthrough = [r for r in carc_rows if not (_is_carc(r) or _is_carc_co(r))]

    # aggregate carcass production per (machine, date, sku) — dates are FINAL (post-split)
    agg: dict = defaultdict(int)
    for r in carc_rows:
        if _is_carc(r):
            agg[(str(r["Machine"]), str(r["Date"])[:10], str(r["SKUCode"]))] += int(round(r.get("Qty", 0) or 0))
    prodmin: dict = defaultdict(float)
    for (m, dt, s), q in agg.items():
        prodmin[(m, dt)] += q * _bld_ct_sec(m, s) / 60.0

    # (1) cross-machine FOLD of sub-threshold (date, sku) fragments onto a busier sibling
    if _CARCASS_CONSOLIDATE_MIN > 0:
        capmin = 3.0 * SHIFT_MINS - 2.0 * 180.0     # leave room for the re-layout COs
        dsm: dict = defaultdict(list)
        for (m, dt, s) in list(agg):
            dsm[(dt, s)].append(m)
        for (dt, s), ms in sorted(dsm.items()):
            if len(ms) < 2:
                continue
            for donor in sorted(ms, key=lambda mm: agg.get((mm, dt, s), 0)):
                dq = agg.get((donor, dt, s), 0)
                if dq <= 0 or dq >= _CARCASS_CONSOLIDATE_MIN:
                    continue
                for recv in sorted((mm for mm in ms
                                    if mm != donor and agg.get((mm, dt, s), 0) > 0),
                                   key=lambda mm: -agg.get((mm, dt, s), 0)):
                    add = dq * _bld_ct_sec(recv, s) / 60.0
                    if prodmin[(recv, dt)] + add > capmin:
                        continue
                    agg[(recv, dt, s)] += dq
                    prodmin[(recv, dt)] += add
                    prodmin[(donor, dt)] -= dq * _bld_ct_sec(donor, s) / 60.0
                    del agg[(donor, dt, s)]
                    break

    # a loop-day hint per date (for the carcass-CO event Day field), from the split CO rows
    dayof: dict = {}
    for r in carc_rows:
        if _is_carc_co(r) and r.get("Day") is not None:
            dayof.setdefault(str(r["Date"])[:10], r["Day"])

    # (2) rebuild rows date-contained per machine, contiguous from 07:00
    mds: dict = defaultdict(lambda: defaultdict(dict))     # machine -> date -> {sku: qty}
    for (m, dt, s), q in agg.items():
        if q > 0:
            mds[m][dt][s] = q
    out_rows: list = list(passthrough)
    out_co: list = []
    for m in sorted(mds):
        cur = ""
        for dt in sorted(mds[m]):
            day = dayof.get(dt)
            if day is None and dayof:
                _a = min(dayof)
                day = dayof[_a] + (_dtp.strptime(str(dt)[:10], "%Y-%m-%d")
                                   - _dtp.strptime(str(_a)[:10], "%Y-%m-%d")).days
            elif day is None:
                day = 1
            day_start = _shift_start_dt(dt, "A")
            cursor = day_start
            def _shlabel(_c):
                return _INV[min(2, max(0, int((_c - day_start).total_seconds() // 60 // SHIFT_MINS)))]
            # continuing SKU first (no leading CO), then biggest-first for fuller rows
            order = sorted(mds[m][dt].items(), key=lambda t: (0 if t[0] == cur else 1, -t[1], t[0]))
            for (s, q) in order:
                q = int(round(q))
                if q <= 0:
                    continue
                if cur not in ("", s):
                    comin = int(_co_cost(m, _si.get(cur, ""), _si.get(s, "")))
                    cot = ("same_size_CO" if _si.get(cur, "") == _si.get(s, "")
                           else "diff_size_CO")
                    sh = _shlabel(cursor)
                    cs = cursor
                    cursor = cursor + timedelta(minutes=comin)
                    out_rows.append({
                        "Machine": m, "Date": dt, "Shift": sh, "SKUCode": "CHANGEOVER",
                        "Qty": 0, "CO_Mins": comin,
                        "StartTime": _fmt_dt(cs), "EndTime": _fmt_dt(cursor),
                        "Machine_Group": _group_label(m), "CO_Type": cot,
                    })
                    out_co.append({
                        "Machine": m, "Date": dt, "Shift": sh, "Day": day,
                        "CO_Day_Index": day, "From_SKU": cur, "Target_SKU": s,
                        "CO_Type": cot, "CO_Cost_Mins": comin,
                        "Status": f"Stage-1 carcass CO ({cot})",
                    })
                ct = _bld_ct_sec(m, s)
                st = cursor
                sh = _shlabel(st)
                cursor = cursor + timedelta(minutes=q * ct / 60.0)
                out_rows.append({
                    "Machine": m, "Date": dt, "Shift": sh, "SKUCode": s,
                    "Qty": q, "CO_Mins": 0,
                    "StartTime": _fmt_dt(st), "EndTime": _fmt_dt(cursor),
                    "Machine_Group": _group_label(m), "CO_Type": "carcass",
                })
                cur = s
    return out_rows, out_co


def _enforce_carcass_shift_cap(rows: list, holiday_windows: "list | None" = None) -> tuple:
    """FINAL carcass-row hard cap — fixes the residual cumulative-rounding artifact.

    ROOT CAUSE: `_split_rows_at_shift_boundaries` apportions a multi-shift block's Qty
    across its per-shift segments as `int(qty_floor * frac)` for every segment EXCEPT
    the LAST, which gets the remainder `qty_floor - already_assigned` (so the per-row
    total reconciles exactly). That remainder INHERITS the flooring loss of every
    earlier segment (up to ~(#segments-1) units), so the last segment's Qty can exceed
    its OWN shift's strict physical capacity by 1-2 units even though its own
    [StartTime,EndTime] span never crosses a shift boundary (that was already fixed —
    see the R11B/over-480-min fix in `_split_rows_at_shift_boundaries`/
    `_consolidate_carcass_rows`). This is a SEPARATE, smaller artifact: over the row's
    own strict floor(available_shift_min*60/ct), not over a whole shift's span.

    For every carcass production row, computes:
        available_shift_min = SHIFT_MINS(480) - PM/MTC downtime overlapping that
            (machine, date, shift) - holiday downtime - any CO minutes the SAME
            machine spent that SAME shift - any OTHER SKU's production minutes
            already occupying that (machine, date, shift) (a machine can run >1 SKU
            per shift after a same-shift CO)
        cap = floor(available_shift_min * 60 / _bld_ct_sec(machine, sku))
    The "OTHER SKU" term is read from a FROZEN pre-mutation snapshot of every row's
    Qty, taken once up front — NOT the live (being-edited) Qty. This keeps the result
    deterministic and order-independent (which (machine,date,sku) group is processed
    first no longer changes the answer) and, since a frozen value is always >= the
    OTHER sku's own eventual (possibly also capped-down) Qty, it is a conservative
    (never-too-generous) budget: this fix can under-fill a shift by a few units versus
    a perfectly joint-optimal packing, but can NEVER let two SKUs' rows in the same
    shift jointly exceed SHIFT_MINS (verified: 0 new R11B/R15B violations after this
    fix, vs 0 before — see the accompanying audit).

    A row with Qty > cap is capped at `cap`; the excess is offered to a SIBLING row of
    the SAME (machine, date, SKU) on ANOTHER shift of the SAME calendar day (carcass
    1-day shelf -> same-day only) that has spare room (its own cap - its own Qty),
    extending that sibling's EndTime to stay Qty*CT-consistent (R18B). Both the capped
    donor row and any topped-up sibling have their EndTime re-derived from their
    (unchanged) StartTime + Qty*CT so every row stays internally consistent. A residual
    excess with no same-day home is a genuinely unbuildable unit that day -> DROPPED
    from the display (carcass is NOT in gt_inventory -> never touches cured/built/
    coverage) and reported.

    Display-only, KPI-neutral, deterministic (all iteration over sorted keys).
    Returns (rows, rows_over_before, dropped_total, dropped_by_day)."""
    _hol = list(holiday_windows or [])

    def _is_carc(r):
        return r.get("CO_Type") == "carcass" and str(r.get("SKUCode")) != "CHANGEOVER"

    def _is_co(r):
        return (str(r.get("SKUCode")) == "CHANGEOVER"
                and r.get("CO_Type") in ("same_size_CO", "diff_size_CO"))

    # FROZEN pre-mutation snapshot: (machine,date,shift) -> CO minutes, and
    # (machine,date,shift) -> [(sku, qty), ...] for every carcass production row. Used
    # ONLY to compute each SKU's OTHER-sku budget below — never mutated, so the cap for
    # one SKU never depends on the processing order of another SKU's fix this pass.
    _frozen_co: dict = defaultdict(float)
    _frozen_prod: dict = defaultdict(list)
    for r in rows:
        key = (str(r.get("Machine")), str(r.get("Date"))[:10], r.get("Shift"))
        if _is_co(r):
            _frozen_co[key] += float(r.get("CO_Mins", 0) or 0)
        elif _is_carc(r):
            _frozen_prod[key].append((str(r.get("SKUCode")), float(r.get("Qty", 0) or 0)))

    def _shift_bounds(date, shift):
        st = _shift_start_dt(date, shift)
        return st, st + timedelta(minutes=SHIFT_MINS)

    def _avail_for_sku(machine, date, shift, sku):
        key = (machine, date, shift)
        st, en = _shift_bounds(date, shift)
        pm = _down_mins(_BLD_DOWN.get(str(machine)) or [], st, en)
        hol = _down_mins(_hol, st, en)
        other = 0.0
        for s, q in _frozen_prod.get(key, ()):
            if s == sku:
                continue
            ct2 = _bld_ct_sec(machine, s)
            if ct2 > 0:
                other += q * ct2 / 60.0
        return max(0.0, SHIFT_MINS - pm - hol - _frozen_co.get(key, 0.0) - other)

    def _cap(machine, date, shift, sku):
        ct2 = _bld_ct_sec(machine, sku)
        if ct2 <= 0:
            return 0
        return int(_avail_for_sku(machine, date, shift, sku) * 60.0 / ct2)

    # group production rows by (machine, date, sku) for same-day redistribution
    by_mds: dict = defaultdict(list)
    for r in rows:
        if _is_carc(r):
            by_mds[(str(r.get("Machine")), str(r.get("Date"))[:10],
                    str(r.get("SKUCode")))].append(r)

    over_before = 0
    dropped_total = 0
    dropped_by_day: dict = defaultdict(int)
    for (m, dt, sku), rs in sorted(by_mds.items()):
        rs = sorted(rs, key=lambda r: str(r.get("Shift", "")))
        ct = _bld_ct_sec(m, sku)
        # caps are computed ONCE from the frozen snapshot (order-independent) — safe to
        # precompute per row before any mutation in this group.
        caps = {id(r): _cap(m, dt, r.get("Shift"), sku) for r in rs}
        for r in rs:
            q = int(round(r.get("Qty", 0) or 0))
            c = caps[id(r)]
            if q <= c:
                continue
            over_before += 1
            excess = q - c
            r["Qty"] = c
            if ct > 0:
                try:
                    _rst = datetime.strptime(str(r["StartTime"]), "%Y-%m-%d %H:%M")
                    _shend = _shift_bounds(dt, r.get("Shift"))[1]
                    r["EndTime"] = _fmt_dt(min(_shend, _rst + timedelta(minutes=c * ct / 60.0)))
                except Exception:
                    pass
            for other in rs:
                if other is r or excess <= 0:
                    continue
                oc = caps[id(other)]
                oq = int(round(other.get("Qty", 0) or 0))
                spare = oc - oq
                if spare <= 0:
                    continue
                take = min(spare, excess)
                other["Qty"] = oq + take
                if ct > 0:
                    try:
                        _ost = datetime.strptime(str(other["StartTime"]), "%Y-%m-%d %H:%M")
                        _oshend = _shift_bounds(dt, other.get("Shift"))[1]
                        other["EndTime"] = _fmt_dt(min(_oshend, _ost + timedelta(minutes=(oq + take) * ct / 60.0)))
                    except Exception:
                        pass
                excess -= take
            if excess > 0:
                dropped_total += excess
                dropped_by_day[dt] += excess
    return rows, over_before, dropped_total, dict(dropped_by_day)


def _daily_capacity_util(cure_shift_rows: list, bld_shift_rows: list,
                         press_stats: dict, planning_days: int) -> list:
    """Per-DAY capacity utilisation for the cloud jkt_plan_capacityUtilisation table
    (30-31 rows per plan). Curing daily util = (production + mould-clean + CO time) /
    total available press-minutes that day — the metric the client requested. Building
    group utils computed symmetrically = (production + CO) / available machine-minutes.
    Aggregates to the same monthly occupancy written to jkt_plan_kpis (sum of daily busy
    over sum of daily available == the monthly figure)."""
    from collections import defaultdict
    from datetime import datetime as _dt
    DAY_AVAIL_PER = 3 * SHIFT_MINS          # 1440 min/machine/day

    def _grp(m: str) -> str:
        m = str(m)
        if m in {"6001","6002","6003","6004","7001","7002","7003","7004"}: return "VMI"
        if m in {"7101","7102","7103","7104","7105","7106","7201"}:        return "BJ"
        if m in {"7501","7502","7503"}:                                    return "UNI_NARROW"
        if m in {"8201","8301","8302","8501","8502","7301"}:               return "STAGE2"
        return "STAGE1"
    GROUP_N = {"VMI": 8, "BJ": 7, "UNI_NARROW": 3, "STAGE2": 6, "STAGE1": 15}

    # ── curing: production (from RUNNING segment duration) + CO + mould-clean per day ──
    cure = defaultdict(lambda: {"prod": 0.0, "co": 0.0, "clean": 0.0})
    for r in cure_shift_rows:
        d = r.get("Date")
        if not d:
            continue
        cure[d]["co"]    += float(r.get("CO_Mins", 0) or 0)
        cure[d]["clean"] += float(r.get("Mould_Clean_Mins", 0) or 0)
        if r.get("_status") == "RUNNING" and (r.get("Qty", 0) or 0) > 0:
            try:
                _st = _dt.strptime(r["StartTime"], "%Y-%m-%d %H:%M")
                _en = _dt.strptime(r["EndTime"],   "%Y-%m-%d %H:%M")
                cure[d]["prod"] += max(0.0, (_en - _st).total_seconds() / 60.0)
            except Exception:
                pass
    n_press    = len(press_stats)
    # Denominator is the FIXED plant roster (CURING_PRESS_COUNT), NOT the count of
    # presses actually simulated (len(press_stats), which tracks the running-moulds
    # snapshot and can drift). Keeps the daily/monthly curing KPI stable and
    # comparable across runs; change the roster size in bc_config.CURING_PRESS_COUNT.
    cure_avail = CURING_PRESS_COUNT * DAY_AVAIL_PER

    # ── building: production (Qty × CT) + CO per day per group ──
    bld = defaultdict(lambda: defaultdict(lambda: {"prod": 0.0, "co": 0.0}))
    for r in bld_shift_rows:
        d = r.get("Date"); m = r.get("Machine"); g = _grp(m)
        if str(r.get("SKUCode", "")) == "CHANGEOVER":
            bld[d][g]["co"]   += float(r.get("CO_Mins", 0) or 0)
        else:
            bld[d][g]["prod"] += (float(r.get("Qty", 0) or 0)
                                  * _bld_ct_sec(m, r.get("SKUCode")) / 60.0)

    out = []
    for d in sorted(cure):
        c = cure[d]
        cu = round(100.0 * (c["prod"] + c["co"] + c["clean"]) / cure_avail, 2) if cure_avail else 0.0
        def _gocc(groups):
            busy  = sum(bld[d][g]["prod"] + bld[d][g]["co"] for g in groups)
            avail = sum(GROUP_N[g] * DAY_AVAIL_PER for g in groups)
            return round(100.0 * busy / avail, 2) if avail else 0.0
        out.append({
            "date":                             d,
            "capacityUtilisation":              cu,                                  # CURING (client ask)
            "building_capacityUtilisation":     _gocc(["VMI","BJ","UNI_NARROW","STAGE2","STAGE1"]),
            "building_s2_capacityUtilisation":  _gocc(["VMI","BJ","UNI_NARROW","STAGE2"]),
            "stage1_capacityUtilisation":       _gocc(["STAGE1"]),
            "vmi_capacityUtilisation":          _gocc(["VMI"]),
            "bj_capacityUtilisation":           _gocc(["BJ"]),
            "uniNarrow_capacityUtilisation":    _gocc(["UNI_NARROW"]),
            # audit fields (not written to DB; used to verify the daily calc)
            "_cure_prod_mins":  round(c["prod"]), "_cure_co_mins": round(c["co"]),
            "_cure_clean_mins": round(c["clean"]), "_cure_avail_mins": cure_avail,
            "_n_presses": n_press,
        })
    return out


def _urgency_score(
    sku: str,
    demand_remaining: dict,
    press_count: dict,
    cure_ct_map: dict,
    days_left: int,
    cavities: int = 2,
) -> float:
    """
    Urgency = demand that CANNOT be covered by current curing presses in remaining horizon.
    Higher → more urgent → becomes primary in pool ranking.

    If days_left ≤ 0 or no presses: return full remaining demand (maximally urgent).
    """
    dem = demand_remaining.get(sku, 0.0)
    if dem <= 0:
        return 0.0
    n  = press_count.get(sku, 0)
    ct = float(cure_ct_map.get(sku, DEFAULT_CURING_CT))
    if n <= 0 or ct <= 0 or days_left <= 0:
        return dem
    rate_per_shift = int(SHIFT_MINS / ct) * cavities   # units one press cures per shift
    max_curable    = n * rate_per_shift * 3 * days_left  # 3 shifts/day
    return max(0.0, dem - max_curable)


def _build_machine_pools(
    machine_skus:    dict,
    sku_inch:        dict,
    demand_dict:     dict,
    press_count:     dict,
    cure_ct_map:     dict,
    planning_days:   int,
    pool_size:       int = POOL_SIZE,
) -> dict:
    """
    Build a fixed pool of up to `pool_size` same-dominant-inch SKUs per machine.
    Pool ordering = urgency descending at Day 1.
    Returns: {machine: [sku_primary, sku_sec1, sku_sec2, ...]}
    """
    pools: dict[str, list[str]] = {}
    for machine, skus in machine_skus.items():
        dom_inch = _MACHINE_DOMINANT_INCH.get(str(machine), "")

        # Prefer same dominant inch + has demand + has active presses
        same_inch = [
            s for s in skus
            if sku_inch.get(s, "") == dom_inch
            and demand_dict.get(s, 0) > 0
            and press_count.get(s, 0) > 0
        ]
        if not same_inch:
            # Fallback: any eligible SKU with demand and presses
            same_inch = [
                s for s in skus
                if demand_dict.get(s, 0) > 0 and press_count.get(s, 0) > 0
            ]

        same_inch.sort(
            key=lambda s: (-_urgency_score(s, demand_dict, press_count, cure_ct_map,
                                           _bc_working_days_left(1, planning_days)), s)  # working-day horizon
        )
        pools[machine] = same_inch[:pool_size]

    return pools


def _cure_qty_per_shift(ct_min: float) -> int:
    return int(SHIFT_MINS / ct_min) * CURING_CAVITIES


def _co_cost(machine: str, from_inch: str, to_inch: str,
             from_sku: str = "", to_sku: str = "") -> int:
    mg = _MACHINE_GROUP.get(str(machine), "VMI")
    if from_inch == to_inch:
        # BJ-only same-size CO exception for specific SKU pairs (direction-agnostic).
        # Applies ONLY to BJ + same-inch + a listed pair; everything else unchanged.
        if mg == "BJ" and from_sku and to_sku:
            _ex = BJ_SAME_SIZE_CO_EXCEPTIONS.get(frozenset({str(from_sku), str(to_sku)}))
            if _ex is not None:
                return int(_ex)
        return BUILDING_CO_SAME_SIZE.get(mg, 60)
    # #2 optimizer-choice: a DIRECT +3/-3 (>2-inch jump) on a non-BJ/non-Stage-2 machine costs the
    # full 8h CO (INCH_PLUS3_CO_MINS) — so the scorer naturally prefers a cheaper TWO-HOP (two ≤2
    # diff-COs, e.g. 15→16→18) whenever the intermediate inch is productive, and only pays the 8h
    # direct jump when the two-hop isn't worthwhile. (BJ has no +3 in its set (#1); Stage-2 exempt.)
    if _VMI_TWO_HOP and mg != "STAGE2":
        _f, _t = _inch_num(from_inch), _inch_num(to_inch)
        if _f is not None and _t is not None and abs(_t - _f) > 2:
            return int(INCH_PLUS3_CO_MINS)
    return BUILDING_CO_DIFF_SIZE.get(mg, 120)


def _inch_num(inch: str):
    """Inch string -> int, or None when it isn't a usable number."""
    try:
        return int(str(inch).strip())
    except (TypeError, ValueError):
        return None


def _inch_ok(to_inch: str, cur_inch: str, anchor: str, used: set) -> bool:
    """Client inch rules — Rule 2 (+/-2 band) ONLY.

    anchor == "" means the machine has not been assigned yet, so its first
    assignment is unconstrained (that assignment sets the anchor).
    Staying on the current inch is always allowed.

    Rule 1a (permanent no-revisit) has been RETIRED: the plant rule is now a 5-day
    minimum inch dwell (MIN_INCH_DWELL_DAYS), so a machine MAY return to an inch it
    left once the dwell/deficit-done leave gate permits. Revisit is therefore legal
    here; the timing is enforced by _may_leave_inch at the leave sites. `used` is kept
    in the signature for callers but no longer gates.
    """
    if not _INCH_RULES_ENABLED:
        return True
    if to_inch == cur_inch:
        return True
    if _INCH_HIST_LOCK_ENABLED:
        # ±2 anchor band DISCONTINUED — the per-machine historical allowed-inch set
        # (enforced by the allowable-machine strip + machine_locked_inches gate) is
        # the sole WHICH constraint, so historically-evidenced +/-3 jumps are legal.
        return True

    # Rule 2 — must stay within anchor +/- _INCH_BAND_WIDTH.
    a, t = _inch_num(anchor), _inch_num(to_inch)
    if a is not None and t is not None and abs(t - a) > _INCH_BAND_WIDTH:
        return False
    return True


def _inch_demand_done(machine: str, cur_inch: str, machine_skus: dict,
                      sku_inch: dict, deficit_fn, buf, rate: float = 0.0,
                      demand_remaining: dict | None = None,
                      projected_gt: dict | None = None) -> bool:
    """Rule 1b — may this machine leave `cur_inch` for a different inch?

    True when no SKU at the machine's current inch still has a deficit it could
    usefully serve. (Per the client clarification this is "no unmet demand it can
    serve now", NOT "all demand at that inch globally exhausted".)

    IMPORTANT — the deficit must be big enough to form a LEGAL campaign. A
    machine needs >= MIN_CAMPAIGN_MINS of work to build anything, so a tiny
    residual deficit (say 10 units) would otherwise pin the machine forever:
    too small to build, too big to leave. That trap was the single largest
    source of idle time under the one-way rule, so a sub-campaign remainder
    counts as "inch finished".
    """
    if not _INCH_RULES_ENABLED or not cur_inch:
        return True
    # STRICT (Rule 1, INCH_STRICT): "done" = the inch's WHOLE servable demand is exhausted
    # on this machine — every same-inch SKU has demand_remaining - projected_gt <= 0. This
    # replaces the momentary buffer-filled deficit test below, so a machine stays single-inch
    # through its dwell window (no temporary off-inch hop) and only CO's away when the inch is
    # truly finished (or 5 days pass, enforced separately in _may_leave_inch).
    if _INCH_STRICT and demand_remaining is not None:
        _pg = projected_gt or {}
        for s in machine_skus.get(machine, ()) or ():
            if sku_inch.get(s, "") != cur_inch:
                continue
            if demand_remaining.get(s, 0.0) - _pg.get(s, 0.0) > 0:
                return False
        return True
    # Threshold below which a leftover deficit is treated as "inch finished".
    # MEASURED RESULT: raising this to a full campaign made every month WORSE
    # (May -13,197 / June -8,550 / July -28,304). Under one-way movement leaving
    # is PERMANENT, so an easier exit just burns the machine's inches sooner and
    # it runs out of legal work. Reluctance to leave is protective here — keep
    # the strict "any deficit blocks departure" default (threshold = 0).
    min_units = 0.0
    if _INCH_GATE_CAMPAIGN_THRESHOLD:
        min_units = (max(MIN_CAMPAIGN_UNITS, MIN_CAMPAIGN_MINS * rate)
                     if rate > 0 else MIN_CAMPAIGN_UNITS)
    for s in machine_skus.get(machine, ()) or ():
        if sku_inch.get(s, "") != cur_inch:
            continue
        try:
            d = deficit_fn(s, buf)
        except TypeError:          # deficit closures that take only the SKU
            d = deficit_fn(s)
        if d > min_units:
            return False
    return True


def _select_dynamic_co_target(
    old_sku: str,
    demand_remaining: dict,
    press_count: dict,
    cure_ct_map: dict,
    priority_score_map: dict,
    gt_inventory: dict,
    horizon_left: int,
    already_targeted: set,
    priority_deadline_map: dict | None = None,   # DELIVERY_PRIORITY: {sku: deadline_day} or None
    day: int = 0,
) -> "str | None":
    """Select the best curing CO target when a press finishes its SKU demand mid-plan.

    The calling press is already idle (old_sku demand = 0), so any production on a
    new SKU is strictly better than idle. Both Class A and Class B targets are eligible.

    Sort key: Class A first (critical, can't meet demand without this press), then Class B.
    Within class: fewest after-CO days → highest priority score → most GT in inventory.
    `already_targeted` prevents multiple dynamic COs going to the same new SKU in
    the same shift.

    DELIVERY_PRIORITY: when priority_deadline_map is non-empty, a committed target is
    ranked FIRST (EARLIEST-DEADLINE-FIRST) and is measured against its OWN deadline for
    the Class-A test (so a behind-schedule committed SKU fires even before GT is banked —
    the building side co-directs its GT). Empty/None → byte-identical to before.
    """
    _pdm = priority_deadline_map or {}
    _prio_on = bool(_pdm)
    candidates = []
    for sku, rem in demand_remaining.items():
        if rem <= 0 or sku == old_sku or sku in already_targeted:
            continue
        n    = press_count.get(sku, 0)
        ct   = cure_ct_map.get(sku, DEFAULT_CURING_CT)
        rate = _cure_qty_per_shift(ct) * 3          # per-day curing rate
        if rate <= 0:
            continue
        current_days = rem / (n * rate) if n > 0 else float("inf")
        _dd = _pdm.get(sku) if _prio_on else None
        _h  = horizon_left if _dd is None else min(horizon_left, max(1, _dd - day + 1))
        # Class A = cannot meet demand without this press; Class B = helpful but not critical.
        urgency_class = 0 if current_days > _h else 1
        # Class B allowed only if GT inventory already covers ≥ 1 shift of curing.
        # Without this guard, a Class B press starts curing next shift but finds zero GT
        # (building never pre-built for it) → starvation event instead of production.
        gt_inv = gt_inventory.get(sku, 0.0)
        if urgency_class == 1:
            ct_sku    = cure_ct_map.get(sku, DEFAULT_CURING_CT)
            shift_need = _cure_qty_per_shift(ct_sku)
            if gt_inv < shift_need:
                continue  # no GT ready → skip Class B to avoid starvation
        after_days = rem / ((n + 1) * rate)
        prio       = priority_score_map.get(sku, 0.0)
        gt_signal  = min(gt_inv, rate)  # cap at 1 day's rate
        if _prio_on:
            _pk = (0, float(_dd)) if _dd is not None else (1, 0.0)   # committed first, EDF
            candidates.append((_pk, urgency_class, after_days, -prio, -gt_signal, sku))
        else:
            candidates.append((urgency_class, after_days, -prio, -gt_signal, sku))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][-1]


def _select_ratio_co_target(
    old_sku: str,
    press: str,
    demand_remaining: dict,
    press_count: dict,
    pending_counts: dict,
    demand_dict: dict,
    cure_ct_map: dict,
    press_to_demand_targets: dict,
    press_total_demand: dict,
    horizon_left: int,
    sku_to_press_count: dict | None = None,
    rich_ranking: bool = False,
) -> "str | None":
    """Select the next SKU for a press whose current SKU demand just hit zero
    (_RATIO_CO_ALLOCATION_ENABLED). Eligibility timing is untouched by
    rich_ranking — this is called ONLY once a press has already fully
    exhausted its current SKU (no urgency/Class A/B/horizon-threshold
    early-CO, per spec); rich_ranking only changes which candidate wins
    among those already eligible right now.

    rich_ranking=False (default): ranked purely by static ratio
    (demand_dict[target]/press_total_demand[press], same formula as the
    building scheduler's NRI ranking, never decremented).

    rich_ranking=True (_RATIO_CO_RICH_RANKING_ENABLED): adds COScheduler's
    other four ranking levels on top of ratio — Class A/B (is the target
    under-resourced relative to its own deadline?), after_days (prefer
    targets closest to completion once this press joins), cycle time
    (faster-curing preferred), and target-side scarcity (SKU with fewer
    compatible presses preferred — protects thin SKUs). Still only reorders
    candidates that pass the presses-needed eligibility gate below; never
    makes a press eligible sooner than "current SKU fully done."

    Before ranking (either mode), a press-allocation tracker skips any SKU
    that already has enough presses (assigned + in-flight/pending this
    evaluation window) to finish its remaining demand within the remaining
    horizon — so a free press is never handed to an already-adequately-served
    SKU merely because it ranks well.

    Compatibility is enforced by construction: only SKUs in
    press_to_demand_targets[press] (the physical mould/allowable-machines
    mapping) are ever considered.
    """
    candidates: list[tuple] = []
    for target in press_to_demand_targets.get(press, []):
        if target == old_sku:
            continue
        rem = demand_remaining.get(target, 0.0)
        if rem <= 0:
            continue
        ct   = cure_ct_map.get(target, DEFAULT_CURING_CT)
        rate = _cure_qty_per_shift(ct) * 3  # per-press per-day rate
        if rate <= 0 or horizon_left <= 0:
            continue
        effective_n = press_count.get(target, 0) + pending_counts.get(target, 0)
        presses_needed = max(1, math.ceil(rem / (rate * horizon_left)))
        if effective_n >= presses_needed:
            continue  # already has enough presses assigned — try next-highest ratio
        ratio = demand_dict.get(target, 0.0) / press_total_demand.get(press, 1e-9)
        if rich_ranking:
            current_days = rem / (effective_n * rate) if effective_n > 0 else float("inf")
            cls = 0 if current_days > horizon_left * CO_CLASS_B_THRESHOLD else 1
            after_days = rem / ((effective_n + 1) * rate)
            scarcity = (sku_to_press_count or {}).get(target, 0)
            candidates.append((cls, -ratio, after_days, ct, scarcity, target))
        else:
            candidates.append((-ratio, target))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][-1]


def _score_co_candidate(
    target: str,
    press: str,
    demand_remaining: dict,
    press_count: dict,
    cure_ct_map: dict,
    priority_score_map: dict,
    demand_dict: dict,
    press_total_demand: dict,
    sku_campaign_tier: dict,
    nri_skus: frozenset,
    horizon_left: int,
) -> tuple:
    """Score a (press, target) CO candidate for the dynamic day-planner
    (_DYNAMIC_CO_PLANNER_ENABLED). Mirrors curing_consumption_dynamic.py's
    _urgency_sort_key/_priority_signal: RI keeps Priority_Score-driven
    urgency untouched, NRI ranked by static demand[target]/press_total_demand[press]
    (confirmed win this session — same formula, duplicated here rather than
    imported, to keep COScheduler's own tie-break ordering completely
    unperturbed). sku_campaign_tier only applied as a tiebreak, and only when
    _CAMPAIGN_TIER_TIEBREAK_ENABLED — day-granularity, never a primary signal.
    Final (press, target) pair is always the last tuple element for full
    determinism regardless of dict/set iteration order upstream.
    """
    rem  = demand_remaining.get(target, 0.0)
    n    = press_count.get(target, 0)
    ct   = cure_ct_map.get(target, DEFAULT_CURING_CT)
    rate = _cure_qty_per_shift(ct) * 3
    current_days = rem / (n * rate) if (n > 0 and rate > 0) else float("inf")
    cls = 0 if current_days > horizon_left * CO_CLASS_B_THRESHOLD else 1
    after_days = rem / ((n + 1) * rate) if rate > 0 else float("inf")

    if target in nri_skus:
        signal = -(demand_dict.get(target, 0.0) / press_total_demand.get(press, 1e-9))
    else:
        signal = -priority_score_map.get(target, 0.0)

    tier = sku_campaign_tier.get(target, 99) if _CAMPAIGN_TIER_TIEBREAK_ENABLED else 0

    return (cls, signal, after_days, tier, press, target)


def _plan_day_cos(
    day: int,
    press_state: dict,
    demand_remaining: dict,
    press_count: dict,
    cure_ct_map: dict,
    priority_score_map: dict,
    demand_dict: dict,
    press_to_demand_targets: dict,
    press_total_demand: dict,
    sku_campaign_tier: dict,
    ri_skus: frozenset,
    nri_skus: frozenset,
    daily_co_count: dict,
    planning_days: int,
) -> list:
    """Dynamic per-day CO planner (_DYNAMIC_CO_PLANNER_ENABLED). Replaces the
    static co_by_day.get(day, []) lookup with a decision grounded in real
    live state, made fresh each day before Shift A — same real-state
    discipline as the existing reactive dynamic_co_tracker mechanism
    (b2c_pipeline.py _select_dynamic_co_target), extended from "one press
    reactively" to "the whole eligible press fleet, once a day."
    """
    horizon_left = _bc_working_days_left(day, planning_days)   # holiday-aware urgency horizon
    slots_left = MAX_CHANGEOVERS_PER_DAY - daily_co_count.get(day, 0)
    if slots_left <= 0:
        return []

    # Presses whose current SKU has no remaining demand — generalizes RO +
    # just-finished-RI into one real-state check (sorted: deterministic seed
    # order regardless of press_state's dict iteration order). NOTE: no
    # early-return when this is empty — the donation pass below doesn't need
    # any press to be pre-idle (it creates capacity by pulling from a healthy
    # RI SKU), and gating it behind newly_free caused a severe regression
    # (670k->593k): most days have zero naturally-idle presses, so the
    # donation pass — the mechanism actually capable of proactive
    # reallocation — never got a chance to run on those days at all.
    newly_free = sorted(
        p for p, st in press_state.items()
        if demand_remaining.get(st["sku"], 0.0) <= 0
    )

    # Local working copies, mutated as assignments are made this pass, so a
    # target already sufficiently covered by an earlier pick this same day
    # stops absorbing further presses (mirrors COScheduler's own
    # "re-check with CURRENT press_count" guard).
    _press_count = dict(press_count)

    candidates = []
    for p in newly_free:
        old_sku = press_state[p]["sku"]
        for target in press_to_demand_targets.get(p, []):
            if target == old_sku:
                continue
            rem = demand_remaining.get(target, 0.0)
            if rem <= 0:
                continue
            is_nri = target in nri_skus
            is_ri  = target in ri_skus
            if is_nri:
                pass  # always eligible
            elif is_ri and _press_count.get(target, 0) > 0:
                n_t = _press_count.get(target, 0)
                ct_t = cure_ct_map.get(target, DEFAULT_CURING_CT)
                rate_t = _cure_qty_per_shift(ct_t) * 3
                current_days = rem / (n_t * rate_t) if rate_t > 0 else float("inf")
                if current_days <= horizon_left:
                    continue  # RI already on track — skip
            else:
                continue
            key = _score_co_candidate(
                target, p, demand_remaining, _press_count, cure_ct_map,
                priority_score_map, demand_dict, press_total_demand,
                sku_campaign_tier, nri_skus, horizon_left,
            )
            candidates.append((key, p, old_sku, target))

    candidates.sort(key=lambda x: x[0])

    today_events: list = []
    assigned_press: set = set()
    for key, p, old_sku, target in candidates:
        if slots_left <= 0:
            break
        if p in assigned_press:
            continue
        rem = demand_remaining.get(target, 0.0)
        if rem <= 0:
            continue
        if target in ri_skus and target not in nri_skus:
            n_t = _press_count.get(target, 0)
            if n_t > 0:
                ct_t = cure_ct_map.get(target, DEFAULT_CURING_CT)
                rate_t = _cure_qty_per_shift(ct_t) * 3
                if rate_t > 0 and rem / (n_t * rate_t) <= horizon_left:
                    continue  # became sufficiently covered by an earlier pick this pass
        today_events.append((p, old_sku, target))
        assigned_press.add(p)
        _press_count[old_sku] = max(0, _press_count.get(old_sku, 0) - 1)
        _press_count[target]  = _press_count.get(target, 0) + 1
        slots_left -= 1

    # Donation pass: any Class-A-critical target (RI or NRI) may pull a spare
    # press from an RI SKU that can safely give one up (n>1, still meets its
    # own full demand within the horizon with n-1) — generalized from an
    # earlier NRI-only version. That narrower version left critically
    # under-served RI SKUs (some presses, but not enough to meet demand
    # within the horizon) with no path to get help, because the main pass
    # above only ever acts on presses that are ALREADY idle today — the root
    # cause of a severe regression (670k->593k) when first tested: reactive
    # "fill an idle press" logic alone never proactively reallocates capacity
    # the way COScheduler's upfront horizon-wide simulation does.
    if slots_left > 0:
        served_targets = {t for _, _, t in today_events}

        def _current_days(sku: str) -> float:
            rem  = demand_remaining.get(sku, 0.0)
            n    = _press_count.get(sku, 0)
            ct   = cure_ct_map.get(sku, DEFAULT_CURING_CT)
            rate = _cure_qty_per_shift(ct) * 3
            if n <= 0 or rate <= 0:
                return float("inf")
            return rem / (n * rate)

        needs_help = sorted(
            (
                s for s in (set(ri_skus) | set(nri_skus))
                if s not in served_targets
                and demand_remaining.get(s, 0.0) > 0
                and _current_days(s) > horizon_left * CO_CLASS_B_THRESHOLD
            ),
            key=lambda s: (-_current_days(s), -demand_remaining.get(s, 0.0), s),
        )
        if os.environ.get("DYNCO_DEBUG"):
            print(f"    [DynCO-debug] Day {day}: newly_free={len(newly_free)} "
                  f"needs_help={len(needs_help)} slots_left={slots_left} "
                  f"sample_needs_help={needs_help[:5]}")
        _donors_found = 0
        for target in needs_help:
            if slots_left <= 0:
                break
            donor = None
            for p in sorted(press_state.keys()):
                if p in assigned_press:
                    continue
                cur_sku = press_state[p]["sku"]
                if cur_sku == target or cur_sku not in ri_skus:
                    continue
                if target not in press_to_demand_targets.get(p, []):
                    continue
                n = _press_count.get(cur_sku, 0)
                if n <= 1:
                    continue
                rem_cur  = demand_remaining.get(cur_sku, 0.0)
                ct_cur   = cure_ct_map.get(cur_sku, DEFAULT_CURING_CT)
                rate_cur = _cure_qty_per_shift(ct_cur) * 3
                if rate_cur > 0 and rem_cur / ((n - 1) * rate_cur) <= horizon_left:
                    donor = (p, cur_sku)
                    break
            if donor is not None:
                p, cur_sku = donor
                today_events.append((p, cur_sku, target))
                assigned_press.add(p)
                _press_count[cur_sku] = max(0, _press_count.get(cur_sku, 0) - 1)
                _press_count[target]  = _press_count.get(target, 0) + 1
                slots_left -= 1
                _donors_found += 1
        if os.environ.get("DYNCO_DEBUG") and needs_help:
            print(f"    [DynCO-debug] Day {day}: donors_found={_donors_found}/{len(needs_help)}")

    return today_events


def _rolling_horizon_co_call(
    day: int,
    planning_days: int,
    press_state: dict,
    press_count: dict,
    demand_remaining: dict,
    demand_dict: dict,
    priority_score_map: dict,
    df_allowable: "pd.DataFrame",
    ct_map: dict,
    dynamic_co_tracker: dict,
    scheduler: "COScheduler",
    max_co_per_day: int = MAX_CHANGEOVERS_PER_DAY,
) -> list:
    """Rolling-Horizon COScheduler call (_ROLLING_HORIZON_CO_ENABLED). Reuses
    the existing, proven COScheduler.schedule() unchanged — only synthesizes
    fresh Day-0-shaped inputs from live rolling-pipeline state each day, with
    a shrinking remaining horizon, and keeps only that call's relative-day-1
    events (the rest of its lookahead is discarded and recomputed tomorrow).

    Finding 2: presses mid-transition in dynamic_co_tracker are excluded from
    this day's snapshot — press_count already reflects them, but press_state
    won't until the shift their transition resolves; including them would let
    schedule() assign them a second, conflicting CO.

    Finding 1: Category mirrors SKUClassifier exactly — press_count>0 is
    always "Runner-In" for a demand SKU (even at zero remaining demand;
    schedule()'s own demand_done_free logic handles that case with the
    any-class bypass it already has). Only a non-demand SKU occupying a press
    is "Runner-Out".
    """
    horizon_left = _bc_working_days_left(day, planning_days)   # holiday-aware urgency horizon
    demand_skus = set(demand_remaining.keys())

    running_rows = [
        {"Machine": p, "SKUCode": st["sku"]}
        for p, st in sorted(press_state.items())
        if p not in dynamic_co_tracker
    ]
    df_running_moulds = pd.DataFrame(running_rows, columns=["Machine", "SKUCode"])

    day0_rows = [
        {
            "SKUCode": sku,
            "Category": "Runner-In" if press_count.get(sku, 0) > 0 else "Non-Runner-In",
            "Running_Press_Count": press_count.get(sku, 0),
        }
        for sku in sorted(demand_skus)
    ]
    ro_counts: dict[str, int] = defaultdict(int)
    for p, st in sorted(press_state.items()):
        if p in dynamic_co_tracker:
            continue
        if st["sku"] not in demand_skus:
            ro_counts[st["sku"]] += 1
    for sku in sorted(ro_counts):
        day0_rows.append({
            "SKUCode": sku, "Category": "Runner-Out",
            "Running_Press_Count": ro_counts[sku],
        })
    df_day0 = pd.DataFrame(day0_rows, columns=["SKUCode", "Category", "Running_Press_Count"])

    df_demand = pd.DataFrame(
        [
            {
                "SKUCode": sku,
                "Quantity": max(0.0, demand_remaining.get(sku, 0.0)),
                "Priority": priority_score_map.get(sku, 0.0),
            }
            for sku in sorted(demand_skus)
        ],
        columns=["SKUCode", "Quantity", "Priority"],
    )

    if os.environ.get("ROLLCO_DEBUG"):
        print(f"    [RollCO-debug] Day {day}: horizon_left={horizon_left} "
              f"presses={len(df_running_moulds)} skus={len(demand_skus)}")

    co_events = scheduler.schedule(
        df_day0=df_day0,
        df_demand=df_demand,
        df_allowable=df_allowable,
        df_running_moulds=df_running_moulds,
        ct_map=ct_map,
        max_co_per_day=max_co_per_day,
        planning_days=horizon_left,
        ratio_demand_map=demand_dict,
    )
    today = sorted(
        (ev["press"], ev["old_sku"], ev["new_sku"])
        for ev in co_events if ev["day"] == 1
    )
    if os.environ.get("ROLLCO_DEBUG"):
        print(f"    [RollCO-debug] Day {day}: {len(today)} events {today}")
    return today


def _assign_building_shift(
    shift_cure_demand:      dict,
    machine_skus:           dict,
    machine_current_sku:    dict,
    sku_inch:               dict,
    demand_remaining:       dict,
    gt_inventory:           dict,
    machine_pool:           dict,
    machine_minutes_on_sku: dict,
    cure_ct_map:            dict,
    press_count:            dict,
    co_target_skus:         frozenset = frozenset(),
    days_left:              int = 31,
    demand_dict:            dict | None = None,
    machine_total_demand:   dict | None = None,
    machine_anchor_inch:    dict | None = None,
    machine_used_inches:    dict | None = None,
    machine_inch_since:     dict | None = None,   # {machine: day its current inch began}
    day:                    int = 1,              # current plan day (for the 5-day dwell)
    machine_day_skus:       dict | None = None,   # {machine: set(SKUs built today)} — 4/day cap
    machine_plus3_used:     set | None = None,    # machines that spent their +3/-3 escape
    machine_last_diff_co_day: dict | None = None, # {machine: last day it did a diff-size CO}
    machine_locked_inches: dict | None = None,   # Part 1: {machine: set(allowed inches)} or None
    machine_left_skus:     dict | None = None,   # SKU_NO_REVERT: {machine: set(SKUs it has LEFT)}
    machine_day_diff_co: dict | None = None,     # Part 2: {machine: #diff-size COs done today}
    machine_day_co: dict | None = None,          # S2_CAMPAIGN: {machine: #building COs done today}
    fixed_escape_used: dict | None = None,       # Lever B: {machine: #escape diff-COs spent}
    machine_step_drift: dict | None = None,      # INCH_STEP_DRIFT: {machine: signed cumulative drift}
    machine_db_skus: dict | None = None,         # INCH_STEP_DRIFT: {machine: un-stripped DB SKUs}
    lookahead_draw: dict | None = None,          # LOOKAHEAD_BUF: {sku: anticipated peak draw today}
    priority_deadline_map: dict | None = None,   # DELIVERY_PRIORITY: {sku: deadline_day} or None
    priority_dated_skus: set | None = None,      # DELIVERY_PRIORITY: committed SKUs that HAVE a Delivery Date
    writeoff_cum: dict | None = None,            # R8B: cumulative expired GT per SKU (cap tightener)
    fwd_work_shifts: int | None = None,          # #2 NO-PERISH: working shifts inside the GT shelf window (None → full 9)
    bridge_shifts: int = 0,                      # #3 BRIDGE: consecutive holiday shifts imminent (0 → no bridge)
    shift_budget_mins: int | None = None,        # #7 HOLIDAY_SHIFTC_CAP: per-shift minute budget override (None → SHIFT_MINS)
    machine_down_mins: dict | None = None,        # PM/MTC: per-machine maintenance minutes lost THIS shift
    sku_curable_ceiling: dict | None = None,      # BLD_CURABLE_CAP: {sku: curable GT-on-hand stock cap} or None
    shift_idx: int = 0,                            # BLD_SEED_PIN_D1A: 0=A,1=B,2=C — for the Day-1 Shift-A seed pin
    pacing_day_built: float = 0.0,                 # PACING: GT already built earlier today (prior shifts)
    pacing_day_skus: set | None = None,            # PACING: SKUs already built today (widen the active set)
    sku_home_group: dict | None = None,            # SAME_GROUP: {sku: home_group} or None (feature inert)
    sku_grp_target: dict | None = None,            # SG_DELIB: {sku: frozenset(allowed groups)} (deliberate) or None
    sku_cur_group: dict | None = None,             # SG_DELIB: {sku: current finer group} (mutable)
    sku_last_group_move: dict | None = None,       # SG_DELIB: {sku: day of last group MOVE} (cooldown)
    machine_plant_set: dict | None = None,         # PLANT_SET_LOCK: {GT-machine: set(plant Days-1-2 SKUs)} or None
) -> dict:
    """
    Greedy per-SHIFT building assignment.

    Plant-accurate behaviour:
    - CO can happen in any shift (A/B/C) whenever a deficit SKU exists and CO cost fits.
    - co_target_skus: NRI SKUs with curing CO firing today — the 30% guard is bypassed
      for urgent ones (0 GT inventory + active demand) in Campaign 2+.
    - Dominant-inch priority: when selecting CO candidates, SKUs matching the
      machine's dominant inch sort before other eligible inches.
    - All machines: MAX_BUILDING_COS_PER_MACHINE_PER_SHIFT COs per shift.

    Returns: {machine: [(sku, qty_int, co_type_str)]}
      co_type: "start" | "same_size_CO" | "diff_size_CO"
    """
    # R8B cap tightener: demand-cap ceilings subtract cumulative expired GT so a SKU is
    # never rebuilt to replace written-off (wasted) GT → total built <= demand.
    _woc = writeoff_cum if writeoff_cum is not None else {}
    # #7 HOLIDAY_SHIFTC_CAP: this shift's per-machine minute budget (None → full SHIFT_MINS).
    _sbud = float(shift_budget_mins) if shift_budget_mins is not None else float(SHIFT_MINS)
    # Client inch-rule state (persisted across shifts by run_rolling_pipeline).
    machine_anchor_inch = machine_anchor_inch if machine_anchor_inch is not None else {}
    machine_used_inches = machine_used_inches if machine_used_inches is not None else {}
    machine_inch_since  = machine_inch_since  if machine_inch_since  is not None else {}
    machine_day_skus    = machine_day_skus    if machine_day_skus    is not None else {}
    machine_plus3_used  = machine_plus3_used  if machine_plus3_used  is not None else set()
    machine_last_diff_co_day = (machine_last_diff_co_day
                                if machine_last_diff_co_day is not None else {})
    machine_locked_inches = machine_locked_inches if machine_locked_inches is not None else {}
    machine_left_skus = machine_left_skus if machine_left_skus is not None else {}
    def _sku_revert_ok(_m: str, _s: str) -> bool:
        # SKU_NO_REVERT: block a machine from re-building a SKU it has already LEFT (its current SKU is
        # never "left", so continuation is always fine). Off → always True (bit-for-bit).
        if not _SKU_NO_REVERT:
            return True
        return _s not in machine_left_skus.get(str(_m), ())
    _sku_home = sku_home_group if sku_home_group is not None else {}
    # ── SG_DELIB (deliberate + stable group allocation) local state ──────────────────
    # Active only when SG_DELIB is on AND a target map was supplied (day 3+). OFF/None →
    # every helper below is inert (returns 0 / no-op) → bit-for-bit.
    _grp_tgt      = sku_grp_target      if sku_grp_target      is not None else {}
    _grp_cur      = sku_cur_group       if sku_cur_group       is not None else {}
    _grp_lastmove = sku_last_group_move if sku_last_group_move is not None else {}
    _sg_delib = bool(_SG_DELIB and sku_grp_target is not None and day > _PLANT_2DAY_DAYS)

    def _sg_multi_ok(_m: str, _s: str) -> bool:
        """A SKU with a sanctioned MULTI-group TARGET SET (≥2 groups, frozen at day 3) may build in
        ALL of its target groups — including SIMULTANEOUSLY in the same shift (planned parallel
        multi-group, e.g. an oversubscribed-inch 13"/15" SKU whose one group saturates). So the
        same-shift HARD guard is BYPASSED for a machine whose group is IN the SKU's multi-group
        target. Stable / no ping-pong: the set is frozen, never widened at runtime."""
        _t = _grp_tgt.get(_s)
        return bool(_sg_delib and _t and len(_t) >= 2 and _sku_group_of(_m) in _t)

    def _group_pen(_m: str, _s: str) -> int:
        """SAME_GROUP soft penalty: 0 if machine `_m`'s finer group == SKU `_s`'s home
        group (or the SKU has no home), else `_SAME_GROUP_PEN`. Returns 0 for every
        candidate when the lever is OFF → inserting it as a tuple slot is order-preserving
        (bit-for-bit). SOFT: it is only ever a LATE tiebreaker (after the deficit/urgency
        tiers), so a starving/urgent SKU is still built cross-group when its home group has
        no free machine this shift — group purity never overrides demand or starvation.
        DELIBERATE mode routes grouping through _sg_move_pen instead → this returns 0."""
        if not _SAME_GROUP_SOFT or _sg_delib or _s in _SG_EXEMPT_SKUS:
            return 0
        _h = _sku_home.get(_s)
        if not _h:
            return 0
        return 0 if _sku_group_of(_m) == _h else _SAME_GROUP_PEN

    def _sg_move_pen(_m: str, _s: str) -> int:
        """DELIBERATE stable-group penalty for pair (machine, sku).
          0  — the machine's finer group is IN the SKU's deliberate target set (sanctioned);
          0  — a cross-group MOVE that the hysteresis ADMITS (SKU is starving in its target
               AND the per-SKU cooldown has elapsed) — a deliberate group-level decision;
          _HYST_BIG — every other cross-group pair (deprioritized so the pair loses → the
               SKU stays in its assigned group; no per-shift churn / ping-pong).
        Inert (0) unless deliberate mode is active and the SKU has a target set.
        Exempt SKUs (SG_EXEMPT_SKUS) are unconstrained → always 0."""
        if not _sg_delib or _s in _SG_EXEMPT_SKUS:
            return 0
        _tgt = _grp_tgt.get(_s)
        if not _tgt:                       # no deliberate target → unconstrained (free)
            return 0
        _g = _sku_group_of(_m)
        if _g in _tgt:                     # sanctioned group → no penalty
            return 0
        # cross-group MOVE: blocked unless the deliberate move-gate is enabled AND the SKU's
        # target group is STRUCTURALLY behind — cumulative monthly gap > _SG_MOVE_GAP_SHIFTS
        # shifts of draw AND starving THIS shift (dead-band) AND the per-SKU cooldown elapsed.
        # Momentary per-shift lag alone is NOT enough (that keeps groups stable).
        if not _SG_MOVE_ADMIT:
            return _HYST_BIG
        _draw = shift_cure_demand.get(_s, 0.0)
        if _draw <= 0.0:
            return _HYST_BIG
        _proj = projected_gt.get(_s, 0.0)
        _starving = _proj < _draw * (1.0 + _SG_MOVE_BAND)
        _gap = demand_remaining.get(_s, 0.0) - _proj
        _behind = _gap > _draw * _SG_MOVE_GAP_SHIFTS
        _cool_ok = (day - _grp_lastmove.get(_s, -10**9)) >= _SG_MOVE_COOLDOWN_DAYS
        if _starving and _behind and _cool_ok:
            return 0                       # deliberate structural move admitted (waived)
        return _HYST_BIG

    def _sg_pair_blocked(_m: str, _s: str) -> bool:
        """HARD group purity: True when pair (machine, sku) is a NON-admitted cross-group move
        and SG_HARD is on → drop it from the candidate pool (machine idles rather than build a
        foreign SKU). Inert when deliberate mode / SG_HARD is off."""
        return bool(_sg_delib and _SG_HARD and _sg_move_pen(_m, _s) >= _HYST_BIG)

    def _sg_move_commit(_m: str, _s: str) -> None:
        """Record an ADMITTED cross-group move: sanction the new group permanently (add to
        the SKU's target set → stable, no ping-pong) and stamp the move day (cooldown)."""
        if not _sg_delib:
            return
        _tgt = _grp_tgt.get(_s)
        if not _tgt:
            return
        _g = _sku_group_of(_m)
        if _g in _tgt:
            return
        _grp_tgt[_s] = frozenset(_tgt | {_g})
        _grp_lastmove[_s] = day
    machine_day_diff_co = machine_day_diff_co if machine_day_diff_co is not None else {}
    machine_day_co = machine_day_co if machine_day_co is not None else {}
    fixed_escape_used = fixed_escape_used if fixed_escape_used is not None else {}
    machine_step_drift = machine_step_drift if machine_step_drift is not None else {}
    machine_db_skus    = machine_db_skus    if machine_db_skus    is not None else {}
    lookahead_draw     = lookahead_draw     if lookahead_draw     is not None else {}

    def _eff_draw(sku: str) -> float:
        """LOOKAHEAD_BUF: the draw used to SIZE pre-build (_dyn_H + forward-buffer gate) — the
        ANTICIPATED peak (running + incoming-CO presses today), not just this shift's draw. OFF →
        the current shift's draw, bit-for-bit."""
        _b = shift_cure_demand.get(sku, 0.0)
        return max(_b, lookahead_draw.get(sku, 0.0)) if _LOOKAHEAD_BUF else _b

    # PACING (small-buffer rotation): the per-SKU buffer HORIZON (shifts a machine banks
    # ahead of draw) shrinks from the full GT shelf (9) to PACING_BUFFER_SHIFTS (1-2) from
    # day 3 on. This is what makes each machine hold only a THIN buffer per SKU and rotate
    # to the next drawn SKU (breadth → more distinct SKUs/day, flatter, less overnight carry).
    # Days 1-2 are the plant replay (never here). OFF → full 9-shift shelf, bit-for-bit.
    _pace_shelf = (min(GT_SHELF_LIFE_SHIFTS, PACING_BUFFER_SHIFTS)
                   if (_PACING_ENABLED and day > _PLANT_2DAY_DAYS)
                   else GT_SHELF_LIFE_SHIFTS)

    def _inch_gate(m: str, to_inch: str, cur_inch: str) -> bool:
        """Rule 2 (+/-2 band) for a candidate (machine, to_inch), PLUS the Part-1 locked
        inch-set: when a machine has a locked set, it may only build/CO to an inch in it."""
        _lset = machine_locked_inches.get(m)
        if _lset and to_inch and to_inch not in _lset:
            return False
        # #5 REVERT-DWELL (anti-flip-flop / long campaigns): a machine may revert to an inch it already
        # LEFT only after dwelling >= REVERT_DWELL_DAYS on its current inch — blocks RAPID Size1→Size2→
        # Size1 flip-flops (measured 52% of unlock diff-COs) while still allowing legitimate rotation
        # after a real campaign, and a fresh step like 13→15→14 (14 never built) is always free. Hard
        # no-revert was measured −25k (strands flexible machines). Stage-2 exempt (max util).
        if (_REVERT_DWELL_DAYS > 0 and to_inch and cur_inch and to_inch != cur_inch
                and _MACHINE_GROUP.get(str(m), "") != "STAGE2"
                and to_inch in machine_used_inches.get(m, set())
                and (day - machine_inch_since.get(m, day)) < _REVERT_DWELL_DAYS):
            return False
        # #2 optimizer-choice: a DIRECT +3/-3 is NOT blocked — it is ALLOWED but priced at the 8h CO
        # (see _co_cost), so the scorer chooses the cheaper TWO-HOP (15→16→18) when the intermediate
        # is productive and only pays the 8h direct jump when it isn't. (BJ +3 already removed from its
        # set by #1; Stage-2 exempt.) No hard block here.
        if (_ONEWAY_INCH_ENABLED and to_inch and cur_inch
                and _MACHINE_GROUP.get(str(m), "") != "STAGE2"):
            # STEP-INCH (non-Stage-2): start at dominant, reach any DB-allowable inch, revisiting
            # allowed (13→15→14 is fine) — the ONLY restriction is that a SINGLE CO may jump at most
            # ONEWAY_MAX_JUMP (=2) inches. A +3/-3 must step through an intermediate inch (13→15→16,
            # each ≤2) OR be taken as an 8h direct CO (INCH_PLUS3). BJ therefore never gets a +3.
            # STAGE-2 skips this entirely (free diff-COs, max util).
            if to_inch != cur_inch:
                _t, _c = _inch_num(to_inch), _inch_num(cur_inch)
                if _t is not None and _c is not None and abs(_t - _c) > _ONEWAY_MAX_JUMP:
                    return False                       # no direct +3/-3 (two-hop or 8h instead)
            return True
        return _inch_ok(to_inch, cur_inch,
                        machine_anchor_inch.get(m, ""),
                        machine_used_inches.get(m, set()))

    def _fixed_esc_block(m: str, to_inch: str, cur_inch: str) -> bool:
        """Lever B — return True to BLOCK a candidate for a FIXED machine's escape rules:
          • same-inch moves are never blocked;
          • off-DOMINANT-inch is allowed ONLY (a) while it still has budget
            (< FIXED_ESCAPE_MAX_COS escape diff-COs spent) AND (b) the machine's own
            dominant-inch demand is DONE (no servable deficit) — else it stays on its inch;
          • once it has spent its budget it may only continue its CURRENT (escaped) inch.
        Non-fixed machines / lever off → never block here."""
        if not _FIXED_ESCAPE_ENABLED or str(m) not in _FIXED_MACHS_HIST:
            return False
        if to_inch == cur_inch:
            return False                     # same inch (incl. same-size CO) always allowed
        _spent = fixed_escape_used.get(str(m), 0)
        if _spent >= _FIXED_ESCAPE_MAX_COS:
            return True                      # budget gone → only its current inch remains
        _dom = _MACHINE_ALLOWED_INCHES.get(str(m), [cur_inch])[0]
        _eff_cur = cur_inch or _dom          # an empty machine is treated as on its dominant
        if to_inch == _dom and _eff_cur == _dom:
            return False                     # seeding/continuing the dominant inch is always fine
        if _eff_cur != _dom:
            return True                      # already off dominant, budget left → no 2nd hop
        # On dominant with budget: may escape only once the dominant inch's whole-month
        # servable demand is truly EXHAUSTED — use the real monthly gap
        # (demand_remaining − projected_gt), NOT the momentary buffered deficit (_defc),
        # so a machine that is merely built-ahead does NOT abandon an inch it still owes.
        _dom_gap = sum(max(0.0, demand_remaining.get(_s, 0.0) - projected_gt.get(_s, 0.0))
                       for _s in machine_skus.get(m, ()) if sku_inch.get(_s, "") == _dom)
        return _dom_gap > MIN_CAMPAIGN_UNITS  # dominant still owes a real campaign → block escape

    def _sku_cap_blocks(m: str, sku: str, shift_skus) -> bool:
        """Plant 4-SKU/day rule: block a CO that would introduce a 5th DISTINCT SKU on
        machine `m` today. `shift_skus` = SKUs already in this shift's plan. Carryover +
        prior shifts live in machine_day_skus[m]. Revisiting an already-built SKU is free."""
        if not _BLD_SKU_CAP_ENABLED:
            return False
        _distinct = machine_day_skus.get(m, set()) | set(shift_skus)
        _cap = MAX_BUILDING_SKUS_PER_DAY
        # S2_CAMPAIGN: a tighter distinct-SKU/day cap on Stage-2 GT machines only —
        # spreads SKUs and cuts round-trips. OFF/non-Stage-2 → plant-wide cap (no-op).
        if _S2_CAMPAIGN and _MACHINE_GROUP.get(str(m), "") == "STAGE2":
            _cap = min(_cap, _S2_SKU_CAP)
        return sku not in _distinct and len(_distinct) >= _cap

    def _s2_co_budget_blocks(m: str) -> bool:
        """S2_CAMPAIGN per-day CO budget: a Stage-2 machine may do at most
        _S2_MAX_CO_PER_DAY building COs per calendar day (0 = disabled). Once spent,
        it can only continue its current SKU. Non-Stage-2 / OFF → never blocks."""
        if not (_S2_CAMPAIGN and _S2_MAX_CO_PER_DAY > 0):
            return False
        if _MACHINE_GROUP.get(str(m), "") != "STAGE2":
            return False
        return machine_day_co.get(m, 0) >= _S2_MAX_CO_PER_DAY

    def _may_leave_inch(m: str, cur_inch: str, defc, buf, rate: float = 0.0) -> bool:
        """Plant 5-day rule: a machine may change to a DIFFERENT inch only if its
        current size's servable demand is already done (deficit-done override → leave
        early), OR it has dwelled >= MIN_INCH_DWELL_DAYS on the current inch."""
        if not _INCH_RULES_ENABLED or not cur_inch:
            return True
        if _JIT_INCH:
            # Part 2: no dwell — a machine may leave its inch any time. Churn is controlled at the
            # candidate level by _jit_diff_ok (urgency margin + per-day budget + amortization),
            # not by a time gate. Leaving is always permitted here.
            return True
        if _inch_demand_done(m, cur_inch, machine_skus, sku_inch, defc, buf, rate,
                             demand_remaining=demand_remaining, projected_gt=projected_gt):
            if os.environ.get("INCH_DEBUG"): _INCH_DBG[0] += 1     # deficit-done exits
            return True                                    # deficit-done → may leave early
        if _INCH_COOLDOWN_RULE:
            # "1 diff-size CO per machine per cooldown window" — clock runs from the machine's
            # last DIFF-size CO (not from arrival on the current inch). A machine that has not
            # changed size in ≥ cooldown days may change now; otherwise it is committed to its
            # current inch (same-inch SKU switches remain free). Re-entry to a left size is thus
            # only possible after the window (a return is itself a diff-CO).
            _ok = (day - machine_last_diff_co_day.get(m, -10**9)) >= _INCH_COOLDOWN_DAYS
        else:
            _ok = (day - machine_inch_since.get(m, day)) >= MIN_INCH_DWELL_DAYS
        if os.environ.get("INCH_DEBUG"): _INCH_DBG[1 if _ok else 2] += 1  # dwell-pass / dwell-block
        return _ok

    def _diff_co_ok(m: str, to_inch: str) -> bool:
        """Diff-size-CO amortization gate: a machine may change to `to_inch` only if it has
        NOT done a diff-size CO in the last DIFF_CO_MIN_DWELL_DAYS days AND the target inch
        has ≥ DIFF_CO_MIN_TARGET_UNITS of sustained servable demand (real remaining demand,
        not the momentary buffer) — so the 88-180 min CO is amortized by a real campaign,
        not a bounce. Same-inch moves never reach here."""
        if day - machine_last_diff_co_day.get(m, -10**9) < DIFF_CO_MIN_DWELL_DAYS:
            return False
        tgt = sum(max(0.0, demand_remaining.get(_s, 0.0) - projected_gt.get(_s, 0.0))
                  for _s in machine_skus.get(m, ()) if sku_inch.get(_s, "") == to_inch)
        return tgt >= DIFF_CO_MIN_TARGET_UNITS

    def _jit_diff_ok(m: str, cur_inch: str, to_inch: str, buf: float) -> bool:
        """Part 2 (JIT) churn control for a DIFF-inch candidate. Allows the switch only when it is
        genuinely worth it, with no time-dwell:
          - per-machine per-day diff-CO BUDGET not yet spent;
          - URGENCY MARGIN: target inch's aggregate curing-draw deficit exceeds the machine's
            CURRENT inch residual deficit by _JIT_URGENCY_MARGIN (demand-adaptive anti-bounce);
          - AMORTIZATION: target inch has ≥ DIFF_CO_MIN_TARGET_UNITS of sustained remaining demand.
        Same-inch moves never reach here."""
        if not _JIT_INCH:
            return True
        # Part 2: BJ + VMI use a tighter margin/budget under the group policy (few diff-COs, keep
        # the ±2 band). US is hard-locked (0). Stage-2 stays on the normal JIT (flexible).
        if _GROUP_INCH_POLICY and _MACHINE_GROUP.get(str(m), "") in ("BJ", "VMI"):
            _budget, _margin = _VMI_MAX_DIFF_CO_PER_DAY, _VMI_JIT_MARGIN
        else:
            _budget, _margin = _MAX_DIFF_CO_PER_MACHINE_PER_DAY, _JIT_URGENCY_MARGIN
        if machine_day_diff_co.get(m, 0) >= _budget:
            return False
        _skus = machine_skus.get(m, ())
        _cur_urg = sum(_defc(_s, buf) for _s in _skus if sku_inch.get(_s, "") == cur_inch)
        _tgt_urg = sum(_defc(_s, buf) for _s in _skus if sku_inch.get(_s, "") == to_inch)
        if _tgt_urg <= _cur_urg + _margin:
            return False
        _sustained = sum(max(0.0, demand_remaining.get(_s, 0.0) - projected_gt.get(_s, 0.0))
                         for _s in _skus if sku_inch.get(_s, "") == to_inch)
        return _sustained >= DIFF_CO_MIN_TARGET_UNITS

    def _max_cos(mach: str) -> int:
        # Flex machines get extra CO budget so they can take an off-inch
        # excursion AFTER exhausting same-inch work (which uses the normal 2).
        if mach in _INCH_FLEX_MACHINES:
            return MAX_BUILDING_COS_PER_MACHINE_PER_SHIFT + _INCH_FLEX_EXTRA_COS
        return MAX_BUILDING_COS_PER_MACHINE_PER_SHIFT
    projected_gt: dict[str, float] = dict(gt_inventory)
    # GT actually carried INTO this shift (before any building) — the base for the
    # end-of-day 10k cap. Base Phase A/B build is cure-neutral (curing consumes it
    # this/next shift, so carry stays flat without a forward buffer); only Phase C's
    # forward buffer adds NET overnight carry, so we bound entry_carry + forward ≤ 10k.
    _entry_carry_gt = sum(v for v in gt_inventory.values() if v > 0)

    # ── DELIVERY_PRIORITY: committed-delivery building boost (EDF, behind-only) ──
    # A committed SKU that still needs GT built (demand_remaining > projected_gt) is
    # ranked FIRST among a machine's eligible pairs, earliest-deadline-first — so the
    # machine builds its GT ahead of its deadline. The boost operates ONLY within the
    # already inch-locked / demand-capped candidate set (it re-orders, never widens),
    # and drops to the identity constant (1,0.0) once the SKU is caught up or off the
    # map → OFF/empty = bit-for-bit baseline. Reads projected_gt live so it stops
    # boosting a SKU already filled this shift (never overbuilds past the cap).
    _pdm_bld: dict = {str(k): int(v) for k, v in (priority_deadline_map or {}).items()}
    _prio_dated_bld: set = {str(s) for s in (priority_dated_skus or _PRIO_DATED_SKUS)}  # committed SKUs WITH a Delivery Date
    _prio_on_bld = _DELIVERY_PRIORITY_ENABLED and _DP_BLD and bool(_pdm_bld)
    def _bld_prio(sku: str) -> tuple:
        if not _prio_on_bld:
            return (1, 0.0)
        dd = _pdm_bld.get(sku)
        if dd is None:
            return (1, 0.0)
        behind = demand_remaining.get(sku, 0.0) - projected_gt.get(sku, 0.0)
        return (0, float(dd)) if behind > 0 else (1, 0.0)

    def _delivery_relax(sku: str) -> bool:
        """AT-RISK soft-rule relaxation for a committed SKU (drives both the gate bypasses and the
        Phase-A preemption). RELAX only when the SKU is at RISK of missing its target under normal
        rules — i.e. it has fallen behind the linear pace to its deadline (built < demand·day/dd).
        While on pace, returns False → normal optimisation + ordering priority only (no relaxation).
        Split by SKU type + gated by the matching (mutually-exclusive) toggle:
          • DATED SKU     → gated by DELIVERY_DATE_ALL_SOFT_RULES_RELAXED, pace toward its date.
          • FLAG/UNDATED  → gated by PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED, pace toward month-end.
        KEEPS allowable/tooling + demand cap + mould feasibility. Never fires once demand is met."""
        if not _prio_on_bld:
            return False
        dd = _pdm_bld.get(sku)
        if dd is None:
            return False
        _rem = demand_remaining.get(sku, 0.0)
        if _rem <= 1e-9:                                   # demand met → nothing to relax
            return False
        # which mode governs this SKU, and is it enabled?
        _is_dated = sku in _prio_dated_bld
        if not (_DELIVERY_DATE_RELAX if _is_dated else _PRIORITY_FLAG_MONTHEND_RELAX):
            return False
        # AT-RISK = behind the linear build pace to the deadline `dd` (1-based plan day).
        _orig = float((demand_dict or {}).get(sku, _rem))
        if _orig <= 0 or day >= dd:                        # deadline reached/passed & still owed → relax
            return True
        _allowed_rem = _orig * (dd - day) / dd             # on-pace remaining by end of `day`
        return _rem > _allowed_rem + 1e-9                  # behind pace → at risk → relax

    if _GLOBAL_ASSIGN_ENABLED:
        # ══ Global machine-SKU scoring assignment (supersedes per-machine greedy) ══
        # Phase A: each machine continues its current SKU (no CO). Phase B: all
        # remaining (machine,SKU) pairs are scored together and assigned best-first,
        # with constraint=min(flex_machine,flex_sku) so constrained machines/SKUs win.
        def _buf_of(m: str) -> float:
            return (GT_BUFFER_SHIFTS_VMI if _MACHINE_GROUP.get(m, "") == "VMI"
                    else GT_BUFFER_SHIFTS_OTHER)

        # ── Phase-1 dynamic buffer precompute (DYN_BUFFER) ──────────────────────
        # feeders_s = building machines eligible for s (invert machine_skus);
        # contention_s = competing active SKUs sharing s's feeders, per feeder
        # (away-time proxy). Both structural → computed once per shift.
        _feeders: dict = defaultdict(set)
        for _mm, _sks in machine_skus.items():
            for _s in _sks:
                _feeders[_s].add(_mm)

        def _inch_penalty_ranked(_m: str, _to_inch: str) -> int:
            """Graded inch penalty from the machine's ranked dominant band: 0 for the
            dominant inch, k for the k-th band inch, len(band)+OFFBAND off-band.
            Replaces the binary 0/1 inch_penalty in the GLOBAL_SCORE_V2 ranking only."""
            _band = _MACHINE_DOMINANT_INCH_RANKED.get(str(_m), [])
            return _band.index(_to_inch) if _to_inch in _band else len(_band) + GS_INCH_OFFBAND

        _dyn_contention: dict = {}
        if _DYN_BUFFER_ENABLED:
            _active = {s for s in shift_cure_demand
                       if shift_cure_demand.get(s, 0.0) > 0
                       and demand_remaining.get(s, 0.0) > 0}
            for _s in _active:
                _fs = _feeders.get(_s, set())
                _comp = set()
                for _mm in _fs:
                    _comp |= (machine_skus.get(_mm, set()) & _active)
                _comp.discard(_s)
                _dyn_contention[_s] = len(_comp) / max(1, len(_fs))

        def _dyn_H(sku: str) -> float:
            """Per-SKU dynamic buffer horizon (shifts): floor·(1+α·contention+β·risk),
            clipped to [floor, GT_SHELF_LIFE_SHIFTS]. floor = VMI floor if any VMI
            feeds the SKU else OTHER floor."""
            _fs = _feeders.get(sku, set())
            floor = (DYN_BUF_FLOOR_VMI
                     if any(_MACHINE_GROUP.get(m, "") == "VMI" for m in _fs)
                     else DYN_BUF_FLOOR_OTHER)
            draw = _eff_draw(sku)                       # LOOKAHEAD_BUF: anticipated peak draw
            if draw <= 0:
                return float(floor)
            risk_short = max(0.0, 1.0 - projected_gt.get(sku, 0.0) / draw)
            H = floor * (1.0 + DYN_BUF_ALPHA * _dyn_contention.get(sku, 0.0)
                              + DYN_BUF_BETA * risk_short)
            return float(min(_pace_shelf, max(floor, round(H))))

        def _defc(sku: str, _b: float) -> float:
            if _DYN_BUFFER_ENABLED:
                _b = _dyn_H(sku)   # per-SKU dynamic horizon overrides the flat buffer
            # PACING: cap the buffer horizon to the thin small-buffer window (day 3+ only).
            # Guarded so OFF is bit-for-bit (never clamps the OFF-path buffers, incl >9 ones).
            if _PACING_ENABLED and day > _PLANT_2DAY_DAYS:
                _b = min(_b, _pace_shelf)
            built_ahead = projected_gt.get(sku, 0.0)
            gap = shift_cure_demand.get(sku, 0.0) * _b - built_ahead
            cap = demand_remaining.get(sku, 0.0) - built_ahead - _woc.get(sku, 0.0)
            return min(max(0.0, gap), max(0.0, cap))

        def _curable_cap(sku: str, _room: float) -> float:
            """BLD_CURABLE_CAP: clamp a proposed build `_room` (this-shift additional units)
            so the SKU's GT-on-hand can never exceed what its eligible curing presses can
            drain over the 3-day shelf → excess would age out as expired waste (invariant #4).

            Curable STOCK cap = (eligible curing presses for the SKU) × cure_rate/shift ×
            GT_SHELF_LIFE_SHIFTS. That is the most GT the SKU's presses could consume before
            it ages out — the max sensible on-hand. Uses the ELIGIBLE-press count (not the
            live running-press draw, so banking a buffer for presses that are between runs is
            preserved → no starvation; and not the mould-pair count, which the DB undercounts).
            Room = cap − current projected on-hand. A captive/single-source machine (ps2 →
            TUXPE, 1 press) is bounded to ~9 shifts of 1-press throughput instead of the full
            demand, so it stops accumulating GT the lone press can't reach; a machine feeding
            many live presses sees a large cap → builds freely. sku_curable_ceiling carries the
            precomputed stock cap per SKU; None ⇒ no cap. Never widens `_room`."""
            if not _BLD_CURABLE_CAP or sku_curable_ceiling is None:
                return _room
            stock_cap = sku_curable_ceiling.get(sku)
            if stock_cap is None:
                return _room
            return max(0.0, min(_room, stock_cap - projected_gt.get(sku, 0.0)))

        def _plus3_direct_ok(_m: str, _cur_inch: str, _to_inch: str, _buf: float) -> bool:
            """#2 EXPLICIT +3 route choice (deterministic, provably cheapest). A DIRECT +3/-3 jump
            (8h CO) is allowed ONLY when NO productive two-hop intermediate exists — i.e. no same-machine
            SKU sits on an inch strictly BETWEEN cur and target with a real deficit. If one does exist the
            machine must TWO-HOP through it (2 cheap ≤2 diff-COs + a productive stop ≪ one 8h CO), so the
            direct +3 candidate is suppressed here and the machine takes the intermediate first. Only
            active for non-BJ/non-Stage-2 when VMI_TWO_HOP is on; BJ has no +3 in its set (#1)."""
            if not _VMI_TWO_HOP or str(_m) in _BJ_MACHINES \
                    or _MACHINE_GROUP.get(str(_m), "") == "STAGE2":
                return True
            _c, _t = _inch_num(_cur_inch), _inch_num(_to_inch)
            if _c is None or _t is None or abs(_t - _c) <= 2:
                return True                                  # not a +3/-3 → unrestricted
            _lo, _hi = (_c, _t) if _t > _c else (_t, _c)
            for _s in machine_skus.get(_m, ()):
                _si = _inch_num(sku_inch.get(_s, ""))
                if _si is not None and _lo < _si < _hi and _defc(_s, _buf) > 0:
                    return False                             # productive intermediate → two-hop, not 8h
            return True                                      # no productive intermediate → 8h direct OK

        def _gt_headroom(fwd_added: float) -> float:
            # Strict room under the end-of-day 10k cap for the forward buffer. We do NOT
            # credit this shift's curing to the forward GT (conservative → even if
            # nothing cures, carry stays ≤ 10k). Base Phase A/B build is cure-neutral and
            # is excluded; only entry_carry + forward-added is bounded. Cap OFF ⇒ inf.
            if not _ENDOFDAY_GT_CAP_ENABLED:
                return float("inf")
            return max(0.0, MAX_ENDOFDAY_GT_INVENTORY - (_entry_carry_gt + fwd_added))

        # ── Phase-1 dynamic-buffer GT-cap guard ──────────────────────────────────
        # The dynamic buffer builds deeper than one shift of draw in Phase A/B; that
        # OVERNIGHT-CARRY EXCESS (build beyond 1 shift of draw) must stay under the 7k
        # end-of-day cap, exactly like the Phase-C forward buffer. _dyn_over accumulates
        # the excess committed this shift; the forward buffer's headroom then subtracts
        # it too (so the two share one budget). No-op when the dynamic buffer is off →
        # OFF-parity bit-for-bit (the flat buffer never breached the cap).
        _dyn_over = [0.0]
        # Curing drains GT during the day, so reserving the full entry-carry under-fills
        # the legal 7k. Credit DYN_BUF_CURE_CREDIT shifts of total curing draw back into
        # the dynamic buffer's headroom (retest verifies end-of-day GT stays ≤ cap).
        _dyn_cure_credit = DYN_BUF_CURE_CREDIT * sum(
            v for v in shift_cure_demand.values() if v > 0)

        def _dyn_headroom() -> float:
            """End-of-day GT headroom for the dynamic buffer, curing-credited."""
            if not _ENDOFDAY_GT_CAP_ENABLED:
                return float("inf")
            return max(0.0, MAX_ENDOFDAY_GT_INVENTORY + _dyn_cure_credit
                       - (_entry_carry_gt + _dyn_over[0]))

        def _dyn_cap_qty(sku: str, qty: int) -> int:
            """Trim a dynamic-buffer build so its overnight-carry excess keeps total
            end-of-day GT within MAX_ENDOFDAY_GT_INVENTORY. Updates _dyn_over. No-op
            when the dynamic buffer is off."""
            if not _DYN_BUFFER_ENABLED or qty <= 0:
                return qty
            _draw  = shift_cure_demand.get(sku, 0.0)
            _prior = projected_gt.get(sku, 0.0)
            _old_x = max(0.0, _prior - _draw)                 # excess already carrying
            _add   = max(0.0, _prior + qty - _draw) - _old_x  # this build's overnight add
            if _add <= 0:
                return qty
            _hr = _dyn_headroom()
            if _add > _hr:
                qty  = max(0, int(qty - (_add - _hr)))
                _add = max(0.0, max(0.0, _prior + qty - _draw) - _old_x)
            _dyn_over[0] += _add
            return qty

        def _tierg(sku: str, m: str, d: float) -> tuple:
            _is_ri = press_count.get(sku, 0) > 0
            _ri_ratio = (_RI_RATIO_ENABLED or _RI_RATIO_GLOBAL) and _is_ri
            if (_BUILDING_RATIO_ENABLED and (press_count.get(sku, 0) <= 0 or _ri_ratio)
                    and demand_dict is not None and machine_total_demand is not None):
                ratio = demand_dict.get(sku, 0.0) / machine_total_demand.get(m, 1e-9)
                return (0 if _is_ri else 1, -ratio)
            return (0, -d)

        machines = [m for m in machine_skus if _MACHINE_GROUP.get(m, "") != "STAGE1"]
        _mdown = machine_down_mins or {}
        stg = {
            m: {
                "remaining": max(0.0, _sbud - float(_mdown.get(str(m), 0.0))),   # PM/MTC downtime
                "co_count": 0, "max_cos": _max_cos(m),
                "cur_sku": machine_current_sku.get(m, ""),
                "rate": _bld_qty_per_shift(m) / SHIFT_MINS,
                # BLD_ACTUAL_SEED: a seeded machine's DOMINANT inch is its seed inch (empty map ⇒ OFF).
                "dom": _BLD_SEED_INCH.get(str(m)) or _MACHINE_DOMINANT_INCH.get(
                    str(m), sku_inch.get(machine_current_sku.get(m, ""), "")),
                "primary_done": True, "campaigns": [],
            }
            for m in machines
        }

        # BLD_SEED_PIN_D1A: GT machines HARD-pinned to their seed SKU this shift (Day-1 Shift-A
        # only). Populated in Phase A; these machines are then LOCKED OUT of Phase B and Phase C
        # so nothing reassigns their seed-SKU production. Empty on every other shift → no effect.
        _d1a_pin_lock: set = set()

        # CONCENTRATION (CONC_ALLOC): per-shift bookkeeping of which machines already serve
        # each SKU and how much has been committed to it this shift. Seeded by Phase A, grown
        # by Phase B/C. Drives _over_prov below. Untouched (and _over_prov returns 0) when the
        # lever is OFF → the selection keys are byte-identical to the committed baseline.
        _sku_shift_machs: dict = defaultdict(set)
        _sku_shift_qty: dict = defaultdict(float)

        def _over_prov(m, sku):
            """CONCENTRATION: 1 if giving `sku` an ADDITIONAL machine this shift would
            over-provision it, else 0. Deferral only (ranks below still-under-served SKUs) —
            never blocks, so a machine whose eligible SKUs are all paced still builds one (no
            forced idle). Rules: the first machine on a SKU this shift, or a machine already
            on it, is always 0. A STARVING SKU (on-hand GT < draw × _CONC_STARV_SHIFTS) is
            always 0 — the deviation override that lets multiple machines rescue a behind SKU.
            Otherwise 1 once this shift's committed build already keeps pace with the draw
            (committed_qty >= draw), which self-limits a SKU to ~ceil(draw/rate) machines."""
            if not _CONCENTRATION:
                return 0
            _machs = _sku_shift_machs.get(sku)
            if not _machs or m in _machs:
                return 0
            _draw = shift_cure_demand.get(sku, 0.0)
            if _draw <= 0:
                return 0
            if projected_gt.get(sku, 0.0) < _draw * _CONC_STARV_SHIFTS:
                return 0                                   # starving → admit extra machines
            return 1 if _sku_shift_qty.get(sku, 0.0) >= _draw else 0

        def _conc_commit(m, sku, qty):
            """Record a committed build against the per-shift concentration trackers."""
            if _CONCENTRATION and qty > 0:
                _sku_shift_machs[sku].add(m)
                _sku_shift_qty[sku] += qty

        # ── PLANT_SET_LOCK helpers: a GT machine is restricted to its plant Days-1-2 SKU set
        # until that set is demand-complete, then released. day<=2 = replay (never here). ────
        _mps = machine_plant_set or {}
        # SOLE-BUILDER orphans: demand SKUs whose ONLY allowable building machines are ALL
        # plant-locked to OTHER SKUs (the SKU is in none of their plant sets). Under the strict
        # spare gate no machine ever builds them → they starve (e.g. STMX0 / TUHL0(75)/(77)
        # allowable only on 7502+7503). These are allowed through the plant gate as spare.
        _sole_builder_skus: set = set()
        if _PLANT_SET_LOCK and day > _PLANT_SET_LOCK_FROM and _mps:
            _sku_feeders: dict = defaultdict(set)
            for _mm, _sk_set in machine_skus.items():
                for _sk in _sk_set:
                    _sku_feeders[_sk].add(str(_mm))
            for _sk, _feeds in _sku_feeders.items():
                if _feeds and all(_f in _mps and _mps[_f] and _sk not in _mps[_f] for _f in _feeds):
                    _sole_builder_skus.add(_sk)
        def _plant_set_done(_m) -> bool:
            _ps = _mps.get(str(_m))
            if not _ps:
                return True
            return all(demand_remaining.get(_s, 0) <= 0 for _s in _ps)
        def _plant_ok(_m, _sku) -> bool:
            # PLANT-FIRST, SPARE→OTHERS: a plant-set SKU is always allowed (fed first, never removed);
            # a NON-plant SKU is allowed only as SPARE — when no plant-set SKU still needs building
            # this shift (all their per-shift draw is met).
            if not (_PLANT_SET_LOCK and day > _PLANT_SET_LOCK_FROM):
                return True
            _ps = _mps.get(str(_m))
            if not _ps or _plant_set_done(_m):
                return True
            if _sku in _ps:
                return True
            # SOLE-BUILDER OVERRIDE: a SKU that is building-allowable ONLY on plant-locked machines
            # (no free builder anywhere) would otherwise be stranded by every such machine's lock —
            # e.g. STMX0 / TUHL0 allowable ONLY on 7502+7503, both of which are plant-locked to OTHER
            # SKUs. Let a plant-locked machine build such a sole-builder SKU as spare so it isn't
            # orphaned. Bounded by no-waste-GT (curable ceiling) as usual.
            if _sku in _sole_builder_skus:
                return True
            # STARVED-FEED (Fix C): feed a SAME-INCH, live-draw, STARVED non-plant SKU even when a
            # plant SKU still shows a nominal deficit (it's draw-capped). Same inch as the machine's
            # current campaign → cheap CO; genuinely starved (press empty) → no waste. Downstream
            # CO-cost / dwell gates still apply.
            if (_PLANT_STARVE_FEED
                    and sku_inch.get(_sku, "") == sku_inch.get(machine_current_sku.get(_m, ""), "\0")
                    and shift_cure_demand.get(_sku, 0.0) > 0
                    and gt_inventory.get(_sku, 0.0) <= 0
                    and projected_gt.get(_sku, 0.0) <= 0
                    and demand_remaining.get(_sku, 0.0) > 0):
                return True
            return not any(_defc(_p, _buf_of(_m)) > 0 for _p in _ps)

        # ── Phase A: continuation anchor (no CO) ──
        # BLD_SEED_STICKY: process a sticky-seed machine (still on its seed SKU, within the
        # sticky window) BEFORE its captive-max peers, so it claims its seed SKU's draw-bounded
        # deficit first and builds a consistent amount — the peers absorb the daily residual.
        def _phaseA_key(m):
            _st = (_BLD_SEED_STICKY and str(m) in _BLD_SEED_MACHINES
                   and machine_current_sku.get(m, "") == _BLD_SEED_SKU.get(str(m), "\0")
                   and day <= _BLD_SEED_STICKY_DAYS)
            return (0 if _st else 1, str(m))
        for m in sorted(machines, key=_phaseA_key):
            s = stg[m]; buf = _buf_of(m); rate = s["rate"]
            eligible = machine_skus.get(m, set()); dom = s["dom"]
            cur = s["cur_sku"]
            if cur and not _plant_ok(m, cur):
                cur = ""; s["cur_sku"] = ""   # PLANT_SET_LOCK: drop a non-plant carryover → re-seed from the set
            # seed empty machine with a dom-inch-preferred deficit SKU (== "start")
            if not cur:
                cands = [x for x in eligible if _defc(x, buf) > 0
                         and not _fixed_esc_block(m, sku_inch.get(x, ""), "")
                         and _sku_revert_ok(m, x) and _plant_ok(m, x)]
                # MID-MONTH N-DAY SET: on plan day 1 restrict the start to SKUs this machine
                # ACTUALLY built in the N days before PLAN_START. It ran all of them recently,
                # so starting on any of them is physical and costs no changeover; the ranking
                # below then picks the one with the best live need (draw / deficit / inventory).
                # Machines with no set (idle in the window) are unrestricted, as before.
                if day == 1 and _MIDMONTH_SET:
                    _ms = _MIDMONTH_SET.get(str(m)) or set()
                    if _ms:
                        _in_set = [x for x in cands if x in _ms]
                        if not _in_set:      # plant-set lock excluded them all -> relax that gate
                            _in_set = [x for x in eligible if x in _ms and _defc(x, buf) > 0
                                       and not _fixed_esc_block(m, sku_inch.get(x, ""), "")
                                       and _sku_revert_ok(m, x)]
                        if _in_set:
                            cands = _in_set
                        else:
                            # Nothing in the machine's recent set shows a deficit this shift.
                            # It was PHYSICALLY RUNNING at PLAN_START 07:00, so continue its
                            # last actual SKU rather than idle it — the plant would not stop
                            # the machine. Still a day-1 "start" ⇒ no changeover charged.
                            _last = _MIDMONTH_LAST.get(str(m), "")
                            if _last and _last in eligible:
                                cands = [_last]
                if cands:
                    # DELIVERY_PRIORITY: a behind committed SKU seeds an empty machine first
                    # (EDF), still dom-inch-filtered. (1,0.0) constant when off → identity.
                    cur = min(cands, key=lambda x: (
                        # DEDICATED_SKU_FIRST: a dedicated SKU outranks pure deficit so its
                        # Stage-1 partner actually gets carcass to build (see bc_config).
                        0 if (_DEDICATED_FIRST and x in _DEDICATED_SKUS) else 1,
                        0 if (day <= _EARLY_DAYS and x in _EARLY_SKUS) else 1,
                        _bld_prio(x),
                        0 if sku_inch.get(x, "") == dom else 1,
                        *_tierg(x, m, _defc(x, buf)), x))
                    s["cur_sku"] = cur
            # BLD_SEED_STICKY: re-anchor a sticky-seed machine onto its seed SKU each shift so a
            # prior-shift Phase-B backfill (which flips machine_current_sku to another SKU) can't
            # displace it. Only while the seed SKU is still same-inch (no phantom diff-CO), eligible,
            # and has curable deficit — so it keeps a consistent seed-SKU output through the window.
            if (_BLD_SEED_PIN and str(m) in _BLD_SEED_MACHINES
                    and day <= _BLD_SEED_STICKY_DAYS):
                _seed = _BLD_SEED_SKU.get(str(m), "")
                if (_seed and _seed != cur and _seed in eligible
                        and sku_inch.get(_seed, "") == sku_inch.get(cur, "")
                        and _defc(_seed, GT_SHELF_LIFE_SHIFTS) > 0):
                    cur = _seed; s["cur_sku"] = cur
            # BLD_SEED_PIN_D1A: HARD Day-1 Shift-A pin. Force cur = seed SKU UNCONDITIONALLY
            # (ignore inch-match / deficit) for a seeded GT machine, and lock it out of Phase B/C
            # below. Room uses the demand-cap (curable-bounded) path so it builds even when the
            # seed SKU has no live curing draw (_defc==0) — the plant was building it — but never
            # past the demand cap or the curable/GT ceiling (no runaway overbuild).
            _pin_d1a_m = False
            if (_BLD_SEED_PIN_D1A and day == 1 and shift_idx == 0
                    and str(m) in _BLD_SEED_MACHINES
                    and _MACHINE_GROUP.get(m, "") != "STAGE1"):
                _seedp = _BLD_SEED_SKU.get(str(m), "")
                if _seedp and _seedp in eligible and demand_remaining.get(_seedp, 0.0) > 0:
                    cur = _seedp; s["cur_sku"] = cur
                    _pin_d1a_m = True
                    _d1a_pin_lock.add(m)
                else:
                    print(f"  [Rolling] BLD_SEED_PIN_D1A: {m} seed {_seedp!r} NOT pinnable "
                          f"(eligible={_seedp in eligible}, "
                          f"dem_left={demand_remaining.get(_seedp, 0.0):.0f}) → normal alloc")
            cur_inch = sku_inch.get(cur, "")
            # round-trip buffer sizing (same as per-machine path)
            eff_buf = buf
            if _ROUND_TRIP_BUFFER_ENABLED and cur and len(eligible) > 1:
                pc = [x for x in eligible if x != cur
                      and demand_remaining.get(x, 0.0) > 0 and _defc(x, buf) > 0]
                if pc and _RT_SAME_INCH:                          # prefer a same-inch rotation partner
                    _si_pc = [x for x in pc if sku_inch.get(x, "") == cur_inch]
                    if _si_pc and (max(_defc(x, buf) for x in _si_pc)
                                   >= _RT_SAME_INCH_FRAC * max(_defc(x, buf) for x in pc)):
                        pc = _si_pc                               # same-inch partner is nearly as needy
                if pc:
                    if m in _INCH_FLEX_MACHINES:
                        partner = max(pc, key=lambda x: (
                            _co_cost(m, cur_inch, sku_inch.get(x, ""))
                            + _co_cost(m, sku_inch.get(x, ""), cur_inch), _defc(x, buf), x))
                    else:
                        partner = max(pc, key=lambda x: (_defc(x, buf), x))
                    p_inch = sku_inch.get(partner, "")
                    if _RT_PARTNER_RT and rate > 0:       # T1: size B's dwell to B's OWN round-trip
                        _a_dwell = max(MIN_CAMPAIGN_MINS, _defc(cur, buf) / rate)
                        _b_rt = _co_cost(m, p_inch, cur_inch) + _a_dwell + _co_cost(m, cur_inch, p_inch)
                        _p_buf = max(buf, _b_rt / SHIFT_MINS)
                    else:
                        _p_buf = buf
                    p_dwell = max(MIN_CAMPAIGN_MINS,
                                  _defc(partner, _p_buf) / rate if rate > 0 else MIN_CAMPAIGN_MINS)
                    rt = _co_cost(m, cur_inch, p_inch) + p_dwell + _co_cost(m, p_inch, cur_inch)
                    eff_buf = max(buf, rt / SHIFT_MINS)
            flex_reclaim = (m in _INCH_FLEX_MACHINES and cur_inch != dom
                            and any(sku_inch.get(x, "") == dom and _defc(x, buf) > 0
                                    for x in eligible)
                            # BLD_ACTUAL_SEED: never abandon the seed on Day-1 (no Day-1 reshuffle).
                            and not (day == 1 and str(m) in _BLD_SEED_MACHINES))
            # Captive-max experiment: a captive machine (only 1 eligible SKU) builds
            # its sole SKU to the full demand cap (not just the buffer) so it runs
            # flat-out and never idles while its SKU has unmet demand.
            _cap_max = (_CAPTIVE_MAX_ENABLED and len(eligible) == 1
                        and _MACHINE_GROUP.get(m, "") != "STAGE1")
            # BLD_SEED_STICKY: a sticky-seed machine on its seed SKU fills its whole shift with
            # that SKU (steady output, like a captive machine) instead of building only a thin
            # draw-share and letting Phase B backfill it with varying other SKUs. The room uses
            # the full 3-day GT-shelf buffer via _defc, which is bounded by BOTH the curable draw
            # over the shelf AND (demand − built − writeoff) — so a curing-limited seed SKU gets
            # only a small room (no overbuild, no extra expiry) while a high-draw one (LSTL0-12)
            # fills the shift steadily. Draw/shelf/demand-cap all respected; window-gated.
            _sticky_now = (_BLD_SEED_STICKY and str(m) in _BLD_SEED_MACHINES
                           and cur == _BLD_SEED_SKU.get(str(m), "\0")
                           and day <= _BLD_SEED_STICKY_DAYS
                           and _MACHINE_GROUP.get(m, "") != "STAGE1")
            s["eff_buf"] = eff_buf                        # T2: stored for the Phase-B rotation gate
            _base_buf = buf if _RT_IMMINENT else eff_buf  # T2: Phase A builds only flat when RT_IMMINENT
            # ── DELIVERY PREEMPTION ────────────────────────────────────────────────
            # A machine ALLOWABLE for an AT-RISK committed SKU, whose CURRENT sku is NOT itself an
            # at-risk committed SKU, ABANDONS its continuation (room→0) so Phase-B CO's it onto the
            # committed SKU this shift (e.g. 7105/7106 drop the huge non-committed SUNE1 to build the
            # committed SXC1T/TUNE6 when they fall behind pace to their deadline). Only fires when a
            # committed SKU is AT RISK (via _delivery_relax) — not for on-pace SKUs. EDF ordering
            # comes from _bld_prio; allowable/tooling + demand cap + mould feasibility unchanged.
            # A machine plant-locked to `cur` cannot actually be CO'd onto the committed SKU
            # (its plant set forbids the switch), so preempting it just idles it (it abandons cur
            # AND can't build the committed SKU — e.g. 7106 locked to SUNE1 while SXC1T/TUNE6 are
            # at-risk). Only preempt if the machine can REACH a committed SKU under the plant-lock
            # (_plant_ok). PLANT_SET_LOCK off → _plant_ok always True → bit-for-bit prior behaviour.
            _preempt_delivery = (_ANY_DELIVERY_RELAX and _prio_on_bld
                                 and not _delivery_relax(cur)
                                 and any(_delivery_relax(x) and _plant_ok(m, x)
                                         for x in machine_skus.get(m, set())))
            _room = (0.0 if _preempt_delivery else
                     max(0.0, demand_remaining.get(cur, 0.0) - projected_gt.get(cur, 0.0) - _woc.get(cur, 0.0))
                     if (_cap_max or _pin_d1a_m) else _defc(cur, GT_SHELF_LIFE_SHIFTS) if _sticky_now
                     else _defc(cur, _base_buf))
            # BLD_CURABLE_CAP: the captive-max / sticky / D1A-pin room can outrun the SKU's curing
            # side (a 1-press SKU on a fast PS machine) → expired GT. Bound it by the
            # curable ceiling. Only NARROWS the room; no-op when the toggle is off.
            if _cap_max or _sticky_now or _pin_d1a_m:
                _room = _curable_cap(cur, _room)
                # ps2-only batch exception: don't dribble a tiny top-up — skip the shift
                # unless a >= _BLD_BATCH_MIN chunk can be built (press drains meanwhile).
                # Curable cap already bounds _room, so batching never re-introduces expiry.
                if (_BLD_BATCH_ENABLED and m in _BLD_BATCH_MACHINES
                        and 0.0 < _room < _BLD_BATCH_MIN):
                    _room = 0.0
            # ── Machine CONTINUATION stickiness (STICKY_MACHINE) ───────────────────
            # If the current SKU still has curable in-demand work (demand left AND a live
            # curing draw) but its thin dynamic buffer is momentarily full (_room≈0 → the
            # machine would RELEASE and Phase B would reshuffle the SKU to another machine
            # at a CO cost), EXTEND the continuation room to a shelf-safe, demand-capped,
            # curable-capped depth so the machine keeps building ITS OWN SKU. Never exceeds
            # the 3-day GT shelf or the demand cap (identical bound math to _defc, but with a
            # deeper horizon that DYN_BUFFER's per-SKU _dyn_H would otherwise cap short). OFF
            # or no live draw → _cont_room=0 → _room unchanged (bit-for-bit).
            if (_STICKY_MACHINE and cur and cur in eligible and not flex_reclaim
                    and not (_cap_max or _sticky_now or _pin_d1a_m) and not _preempt_delivery
                    and demand_remaining.get(cur, 0.0) > 0
                    and shift_cure_demand.get(cur, 0.0) > 0):
                _cont_b = min(_STICKY_CONT_SHIFTS, _pace_shelf)
                _cont_gap = shift_cure_demand.get(cur, 0.0) * _cont_b - projected_gt.get(cur, 0.0)
                _cont_cap = (demand_remaining.get(cur, 0.0)
                             - projected_gt.get(cur, 0.0) - _woc.get(cur, 0.0))
                _cont_room = min(max(0.0, _cont_gap), max(0.0, _cont_cap))
                _cont_room = _curable_cap(cur, _cont_room)   # no waste past what presses can drain
                if _cont_room > _room:
                    _room = _cont_room
            # ── NARROW day-2→day-3 BOUNDARY continuation (STICKY_HANDOFF) ───────────
            # ONE-TIME, at day 3 Shift A only: if the SKU carried from the end of day 2 still
            # has curable in-demand work, KEEP the machine on it (thin continuation room) and
            # LOCK it out of Phase-B/C reassignment this shift — so the avoidable boundary CO
            # (abandoning a still-unmet day-2 SKU) never happens. Independent of _STICKY_MACHINE.
            if (_STICKY_HANDOFF and day == _PLANT_2DAY_DAYS + 1 and shift_idx == 0
                    and cur and cur in eligible and not _pin_d1a_m
                    and demand_remaining.get(cur, 0.0) > 0
                    and shift_cure_demand.get(cur, 0.0) > 0):
                _ho_b = min(_STICKY_HANDOFF_SHIFTS, _pace_shelf)
                _ho_gap = shift_cure_demand.get(cur, 0.0) * _ho_b - projected_gt.get(cur, 0.0)
                _ho_cap = (demand_remaining.get(cur, 0.0)
                           - projected_gt.get(cur, 0.0) - _woc.get(cur, 0.0))
                _ho_room = _curable_cap(cur, min(max(0.0, _ho_gap), max(0.0, _ho_cap)))
                if _ho_room > _room:
                    _room = _ho_room
                flex_reclaim = False        # keep it on its carried SKU (don't reclaim dominant)
                if _STICKY_HANDOFF_LOCK:
                    _d1a_pin_lock.add(m)    # (optional) Phase B/C skip it → hard no-swap this shift
            # SG_DELIB HARD guard (Phase A): don't continue `cur` if another machine has
            # already committed it THIS shift in a DIFFERENT finer group — release m to
            # Phase B/C instead so the SKU stays in ONE group this shift (invariant #3).
            _sg_cont_block = False
            if _sg_delib and cur and cur not in _SG_EXEMPT_SKUS and not _sg_multi_ok(m, cur):
                _mg_c = _sku_group_of(m)
                # (i) SKU already committed to a different group this shift, OR
                # (ii) HARD purity: m's group is not in cur's sanctioned target set.
                if _sg_pair_blocked(m, cur):
                    _sg_cont_block = True
                else:
                    for _mm in machines:
                        if _mm == m:
                            continue
                        for (_cs, _cq, _ct) in stg[_mm]["campaigns"]:
                            if _cq > 0 and _cs == cur and _sku_group_of(_mm) != _mg_c:
                                _sg_cont_block = True
                                break
                        if _sg_cont_block:
                            break
            if cur in eligible and _room > 0 and not flex_reclaim and not _sg_cont_block:
                _ra = _bld_qty_per_shift(m, cur) / SHIFT_MINS   # per-SKU CT rate
                mins = min(s["remaining"], _room / _ra if _ra > 0 else s["remaining"])
                # CONTINUOUS BUILD CARRY (plant rule): use the EXACT (non-floored) unit rate and add
                # the fractional unit carried from the prior shift IF this machine continued `cur`.
                # Only carries OUT when this continuation is TIME-bound (spent the whole shift) —
                # a demand/shelf-bound campaign has spare time, no in-progress unit at the boundary.
                _ctsec_cur = _bld_ct_sec(m, cur)
                _exact_u = (mins * 60.0 / _ctsec_cur) if _ctsec_cur > 0 else 0.0
                _prevc   = _BLD_CARRY_UNITS.get(str(m)) if _BLD_CYCLE_CARRY else None
                _carry_in_u = _prevc[1] if (_prevc and _prevc[0] == cur) else 0.0
                _tot_u = _exact_u + _carry_in_u
                qty = int(_tot_u) if _BLD_CYCLE_CARRY else int(mins * _ra)
                _time_bound = mins >= s["remaining"] - 1e-6
                if _BLD_CYCLE_CARRY:
                    _BLD_CARRY_UNITS[str(m)] = (cur, (_tot_u - qty)) if (_time_bound and qty > 0) else (cur, 0.0)
                _qc = _dyn_cap_qty(cur, qty)                    # bound overnight excess by 7k cap
                if _qc != qty:
                    qty = _qc
                    mins = qty / _ra if _ra > 0 else mins
                if mins >= MIN_CAMPAIGN_MINS and qty > 0 and not _GLOBAL_SCORE_V2:
                    # V2: the continuation is scored in the unified Phase-B pool instead
                    # of being force-built here (no primary/secondary split).
                    s["campaigns"].append((cur, qty, "start"))
                    projected_gt[cur] = projected_gt.get(cur, 0.0) + qty
                    s["remaining"] -= mins
                    _conc_commit(m, cur, qty)          # CONCENTRATION: count the continuation
                s["primary_done"] = _defc(cur, _base_buf) <= 0

        # ── Phase B: global pair-scoring greedy for remaining capacity ──
        _guard = 0
        while _guard < 100000:
            _guard += 1
            pairs = []
            flex_m: dict = {}
            flex_s: dict = {}
            # SG_DELIB HARD guard: which finer group each SKU is ALREADY committed to THIS
            # shift (Phase-A continuations + earlier Phase-B picks). Recomputed each greedy
            # iteration; a pair that would build the SKU in a DIFFERENT group is skipped →
            # a SKU is never produced in >1 group in one shift (invariant #3). Inert when OFF.
            _shift_grp: dict = {}
            if _sg_delib:
                for _mm in machines:
                    _mg = _sku_group_of(_mm)
                    for (_cs, _cq, _ct) in stg[_mm]["campaigns"]:
                        if _cq > 0 and _cs not in _shift_grp:
                            _shift_grp[_cs] = _mg
            for m in machines:
                if m in _d1a_pin_lock:
                    continue   # BLD_SEED_PIN_D1A: hard-pinned machine keeps ONLY its seed build
                s = stg[m]
                if s["remaining"] < MIN_CAMPAIGN_MINS or s["co_count"] >= s["max_cos"]:
                    continue
                # S2_CAMPAIGN: a Stage-2 machine that has spent its per-day CO budget
                # takes no more switches this day (it may only continue its current SKU).
                if _s2_co_budget_blocks(m):
                    continue
                buf = _buf_of(m); rate = s["rate"]; dom = s["dom"]
                cur = s["cur_sku"]; cur_inch = sku_inch.get(cur, "")
                for sku in machine_skus.get(m, set()):
                    if not _plant_ok(m, sku):
                        continue   # PLANT_SET_LOCK: restrict to the plant Days-1-2 set until it's demand-complete
                    if sku == cur and not _GLOBAL_SCORE_V2:
                        continue   # V2: fold continuation into the pool as a CO=0 candidate
                    _dlvr = _delivery_relax(sku)   # delivery-date SKU → relax soft rules (keep allowable)
                    if _sg_delib and sku not in _SG_EXEMPT_SKUS and not _dlvr and not _sg_multi_ok(m, sku):
                        _cg = _shift_grp.get(sku)
                        if (_SG_SAMESHIFT_HARD and _cg is not None
                                and _cg != _sku_group_of(m)):
                            continue   # SG_DELIB: SKU already built in another group this shift
                        if _sg_pair_blocked(m, sku):
                            continue   # SG_HARD: non-admitted cross-group pair → machine idles
                    if sku != cur and not _sku_revert_ok(m, sku) and not _dlvr:
                        continue   # SKU_NO_REVERT: never re-build a SKU this machine has left
                    if sku != cur and not _plus3_direct_ok(
                            m, sku_inch.get(cur, ""), sku_inch.get(sku, ""), buf) and not _dlvr:
                        continue   # #2: two-hop through a productive intermediate, not a direct 8h +3
                    d = _defc(sku, buf)
                    if d <= 0:
                        continue
                    # GT-cap guard: a fully-buffered SKU (built ≥ 1 shift of draw) with
                    # no end-of-day headroom left would clamp to 0 → skip it so the
                    # Phase-B loop can't re-select it forever (dynamic buffer only).
                    if (_DYN_BUFFER_ENABLED
                            and projected_gt.get(sku, 0.0) >= shift_cure_demand.get(sku, 0.0)
                            and _dyn_headroom() <= 0):
                        continue
                    # 4-SKU/day cap: skip a CO that would be the 5th distinct SKU today.
                    if _sku_cap_blocks(m, sku, (c[0] for c in s["campaigns"])) and not _dlvr:
                        continue
                    to_inch = sku_inch.get(sku, "")
                    # ── Fixed-machine escape gate (Lever B): a fixed machine may reach an
                    # off-dominant inch only after its own inch is done, <= 1 diff-CO. ──
                    if _fixed_esc_block(m, to_inch, cur_inch) and not _dlvr:
                        continue
                    # ── Client inch rules (Rule 1a no-revisit + Rule 2 band) ──
                    if not _inch_gate(m, to_inch, cur_inch) and not _dlvr:
                        continue
                    # ── Leave gate: 5-day inch dwell OR deficit-done override ──
                    if (_INCH_RULES_ENABLED and to_inch != cur_inch
                            and not _may_leave_inch(m, cur_inch, _defc, buf, rate) and not _dlvr):
                        continue
                    # ── Diff-size-CO amortization gate: block wasteful inch-hop churn ──
                    if _DIFF_CO_GATE and to_inch != cur_inch and not _diff_co_ok(m, to_inch) and not _dlvr:
                        continue
                    # ── Part 2 JIT churn control: urgency margin + per-day budget + amortization ──
                    if to_inch != cur_inch and not _jit_diff_ok(m, cur_inch, to_inch, buf) and not _dlvr:
                        continue
                    if (m in (_SOFT_LOCK_MACHINES | _INCH_FLEX_MACHINES)
                            and to_inch != dom and not s["primary_done"]
                            and not _INCH_RULES_ENABLED):
                        continue
                    cost = 0.0 if (_GLOBAL_SCORE_V2 and sku == cur) else _co_cost(m, cur_inch, to_inch, from_sku=cur, to_sku=sku)
                    if s["remaining"] - cost < MIN_CAMPAIGN_MINS and not _dlvr:
                        continue
                    is_urgent = (sku in co_target_skus and projected_gt.get(sku, 0.0) == 0
                                 and demand_remaining.get(sku, 0.0) > 0)
                    # STARVING-PRESS bypass: same-inch CO onto a SKU whose curing presses are RUNNING
                    # but STARVED (live draw, no GT on hand/built this shift, demand left). Keyed on
                    # LIVE starvation (not planned-CO membership like is_urgent), so a machine feeds
                    # its own empty presses instead of idling its residual minutes.
                    is_starved = (_CO_STARVE_BYPASS and to_inch == cur_inch
                                  and shift_cure_demand.get(sku, 0.0) > 0
                                  and gt_inventory.get(sku, 0.0) <= 0
                                  and projected_gt.get(sku, 0.0) <= 0
                                  and demand_remaining.get(sku, 0.0) > 0)
                    _flex_off_ok = (m in _INCH_FLEX_MACHINES and to_inch != dom
                                    and s["primary_done"])
                    # With the client inch rules a diff-inch move has already passed
                    # the Rule-1b gate (nothing left to serve at the current inch), so
                    # the 30% cost guard must not block it — the machine would idle.
                    if _INCH_RULES_ENABLED and to_inch != cur_inch:
                        _flex_off_ok = True
                    if (cost > 0.30 * s["remaining"] and not is_urgent and not is_starved
                            and not _flex_off_ok and not _dlvr):
                        continue
                    avail = s["remaining"] - cost
                    _ra = _bld_qty_per_shift(m, sku) / SHIFT_MINS   # per-SKU CT rate
                    mins = min(avail, d / _ra if _ra > 0 else avail)
                    qty = int(mins * _ra)
                    # S2_CAMPAIGN: a Stage-2 CO to a NEW sku must yield a LONG campaign
                    # (>= S2_MIN_CAMPAIGN_MINS) — block short churn switches so the machine
                    # idles instead. Same-sku / non-Stage-2 / OFF → plant MIN (no-op).
                    _min_camp = MIN_CAMPAIGN_MINS
                    if (_S2_CAMPAIGN and sku != cur
                            and _MACHINE_GROUP.get(str(m), "") == "STAGE2"):
                        _min_camp = max(_min_camp, _S2_MIN_CAMPAIGN_MINS)
                    if (mins < _min_camp and not _dlvr) or qty <= 0:
                        continue
                    tier, primary = _tierg(sku, m, d)
                    # Prefer staying on the CURRENT inch (cheap same-size CO) once the
                    # inch rules are live — "dominant inch" is superseded by the anchor
                    # band, so the old dominant-inch preference no longer applies.
                    inch_penalty = ((0 if to_inch == cur_inch else 1) if _INCH_RULES_ENABLED
                                    else (0 if to_inch == dom else 1))
                    pairs.append((m, sku, cost, to_inch, tier, primary, qty, mins, inch_penalty))
                    flex_m[m] = flex_m.get(m, 0) + 1
                    flex_s[sku] = flex_s.get(sku, 0) + 1
            if not pairs:
                break

            # STICKY_MACHINE pickup damper: how many machines currently HOLD each SKU (their
            # carryover/continued SKU this shift). A Phase-B pair (m, sku) with sku != m's own
            # current SKU that PICKS UP a SKU another machine is already validly building is a
            # migration/duplication — softly demoted so the deficit-driven need still wins when
            # a 2nd machine is genuinely required, but a gratuitous swap does not. Empty/OFF → 0.
            _held_ct: dict = {}
            if _STICKY_MACHINE and _STICKY_PICKUP_PEN:
                for _mm in machines:
                    _cs = stg[_mm]["cur_sku"]
                    if _cs:
                        _held_ct[_cs] = _held_ct.get(_cs, 0) + 1

            def _pickup_pen(_m: str, _sku: str) -> int:
                """Demote a pair that CO's `_m` onto a SKU another machine already holds."""
                if not (_STICKY_MACHINE and _STICKY_PICKUP_PEN):
                    return 0
                if _sku == stg[_m]["cur_sku"]:
                    return 0                        # continuation of its own SKU — never a pickup
                return _STICKY_PICKUP_PEN if _held_ct.get(_sku, 0) >= 1 else 0

            # Lever A (FLEX_SCARCE_INCH): rank this iteration's inches by live curing-draw
            # shortfall (min(draw, demand_left) − projected_gt), scarcest = rank 0. A FLEXIBLE
            # machine then prefers its scarcest allowed inch (used in _key below). Recomputed
            # every pick, so it self-balances as projected_gt updates. None ⇒ lever off.
            _flex_scar_rank = None
            _flex_ival = None       # #2 MCV: blended per-inch value (for the hysteresis in _key)
            _flex_val_scale = 1.0
            if _FLEX_SCARCE_INCH and _FLEX_MACHS_HIST:
                _ishort: dict = {}
                _imon: dict = {}
                for _p in pairs:
                    _sk = _p[1]; _ti = _p[3]
                    _dr = shift_cure_demand.get(_sk, 0.0)
                    if _dr > 0:                              # this-shift draw shortfall (S_now)
                        _sh = max(0.0, min(_dr, demand_remaining.get(_sk, 0.0))
                                  - projected_gt.get(_sk, 0.0))
                        _ishort[_ti] = _ishort.get(_ti, 0.0) + _sh
                    if _FLEX_MCV_ENABLED:                    # cumulative monthly unmet units (G_mon)
                        _imon[_ti] = _imon.get(_ti, 0.0) + max(
                            0.0, demand_remaining.get(_sk, 0.0) - projected_gt.get(_sk, 0.0))
                if _FLEX_MCV_ENABLED:
                    # #2: blended value = w_now·shortfall + w_mon·monthly-gap; rank by it.
                    _flex_ival = {_i: _FLEX_MCV_W_NOW * _ishort.get(_i, 0.0)
                                     + _FLEX_MCV_W_MON * _imon.get(_i, 0.0)
                                  for _i in (set(_ishort) | set(_imon))}
                    if _flex_ival:
                        _flex_val_scale = max(_flex_ival.values()) or 1.0
                        _flex_scar_rank = {_i: _r for _r, _i in enumerate(
                            sorted(_flex_ival, key=lambda z: (-_flex_ival[z], z)))}
                elif _ishort:
                    # Lever A (MCV off): pure shortfall rank. Tiebreak on the inch string so
                    # equal-shortfall inches rank deterministically (set-iteration order is
                    # hash-seed dependent; the inch tiebreak removes that).
                    _flex_scar_rank = {_i: _r for _r, _i in
                                       enumerate(sorted(_ishort, key=lambda z: (-_ishort[z], z)))}

            if _GLOBAL_SCORE_V2:
                # ══ Phase 2: single unified utility U over the current pairs set ══
                _eps = 1e-9
                def _r_starv(p):
                    _d = shift_cure_demand.get(p[1], 0.0)
                    return 1.0 / (((projected_gt.get(p[1], 0.0) / _d) if _d > 0 else 1e18) + _eps)
                def _starving(p):
                    _d = shift_cure_demand.get(p[1], 0.0)
                    return 1.0 if (_d > 0 and projected_gt.get(p[1], 0.0) < _d) else 0.0
                def _over(p):
                    _d = shift_cure_demand.get(p[1], 0.0)
                    _H = _dyn_H(p[1]) if _DYN_BUFFER_ENABLED else _buf_of(p[0])
                    return 1.0 if (_d > 0 and projected_gt.get(p[1], 0.0) >= _d * _H) else 0.0
                def _mm(vals):
                    _lo, _hi = min(vals), max(vals); _rng = _hi - _lo
                    return [0.0] * len(vals) if _rng <= _eps else [(v - _lo) / _rng for v in vals]
                _nd = _mm([_defc(p[1], _buf_of(p[0])) for p in pairs])                        # deficit
                _ns = _mm([_r_starv(p) for p in pairs])                                        # starvation
                _ng = _mm([max(0.0, demand_remaining.get(p[1], 0.0)
                               - projected_gt.get(p[1], 0.0)) for p in pairs])                 # monthly gap
                _nc = _mm([float(p[2]) for p in pairs])                                        # CO minutes
                _ni = _mm([float(_inch_penalty_ranked(p[0], p[3])) for p in pairs])            # inch band
                _U = {}
                for _i, p in enumerate(pairs):
                    _nf = max(1, len(_feeders.get(p[1], ())))
                    _U[(p[0], p[1])] = (
                          GS_W_DEF    * _nd[_i]  + GS_W_STARV * _ns[_i] + GS_W_GAP * _ng[_i]
                        + GS_W_SCARCE * (1.0 / _nf) * _starving(p)
                        - GS_W_CO     * _nc[_i]  - GS_W_INCH  * _ni[_i] - GS_W_OVER * _over(p))
                # greedy: max U (SAME_GROUP home-group tiebreak, then lower CO, m, sku).
                # -_group_pen is a constant 0 for every pair when the lever is OFF → the
                # ordering is unchanged (bit-for-bit).
                # SG_DELIB: the deliberate stable-group term dominates U (a blocked cross-group
                # move → -_HYST_BIG, always beaten by any sanctioned/admitted pair). 0 for all
                # pairs (identity) when deliberate mode is OFF → bit-for-bit.
                m, sku, cost, to_inch, tier, primary, qty, mins, inch_penalty = max(
                    pairs, key=lambda p: ((-_sg_move_pen(p[0], p[1]),) if _sg_delib else ()) + (
                                          _U[(p[0], p[1])],
                                          -(_pickup_pen(p[0], p[1]) + _group_pen(p[0], p[1])),
                                          -p[2], p[0], p[1]))
                s = stg[m]
                _pick_done = True
            else:
                _pick_done = False

            def _key_base(p):
                m, sku, cost, to_inch, tier, primary, qty, mins, inch_penalty = p
                # Combined soft stickiness: machine-level pickup damper (migration waste) +
                # coarse group-home penalty. 0 for every candidate when both levers are OFF →
                # inserting it as a tuple slot is order-preserving (bit-for-bit).
                _gp = _pickup_pen(m, sku) + _group_pen(m, sku)
                # Lever A: for a flexible machine, replace the same-inch stickiness penalty
                # with the inch's scarcity rank (0 = scarcest allowed inch) so it feeds the
                # about-to-starve scarce inch first. Fixed machines keep the sticky penalty.
                if _flex_scar_rank is not None and str(m) in _FLEX_MACHS_HIST:
                    inch_penalty = _flex_scar_rank.get(to_inch, 99)
                    # #2 hysteresis: a flex machine leaves its CURRENT inch only for a
                    # MEANINGFULLY better one (symmetric dead-band Δ = HYS_BAND·scale +
                    # CO_LAMBDA·forgone-CO-production) and not within the switch cooldown —
                    # kills A→B→A oscillation while still reacting to a real gain.
                    if _FLEX_MCV_ENABLED and _flex_ival is not None:
                        _cur_i = sku_inch.get(stg[m]["cur_sku"], "")
                        if to_inch != _cur_i:
                            _delta = (_FLEX_MCV_HYS_BAND * _flex_val_scale
                                      + _FLEX_MCV_CO_LAMBDA * cost * stg[m]["rate"])
                            _meaningful = (_flex_ival.get(to_inch, 0.0)
                                           > _flex_ival.get(_cur_i, 0.0) + _delta)
                            _cooldown = ((day - machine_last_diff_co_day.get(m, -10**9))
                                         < _FLEX_MCV_COOLDOWN)
                            if (not _meaningful) or _cooldown:
                                inch_penalty += _HYST_BIG
                # §2 18-inch exception: on the 18"-machine (7003), build 18" FIRST whenever a
                # curing press is drawing that 18" SKU now (shift_cure_demand>0) — feeds the
                # already-committed Day-0 18" presses immediately (kills the ~22-day misalign).
                # When 18" has NO live draw it is not a build candidate (no-waste coupling), so
                # 7003 falls back to 15" automatically — no explicit "15"-first branch needed.
                if (_INCH18_EXC and str(m) == _INCH18_MACHINE and to_inch == "18"
                        and shift_cure_demand.get(sku, 0.0) > 0):
                    inch_penalty = -1                          # strictly ahead of any 15" (rank >= 0)
                constraint = min(flex_m[m], flex_s[sku])
                if _BLD_SEC_ORDER != "baseline":
                    # Explicit four-factor ordering (same-size first, cost/m/sku tail).
                    _draw = shift_cure_demand.get(sku, 0.0)
                    STARV = 0 if (_draw > 0 and projected_gt.get(sku, 0.0) < _draw) else 1
                    URG   = 0 if _urgency_score(sku, demand_remaining, press_count,
                                                cure_ct_map, days_left) > 0 else 1
                    DEFC  = -_defc(sku, _buf_of(m))
                    _tail = (_gp, cost, m, sku)   # SAME_GROUP: home-group before the cost/id tail
                    if _BLD_SEC_ORDER == "UD":
                        return (inch_penalty, URG, DEFC) + _tail
                    if _BLD_SEC_ORDER == "USD":
                        return (inch_penalty, URG, STARV, DEFC) + _tail
                    if _BLD_SEC_ORDER == "SUD":
                        return (inch_penalty, STARV, URG, DEFC) + _tail
                    if _BLD_SEC_ORDER == "SD":
                        return (inch_penalty, STARV, DEFC) + _tail
                    if _BLD_SEC_ORDER == "DSU":
                        return (inch_penalty, DEFC, STARV, URG) + _tail
                    if _BLD_SEC_ORDER == "INS_S":   # promote STARV, keep committed load-balancing tail
                        return (inch_penalty, STARV, tier, primary, _gp, constraint, cost, m, sku)
                    # unknown code → fall through to committed baseline
                if _GLOBAL_CONSTRAINT_MODE == "below":
                    if _CONCENTRATION:
                        # Defer a redundant machine (over_prov=1) below any still-under-served
                        # SKU, but keep inch stickiness first (never force a diff-CO to spread).
                        return ((0 if (_DEDICATED_FIRST and sku in _DEDICATED_SKUS) else 1,
                                 0 if (day <= _EARLY_DAYS and sku in _EARLY_SKUS) else 1,
                                 inch_penalty, _over_prov(m, sku), tier, _gp, primary,
                                 constraint, cost, m, sku) if _SG_STRONG else
                                (0 if (_DEDICATED_FIRST and sku in _DEDICATED_SKUS) else 1,
                                 0 if (day <= _EARLY_DAYS and sku in _EARLY_SKUS) else 1,
                                 inch_penalty, _over_prov(m, sku), tier, primary,
                                 _gp, constraint, cost, m, sku))
                    return ((inch_penalty, tier, _gp, primary, constraint, cost, m, sku)
                            if _SG_STRONG else
                            (inch_penalty, tier, primary, _gp, constraint, cost, m, sku))
                elif _GLOBAL_CONSTRAINT_MODE == "captive":
                    return (inch_penalty, 0 if constraint <= 1 else 1,
                            tier, primary, _gp, cost, m, sku)
                return (inch_penalty, constraint, tier, primary, _gp, cost, m, sku)  # "above"

            def _key(p):
                # SG_DELIB: prepend the DELIBERATE stable-group penalty as the DOMINANT term
                # (0 = sanctioned/target group or admitted move; _HYST_BIG = blocked cross-group
                # move → the pair loses), so a SKU is built by its assigned group whenever that
                # group can serve it — group purity above urgency (priority #7, coverage cost
                # accepted). Inert (identity, bit-for-bit) when deliberate mode is OFF.
                _b = _key_base(p)
                return ((_sg_move_pen(p[0], p[1]),) + _b) if _sg_delib else _b

            if not _pick_done:
                # DELIVERY_PRIORITY: prepend the committed-delivery EDF rank so a behind
                # committed SKU wins its eligible machines first. Identity when inactive.
                _key_sel = _key if not _prio_on_bld else (lambda p: (_bld_prio(p[1]), *_key(p)))
                m, sku, cost, to_inch, tier, primary, qty, mins, inch_penalty = min(pairs, key=_key_sel)
                s = stg[m]
            # T2 departure gate: before rotating away from cur, ensure cur holds its round-trip cushion
            # (eff_buf). If short by ≥ a campaign, top cur up toward eff_buf and DEFER the rotation
            # (re-evaluate next iteration); a sub-campaign shortfall is "close enough" → allow rotation.
            if _RT_IMMINENT:
                _curm = s["cur_sku"]
                if _curm and _curm != sku:
                    _curdef = _defc(_curm, s.get("eff_buf", _buf_of(m)))
                    if _curdef > 0:
                        _r = s["rate"]
                        _tmins = min(s["remaining"], _curdef / _r if _r > 0 else 0.0)
                        _tqty = int(_tmins * _r)
                        _tqc = _dyn_cap_qty(_curm, _tqty)       # bound overnight excess by 7k cap
                        if _tqc != _tqty:
                            _tqty = _tqc
                            _tmins = _tqty / _r if _r > 0 else _tmins
                        if _tmins >= MIN_CAMPAIGN_MINS and _tqty > 0:
                            s["campaigns"].append((_curm, _tqty, "start"))
                            projected_gt[_curm] = projected_gt.get(_curm, 0.0) + _tqty
                            s["remaining"] -= _tmins
                            _conc_commit(m, _curm, _tqty)   # CONCENTRATION: count the top-up
                            continue                 # cur topped up → re-evaluate (defer rotation)
            _qc = _dyn_cap_qty(sku, qty)                        # bound overnight excess by 7k cap
            if _qc != qty:
                _raw = _bld_qty_per_shift(m, sku) / SHIFT_MINS
                qty = _qc
                mins = qty / _raw if _raw > 0 else mins
                if qty <= 0 or mins < MIN_CAMPAIGN_MINS:
                    s["remaining"] -= cost           # pay the CO time, build nothing (cap full)
                    s["co_count"] += 1
                    machine_day_co[m] = machine_day_co.get(m, 0) + 1   # S2_CAMPAIGN per-day budget
                    s["cur_sku"] = sku
                    continue
            _is_cont = _GLOBAL_SCORE_V2 and sku == s["cur_sku"]   # continuation = CO=0, no CO charged
            co_type = ("start" if _is_cont
                       else "same_size_CO" if to_inch == sku_inch.get(s["cur_sku"], "")
                       else "diff_size_CO")
            if co_type == "diff_size_CO":
                machine_last_diff_co_day[m] = day
                machine_day_diff_co[m] = machine_day_diff_co.get(m, 0) + 1   # Part 2 per-day budget
                if _FIXED_ESCAPE_ENABLED and str(m) in _FIXED_MACHS_HIST:
                    fixed_escape_used[str(m)] = fixed_escape_used.get(str(m), 0) + 1  # Lever B budget
            s["campaigns"].append((sku, qty, co_type))
            _sg_move_commit(m, sku)                  # SG_DELIB: sanction an admitted cross-group move
            projected_gt[sku] = projected_gt.get(sku, 0.0) + qty
            s["remaining"] -= (cost + mins)
            _conc_commit(m, sku, qty)                # CONCENTRATION: count the Phase-B pick
            if not _is_cont:
                s["co_count"] += 1
                machine_day_co[m] = machine_day_co.get(m, 0) + 1   # S2_CAMPAIGN per-day budget
            s["cur_sku"] = sku

        # ── Phase B2: one-time +3/-3 inch escape for STRANDED machines (experiment) ──
        # A machine whose ±2 in-band work is exhausted this shift may make ONE inch jump
        # of exactly 3 (beyond the band) for the whole month, at an 8h building CO
        # (INCH_PLUS3_CO_MINS). Gated to real +3/-3 demand + enough days to amortise —
        # UNLESS its in-band demand is fully completed (take it immediately; idle is worse).
        # ── INCH18_DEFER forced 15→18 switch (8h CO-only shift) ───────────────────────────
        # 7003's per-day lock flips to {18} on switch_day; here we force the direct 8h CO
        # (qty=0, 480 min, production next shift) that the normal Phase-B path skips ("no room to
        # build same shift"). One-way: the lock={18} blocks any revert to 15".
        if _INCH18_DEFER:
            _m18 = _INCH18_MACHINE
            _s18 = stg.get(_m18)
            _cur18n = sku_inch.get(_s18["cur_sku"], "") if _s18 is not None else ""
            # Only force the 8h CO-only shift for a +3 jump (e.g. 15→18); a +2 jump (16→18) is a
            # normal ~120-min diff-CO that the Phase-B path builds same-shift → no forced block needed.
            if (_s18 is not None and machine_locked_inches.get(_m18) == {"18"}
                    and _cur18n.isdigit() and abs(int(_cur18n) - 18) == 3
                    and not _s18["campaigns"] and _s18["remaining"] >= INCH_PLUS3_CO_MINS):
                _b18 = _buf_of(_m18)
                _c18 = [x for x in machine_skus.get(_m18, set())
                        if sku_inch.get(x, "") == "18" and demand_remaining.get(x, 0.0) > 0]
                if _c18:
                    _best18 = max(_c18, key=lambda x: (_defc(x, _b18), x))
                    _s18["campaigns"].append((_best18, 0, "plus3_CO"))   # 8h CO-only (480 min)
                    _s18["remaining"] -= INCH_PLUS3_CO_MINS
                    _s18["co_count"] += 1
                    _s18["cur_sku"] = _best18

        if _INCH_PLUS3_ENABLED:
            for m in sorted(machines):
                if m in machine_plus3_used:
                    continue
                s = stg[m]; rate = s["rate"]
                # The 8h CO fills a whole shift → the machine must be idle this shift
                # (a full free shift for the CO; production begins next shift).
                if rate <= 0 or s["campaigns"] or s["remaining"] < INCH_PLUS3_CO_MINS:
                    continue
                if _P3DBG: _PLUS3_DBG[0] += 1              # idle this shift + escape available
                _a = _inch_num(machine_anchor_inch.get(m, ""))
                if _a is None:
                    continue
                _buf = _buf_of(m); _elig = machine_skus.get(m, set())
                _cur_inch = sku_inch.get(s["cur_sku"], "")
                # STRANDED: no in-band (±2) deficit left this shift (Phase A/B used it all)
                if any(_defc(x, _buf) > 0 and _inch_num(sku_inch.get(x, "")) is not None
                       and abs(_inch_num(sku_inch.get(x, "")) - _a) <= _INCH_BAND_WIDTH
                       for x in _elig):
                    continue
                if _P3DBG: _PLUS3_DBG[1] += 1              # stranded (no in-band ±2 deficit)
                # the +3/-3 move is still a diff-inch CO → respect the 5-day dwell gate
                if not _may_leave_inch(m, _cur_inch, _defc, _buf, rate):
                    continue
                if _P3DBG: _PLUS3_DBG[2] += 1              # passed dwell gate
                _has3 = any(_inch_num(sku_inch.get(x, "")) is not None
                            and abs(_inch_num(sku_inch.get(x, "")) - _a) == 3 for x in _elig)
                if _P3DBG and _has3: _PLUS3_DBG[3] += 1    # has ANY ±3 eligible SKU (data)
                # AMORTISE: need ≥ N days left, UNLESS in-band demand is fully completed
                # for the month (machine would otherwise idle → take the escape now).
                _inband_demand_left = any(
                    demand_remaining.get(x, 0.0) > 0
                    and _inch_num(sku_inch.get(x, "")) is not None
                    and abs(_inch_num(sku_inch.get(x, "")) - _a) <= _INCH_BAND_WIDTH
                    for x in _elig)
                if days_left < INCH_PLUS3_MIN_DAYS_LEFT and _inband_demand_left:
                    continue
                # TARGET: exactly 3 inches from anchor, real deficit + unmet demand
                _cands = [x for x in _elig
                          if _defc(x, _buf) > 0 and demand_remaining.get(x, 0.0) > 0
                          and _inch_num(sku_inch.get(x, "")) is not None
                          and abs(_inch_num(sku_inch.get(x, "")) - _a) == 3]
                if not _cands:
                    continue
                _best  = max(_cands, key=lambda x: (_defc(x, _buf), x))
                # The 8h CO occupies the WHOLE shift (INCH_PLUS3_CO_MINS = one shift), so
                # this shift produces nothing; the machine switches to the +3 target and
                # builds it from the NEXT shift (Phase A continuation). Record a CO-only
                # campaign (qty=0) carrying its real 480-min cost.
                s["campaigns"].append((_best, 0, "plus3_CO"))
                s["remaining"] -= INCH_PLUS3_CO_MINS
                s["co_count"] += 1
                s["cur_sku"]   = _best
                machine_plus3_used.add(m)                  # one escape per machine per month

        # ── Phase B3: stepwise inch-DRIFT for STRANDED idle machines (INCH_STEP_DRIFT) ──
        # A machine Phase A/B left IDLE this shift whose current eligible set has no deficit may
        # migrate ONE inch step (±1) via a normal diff-size CO to a DB-CERTIFIED adjacent inch with
        # real deficit + demand. One-way (once a direction is chosen it never reverses), cumulative
        # reach capped at _INCH_STEP_MAX from the historical anchor, NEVER a direct ±3 (must step
        # 14→15→16). Extends the machine's runtime eligibility + locked-inch set so Phase A/B keeps
        # building the drifted inch next shift. OFF → no-op (bit-for-bit today's plan).
        if _INCH_STEP_DRIFT and machine_db_skus:
            for m in sorted(machines):
                s = stg[m]; rate = s["rate"]
                if rate <= 0 or s["campaigns"]:            # only a machine idle THIS shift may drift
                    continue
                _base = _inch_num(machine_anchor_inch.get(m, "")
                                  or _MACHINE_DOMINANT_INCH.get(str(m), ""))
                if _base is None:
                    continue
                _drift = machine_step_drift.get(m, 0)
                if abs(_drift) >= _INCH_STEP_MAX:          # already at the cumulative cap
                    continue
                if _step_dbg: _STEP_DBG[0] += 1
                _buf = _buf_of(m)
                # STRANDED: the machine's CURRENT eligible set has no WHOLE-MONTH remaining demand
                # (demand_remaining − projected_gt), NOT just this-shift buffered deficit. The buffered
                # version premature-abandons a machine whose inch still has real month demand (CLAUDE.md
                # Lever-B finding: buffered _defc → −31.7k). Only a genuinely-DONE inch may drift.
                if any(demand_remaining.get(x, 0.0) - projected_gt.get(x, 0.0) > 0
                       for x in machine_skus.get(m, set())):
                    continue
                if _step_dbg: _STEP_DBG[1] += 1
                _cur_inch = sku_inch.get(s["cur_sku"], "")
                _reach = _base + _drift                    # the furthest inch it currently builds
                _dirs = [1, -1] if _drift == 0 else [1 if _drift > 0 else -1]   # one-way after step 1
                _best = None; _best_key = None; _best_ni = None; _best_dir = 0
                for _dir in _dirs:
                    if abs(_drift + _dir) > _INCH_STEP_MAX:
                        continue
                    _nis = str(_reach + _dir)              # the single adjacent inch (±1 step)
                    for x in machine_db_skus.get(m, set()):
                        if sku_inch.get(x, "") != _nis or demand_remaining.get(x, 0.0) <= 0:
                            continue
                        _d = _defc(x, _buf)
                        if _d <= 0:
                            continue
                        _k = (_d, x)
                        if _best is None or _k > _best_key:
                            _best = x; _best_key = _k; _best_ni = _reach + _dir; _best_dir = _dir
                if _best is None:
                    continue
                _co = _co_cost(m, _cur_inch or str(_reach), str(_best_ni))
                if s["remaining"] < _co:                   # the diff-size CO must fit this shift
                    continue
                # take the ±1 step: CO-only this shift (production begins next shift via Phase A)
                s["campaigns"].append((_best, 0, "diff_size_CO"))
                s["remaining"] -= _co
                s["co_count"] += 1
                s["cur_sku"] = _best
                machine_last_diff_co_day[m] = day
                machine_step_drift[m] = _drift + _best_dir
                # extend runtime eligibility so Phase A/B keeps building the drifted inch
                for x in machine_db_skus.get(m, set()):
                    if sku_inch.get(x, "") == str(_best_ni):
                        machine_skus.setdefault(m, set()).add(x)
                _ls = machine_locked_inches.get(m)
                if _ls is not None:
                    _ls.add(str(_best_ni))
                if _step_dbg: _STEP_DBG[2] += 1

        # ── Phase C: forward-buffer slack-fill (level-loading) ──
        # Use building capacity that Phase A/B left idle to PRE-BUILD a shelf-life-safe
        # forward buffer for SKUs that WILL be cured in the next 3 days. Builds ONLY
        # required GT: a candidate must have a LIVE cure-draw (shift_cure_demand>0 → a
        # press is actively pulling it — never a random SKU) and unmet demand. Per-SKU
        # forward target = min(demand_remaining, 3-day cure-draw); the shelf-safe cap
        # auto-targets building-limited SKUs (high draw) and auto-skips press-limited
        # ones (tiny draw → nothing to pre-build). Bounded by the 10k end-of-day cap.
        if _FORWARD_BUFFER_ENABLED:
            _fwd_added = 0.0   # total forward GT queued this shift (across all machines)
            # ── PACING (small-buffer rotation): compute the day's FLAT build target and the
            # SKUs already built today, so Phase C stops banking once the day hits its flat
            # pace and prefers WIDENING the active SKU set. day 1-2 = plant replay (never here).
            _pacing_target = None
            _pace_served = None
            if _PACING_ENABLED and day > _PLANT_2DAY_DAYS:
                if PACING_TARGET_MODE == "fixed":
                    _pacing_target = PACING_DAILY_TARGET
                else:
                    _rem_dem = sum(v for v in demand_remaining.values() if v > 0)
                    _wdl = max(1, int(days_left))
                    _cap = sum(v for v in shift_cure_demand.values() if v > 0) * 3.0
                    _pace_pace = _rem_dem / _wdl
                    _pacing_target = min(_pace_pace, _cap) if _cap > 0 else _pace_pace
                # base (Phase A/B, cure-neutral) GT already committed this shift — counts
                # toward the flat daily total but is NEVER throttled (only forward banking is).
                _base_committed = float(sum(q for _m in machines
                                            for (_sk, q, _ct) in stg[_m]["campaigns"]))
                _pace_served = set(pacing_day_skus) if pacing_day_skus else set()
                for _m in machines:
                    for (_sk, q, _ct) in stg[_m]["campaigns"]:
                        if q > 0:
                            _pace_served.add(_sk)
            for m in sorted(machines):
                if m in _d1a_pin_lock:
                    continue   # BLD_SEED_PIN_D1A: hard-pinned machine keeps ONLY its seed build
                s = stg[m]; rate = s["rate"]
                if rate <= 0:
                    continue
                _cguard = 0
                while s["remaining"] >= MIN_CAMPAIGN_MINS and _cguard < 1000:
                    _cguard += 1
                    hr = _gt_headroom(_fwd_added + _dyn_over[0])   # share the 7k budget with the dynamic buffer
                    if _ENDOFDAY_GT_CAP_ENABLED and hr <= 0:
                        break
                    dom = s["dom"]; cur = s["cur_sku"]; cur_inch = sku_inch.get(cur, "")
                    # SG_DELIB HARD guard: groups already committed to each SKU this shift.
                    _shift_grp_c: dict = {}
                    if _sg_delib:
                        for _mm in machines:
                            _mg = _sku_group_of(_mm)
                            for (_cs, _cq, _ct) in stg[_mm]["campaigns"]:
                                if _cq > 0 and _cs not in _shift_grp_c:
                                    _shift_grp_c[_cs] = _mg
                    best = None; best_key = None; best_room = 0.0
                    for sku in machine_skus.get(m, set()):
                        _dlvr = _delivery_relax(sku)   # delivery-date SKU → relax soft rules (Phase C)
                        if sku != cur and not _sku_revert_ok(m, sku) and not _dlvr:
                            continue   # SKU_NO_REVERT (Phase C): don't pre-build a left SKU
                        if _sg_delib and sku not in _SG_EXEMPT_SKUS and not _dlvr and not _sg_multi_ok(m, sku):
                            _cg = _shift_grp_c.get(sku)
                            if (_SG_SAMESHIFT_HARD and _cg is not None
                                    and _cg != _sku_group_of(m)):
                                continue   # SG_DELIB: SKU already built in another group this shift
                            if _sg_pair_blocked(m, sku):
                                continue   # SG_HARD: non-admitted cross-group pair → machine idles
                        if sku != cur and not _plus3_direct_ok(m, cur_inch, sku_inch.get(sku, ""), _buf_of(m)):
                            continue   # #2: two-hop through a productive intermediate, not direct 8h
                        draw = _eff_draw(sku)              # LOOKAHEAD_BUF: anticipated peak draw
                        if draw <= 0:                      # not needed soon → not "required"
                            continue
                        dr = demand_remaining.get(sku, 0.0)
                        if dr <= 0:
                            continue
                        # starvation-risk gate: skip SKUs that already hold enough GT
                        # (>= _FWD_RISK_SHIFTS shifts of draw) — they are NOT about to
                        # starve, so pre-building them would only front-load early month.
                        # IDLE_UNMET relaxes this: aim idle capacity at the biggest unmet
                        # gaps up to the shelf-safe target regardless of current on-hand.
                        # IDLE_GAP_FILL bypasses the gate for BUILDING-LIMITED SKUs only
                        # (_urgency_score==0 → curing can still cover the demand in the
                        # horizon, so the shelf-capped extra GT will be cured, not clogged):
                        # an idle machine builds toward the cumulative monthly gap even when
                        # the SKU is not about to starve THIS shift (the 7502 case).
                        _bld_limited = (_IDLE_GAP_FILL
                                        and demand_remaining.get(sku, 0.0) > 0
                                        and _urgency_score(sku, demand_remaining, press_count,
                                                           cure_ct_map, days_left) == 0)
                        # DELIVERY_PRIORITY: a behind committed SKU bypasses the starvation-risk
                        # gate so idle capacity banks its GT AHEAD of the deadline even when it
                        # isn't about to run dry this shift. Still bounded by the shelf-safe target
                        # (draw·9), the end-of-day GT cap, and the demand cap → no overbuild.
                        _prio_behind = (_prio_on_bld and _pdm_bld.get(sku) is not None
                                        and (demand_remaining.get(sku, 0.0)
                                             - projected_gt.get(sku, 0.0)) > 0)
                        # #3 BRIDGE: a bridgeable holiday is imminent → bank this live-draw SKU's
                        # GT ahead (bypass the starvation-risk gate exactly like _prio_behind).
                        _bridge_need = (_HOLIDAY_BRIDGE and bridge_shifts > 0)
                        if (not _bld_limited and not _prio_behind and not _bridge_need
                                and (not _IDLE_UNMET_ENABLED or _IDLE_UNMET_KEEP_GATE)
                                and _FWD_RISK_SHIFTS > 0
                                and projected_gt.get(sku, 0.0) >= draw * _FWD_RISK_SHIFTS):
                            continue
                        need_co = (sku != cur)
                        if need_co and s["co_count"] >= s["max_cos"] and not _dlvr:
                            continue
                        # S2_CAMPAIGN per-day CO budget: no more Stage-2 switches once spent.
                        if need_co and _s2_co_budget_blocks(m):
                            continue
                        # 4-SKU/day cap: don't let the forward buffer spend the day's 5th SKU.
                        if need_co and _sku_cap_blocks(m, sku, (c[0] for c in s["campaigns"])):
                            continue
                        to_inch = sku_inch.get(sku, "")
                        # Lever B: the forward buffer must not spend a fixed machine's escape
                        # (escapes are deliberate Phase-B decisions, marked there) — keep the
                        # opportunistic Phase-C pre-build on the fixed machine's current inch.
                        if (_FIXED_ESCAPE_ENABLED and str(m) in _FIXED_MACHS_HIST
                                and to_inch != cur_inch):
                            continue
                        # ── Client inch rules (same gates as Phase B) ──
                        # Under one-way inch movement every inch change is a
                        # ONE-TIME door. The forward buffer is opportunistic
                        # pre-building, so letting it spend that door permanently
                        # closes an inch the machine may need for real demand
                        # later. Restrict Phase C to the machine's current inch.
                        if (_INCH_RULES_ENABLED and _INCH_RULES_PHASE_C_SAME_INCH
                                and to_inch != cur_inch):
                            continue
                        if not _inch_gate(m, to_inch, cur_inch):
                            continue
                        if (_INCH_RULES_ENABLED and to_inch != cur_inch
                                and not _may_leave_inch(m, cur_inch, _defc, _buf_of(m), rate)):
                            continue
                        # ── Diff-size-CO amortization gate (same as Phase B) ──
                        if _DIFF_CO_GATE and to_inch != cur_inch and not _diff_co_ok(m, to_inch):
                            continue
                        # ── Part 2 JIT churn control (same as Phase B) ──
                        if to_inch != cur_inch and not _jit_diff_ok(m, cur_inch, to_inch, _buf_of(m)):
                            continue
                        # respect the flex/soft-lock off-dominant-inch gate (as Phase B)
                        if (m in (_SOFT_LOCK_MACHINES | _INCH_FLEX_MACHINES)
                                and to_inch != dom and not s["primary_done"]
                                and not _INCH_RULES_ENABLED):
                            continue
                        # Small-buffer rotation: shrink the per-SKU forward buffer to
                        # PACING_BUFFER_SHIFTS (1-2 shifts of draw) so the machine banks only a
                        # thin buffer, then rotates to the next drawn SKU. OFF → full 9 shifts.
                        _fwd_shifts = (min(GT_SHELF_LIFE_SHIFTS, PACING_BUFFER_SHIFTS)
                                       if _PACING_ENABLED else GT_SHELF_LIFE_SHIFTS)
                        # #2 NO-PERISH: cap the forward window to the WORKING shifts reachable before
                        # a holiday (holiday shifts cure nothing → GT built for them would age out).
                        # None → full 9-shift shelf (bit-for-bit).
                        if _HOLIDAY_NO_PERISH and fwd_work_shifts is not None:
                            _fwd_shifts = min(_fwd_shifts, fwd_work_shifts)
                        # #3 BRIDGE: raise the floor to cover the holiday + first post-holiday shift so
                        # presses keep feeding through the idle day (shelf-safe holidays only; the call
                        # site zeroes bridge_shifts when the holiday is too long for the 3-day shelf).
                        _floor_shifts = (bridge_shifts + 1) if _bridge_need else 0
                        target = min(dr, draw * max(_fwd_shifts, _floor_shifts))
                        room = target - projected_gt.get(sku, 0.0) - _woc.get(sku, 0.0)
                        if room <= 0:
                            continue
                        # PACING daily-cap + spread. Once the day's TOTAL building hits the flat
                        # target, stop banking ahead — EXCEPT a SKU about to starve THIS shift
                        # (on-hand < its draw) is still fed (no new starvation). Both rank terms
                        # are 0 when pacing OFF → order-preserving (bit-for-bit).
                        _pace_star_rank = 0
                        _pace_spread = 0
                        if _pacing_target is not None:
                            _pace_starving = projected_gt.get(sku, 0.0) < draw
                            _pace_room_day = (_pacing_target - pacing_day_built
                                              - _base_committed - _fwd_added)
                            if _pace_room_day <= 0 and not _pace_starving:
                                continue
                            if not _pace_starving:
                                room = min(room, _pace_room_day)
                                if room <= 0:
                                    continue
                            _pace_star_rank = 0 if _pace_starving else 1
                            _pace_spread = (0 if (_pace_served is not None
                                                  and sku not in _pace_served) else 1)
                        key = (# SG_DELIB: DOMINANT stable-group term (0 = target/admitted move;
                               # _HYST_BIG = blocked cross-group → the SKU loses this idle machine).
                               # 0 for all pairs when deliberate mode is OFF → order-preserving.
                               _sg_move_pen(m, sku),
                               # DELIVERY_PRIORITY: a behind committed SKU is pre-built first
                               # (EDF). (1,0.0) constant when inactive → order-preserving.
                               _bld_prio(sku),
                               # CONCENTRATION: defer an idle machine from piling onto an
                               # already-paced SKU (0 constant when OFF → order-preserving).
                               _over_prov(m, sku),
                               # Lever C: under the inch rules, keep the machine on its
                               # CURRENT inch first — only move to another in-band inch when
                               # no current-inch SKU is starving (i.e. move only for starvation).
                               (0 if to_inch == cur_inch else 1) if _INCH_RULES_ENABLED else 0,
                               0 if to_inch == dom else 1,        # dominant inch first
                               0 if sku == cur else 1,            # avoid a CO if possible
                               # SAME_GROUP: prefer this SKU's home group for the forward
                               # buffer too (0 for all pairs when the lever is OFF →
                               # order-preserving). Placed after avoid-CO so an about-to-
                               # starve SKU is still pre-built cross-group when needed.
                               _group_pen(m, sku),
                               # PACING: feed a starving SKU first, then WIDEN (a SKU not yet
                               # built today) — spreads building across more distinct SKUs/day.
                               _pace_star_rank, _pace_spread,
                               # PACING → nearest-to-starve (thin, steady); else IDLE_UNMET
                               # biggest-gap concentration; else nearest-to-starve.
                               (-(draw / (projected_gt.get(sku, 0.0) + 1.0)) if _pacing_target is not None
                                else (-dr if _IDLE_UNMET_ENABLED
                                      else -(draw / (projected_gt.get(sku, 0.0) + 1.0)))),
                               -room, sku)
                        if best is None or key < best_key:
                            best = sku; best_key = key; best_room = room
                    if best is None:
                        break
                    to_inch = sku_inch.get(best, "")
                    cost = 0.0 if best == cur else _co_cost(m, cur_inch, to_inch)
                    if s["remaining"] - cost < MIN_CAMPAIGN_MINS:
                        break
                    room = min(best_room, hr) if _ENDOFDAY_GT_CAP_ENABLED else best_room
                    avail = s["remaining"] - cost
                    mins = min(avail, room / rate)
                    qty = int(mins * rate)
                    # hard demand-cap clamp (sacred invariant): never build past the
                    # SKU's remaining demand headroom, even by a min-campaign rounding.
                    qty = min(qty, max(0, int(demand_remaining.get(best, 0.0)
                                              - projected_gt.get(best, 0.0))))
                    # S2_CAMPAIGN: same long-campaign floor for a Stage-2 forward-buffer CO
                    # to a NEW sku (best != cur) — no short Phase-C churn switches.
                    _min_camp = MIN_CAMPAIGN_MINS
                    if (_S2_CAMPAIGN and best != cur
                            and _MACHINE_GROUP.get(str(m), "") == "STAGE2"):
                        _min_camp = max(_min_camp, _S2_MIN_CAMPAIGN_MINS)
                    if mins < _min_camp or qty <= 0:
                        break
                    co_type = ("start" if best == cur
                               else ("same_size_CO" if to_inch == cur_inch else "diff_size_CO"))
                    if co_type == "diff_size_CO":
                        machine_last_diff_co_day[m] = day
                        machine_day_diff_co[m] = machine_day_diff_co.get(m, 0) + 1   # Part 2 budget
                    s["campaigns"].append((best, qty, co_type))
                    _sg_move_commit(m, best)         # SG_DELIB: sanction an admitted cross-group move
                    projected_gt[best] = projected_gt.get(best, 0.0) + qty
                    _fwd_added += qty
                    if _pace_served is not None:
                        _pace_served.add(best)       # PACING: mark served → widen next picks
                    s["remaining"] -= (cost + mins)
                    _conc_commit(m, best, qty)       # CONCENTRATION: count the Phase-C pre-build
                    if best != cur:
                        s["cur_sku"] = best
                        s["co_count"] += 1
                        machine_day_co[m] = machine_day_co.get(m, 0) + 1   # S2_CAMPAIGN per-day budget

        return {m: stg[m]["campaigns"] for m in machines if stg[m]["campaigns"]}

    if _MACHINE_ORDER_REVERSED:
        _pri = {"STAGE2": 0, "UNISTAGE": 1, "BJ": 2, "VMI": 3, "STAGE1": 9}
    else:
        _pri = {"VMI": 0, "BJ": 1, "UNISTAGE": 2, "STAGE2": 3, "STAGE1": 9}
    if _RI_RATIO_GLOBAL and demand_dict is not None and machine_total_demand is not None:
        # Approach 2: ratio-driven machine order — replace the fixed VMI->BJ->US
        # ->Stage2 group order with DESCENDING best-ratio order, so the machine
        # whose best-fit deficit SKU has the highest ratio claims first. A SKU is
        # "in play" for this ordering if it has curing demand this shift and unmet
        # demand remaining. Stage-1 last (carcass). Group priority + machine name
        # break ties deterministically.
        def _best_ratio(m: str) -> float:
            mtd  = machine_total_demand.get(m, 1e-9)
            best = 0.0
            for s in machine_skus.get(m, ()):
                if shift_cure_demand.get(s, 0.0) > 0 and demand_remaining.get(s, 0.0) > 0:
                    r = demand_dict.get(s, 0.0) / mtd
                    if r > best:
                        best = r
            return best
        sorted_machines = sorted(
            machine_skus.keys(),
            key=lambda m: (
                9 if _MACHINE_GROUP.get(m, "") == "STAGE1" else 0,
                -_best_ratio(m),
                _pri.get(_MACHINE_GROUP.get(m, ""), 9),
                m,
            ),
        )
    elif _SCARCITY_ORDER_ENABLED:
        # Scarcity-first: machines with the FEWEST eligible SKUs claim their
        # shared-SKU deficit before flexible machines can poach it. A captive
        # machine (e.g. 7301, eligible for only LSTL0) can build nothing else,
        # so it must lock in its specialty first — freeing flexible machines
        # (which share that SKU but have many others) for SKUs only they serve.
        # Stage-1 stays last (carcass, not GT). Eligible-count is the primary
        # key; group priority + machine name break ties deterministically.
        sorted_machines = sorted(
            machine_skus.keys(),
            key=lambda m: (
                9 if _MACHINE_GROUP.get(m, "") == "STAGE1" else 0,
                len(machine_skus.get(m, ())),
                _pri.get(_MACHINE_GROUP.get(m, ""), 9),
                m,
            ),
        )
    else:
        sorted_machines = sorted(
            machine_skus.keys(),
            key=lambda m: (_pri.get(_MACHINE_GROUP.get(m, ""), 9), m),
        )

    # Captive-first: stable re-sort putting non-Stage-1 captive machines (exactly
    # one eligible SKU) at the FRONT so they claim their sole SKU before flexible
    # machines drain its deficit. Stable → preserves the base order within groups.
    if _CAPTIVE_FIRST_ENABLED:
        sorted_machines = sorted(
            sorted_machines,
            key=lambda m: 0 if (len(machine_skus.get(m, ())) == 1
                                and _MACHINE_GROUP.get(m, "") != "STAGE1") else 1,
        )

    plan: dict[str, list] = {}

    for machine in sorted_machines:
        group = _MACHINE_GROUP.get(machine, "")
        _buf = GT_BUFFER_SHIFTS_VMI if group == "VMI" else GT_BUFFER_SHIFTS_OTHER

        def _deficit(sku: str, _b: float = _buf) -> float:
            built_ahead = projected_gt.get(sku, 0.0)
            need = shift_cure_demand.get(sku, 0.0) * _b
            gap  = need - built_ahead
            # Hard demand cap — subtract projected_gt (= gt_inventory + GT already
            # built THIS shift by earlier machines). demand_remaining is cured-
            # decremented and projected_gt is build-tracked, so
            # (demand_remaining - projected_gt) = D - total_GT_built_so_far.
            # Without the projected_gt term, two machines eligible for the same
            # SKU each build up to the full remaining demand in one shift → the
            # 1-4% overbuild seen across ~26 SKUs. This keeps total build ≤ demand.
            cap  = demand_remaining.get(sku, 0.0) - built_ahead - _woc.get(sku, 0.0)
            return min(max(0.0, gap), max(0.0, cap))

        def _priority_tier(sku: str, d: float) -> tuple:
            # RI (has a live/CO'd press) always outranks NRI — tier 0 < tier 1.
            # NRI candidates are ranked by static demand[sku]/machine_total_demand[machine]
            # (never decremented) instead of raw deficit magnitude, so thin NRI SKUs aren't
            # structurally starved by high-volume ones competing for the same machine.
            _is_ri = press_count.get(sku, 0) > 0
            # Approach 1/2: when RI-ratio is on, RI candidates are ALSO ranked by
            # ratio (not raw deficit). RI keeps tier 0 (outranks NRI); only the
            # within-RI ordering switches from -deficit to -ratio.
            _ri_ratio = (_RI_RATIO_ENABLED or _RI_RATIO_GLOBAL) and _is_ri
            if (
                _BUILDING_RATIO_ENABLED
                and (press_count.get(sku, 0) <= 0 or _ri_ratio)
                and demand_dict is not None
                and machine_total_demand is not None
            ):
                ratio = demand_dict.get(sku, 0.0) / machine_total_demand.get(machine, 1e-9)
                return (0 if _is_ri else 1, -ratio)
            return (0, -d)

        eligible = machine_skus.get(machine, set())
        if not any(_deficit(s) > 0 for s in eligible):
            continue

        remaining = max(0.0, _sbud - float((machine_down_mins or {}).get(str(machine), 0.0)))   # PM/MTC
        co_count  = 0
        MAX_COS   = _max_cos(machine)
        cur_sku   = machine_current_sku.get(machine, "")
        cur_inch  = sku_inch.get(cur_sku, "")
        dom_inch  = _MACHINE_DOMINANT_INCH.get(str(machine), cur_inch)
        rate      = _bld_qty_per_shift(machine) / SHIFT_MINS
        campaigns: list[tuple] = []

        # ── Round-trip buffer sizing for cur_sku ────────────────────────────
        # Skip conditions (fall back to flat _buf): (1) machine has only one
        # eligible SKU — len(eligible)<=1, nowhere to rotate to; (2) no other
        # eligible SKU has unmet demand — demand_remaining<=0 filters it out;
        # (3) no other eligible SKU currently has a real curing-driven deficit
        # — _deficit(s)<=0 (flat-buf, no circularity) filters it out. When a
        # genuine rotation partner exists, size cur_sku's buffer to survive
        # CO(cur->partner) + partner's own deficit-driven dwell (floored at
        # MIN_CAMPAIGN_MINS) + CO(partner->cur), so cur_sku's press doesn't
        # starve while this machine is away serving the partner.
        effective_buf = _buf
        if _ROUND_TRIP_BUFFER_ENABLED and cur_sku and len(eligible) > 1:
            partner_candidates = [
                s for s in eligible
                if s != cur_sku
                and demand_remaining.get(s, 0.0) > 0
                and _deficit(s) > 0
            ]
            if partner_candidates and _RT_SAME_INCH:             # prefer a same-inch rotation partner
                _si_pc = [s for s in partner_candidates if sku_inch.get(s, "") == cur_inch]
                if _si_pc and (max(_deficit(s) for s in _si_pc)
                               >= _RT_SAME_INCH_FRAC * max(_deficit(s) for s in partner_candidates)):
                    partner_candidates = _si_pc                  # same-inch partner is nearly as needy
            if partner_candidates:
                if machine in _INCH_FLEX_MACHINES:
                    # Flex machines may rotate to an OFF-inch partner (expensive
                    # diff-inch CO both ways). Size the own-inch buffer for the
                    # worst-case round trip so the own-inch press never starves
                    # while the machine is away — pick the partner with the
                    # largest CO round-trip cost, then deficit, then name.
                    partner = max(
                        partner_candidates,
                        key=lambda s: (
                            _co_cost(machine, cur_inch, sku_inch.get(s, ""))
                            + _co_cost(machine, sku_inch.get(s, ""), cur_inch),
                            _deficit(s), s,
                        ),
                    )
                else:
                    partner = max(partner_candidates, key=lambda s: (_deficit(s), s))
                partner_inch = sku_inch.get(partner, "")
                partner_dwell = max(
                    MIN_CAMPAIGN_MINS,
                    _deficit(partner) / rate if rate > 0 else MIN_CAMPAIGN_MINS,
                )
                round_trip_mins = (
                    _co_cost(machine, cur_inch, partner_inch)
                    + partner_dwell
                    + _co_cost(machine, partner_inch, cur_inch)
                )
                effective_buf = max(_buf, round_trip_mins / SHIFT_MINS)

        # ── Campaign 1: continue current SKU (no CO cost) ──────────────────
        if not cur_sku:
            best_start = min(
                (s for s in eligible if _deficit(s) > 0),
                key=lambda s: (
                    0 if sku_inch.get(s, "") == dom_inch else 1,
                    *_priority_tier(s, _deficit(s)),
                    s,
                ),
                default=None,
            )
            if best_start is not None:
                cur_sku  = best_start
                cur_inch = sku_inch.get(cur_sku, "")

        # urgent_co_set: co_target SKUs eligible on this machine with zero GT inventory.
        # These bypass allow_new_co=False guard in Campaign 2+ and the 30% cost guard.
        urgent_co_set = frozenset(
            s for s in co_target_skus
            if s in eligible and s != cur_sku
            and projected_gt.get(s, 0.0) == 0
            and demand_remaining.get(s, 0.0) > 0
        )

        # starving_set: eligible SKUs whose curing presses are running THIS shift
        # but have zero GT to cure (they will starve). Generalizes urgent_co_set
        # to any starving SKU so an otherwise-idle machine can feed it as a last
        # resort — see _STARVATION_FEED_ENABLED. Only relaxes the 30% cost guard;
        # inch preference and MIN_CAMPAIGN feasibility are untouched.
        starving_set = frozenset(
            s for s in eligible
            if s != cur_sku
            and shift_cure_demand.get(s, 0.0) > 0
            and projected_gt.get(s, 0.0) <= 0
            and demand_remaining.get(s, 0.0) > 0
        ) if _STARVATION_FEED_ENABLED else frozenset()

        # Default: allow non-dominant inch in Campaign 2+ unless Campaign 1 had unfinished demand.
        primary_demand_done = True

        # Reclamation guard (flex machines): if the machine carried over onto an
        # OFF-inch SKU but a dominant-inch SKU now has real deficit, do NOT
        # continue the off-inch campaign — force a return to the dominant inch
        # via Campaign 2+ (merged sort prefers dominant). This is the anti-
        # carry-over protection that lets us generalize the 7001/7003 soft-lock
        # to the high-demand VMIMAXX inches without the prior regression.
        # Disabled under the client inch rules: forcing a return to the dominant
        # inch IS a revisit, which Rule 1a forbids. (Expect the inch-flex gain to
        # go with it — that is the cost of one-way inch movement.)
        _flex_reclaim = (
            not _INCH_RULES_ENABLED
            and machine in _INCH_FLEX_MACHINES
            and cur_inch != dom_inch
            and any(sku_inch.get(s, "") == dom_inch and _deficit(s) > 0 for s in eligible)
        )
        if (cur_sku in eligible and _deficit(cur_sku, effective_buf) > 0
                and not _flex_reclaim):
            _ra = _bld_qty_per_shift(machine, cur_sku) / SHIFT_MINS   # per-SKU CT rate
            mins = min(remaining, _deficit(cur_sku, effective_buf) / _ra if _ra > 0 else remaining)
            qty  = int(mins * _ra)
            if mins >= MIN_CAMPAIGN_MINS and qty > 0:
                campaigns.append((cur_sku, qty, "start"))
                projected_gt[cur_sku] = projected_gt.get(cur_sku, 0.0) + qty
                remaining -= mins

            # Track whether Campaign 1 exhausted demand (vs cut by shift time).
            # Soft-lock machines may serve non-dominant inch only when primary is done.
            primary_demand_done = _deficit(cur_sku, effective_buf) <= 0

        # ── Campaign 2+: CO to deficit SKUs ──────────────────────────────
        _is_flex = machine in _INCH_FLEX_MACHINES
        while remaining >= MIN_CAMPAIGN_MINS and co_count < MAX_COS:
            same_cands: list = []
            diff_cands: list = []
            flex_cands: list = []
            seen_in_plan = {sku for sku, _, _ in campaigns}

            for sku in eligible:
                d = _deficit(sku)
                if sku == cur_sku or d <= 0:
                    continue
                # 4-SKU/day cap (legacy path): block a CO that would be the 5th SKU today.
                if _sku_cap_blocks(machine, sku, seen_in_plan):
                    continue
                to_inch = sku_inch.get(sku, "")
                is_urgent = sku in urgent_co_set or sku in starving_set
                # Off-inch gate: soft-lock AND flex machines serve non-dominant
                # inch ONLY when primary (dominant, buffer-sized) demand is done
                # this shift. BJ/other machines were never locked — unchanged.
                if (machine in (_SOFT_LOCK_MACHINES | _INCH_FLEX_MACHINES)
                        and to_inch != dom_inch and not primary_demand_done):
                    continue
                cost = _co_cost(machine, cur_inch, to_inch, from_sku=cur_sku, to_sku=sku)
                if remaining - cost < MIN_CAMPAIGN_MINS:
                    continue
                # 30% cost guard normally blocks expensive COs. Bypass for urgent
                # co_target SKUs, AND for a flex machine taking an off-inch CO once
                # its own inch is done this shift (primary_demand_done): the diff-
                # inch CO (up to 180 min) exceeds 30% of a shift, but the machine
                # would otherwise IDLE the whole shift, so any production beats idle.
                _flex_offinch_ok = (_is_flex and to_inch != dom_inch and primary_demand_done)
                if cost > 0.30 * remaining and not is_urgent and not _flex_offinch_ok:
                    continue
                revisit_penalty = 1 if sku in seen_in_plan else 0
                inch_penalty    = 0 if to_inch == dom_inch else 1
                tier, primary   = _priority_tier(sku, d)
                if _is_flex:
                    # Flex tuple puts inch_penalty FIRST so dominant inch always
                    # wins (strict reclamation), then off_key orders WITHIN the
                    # off-inch group only (dominant candidates get neutral 0).
                    if inch_penalty == 0:
                        off_key = 0.0
                    elif _INCH_FLEX_OFFINCH_ORDER == "starving_first":
                        _starving = (shift_cure_demand.get(sku, 0.0) > 0
                                     and projected_gt.get(sku, 0.0) <= 0
                                     and demand_remaining.get(sku, 0.0) > 0)
                        off_key = 0.0 if _starving else 1.0
                    else:  # "demand_first" — largest deficit first (best CO amortization)
                        off_key = -float(d)
                    flex_cands.append((inch_penalty, off_key, tier, primary,
                                       revisit_penalty, cost, sku))
                else:
                    bucket = same_cands if to_inch == cur_inch else diff_cands
                    bucket.append((tier, primary, inch_penalty, revisit_penalty, cost, sku))

            if _is_flex:
                if flex_cands:
                    flex_cands.sort()
                    best_cost, best_sku = flex_cands[0][-2], flex_cands[0][-1]
                    co_type = ("same_size_CO"
                               if sku_inch.get(best_sku, "") == cur_inch else "diff_size_CO")
                    if os.environ.get("FLEX_DEBUG"):
                        _noff = sum(1 for c in flex_cands if c[0] == 1)  # inch_penalty==1
                        if _noff:
                            print(f"    [FLEX] {machine} cur_inch={cur_inch} "
                                  f"off_cands={_noff} best={best_sku}({sku_inch.get(best_sku,'')}) "
                                  f"co={co_type} rem={remaining:.0f}")
                else:
                    break
            elif same_cands:
                same_cands.sort()
                _, _, _, _, best_cost, best_sku = same_cands[0]
                co_type = "same_size_CO"
            elif diff_cands:
                diff_cands.sort()
                _, _, _, _, best_cost, best_sku = diff_cands[0]
                co_type = "diff_size_CO"
            else:
                break

            avail = remaining - best_cost
            _ra = _bld_qty_per_shift(machine, best_sku) / SHIFT_MINS   # per-SKU CT rate
            mins  = min(avail, _deficit(best_sku) / _ra if _ra > 0 else avail)
            qty   = int(mins * _ra)
            if mins < MIN_CAMPAIGN_MINS or qty <= 0:
                break

            campaigns.append((best_sku, qty, co_type))
            projected_gt[best_sku] = projected_gt.get(best_sku, 0.0) + qty
            remaining -= (best_cost + mins)
            co_count  += 1
            cur_sku    = best_sku
            cur_inch   = sku_inch.get(cur_sku, "")

        if campaigns:
            plan[machine] = campaigns

    return plan


def _writeoff_stale_gt(gt_inventory, last_build_day, current_day, shelf_days=GT_SHELF_LIFE_DAYS):
    total = 0.0
    for sku in list(gt_inventory.keys()):
        qty = gt_inventory[sku]
        if qty > 0 and (current_day - last_build_day.get(sku, 0)) > shelf_days:
            total += qty
            gt_inventory[sku] = 0.0
    return total


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT WRITERS — same sheet names as legacy pipeline
# ══════════════════════════════════════════════════════════════════════════════

def _xl_header(ws, row: int, cols: list, bg="1F3864", fg="FFFFFF"):
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, size=10, color=fg)
    aln  = Alignment(horizontal="center", vertical="center")
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill, c.font, c.alignment = fill, font, aln


def _xl_fill(ws, row_num: int, n_cols: int, hex_color: str):
    from openpyxl.styles import PatternFill
    fill = PatternFill("solid", fgColor=hex_color)
    for ci in range(1, n_cols + 1):
        ws.cell(row=row_num, column=ci).fill = fill


def _working_days_count(plan_start, planning_days) -> int:
    """planning_days minus the bc_config.PLANT_HOLIDAYS dates that fall in [1..planning_days].
    No holidays ⇒ returns planning_days (parity). Used for utilization/availability denominators."""
    try:
        base = plan_start.date() if hasattr(plan_start, "date") else plan_start
        n = 0
        for _h in (getattr(_bc_cfg, "PLANT_HOLIDAYS", []) or []):
            _idx = (datetime.strptime(str(_h).strip(), "%Y-%m-%d").date() - base).days + 1
            if 1 <= _idx <= planning_days:
                n += 1
        return planning_days - n
    except Exception:
        return planning_days


def _bc_holiday_day_set(planning_days) -> set:
    """1-based day indices in [1..planning_days] that are plant holidays (bc_config.PLANT_HOLIDAYS)."""
    out: set = set()
    try:
        ps = getattr(_bc_cfg, "PLAN_START", None)
        base = ps.date() if hasattr(ps, "date") else ps
        for _h in (getattr(_bc_cfg, "PLANT_HOLIDAYS", []) or []):
            _idx = (datetime.strptime(str(_h).strip(), "%Y-%m-%d").date() - base).days + 1
            if 1 <= _idx <= planning_days:
                out.add(_idx)
    except Exception:
        pass
    return out


def _bc_working_days_left(day, planning_days) -> int:
    """Working (non-holiday) days in [day..planning_days] inclusive. Urgency/CO-timing horizon.
    No holidays ⇒ planning_days - day + 1 (byte-for-byte parity)."""
    hol = _bc_holiday_day_set(planning_days)
    if not hol:
        return planning_days - day + 1
    return sum(1 for _d in range(day, planning_days + 1) if _d not in hol)


def _split_rows_at_shift_boundaries(rows, mkey="Machine", even_qty=False, mpq=0,
                                     holiday_windows=None):
    """Split each row's [StartTime,EndTime] run at the plant shift boundaries
    (07:00 / 15:00 / 23:00) into one row per shift, and remove cross-shift OVERLAP by
    sequencing each machine/press's rows on a continuous wall-clock cursor.
    Qty carries NO decimals (floored). `even_qty=True` (CURING) makes every shift's Qty EVEN
    by counting whole CYCLES (2 cavities = 2 tyres): a cycle straddling a shift boundary is
    credited to its COMPLETION shift, so no half-cycle (odd) output ever appears.
    Totals are preserved (remainder on the last/completion segment).

    `holiday_windows` (optional): a list of [start_dt, end_dt) plan-day-index-derived plant
    holiday windows (e.g. one holiday plan-day = [holiday_date 07:00, next_date 07:00)).
    HOLIDAY FIX: a machine (e.g. Stage-1 carcass) whose cumulative wall-clock cursor has
    drifted behind its nominal per-shift schedule (an earlier PM/MTC window it had to skip
    OVER, see below) can otherwise have that drift carry it straight through a later holiday
    gap in its own row list and land production inside the holiday's wall-clock window — even
    though no row is ever dated/produced on the holiday plan-day itself. Every holiday window
    is treated exactly like a per-machine PM/MTC window (skip OVER it, no quantity dropped,
    resume on the next working shift) but applies to EVERY machine, regardless of the
    `_PM_MTC_NO_OVERLAP` toggle. Empty/None ⇒ identical to the old behavior (bit-for-bit)."""
    from datetime import datetime as _dt, timedelta as _td
    _hol_w = list(holiday_windows or [])

    def _p(s):
        try:
            return _dt.strptime(str(s), "%Y-%m-%d %H:%M")
        except Exception:
            return None

    def _f(dt):
        return dt.strftime("%Y-%m-%d %H:%M")

    def _shift_of(dt):
        h = dt.hour
        if 7 <= h < 15:
            return dt.strftime("%Y-%m-%d"), "A"
        if 15 <= h < 23:
            return dt.strftime("%Y-%m-%d"), "B"
        if h >= 23:
            return dt.strftime("%Y-%m-%d"), "C"
        return (dt - _td(days=1)).strftime("%Y-%m-%d"), "C"

    def _win_end(dt):
        h = dt.hour
        d = dt.replace(minute=0, second=0, microsecond=0)
        if 7 <= h < 15:
            return d.replace(hour=15)
        if 15 <= h < 23:
            return d.replace(hour=23)
        if h >= 23:
            return (d + _td(days=1)).replace(hour=7)
        return d.replace(hour=7)

    by = defaultdict(list)
    for r in rows:
        by[str(r.get(mkey, ""))].append(r)
    out = []
    for _m, rr in by.items():
        rr = sorted(rr, key=lambda r: str(r.get("StartTime", "")))
        cursor = None
        carry = {}   # CURING even-Qty: per-SKU odd tyre carried to this press's next cure of the SAME sku
        for r in rr:
            # 2-day plant playback: replay rows carry their exact plant (Date, Shift, Qty) and must
            # NOT be re-sequenced/split — the plant qty may fill a whole shift AND carry a CO, which
            # would else push the overflow into the next shift. Pass through untouched (no cursor).
            if r.get("_replay"):
                out.append(r)
                continue
            s = _p(r.get("StartTime"))
            e = _p(r.get("EndTime"))
            if s is None or e is None or e <= s:
                out.append(r)
                continue
            if cursor is not None and s < cursor:          # remove cross-shift overlap
                dur = e - s
                s = cursor
                e = s + dur
            # PM/MTC NO-OVERLAP (building only): the cross-shift cursor above can bump a building
            # machine's production/CO INTO a maintenance window (e.g. a prior campaign's cross-midnight
            # tail pushes the cursor past the window start). Keep the emitted [StartTime,EndTime] clear
            # of the window WITHOUT dropping any quantity: production skips OVER a window (the window is
            # an idle gap; the same production minutes resume after it, extending wall-clock — no clamp,
            # so the sheet still reconciles to the built/cured KPI and mould feasibility). A CO
            # (indivisible) is moved to start just AFTER any window it would collide with.
            _mw = None
            if e > s:
                _pm_w = (_BLD_DOWN.get(str(r.get(mkey, "")).split(".")[0]) or [])
                _w = (list(_pm_w) if (_PM_MTC_NO_OVERLAP and not even_qty) else []) + _hol_w
                if _w and _down_mins(_w, s, e) > 0:      # only reshape a row that ACTUALLY hits a window
                    _mw = _w
            _is_co = _mw is not None and float(r.get("CO_Mins", 0) or 0) > 0 \
                     and int(float(r.get("Qty", 0) or 0)) == 0
            if _mw and _is_co:
                _co0 = float(r.get("CO_Mins", 0) or 0)
                for _ws, _we, *_rest in sorted(_mw):
                    if s < _we and _ws < s + timedelta(minutes=_co0):   # CO body overlaps window → after it
                        s = _we
                e = s + timedelta(minutes=_co0)
            total = (e - s).total_seconds() / 60.0
            qty = float(r.get("Qty", 0) or 0)
            qfloor = int(qty)          # FLOOR: the Qty column carries NO decimals (Qty >= 0)
            co = float(r.get("CO_Mins", 0) or 0)
            clean = float(r.get("Mould_Clean_Mins", 0) or 0)
            segs = []
            cur = s
            if _mw and not _is_co:
                # Lay `total` production minutes from s into the free (non-maintenance) time, skipping
                # OVER every maintenance window (a window is an idle gap) and breaking at shift
                # boundaries. No quantity is dropped — production resumes AFTER a window, extending
                # wall-clock into later shifts when the window leaves no room in this one (used for
                # Stage-1 carcass, whose long windows can cover a whole shift; GT machines are already
                # capped to post-maintenance minutes upstream so a GT run never enters this branch).
                # A window covering carcass same-day keeps its DAY-granular 1-day aging unchanged.
                _rem = total
                _guard = 0
                while _rem > 1e-9 and _guard < 500:
                    _guard += 1
                    _jump = None
                    for _ws, _we, *_rest in sorted(_mw):        # inside a window → jump to its end
                        if _ws <= cur < _we:
                            _jump = _we; break
                    if _jump is not None:
                        cur = _jump; continue
                    _lim = _win_end(cur)                         # next shift boundary
                    for _ws, _we, *_rest in sorted(_mw):        # or next window open, whichever first
                        if cur < _ws < _lim:
                            _lim = _ws; break
                    _avail = (_lim - cur).total_seconds() / 60.0
                    _take = min(_avail, _rem)
                    if _take > 1e-9:
                        segs.append((cur, cur + timedelta(minutes=_take)))
                        _rem -= _take
                    cur = _lim if _take >= _avail - 1e-9 else cur + timedelta(minutes=_take)
                if not segs:                                    # defensive: nothing placed → keep 1 row
                    segs.append((s, s))
                e = segs[-1][1]                                 # final wall-clock end (for the cursor)
            else:
                while cur < e:
                    we = min(_win_end(cur), e)
                    segs.append((cur, we))
                    cur = we
            # CURING even-Qty: fold in the odd tyre carried from this press's previous cure, emit
            # an EVEN Qty (whole cycles), and carry any new odd tyre forward. This preserves the
            # press's TOTAL (no per-row loss — only the very last odd tyre can strand), so the KPI
            # and GT mass-balance stay intact. A boundary cycle is credited to its COMPLETION shift.
            if even_qty and qfloor > 0:
                _sku = str(r.get("SKUCode", ""))
                _avail = qfloor + carry.get(_sku, 0)
                qeven = _avail - (_avail % 2)
                carry[_sku] = _avail % 2
            else:
                qeven = qfloor
            cyc_total = (qeven // 2) if even_qty else 0
            cyc_time = (total / cyc_total) if (even_qty and cyc_total > 0) else 0.0
            _qa = _coa = _cla = 0.0
            _cum = 0.0
            _cyc_prev = 0
            _pieces = []
            for _i, (ss, ee) in enumerate(segs):
                last = (_i == len(segs) - 1)
                seg_min = (ee - ss).total_seconds() / 60.0
                _cum += seg_min
                frac = seg_min / total if total > 0 else 1.0
                d2, sh2 = _shift_of(ss)
                nr = dict(r)
                nr["Date"], nr["Shift"] = d2, sh2
                nr["StartTime"], nr["EndTime"] = _f(ss), _f(ee)
                if qty:
                    if even_qty:
                        if last:
                            nr["Qty"] = qeven - _qa               # remainder to completion shift (even)
                        else:
                            _cyc_by = min(cyc_total, int(_cum / cyc_time)) if cyc_time > 0 else 0
                            nr["Qty"] = 2 * max(0, _cyc_by - _cyc_prev)   # whole completed cycles → even
                            _cyc_prev = _cyc_by
                    else:
                        nr["Qty"] = (qfloor - _qa) if last else int(qfloor * frac)   # floor, no decimals
                    _qa += nr["Qty"]
                if "CO_Mins" in r:
                    nr["CO_Mins"] = round(co - _coa, 1) if last else round(co * frac, 1)
                    _coa += nr["CO_Mins"]
                if "Mould_Clean_Mins" in r:
                    nr["Mould_Clean_Mins"] = round(clean - _cla, 1) if last else round(clean * frac, 1)
                    _cla += nr["Mould_Clean_Mins"]
                _pieces.append(nr)
            # MPQ (shift-level floor): a run split across a shift boundary can leave a sub-floor
            # wall-clock fragment. Fold each <mpq fragment into the adjacent piece of the SAME run
            # (total-preserving) so no sub-floor production block is EMITTED. (A whole run < mpq is
            # already prevented by the campaign-level MPQ guard; only split fragments reach here.)
            # SKIP the fold for a PM/MTC maintenance-machine run: the fold sets StartTime=min /
            # EndTime=max, which would merge two pieces ACROSS the maintenance-window gap and
            # re-create the very overlap the window-skip above removed. Leaving the (few) sub-MPQ
            # fragments as separate rows is total-preserving and keeps every row clear of the window.
            # BUG (found auditing Stage-1 carcass over-capacity rows, e.g. machine 6803 /
            # 2026-09-14 / SKU 1325119815106QBRQ0 emitted Qty=161 spanning 484 min into the
            # next shift): the fold below used to merge a sub-mpq fragment into ITS NEIGHBOUR
            # UNCONDITIONALLY, incl. across a shift boundary — that recombines exactly the two
            # pieces the boundary-split above exists to keep apart, reproducing a row whose
            # Qty exceeds its own shift's physical capacity (floor(SHIFT_MINS*60/ct)) and whose
            # EndTime spills past the shift. Now the merge only fires when the COMBINED
            # [min(StartTime), max(EndTime)] stays inside ONE shift window (`_shift_of` agrees
            # for the merged start and the merged end's last minute) — i.e. it may only fold
            # fragments that were already in the SAME shift (e.g. from overlap resequencing),
            # never a genuine cross-shift split artifact. If no neighbour is safe, the sub-mpq
            # fragment is left as its own (small) row rather than fabricating an over-capacity
            # one — total Qty is still preserved exactly, just not folded away cosmetically.
            if mpq > 0 and qfloor > 0 and len(_pieces) > 1 and not _mw:
                _guard = 0
                while len(_pieces) > 1 and _guard < 64:
                    _guard += 1
                    _sm = min(range(len(_pieces)), key=lambda k: int(_pieces[k].get("Qty", 0) or 0))
                    if int(_pieces[_sm].get("Qty", 0) or 0) >= mpq:
                        break
                    _cands = [k for k in (_sm - 1, _sm + 1) if 0 <= k < len(_pieces)]
                    _cands.sort(key=lambda k: -int(_pieces[k].get("Qty", 0) or 0))
                    _nb = None
                    for _c in _cands:
                        _ns = _p(min(_pieces[_c]["StartTime"], _pieces[_sm]["StartTime"]))
                        _ne = _p(max(_pieces[_c]["EndTime"],   _pieces[_sm]["EndTime"]))
                        if _ns is None or _ne is None:
                            continue
                        if _shift_of(_ns) == _shift_of(max(_ns, _ne - _td(minutes=1))):
                            _nb = _c
                            break
                    if _nb is None:
                        break     # no shift-safe neighbour — keep the sub-mpq fragment as-is
                    _pieces[_nb]["Qty"] = int(_pieces[_nb]["Qty"]) + int(_pieces[_sm].get("Qty", 0) or 0)
                    _pieces[_nb]["StartTime"] = min(_pieces[_nb]["StartTime"], _pieces[_sm]["StartTime"])
                    _pieces[_nb]["EndTime"]   = max(_pieces[_nb]["EndTime"], _pieces[_sm]["EndTime"])
                    for _mk in ("CO_Mins", "Mould_Clean_Mins"):
                        if _mk in _pieces[_sm]:
                            _pieces[_nb][_mk] = round(float(_pieces[_nb].get(_mk, 0) or 0)
                                                      + float(_pieces[_sm].get(_mk, 0) or 0), 1)
                    _pieces.pop(_sm)
            out.extend(_pieces)
            cursor = e
    return sorted(out, key=lambda r: (str(r.get("StartTime", "")), str(r.get(mkey, ""))))


def _write_rolling_building_excel(
    output_path: str,
    bld_shift_rows: list,          # per-shift rows (includes CO sentinels)
    bld_co_events: list,           # building machine CO events
    df_day0: "pd.DataFrame",       # Day 0 curing consumption (SKU classification)
    sku_machine_map: dict,         # {sku: set(machines)} for eligibility
    opening_gt: dict,              # opening GT inventory
    demand_dict: dict,             # {sku: demand_qty} from demand file
    planning_days: int,
    working_days: int = None,      # planning_days − holidays; None → planning_days (parity)
    n_curing_cos: int = 0,         # curing press CO count (from co_events)
    endday_gt_by_date: "dict | None" = None,  # {date_str: end-of-day total GT inventory}
    cure_ct_map: "dict | None" = None,        # {sku: ct} — for Skip_Reason data checks
    curing_allowable: "dict | None" = None,   # {sku: [presses]} — for Skip_Reason
    sku_moulds: "dict | None" = None,         # {sku: set(moulds)} — for Skip_Reason
    gt_waste_map: "dict | None" = None,       # {sku: expired GT units}  → "expired GT/carcass" col
    carcass_waste_map: "dict | None" = None,  # {sku: expired carcass units}
    expiry_rows: "list | None" = None,        # per-(day,shift,SKU) expired GT/carcass display rows
    holiday_dates: "set | None" = None,       # #4: plant-holiday date strings → Holiday flag + idle-day rows
    pm_mtc_rows: "list | None" = None,         # DISPLAY-ONLY MTC maintenance rows (CO_Type="MTC")
    bld_matrix_skus: "set | None" = None,      # SKUs present in RAW building allowable matrix (Skip_Reason)
    cur_master_skus: "set | None" = None,      # SKUs present in curing allowable master (Skip_Reason)
) -> None:
    """
    Write building Excel matching the legacy bc_building_schedule output.

    Sheets:
      1. Shift Schedule         — per-shift rows (production + carcass + CHANGEOVER; title row 1, header row 3)
      2. Changeover Plan        — building machine CO events
      3. SKU Classification     — category summary from Day 0 consumption
      4. Daily GT & Carcass     — daily GT and carcass totals
      5. Demand Fulfillment (B2C) — per-SKU demand vs planned GT
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _NAVY  = "1F3864"; _WHITE = "FFFFFF"; _GREEN = "C6E0B4"
    _AMBER = "FFE699"; _RED   = "FFE0E0"; _GREY  = "D3D3D3"
    _CO    = "FFC000"

    def _fill(h):   return PatternFill("solid", fgColor=h)
    def _bold(s=10): return Font(bold=True, size=s)
    def _ctr():      return Alignment(horizontal="center", vertical="center")

    # Present the shift schedule as a chronological timeline: ascending StartTime,
    # then Machine as a stable tiebreak. StartTime is "YYYY-MM-DD HH:MM" so its
    # lexical order IS chronological (post-midnight Shift-C rows carry the real
    # next-day date, so they sort correctly). Display-only — every aggregate
    # sheet below sums over the rows and is order-independent.
    bld_shift_rows = sorted(
        bld_shift_rows,
        key=lambda r: (str(r.get("StartTime", "")), str(r.get("Machine", ""))),
    )
    # SINGLE SOURCE OF TRUTH: split every row at shift boundaries (no cross-shift overlap)
    # ONCE here, so ALL sheets below (Shift Schedule, Daily GT & Carcass, Demand Fulfillment,
    # Machine Utilization) read the same shift-accurate rows. The split preserves per-
    # (machine,SKU) Qty/CO totals exactly, so per-SKU / per-machine sums and the KPIs are
    # unchanged; only cross-midnight rows attribute their tail to the correct calendar day.
    # HOLIDAY FIX (backstop): re-derive each holiday's wall-clock window ([date 07:00, next
    # date 07:00)) from `holiday_dates` (calendar date strings, plan-day-index-derived by the
    # caller) so a machine whose cumulative cursor has drifted behind schedule can never have
    # production land inside a holiday's window here either — see _split_rows_at_shift_
    # boundaries' HOLIDAY FIX docstring. Empty holiday_dates ⇒ bit-for-bit parity.
    _hol_wc_windows = []
    for _hd in (holiday_dates or ()):
        try:
            _hs = datetime.strptime(str(_hd)[:10], "%Y-%m-%d").replace(hour=7)
            _hol_wc_windows.append((_hs, _hs + timedelta(days=1)))
        except Exception:
            pass
    bld_shift_rows = _split_rows_at_shift_boundaries(bld_shift_rows, "Machine", mpq=_BUILDING_MPQ,
                                                       holiday_windows=_hol_wc_windows)

    _SENTINEL = {"CHANGEOVER", "MOULD_CLEAN", "C/O", "CO"}

    # Production tail-fold (DISPLAY-ONLY, KPI-NEUTRAL, all HARD rules preserved): the shift-boundary
    # split above can leave a sub-threshold tail slice (e.g. a block ending 15:20 → a tiny Shift-B
    # row) on both GT and carcass production. Fold each tail (< _CARCASS_CONSOLIDATE_MIN) into the
    # LARGEST sibling slice of the SAME (machine, date, SKU) that day, dropping the tail row. The
    # per-(date, machine, SKU) Qty total is preserved EXACTLY, so cured / coverage / demand-cap /
    # mass-balance and every HARD feasibility rule (R11B per-shift ≤480, R3B/R4/R8/R14/R17 …) are
    # untouched — only the display shift attribution of a few units moves. (The receiver's stated
    # Qty vs its unchanged time span shows as extra rows in the BENIGN R18B display-CT check, which
    # already fails on the Stage-1 carcass rows regardless.) CHANGEOVER / mould-clean / PM / MTC rows
    # are never touched. CARCASS_CONSOLIDATE=0 disables.
    if _CARCASS_CONSOLIDATE and _CARCASS_CONSOLIDATE_MIN > 0:
        from collections import defaultdict as _dd
        _MINF = _CARCASS_CONSOLIDATE_MIN
        def _is_prod(_r):
            return (int(_r.get("Qty", 0) or 0) > 0
                    and int(_r.get("CO_Mins", 0) or 0) == 0
                    and str(_r.get("SKUCode")) not in _SENTINEL)
        def _pdt(_s):
            try:
                return datetime.strptime(str(_s), "%Y-%m-%d %H:%M")
            except Exception:
                return None
        _grp = _dd(list)   # (machine, date, sku) -> [row,...]
        for r in bld_shift_rows:
            if _is_prod(r):
                _grp[(str(r["Machine"]), str(r["Date"])[:10], str(r["SKUCode"]))].append(r)
        _drop = set()
        for (_mch, _dt0, _sk), _rs in _grp.items():
            if len(_rs) < 2:
                continue
            if not any(0 < int(x.get("Qty", 0) or 0) < _MINF for x in _rs):
                continue
            # BUG (found auditing over-capacity carcass rows, e.g. machine 6803 / 2026-09-14
            # / SKU 1325119815106QBRQ0: two sub-threshold tails (27 + 2 units) both folded
            # BLINDLY into the day's single biggest slice — 132(cap) + 27 + 2 = 161, a Qty no
            # 480-min shift can physically build for this SKU's cycle time). The fold now only
            # accepts a tail into a sibling that has SPARE room in its OWN recorded shift span
            # (sibling's existing [StartTime,EndTime] duration + the tail's own production time
            # at this (machine,SKU) cycle time ≤ SHIFT_MINS) — never past a full shift, and
            # never past whatever less-than-full span the sibling already reflects (e.g. a
            # PM/MTC-shortened shift). A tail with no capacity-safe sibling is left as its own
            # (small) display row instead of being folded into a physically-impossible one.
            _ct = _bld_ct_sec(_mch, _sk)
            for _r in _rs:
                _q = int(_r.get("Qty", 0) or 0)
                if not (0 < _q < _MINF) or id(_r) in _drop:
                    continue
                _tail_min = (_q * _ct / 60.0) if _ct > 0 else 0.0
                _recv = None
                for _cand in sorted((x for x in _rs if x is not _r and id(x) not in _drop),
                                     key=lambda x: -int(x.get("Qty", 0) or 0)):
                    _cs, _ce = _pdt(_cand.get("StartTime")), _pdt(_cand.get("EndTime"))
                    _cand_span = (_ce - _cs).total_seconds() / 60.0 if (_cs and _ce) else 0.0
                    if _ct <= 0 or _cand_span + _tail_min <= SHIFT_MINS + 1e-6:
                        _recv = _cand
                        break
                if _recv is None:
                    continue     # no shift-capacity-safe sibling — keep this tail row as-is
                _recv["Qty"] = int(_recv.get("Qty", 0) or 0) + _q
                _drop.add(id(_r))
        if _drop:
            bld_shift_rows = [r for r in bld_shift_rows if id(r) not in _drop]

    # Carcass FINAL shift-cap enforcement (residual cumulative-rounding fix — see
    # `_enforce_carcass_shift_cap` docstring). Runs AFTER the boundary split + tail-fold
    # above so it sees the truly final rows; catches the 1-2-unit-over-strict-floor-cap
    # artifact the earlier fixes (over-480-min-span, tail-fold) did not target. Display-
    # only / KPI-neutral (carcass not in gt_inventory). CARCASS_SHIFT_CAP=0 disables.
    if _CARCASS_SHIFT_CAP_ENFORCE and any(r.get("CO_Type") == "carcass" for r in bld_shift_rows):
        bld_shift_rows, _cap_over, _cap_dropped, _cap_dropped_by_day = _enforce_carcass_shift_cap(
            bld_shift_rows, holiday_windows=_hol_wc_windows)
        # a row capped all the way down to 0 (its whole shift was PM/MTC/CO-consumed) is a
        # zero-production stub — drop it from the display rather than showing a 0-Qty row.
        bld_shift_rows = [r for r in bld_shift_rows
                           if not (r.get("CO_Type") == "carcass"
                                   and str(r.get("SKUCode")) != "CHANGEOVER"
                                   and int(round(r.get("Qty", 0) or 0)) <= 0)]
        if _cap_over:
            print(f"  [Stage-1 carcass] shift-cap enforce: {_cap_over:,} row(s) were over their "
                  f"strict per-shift floor cap (residual rounding artifact) -> re-capped"
                  + (f"; {_cap_dropped:,} unit(s) unbuildable that day (dropped, display-only): "
                     + ", ".join(f"{d}={q}" for d, q in sorted(_cap_dropped_by_day.items()))
                     if _cap_dropped else "; 0 units dropped (all excess re-homed same-day)") + ".")

    # HARD min-carcass floor on the display rows: fold any sub-MIN carcass slice left by the
    # renderer / shift-cap redistribution into a same-(machine,date,SKU) sibling (preserves the
    # day total → GT/curing sync + R5 untouched). Runs before the PM relocate so re-timed rows
    # still get maintenance-cleared. OFF (CARCASS_MIN_ENFORCE=0) → bit-for-bit.
    bld_shift_rows = _enforce_carcass_min_qty(bld_shift_rows)

    # PM/MTC no-overlap for carcass rows: qty was already reduced for maintenance upstream, but
    # the emitted time span could still straddle a window (a naive shift-start cursor). Relocate
    # each straddling carcass row into its shift's free sub-interval(s). Display-only / KPI-neutral
    # (mirrors the curing-side _pm_relocate_curing_rows). OFF (PM_MTC off) = bit-for-bit.
    bld_shift_rows = _pm_relocate_carcass_rows(bld_shift_rows)

    # ── SYNC: cap Stage-2 GT to displayed carcass per SKU (drop the extra GT) ─────────
    # After all carcass finalization (min-10 fold/drop, shift-cap, PM), the displayed carcass
    # per SKU is FINAL. Cap each Stage-2 SKU's total GT display rows to (carcass + opening
    # carcass) so no GT is shown without carcass backing → R5 (GT ≤ carcass) holds. Report the
    # per-SKU carcass cap so the caller reduces cured/coverage in sync (drop carcass+GT+cured
    # together). Non-Stage-2 GT (VMI/BJ/Unistage — no carcass) is untouched. OFF → bit-for-bit.
    _carcass_cap_by_sku: dict = {}
    _gt_drop_by_sku: dict = {}
    if _CARCASS_MIN_ENFORCE:
        from collections import defaultdict as _dd2
        _carc_tot = _dd2(float); _gt_rows = _dd2(list)
        _gt_drop_by_sku = _dd2(float)
        for r in bld_shift_rows:
            _sk = str(r.get("SKUCode", "")); _grp = _MACHINE_GROUP.get(str(r.get("Machine", "")), "")
            if r.get("CO_Type") == "carcass" and _sk != "CHANGEOVER":
                _carc_tot[_sk] += float(r.get("Qty", 0) or 0)
            elif (_grp == "STAGE2" and _sk not in _SENTINEL
                  and int(r.get("CO_Mins", 0) or 0) == 0 and float(r.get("Qty", 0) or 0) > 0):
                _gt_rows[_sk].append(r)
        for _sk, _rows in _gt_rows.items():
            _cap = _carc_tot.get(_sk, 0.0)                # displayed carcass (opening ignored → strict)
            _carcass_cap_by_sku[_sk] = _cap
            _tot = sum(float(r.get("Qty", 0) or 0) for r in _rows)
            _excess = _tot - _cap
            if _excess <= 1e-6:
                continue
            _gt_drop_by_sku[_sk] = _excess
            for r in sorted(_rows, key=lambda x: str(x.get("StartTime", "")), reverse=True):
                if _excess <= 1e-6:
                    break
                _q = float(r.get("Qty", 0) or 0); _d = min(_q, _excess)
                _newq = int(round(_q - _d)); _excess -= _d
                _ctm = _bld_ct_sec(str(r.get("Machine", "")), _sk)
                r["Qty"] = _newq
                _cs = None
                try:
                    _cs = datetime.strptime(str(r.get("StartTime")), "%Y-%m-%d %H:%M")
                except (TypeError, ValueError):
                    _cs = None
                if _cs and _ctm > 0:
                    r["EndTime"] = _fmt_dt(_cs + timedelta(minutes=_newq * _ctm / 60.0))
        # drop now-zero Stage-2 GT rows
        _zids = {id(r) for _rows in _gt_rows.values() for r in _rows
                 if int(round(r.get("Qty", 0) or 0)) <= 0}
        if _zids:
            bld_shift_rows = [r for r in bld_shift_rows if id(r) not in _zids]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Shift Schedule (header at row 3 — matches legacy header=2 read) ─
    ws = wb.create_sheet("Shift Schedule")
    ws.cell(row=1, column=1, value="BC Building Schedule (Rolling Pipeline)").font = _bold(12)
    # Row 2: blank
    # Row 3: headers
    # Qty = units produced (0 on CHANGEOVER rows — a CO makes no tyres).
    # CO_Mins = changeover duration in minutes (0 on production/carcass rows).
    bld_cols = ["Machine", "Date", "Shift", "SKUCode", "Qty", "CO_Mins",
                "StartTime", "EndTime", "Machine_Group", "CO_Type"]
    _xl_header(ws, 3, bld_cols)
    # Expired GT/carcass waste rows are NO LONGER interleaved here — they live in the
    # dedicated "expired" sheet (added below). The Shift Schedule now shows production +
    # building CO rows only. (Daily GT & Carcass Expired_GT/Expired_Carcass columns are a
    # separate aggregation over expiry_rows and are unaffected.)
    _EXP_TYPES = {"expired_GT", "expired_carcass"}
    # PM/MTC: DISPLAY-ONLY MTC maintenance rows (CO_Type="MTC"), merged into the Shift
    # Schedule for visibility only — NOT in prod_rows / KPIs / feasibility.
    _display_rows = sorted(
        list(bld_shift_rows) + list(pm_mtc_rows or []),
        key=lambda r: (str(r.get("StartTime", "")), str(r.get("Machine", ""))),
    )
    for ri, row in enumerate(_display_rows, 4):     # production + CO + MTC rows
        is_co  = str(row.get("SKUCode", "")).upper() in _SENTINEL
        is_mtc = str(row.get("CO_Type", "")) in ("MTC", "PM")
        for ci, col in enumerate(bld_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.alignment = _ctr()
            if is_mtc:
                cell.fill = _fill(_GREY)
                cell.font = Font(bold=True)
            elif is_co:
                cell.fill = _fill(_CO)
                cell.font = Font(bold=True)
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 38)

    # ── Sheet 2: Changeover Plan ───────────────────────────────────────────────
    ws_co = wb.create_sheet("Changeover Plan")
    co_cols = ["Machine", "Date", "Day", "From_SKU", "Target_SKU",
               "CO_Type", "CO_Cost_Mins", "CO_Day_Index", "Status"]
    _xl_header(ws_co, 1, co_cols)
    for ri, row in enumerate(bld_co_events, 2):
        for ci, col in enumerate(co_cols, 1):
            ws_co.cell(row=ri, column=ci, value=row.get(col, "")).alignment = _ctr()
    for col in ws_co.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws_co.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 38)

    # ── Sheet 3: SKU Classification ────────────────────────────────────────────
    ws_cat = wb.create_sheet("SKU Classification")
    cat_counts: dict[str, dict] = defaultdict(lambda: {"SKU_Count": 0, "Total_Demand": 0, "Avg_Priority": []})
    if df_day0 is not None and not df_day0.empty:
        for _, r in df_day0.iterrows():
            cat = str(r.get("Category", "Unknown"))
            sku = str(r.get("SKUCode", ""))
            dem = float(r.get("Demand_Qty", 0) or 0)
            pri = float(r.get("Priority_Score", 0) or 0)
            cat_counts[cat]["SKU_Count"]    += 1
            cat_counts[cat]["Total_Demand"] += dem
            cat_counts[cat]["Avg_Priority"].append(pri)
    cat_data = [
        {"Category": cat, "SKU_Count": v["SKU_Count"],
         "Total_Demand": int(v["Total_Demand"]),
         "Avg_Priority": round(sum(v["Avg_Priority"]) / max(len(v["Avg_Priority"]), 1), 4)}
        for cat, v in sorted(cat_counts.items())
    ]
    cat_cols = ["Category", "SKU_Count", "Total_Demand", "Avg_Priority"]
    _xl_header(ws_cat, 1, cat_cols)
    for ri, row in enumerate(cat_data, 2):
        for ci, col in enumerate(cat_cols, 1):
            ws_cat.cell(row=ri, column=ci, value=row.get(col, "")).alignment = _ctr()
    # KPI footer
    n_bld_co = len(bld_co_events)   # match the Changeover Plan sheet (actual CO events), not sentinel shift-rows
    ws_cat.cell(row=len(cat_data) + 3, column=1, value="Building COs scheduled").font = _bold()
    ws_cat.cell(row=len(cat_data) + 3, column=2, value=n_bld_co)
    ws_cat.column_dimensions["A"].width = 22
    for ltr in "BCD":
        ws_cat.column_dimensions[ltr].width = 16

    # Production-only rows (CO/clean sentinels excluded) — used by the aggregate
    # sheets below (Daily GT & Carcass, Demand Fulfillment, Machine Utilization).
    # The separate "Shift Schedule (Clean)" sheet was removed; the full "Shift
    # Schedule" sheet (with production + carcass + CHANGEOVER rows) is the one output.
    prod_rows = [r for r in bld_shift_rows if str(r.get("SKUCode","")).upper() not in _SENTINEL]

    # ── Sheet 4: Daily GT & Carcass ────────────────────────────────────────────
    ws_daily = wb.create_sheet("Daily GT & Carcass")
    daily_agg: dict[str, dict] = defaultdict(lambda: {"GT_Produced": 0, "Carcass_Produced": 0,
                                                       "Total_Units": 0, "Active_SKUs": set()})
    for row in prod_rows:
        d    = str(row.get("Date", ""))
        mach = str(row.get("Machine", ""))
        qty  = int(row.get("Qty", 0) or 0)
        sku  = str(row.get("SKUCode", ""))
        if mach in _S1_MACHINES:
            daily_agg[d]["Carcass_Produced"] += qty
        else:
            daily_agg[d]["GT_Produced"] += qty
        daily_agg[d]["Total_Units"]  += qty
        daily_agg[d]["Active_SKUs"].add(sku)
    # Per-day expired GT / carcass (from the waste rows — NOT production, so kept out of
    # GT_Produced/Carcass_Produced above). Shows how much aged out each day (built-then-aged
    # + any opening Day-0 stock that expired), giving a shift/day-level view of waste.
    exp_by_day: dict[str, dict] = defaultdict(lambda: {"GT": 0, "Carcass": 0})
    for _er in (expiry_rows or []):
        _d = str(_er.get("Date", ""))
        if _er.get("CO_Type") == "expired_GT":
            exp_by_day[_d]["GT"] += int(_er.get("Qty", 0) or 0)
        elif _er.get("CO_Type") == "expired_carcass":
            exp_by_day[_d]["Carcass"] += int(_er.get("Qty", 0) or 0)
    # EndDay_GT_Inventory: total GT held overnight (all SKUs, after curing + writeoff)
    # — audits the MAX_ENDOFDAY_GT_INVENTORY plant cap directly in the sheet.
    _eod = endday_gt_by_date or {}
    # #4 Holiday reporting: a plant holiday builds/cures nothing, so it has NO production rows and
    # was previously MISSING from this sheet entirely (GT inventory carried but was invisible). Add
    # a Holiday flag column and force a row for every holiday date so the idle day, its carried GT,
    # and any GT/carcass that aged out DURING the holiday are all visible in day/shift context.
    # Per-day BUILDING occupancy (CU) — production + CO wall-clock span (idle & PM/MTC
    # maintenance excluded), over the full 41-machine roster (41×3×480 min/day). Same rule
    # used in the ad-hoc CU analysis. Read from bld_shift_rows (already shift-split).
    from datetime import datetime as _dt_occ
    _NB = 41
    _bld_busy: dict[str, float] = defaultdict(float)
    for _r in bld_shift_rows:
        if str(_r.get("CO_Type", "")) in ("MTC", "PM"):
            continue                                   # maintenance downtime is NOT busy
        _d = str(_r.get("Date", ""))[:10]
        try:
            _s = _dt_occ.strptime(str(_r.get("StartTime", "")).strip(), "%Y-%m-%d %H:%M")
            _e = _dt_occ.strptime(str(_r.get("EndTime", "")).strip(), "%Y-%m-%d %H:%M")
            _mn = (_e - _s).total_seconds() / 60.0
        except (TypeError, ValueError):
            _mn = float(_r.get("CO_Mins", 0) or 0)
        if _mn > 0:
            _bld_busy[_d] += min(_mn, 480.0)
    _bld_occ = {d: round(100.0 * b / (_NB * 3 * 480), 1) for d, b in _bld_busy.items()}
    _hol_set = set(holiday_dates or [])
    _all_dates = sorted(set(daily_agg.keys()) | _hol_set)
    daily_cols = ["Date", "Holiday", "GT_Produced", "Carcass_Produced", "Expired_GT", "Expired_Carcass",
                  "Total_Units", "Active_SKUs", "Cumulative_GT", "EndDay_GT_Inventory", "Building_Occupancy_%"]
    _xl_header(ws_daily, 1, daily_cols)
    cum_gt = 0
    for ri, date in enumerate(_all_dates, 2):
        v = daily_agg.get(date, {"GT_Produced": 0, "Carcass_Produced": 0,
                                 "Total_Units": 0, "Active_SKUs": set()})
        _is_hol = date in _hol_set
        cum_gt += v["GT_Produced"]
        vals = [date, ("Y" if _is_hol else ""), v["GT_Produced"], v["Carcass_Produced"],
                exp_by_day[date]["GT"], exp_by_day[date]["Carcass"],
                v["Total_Units"], len(v["Active_SKUs"]), cum_gt,
                int(round(_eod.get(date, 0))), _bld_occ.get(date, 0.0)]
        for ci, val in enumerate(vals, 1):
            cell = ws_daily.cell(row=ri, column=ci, value=val)
            cell.alignment = _ctr()
            if _is_hol:
                cell.fill = _fill(_GREY)
    ws_daily.column_dimensions["A"].width = 14
    for ltr in "BCDEFGHIJK":
        ws_daily.column_dimensions[ltr].width = 16

    # ── Sheet 6: Demand Fulfillment (B2C) ─────────────────────────────────────
    ws_dem = wb.create_sheet("Demand Fulfillment (B2C)")
    prod_by_sku: dict[str, int] = defaultdict(int)
    for row in prod_rows:
        sku = str(row.get("SKUCode", ""))
        if sku and sku.upper() not in _SENTINEL and str(row.get("Machine","")) not in _S1_MACHINES:
            prod_by_sku[sku] += int(row.get("Qty", 0) or 0)

    dem_cols = ["SKUCode", "Category", "Priority", "Demand", "GT_Inventory",
                "expired GT/carcass",                      # col F — right after GT_Inventory
                "Planned_Units", "Planned+GT", "Gap", "Fulfillment_Pct", "Status",
                "CycleTime_min", "Eligible_Machines", "Presses_Needed",
                "Skip_Reason"]
    _xl_header(ws_dem, 1, dem_cols)
    _gtw = gt_waste_map or {}
    _carw = carcass_waste_map or {}

    def _waste_cell(sku):
        # "50C" = 50 carcass aged-out; "100GT" = 100 GT aged-out; both → "50C, 100GT".
        c = int(round(_carw.get(sku, 0) or 0))
        g = int(round(_gtw.get(sku, 0) or 0))
        parts = []
        if c > 0:
            parts.append(f"{c}C")
        if g > 0:
            parts.append(f"{g}GT")
        return ", ".join(parts)

    cat_map_d0: dict[str, str]   = {}
    pri_map_d0: dict[str, float] = {}
    if df_day0 is not None and not df_day0.empty:
        for _, r in df_day0.iterrows():
            s = str(r.get("SKUCode","")).strip()
            cat_map_d0[s] = str(r.get("Category",""))
            pri_map_d0[s] = float(r.get("Priority_Score", 0) or 0)

    # Average building CT per SKU (seconds → minutes), across eligible machines
    _PRESS_NORM = 86_400.0  # 60 days × 24h × 60min (presses_needed normalisation)
    def _avg_bld_ct(sku: str):
        machs = sku_machine_map.get(sku, set())
        cts   = [_bld_ct_sec(m, sku) / 60.0 for m in machs]
        return round(sum(cts) / len(cts), 1) if cts else None

    dem_rows_out = []
    for sku, dem in sorted(demand_dict.items(), key=lambda x: -x[1]):
        planned_orig = float(prod_by_sku.get(sku, 0))
        gt_inv   = float(opening_gt.get(sku, 0))       # DB opening — shown as-is in GT_Inventory
        # EXPIRED GT is WASTE — it must NOT count as usable supply. Subtract the aged-out GT
        # (shown separately in the "expired GT/carcass" column) from Planned and Planned+GT so
        # they reflect only USABLE GT. Expiry is charged to built production first, then to the
        # opening inventory (both drawn from the same per-SKU pool). GT_Inventory keeps the true
        # DB opening. Carcass expiry is upstream (Stage-1), already reflected in built GT, so it
        # shows only in the waste column.
        _gt_exp  = float((gt_waste_map or {}).get(sku, 0.0))
        planned  = max(0.0, planned_orig - _gt_exp)                        # usable built GT
        _gt_usable = max(0.0, gt_inv - max(0.0, _gt_exp - planned_orig))   # usable opening GT
        planned_plus_gt = planned + _gt_usable
        gap      = max(0, int(dem) - int(planned_plus_gt))
        fill_pct = round(100 * planned_plus_gt / dem, 1) if dem > 0 else 0.0
        status   = ("FULLY MET" if planned_plus_gt >= dem * 0.95
                    else "PARTIAL" if planned_plus_gt > 0 else "UNMET")
        avg_ct   = _avg_bld_ct(sku)
        p_needed = (round(dem * avg_ct / _PRESS_NORM, 2) if avg_ct else "NA")
        dem_rows_out.append({
            "SKUCode": sku, "Category": cat_map_d0.get(sku, ""),
            "Priority": round(pri_map_d0.get(sku, 0), 7),
            "Demand": int(dem),
            "GT_Inventory": int(gt_inv),
            "Planned_Units": int(planned), "Planned+GT": int(planned_plus_gt),
            "Gap": gap,
            "Fulfillment_Pct": f"{fill_pct}%", "Status": status,
            "CycleTime_min": avg_ct if avg_ct is not None else "NA",
            "Eligible_Machines": len(sku_machine_map.get(sku, set())) or "NA",
            "Presses_Needed": p_needed,
            "expired GT/carcass": _waste_cell(sku),
            "Skip_Reason": _sku_data_skip_reasons(
                sku, sku_machine_map, cure_ct_map, curing_allowable, sku_moulds,
                bld_matrix_skus=bld_matrix_skus, cur_master_skus=cur_master_skus),
        })
    status_colors = {"FULLY MET": _GREEN, "PARTIAL": _AMBER, "UNMET": _RED}
    pu_col_idx = dem_cols.index("Planned_Units") + 1
    for ri, row in enumerate(dem_rows_out, 2):
        color = status_colors.get(row["Status"], _GREY)
        for ci, col in enumerate(dem_cols, 1):
            cell = ws_dem.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.fill = _fill(color)
            cell.alignment = _ctr()
            if ci == pu_col_idx:
                cell.font = Font(bold=True)
    ws_dem.column_dimensions["A"].width = 34
    for ltr in "BCDEFGHIJKLMN":
        ws_dem.column_dimensions[ltr].width = 15

    # KPI footer (matches building_b2c.py _append_b2c_sheets format)
    n_full  = sum(1 for r in dem_rows_out if r["Status"] == "FULLY MET")
    n_part  = sum(1 for r in dem_rows_out if r["Status"] == "PARTIAL")
    n_unmet = sum(1 for r in dem_rows_out if r["Status"] == "UNMET")
    tot_bld = sum(r["Planned_Units"] for r in dem_rows_out)
    tot_dem = sum(r["Demand"]        for r in dem_rows_out)
    tot_avail = sum(r["Planned+GT"]  for r in dem_rows_out)  # built + opening GT
    kpi_pct = round(100 * tot_bld / tot_dem, 1) if tot_dem else 0.0
    kpi_pct_avail = round(100 * tot_avail / tot_dem, 1) if tot_dem else 0.0
    n_co_bld = len(bld_co_events)   # same source as the Changeover Plan sheet (was sentinel shift-rows → off-by-one vs the sheet)
    footer = len(dem_rows_out) + 3
    ws_dem.cell(row=footer,   column=1, value="KPI SUMMARY").font = Font(bold=True)
    ws_dem.cell(row=footer+1, column=1, value="Total Customer Demand (units)")
    ws_dem.cell(row=footer+1, column=2, value=tot_dem)
    ws_dem.cell(row=footer+2, column=1, value="Total GT Built (units)")
    ws_dem.cell(row=footer+2, column=2, value=tot_bld)
    _kv = ws_dem.cell(row=footer+3, column=1, value="KPI — GT Built / Customer Demand")
    _kv.font = Font(bold=True)
    _kv2 = ws_dem.cell(row=footer+3, column=2, value=f"{kpi_pct}%")
    _kv2.font = Font(bold=True)
    _kv3 = ws_dem.cell(row=footer+4, column=1,
                       value="KPI — (GT Built + Opening GT) / Customer Demand")
    _kv3.font = Font(bold=True)
    _kv4 = ws_dem.cell(row=footer+4, column=2, value=f"{kpi_pct_avail}%")
    _kv4.font = Font(bold=True)
    ws_dem.cell(row=footer+5, column=1, value="Total SKUs in demand file")
    ws_dem.cell(row=footer+5, column=2, value=len(dem_rows_out))
    ws_dem.cell(row=footer+6, column=1, value="Fully Met (≥95% of demand, built+opening GT)")
    ws_dem.cell(row=footer+6, column=2, value=n_full)
    ws_dem.cell(row=footer+7, column=1, value="Partial (0 < built+opening GT < 95%)")
    ws_dem.cell(row=footer+7, column=2, value=n_part)
    ws_dem.cell(row=footer+8, column=1, value="Unmet (built+opening GT = 0)")
    ws_dem.cell(row=footer+8, column=2, value=n_unmet)
    ws_dem.cell(row=footer+9, column=1, value="Total Building COs")
    ws_dem.cell(row=footer+9, column=2, value=n_co_bld)
    ws_dem.cell(row=footer+10, column=1, value=f"Curing COs scheduled (≤{MAX_CHANGEOVERS_PER_DAY}/day)")
    ws_dem.cell(row=footer+10, column=2, value=n_curing_cos)

    # ── Sheet 7: Machine Utilization ──────────────────────────────────────────
    ws_util = wb.create_sheet("Machine Utilization")

    _GREEN_U = "C6E0B4"; _AMBER_U = "FFE699"; _RED_U = "FFE0E0"

    def _mgroup(m: str) -> str:
        if m in {"6001","6002","6003","6004","7001","7002","7003","7004"}: return "VMI"
        if m in {"7101","7102","7103","7104","7105","7106","7201"}:        return "BJ"
        if m in {"7501","7502","7503"}:                                    return "UNI_NARROW"
        if m in {"ps2","ps3","ps4"}:                                       return "PS"
        if m in {"8201","8301","8302","8501","8502","7301"}:               return "Stage-2"
        return "Stage-1"

    _wd = working_days if working_days is not None else planning_days   # holidays excluded
    avail_per_mach = _wd * 3 * SHIFT_MINS  # working-day availability (holiday shifts dropped)

    # Production time per machine
    mach_prod_mins: dict[str, float] = defaultdict(float)
    mach_carcass:   dict[str, int]   = defaultdict(int)
    mach_gt:        dict[str, int]   = defaultdict(int)
    mach_skus:      dict[str, set]   = defaultdict(set)
    for row in prod_rows:
        m   = str(row.get("Machine", ""))
        qty = int(row.get("Qty", 0) or 0)
        sku = str(row.get("SKUCode", ""))
        ct_sec = _bld_ct_sec(m, sku)
        mach_prod_mins[m] += qty * ct_sec / 60.0
        if m in _S1_MACHINES:
            mach_carcass[m] += qty
        else:
            mach_gt[m] += qty
        mach_skus[m].add(sku)

    # CO time per machine from bld_co_events
    mach_co_mins: dict[str, float] = defaultdict(float)
    mach_co_count: dict[str, int]  = defaultdict(int)
    for ev in bld_co_events:
        m    = str(ev.get("Machine", ""))
        cost = float(ev.get("CO_Cost_Mins", 0) or 0)
        mach_co_mins[m]  += cost
        mach_co_count[m] += 1

    # All 38 building machines — explicit set so zero-production machines (e.g. 8101)
    # are always included regardless of whether they appear in production or CO dicts.
    # (6801/bj1stage1 removed — plant retired it → 14 Stage-1.)
    _ALL_BUILDING_MACHINES = frozenset({
        # 6801 stays retired (excluded); ps3/ps4 LIVE 2026-09.
        "6802","6803","6909","6911","7601","7701",
        "7801","7802","7803","7804","8001","8002","8003","8101",  # Stage-1 (14)
        "8201","8301","8302","8501","8502","7301",                # Stage-2 (6)
        "7001","7002","7003","7004","6001","6002","6003","6004",  # VMI (8)
        "7101","7102","7103","7104","7105","7106","7201",         # BJ (7)
        "7501","7502","7503",                                     # UNI_NARROW (3)
        "ps2","ps3","ps4",                                        # NEW GT machines (3) — ps2=13", ps3=15", ps4=16"
    })
    all_machines = sorted(
        _ALL_BUILDING_MACHINES,
        key=lambda m: (
            {"VMI":0,"BJ":1,"UNI_NARROW":2,"PS":3,"Stage-2":4,"Stage-1":5}.get(_mgroup(m), 6),
            m
        )
    )

    # Summary header
    def _u_avg(machines):
        vals = [mach_prod_mins[m] / avail_per_mach for m in machines]
        return sum(vals) / len(vals) if vals else 0
    avg_gt_mach = [m for m in all_machines if _mgroup(m) != "Stage-1"]
    avg_util = _u_avg(avg_gt_mach) if avg_gt_mach else 0
    high = sum(1 for m in avg_gt_mach if mach_prod_mins[m]/avail_per_mach >= 0.80)
    low  = sum(1 for m in avg_gt_mach if mach_prod_mins[m]/avail_per_mach < 0.40)
    ws_util.cell(row=1, column=1, value=(
        f"Avg GT-machine util (prod only): {avg_util:.1%}  |  "
        f"High(≥80%): {high}  |  Low(<40%): {low}  |  "
        f"Note: Stage-1 always <77% by design (15 machines for 11.5-equiv S2 demand)"
    )).font = Font(bold=True, size=10)

    util_cols = [
        "Machine", "Machine_Group", "Available_Mins",
        "GT_Built", "Carcass_Built",
        "Prod_Mins", "CO_Mins", "Idle_Mins",
        "Util_Pct", "CO_Pct", "Idle_Pct", "Occupancy_Pct",
        "SKUs_Served", "COs_Done",
    ]
    # Util_Pct = production only. Occupancy_Pct = (prod + CO)/available — this is
    # the metric stored in jkt_plan_kpis / jkt_plan_capacityUtilisation.
    _xl_header(ws_util, 2, util_cols)

    for ri, m in enumerate(all_machines, 3):
        prod  = mach_prod_mins[m]
        co    = mach_co_mins[m]
        idle  = max(0.0, avail_per_mach - prod - co)
        util_pct = prod / avail_per_mach
        co_pct   = co   / avail_per_mach
        idle_pct = idle / avail_per_mach
        grp = _mgroup(m)
        # Color: red <40%, amber 40-85%, green ≥80% (Stage-1 always amber by design)
        if grp == "Stage-1":
            color = _AMBER_U
        elif util_pct >= 0.85:
            color = _GREEN_U
        elif util_pct >= 0.40:
            color = _AMBER_U
        else:
            color = _RED_U
        vals = [
            m, grp, avail_per_mach,
            mach_gt.get(m, 0), mach_carcass.get(m, 0),
            round(prod), round(co), round(idle),
            util_pct, co_pct, idle_pct, util_pct + co_pct,
            len(mach_skus[m]), mach_co_count[m],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws_util.cell(row=ri, column=ci, value=val)
            cell.fill = _fill(_GREEN_U if grp != "Stage-1" and util_pct >= 0.80 else
                              _AMBER_U if util_pct >= 0.40 or grp == "Stage-1" else _RED_U)
            cell.alignment = _ctr()
            if ci in (9, 10, 11, 12):  # percent columns
                cell.number_format = "0.0%"

    # Totals row
    tot_row = len(all_machines) + 3
    ws_util.cell(row=tot_row, column=1, value="TOTAL / AVERAGE").font = Font(bold=True)
    ws_util.cell(row=tot_row, column=3, value=avail_per_mach * len(all_machines)).font = Font(bold=True)
    ws_util.cell(row=tot_row, column=4, value=sum(mach_gt.values())).font = Font(bold=True)
    ws_util.cell(row=tot_row, column=5, value=sum(mach_carcass.values())).font = Font(bold=True)
    ws_util.cell(row=tot_row, column=6, value=round(sum(mach_prod_mins.values()))).font = Font(bold=True)
    ws_util.cell(row=tot_row, column=7, value=round(sum(mach_co_mins.values()))).font = Font(bold=True)
    tot_idle = max(0, avail_per_mach * len(all_machines)
                   - sum(mach_prod_mins.values()) - sum(mach_co_mins.values()))
    ws_util.cell(row=tot_row, column=8, value=round(tot_idle)).font = Font(bold=True)
    avg_pct = sum(mach_prod_mins.values()) / (avail_per_mach * len(all_machines))
    cell = ws_util.cell(row=tot_row, column=9, value=avg_pct)
    cell.font = Font(bold=True); cell.number_format = "0.0%"

    for ltr_idx, w in enumerate([14,16,16,12,14,12,10,10,10,10,10,12,12,10], 1):
        ws_util.column_dimensions[get_column_letter(ltr_idx)].width = w
    ws_util.freeze_panes = "A3"

    # ── Sheet: expired ─────────────────────────────────────────────────────────
    # All aged-out GT / carcass, moved OUT of the Shift Schedule into their own sheet.
    # expired_GT is SKU-level FIFO lot aging and expired_carcass is per-SKU bank aging —
    # neither is tied to a physical machine, so Machine / Machine Name render "—".
    ws_exp = wb.create_sheet("expired")
    # NOTE: "SKUCode Description" (after SKUCode) and "Machine Name" (after Machine) are added
    # automatically by _inject_label_columns post-write, so they are NOT listed here.
    exp_cols = ["Date", "Shift", "SKUCode", "Type", "Qty", "Machine"]
    _xl_header(ws_exp, 1, exp_cols)
    _type_lbl = {"expired_GT": "Expired GT", "expired_carcass": "Expired Carcass"}
    _exp_sorted = sorted(list(expiry_rows or []),
                         key=lambda r: (str(r.get("Date", "")), str(r.get("Shift", "")),
                                        str(r.get("CO_Type", "")), str(r.get("SKUCode", ""))))
    for ri, row in enumerate(_exp_sorted, 2):
        _mach = row.get("Machine", "—") or "—"
        vals = {"Date": row.get("Date", ""), "Shift": row.get("Shift", ""),
                "SKUCode": row.get("SKUCode", ""),
                "Type": _type_lbl.get(row.get("CO_Type", ""), row.get("CO_Type", "")),
                "Qty": row.get("Qty", 0), "Machine": _mach}
        for ci, col in enumerate(exp_cols, 1):
            cell = ws_exp.cell(row=ri, column=ci, value=vals.get(col, ""))
            cell.alignment = _ctr()
            cell.fill = _fill(_RED)   # waste marker
    for col in ws_exp.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws_exp.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 38)
    ws_exp.freeze_panes = "A2"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"  [Rolling] Building output → {output_path}")
    return _carcass_cap_by_sku, dict(_gt_drop_by_sku)   # carcass cap + GT dropped, per SKU (caller cured-sync)


def _write_rolling_curing_excel(
    output_path: str,
    cure_shift_rows: list,         # per-shift press events
    cure_co_events: list,          # curing press CO events (Planned + Dynamic) — Changeover Plan sheet
    mould_clean_events: list,      # curing press mould-clean events — Changeover Plan sheet
    press_stats: dict,             # {press: {running_mins, co_mins, clean_mins, skus, cycles, units}}
    press_sku_stats: dict,         # {(press,sku): {cycles, units, mins_used}}
    daily_cured: dict,             # {date_str: qty}
    sku_cured: dict,               # {sku: qty}
    closing_gt_bal: dict,          # {sku: gt_remaining}
    build_by_shift_sku: dict,      # {(date,shift): {sku: qty}} — for GT diagnostic
    opening_gt: dict,
    demand_dict: dict,             # {sku: demand_qty}
    cure_ct_map: dict,             # {sku: ct_min}
    curing_allowable: dict,        # {sku: [press_ids]}
    planning_days: int,
    plan_start: datetime,
    df_day0: "pd.DataFrame | None" = None,  # Day 0 consumption (has Priority_Score, Category)
    mould_life: "dict | None" = None,       # {press: remaining mould life (cycles) at horizon end}
    mould_info: "dict | None" = None,       # end-of-plan mould state for the Mould Tracker sheet
    sku_desc_map: "dict | None" = None,     # {sku: description} for the MouldInUse sheet
    sku_machine_map: "dict | None" = None,  # {sku: set(building machines)} — for Skip_Reason
    pm_mtc_rows: "list | None" = None,       # DISPLAY-ONLY PM maintenance rows (Remarks="PM Schedule")
    bld_matrix_skus: "set | None" = None,    # SKUs present in RAW building allowable matrix (Skip_Reason)
    cur_master_skus: "set | None" = None,    # SKUs present in curing allowable master (Skip_Reason)
) -> None:
    """
    Write curing Excel matching the legacy bc_curing_b2c output.

    Sheets:
      1. Demand Fulfillment  — per-SKU demand vs cured + fulfillment %
      2. Machine Utilization — per-press running / idle / CO minutes + utilization %
      3. Shift Schedule      — per-shift press events (RUNNING / CO / MOULD_CLEAN)
      4. Mould Tracker       — placeholder (mould cycle not tracked in rolling)
      5. Machine Schedule    — per (press, SKU) summary
      6. Daily Cured tyres   — daily cured totals
      7. GT Gap Diagnostic   — closing GT balance by SKU
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    _GREEN = "A9D08E"; _AMBER = "FFD966"; _RED   = "FFC7CE"; _LGREY = "D9D9D9"
    _NAVY  = "1F3864"; _WHITE = "FFFFFF"; _BLUE  = "DCE6F1"; _LYELL = "FFE699"
    _DGREY = "F2F2F2"; _ORANGE= "FFC000"

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _bold(s=10, color="000000"): return Font(bold=True, size=s, color=color)
    def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _hdr(ws, row, cols, bg=_NAVY, fg=_WHITE):
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = _fill(bg); c.font = _bold(10, fg); c.alignment = _ctr()

    # SINGLE SOURCE OF TRUTH: split every curing row at shift boundaries (no cross-shift
    # overlap) ONCE, so the Shift Schedule + MouldInUse sheets read shift-accurate rows.
    # even_qty=True → every shift's cured Qty is EVEN (2 cavities = 2 tyres/cycle; a boundary
    # cycle is credited to its completion shift). Preserves per-(press,SKU) totals.
    cure_shift_rows = _split_rows_at_shift_boundaries(cure_shift_rows, "Machine", even_qty=True, mpq=_CURING_MPQ)

    # CONTINUOUS CARRY display: the shift-boundary split can emit a press's continuous-cured shift
    # as a main row + a small straddling-cycle completion row (30 + 2 → 32). Consolidate same-
    # (press, date, shift, SKU) PRODUCTION rows into ONE row (sum Qty; EndTime = Start + Qty/2·CT)
    # so each shift shows the true per-shift cured (e.g. 32) and the per-shift feasibility rules see
    # one press = 2 mould cavities. Quantity-preserving; CO / clean / PM rows untouched.
    if _CURE_CYCLE_CARRY:
        def _p_dt(_s):
            try:
                return datetime.strptime(str(_s), "%Y-%m-%d %H:%M")
            except (TypeError, ValueError):
                return None
        _seen_pk: dict = {}; _cons_out = []
        for r in cure_shift_rows:
            _q = float(r.get("Qty", 0) or 0)
            if _q > 0 and str(r.get("_status", "RUNNING")) == "RUNNING" \
                    and str(r.get("SKUCode", "")) not in ("", "—"):
                _pk = (str(r.get("Machine")), str(r.get("Date"))[:10], r.get("Shift"),
                       str(r.get("SKUCode")))
                if _pk in _seen_pk:
                    _b = _seen_pk[_pk]
                    _b["Qty"] = float(_b.get("Qty", 0) or 0) + _q
                    _bs = _p_dt(_b.get("StartTime"))
                    _ctm = float(cure_ct_map.get(str(r.get("SKUCode")), DEFAULT_CURING_CT) or DEFAULT_CURING_CT)
                    if _bs is not None and _ctm > 0:
                        _b["EndTime"] = _fmt_dt(_bs + timedelta(minutes=(_b["Qty"] / CURING_CAVITIES) * _ctm))
                    continue
                _seen_pk[_pk] = r
            _cons_out.append(r)
        cure_shift_rows = _cons_out

    # PM/MTC no-overlap: relocate curing (Qty>0) rows that overlap a press's maintenance window
    # into the free minutes of their shift (post-split; plan-neutral, quantity-preserving).
    cure_shift_rows = _pm_relocate_curing_rows(cure_shift_rows)

    # ── Build priority + category lookup from df_day0 ─────────────────────────
    pri_map: dict[str, float] = {}
    cat_map: dict[str, str]   = {}
    if df_day0 is not None and not df_day0.empty:
        for _, r in df_day0.iterrows():
            s = str(r.get("SKUCode", "")).strip()
            pri_map[s] = float(r.get("Priority_Score", 0) or 0)
            cat_map[s] = str(r.get("Category", ""))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    avail_mins = _working_days_count(plan_start, planning_days) * 3 * SHIFT_MINS  # holiday shifts dropped

    # ── Sheet 1: Demand Fulfillment ───────────────────────────────────────────
    ws = wb.create_sheet("Demand Fulfillment")
    # CT_available: the SKU's real curing cycle time if it exists in the curing CT
    # table (cure_ct_map), else "NA" — meaning no data, so CycleTime_min fell back to
    # the DEFAULT_CURING_CT (17 min).
    cols = ["SKUCode", "Priority", "Demand", "GT_Inventory", "Planned_Units",
            "Gap", "Fulfillment_Pct", "Status", "CycleTime_min", "CT_available",
            "Eligible_Machines", "Presses_Needed", "Skip_Reason"]
    _hdr(ws, 1, cols)
    status_fill = {"FULLY MET": _GREEN, "PARTIAL": _AMBER, "UNMET": _RED, "NO DATA": _LGREY}
    rows_out = []
    for sku, dem in sorted(demand_dict.items(), key=lambda x: -x[1]):
        planned = float(sku_cured.get(sku, 0))
        gap     = max(0, dem - planned)
        pct     = planned / dem if dem > 0 else 0.0
        ct      = cure_ct_map.get(sku, DEFAULT_CURING_CT)
        cap_day = _cure_qty_per_shift(ct) * 3 * _working_days_count(plan_start, planning_days)
        p_needed= max(1, round(dem / cap_day)) if cap_day > 0 else "-"
        status  = ("FULLY MET" if planned >= dem * 0.999
                   else "PARTIAL" if planned > 0
                   else ("NO DATA" if dem <= 0 else "UNMET"))
        rows_out.append({
            "SKUCode": sku, "Priority": round(pri_map.get(sku, 0.0), 4),
            "Demand": int(dem),
            "GT_Inventory": int(opening_gt.get(sku, 0)),
            "Planned_Units": int(planned), "Gap": int(gap),
            "Fulfillment_Pct": pct, "Status": status,
            "CycleTime_min": round(ct, 2),
            "CT_available": round(cure_ct_map[sku], 2) if sku in cure_ct_map else "NA",
            "Eligible_Machines": len(curing_allowable.get(sku, [])),
            "Presses_Needed": p_needed,
            "Skip_Reason": _sku_data_skip_reasons(
                sku, sku_machine_map, cure_ct_map, curing_allowable,
                (mould_info or {}).get("sku_moulds"),
                bld_matrix_skus=bld_matrix_skus, cur_master_skus=cur_master_skus),
        })
    for ri, r in enumerate(rows_out, 2):
        f = _fill(status_fill.get(r["Status"], _WHITE))
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r[h])
            cell.fill = f; cell.alignment = _ctr()
            if h == "Fulfillment_Pct":
                cell.number_format = "0.0%"
    tr = len(rows_out) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = _bold(11)
    ws.cell(row=tr, column=3, value=sum(r["Demand"]        for r in rows_out)).font = _bold(11)
    ws.cell(row=tr, column=5, value=sum(r["Planned_Units"] for r in rows_out)).font = _bold(11)
    ws.cell(row=tr, column=6, value=sum(r["Gap"]           for r in rows_out)).font = _bold(11)
    tot_d = sum(r["Demand"] for r in rows_out); tot_p = sum(r["Planned_Units"] for r in rows_out)
    tc = ws.cell(row=tr, column=7, value=tot_p / tot_d if tot_d else 0)
    tc.font = _bold(11); tc.number_format = "0.0%"
    ws.column_dimensions["A"].width = 34
    for ltr in "BCDEFGHIJKL": ws.column_dimensions[ltr].width = 15
    ws.freeze_panes = "A2"

    # ── Sheet 2: Machine Utilization ──────────────────────────────────────────
    ws = wb.create_sheet("Machine Utilization")
    all_presses = sorted(press_stats)
    if all_presses:
        # Overall occupancy = (Σ Used + Σ CO + Σ Mould-Clean) / Σ Available across all
        # presses for the whole month (all presses share the same monthly Available_Mins).
        _busy    = sum(press_stats[p]["running_mins"] + press_stats[p]["co_mins"]
                       + press_stats[p]["clean_mins"] for p in all_presses)
        avg_u    = _busy / (len(all_presses) * avail_mins) if avail_mins else 0.0
        total_co = sum(press_stats[p]["co_mins"] for p in all_presses)
        _occ_p   = lambda p: ((press_stats[p]["running_mins"] + press_stats[p]["co_mins"]
                               + press_stats[p]["clean_mins"]) / avail_mins if avail_mins else 0.0)
        high     = sum(1 for p in all_presses if _occ_p(p) >= 0.90)
        low      = sum(1 for p in all_presses if _occ_p(p) < 0.05)
        ws.cell(row=1, column=1,
                value=(f"Avg util (occupancy = used+CO+clean): {avg_u:.1%}  |  "
                       f"High(≥90%): {high}  |  Idle(<5%): {low}  |  "
                       f"Presses: {len(all_presses)}  |  Total CO_Mins: {int(total_co):,}")
                ).font = _bold(10)
    _ml = mould_life or {}
    # Utilization_Pct = production only. Occupancy_Pct = (used + CO + mould-clean)/
    # available — the metric stored in jkt_plan_kpis / jkt_plan_capacityUtilisation.
    u_cols = ["Machine", "Available_Mins", "Used_Mins", "CO_Mins", "Mould_Clean_Mins",
              "Idle_Mins", "Utilization_Pct", "CO_Pct", "Mould_Clean_Utilization_%",
              "Idle_Pct", "Occupancy_Pct", "SKUs_Count", "total_cycle", "Total_Units",
              "Remaining_Mould_Life"]
    _hdr(ws, 2, u_cols)
    for ri, press in enumerate(all_presses, 3):
        s     = press_stats[press]
        used  = s["running_mins"]
        co    = s["co_mins"]
        clean = s["clean_mins"]
        idle  = max(0, avail_mins - used - co - clean)
        pct       = used  / avail_mins if avail_mins else 0.0
        co_pct    = co    / avail_mins if avail_mins else 0.0
        clean_pct = clean / avail_mins if avail_mins else 0.0
        idle_pct  = idle  / avail_mins if avail_mins else 0.0
        color = _GREEN if pct >= 0.90 else (_AMBER if pct >= 0.60 else _RED)
        vals  = [press, avail_mins, round(used), round(co), round(clean), round(idle),
                 pct, co_pct, clean_pct, idle_pct, pct + co_pct + clean_pct,
                 len(s["skus"]), s["cycles"], s["units"],
                 _ml.get(press, MOULD_CLEAN_CYCLES)]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = _fill(color); cell.alignment = _ctr()
            if ci in (7, 8, 9, 10, 11): cell.number_format = "0.0%"
    for ci in range(1, len(u_cols) + 1):
        ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = 18
    ws.freeze_panes = "A3"

    # ── Sheet 3: Shift Schedule ────────────────────────────────────────────────
    ws = wb.create_sheet("Shift Schedule")
    # Qty = tyres cured. CO_Mins / Mould_Clean_Mins = minutes this press-shift spent
    # in changeover / mould clean (covers planned full-shift COs, dynamic mid-shift
    # COs and their overhang) — so every CO is visible here, not only in the
    # Changeover Plan sheet.
    ss_cols = ["Date", "Shift", "Machine", "SKUCode", "StartTime", "EndTime",
               "Qty", "CO_Mins", "Mould_Clean_Mins",
               "CycleTime_min", "GT_Inventory", "Remarks"]
    _hdr(ws, 1, ss_cols)
    s_fill = {"A": _fill(_BLUE), "B": _fill(_LYELL), "C": _fill(_DGREY)}
    # PM/MTC: DISPLAY-ONLY PM maintenance rows (Remarks="PM Schedule"), merged for
    # visibility only — NOT in any KPI / utilization / feasibility computation above.
    _cur_display = sorted(
        list(cure_shift_rows) + list(pm_mtc_rows or []),
        key=lambda r: (str(r.get("StartTime", "")), str(r.get("Machine", ""))),
    )
    for ri, r in enumerate(_cur_display, 2):         # already split (single source above)
        st = r.get("_status", "RUNNING")
        if st == "CHANGEOVER":
            f = _fill(_ORANGE)
        elif st == "MOULD_CLEAN":
            f = _fill(_AMBER)
        elif st == "PM_MTC":
            f = _fill(_LGREY)
        else:
            f = s_fill.get(r.get("Shift", ""), _fill(_WHITE))
        for ci, h in enumerate(ss_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r.get(h, ""))
            cell.fill = f; cell.alignment = _ctr()
            if st in ("CHANGEOVER", "MOULD_CLEAN", "PM_MTC"):
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["D"].width = 32
    ws.column_dimensions["I"].width = 16; ws.freeze_panes = "A2"

    # ── Sheet 3b: Changeover Plan — every curing press CO (Planned + Dynamic) AND
    # every mould clean (each 8h / 480 min). Both consume a shift, so both are shown
    # here with their Mins, not just in the console log.
    ws = wb.create_sheet("Changeover Plan")
    co_cols = ["Date", "Day", "Shift", "Press", "From_SKU", "Target_SKU", "CO_Type", "Mins"]
    _hdr(ws, 1, co_cols)
    _co_events   = list(cure_co_events or [])
    _clean_events = list(mould_clean_events or [])
    _all_events = sorted(
        _co_events + _clean_events,
        key=lambda e: (int(e.get("Day", 0)),
                       {"A": 0, "B": 1, "C": 2}.get(e.get("Shift", "A"), 0),
                       str(e.get("Press", ""))),
    )
    for ri, e in enumerate(_all_events, 2):
        _ct = e.get("CO_Type")
        if _ct == "Mould Clean":
            f = _fill("D9E1F2")                      # blue — mould clean
        elif _ct == "Dynamic":
            f = _fill(_ORANGE)
        else:
            f = _fill(_LYELL)                        # planned CO
        _mins = e.get("Mins", CURING_CO_CHANGEOVER_MINS if _ct != "Mould Clean" else MOULD_CLEAN_MINS)
        for ci, h in enumerate(co_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=(_mins if h == "Mins" else e.get(h, "")))
            cell.fill = f; cell.alignment = _ctr()
            if h == "CO_Type":
                cell.font = Font(bold=True)
    _n_plan  = sum(1 for e in _co_events if e.get("CO_Type") == "Planned")
    _n_dyn   = sum(1 for e in _co_events if e.get("CO_Type") in ("Dynamic", "Early-CO"))
    _n_clean = len(_clean_events)
    _foot = len(_all_events) + 3
    ws.cell(row=_foot, column=1,
            value=f"Total curing COs: {len(_co_events)}  "
                  f"(Planned: {_n_plan}, Dynamic: {_n_dyn})   |   "
                  f"Mould cleans: {_n_clean}  (each {MOULD_CLEAN_MINS} min = 8h)"
            ).font = Font(bold=True)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["E"].width = 32; ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A2"

    # ── Sheet 4: Mould Tracker ────────────────────────────────────────────────
    # End-of-plan mould state per curing press: which 2 moulds are mounted, whether
    # they are eligible for the SKU the press ends on (feasibility proof), and how
    # over-subscribed that SKU's mould pool is (contention — explains unmet demand).
    ws = wb.create_sheet("Mould Tracker")
    if not mould_info:
        _hdr(ws, 1, ["MouldNo", "Compatible_SKUs", "Life_Remaining", "Assigned_Machine"])
        ws.cell(row=2, column=1,
                value="Mould gate OFF — no mould state to report (run with MOULD_GATE=1)"
                ).font = Font(italic=True, color="888888")
        ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 44
    else:
        _pm  = mould_info.get("press_moulds", {})
        _fs  = mould_info.get("final_sku", {})
        _msk = mould_info.get("mould_skus", {})
        _sm  = mould_info.get("sku_moulds", {})
        _pc  = mould_info.get("press_count", {})
        _sc  = mould_info.get("scorer") or {}
        _ev  = mould_info.get("events", []) or []
        _asg = mould_info.get("assignments", []) or []
        # swaps per press (for the header "never swapped" count)
        _swaps: dict = defaultdict(int)
        for _e in _ev:
            _swaps[_e["press"]] += 1
        # summary header — CO provenance + contention + movement totals for the client
        _n_moulds = sum(len(v) for v in _pm.values())
        _n_shared = sum(1 for m, s in _msk.items() if len(s) > 1)
        _n_fixed  = sum(1 for _p in _pm if _swaps.get(_p, 0) == 0)
        ws.cell(row=1, column=1, value=(
            f"Moulds mounted: {_n_moulds}  |  Shared moulds (serve >1 SKU): {_n_shared}  |  "
            f"Total mould swaps this month: {len(_ev)}  |  "
            f"Presses that never swapped: {_n_fixed}/{len(_pm)}  |  "
            f"Day-0 2nd-mould top-ups: {mould_info.get('day0_topups', 0)}  |  "
            f"Mould-blocked COs: {mould_info.get('blocked', 0)}  |  "
            f"Retargeted: {mould_info.get('retargeted', 0)}"
            + (f"  |  Pulled-forward: {_sc.get('pullfwd', 0)}  |  "
               f"Idle-fill/dynamic: {_sc.get('dynamic', 0)}" if _sc else "")
        )).font = _bold(10)
        # EXPANDED: one row per (press, mould) building a SKU — Day-0 opening plus
        # every changeover mount. So a press that runs one SKU all month = 2 rows
        # (its 2 moulds); a press that swaps N times = 2 × (N+1) rows. Total >> 167.
        mt_cols = ["Press", "Mould", "SKU_Built", "Day_Mounted", "Mould_Eligible_For_SKU",
                   "Shared_Mould", "SKU_Eligible_Mould_Pool", "Presses_On_SKU",
                   "SKU_Tooling_Pressure"]
        _hdr(ws, 2, mt_cols)
        _rows = []
        for _a in _asg:
            _p   = _a["press"]; _sku = _a["sku"]; _day = _a["day"]
            _pool = len(_sm.get(_sku, set()))
            _npr  = int(_pc.get(_sku, 0))
            _tp   = round((_npr * 2) / _pool, 2) if _pool else ""
            for _m in _a.get("moulds", []):
                _elig   = "Yes" if _sku in _msk.get(_m, set()) else "No"
                _shared = "Yes" if len(_msk.get(_m, set())) > 1 else "No"
                _rows.append([_p, _m, _sku, _day, _elig, _shared, _pool, _npr, _tp])
        _rows.sort(key=lambda r: (str(r[0]), int(r[3]), str(r[1])))   # press, day, mould
        for _ri, _r in enumerate(_rows, 3):
            for _ci, _v in enumerate(_r, 1):
                _cell = ws.cell(row=_ri, column=_ci, value=_v)
                _cell.alignment = _ctr()
                if _ci == 5 and _v == "No":       # mould not eligible for the SKU — flag red
                    _cell.font = _bold(10, "CC0000")
        ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value + \
            f"  |  Rows (press×mould×SKU-run): {len(_rows)}"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 30
        for _ltr in ("D", "E", "F", "G", "H", "I"):
            ws.column_dimensions[_ltr].width = 16
        ws.freeze_panes = "A3"

        # ── Sheet 4b: Mould Movement (one row per swap — the actual "tracker") ────
        ws2 = wb.create_sheet("Mould Movement")
        ws2.cell(row=1, column=1, value=(
            f"Every mould change during the month ({len(_ev)} swaps). A press keeps its "
            f"moulds until it changes over to an SKU those moulds cannot cure, then swaps."
        )).font = _bold(10)
        mv_cols = ["Day", "Press", "New_SKU", "Moulds_Mounted", "Moulds_Removed"]
        _hdr(ws2, 2, mv_cols)
        for _ri, _e in enumerate(
                sorted(_ev, key=lambda e: (e["day"], str(e["press"]))), 3):
            for _ci, _v in enumerate([
                    _e["day"], _e["press"], _e["sku"],
                    ", ".join(_e.get("added", [])),
                    ", ".join(_e.get("removed", []))], 1):
                ws2.cell(row=_ri, column=_ci, value=_v).alignment = _ctr()
        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 30
        ws2.column_dimensions["D"].width = 40
        ws2.column_dimensions["E"].width = 40
        ws2.freeze_panes = "A3"

    # ── Sheet 4c: MouldInUse ──────────────────────────────────────────────────
    # Fixed daily grid: one row per (calendar day, demand SKU) = PLANNING_DAYS ×
    # #demand-SKUs rows. "Mould in USE" = the SKU's moulds OCCUPIED that day = moulds on
    # presses COMMITTED to the SKU (holding its moulds), counted on every held day incl.
    # dry / GT-starved / idle days; the count drops only when a press CO's AWAY, not when
    # it runs dry. "Total Eligible Moulds" = the SKU's eligible pool size, constant (0 if
    # none). Days with no press holding the SKU → 0. See _mould_in_use_rows.
    ws = wb.create_sheet("MouldInUse")
    miu_cols = ["Date", "SKU Code", "Description",
                "Mould in USE", "Total Eligible Moulds"]
    _sdesc = sku_desc_map or {}
    _miu = _mould_in_use_rows(cure_shift_rows, mould_info, demand_dict,
                              planning_days, plan_start)
    ws.cell(row=1, column=1, value=(
        f"Daily mould occupancy per SKU (grid: {planning_days} days × "
        f"{len(demand_dict)} demand SKUs = {len(_miu)} rows)  |  "
        "Mould in USE = the SKU's moulds OCCUPIED that day = moulds on presses committed to the SKU "
        "(holding its moulds) — counted on every held day incl. dry/GT-starved/idle days; the count "
        "drops only when a press CO's AWAY, not when it runs dry. Total Eligible Moulds = size of the "
        "SKU's eligible mould pool (Master_Mapping_Mould_SKU), constant per SKU."
        if _miu else
        "Mould gate OFF — no mould state to report (run with MOULD_GATE=1)."
    )).font = _bold(10)
    _hdr(ws, 2, miu_cols)
    for _ri, _r in enumerate(_miu, 3):
        _r = {**_r, "Description": _sdesc.get(_r["SKU Code"], "NA")}
        for _ci, _h in enumerate(miu_cols, 1):
            ws.cell(row=_ri, column=_ci, value=_r.get(_h, "")).alignment = _ctr()
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A3"

    # ── Sheet 5: Machine Schedule ─────────────────────────────────────────────
    ws = wb.create_sheet("Machine Schedule")
    ms_rows = []
    for (press, sku), s in sorted(press_sku_stats.items()):
        if s["units"] == 0:
            continue
        ct        = cure_ct_map.get(sku, DEFAULT_CURING_CT)
        days_used = s["mins_used"] / (3 * SHIFT_MINS) if s["mins_used"] else 0
        ms_rows.append({
            "Machine": press, "SKUCode": sku,
            "Priority": round(pri_map.get(sku, 0.0), 4),
            "CycleTime_min": round(ct, 2),
            "Cycles": s["cycles"], "Units_Planned": s["units"],
            "Mins_Used": round(s["mins_used"]), "Days_Used": round(days_used, 2),
        })
    ms_rows.sort(key=lambda r: (r["Machine"], -r["Units_Planned"]))
    tot_u = sum(r["Units_Planned"] for r in ms_rows)
    tot_c = sum(r["Cycles"] for r in ms_rows)
    ws.cell(row=1, column=1,
            value=f"Press-SKU pairs: {len(ms_rows)}  |  Total Units: {tot_u:,}  |  Total Cycles: {tot_c:,}"
            ).font = _bold(10)
    ms_cols = ["Machine", "SKUCode", "Priority", "CycleTime_min",
               "Cycles", "Units_Planned", "Mins_Used", "Days_Used"]
    _hdr(ws, 2, ms_cols)
    for ri, r in enumerate(ms_rows, 3):
        for ci, h in enumerate(ms_cols, 1):
            ws.cell(row=ri, column=ci, value=r.get(h, "")).alignment = _ctr()
    ws.column_dimensions["A"].width = 12; ws.column_dimensions["B"].width = 34
    for ltr in "CDEFGH": ws.column_dimensions[ltr].width = 14
    ws.freeze_panes = "A3"

    # ── Sheet 6: Daily Cured tyres ────────────────────────────────────────────
    # Per-day CURING occupancy (CU) — actual cure time (Qty/2)×CT + CO + mould-clean
    # (idle & PM/MTC maintenance excluded), over the full press roster (Npress×3×480/day).
    # Same rule as the ad-hoc CU analysis. Read from cure_shift_rows.
    _cur_busy: dict[str, float] = defaultdict(float)
    _presses: set = set()
    for _r in cure_shift_rows:
        if str(_r.get("Remarks", "")) in ("PM Schedule", "MTC Schedule"):
            continue                                   # maintenance downtime is NOT busy
        _p = str(_r.get("Machine", ""))
        if _p and _p not in ("—", "nan", ""):
            _presses.add(_p)
        _d = str(_r.get("Date", ""))[:10]
        try: _q = float(_r.get("Qty", 0) or 0)
        except (TypeError, ValueError): _q = 0.0
        try: _ct = float(_r.get("CycleTime_min", 0) or 0)
        except (TypeError, ValueError): _ct = 0.0
        _co = float(_r.get("CO_Mins", 0) or 0); _cl = float(_r.get("Mould_Clean_Mins", 0) or 0)
        _busy = (_q / 2.0) * _ct + _co + _cl
        if _busy > 0:
            _cur_busy[_d] += min(_busy, 480.0)
    _NC = len(_presses) or 1
    _cur_occ = {d: round(100.0 * b / (_NC * 3 * 480), 1) for d, b in _cur_busy.items()}
    ws = wb.create_sheet("Daily Cured tyres")
    _hdr(ws, 1, ["Date", "Cured_Qty", "Curing_Occupancy_%"])
    total_c = 0
    for d in range(planning_days):
        date_str = (plan_start + timedelta(days=d)).strftime("%Y-%m-%d")
        qty      = int(daily_cured.get(date_str, 0))
        ws.cell(row=d + 2, column=1, value=date_str).alignment = _ctr()
        c = ws.cell(row=d + 2, column=2, value=qty)
        c.alignment = _ctr()
        c.fill = _fill(_BLUE) if qty > 0 else _fill(_RED)
        ws.cell(row=d + 2, column=3, value=_cur_occ.get(date_str, 0.0)).alignment = _ctr()
        total_c += qty
    tr = planning_days + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = _bold(11)
    t = ws.cell(row=tr, column=2, value=total_c)
    t.font = _bold(11); t.fill = _fill(_GREEN)
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    # ── Sheet 7: GT Gap Diagnostic ────────────────────────────────────────────
    ws = wb.create_sheet("GT Gap Diagnostic")
    _hdr(ws, 1, ["SKUCode", "GT_Built", "GT_Cured", "Closing_Balance", "Reason"])
    built_per_sku: dict[str, float] = defaultdict(float)
    for sku_qty in build_by_shift_sku.values():
        for sku, qty in sku_qty.items():
            built_per_sku[str(sku)] += float(qty)
    for sku, qty in opening_gt.items():
        built_per_sku[str(sku)] += float(qty)
    press_skus = set(sku_cured.keys())
    ri = 2
    for sku in sorted(closing_gt_bal, key=lambda s: -closing_gt_bal[s]):
        bal = closing_gt_bal[sku]
        if bal < 0.5:
            continue
        built  = built_per_sku.get(sku, 0.0)
        cured  = float(sku_cured.get(sku, 0))
        if sku not in press_skus:
            reason = "NO_PRESS"; fill = _fill(_RED)
        elif cured > 0:
            reason = "DEMAND_MET"; fill = _fill(_AMBER)
        else:
            reason = "RESIDUAL"; fill = _fill(_LGREY)
        for ci, val in enumerate([sku, round(built), round(cured), round(bal), reason], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if ci == 4: c.fill = fill
            c.alignment = _ctr()
        ri += 1
    # Summary + note (matches curing_b2c.py format)
    ws.cell(row=ri + 1, column=1, value="TOTAL").font = _bold(11)
    total_built_diag  = sum(round(built_per_sku.get(s, 0)) for s in closing_gt_bal if closing_gt_bal[s] >= 0.5)
    total_cured_diag  = sum(round(float(sku_cured.get(s, 0))) for s in closing_gt_bal if closing_gt_bal[s] >= 0.5)
    total_bal_diag    = sum(round(closing_gt_bal[s]) for s in closing_gt_bal if closing_gt_bal[s] >= 0.5)
    ws.cell(row=ri + 1, column=2, value=total_built_diag).font = _bold(11)
    ws.cell(row=ri + 1, column=3, value=total_cured_diag).font = _bold(11)
    tc = ws.cell(row=ri + 1, column=4, value=total_bal_diag)
    tc.font = _bold(11); tc.fill = _fill(_RED)
    note = ws.cell(row=ri + 3, column=1,
        value=("NO_PRESS = built but no curing press (main gap cause)  |  "
               "DEMAND_MET = carry-over to next month  |  "
               "RESIDUAL = last-shift lag (next month opening inventory)"))
    note.font = _bold(9)
    ws.column_dimensions["A"].width = 34
    for ltr in "BCDE": ws.column_dimensions[ltr].width = 16

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"  [Rolling] Curing output  → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# ROLLING PIPELINE — main function
# ══════════════════════════════════════════════════════════════════════════════

def _build_sku_desc_map(demand_path: str) -> dict:
    """SKU code → description from the demand file, consolidated across market rows.
    Rule (client): among a SKU's rows take the description from the largest-Requirement
    row that HAS a non-empty description; if none has one, "NA". Cosmetic only (does not
    affect the plan) — any read failure degrades to {} (all-"NA" downstream)."""
    try:
        df = (pd.read_csv(demand_path) if str(demand_path).lower().endswith(".csv")
              else pd.read_excel(demand_path))
    except Exception:
        return {}
    low = {str(c).strip().lower(): c for c in df.columns}
    def _find(cands):
        for c in cands:
            if c in low:
                return low[c]
        return None
    sku_c = _find(["skucode", "sku_code", "sapcode", "sku"])
    qty_c = _find(["requirement", "updated_requirement", "quantity", "qty"])
    dsc_c = _find(["sku description", "skudescription", "sku_description", "description"])
    if sku_c is None or dsc_c is None:
        return {}
    keep = [sku_c, dsc_c] + ([qty_c] if qty_c else [])
    tmp = df[keep].copy()
    tmp[sku_c] = tmp[sku_c].astype(str).str.strip()
    out: dict = {}
    for sku, g in tmp.groupby(sku_c):
        if not sku or sku.lower() == "nan":
            continue
        d = g[dsc_c].astype(str).str.strip()
        valid = g[d.notna() & (d != "") & (d.str.lower() != "nan")]
        if len(valid):
            if qty_c:
                valid = valid.sort_values(qty_c, ascending=False)
            out[sku] = str(valid.iloc[0][dsc_c]).strip()
        else:
            out[sku] = "NA"
    return out


def _build_priority_deadline_map(demand_path: str, plan_start, planning_days: int):
    """Parse the optional committed-delivery columns from the demand file.

    Columns (header-normalized, so "Delivery Date"/"Delivery date"/"delivery_date"
    all match; values are strings): "Priority Flag" ("0"/"1"/"Yes") and
    "Delivery Date" ("DD/MM/YY"). A SKU is DELIVERY-COMMITTED when its flag is set
    (``str.lower() in {"1","yes"}``) OR it carries a valid delivery date — a date
    implies commitment even if the flag reads "No"/"0"/blank (client rule).

    Returns ``(priority_deadline_map, priority_meta)`` where
      • ``priority_deadline_map`` = ``{sku: deadline_day_index}`` (1-based within the
        horizon) for committed SKUs only — the single structure threaded through the
        pipeline; and
      • ``priority_meta`` = ``{sku: {demand, deadline_date, deadline_day, undated,
        past_start, beyond_month}}`` for the feasibility report.

    A committed SKU with no date maps to ``planning_days`` (end of month) when
    ``PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED`` is on. A date before the plan start
    clamps to day 1 (``past_start``); a date beyond month-end clamps to
    ``planning_days`` (``beyond_month``). Any read/parse failure → ``({}, {})`` so the
    feature is simply inert (never breaks a run). Non-committed SKUs are absent from
    both maps → treated as normal everywhere (identity)."""
    empty: tuple[dict, dict] = ({}, {})
    # feature active iff EITHER merged mode toggle is on (DELIVERY_PRIORITY_ENABLED merged away)
    if not (bool(getattr(_bc_cfg, "DELIVERY_DATE_ALL_SOFT_RULES_RELAXED", False))
            or bool(getattr(_bc_cfg, "PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED", False))):
        return empty
    try:
        df = (pd.read_csv(demand_path) if str(demand_path).lower().endswith(".csv")
              else pd.read_excel(demand_path))
    except Exception:
        return empty
    low = {str(c).strip().lower().replace("_", " "): c for c in df.columns}
    def _find(cands):
        for c in cands:
            if c in low:
                return low[c]
        return None
    sku_c  = _find(["skucode", "sku code", "sapcode", "sku"])
    flag_c = _find(["priority flag", "priorityflag", "priority"])
    date_c = _find(["delivery date", "deliverydate"])
    qty_c  = _find(["requirement", "updated requirement", "quantity", "qty", "demand"])
    if sku_c is None or (flag_c is None and date_c is None):
        return empty  # no committed-delivery columns at all → inert

    def _flag_set(v) -> bool:
        if pd.isna(v):
            return False
        return str(v).strip().lower() in {"1", "1.0", "yes", "y", "true"}

    def _parse_date(v):
        if v is None or (not isinstance(v, str) and pd.isna(v)):
            return None
        if isinstance(v, (pd.Timestamp, datetime)):
            return pd.Timestamp(v)
        s = str(v).strip()
        if not s or s.lower() == "nan":
            return None
        d = pd.to_datetime(s, format="%d/%m/%y", errors="coerce")
        if pd.isna(d):                       # tolerate a stray full-year / other order
            d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return None if pd.isna(d) else pd.Timestamp(d)

    undated_to_monthend = bool(
        getattr(_bc_cfg, "PRIORITY_FLAG_MONTHEND_ALL_SOFT_RULES_RELAXED", False))
    start_date = plan_start.date() if hasattr(plan_start, "date") else plan_start

    df = df.copy()
    df["_sku_norm"] = df[sku_c].astype(str).str.strip()
    pmap: dict[str, int] = {}
    meta: dict[str, dict] = {}
    for sku, g in df.groupby("_sku_norm"):
        if not sku or sku.lower() == "nan":
            continue
        flag_any = bool(g[flag_c].map(_flag_set).any()) if flag_c else False
        dates = [d for d in (g[date_c].map(_parse_date) if date_c else []) if d is not None and not pd.isna(d)]
        has_date = len(dates) > 0
        if not (flag_any or has_date):
            continue  # normal SKU
        demand = 0.0
        if qty_c:
            demand = float(pd.to_numeric(g[qty_c], errors="coerce").fillna(0).sum())
        undated = not has_date
        past_start = beyond_month = False
        if undated:
            if not undated_to_monthend:
                continue  # flagged-but-undated treated as normal when toggle off
            day_idx = int(planning_days)
            deadline_date = None
        else:
            deadline = min(dates)                     # earliest date = EDF-conservative
            deadline_date = deadline
            raw = (deadline.date() - start_date).days + 1
            past_start   = raw < 1
            beyond_month = raw > int(planning_days)
            day_idx = max(1, min(int(planning_days), raw))
        pmap[sku] = day_idx
        meta[sku] = {
            "demand":        demand,
            "deadline_date": (deadline_date.strftime("%Y-%m-%d") if deadline_date is not None else None),
            "deadline_day":  day_idx,
            "undated":       undated,
            "past_start":    past_start,
            "beyond_month":  beyond_month,
        }
    return pmap, meta


def _priority_feasibility_precheck(pmap, meta, demand_dict, sku_presses, sku_moulds,
                                   sku_bld_machines, cure_ct_map, plan_start, planning_days):
    """Static, best-effort feasibility ceiling for each committed-delivery SKU.

    For each priority SKU, compute the MOST it could possibly be cured by its
    deadline given ONLY its DB-allowable presses + DB-eligible moulds (2 per press,
    contention) and its inch-eligible building machines — never inventing a pair.
    This is an upper bound (ignores contention with other SKUs), so it is a
    conservative warning, never a hard-stop. Also derives the earliest date the SKU
    could be fully completed at full dedication from day 1 (the 'relax-report' half
    of the infeasibility decision). Returns a list of report rows, EDF-ordered."""
    rows = []
    for sku in sorted(pmap, key=lambda s: (meta[s]["deadline_day"], s)):
        dd     = pmap[sku]
        shifts = dd * 3                                       # A/B/C shifts per day
        demand = float(demand_dict.get(sku, meta[sku]["demand"]))
        n_press = len(sku_presses.get(sku, ()) or ())
        n_pairs = len(sku_moulds.get(sku, ()) or ()) // CURING_CAVITIES
        simul   = min(n_press, n_pairs)                       # presses that can run concurrently
        ct        = float(cure_ct_map.get(sku, DEFAULT_CURING_CT))
        cure_rate = _cure_qty_per_shift(ct)                   # per press per shift
        cure_ceiling = simul * cure_rate * shifts
        bld_rate = sum(_bld_qty_per_shift(m, sku)             # GT/shift over eligible machines
                       for m in (sku_bld_machines.get(sku, ()) or ()))
        bld_ceiling = bld_rate * shifts
        feasible  = min(demand, cure_ceiling, bld_ceiling)
        shortfall = max(0.0, demand - feasible)
        per_day = min(simul * cure_rate * 3, bld_rate * 3)
        if per_day > 0:
            e_days   = int(math.ceil(demand / per_day))
            earliest = (plan_start + timedelta(days=e_days - 1)).date().isoformat()
        else:
            e_days, earliest = None, None
        rows.append({
            "sku": sku, "demand": demand, "deadline_day": dd,
            "deadline_date": meta[sku]["deadline_date"], "undated": meta[sku]["undated"],
            "past_start": meta[sku]["past_start"], "beyond_month": meta[sku]["beyond_month"],
            "n_presses": n_press, "n_mould_pairs": n_pairs, "simul_presses": simul,
            "cure_ceiling": cure_ceiling, "bld_ceiling": bld_ceiling,
            "feasible_by_deadline": feasible, "static_shortfall": shortfall,
            "earliest_feasible_days": e_days, "earliest_feasible_date": earliest,
            "on_time_possible": (e_days is not None and e_days <= dd),
            "structurally_infeasible": (simul == 0 or bld_rate == 0),
        })
    return rows


def _inject_label_columns(xlsx_path: str, sku_desc: dict,
                          machine_names: dict | None = None) -> None:
    """Post-process a finished workbook: after every SKU-code column insert a
    "<col> Description" column (value from sku_desc, "NA" if unknown), and — only
    when machine_names is given (the building workbook) — after every building
    Machine column insert a "<col> Name" column. Header-driven so it covers every
    sheet and tolerates the sheets whose header is not on row 1. Cosmetic-only:
    on any error the workbook is left exactly as written. The inserted column copies
    the SOURCE column's full styling (header fill/font, borders, alignment, number
    format, column width) so the description reads like the SKU-code column and the
    name reads like the machine-code column."""
    import openpyxl, copy as _copy
    from openpyxl.utils import get_column_letter
    _SKU_HDRS  = {"skucode", "from_sku", "target_sku", "new_sku", "sku_built"}
    _MACH_HDRS = {"machine"}
    _SENT = {"CHANGEOVER", "MOULD_CLEAN", "MOULD CLEAN", "C/O", "CO"}
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception:
        return
    for ws in wb.worksheets:
        # header row = first of rows 1..4 that holds any known code header
        hdr = None
        for r in range(1, 5):
            vals = {str(ws.cell(r, c).value or "").strip().lower()
                    for c in range(1, ws.max_column + 1)}
            if vals & (_SKU_HDRS | _MACH_HDRS):
                hdr = r
                break
        if hdr is None:
            continue
        targets = []                                   # (col_idx, kind, orig_header)
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(hdr, c).value or "").strip().lower()
            if h in _SKU_HDRS:
                targets.append((c, "desc", ws.cell(hdr, c).value))
            elif machine_names is not None and h in _MACH_HDRS:
                targets.append((c, "name", ws.cell(hdr, c).value))
        # snapshot each column's width by header text — insert_cols does NOT shift
        # column widths, so we re-apply them by name after all inserts are done.
        orig_w = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(hdr, c).value or "")
            wdt = ws.column_dimensions[get_column_letter(c)].width
            if h and wdt:
                orig_w[h] = wdt
        # insert right-to-left so already-collected column indices stay valid
        for c, kind, orig in sorted(targets, key=lambda t: -t[0]):
            ws.insert_cols(c + 1)
            new_hdr = f"{orig} Description" if kind == "desc" else f"{orig} Name"
            # clone the source column's styling row-by-row, then write values
            for r in range(1, ws.max_row + 1):
                src, dst = ws.cell(r, c), ws.cell(r, c + 1)
                dst.font          = _copy.copy(src.font)
                dst.fill          = _copy.copy(src.fill)
                dst.border        = _copy.copy(src.border)
                dst.alignment     = _copy.copy(src.alignment)
                dst.number_format = src.number_format
                if r == hdr:
                    dst.value = new_hdr
                elif r > hdr:
                    code = src.value
                    if code is None or str(code).strip() == "":
                        continue
                    key = str(code).strip()
                    if kind == "desc":
                        dst.value = key if key.upper() in _SENT else sku_desc.get(key, "NA")
                    else:
                        dst.value = machine_names.get(key, "NA")
        # re-apply widths by header — each label column inherits its source's width
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(hdr, c).value or "")
            src_h = (h[:-len(" Description")] if h.endswith(" Description")
                     else h[:-len(" Name")] if h.endswith(" Name") else h)
            if src_h in orig_w:
                ws.column_dimensions[get_column_letter(c)].width = orig_w[src_h]
        # extend multi-column title merges (rows above the header) to the full new
        # width so the title bar stays solid across the inserted columns
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row < hdr and mr.max_col > mr.min_col and mr.max_col < ws.max_column:
                r0, c0, r1 = mr.min_row, mr.min_col, mr.max_row
                try:
                    ws.unmerge_cells(start_row=r0, start_column=c0,
                                     end_row=r1, end_column=mr.max_col)
                    ws.merge_cells(start_row=r0, start_column=c0,
                                   end_row=r1, end_column=ws.max_column)
                except Exception:
                    pass
    try:
        wb.save(xlsx_path)
    except Exception:
        pass


# ── Mid-month plan start: deduct already-completed production from demand ──────────────
# Feature: a plan may start on any date. For a mid-month start we first run a full-month plan
# (Run 1) to SIMULATE the production the plant already made on days 1..(start-1), deduct that
# (per SKU) from the original demand, then run the ACTUAL plan (Run 2) for start..month-end on
# the reduced demand. day==1 or toggle OFF → single run, bit-for-bit unchanged. Local-only for
# now (wired in local_main.py); the cloud path (main.run_plan) will call this in a later step.
_MIDMONTH_DEDUCT = os.environ.get("MIDMONTH_DEDUCT", "0") != "0"   # 2-pass SIMULATION path (legacy); OFF
# (inert for a 1st-of-month PLAN_START → single run, bit-for-bit; only local_main uses the 2pass wrapper)
_MIDMONTH_SIM_LO = float(os.environ.get("MIDMONTH_SIM_LO", "0.90"))
_MIDMONTH_SIM_HI = float(os.environ.get("MIDMONTH_SIM_HI", "1.05"))
# ADOPTED going forward — SINGLE-RUN with the plant's ACTUAL SAP production (ACTUAL_PROD, default ON).
# For a plan starting after the 1st: deduct the real SAP-reported CURED production (Plant 1300,
# Mtart=ZFGS) for days 1..(start-1) from demand — NO 90-105% simulation and NO Run 1 — then run the
# plan ONCE from the start date, seeded from THAT date's real running-moulds/GT/carcass (date-filtered
# ETL, `date`=PLAN_DATE). Supersedes the 2-pass simulation above; ACTUAL_PROD=0 falls back to it.
# NOTE: fetches SAP live at plan time (needs corporate VPN). See api/sap_production_data.py.
_ACTUAL_PROD_DEDUCT = os.environ.get("ACTUAL_PROD", "0") != "0"   # SAP live fetch + deduction; default OFF

# Option B — GENERATE FROM TODAY (TODAY_START, default ON). The nominal plan window (plan_start..end
# from bc_config / jkt_plan_params) stays the FULL month; the engine sets the EFFECTIVE start =
# datetime.now() (clamped into [nominal_start, month_end]) and shrinks planning_days to the same
# month-end. So a plan CREATED on the 21st runs 21..month-end, seeded from the 21st's real state and
# with days 1..20 deducted from demand (ACTUAL_PROD). TODAY_START=0 keeps the nominal window verbatim
# (used by month-specific tests / historical re-runs — the harnesses set it 0).
_TODAY_START = os.environ.get("TODAY_START", "0") != "0"   # generate-from-today clamp; default OFF


def _sync_plan_keys(plan_start) -> None:
    """Re-point the date/month ETL filter keys at `plan_start` across bc_config + every already-imported
    engine module (they import PLAN_DATE / PLAN_MONTH / RUNNING_MOULDS_MONTH BY VALUE). Called by the
    today-start override so the date-filtered ETL reads the EFFECTIVE start date's running-moulds /
    GT / carcass. Mirrors main._set_plan_month for the local path."""
    pm = plan_start.strftime("%Y-%m")
    pd_ = plan_start.strftime("%Y-%m-%d")
    os.environ["PLAN_MONTH"] = pm
    os.environ["RUNNING_MOULDS_MONTH"] = pm
    os.environ["PLAN_DATE"] = pd_
    try:
        import bc_config as _bcc
        for _a, _v in (("PLAN_MONTH", pm), ("RUNNING_MOULDS_MONTH", pm),
                       ("PLAN_DATE", pd_), ("PLAN_START", plan_start)):
            setattr(_bcc, _a, _v)
    except Exception:
        pass
    for _mod in ("connection", "curing_consumption_dynamic", "curing_b2c",
                 "building", "building_b2c"):
        try:
            _m = __import__(_mod)
        except Exception:
            continue
        for _a, _v in (("PLAN_MONTH", pm), ("RUNNING_MOULDS_MONTH", pm), ("PLAN_DATE", pd_)):
            if hasattr(_m, _a):
                setattr(_m, _a, _v)


def _midmonth_sim_factor(sku: str) -> float:
    """Deterministic per-SKU production-simulation factor in [LO, HI]. Uses hashlib (NOT random()
    / builtin hash()) so it is identical across processes and PYTHONHASHSEED values — matching the
    pipeline's reproducibility guarantee."""
    _h = hashlib.md5(str(sku).encode("utf-8")).hexdigest()
    _frac = int(_h[:8], 16) / 0xFFFFFFFF                      # deterministic [0,1]
    return _MIDMONTH_SIM_LO + _frac * (_MIDMONTH_SIM_HI - _MIDMONTH_SIM_LO)


def _extract_cured_through_day(curing_xlsx: str, run_start: datetime, k_days: int) -> dict:
    """Per-SKU CURED qty over days 1..k_days of a completed run, read from its curing Shift
    Schedule sheet (production rows = Qty>0; CO/clean rows are Qty 0)."""
    cutoff = (run_start.date() + timedelta(days=k_days - 1))   # last produced day (inclusive)
    df = pd.read_excel(curing_xlsx, sheet_name="Shift Schedule")
    df["_d"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    df["_q"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0.0)
    m = (df["_d"].notna()) & (df["_d"] <= cutoff) & (df["_q"] > 0)
    return df.loc[m].groupby("SKUCode")["_q"].sum().to_dict()


def _write_deducted_demand(demand_path: str, produced: dict, out_path: str,
                           apply_factor: bool = True) -> dict:
    """Write a demand workbook = original with each SKU's qty reduced by production, floored at 0.
    apply_factor=True → production = planned_cured × per-SKU [0.90,1.05] factor (the SIMULATION path);
    apply_factor=False → production = `produced` as-is (the ACTUAL SAP-production path). Scales every
    row of a SKU by the SKU's updated/original ratio so multi-row files and all other columns
    (Priority Flag / Delivery Date) are preserved. Returns the per-SKU deducted-production dict."""
    df = pd.read_excel(demand_path)
    sku_col = next((c for c in ("SKUCode", "skuCode", "sku_code", "Sapcode") if c in df.columns),
                   df.columns[0])
    qty_col = next((c for c in ("Quantity", "Updated_Requirement", "Requirement") if c in df.columns),
                   None)
    if qty_col is None:
        raise ValueError(f"[midmonth] no qty column in {demand_path} (cols={list(df.columns)})")
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0.0)
    orig_tot = df.groupby(sku_col)[qty_col].sum().to_dict()
    simulated = ({s: float(q) * _midmonth_sim_factor(s) for s, q in produced.items()}
                 if apply_factor else {str(s): float(q) for s, q in produced.items()})
    scale = {}
    for s, o in orig_tot.items():
        upd = max(0.0, o - simulated.get(s, 0.0))
        scale[s] = (upd / o) if o > 0 else 0.0
    df[qty_col] = df.apply(lambda r: r[qty_col] * scale.get(r[sku_col], 1.0), axis=1)
    df.to_excel(out_path, index=False)
    return simulated


def run_rolling_pipeline_2pass(
    demand_path:    str | None = None,
    plan_start:     datetime | None = None,
    planning_days:  int | None = None,
    build_output:   str | None = None,
    curing_output:  str | None = None,
    sku_desc_map:   dict | None = None,
) -> dict:
    """Mid-month wrapper around run_rolling_pipeline (see block comment above). When
    _MIDMONTH_DEDUCT is ON and plan_start is not the 1st, runs Run 1 (full month, original demand)
    to simulate days 1..(start-1) production, deducts it, then runs + returns Run 2 (start..end,
    reduced demand). Otherwise a single normal run (bit-for-bit)."""
    demand_path   = demand_path   or DEMAND_FILE
    plan_start    = plan_start    or PLAN_START
    planning_days = planning_days or PLANNING_DAYS
    if _TODAY_START:
        # Option B: nominal window is the full month; the EFFECTIVE start is today (datetime.now()),
        # clamped into [nominal_start, month_end]. Re-point the ETL date keys at the effective date so
        # the run reads TODAY's running-moulds / GT / carcass and plans today..month-end.
        _month_end = plan_start + timedelta(days=planning_days - 1)
        _now = datetime.now()
        _eff = min(max(plan_start.date(), _now.date()), _month_end.date())
        if _eff != plan_start.date():
            planning_days = (_month_end.date() - _eff).days + 1
            plan_start = datetime(_eff.year, _eff.month, _eff.day, 7, 0, 0)
            _sync_plan_keys(plan_start)
            print(f"[today-start] nominal {_month_end.replace(day=1).date()}..{_month_end.date()} → "
                  f"EFFECTIVE start {plan_start.date()} (+{planning_days}d), state date={plan_start.date()}")
    if plan_start.day == 1 or (not _MIDMONTH_DEDUCT and not _ACTUAL_PROD_DEDUCT):
        return run_rolling_pipeline(demand_path, plan_start, planning_days,
                                    build_output, curing_output, sku_desc_map)
    if _ACTUAL_PROD_DEDUCT:
        # ── SINGLE RUN with ACTUAL SAP production (adopted path) ─────────────────────────────
        # Deduct real cured production for days 1..(start-1); run ONCE from the start date, whose
        # running-moulds/GT/carcass the date-filtered ETL reads (date=PLAN_DATE). No Run 1, no factor.
        from api.sap_production_data import production_by_sku
        _m1 = plan_start.replace(day=1).strftime("%Y-%m-%d")
        _prev = (plan_start - timedelta(days=1)).strftime("%Y-%m-%d")
        produced = production_by_sku(_m1, _prev)                  # {sku: actual cured, days 1..start-1}
        _tmpd = tempfile.mkdtemp(prefix="actualprod_")
        _upd = os.path.join(_tmpd, "updated_demand.xlsx")
        deducted = _write_deducted_demand(demand_path, produced, _upd, apply_factor=False)
        print(f"[actual-prod] SAP actual cured days 1..{plan_start.day - 1} = "
              f"{sum(deducted.values()):,.0f} over {len(deducted)} SKUs deducted → single run "
              f"{plan_start.date()} +{planning_days}d (state date={plan_start.date()})")
        res = run_rolling_pipeline(_upd, plan_start, planning_days, build_output, curing_output,
                                   sku_desc_map)
        res["actual_prod"] = {"deducted_total": sum(deducted.values()), "sku_count": len(deducted)}
        return res
    k_days = plan_start.day - 1                                   # already-produced days 1..k
    run1_start = plan_start.replace(day=1)
    run1_days = k_days + planning_days                            # 1st .. same end date
    _tmp = tempfile.mkdtemp(prefix="midmonth_")
    r1_bld = os.path.join(_tmp, "run1_bld.xlsx")
    r1_cur = os.path.join(_tmp, "run1_cur.xlsx")
    print(f"[midmonth] mid-month start day={plan_start.day}: Run 1 full month "
          f"({run1_start.date()} +{run1_days}d), snapshot state at day {plan_start.day} (07:00)")
    res1 = run_rolling_pipeline(demand_path, run1_start, run1_days, r1_bld, r1_cur,
                                sku_desc_map, snapshot_at_day=plan_start.day)
    snap = res1.get("state_snapshot")
    if snap is None:
        raise RuntimeError("[midmonth] Run 1 returned no state_snapshot (snapshot_at_day missed)")
    # Cured on days 1..k (finished tyres already produced) = original demand − demand still
    # remaining at the snapshot. This is what gets simulated (×factor) and deducted.
    _dd, _drem = snap["demand_dict"], snap["demand_remaining"]
    produced = {s: max(0.0, _dd.get(s, 0.0) - _drem.get(s, 0.0)) for s in _dd}
    produced = {s: q for s, q in produced.items() if q > 0}
    upd_path = os.path.join(_tmp, "updated_demand.xlsx")
    simulated = _write_deducted_demand(demand_path, produced, upd_path)
    print(f"[midmonth] cured 1..{k_days} = {sum(produced.values()):,.0f}; simulated (×0.90-1.05) "
          f"= {sum(simulated.values()):,.0f} tyres over {len(simulated)} SKUs deducted → "
          f"Run 2 ({plan_start.date()} +{planning_days}d) seeded from day-{plan_start.day} state")
    res = run_rolling_pipeline(upd_path, plan_start, planning_days,
                              build_output, curing_output, sku_desc_map,
                              initial_state=snap)
    res["midmonth"] = {"k_days": k_days, "run1_start": run1_start, "run1_days": run1_days,
                       "sku_count": len(simulated),
                       "cured_prior": sum(produced.values()),
                       "simulated_total": sum(simulated.values())}
    # Export the day-start opening (aged GT lots + carcass totals) so the feasibility auditor can
    # seed from the ACTUAL Run-2 opening (--midmonth-opening) instead of the 1st-of-month DB.
    # age_days = snap_day − build_day (GT lot age at the start date); auditor seeds each as (−age).
    try:
        import json as _json
        _snap_day = int(snap["_snap_day"])
        _mm_open = {"snap_day": _snap_day, "plan_start": str(plan_start.date()),
                    "gt_lots": {}, "carcass": {}}
        for _s, _lots in snap["gt_lots"].items():
            _ls = [[max(0, _snap_day - int(_bd)), float(_q)] for _bd, _q in _lots if _q > 0]
            if _ls:
                _mm_open["gt_lots"][str(_s)] = _ls
        for _s, _b in snap["carcass_bank"].items():
            _tot = float(sum(_q for _a, _q in _b))
            if _tot > 0:
                _mm_open["carcass"][str(_s)] = _tot
        _open_json = os.path.splitext(build_output or BUILD_OUTPUT)[0] + "_midmonth_opening.json"
        with open(_open_json, "w") as _f:
            _json.dump(_mm_open, _f)
        res["midmonth_opening_file"] = _open_json
        print(f"[midmonth] day-{plan_start.day} opening exported for audit → {_open_json}")
    except Exception as _e:
        print(f"[midmonth] opening export failed ({_e})")
    return res


def run_rolling_pipeline(
    demand_path:    str | None = None,
    plan_start:     datetime | None = None,
    planning_days:  int | None = None,
    build_output:   str | None = None,
    curing_output:  str | None = None,
    sku_desc_map:   dict | None = None,
    snapshot_at_day: int | None = None,
    initial_state:   dict | None = None,
) -> dict:
    """
    Rolling day-by-day B2C pipeline.

    Press roster = the UNION of the Day-0 running-moulds snapshot and the allowable
    matrix: every snapshot press is used (running ⊆ allowable), and allowable presses
    absent from the snapshot are cold-started as production (IDLE_PRESS_ACTIVATE). The
    old "drop a running press not in the allowable matrix" restriction was removed.

    Generates building and curing schedules simultaneously:
      - Building machines are assigned based on actual GT deficit each day
      - Curing presses cure only what GT is available (GT-limited)
      - Both schedules written to the same Excel format as the legacy pipeline
    """
    demand_path   = demand_path   or DEMAND_FILE
    plan_start    = plan_start    or PLAN_START
    planning_days = planning_days or PLANNING_DAYS
    build_output  = build_output  or BUILD_OUTPUT
    curing_output = curing_output or CURING_OUTPUT
    # #5 robustness: THIS run's plan_start is the single source of truth for the holiday
    # calendar. The CO scheduler (curing_consumption_dynamic) and the urgency helpers
    # (_bc_holiday_day_set) each re-derive holiday day-indices from their OWN imported-by-value
    # PLAN_START global — the same "imported BY VALUE" hazard CLAUDE.md documents for
    # RUNNING_MOULDS_MONTH. Align both globals to plan_start here so all three derivations agree;
    # warn if they had diverged (they never should on the local/cloud paths).
    try:
        import curing_consumption_dynamic as _ccd_mod
        _prev_ps = getattr(_bc_cfg, "PLAN_START", None)
        if _prev_ps is not None and _prev_ps != plan_start:
            print(f"  [Rolling] #5 WARNING: bc_config.PLAN_START ({_prev_ps}) != run plan_start "
                  f"({plan_start}); aligning holiday calendar to plan_start.")
        _bc_cfg.PLAN_START = plan_start
        _ccd_mod.PLAN_START = plan_start
    except Exception as _e:
        print(f"  [Rolling] #5 PLAN_START sync skipped ({_e})")
    # SKU code → description for the output sheets. Cloud passes it (from the DB
    # master); local builds it from the demand file's "SKU Description" column,
    # consolidated across market rows. Missing → "NA" downstream. Purely cosmetic
    # (never touches the plan); LABELS=0 skips it to reproduce label-free sheets.
    _labels_on = os.environ.get("LABELS", "1") != "0"
    if _labels_on and sku_desc_map is None:
        sku_desc_map = _build_sku_desc_map(demand_path)

    # Delivery-date / priority-flag committed-delivery SKUs (DELIVERY_PRIORITY).
    # Built ONCE here so both the Phase-0 curing-CO scheduler and building assignment
    # share the same {sku: deadline_day} map. Empty (June, cloud, toggle off) →
    # _prio_active is False → every priority insertion collapses to identity.
    priority_deadline_map, priority_meta = _build_priority_deadline_map(
        demand_path, plan_start, planning_days)
    _prio_active = _DELIVERY_PRIORITY_ENABLED and bool(priority_deadline_map)
    # SKUs committed WITH a real Delivery Date (vs flag-only month-end) — drives the split
    # DATE-mode vs MONTHEND-mode relaxation gating in _assign_building_shift._delivery_relax.
    _prio_dated = {str(_s) for _s, _m in priority_meta.items() if not _m["undated"]}
    _PRIO_DATED_SKUS.clear(); _PRIO_DATED_SKUS.update(_prio_dated)   # seen by _assign_building_shift
    _BLD_CARRY_UNITS.clear()   # continuous-build carry: fresh per run
    if _prio_active:
        _n_dated = len(_prio_dated)
        _relax_mode = ("DATE at-risk" if _DELIVERY_DATE_RELAX
                       else "MONTHEND at-risk" if _PRIORITY_FLAG_MONTHEND_RELAX else "ordering-only")
        print(f"  [Rolling] DELIVERY_PRIORITY ON — {len(priority_deadline_map)} committed SKUs "
              f"({_n_dated} dated, {len(priority_deadline_map) - _n_dated} end-of-month); EDF order; "
              f"relax={_relax_mode}")

    print("\n" + "=" * 70)
    print("  ROLLING PIPELINE — Pre-computation")
    print("=" * 70)

    # ── A: CO schedule ────────────────────────────────────────────────────────
    print("  [Rolling] Computing CO schedule …")
    # Surplus-release 5b guard needs a per-SKU building-supply estimate (curing
    # scheduler is otherwise building-independent). Compute it only when the
    # feature is on; failure falls back to None (guard becomes a no-op).
    _buildable_rate = None
    if _SURPLUS_RELEASE_ENABLED:
        try:
            from bc_config import make_engine as _mk
            _buildable_rate = _compute_buildable_rate(_mk(), demand_path)
            print(f"  [Rolling] Surplus-release ON — buildable_rate for "
                  f"{len(_buildable_rate)} SKUs (5b guard)")
        except Exception as _e:
            print(f"  [Rolling] buildable_rate computation failed ({_e}); 5b guard disabled")
    # Part 1: compute sku_inch EARLY (only when curing-align is on) so the Phase-0 CO scheduler
    # can prefer same-inch targets. Gated → OFF path is zero-cost and bit-for-bit.
    _early_sku_inch = None
    if _CURING_INCH_ALIGN or _GROUP_INCH_POLICY:
        try:
            from bc_config import make_engine as _mk_si
            from connection import B2C_ETL as _BETL_si
            _etl_si = _BETL_si(_mk_si())
            _early_sku_inch = {str(k): str(v).strip().replace('"', "")
                               for k, v in _etl_si.load_sku_sizes().items()}
            for _, _row in _etl_si.load_machine_allowable().iterrows():
                _s = str(_row["SKUCode"])
                if (_s not in _early_sku_inch or not _early_sku_inch[_s]) and len(_s) >= 10:
                    _early_sku_inch[_s] = _s[8:10]
            print(f"  [Rolling] early sku_inch for {len(_early_sku_inch)} SKUs")
        except Exception as _e:
            print(f"  [Rolling] early sku_inch failed ({_e})")
            _early_sku_inch = None
    # Co-planning pre-solve (Part A+B): BJ/US locks + building_inch_capacity + LOCK-AWARE
    # buildable_rate, computed BEFORE curing so the Phase-0 scheduler can supply-match its draw.
    _coplan_lock: dict = {}
    _building_inch_capacity: dict | None = None
    _feed_ctx: dict | None = None                   # PERSKU_FEED: lock-aware SKU->machines + GT/day
    _coplan_inch_dem: dict = {}                      # INCH18_DEFER: per-inch total demand
    if _GROUP_INCH_POLICY and _early_sku_inch:
        try:
            from bc_config import make_engine as _mk_cp
            _cp = _co_plan_supply(_mk_cp(), demand_path, _early_sku_inch, planning_days)
            _coplan_lock = _cp["bjus_lock"]
            _building_inch_capacity = _cp["building_inch_capacity"]
            _buildable_rate = _cp["buildable"]      # override with the LOCK-AWARE rate
            _feed_ctx = _cp.get("feed_ctx")
            _coplan_inch_dem = _cp.get("inch_dem", {})
            print(f"  [Rolling] CO-PLAN: {len(_coplan_lock)} BJ/US locks; building_inch_capacity "
                  f"{ {k: round(v) for k, v in sorted(_building_inch_capacity.items())} }")
        except Exception as _e:
            print(f"  [Rolling] co-plan supply failed ({_e})")
    # Per-machine ONE-WAY inch switches (kept even when the GENERAL rule is OFF): 7002 14"→16" and
    # 7501 12"→13" build their dominant inch until its share is done, then flip ONCE and STAY (no
    # revert) — stops the 14↔16 / 12↔13 ping-pong that Lever-A re-ranking otherwise causes.
    switch_day_7002 = switch_day_7501 = None
    if _building_inch_capacity:
        def _oneway_switch_day(mach, dom, others_rate):
            _d = float(_coplan_inch_dem.get(dom, 0.0))
            _r = max(1.0, _bld_qty_per_shift(mach) * 3.0)
            _short = max(0.0, _d - others_rate * planning_days)
            return min(planning_days, max(1, math.ceil(_short / _r) + 1))
        # 7002: LEFT FLEXIBLE (14"/16"), NOT one-way. Measured: its 14↔16 flexibility is PRODUCTIVE
        # (July 705k vs 699.5k if pinned) because 14" keeps a real gap all month; an upfront one-way
        # switch either strands 7002 on 16" (co-plan overestimates 14" coverage → day-1 switch) or on
        # 14". The only cost of the flex is the ~585 CO-min of day-to-day churn (a runtime inch-dwell,
        # not a one-way lock, is the right way to damp that — see the 7002 note). switch_day_7002=None.
        # 7501: 12" until the OTHER 12" machines can cover the rest → 13" (one-way; the user's rule).
        _cap12 = float((_building_inch_capacity or {}).get("12", 0.0))
        switch_day_7501 = _oneway_switch_day(
            "7501", "12", max(0.0, _cap12 - _bld_qty_per_shift("7501") * 3.0))
        print(f"  [INCH_ONEWAY] 7501: 12\"→13\" one-way @day {switch_day_7501} (7002 left flexible 14/16)")

    cc_result = run_dynamic_consumption(
        demand_path=demand_path, output_path=CC_OUTPUT,
        plan_start=plan_start, planning_days=planning_days,
        max_co_per_day=MAX_CHANGEOVERS_PER_DAY,
        buildable_rate=_buildable_rate,
        sku_inch=_early_sku_inch,
        building_inch_capacity=_building_inch_capacity,
        feed_ctx=_feed_ctx,
        priority_deadline_map=(priority_deadline_map if _prio_active else None),
        reactive_only=_REACTIVE_ONLY,   # Part B: skip planned schedule + CC workbook
        switch_day18=None,   # 18" defer REMOVED — 18" is built + cured from day 1 (general one-way rule)
        # STAGE 3 (mid-month): seed the CO planner from the carried day-K press positions (same as
        # the day-loop injection) so plan and execution agree. None on a normal run → Day-0 snapshot.
        initial_press_state=({
            "press_to_sku": {p: v["sku"] for p, v in initial_state["press_state"].items()},
            "mould_life": initial_state.get("mould_life", {}),
            "press_moulds": {p: list(m) for p, m in initial_state.get("press_moulds", {}).items()},
        } if initial_state is not None else None),
    )
    co_events = cc_result["co_events"]
    df_day0   = cc_result["df_day0"]
    # CAMPAIGN active-set: {sku: {day: target_presses}} from the Phase-0 planner. When present,
    # the curing sim caps per-SKU RUNNING presses to the day's target and IDLES the excess (so
    # per-inch draw ≤ building supply → building concentrates, cureRUN drops, no fake-busy starve).
    _campaign_target_b2c: dict = cc_result.get("campaign_target", {}) or {}
    _campaign_on = bool(_campaign_target_b2c)
    print(f"  [Rolling] {len(co_events)} curing CO events pre-computed"
          f"{'  | CAMPAIGN active-set ON' if _campaign_on else ''}")

    co_by_day: dict[int, list] = defaultdict(list)
    for ev in co_events:
        co_by_day[int(ev["day"])].append((ev["press"], ev["old_sku"], ev["new_sku"]))

    # Dynamic planner takes over CO decisions entirely — discard the static
    # schedule's events so every remaining co_by_day consumer (daily_co_count
    # seeding, today_cos lookup, the reactive mechanism's "tomorrow" lookahead)
    # is neutralized by this single reset. df_day0/co_events themselves are
    # still needed (SKU classification) so run_dynamic_consumption still runs.
    if (_DYNAMIC_CO_PLANNER_ENABLED or _ROLLING_HORIZON_CO_ENABLED or _RATIO_CO_ALLOCATION_ENABLED
            or _REACTIVE_ONLY):
        co_by_day = {}

    # ── B: Master data ────────────────────────────────────────────────────────
    from bc_config import make_engine
    engine = make_engine()

    cetl = ConsumptionETL(engine)
    df_ct_raw = cetl.load_cycle_times()
    cure_ct_map: dict[str, float] = {
        str(r["SKUCode"]): float(r["CycleTime_min"])
        for _, r in df_ct_raw.iterrows() if r.get("CycleTime_min")
    }

    # Building CTs are hardcoded in _BLD_CT_SEC above (sourced from plant data)

    # ── _ROLLING_HORIZON_CO_ENABLED pre-computation (zero cost when off) ───────
    # Static, reused unchanged for every one of the 31 daily schedule() calls —
    # physical press<->SKU compatibility doesn't change day to day. Uses
    # cc_result["ct_map"], not cure_ct_map above, which has no NaN guard/default
    # (see plan Finding 3) — schedule() has no NaN guard of its own.
    _df_curing_allow_static: "pd.DataFrame | None" = None
    _rolling_co_scheduler: "COScheduler | None" = None
    if _ROLLING_HORIZON_CO_ENABLED:
        _df_curing_allow_static = cetl.load_curing_allowable()
        _rolling_co_scheduler = COScheduler()

    machine_skus: dict[str, set]  = defaultdict(set)
    sku_machine_map: dict[str, set] = defaultdict(set)
    # INCH_STEP_DRIFT: the UN-stripped DB-allowable GT eligibility per machine (before the
    # hist-lock inch strip below) — the source the stepwise drift re-adds from. Empty when OFF.
    _machine_db_skus: dict[str, set] = defaultdict(set)
    _raw_machine_skus: dict[str, set] = defaultdict(set)   # BLD_ACTUAL_SEED: pre-inch-strip certified map
    s1_sku_to_machines: dict[str, set] = defaultdict(set)  # Stage-1 carcass eligibility (kept
    # OUT of machine_skus/sku_machine_map so it never feeds the Stage-2 deficit signal — see
    # the STAGE1 skip below. Used only for Step 3b carcass-utilization simulation.
    sku_inch: dict[str, str] = {}
    try:
        from connection import B2C_ETL as _BETL
        _etl = _BETL(engine)
        df_allow    = _etl.load_machine_allowable()
        sku_to_size = _etl.load_sku_sizes()
        sku_inch = {str(k): str(v).strip().replace('"', "") for k, v in sku_to_size.items()}

        # Fallback: derive inch from characters 9–10 of SKU code (1-indexed) for
        # any SKU missing from the size master.  E.g. "1325216814085SURL0"[8:10] = "14".
        for _, row in df_allow.iterrows():
            sku = str(row["SKUCode"])
            if sku not in sku_inch or not sku_inch[sku]:
                if len(sku) >= 10:
                    sku_inch[sku] = sku[8:10]

        # Hard-inch filter: restricts each machine to its dominant inch(es).
        # Prevents carry-over locking: without this, a machine that does a diff_size_CO
        # in Shift A gets locked onto the non-dominant inch for the rest of the month
        # via machine_current_sku carry-over, starving dominant-inch SKUs.
        # VMIMAXX inches confirmed from May 2026 plant inch-run study (CLAUDE.md §inch-run).
        _HARD = {
            # VMIMAXX — dominant-inch locks (each machine serves its primary inch group)
            # 7001 (dom=16") and 7003 (dom=15") are SOFT-locked: no hard filter here.
            # They prefer dominant inch via inch_penalty sort key; serve others when primary done.
            "6001": {"14"},             # dom=14" — 3 machines share 14" (6001/7002/7004)
            "6002": {"15"},             # dom=15" — 2 machines share 15" (6002/7003)
            "6003": {"17", "18"},       # dom=17" — sole 17"/18" machine; serves HURL1/HRHT0
            "6004": {"16"},             # dom=16" — primary 16" machine; 7001 is secondary
            "7002": {"14"},             # dom=14"
            "7004": {"14"},             # dom=14"
            # BJ — no hard filter; dominant-inch preference via _MACHINE_DOMINANT_INCH + inch_penalty.
            # Removing hard locks allows BJ machines to serve off-dominant SKUs in their
            # allowable table (e.g. 7103 can produce HURL0 at 14" when 13" demand is met).
            # UNI_NARROW — physically cannot run 14"+ (genuine hard constraint)
            # 7501 dominant=12" but can also run 13" (confirmed allowable)
            "7501":{"12","13"},  "7502":{"13"},        "7503":{"13"},
        }
        # Client inch rules, variant A: the +/-2 anchor band REPLACES the
        # dominant-inch hard locks, so every machine starts from its full DB
        # allowable and is bound only by anchor+/-2 and the no-revisit rule.
        # Variant B (_INCH_BAND_REPLACES_HARD=False) keeps _HARD as well, so the
        # machine is bound by the INTERSECTION (most restrictive) — this also
        # preserves the UNI_NARROW 12"/13" physical limit.
        if _INCH_RULES_ENABLED and _INCH_BAND_REPLACES_HARD:
            _HARD = {}
        # Inch-flexibility: drop flex machines from _HARD so their full DB
        # allowable (all eligible inches) passes through — off-inch access is
        # then gated at Campaign-2+ (primary_demand_done) + reclamation guard.
        for _fm in _INCH_FLEX_MACHINES:
            _HARD.pop(_fm, None)
        # BLD_ACTUAL_SEED: RAW (pre-inch-strip) allowable {machine: set(SKUs)} — the
        # "Yuvraj" certified map. Used to validate a plant seed whose inch the hist-lock
        # would otherwise strip (the whole point: the seed inch may differ from history).
        for idx, row in df_allow.iterrows():
            sku = str(row["SKUCode"]); si = sku_inch.get(sku, "")
            ml  = list(row.get("Machines", []) or [])
            if _BLD_ACTUAL_SEED:                        # capture full certified eligibility (incl STAGE1)
                for _m in ml:
                    _raw_machine_skus[str(_m)].add(sku)
            if _INCH_STEP_DRIFT:                        # capture DB eligibility BEFORE the hist-lock strip
                for _m in ml:
                    if _MACHINE_GROUP.get(str(_m), "") != "STAGE1":
                        _machine_db_skus[str(_m)].add(sku)
            df_allow.at[idx, "Machines"] = [
                m for m in ml
                if (str(m) not in _HARD or si in _HARD[str(m)])
                # INCH_HIST_LOCK: strip SKUs whose inch is not in the machine's
                # historical allowed-inch set → FIXED machines become single-inch,
                # FLEXIBLE machines are bounded to their ranked historical inches
                # (this is what makes Phase A/C + the pool builder respect the lock,
                # not just the Phase-B _inch_gate). OFF → passes everything.
                and (not _INCH_HIST_LOCK_ENABLED
                     # ONE-WAY: keep the FULL DB-allowable set eligible so any inch is reachable;
                     # the one-way + ≤2-jump discipline is enforced at runtime in _inch_gate.
                     or _ONEWAY_INCH_ENABLED
                     # GENERAL one-way rule: keep every CT-allowable SKU eligible; machine_locked_inches
                     # (seeded to the dominant, flipped once when done) does the per-day restriction.
                     or _ONEWAY_INCH_GENERAL
                     or str(m) not in _MACHINE_ALLOWED_INCH_SET
                     or si in _MACHINE_ALLOWED_INCH_SET[str(m)]
                     # Lever B: keep a fixed machine's full DB-allowable set eligible so it
                     # CAN escape to a certified off-inch; the escape is gated at runtime.
                     or (_FIXED_ESCAPE_ENABLED and str(m) in _FIXED_MACHS_HIST))
            ]
        for _, row in df_allow.iterrows():
            sku = str(row["SKUCode"])
            for m in (row.get("Machines") or []):
                m_str = str(m)
                # Rolling pipeline tracks GT only (not carcass).
                # Stage-1 machines produce carcass that feeds Stage-2 implicitly.
                # Excluding them from machine_skus prevents:
                #   (a) phantom projected_gt updates inside _assign_building_shift
                #       that fool Stage-2 into seeing deficit=0 for Stage-2 SKUs
                #   (b) Stage-1 machines "wasting" shifts on carcass campaigns whose
                #       output (correctly) doesn't add to gt_inventory (line 1388)
                # The planning assumption: Stage-1 always has capacity to supply
                # whatever Stage-2 needs (validated: Stage-1 util ≈ 33% by design).
                if _MACHINE_GROUP.get(m_str, "") == "STAGE1":
                    if m_str != "6801":          # 6801 is plant-RETIRED — not an eligible carcass
                        s1_sku_to_machines[sku].add(m_str)   # machine (R14 expects the 14 live S1)
                    continue
                machine_skus[m_str].add(sku)
                sku_machine_map[sku].add(m_str)
        print(f"  [Rolling] Allowable map: {len(machine_skus)} machines")
    except Exception as _e:
        print(f"  [Rolling] Allowable map: failed ({_e})")

    # ── C: Press state ────────────────────────────────────────────────────────
    # PRESS ROSTER = UNION of the Day-0 running-moulds snapshot AND the allowable matrix.
    # Every press in the running-moulds snapshot is used (running presses are always in the
    # allowable matrix); allowable presses absent from the snapshot are cold-started as
    # production via IDLE_PRESS_ACTIVATE below (all new presses are client-confirmed ready).
    # (The old "drop a running press not in the allowable matrix" restriction was removed —
    # running ⊆ allowable, so it dropped nothing.)
    df_moulds = cetl.load_running_moulds()
    # DIAGNOSTIC (env EXCLUDE_PRESS_PREFIX): also drop matching presses from the running snapshot
    # so a run can fully reproduce the roster before a set of presses existed (default no-op).
    _xpfx = tuple(p for p in os.environ.get("EXCLUDE_PRESS_PREFIX", "").split(",") if p)
    if _xpfx:
        df_moulds = df_moulds[~df_moulds["Machine"].astype(str).str.startswith(_xpfx)].reset_index(drop=True)
    press_state: dict[str, dict] = {}
    for _, r in df_moulds.iterrows():
        press_state[str(r["Machine"])] = {"sku": str(r["SKUCode"]), "status": "RUNNING"}

    press_count: dict[str, int] = defaultdict(int)
    for st in press_state.values():
        press_count[st["sku"]] += 1

    # PRESS_RETURN_BLOCK: per-press set of SKUs this press has CO'd AWAY from during THIS run's
    # simulation (recorded at every planned + dynamic transition). Used by the planned-CO retarget
    # to avoid sending a press back to a SKU it just left. Empty at run start (a press has not "left"
    # its Day-0 SKU); per-run local so the 2pass mid-month Run-2 starts fresh.
    press_ran: dict[str, set] = defaultdict(set)

    # IDLE_PRESS_ACTIVATE: roster presses (the 170 in the allowable matrix) that are NOT in the
    # Day-0 running snapshot. Seeded into press_state via a Day-1 cold-start CO in the day loop
    # (they are NOT added to press_state here — so pre-day-loop state stays identical to OFF).
    _idle_presses: list = []
    if _IDLE_PRESS_ACTIVATE:
        try:
            _roster_ids = {str(p) for p in cetl.load_allowable_press_ids()}
            _idle_presses = sorted(p for p in _roster_ids if p not in press_state)
            print(f"  [Rolling] IDLE_PRESS_ACTIVATE ON — {len(_idle_presses)} roster press(es) "
                  f"absent from Day-0 snapshot to cold-start Day-1 Shift A: {_idle_presses}")
        except Exception as _e:
            _idle_presses = []
            print(f"  [Rolling] IDLE_PRESS_ACTIVATE: roster load failed ({_e}); none activated")

    # FULL_PRESS_ROSTER: independently union EVERY allowable-matrix press absent from the Day-0
    # snapshot into the cold-start idle set (idempotent with IDLE_PRESS_ACTIVATE when that is ON).
    # Lets a run force the full 185-press roster even when IDLE_PRESS_ACTIVATE is off. OFF = no-op.
    if _FULL_PRESS_ROSTER:
        try:
            _roster_ids = {str(p) for p in cetl.load_allowable_press_ids()}
            _extra = sorted(p for p in _roster_ids
                            if p not in press_state and p not in _idle_presses)
            if _extra:
                _idle_presses = sorted(set(_idle_presses) | set(_extra))
                print(f"  [Rolling] FULL_PRESS_ROSTER ON — added {len(_extra)} roster press(es) "
                      f"absent from Day-0 snapshot as cold-start idle: {_extra}")
            else:
                print(f"  [Rolling] FULL_PRESS_ROSTER ON — no additional presses "
                      f"(IDLE_PRESS_ACTIVATE already covers the full roster); idle set unchanged")
        except Exception as _e:
            print(f"  [Rolling] FULL_PRESS_ROSTER: roster load failed ({_e}); no extra presses")

    # ── C2: Mould pool (client mould-availability gate) ───────────────────────
    # sku_moulds[sku] = eligible mould IDs; mould_owner[mould] = press it's mounted
    # on (None = free in storage). Day-0 mounting comes from the running-moulds
    # MouldNos list (2 moulds per press). A press CO/run is feasible only if 2
    # eligible moulds are free (or already on this press). See _MOULD_GATE_ENABLED.
    _sku_moulds: dict[str, set] = {}
    _mould_skus: dict[str, set] = {}
    _mould_owner: dict[str, str] = {}
    _press_moulds: dict[str, set] = {}
    mould_blocked_cos = 0
    mould_day0_topups = 0                        # Day-0 single-mould presses given a 2nd mould
    # Full (press, SKU, 2-moulds) timeline — Day-0 opening + every changeover mount.
    # Expanded in the Mould Tracker sheet to one row per (press, mould) building a SKU.
    mould_assignments: list = []
    _mould_selfcheck = [0]                      # debug: count RUNNING-without-2-moulds
    _mould_gate = _MOULD_GATE_ENABLED          # local (may disable on load failure)
    _mould_opt  = _MOULD_OPT_ENABLED           # Phase-2 optimisation (needs the gate)
    mould_retargeted_cos = 0                    # planned COs saved by retarget-on-block
    if _mould_gate:
        try:
            _elig = cetl.load_mould_eligibility()
            _sku_moulds = {k: set(v) for k, v in _elig["sku_moulds"].items()}
            _mould_skus = {k: set(v) for k, v in _elig["mould_skus"].items()}
        except Exception as _e:
            print(f"  [Rolling] mould eligibility load FAILED ({_e}); gate disabled")
            _mould_gate = False
    if _mould_gate:
        # Seed Day-0 mounted moulds from the running-moulds MouldNos list.
        for _, r in df_moulds.iterrows():
            _p = str(r["Machine"]); _sku0 = str(r["SKUCode"])
            _mn = r.get("MouldNos", []) or []
            for _m in _mn:
                _m = str(_m).strip()
                if not _m or _m.lower() == "nan":
                    continue
                _mould_owner[_m] = _p
                _press_moulds.setdefault(_p, set()).add(_m)
                # Fold Day-0 orphan pairs into eligibility so the seed never
                # self-violates (5 mounted moulds aren't listed for their SKU).
                _sku_moulds.setdefault(_sku0, set()).add(_m)
                _mould_skus.setdefault(_m, set()).add(_sku0)

        # Day-0 second-mould top-up: a few presses list only 1 mould in
        # Daily_Running_Moulds (e.g. 75214, 9404). A press physically has 2 cavities,
        # so give each such press a 2nd COMPATIBLE FREE mould for its Day-0 SKU (the
        # realistic floor state). Deterministic (sorted free pool). Runs AFTER all
        # presses are seeded so the free pool is complete.
        _topped = 0
        for _, r in df_moulds.iterrows():
            _p = str(r["Machine"]); _sku0 = str(r["SKUCode"])
            _have = _press_moulds.get(_p, set())
            if len(_have) >= 2:
                continue
            _elig0 = _sku_moulds.get(_sku0, set())
            _free0 = sorted(m for m in _elig0 if _mould_owner.get(m) is None)
            for _m in _free0:
                if len(_press_moulds.get(_p, set())) >= 2:
                    break
                _mould_owner[_m] = _p
                _press_moulds.setdefault(_p, set()).add(_m)
                _topped += 1
        mould_day0_topups = _topped
        if _topped:
            print(f"  [Rolling] Day-0 second-mould top-up: {_topped} moulds "
                  f"assigned to single-mould presses")

        # Record each press's Day-0 opening (press, SKU, 2 moulds) so the expanded
        # tracker timeline starts from the floor state before any changeover.
        for _, r in df_moulds.iterrows():
            _p = str(r["Machine"]); _sku0 = str(r["SKUCode"])
            mould_assignments.append({
                "day": 1, "press": _p, "sku": _sku0,
                "moulds": sorted(_press_moulds.get(_p, set())),
            })

    # Old moulds whose release is deferred to end-of-day: a planned CO placed in
    # shift B/C keeps running its OLD sku (needs its old moulds) until the CO fires,
    # so the old moulds must stay reserved to the press until the day completes —
    # freeing them at day-start let another press grab them and made the OLD sku's
    # pre-CO production physically mould-less. {press: set(old moulds to free)}.
    _deferred_free: dict[str, set] = {}

    # Mould-movement log: every time a press mounts a DIFFERENT mould (a genuine swap
    # on a changeover), record it so the output can show moulds moving over the month
    # (the tracker sheet's end-state snapshot alone hides this). {day, press, sku,
    # added, removed}. _cur_day is set at the top of each day loop iteration.
    mould_events: list = []
    _cur_day = [0]

    # ── Mould-contention gate (MONOTONICITY FIX) support ──────────────────────────
    # Reverse map: mould -> set of DEMANDED SKUs eligible for it (static eligibility).
    _mould_to_skus: dict[str, set] = defaultdict(set)
    if _MOULD_CONTENTION_GATE:
        for _sku_e, _ms_e in _sku_moulds.items():
            for _m_e in _ms_e:
                _mould_to_skus[_m_e].add(_sku_e)

    def _avail_for_sku(sku: str, exclude: set) -> int:
        """How many of `sku`'s eligible moulds remain USABLE for it if `exclude` moulds
        are taken away: a mould counts if it is free OR already owned by a press that is
        currently running `sku` (so `sku` keeps it). ≥2 ⇒ sku can still hold a pair."""
        cnt = 0
        for _m in _sku_moulds.get(sku, ()):  # small (≈ that SKU's mould list)
            if _m in exclude:
                continue
            _o = _mould_owner.get(_m)
            if _o is None or press_state.get(_o, {}).get("sku") == sku:
                cnt += 1
                if cnt >= 2:
                    return cnt
        return cnt

    def _strips_claimant(chosen: list, new_sku: str) -> bool:
        """True if mounting `chosen` for `new_sku` would drop a DIFFERENT demanded SKU that
        shares one of these moulds below 2 usable moulds (i.e. displace a productive press)."""
        _ch = set(chosen)
        _seen: set = set()
        for _m in chosen:
            for _S in _mould_to_skus.get(_m, ()):
                if _S == new_sku or _S in _seen:
                    continue
                if demand_remaining.get(_S, 0.0) <= 0:
                    continue
                _seen.add(_S)
                if _avail_for_sku(_S, _ch) < 2:
                    return True
        return False

    def _try_mount(press: str, new_sku: str, defer_free: bool = False) -> bool:
        """Allocate 2 eligible moulds for `new_sku` on `press`, or return False.

        Prefers moulds already on this press (shared → no movement, leaves more free
        moulds for others). `defer_free`=True (planned COs): keep the press's OLD
        moulds reserved to it until end-of-day (they still serve the old SKU in
        pre-CO shifts). `defer_free`=False (reactive mid-shift CO): free old moulds
        immediately (the press is CHANGEOVER that shift, not producing).
        """
        if not _mould_gate:
            return True
        elig = _sku_moulds.get(new_sku, set())
        if len(elig) < 2:
            return False
        own = _press_moulds.get(press, set())
        # candidates: eligible moulds free OR already on this press. SORTED — set
        # iteration order is hash-randomised → would break determinism.
        reuse = sorted(m for m in own if m in elig)
        free  = sorted(m for m in elig if _mould_owner.get(m) is None and m not in reuse)
        if _MOULD_CONTENTION_GATE:
            # Prefer the LEAST-contended free moulds (fewest other demanded claimants), then refuse
            # the mount if the resulting pair would strip a mould-sharing under-served SKU's last
            # pair (displace a productive press). reuse moulds are already this press's → never strip.
            free = sorted(free, key=lambda m: (
                sum(1 for _S in _mould_to_skus.get(m, ())
                    if _S != new_sku and demand_remaining.get(_S, 0.0) > 0), m))
            chosen = (reuse + free)[:2]
            if len(chosen) < 2:
                return False
            if _strips_claimant(chosen, new_sku):
                _cand = reuse + free
                _ok = None
                for _i in range(len(_cand)):
                    for _j in range(_i + 1, len(_cand)):
                        _pair = [_cand[_i], _cand[_j]]
                        if not _strips_claimant(_pair, new_sku):
                            _ok = _pair
                            break
                    if _ok:
                        break
                if _ok is None:
                    return False           # any pair would displace a productive press → stay idle
                chosen = _ok
        else:
            chosen = (reuse + free)[:2]
            if len(chosen) < 2:
                return False
        old_extra = [m for m in own if m not in chosen]
        if defer_free:
            # keep old moulds reserved to the press (still serving old sku); free
            # them at end-of-day. The press temporarily "holds" old ∪ new.
            _deferred_free.setdefault(press, set()).update(old_extra)
            _press_moulds[press] = set(own) | set(chosen)
        else:
            for m in old_extra:
                _mould_owner[m] = None
            _press_moulds[press] = set(chosen)
        for m in chosen:
            _mould_owner[m] = press
        # Full assignment: this press now builds `new_sku` on these 2 moulds.
        mould_assignments.append({
            "day": _cur_day[0], "press": press, "sku": new_sku,
            "moulds": list(chosen),
        })
        _added = [m for m in chosen if m not in own]
        if _added:                                  # a genuine mould swap (not pure reuse)
            mould_events.append({
                "day":     _cur_day[0],
                "press":   press,
                "sku":     new_sku,
                "added":   _added,
                "removed": list(old_extra),
            })
        return True

    def _n_free_for(new_sku: str, press: str) -> int:
        """How many eligible moulds could `press` mount for `new_sku` right now
        (own moulds reusable for free + currently-free eligible moulds). ≥2 ⇒ a
        _try_mount would succeed."""
        elig = _sku_moulds.get(new_sku, set())
        if len(elig) < 2:
            return 0
        own = _press_moulds.get(press, set())
        reuse = sum(1 for m in own if m in elig)
        free  = sum(1 for m in elig if _mould_owner.get(m) is None and m not in own)
        return reuse + free

    def _pick_retarget(press: str, avoid: set = None):
        """Phase 2b: a planned CO whose new-SKU has no free moulds would idle the
        press on its (usually demand-done) old SKU. Instead retarget it to the
        NEEDIEST SKU the press is allowable for that still has 2 free moulds.
        Returns the SKU or None. Deterministic (sorted candidate list, tuple key).
        `avoid` (PRESS_RETURN_BLOCK): SKUs to skip (ones this press already left) so the
        retarget doesn't boomerang back; callers fall back to avoid=None if this returns None."""
        _cur = press_state.get(press, {}).get("sku")
        best = None
        best_key = None
        for s in press_allow_skus.get(press, ()):     # pre-sorted list
            if s == _cur:
                continue
            if avoid and s in avoid:
                continue                               # PRESS_RETURN_BLOCK: don't return to a left SKU
            rem = demand_remaining.get(s, 0.0)
            if rem <= 0:
                continue
            if _n_free_for(s, press) < 2:
                continue
            npr = max(1, press_count.get(s, 0))
            key = (rem / npr, rem, s)                  # neediest per serving-press first
            if best_key is None or key > best_key:
                best_key = key
                best = s
        return best

    def _release_deferred():
        """Free the deferred old moulds at end-of-day (their CO has now completed).

        `_olds` is exactly the set of old moulds NOT kept as current (computed as
        old_extra in _try_mount), so free them unconditionally: drop from the
        press's owned set AND release ownership so they return to the free pool.
        """
        for _p, _olds in _deferred_free.items():
            _cur = _press_moulds.get(_p, set())
            for _m in _olds:
                if _mould_owner.get(_m) == _p:
                    _mould_owner[_m] = None
                _cur.discard(_m)
            _press_moulds[_p] = _cur
        _deferred_free.clear()

    # Curing allowable: {sku: [press_ids]} for demand fulfillment sheet
    curing_allowable: dict[str, list] = defaultdict(list)
    for press, st in press_state.items():
        curing_allowable[st["sku"]].append(press)

    # ── D: Opening GT inventory ───────────────────────────────────────────────
    try:
        from connection import _load_opening_gt
        opening_gt = _load_opening_gt(engine)
    except Exception:
        opening_gt = {}
    gt_inventory: dict[str, float] = defaultdict(float, opening_gt)
    # ── GT strict per-lot FIFO (R9G / no-waste fix) ───────────────────────────────
    # GT held as DATED LOTS per SKU: gt_lots[sku] = [[build_day, qty], ...]. Curing
    # consumes OLDEST-first; a lot older than GT_SHELF_LIFE_DAYS is dropped as WASTE at
    # day-start and can NEVER be cured (mirrors the carcass bank). gt_inventory (scalar)
    # is kept as the running SUM of live lots so every existing reader stays correct.
    # writeoff_cum feeds the demand-cap fix (built<=demand) + the "expired GT/carcass" col.
    gt_lots: dict[str, list] = {s: [[0, float(q)]] for s, q in dict(opening_gt).items() if q > 0}
    writeoff_cum: dict[str, float] = defaultdict(float)   # cumulative expired GT per SKU
    carcass_waste: dict[str, float] = defaultdict(float)  # cumulative expired carcass per SKU
    # Per-(day, shift, SKU) expired GT/carcass rows for the building Shift Schedule (display +
    # per-day Daily-GT-&-Carcass Expired columns). Kept OUT of bld_shift_rows / prod_rows so no
    # KPI, utilization, CO-count, or feasibility production sum is affected — they are waste
    # markers only (Qty visible, CO_Mins=0, Machine="—").
    expiry_rows: list[dict] = []

    def _gt_consume_lots(_s, _q):
        """Consume _q units of GT for SKU _s from the oldest lots first (FIFO)."""
        _lots = gt_lots.get(_s)
        if not _lots:
            return
        _r = float(_q)
        while _r > 1e-9 and _lots:
            if _lots[0][1] <= _r + 1e-9:
                _r -= _lots[0][1]
                _lots.pop(0)
            else:
                _lots[0][1] -= _r
                _r = 0.0

    def _gt_expire_lots(_today, _date_str=None):
        """Drop every GT lot older than GT_SHELF_LIFE_DAYS as waste (return total).
        When _date_str is given, also emit one expired_GT Shift-Schedule row per SKU
        (attributed to Shift A = day start, when the lot is dropped)."""
        _tot = 0.0
        _per: dict[str, float] = defaultdict(float)
        for _s in list(gt_lots.keys()):
            _keep = []
            for _bd, _q in gt_lots[_s]:
                if _q <= 1e-9:
                    continue
                if (_today - _bd) > GT_SHELF_LIFE_DAYS:
                    _tot += _q
                    writeoff_cum[_s] += _q
                    _per[_s] += _q
                    gt_inventory[_s] = max(0.0, gt_inventory.get(_s, 0.0) - _q)
                else:
                    _keep.append([_bd, _q])
            gt_lots[_s] = _keep
        if _date_str is not None:
            _st = _fmt_dt(_shift_start_dt(_date_str, "A"))
            for _s, _q in _per.items():
                if _q >= 0.5:
                    expiry_rows.append({"Machine": "—", "Date": _date_str, "Shift": "A",
                                        "SKUCode": _s, "Qty": int(round(_q)), "CO_Mins": 0,
                                        "StartTime": _st, "EndTime": _st,
                                        "Machine_Group": "", "CO_Type": "expired_GT"})
        return _tot

    # #1: Opening carcass inventory (consumed first by the Stage-1 carcass pass; KPI-neutral).
    opening_carcass: dict[str, float] = {}
    if _CARCASS_INV_ENABLED:
        try:
            from connection import _load_opening_carcass
            opening_carcass = _load_opening_carcass(engine)
        except Exception:
            opening_carcass = {}

    # ── D2: Mould-clean state (per press) ─────────────────────────────────────
    # remaining_mould_life = cycles a press may still run before a mandatory clean.
    # v1: everyone opens fresh at MOULD_CLEAN_CYCLES. v2 (_MOULD_LIFE_FROM_DB): seed the
    # OPENING life from the real DB remaining (3000 − consumed cycles), taken as the MIN
    # over the press's 2 moulds so both clean together — already computed by
    # load_running_moulds() as MouldLife_remaining. clean_carry = minutes of an in-progress
    # clean owed at the start of a press's next shift.
    mould_life:  dict[str, int]   = defaultdict(lambda: MOULD_CLEAN_CYCLES)
    if _MOULD_LIFE_FROM_DB and _MOULD_CLEAN_ENABLED and "MouldLife_remaining" in df_moulds.columns:
        _n_seeded = _n_low = 0
        for _, _r in df_moulds.iterrows():
            _rem = _r.get("MouldLife_remaining")
            if _rem is None or pd.isna(_rem):
                continue
            _rem = max(0, min(MOULD_CLEAN_CYCLES, int(_rem)))   # 3000 − consumed, clamped
            mould_life[str(_r["Machine"])] = _rem
            _n_seeded += 1
            if _rem < MOULD_CLEAN_CYCLES:
                _n_low += 1
        print(f"  [Rolling] Mould life v2: seeded {_n_seeded} presses from DB "
              f"({_n_low} open below 3000 → earlier cleans)")
    clean_carry: dict[str, float] = defaultdict(float)
    # Minutes of an in-progress CURING CHANGEOVER owed at the start of a press's
    # next shift. A dynamic (instant) CO fires mid-shift the moment demand is met,
    # so its CURING_CO_CHANGEOVER_MINS overhang spills past the shift boundary —
    # the new SKU therefore starts MID-shift, not at the boundary.
    co_carry: dict[str, float] = defaultdict(float)
    # CONTINUOUS CYCLE CARRY (PLANT RULE — permanent): leftover production-eligible minutes carried
    # from a press's PREVIOUS shift — the fractional cure cycle in progress at the shift boundary.
    # Curing is continuous, so a press that ran flat-out (TIME-limited) does NOT reset to a fresh
    # 480 min each shift; its partial cycle finishes early next shift. Added to _avail before the
    # cap so `int((_avail+carry)/ct)` recovers the fraction (30,32,32,… not 30,30,30,…). Reset to 0
    # on CO / mould-clean / holiday / GT-or-demand-limited (idle at the boundary → no cycle to carry).
    _cure_carry_min: dict[str, float] = defaultdict(float)

    # ── E: Demand ─────────────────────────────────────────────────────────────
    demand_df = pd.read_excel(demand_path)
    sku_col = next((c for c in demand_df.columns if "SKU"  in str(c)), demand_df.columns[0])
    qty_col = next(
        (c for c in demand_df.columns
         if any(x in str(c) for x in ("Requirement","Demand","Qty","Quantity"))),
        demand_df.columns[1],
    )
    # Sum duplicate SKU rows (a SKU may appear on several demand line-items).
    # MUST match curing_consumption_dynamic.py's groupby-sum, otherwise the building
    # demand universe silently diverges from curing's: a plain dict keyed by
    # SKUCode would keep only the LAST duplicate row and drop the rest (e.g.
    # LSTL0 71,000 + 20,680 → only 20,680 survives), capping building far below
    # the real demand while curing presses still pull the full amount → mass
    # starvation on the biggest SKUs.
    _dq = demand_df[[sku_col, qty_col]].copy()
    _dq[sku_col] = _dq[sku_col].astype(str).str.strip()
    _dq[qty_col] = pd.to_numeric(_dq[qty_col], errors="coerce")
    _dq = _dq.dropna(subset=[qty_col])
    demand_dict: dict[str, float] = _dq.groupby(sku_col)[qty_col].sum().to_dict()
    if os.environ.get("DEMAND_INT_NORMALIZE", "1") != "0":
        # Demand is physically integer. Strip float dust baked into some xlsx cells
        # (e.g. 13750.000000000002) so demand_remaining drains to EXACTLY 0 and the
        # `_demand_done` (<= 0) reactive-CO test fires identically to the DB-int path.
        # LOCAL↔CLOUD parity: cloud reads jkt_demand.requirement (int) → round() is a no-op,
        # so cloud output is byte-unchanged; local converges to cloud. DEMAND_INT_NORMALIZE=0 reverts.
        demand_dict = {k: float(round(v)) for k, v in demand_dict.items()}
    _n_rows_raw = len(_dq)
    demand_remaining: dict[str, float] = dict(demand_dict)
    total_demand = sum(demand_dict.values())
    if _n_rows_raw != len(demand_dict):
        print(f"  [Rolling] Demand: collapsed {_n_rows_raw} rows → "
              f"{len(demand_dict)} unique SKUs (summed {_n_rows_raw - len(demand_dict)} duplicate rows)")
    print(f"  [Rolling] Demand: {len(demand_dict)} SKUs, {total_demand:,.0f} units")

    # Static per-machine total demand for _BUILDING_RATIO_ENABLED — computed once
    # from the fixed demand file, never decremented (see _priority_tier).
    machine_total_demand: dict[str, float] = {
        m: sum(demand_dict.get(s, 0.0) for s in skus)
        for m, skus in machine_skus.items()
    }

    # ── _DYNAMIC_CO_PLANNER_ENABLED pre-computation (zero cost when off) ───────
    # Curing press <-> SKU physical compatibility (mould/allowable-machines) —
    # run_rolling_pipeline never loaded this before; the existing reactive
    # dynamic_co_tracker mechanism has no compatibility check at all (fine at
    # its current one-press-at-a-time scale, not fine once scaled to the whole
    # press fleet once a day, so this is required here).
    press_to_demand_targets: dict[str, list] = {}
    press_total_demand: dict[str, float] = {}
    sku_to_press_count: dict[str, int] = {}
    ri_skus: frozenset = frozenset()
    nri_skus: frozenset = frozenset()
    if _DYNAMIC_CO_PLANNER_ENABLED or _RATIO_CO_ALLOCATION_ENABLED or _EARLY_CO_ENABLED:
        df_curing_allow = cetl.load_curing_allowable()
        _all_demand_skus = set(demand_dict.keys())
        sku_to_presses: dict[str, set] = {}
        for _, _row in df_curing_allow.iterrows():
            _sku = str(_row["SKUCode"]).strip()
            if _sku in _all_demand_skus:
                _machines = _row.get("Machines", [])
                if _machines:
                    sku_to_presses[_sku] = {str(_p) for _p in _machines}
        for _sku, _presses in sku_to_presses.items():
            for _p in _presses:
                press_to_demand_targets.setdefault(_p, []).append(_sku)
        press_total_demand = {
            _p: sum(demand_dict.get(_s, 0.0) for _s in _targets)
            for _p, _targets in press_to_demand_targets.items()
        }
        # Target-side scarcity for _RATIO_CO_RICH_RANKING_ENABLED — how many
        # presses could physically serve this SKU (fewer = protect it).
        sku_to_press_count = {_s: len(_p) for _s, _p in sku_to_presses.items()}
        ri_skus  = frozenset(df_day0.loc[df_day0["Category"] == "Runner-In",     "SKUCode"])
        nri_skus = frozenset(df_day0.loc[df_day0["Category"] == "Non-Runner-In", "SKUCode"])

    # Priority score for dynamic CO target selection (higher = serve first).
    # ConsolidatedPriorityScore (v1) — min-max normalise the per-SKU requirement,
    # computed from demand_dict (already summed per SKU) so it matches
    # curing_consumption_dynamic.load_demand exactly and is identical whether the demand
    # came from a local Excel or the cloud DB `jkt_demand` table (no priority
    # column). v1 uses REQUIREMENT ONLY; any priority column in the file is ignored.
    #   score = (q - q_min) / (q_max - q_min)   (1.0 for all if q_max == q_min)
    priority_score_map: dict[str, float] = {}
    if demand_dict:
        _qs = list(demand_dict.values())
        _qmin, _qmax = min(_qs), max(_qs)
        _span = _qmax - _qmin
        priority_score_map = {
            _s: ((_q - _qmin) / _span if _span > 0 else 1.0)
            for _s, _q in demand_dict.items()
        }

    # ── F: Machine current SKU ────────────────────────────────────────────────
    machine_current_sku: dict[str, str] = {}
    if _BLD_START_FREE:
        # All building machines start free → first shift seeds each as a "start"
        # campaign (0 CO); no initial same/diff-size building COs.
        print("  [Rolling] BLD_START_FREE: all building machines start free "
              "(no Day-0 running SKU, no initial building COs)")
    else:
        try:
            df_running_bld = _etl.load_running_machines()
            machine_current_sku = {str(r["Machine"]): str(r["SKUCode"]) for _, r in df_running_bld.iterrows()}
        except Exception:
            pass
    if _SEED_FROM_PLANT_RUNNING and not _BLD_START_FREE:
        try:
            _pr = pd.read_excel(_PLANT_RUNNING_FILE)
            _seed = {}
            for _, r in _pr.iterrows():
                _m   = str(r["Machine_Code"]).strip()
                _sku = str(r["SKUCode"]).strip()
                if not _sku or _sku.lower() == "nan" or "IDLE" in _sku:
                    continue
                _seed[_m] = _sku
            machine_current_sku = _seed
            print(f"  [Rolling] Seeded machine_current_sku from plant running file: {len(_seed)} machines")
        except Exception as _e:
            print(f"  [Rolling] Plant-running seed FAILED: {_e}")

    # ── BLD_ACTUAL_SEED: Day-1 machine→SKU seed from ACTUAL plant production ──────
    # Each building machine (Stage-1 + Stage-2/GT) starts Day-1 on the SKU it was ACTUALLY
    # building latest, so Day-1 does not reshuffle (Phase-A continues it at 0 CO). Takes
    # PRECEDENCE over BLD_START_FREE. Only the full-month / Run-1 Day-1 is seeded — a
    # mid-month 2-pass carries its own state (initial_state), which wins (injected at day 1).
    _BLD_SEED_MACHINES.clear(); _BLD_SEED_INCH.clear(); _BLD_SEED_SKU.clear()
    _seed_gt: dict[str, str] = {}                     # GT machines actually seeded (for inch override below)
    _seed_s1: dict[str, str] = {}                     # Stage-1 machines actually seeded
    if _BLD_ACTUAL_SEED and initial_state is None:
        # MID-MONTH: when the plan does NOT start on the 1st, the plant has already run the
        # first days and the replay is off (its Day column is a plan-day index) — so the
        # carry-in state is the END of the plant's last replayed day, derived from
        # PLANT_2DAY_SCHEDULE_FILE. Falls back to the static seed file if that is unavailable.
        # A 1st-of-month start keeps the static Day-1 seed exactly as before (bit-for-bit).
        # MID-MONTH N-DAY SET: each machine's recent-SKU set from the full-month baseline
        # plan. Day-1 Shift A then picks the best-need SKU from its OWN set, no CO charged.
        _MIDMONTH_SET.clear()
        if int(getattr(plan_start, "day", 1) or 1) != 1:
            _MIDMONTH_SET.update(_derive_midmonth_sets(plan_start))
            # HARD DAY-1 START OVERRIDE (bc_config.DAY1_FORCE_SKU): plant instruction — these
            # machines MUST start day 1 on exactly this SKU. Replace their derived set with the
            # single forced SKU and make it eligible (the hist-lock strip may have removed it).
            # DEDICATED MACHINES: reduce the machine's allowable set to its single SKU for the
            # WHOLE month (not just day 1), and guarantee that SKU is eligible on it.
            _EARLY_SKUS.clear()
            if _EARLY_MACHINES and _EARLY_DAYS > 0:
                for _em in _EARLY_MACHINES:
                    # .update(), NOT |= — an augmented assignment would make _EARLY_SKUS a local
                    _EARLY_SKUS.update({k for k, v in s1_sku_to_machines.items() if _em in v})
                    _EARLY_SKUS.update(machine_skus.get(_em, set()))
                if _EARLY_SKUS:
                    print(f"  [EARLY_FULL_LOAD] {_EARLY_MACHINES} → first call on "
                          f"{len(_EARLY_SKUS)} SKU(s) for plan days 1..{_EARLY_DAYS}")
            for _dm, _ds in _DEDICATED.items():
                if _dm in _S1_MACHINES:
                    for _other, _ms in list(s1_sku_to_machines.items()):
                        if _other != _ds:
                            _ms.discard(_dm)
                    s1_sku_to_machines[_ds].add(_dm)
                else:
                    machine_skus[_dm] = {_ds}
                    for _other, _ms in list(sku_machine_map.items()):
                        if _other != _ds:
                            _ms.discard(_dm)
                    sku_machine_map[_ds].add(_dm)
                print(f"  [DEDICATED] {_dm} restricted to {_ds} for the whole month")
            for _fm, _fs in (getattr(_bc_cfg, "DAY1_FORCE_SKU", {}) or {}).items():
                _fm, _fs = str(_fm), str(_fs)
                _MIDMONTH_SET[_fm] = {_fs}
                _MIDMONTH_LAST[_fm] = _fs
                if _fm in _S1_MACHINES:
                    s1_sku_to_machines[_fs].add(_fm)
                else:
                    machine_skus[_fm].add(_fs); sku_machine_map[_fs].add(_fm)
                print(f"  [DAY1_FORCE] {_fm} pinned to {_fs} on plan day 1")
            if _MIDMONTH_SET:
                _sz = sorted(len(v) for v in _MIDMONTH_SET.values())
                print(f"  [midmonth-set] {len(_MIDMONTH_SET)} machines from "
                      f"{os.path.basename(getattr(_bc_cfg, 'MIDMONTH_BASELINE_PLAN', ''))} "
                      f"({getattr(_bc_cfg, 'MIDMONTH_SET_DAYS', 3)}d before {plan_start.date()}); "
                      f"set size min/med/max {_sz[0]}/{_sz[len(_sz)//2]}/{_sz[-1]}")
        _sd = {}
        if _MIDMONTH_SET:
            # N-DAY SET ACTIVE: do NOT pin a single carry-in SKU. Leaving machine_current_sku
            # empty is what lets the Phase-A day-1 "start" pick the best-need SKU from the
            # machine's own recent set (and be charged no changeover). Pinning one SKU here
            # would make `cur` non-empty and the set would never be consulted.
            pass
        elif int(getattr(plan_start, "day", 1) or 1) != 1:
            # Explicit mid-month carry-in file (machine -> SKU at the END of the last
            # actually-run day) takes priority; else derive from the plant 2-day replay.
            _mm = (os.environ.get("MIDMONTH_SEED_FILE", "")
                   or getattr(_bc_cfg, "MIDMONTH_BUILDING_SEED_FILE", "") or "")
            if _mm:
                _sd = _load_actual_seed(_mm)
                print(f"[midmonth-seed] {len(_sd)} machines from {os.path.basename(_mm)}")
            if not _sd:
                _sd = _derive_seed_from_plant_2day()
        if not _sd and not _MIDMONTH_SET:
            # (skipped when the N-day SET is active — a 1st-of-month static seed would pin
            #  machine_current_sku and the set would never get to choose)
            _sd = _load_actual_seed(_BLD_ACTUAL_SEED_FILE)
        _n_seed = _n_drop_allow = _n_drop_dem = 0
        for _m, _sku in _sd.items():
            _m = str(_m); _sku = str(_sku)
            # (a) certified on this machine (raw pre-inch-strip allowable) AND
            # (b) the SKU still has demand to build.
            if _sku not in _raw_machine_skus.get(_m, set()):
                _n_drop_allow += 1
                continue
            if demand_remaining.get(_sku, 0.0) <= 0:
                _n_drop_dem += 1
                continue
            _si = sku_inch.get(_sku, "")
            machine_current_sku[_m] = _sku
            _BLD_SEED_MACHINES.add(_m)
            _BLD_SEED_SKU[_m] = _sku                  # BLD_SEED_STICKY: remember the seed SKU
            if _si:
                _BLD_SEED_INCH[_m] = _si
            if _m in _S1_MACHINES:
                _seed_s1[_m] = _sku
                # Re-include the seed SKU (+ its inch-siblings) as carcass-feedable for this S1
                # machine even if the hist-lock inch-strip removed them.
                s1_sku_to_machines[_sku].add(_m)
                for _sib in _raw_machine_skus.get(_m, set()):
                    if sku_inch.get(_sib, "") == _si:
                        s1_sku_to_machines[_sib].add(_m)
            else:
                _seed_gt[_m] = _sku
                # Re-include the seed SKU (+ same-inch certified siblings) into the GT
                # eligibility that the hist-lock strip may have removed, so the machine can
                # actually build its seed inch (Phase-A continuation needs sku in machine_skus).
                machine_skus[_m].add(_sku); sku_machine_map[_sku].add(_m)
                if _si:
                    for _sib in _raw_machine_skus.get(_m, set()):
                        if sku_inch.get(_sib, "") == _si:
                            machine_skus[_m].add(_sib); sku_machine_map[_sib].add(_m)
            _n_seed += 1
        print(f"  [Rolling] BLD_ACTUAL_SEED: seeded {_n_seed} machines "
              f"({len(_seed_gt)} GT + {len(_seed_s1)} Stage-1) from actual plant production; "
              f"dropped {_n_drop_allow} not-allowable + {_n_drop_dem} no-demand")

    def _anchor_seed_inch(_m) -> str:
        """Inch the machine 'starts on' for initial anchoring (lock allocation +
        INIT_HYBRID). Start-free → its DOMINANT inch (Phase-5 file / hardcoded map),
        per the client's 'start with the most dominant inch'; else its Day-0 running
        SKU's inch."""
        if _BLD_START_FREE:
            return _MACHINE_DOMINANT_INCH.get(str(_m), "")
        return sku_inch.get(str(machine_current_sku.get(_m, "")), "")

    # ── G: Machine pools + cross-shift minute tracker ─────────────────────────
    # Pool: fixed 2–3 same-inch SKUs per machine, ordered by Day-1 urgency.
    # Replaced only when a pool SKU's demand_remaining hits 0.
    machine_pool: dict[str, list[str]] = _build_machine_pools(
        machine_skus=machine_skus,
        sku_inch=sku_inch,
        demand_dict=demand_dict,
        press_count=press_count,
        cure_ct_map=cure_ct_map,
        planning_days=planning_days,
    )
    print(f"  [Rolling] Pools: {sum(len(v) for v in machine_pool.values())} SKU slots "
          f"across {len(machine_pool)} machines")

    # machine_minutes_on_sku: cumulative minutes each machine has spent on its
    # current SKU without a CO.  Reset to 0 on CO; incremented per campaign.
    # Guards against micro-campaigns across shift boundaries (< MIN_CAMPAIGN_MINS).
    machine_minutes_on_sku: dict[str, float] = {m: 0.0 for m in machine_skus}

    # ── Client inch-rule state (persists across the whole horizon) ────────────
    # machine_anchor_inch: the inch of the machine's FIRST assignment — fixes its
    #   +/-_INCH_BAND_WIDTH band for the month (Rule 2).
    # machine_used_inches: every inch the machine has run — an inch it has left
    #   can never be re-used (Rule 1a).
    machine_anchor_inch: dict[str, str] = {}
    # Part 1 (LOCK_INCH_SET): each GT machine's locked inch-SET (mostly one). A machine may only
    # build/CO to an inch in its set; empty/absent → unconstrained (±2 band). Computed below.
    machine_locked_inches: dict[str, set] = {}
    # Lever B: lifetime (whole-month) count of escape diff-COs each fixed machine has spent.
    # Persists across the day loop (NOT reset per day, unlike machine_day_diff_co).
    fixed_escape_used: dict[str, int] = {}
    machine_used_inches: dict[str, set] = {}
    machine_left_skus:   dict[str, set] = {}   # SKU_NO_REVERT: SKUs each machine has built-then-left
    _oneway_switched:    set = set()           # ONEWAY_INCH_GENERAL: machines that already took their 1 inch-switch
    _buildable_skus:     set = (set().union(*machine_skus.values()) if machine_skus else set())  # SKUs with ≥1 allowable machine
    # machine_inch_now / machine_inch_since: the machine's CURRENT inch and the day
    # that inch campaign began — the 5-day-dwell clock (Rule: min 5 days per size).
    machine_inch_now:   dict[str, str] = {}
    machine_inch_since: dict[str, int] = {}
    # Last day each machine performed a diff-size CO (diff-CO amortization gate, DIFF_CO_GATE).
    machine_last_diff_co_day: dict[str, int] = {}
    # Distinct SKUs each building machine has produced TODAY (reset per day, seeded with
    # the carryover SKU) — enforces the max-4-SKUs-per-machine-per-day plant rule.
    machine_day_skus:   dict[str, set] = {}
    # Machines that have spent their one-time +3/-3 inch escape this month (experiment).
    machine_plus3_used: set = set()
    # INCH_STEP_DRIFT: signed cumulative inch drift per machine (−MAX..+MAX from its historical
    # anchor). Persists across the horizon; one-way once a direction is chosen. {} when OFF.
    machine_step_drift: dict[str, int] = {}
    # Stage-1 carcass machines are scheduled in Step 3b, not in
    # _assign_building_shift, so they need their own current-inch tracker.
    s1_current_inch: dict[str, str] = {}
    # Stage-1 single-inch lock (Rule 2, S1_SINGLE_INCH): {machine: its one month inch}.
    # Empty when the toggle is off (Step-3b then falls back to the band gate).
    s1_locked_inch: dict[str, str] = {}

    # Anchor the +/-2 band to the machine's REAL Day-0 inch whenever the plant
    # running state is known (machine_current_sku seeded above). Without this the
    # anchor is whatever the Day-1 greedy happens to pick first, which then locks
    # the machine for the whole month — the plant state is a far better anchor
    # (e.g. the plant runs a 17" machine that the greedy abandoned entirely).
    if _INCH_RULES_ENABLED and machine_current_sku:
        for _m0, _sku0 in machine_current_sku.items():
            _i0 = sku_inch.get(str(_sku0), "")
            if not _i0:
                continue
            machine_anchor_inch.setdefault(str(_m0), _i0)
            machine_used_inches.setdefault(str(_m0), set()).add(_i0)
            machine_inch_now.setdefault(str(_m0), _i0)
            machine_inch_since.setdefault(str(_m0), 1)     # Day-0 inch clock starts day 1
            if str(_m0) in _S1_MACHINES:
                s1_current_inch.setdefault(str(_m0), _i0)
        print(f"  [Rolling] Inch anchors seeded from Day-0 machine state: "
              f"{len(machine_anchor_inch)} machines")

    # ── Demand-optimal machine→inch pre-solve (Rules C + 2, default OFF) ───────
    # Assigns each machine one inch by matching capacity to per-inch demand, ignoring
    # running-machine state (the TBMStage1/2 production-event tables are NOT used).
    # AUTHORITATIVE: overrides any Day-0 seed above so running state never influences
    # anchors under these toggles. See _optimal_inch_assignment.
    if _INCH_ANCHOR_OPT:
        _gt_machines = list(machine_skus.keys())
        _gt_skus = set().union(*machine_skus.values()) if machine_skus else set()
        _inch_dem_gt: dict[str, float] = defaultdict(float)
        for _s in _gt_skus:
            _i = sku_inch.get(_s, "")
            if _i:
                _inch_dem_gt[_i] += demand_dict.get(_s, 0.0)
        _cap = {_m: _bld_qty_per_shift(_m) * 3 * planning_days for _m in _gt_machines}
        _elig = {_m: {sku_inch.get(_s, "") for _s in machine_skus[_m] if sku_inch.get(_s, "")}
                 for _m in _gt_machines}
        # Curing-aware target: cap each inch by what curing can absorb (mould-feasible presses)
        _inch_skus_gt: dict[str, list] = defaultdict(list)
        for _s in _gt_skus:
            _i = sku_inch.get(_s, "")
            if _i:
                _inch_skus_gt[_i].append(_s)
        _ceil = _curing_inch_ceiling(_inch_skus_gt, _sku_moulds, cure_ct_map,
                                     planning_days, len(press_state))
        _tgt = ({_i: (min(_d, _ceil[_i]) if _i in _ceil else _d)
                 for _i, _d in _inch_dem_gt.items()} if _ANCHOR_CURE_CAP else None)
        _asg = _optimal_inch_assignment(_gt_machines, _elig, _cap, _inch_dem_gt, target=_tgt)
        for _m, _i in _asg.items():
            machine_anchor_inch[_m] = _i          # OVERRIDE (authoritative)
            machine_inch_now[_m]    = _i
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_i)
        _INCH_OPT_DBG[0] = len(_asg)
        print(f"  [Rolling] INCH_ANCHOR_OPT: demand-optimal anchors for {len(_asg)} GT machines")
        if os.environ.get("ANCHOR_DEBUG"):
            from collections import Counter as _Ctr
            _cnt = _Ctr(_asg.values())
            _capby: dict[str, float] = defaultdict(float)
            for _m, _i in _asg.items():
                _capby[_i] += _cap.get(_m, 0.0)
            print("  [ANCHOR_DEBUG] inch | demand | ceiling | target | #mach | anchored_cap | cap/target")
            for _i in sorted(set(_inch_dem_gt) | set(_cnt), key=lambda z: -_inch_dem_gt.get(z, 0)):
                _d = _inch_dem_gt.get(_i, 0.0); _c = _ceil.get(_i, float('inf'))
                _t = (_tgt or _inch_dem_gt).get(_i, 0.0)
                _cp = _capby.get(_i, 0.0)
                print(f"    {_i:>3} | {_d:8.0f} | {_c if _c!=float('inf') else 0:8.0f} | {_t:8.0f} "
                      f"| {_cnt.get(_i,0):3d} | {_cp:9.0f} | {_cp/_t if _t else 0:5.2f}")
            print("  [ANCHOR_DEBUG] per-machine: anchor <- eligible inches (breadth)")
            for _m in sorted(_asg, key=lambda z: (_asg[z], z)):
                _ei = sorted(_elig.get(_m, set()), key=lambda z: int(z) if str(z).isdigit() else 99)
                print(f"    {_m}: anchor={_asg[_m]:>3}  elig={_ei}")

    # ── Part 1: single-inch-majority locked inch-SET (LOCK_INCH_SET) ──────────────
    # Mathematical minimum-assignment covering: most GT machines get ONE inch, a few 2-3 (only
    # where an inch's demand can't be covered by whole machines). HYBRID-seeded from the plant
    # running-machine sizes (a running machine whose real size has demand is PINNED). Fully
    # demand-dynamic (re-solved per month from demand_dict). Sets machine_locked_inches.
    if _LOCK_INCH_SET:
        _lg_machines = list(machine_skus.keys())
        _lg_skus = set().union(*machine_skus.values()) if machine_skus else set()
        _dem_by_inch: dict[str, float] = defaultdict(float)
        _skus_by_inch: dict[str, list] = defaultdict(list)
        for _s in _lg_skus:
            _i = sku_inch.get(_s, "")
            if _i:
                _dem_by_inch[_i] += demand_dict.get(_s, 0.0)
                _skus_by_inch[_i].append(_s)
        _lcap = {_m: _bld_qty_per_shift(_m) * 3 * planning_days for _m in _lg_machines}
        _lelig = {_m: {sku_inch.get(_s, "") for _s in machine_skus[_m] if sku_inch.get(_s, "")}
                  for _m in _lg_machines}
        _lceil = _curing_inch_ceiling(_skus_by_inch, _sku_moulds, cure_ct_map,
                                      planning_days, len(press_state))
        _ltgt = {_i: (min(_d, _lceil[_i]) if _i in _lceil else _d)
                 for _i, _d in _dem_by_inch.items()}
        # HYBRID pins: a running machine whose real running inch has demand ≥ floor keeps it.
        _forced: dict[str, str] = {}
        for _m in _lg_machines:
            _ri = _anchor_seed_inch(_m)
            if _ri and _ri in _lelig.get(_m, set()) \
                    and _dem_by_inch.get(_ri, 0.0) >= _LOCK_PIN_DEMAND_FRAC * _lcap.get(_m, 0.0):
                _forced[_m] = _ri
        machine_locked_inches = _min_assignment_inch_sets(
            _lg_machines, _lelig, _lcap, _ltgt, _forced)
        # Anchor / inch clock = the machine's primary (pinned, else highest-demand) locked inch.
        for _m, _iset in machine_locked_inches.items():
            _primary = (_forced.get(_m)
                        if _forced.get(_m) in _iset
                        else max(_iset, key=lambda _i: (_dem_by_inch.get(_i, 0.0), _i)))
            machine_anchor_inch[_m] = _primary
            machine_inch_now[_m]    = _primary
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_primary)
        _multi = sum(1 for _s in machine_locked_inches.values() if len(_s) > 1)
        print(f"  [Rolling] LOCK_INCH_SET: {len(machine_locked_inches)} GT machines "
              f"({len(machine_locked_inches)-_multi} single-inch, {_multi} multi-inch, "
              f"{len(_forced)} pinned to real running size)")
        if os.environ.get("LOCK_DEBUG"):
            for _m in sorted(machine_locked_inches, key=lambda z: int(z) if str(z).isdigit() else 0):
                _pin = " (pinned)" if _m in _forced else ""
                _ss = ", ".join(sorted(machine_locked_inches[_m],
                                       key=lambda z: int(z) if str(z).isdigit() else 99))
                print(f"    {_m}: {{{_ss}}}{_pin}")
            _under = {_i: _ltgt[_i] - sum(_lcap[_m] for _m in machine_locked_inches
                                          if _i in machine_locked_inches[_m])
                      for _i in _ltgt}
            print("    [coverage] short inches:",
                  {_i: round(_v) for _i, _v in _under.items() if _v > 1})

    # ── Part A: dynamic hybrid initial allocation (INIT_HYBRID) ──────────────────
    # Anchor = real running size where it has demand this month (pinned), else demand-optimal.
    # Soft (sets anchor + ±2 band only); JIT flexes within it. Fully demand-dynamic.
    if _INIT_ALLOC_HYBRID:
        _hm = list(machine_skus.keys())
        _hskus = set().union(*machine_skus.values()) if machine_skus else set()
        _hdem: dict[str, float] = defaultdict(float)
        _hby: dict[str, list] = defaultdict(list)
        for _s in _hskus:
            _i = sku_inch.get(_s, "")
            if _i:
                _hdem[_i] += demand_dict.get(_s, 0.0)
                _hby[_i].append(_s)
        _hcap = {_m: _bld_qty_per_shift(_m) * 3 * planning_days for _m in _hm}
        _helig = {_m: {sku_inch.get(_s, "") for _s in machine_skus[_m] if sku_inch.get(_s, "")}
                  for _m in _hm}
        _hceil = _curing_inch_ceiling(_hby, _sku_moulds, cure_ct_map, planning_days, len(press_state))
        _htgt = {_i: (min(_d, _hceil[_i]) if _i in _hceil else _d) for _i, _d in _hdem.items()}
        # Demand-optimal, May-preferring assignment: process most-constrained machines first;
        # keep a machine's real running inch ONLY while that inch still has UNMET target capacity
        # (i.e. it is not already over-provisioned this month) — otherwise re-anchor it to the
        # neediest reachable inch. This fires real reassignments when May sizes do not fit the
        # month's demand (e.g. July 13″-heavy), which is the dynamic re-allocation.
        def _num(i):
            try: return int(i)
            except Exception: return 99
        _hresid = {_i: float(_t) for _i, _t in _htgt.items()}
        _horder = sorted(_hm, key=lambda _m: (len(_helig.get(_m, ()) or ()), -_hcap.get(_m, 0.0), _m))
        _hasg: dict[str, str] = {}
        _hpin: list = []
        _hreass: list = []
        for _m in _horder:
            _ri = _anchor_seed_inch(_m)
            _el = _helig.get(_m, set())
            if _ri and _ri in _el and _hresid.get(_ri, 0.0) > 0:
                _chosen = _ri; _hpin.append(_m)         # May inch still needed → keep it
            else:
                _opts = [_i for _i in _el if _i in _hresid]
                if _opts:
                    _chosen = max(_opts, key=lambda _i: (_hresid.get(_i, 0.0),
                                                         _hdem.get(_i, 0.0), -_num(_i)))
                elif _el:
                    _chosen = sorted(_el, key=_num)[0]
                else:
                    continue
                (_hpin if _chosen == _ri else _hreass).append(_m)
            _hasg[_m] = _chosen
            _hresid[_chosen] = _hresid.get(_chosen, 0.0) - _hcap.get(_m, 0.0)
        for _m, _i in _hasg.items():
            machine_anchor_inch[_m] = _i
            machine_inch_now[_m]    = _i
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_i)
        print(f"  [Rolling] INIT_HYBRID: {len(_hasg)} GT anchors "
              f"({len(_hpin)} kept real running size, {len(_hreass)} re-anchored to demand)")
        if os.environ.get("LOCK_DEBUG"):
            for _m in sorted(_hasg, key=lambda z: int(z) if str(z).isdigit() else 0):
                _ri = sku_inch.get(str(machine_current_sku.get(_m, "")), "?")
                _tag = "pinned" if _m in _hpin else f"reassigned (real {_ri}\")"
                print(f"    {_m}: anchor={_hasg[_m]}\"  [{_tag}]")

    # ── Part 2: per-group inch policy — HARD single-inch on BJ + UNISTAGE(US) ─────
    # Lock each BJ / UNISTAGE machine to its anchor inch → the _inch_gate lock check forces
    # ZERO different-size CO on those groups (the plant pattern). VMI stays on JIT (tightened
    # separately); Stage-2 stays flexible. Requires an anchor (from INIT_HYBRID or the seed).
    if _GROUP_INCH_POLICY:
        # Co-planning consistency: use the SAME BJ/US locks the curing scheduler was told about
        # (_coplan_lock, computed pre-curing from demand-optimal allocation). Curing supply-matched
        # its draw to exactly this allocation, so building must lock to it → the two agree and the
        # locked machines stay fed. Fall back to the demand-anchor if the co-plan didn't run.
        _bjus = sorted(m for m in machine_skus
                       if _MACHINE_GROUP.get(str(m), "") == "UNISTAGE")   # HARD-lock US only
        _bj_asg = dict(_coplan_lock) if _coplan_lock else {}
        _locked_grp = 0
        for _m in _bjus:
            _li = _bj_asg.get(str(_m)) or machine_anchor_inch.get(str(_m)) or machine_inch_now.get(str(_m))
            if _li:
                machine_locked_inches[str(_m)] = {_li}
                machine_anchor_inch[str(_m)] = _li      # keep anchor/band consistent with the lock
                machine_inch_now[str(_m)]    = _li
                machine_used_inches.setdefault(str(_m), set()).add(_li)
                _locked_grp += 1
        print(f"  [Rolling] GROUP_INCH_POLICY: hard single-inch on {_locked_grp} US machines (0 ongoing "
              f"diff-CO); BJ+VMI tight JIT margin={_VMI_JIT_MARGIN}/budget={_VMI_MAX_DIFF_CO_PER_DAY}; "
              f"Stage-2 flexible")
        if os.environ.get("LOCK_DEBUG"):
            print(f"    BJ/US locks: {dict(sorted((m, next(iter(machine_locked_inches[m]))) for m in _bjus if m in machine_locked_inches))}")

    if _STAGE1_SINGLE_INCH:
        _s1_machines = sorted(_S1_MACHINES)
        _s1_msku: dict[str, set] = defaultdict(set)          # Stage-1 machine → feedable SKUs
        for _sku, _ms in s1_sku_to_machines.items():
            for _m in _ms:
                _s1_msku[_m].add(_sku)
        _inch_dem_s1: dict[str, float] = defaultdict(float)  # Stage-2 carcass demand per inch
        for _sku in s1_sku_to_machines:
            _i = sku_inch.get(_sku, "")
            if _i:
                _inch_dem_s1[_i] += demand_dict.get(_sku, 0.0)
        _cap_s1 = {_m: _bld_qty_per_shift(_m) * 3 * planning_days for _m in _s1_machines}
        _elig_s1 = {_m: {sku_inch.get(_s, "") for _s in _s1_msku.get(_m, ()) if sku_inch.get(_s, "")}
                    for _m in _s1_machines}
        # Curing-aware target: Stage-2 carcass demand per inch capped by curing throughput
        _inch_skus_s1: dict[str, list] = defaultdict(list)
        for _sku in s1_sku_to_machines:
            _i = sku_inch.get(_sku, "")
            if _i:
                _inch_skus_s1[_i].append(_sku)
        _ceil_s1 = _curing_inch_ceiling(_inch_skus_s1, _sku_moulds, cure_ct_map,
                                        planning_days, len(press_state))
        _tgt_s1 = ({_i: (min(_d, _ceil_s1[_i]) if _i in _ceil_s1 else _d)
                    for _i, _d in _inch_dem_s1.items()} if _ANCHOR_CURE_CAP else None)
        if _S1_BALANCED_INCH:
            _asg_s1 = _balanced_inch_assignment(_s1_machines, _elig_s1, _cap_s1,
                                                _tgt_s1 if _tgt_s1 is not None else _inch_dem_s1)
        else:
            _asg_s1 = _optimal_inch_assignment(_s1_machines, _elig_s1, _cap_s1, _inch_dem_s1,
                                               target=_tgt_s1)
        for _m, _i in _asg_s1.items():
            s1_locked_inch[_m]      = _i
            machine_anchor_inch[_m] = _i
            s1_current_inch[_m]     = _i
            machine_inch_now[_m]    = _i
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_i)
        _INCH_OPT_DBG[1] = len(_asg_s1)
        print(f"  [Rolling] S1_SINGLE_INCH: locked {len(_asg_s1)} Stage-1 machines to one inch "
              f"({dict(sorted(_asg_s1.items()))})")

    # ── Historical inch-LOCK (INCH_HIST_LOCK): AUTHORITATIVE per-machine allowed-inch
    # sets from the 4-month plant report. Runs LAST so it overrides the INIT_HYBRID /
    # GROUP demand anchors for the GT machines with the plant's real history: FIXED
    # machines (single historical inch) → locked to it (0 diff-size CO ever); FLEXIBLE
    # machines → their ranked historical inches only. machine_locked_inches gates every
    # Phase-B CO candidate via _inch_gate; the ±2 band is off (see _inch_ok). ──────────
    if _INCH_HIST_LOCK_ENABLED and _MACHINE_ALLOWED_INCHES:
        _hl_fixed = _hl_flex = 0
        for _m in machine_skus:                      # GT machines (VMI/BJ/UNI/Stage-2)
            _al = _MACHINE_ALLOWED_INCHES.get(str(_m))
            if not _al:
                continue
            _elig = {sku_inch.get(_s, "") for _s in machine_skus[_m]}
            if _ONEWAY_INCH_GENERAL:
                # GENERAL one-way rule: start LOCKED to the dominant inch only; the day loop flips the
                # lock to the neediest other allowable inch once the dominant inch's demand is done.
                _set = {_al[0]} if _al[0] in _elig else ({next(iter(_elig))} if _elig else {_al[0]})
            elif ((_FIXED_ESCAPE_ENABLED and str(_m) in _FIXED_MACHS_HIST)
                    or _ONEWAY_INCH_ENABLED):
                # Lever B / ONE-WAY: permit ALL DB-allowable inches so the WHICH-gate lets the
                # transition through; the one-way (no-revisit) + ≤2-jump discipline is enforced
                # at runtime in _inch_gate (STAGE-2 exempt there). Anchor stays the dominant inch.
                _set = set(_elig) or {_al[0]}
            else:
                _set = {i for i in _al if i in _elig} or {_al[0]}
            machine_locked_inches[str(_m)] = set(_set)
            _prim = _al[0] if _al[0] in _elig else next(iter(_set))   # anchor = historical dominant
            machine_anchor_inch[str(_m)] = _prim
            machine_inch_now[str(_m)]    = _prim
            machine_inch_since[str(_m)]  = 1
            machine_used_inches.setdefault(str(_m), set()).add(_prim)
            if len(_al) <= 1:
                _hl_fixed += 1
            else:
                _hl_flex += 1
        print(f"  [Rolling] INCH_HIST_LOCK: {_hl_fixed} GT machines FIXED to one historical "
              f"inch (0 diff-CO), {_hl_flex} FLEXIBLE to their ranked historical inches; "
              f"±2 band discontinued")

    # ── BLD_ACTUAL_SEED: make each seeded machine's SEED inch its DOMINANT/anchor inch ──
    # Runs LAST (after GROUP_INCH_POLICY / STAGE1_SINGLE_INCH / INCH_HIST_LOCK) so it is
    # AUTHORITATIVE for the seeded machines and cannot be clobbered. The seed inch becomes the
    # anchor + first allowed inch; the machine's historical/certified inches are KEPT in the
    # locked set so the deficit-done + diff-CO machinery can move it to another needy inch once
    # the seed inch's demand is done (exactly the current flexible behaviour). Non-seeded
    # machines and the OFF path are untouched (bit-for-bit).
    if _BLD_ACTUAL_SEED and (_seed_gt or _seed_s1):
        for _m, _sku in _seed_gt.items():
            _si = _BLD_SEED_INCH.get(_m) or sku_inch.get(_sku, "")
            if not _si:
                continue
            _hist = set(_MACHINE_ALLOWED_INCHES.get(_m, []) or [])   # historical/certified inches
            _prev = set(machine_locked_inches.get(_m, set()))         # whatever earlier blocks set
            _elig = {sku_inch.get(_s, "") for _s in machine_skus.get(_m, set())}
            _elig.discard("")
            machine_locked_inches[_m] = ({_si} | (_hist & _elig) | (_prev & _elig)) or {_si}
            machine_anchor_inch[_m] = _si
            machine_inch_now[_m]    = _si
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_si)
        for _m, _sku in _seed_s1.items():
            _si = _BLD_SEED_INCH.get(_m) or sku_inch.get(_sku, "")
            if not _si:
                continue
            s1_locked_inch[_m]      = _si       # override the demand-optimal S1 single-inch with the real seed
            s1_current_inch[_m]     = _si
            machine_anchor_inch[_m] = _si
            machine_inch_now[_m]    = _si
            machine_inch_since[_m]  = 1
            machine_used_inches.setdefault(_m, set()).add(_si)
        print(f"  [Rolling] BLD_ACTUAL_SEED: seed inch made DOMINANT for "
              f"{len(_seed_gt)} GT + {len(_seed_s1)} Stage-1 machines (historical inches kept "
              f"in the locked set for later diff-CO)")

    # ══════════════════════════════════════════════════════════════════════════
    # Data accumulators (matching output sheet formats)
    # ══════════════════════════════════════════════════════════════════════════
    bld_shift_rows:  list[dict] = []   # building Shift Schedule rows (+ CO sentinels)
    bld_co_events:   list[dict] = []   # building machine CO events
    cure_shift_rows: list[dict] = []   # curing Shift Schedule rows
    cure_co_events:  list[dict] = []   # curing press CO events (Planned + Dynamic)
    mould_clean_events: list[dict] = []  # curing press mould-clean events (each = 8h/480min)
    press_stats:     dict = defaultdict(lambda: {
        "running_mins": 0.0, "co_mins": 0.0, "clean_mins": 0.0,
        "skus": set(), "cycles": 0, "units": 0,
    })
    press_sku_stats: dict = defaultdict(lambda: {"cycles": 0, "units": 0, "mins_used": 0.0})
    daily_cured:     dict[str, int] = defaultdict(int)
    sku_cured:       dict[str, int] = defaultdict(int)
    build_by_shift_sku: dict = {}      # {(date,shift): {sku: qty}} for GT diagnostic
    last_build_day:  dict[str, int] = {}
    daily_summary:   list[dict] = []
    writeoff_total   = 0.0
    SHIFTS           = ["A", "B", "C"]

    # ── Dynamic CO infrastructure ─────────────────────────────────────────────
    # Tracks instant COs triggered when a press finishes its SKU demand mid-plan.
    # CO starts in the SAME SHIFT demand is met (remaining time = CHANGEOVER).
    # Next shift: PRODUCTION for new SKU (no MOULD_CLEAN).
    # Key = press, value = (global_shift_idx_of_co_start, new_sku)
    # global_shift_idx = (day − 1) × 3 + shift_idx  (A=0, B=1, C=2)
    dynamic_co_tracker: dict[str, tuple[int, str]] = {}
    # CURING_ADAPT_CO: consecutive shifts each press received 0 GT while RUNNING with
    # unmet demand (a starvation run). Reset to 0 whenever it cures >0 or COs. At
    # ≥ _CURING_STARV_SWITCH_SHIFTS the press's SKU is building-limited → switch it.
    _consec_zero_gt: dict[str, int] = defaultdict(int)

    # Track daily CO counts (pre-planned + dynamic combined) to enforce the cap.
    daily_co_count: dict[int, int] = defaultdict(int)
    for _dco_day, _dco_evs in co_by_day.items():
        daily_co_count[_dco_day] += len(_dco_evs)

    # sku_campaign_tier: SKU -> best (min) campaign-list position it received
    # this most recent day (0=primary, 1=secondary via first CO, 2+=tertiary+).
    # Replaced (not merged) each day — see Step 3 / end-of-day finalize below.
    # Only accumulated when _DYNAMIC_CO_PLANNER_ENABLED (zero cost otherwise).
    sku_campaign_tier: dict[str, int] = {}

    if os.environ.get("ROLLCO_STAGE2_CHECK"):
        _stage2_scheduler = COScheduler()
        _stage2_allow = cetl.load_curing_allowable()
        _rolling_today = _rolling_horizon_co_call(
            day=1, planning_days=planning_days,
            press_state=press_state, press_count=press_count,
            demand_remaining=demand_remaining, demand_dict=demand_dict,
            priority_score_map=priority_score_map,
            df_allowable=_stage2_allow, ct_map=cc_result["ct_map"],
            dynamic_co_tracker=dynamic_co_tracker,
            scheduler=_stage2_scheduler, max_co_per_day=MAX_CHANGEOVERS_PER_DAY,
        )
        _legacy_today = sorted(
            (ev["press"], ev["old_sku"], ev["new_sku"])
            for ev in co_events if int(ev["day"]) == 1
        )
        print(f"\n  [Stage2Check] Rolling Day-1: {len(_rolling_today)} events")
        print(f"  [Stage2Check] Legacy  Day-1: {len(_legacy_today)} events")
        _rolling_set = set(_rolling_today)
        _legacy_set  = set(_legacy_today)
        print(f"  [Stage2Check] Match: {_rolling_set == _legacy_set}")
        print(f"  [Stage2Check] Rolling-only: {sorted(_rolling_set - _legacy_set)}")
        print(f"  [Stage2Check] Legacy-only : {sorted(_legacy_set - _rolling_set)}")
        import sys; sys.exit(0)

    # Phase 2b — press → allowable demand-SKUs (sorted), the retarget candidate
    # universe. Built once, only when the optimisation is on (a DB read otherwise
    # skipped). Restricted to demand SKUs so retarget never chases a zero-demand SKU.
    press_allow_skus: dict[str, list] = {}
    if _mould_gate and (_mould_opt or _CO_SCORER_ENABLED):
        try:
            _dfca = cetl.load_curing_allowable()
            _dem_skus = set(demand_dict.keys())
            _tmp: dict[str, set] = {}
            for _, _row in _dfca.iterrows():
                _sk = str(_row["SKUCode"]).strip()
                if _sk in _dem_skus:
                    for _mp in (_row.get("Machines", []) or []):
                        _tmp.setdefault(str(_mp), set()).add(_sk)
            press_allow_skus = {_p: sorted(_ss) for _p, _ss in _tmp.items()}
        except Exception as _e:
            print(f"  [Rolling] curing-allowable load for retarget FAILED ({_e}); "
                  f"retarget disabled")
            press_allow_skus = {}

    # ── Phase 4: Global mould optimiser (experiment, MOULD_GLOBAL_OPT) ─────────
    # Reverse of press_allow_skus: {sku: [presses eligible to cure it]} (sorted).
    sku_allow_presses: dict[str, list] = {}
    for _p_al, _ss_al in press_allow_skus.items():
        for _s_al in _ss_al:
            sku_allow_presses.setdefault(_s_al, []).append(_p_al)
    for _s_al in sku_allow_presses:
        sku_allow_presses[_s_al].sort()
    mould_global_stats = {"add": 0, "lib": 0}

    def _amort_value(sku: str, hleft: int) -> float:
        """Amortised marginal value of +1 press on `sku`: how much it could cure over
        the remaining horizon, capped by remaining demand. Ranks targets and gates
        whether a full_evict is worth the production it sacrifices."""
        rem = demand_remaining.get(sku, 0.0)
        if rem <= 0:
            return 0.0
        draw = _cure_qty_per_shift(cure_ct_map.get(sku, DEFAULT_CURING_CT))
        return min(rem, draw * 3 * max(1, hleft))

    def _global_mould_boost(day: int, today_cos: list) -> list:
        """Move scarce moulds toward the most-under-served SKUs, once per day, before
        today's COs drive the sim. Returns today_cos + extra proactive COs (their
        moulds already reserved via _try_mount, daily_co_count already bumped).
        No-op unless _MOULD_GLOBAL_OPT_ENABLED. Respects the daily CO cap. Two modes:
          ro_only    — only sacrifice/evict presses whose current SKU demand is DONE.
          full_evict — also evict a RUNNING press when the target's amortised value
                       strictly exceeds the current SKU's.
        """
        if not (_MOULD_GLOBAL_OPT_ENABLED and _mould_gate):
            return today_cos
        hleft = _working_days_left(day) - 1   # working days after today (holiday-aware); OFF by default
        slots = MAX_CHANGEOVERS_PER_DAY - daily_co_count.get(day, 0)
        if slots <= 0:
            return today_cos
        full_evict = (_MOULD_GLOBAL_OPT_MODE == "full_evict")
        busy = {p for p, _, _ in today_cos}          # presses already CO'ing today
        added: list = []

        # Under-served scarce targets: real remaining demand the CURRENT presses cannot
        # cure over the horizon, biggest-gap first (scarcest-mould tiebreak → 15"/13").
        targets = []
        for T in demand_remaining:
            rem = demand_remaining.get(T, 0.0)
            if rem <= MIN_CAMPAIGN_UNITS:
                continue
            if len(_sku_moulds.get(T, ())) < 2:
                continue
            draw = _cure_qty_per_shift(cure_ct_map.get(T, DEFAULT_CURING_CT))
            cur_cap = press_count.get(T, 0) * draw * 3 * max(1, hleft)
            if rem <= cur_cap:                        # current presses already suffice
                continue
            targets.append(T)
        targets.sort(key=lambda s: (-demand_remaining.get(s, 0.0),
                                    len(_sku_moulds.get(s, ())), s))

        for T in targets:
            if slots <= 0:
                break
            t_val = _amort_value(T, hleft)
            if t_val <= 0:
                continue
            # eligible presses for T we may sacrifice, least-sacrifice first
            cands = []
            for P in sku_allow_presses.get(T, ()):
                if P in busy:
                    continue
                st = press_state.get(P)
                if not st or st.get("status") != "RUNNING":
                    continue
                cur = st.get("sku")
                if not cur or cur == T:
                    continue
                cur_rem = demand_remaining.get(cur, 0.0)
                if cur_rem > 0:
                    if not full_evict:                # ro_only: never evict a productive press
                        continue
                    c_val = _amort_value(cur, hleft)
                    if t_val <= c_val:                # not worth the sacrifice
                        continue
                    cands.append((c_val, str(P), P))
                else:
                    cands.append((0.0, str(P), P))    # demand-done → free to reassign
            cands.sort()

            # (1) DIRECT ADD — a sacrificeable press that can already mount 2 T-moulds.
            done = False
            for _sac, _ps, P in cands:
                if _n_free_for(T, P) >= 2 and _try_mount(P, T, defer_free=True):
                    added.append((P, press_state[P]["sku"], T))
                    busy.add(P); slots -= 1
                    daily_co_count[day] = daily_co_count.get(day, 0) + 1
                    mould_global_stats["add"] += 1
                    done = True
                    break
            if done or slots <= 0:
                continue
            if not cands:                             # nobody would use a freed mould → skip
                continue

            # (2) LIBERATION — T's moulds are stuck. Free ONE scarce T-mould from a
            # sacrificeable holder by CO'ing it to a needy SKU that does NOT reuse that
            # mould, so a later day can mount the freed mould on T.
            for X in sorted(_sku_moulds.get(T, ())):
                H = _mould_owner.get(X)
                if not H or H in busy:
                    continue
                st = press_state.get(H)
                if not st or st.get("status") != "RUNNING":
                    continue
                hcur = st.get("sku")
                if not hcur or hcur == T:             # already serving T — leave it
                    continue
                hrem = demand_remaining.get(hcur, 0.0)
                if hrem > 0 and not (full_evict and t_val > _amort_value(hcur, hleft)):
                    continue                          # holder still productive / not worth it
                # neediest retarget SKU for H that does NOT need mould X (so X frees)
                HT = None; HT_key = None
                for s in press_allow_skus.get(H, ()):
                    if s == hcur or demand_remaining.get(s, 0.0) <= 0:
                        continue
                    if X in _sku_moulds.get(s, ()):   # would keep X on H
                        continue
                    if _n_free_for(s, H) < 2:
                        continue
                    key = (demand_remaining.get(s, 0.0), s)
                    if HT_key is None or key > HT_key:
                        HT_key = key; HT = s
                if HT is None or not _try_mount(H, HT, defer_free=True):
                    continue
                added.append((H, press_state[H]["sku"], HT))
                busy.add(H); slots -= 1
                daily_co_count[day] = daily_co_count.get(day, 0) + 1
                mould_global_stats["lib"] += 1
                break                                 # one liberation per target per day
        return today_cos + added

    # ── Phase 3: Unified CO scorer ────────────────────────────────────────────
    # counters (provenance of every committed CO + why some were blocked)
    co_scorer_stats = {"planned": 0, "pullfwd": 0, "dynamic": 0, "retarget": 0,
                       "idle": 0, "cancelled": 0, "build_blocked": 0, "forced": 0}
    _CO_COST_UNITS   = float(os.environ.get("CO_COST_UNITS", "0"))  # 0 = cost folded into shift-draw floor
    _DEFAULT_BLD_CT  = 120.0

    def _bld_free_min_shift():
        """Conservative per-machine spare building minutes THIS shift: SHIFT_MINS minus
        the minutes each GT-producing machine already owes its currently-RUNNING SKUs
        (each RUNNING SKU's per-shift draw spread evenly over its eligible machines).
        Returns {machine: free_min}. Only clearly-free minutes count (never negative)."""
        _committed: dict[str, float] = defaultdict(float)
        for _pr, _st in press_state.items():
            if _st.get("status") != "RUNNING":
                continue
            _s = _st["sku"]
            if demand_remaining.get(_s, 0.0) <= 0:
                continue
            _ms = sku_machine_map.get(_s)
            if not _ms:
                continue
            _draw = _cure_qty_per_shift(cure_ct_map.get(_s, DEFAULT_CURING_CT))
            _per  = _draw / len(_ms)
            for _m in _ms:
                _committed[_m] += _per * (_bld_ct_sec(_m, _s) / 60.0)
        return {str(_m): max(0.0, float(SHIFT_MINS) - _committed.get(str(_m), 0.0))
                for _m in machine_skus}

    def _bld_capacity(sku: str, bld_free: dict) -> float:
        """Units/shift of `sku` GT that the currently-spare building machines could add."""
        _ms = sku_machine_map.get(sku)
        if not _ms:
            return 0.0
        return sum(bld_free.get(str(_m), 0.0) / (_bld_ct_sec(_m, sku) / 60.0)
                   for _m in _ms)

    def _supply_ok(sku: str, bld_free: dict) -> bool:
        """P1 reactive building-supply test (only consulted when _REACTIVE_CO is ON).
        True iff a curing CO to `sku` can actually be FED this shift — either curable GT
        is already banked (≥1 shift of draw) OR a building machine can be reserved to
        produce ≥1 shift of draw (same-inch/flex/Stage-1, via _bld_capacity). Prevents a
        reactive CO to a SKU whose GT never arrives."""
        _draw = _cure_qty_per_shift(cure_ct_map.get(sku, DEFAULT_CURING_CT))
        if _draw <= 0:
            return True
        if gt_inventory.get(sku, 0.0) >= _draw:        # GT already in the pool
            return True
        return _bld_capacity(sku, bld_free) >= _draw   # a building machine reservable

    def _bld_commit(sku: str, units: float, bld_free: dict) -> None:
        """Live-decrement the shared building minutes when a CO to `sku` is committed."""
        _ms = sku_machine_map.get(sku)
        if not _ms:
            return
        _per = units / len(_ms)
        for _m in _ms:
            _ms_key = str(_m)
            bld_free[_ms_key] = max(0.0, bld_free.get(_ms_key, 0.0)
                                    - _per * (_bld_ct_sec(_ms_key, sku) / 60.0))

    def _co_utility(press: str, target: str, horizon_left: int, bld_cap: float) -> float:
        """One utility for 'press → target', in units. Higher = more worth a changeover.
        expected extra cured = min(per-press residual demand load, what the press can
        physically cure of target over the horizon, what building can actually feed),
        minus the changeover cost (one shift of the target's own production)."""
        _rem = demand_remaining.get(target, 0.0)
        if _rem <= 0:
            return -1.0
        _npr  = max(1, press_count.get(target, 0))
        _draw = _cure_qty_per_shift(cure_ct_map.get(target, DEFAULT_CURING_CT))
        _residual_load = _rem / _npr
        _horizon_cure  = _draw * 3 * max(0, horizon_left)     # this press's cure over horizon
        _extra = min(_residual_load, _horizon_cure, bld_cap)
        _co_cost = _draw + _CO_COST_UNITS                     # 1 lost shift + optional constant
        return _extra - _co_cost

    def _best_alt(press, bld_free, horizon_left, exclude=None, check_build=True):
        """Highest-utility allowable target for `press` with 2 free moulds (and, if
        check_build, enough building feed). Deterministic. Returns SKU or None."""
        _cur = press_state.get(press, {}).get("sku")
        _best = None
        _best_key = None
        for s in press_allow_skus.get(press, ()):          # pre-sorted
            if s == _cur or s == exclude:
                continue
            if demand_remaining.get(s, 0.0) <= 0:
                continue
            if _n_free_for(s, press) < 2:
                continue
            _cap = _bld_capacity(s, bld_free)
            _draw = _cure_qty_per_shift(cure_ct_map.get(s, DEFAULT_CURING_CT))
            if check_build and _cap < _draw:
                continue
            _u = _co_utility(press, s, horizon_left, _cap)
            if _u <= 0:                                     # only worthwhile changeovers
                continue
            _key = (_u, s)                                  # util desc, SKU tiebreak
            if _best_key is None or _key > _best_key:
                _best_key = _key
                _best = s
        return _best

    def _solve_day_cos(day, today_planned):
        """Global CO solve for one day. Returns the final list of committed COs
        [(press, old_sku, new_sku)], reserving moulds via _try_mount(defer_free) and
        pruning any pulled-forward COs out of tomorrow's co_by_day. Two modes:
        ADDITIVE (keep planned COs, add pull-forward + idle-fill) and FULL_REOPT
        (global utility scoring; planned COs may be cancelled/replaced)."""
        horizon_left     = _working_days_left(day)   # holiday-aware urgency horizon
        bld_free         = _bld_free_min_shift()
        tomorrow_planned = list(co_by_day.get(day + 1, []))
        planned_tom      = {p: ns for (p, _o, ns) in tomorrow_planned}
        committed: dict[str, tuple] = {}     # press -> (target, provenance)
        pulled: list[str] = []
        slots = [MAX_CHANGEOVERS_PER_DAY]

        def _commit(press, target, prov, check_build=True):
            if slots[0] <= 0 or press in committed:
                return False
            if _n_free_for(target, press) < 2:
                return False
            _draw = _cure_qty_per_shift(cure_ct_map.get(target, DEFAULT_CURING_CT))
            if check_build and _bld_capacity(target, bld_free) < _draw:
                co_scorer_stats["build_blocked"] += 1
                return False
            if not _try_mount(press, target, defer_free=True):
                return False
            _bld_commit(target, _draw, bld_free)
            committed[press] = (target, prov)
            slots[0] -= 1
            co_scorer_stats[prov] += 1
            return True

        if not _SCORER_FULL_REOPT:
            # ADDITIVE — planned COs kept (mould-gate + retarget-on-block, no build veto
            # so this path is a superset of Phase-2), then NEW pull-forward + idle-fill.
            for (press, _o, ns) in today_planned:
                if _commit(press, ns, "planned", check_build=False):
                    continue
                # retarget-on-block — EXACT Phase-2 (_pick_retarget, no build veto) so
                # the planned+retarget layer matches the locked mould baseline bit-for-bit.
                # PRESS_RETURN_BLOCK: prefer a target this press has NOT left (avoid boomerang);
                # fall back to the unguarded pick if none exists (never strand → parity in that case).
                if _PRESS_RETURN_BLOCK:
                    _avoid = press_ran.get(press) or set()
                    if _prio_active and priority_deadline_map:
                        _avoid = _avoid - set(priority_deadline_map)
                    _alt = _pick_retarget(press, avoid=_avoid)
                    if _alt is None:
                        _alt = _pick_retarget(press)
                else:
                    _alt = _pick_retarget(press)
                if not (_alt is not None and _commit(press, _alt, "retarget", check_build=False)):
                    co_scorer_stats["cancelled"] += 1
            for (press, _o, ns) in tomorrow_planned:
                if press in committed:
                    continue
                if demand_remaining.get(press_state.get(press, {}).get("sku"), 1.0) > 0:
                    continue                                  # press still busy today
                if _commit(press, ns, "pullfwd"):
                    pulled.append(press)
            for press in sorted(press_state):
                if press in committed:
                    continue
                if demand_remaining.get(press_state[press]["sku"], 1.0) > 0:
                    continue                                  # not idle — leave it running
                _alt = _best_alt(press, bld_free, horizon_left)
                if _alt is not None:
                    _commit(press, _alt, "dynamic")
        else:
            # FULL RE-OPT — every eligible press competes; planned COs are candidates.
            planned_today = {p: ns for (p, _o, ns) in today_planned}
            elig = set(planned_today) | set(planned_tom)
            for press, st in press_state.items():
                if demand_remaining.get(st["sku"], 1.0) <= 0:
                    elig.add(press)
            pairs = []
            for press in sorted(elig):
                _cur = press_state.get(press, {}).get("sku")
                _cands: dict[str, str] = {}
                if press in planned_today:
                    _cands[planned_today[press]] = "planned"
                if press in planned_tom:
                    _cands.setdefault(planned_tom[press], "pullfwd")
                for s in press_allow_skus.get(press, ()):
                    if demand_remaining.get(s, 0.0) > 0:
                        _cands.setdefault(s, "retarget")
                for t, prov in _cands.items():
                    if t == _cur:
                        continue
                    _u = _co_utility(press, t, horizon_left, _bld_capacity(t, bld_free))
                    pairs.append((_u, press, t, prov))
            pairs.sort(key=lambda x: (-x[0], x[1], x[2]))
            for _u, press, t, prov in pairs:
                if _u <= 0:
                    break
                if press in committed:
                    continue
                if _commit(press, t, prov) and prov == "pullfwd":
                    pulled.append(press)

        if pulled:
            _pset = set(pulled)
            co_by_day[day + 1] = [(p, o, n) for (p, o, n) in tomorrow_planned
                                  if p not in _pset]
            daily_co_count[day + 1] = max(0, daily_co_count.get(day + 1, 0) - len(_pset))

        co_scorer_stats["idle"] += sum(
            1 for pr, st in press_state.items()
            if pr not in committed and demand_remaining.get(st["sku"], 1.0) <= 0)

        final = [(pr, press_state.get(pr, {}).get("sku"), tgt)
                 for pr, (tgt, _prov) in committed.items()]
        daily_co_count[day] = len(final)
        return final

    # Env-gated Day-0 diagnostic (OFF by default; no effect on scheduling). Captures
    # the Day-0 building primary seed, Day-0 curing press SKUs, building eligibility,
    # and demand so the primary/secondary/never-built cross-tab can be computed offline.
    _day0_dump_path = os.environ.get("DAY0_DUMP", "")
    if _day0_dump_path:
        import json as _json
        _cure_day0 = {str(p): st["sku"] for p, st in press_state.items()}
        _pc0: dict = {}
        for _st in press_state.values():
            _pc0[_st["sku"]] = _pc0.get(_st["sku"], 0) + 1
        _dd = {
            "bld_day0_primary": {str(m): s for m, s in machine_current_sku.items()},
            "cure_day0":        _cure_day0,
            "press_count_day0": _pc0,
            "machine_eligible": {str(m): sorted(str(s) for s in ss)
                                 for m, ss in machine_skus.items()},
            "demand":           {str(s): float(q) for s, q in demand_dict.items()},
            "sku_inch":         {str(s): sku_inch.get(s, "") for s in demand_dict},
            "s1_sku_to_machines": {str(s): sorted(str(x) for x in ms)
                                   for s, ms in s1_sku_to_machines.items()},
            "s1_locked_inch":   {str(m): i for m, i in s1_locked_inch.items()},
        }
        with open(_day0_dump_path, "w") as _f:
            _json.dump(_dd, _f)
        print(f"  [DAY0_DUMP] wrote Day-0 state → {_day0_dump_path}")

    # ── DELIVERY_PRIORITY: static feasibility pre-check (best-effort + relax-report) ──
    # Computes, per committed SKU, the most it could be cured by its deadline using ONLY
    # DB-allowable presses/moulds + inch-eligible building machines, plus the earliest
    # date it could be fully completed. Read-only; captured now, reported after the run.
    priority_precheck: list = []
    cured_by_deadline: dict[str, float] = {}
    if _prio_active:
        _sku_presses_pc: dict[str, set] = {}
        try:
            _dfca_pc = cetl.load_curing_allowable()
            _dem_skus_pc = set(demand_dict.keys())
            for _, _rw in _dfca_pc.iterrows():
                _s = str(_rw["SKUCode"]).strip()
                if _s in _dem_skus_pc:
                    _ms = _rw.get("Machines", []) or []
                    if _ms:
                        _sku_presses_pc[_s] = {str(_p) for _p in _ms}
        except Exception as _e:
            print(f"  [Rolling] DELIVERY_PRIORITY pre-check: curing-allowable load failed ({_e})")
        priority_precheck = _priority_feasibility_precheck(
            priority_deadline_map, priority_meta, demand_dict,
            _sku_presses_pc, _sku_moulds, sku_machine_map, cure_ct_map,
            plan_start, planning_days)

    print("\n" + "=" * 70)
    print("  ROLLING PIPELINE — Day-by-day simulation")
    print("=" * 70)

    # Stage-2 carcass gate: rolling per-SKU carcass BANK (list of [age_shifts_left,
    # qty]) seeded from opening carcass. Persists across all shifts/days; only used
    # when _STAGE2_CARCASS_GATE. Empty/unused otherwise (OFF path untouched).
    _carcass_bank: dict[str, list] = defaultdict(list)
    if _STAGE2_CARCASS_GATE:
        _seed_age = max(1, _STAGE1_CARCASS_LEAD + 1)
        for _s, _q in (opening_carcass or {}).items():
            if _q and float(_q) > 0:
                _carcass_bank[str(_s)].append([_seed_age, float(_q)])

    # STAGE1_CO: the carcass SKU each Stage-1 machine is currently set up for, carried
    # ACROSS shifts. A switch to a different carcass SKU costs a building CO (60 same-inch
    # / 180 diff-inch, charged in _gate_build). "" = never assigned yet (free first start,
    # like a GT machine's "start"). Only used when _STAGE1_CO; empty otherwise.
    machine_cur_carcass: dict[str, str] = {}
    # STAGE1_CO Site 2: per-(day,shift,machine,sku) carcass the gate actually built, in
    # build order. The CO-charged carcass rows are built from THIS (capped per SKU at real
    # Stage-2 consumption, so the gate's pre-build over-production is dropped) with CO
    # recomputed on the consolidated sequence → carcass total == Stage-2 GT, no overbook.
    # Only populated when _STAGE1_CO; the OFF path uses the max-flow re-derivation unchanged.
    _s1_prod_log: list = []
    _carcass_eod: list = []          # STAGE1_CO: end-of-day carcass buffer (bank after shift C)
    # S1_INCH_FLEX: the inches each Stage-1 machine is eligible to build carcass for (its
    # Stage-2-carcass SKUs' inches), for the one-way inch-advance when its inch is done.
    _s1_elig_inches: dict = defaultdict(set)
    _s1_visited: dict = defaultdict(set)     # inches each machine has been on (one-way guard)
    _s1_diff_co_count: dict = defaultdict(int)   # S1_DIFF_CO: diff-COs each machine has taken this month
    if _STAGE1_CO and _S1_INCH_FLEX:
        for _s2s, _ms in s1_sku_to_machines.items():
            _ii = sku_inch.get(str(_s2s), "")
            if _ii:
                for _mm in _ms:
                    if str(_mm) in _S1_MACHINES:
                        _s1_elig_inches[str(_mm)].add(_ii)
        for _mm in _S1_MACHINES:
            _s1_visited[str(_mm)].add(s1_locked_inch.get(str(_mm), ""))

    # ps3/ps4 hard monthly build cap (bc_config.PS_MAX_BUILD). Track cumulative build; once a
    # capped machine hits its cap it is dropped from that shift's eligibility so it stops building.
    try:
        _PS_MAX_BUILD = dict(getattr(_bc_cfg, "PS_MAX_BUILD", {}) or {})
    except Exception:
        _PS_MAX_BUILD = {}
    _ps_built_sofar: dict = defaultdict(float)

    # ── Plant holidays (bc_config.PLANT_HOLIDAYS) — non-working days inside the FIXED
    #    calendar span. Map each holiday DATE → 1-based day-index. Empty ⇒ feature INERT
    #    (no holiday day ⇒ output bit-for-bit identical). TWO CLOCKS: the loop still
    #    iterates ALL calendar days so aging stays CALENDAR-based; a holiday day produces
    #    nothing (0 build / 0 cure) and is excluded from working-day counts + util denoms,
    #    while in-flight CO/clean carries still complete (setup crew works the idle day). ──
    _holiday_days: set = set()
    try:
        for _h in (getattr(_bc_cfg, "PLANT_HOLIDAYS", []) or []):
            _hd = datetime.strptime(str(_h).strip(), "%Y-%m-%d").date()
            _idx = (_hd - plan_start.date()).days + 1
            if 1 <= _idx <= planning_days:
                _holiday_days.add(_idx)
    except Exception as _e:
        print(f"  [Rolling] PLANT_HOLIDAYS parse failed ({_e}); no holidays applied")
        _holiday_days = set()

    # Wall-clock [start,end) window per holiday PLAN-DAY (shift-based, NOT calendar date):
    # plan-day D = the 3 shifts starting plan_start+(D-1) days 07:00 and ending plan_start+D
    # days 07:00 (Shift C crosses midnight, so this correctly spans the wall-clock tail too).
    # Passed to `_split_rows_at_shift_boundaries` (holiday_windows=) so a machine whose
    # cumulative cursor has drifted behind schedule (an earlier PM/MTC skip-over) cannot have
    # that drift carry its production into the holiday's wall-clock window — see the HOLIDAY
    # FIX docstring on that function. Empty when PLANT_HOLIDAYS is empty ⇒ bit-for-bit parity.
    _holiday_wallclock_windows = [
        (plan_start + timedelta(days=_d - 1), plan_start + timedelta(days=_d))
        for _d in sorted(_holiday_days)
    ]

    def _is_holiday(_d: int) -> bool:
        return _d in _holiday_days

    working_days = planning_days - len([_d for _d in _holiday_days if 1 <= _d <= planning_days])

    def _working_days_left(_d: int) -> int:      # working days in [_d .. planning_days] inclusive
        return sum(1 for _x in range(_d, planning_days + 1) if _x not in _holiday_days)

    # First working day (for Day-1-only cold-start COs when day 1 is itself a holiday).
    _first_working_day = next((_d for _d in range(1, planning_days + 1) if _d not in _holiday_days), 1)
    if _holiday_days:
        print(f"  [Rolling] PLANT HOLIDAYS: day-index {sorted(_holiday_days)} → "
              f"{len(_holiday_days)} idle day(s); {working_days} working days; "
              f"first working day = {_first_working_day}")

    # ── Part B (REACTIVE_ONLY): the SINGLE reactive curing-CO rule ──────────────
    _reactive_surplus: dict = defaultdict(int)     # B-3: consecutive-shift surplus counter/press

    def _r_inch(_s):
        return sku_inch.get(_s, _s[8:10] if len(_s) >= 10 else "")

    def _reactive_co(day, shift, cur_shift_global, date_str):
        """Run once per shift AFTER building assignment. Collect presses wanting a CO
        (their SKU's demand done, or B-3 surplus), build each press's legal target list
        (allowable + 2 free moulds + demand>0 + building-supply test), try a depth-1
        machine-swap to unblock a mould-contended target, then fire best-first
        (delivery-EDF > GT-in-pool > same-size CO > larger unmet demand > SKU tiebreak),
        re-scoring after EVERY fire, with a forced-CO fallback that bypasses ONLY the
        supply test (never mould/allowable). Deterministic. A fired CO rides
        dynamic_co_tracker: CHANGEOVER this shift, RUNNING the new SKU next shift."""
        if _is_holiday(day):                                   # B-5: no NEW CO on a holiday
            return
        _wdl = max(1, _working_days_left(day))

        def _apply(press, target, forced=False, swap=False):
            _old = press_state[press]["sku"]
            press_count[_old] = max(0, press_count.get(_old, 0) - 1)
            dynamic_co_tracker[press] = (cur_shift_global, target)
            daily_co_count[day] += 1
            mould_life[press]  = MOULD_CLEAN_CYCLES
            clean_carry[press] = 0.0
            _reactive_surplus[press] = 0
            if forced: co_scorer_stats["forced"] += 1
            elif swap: co_scorer_stats["swap"] = co_scorer_stats.get("swap", 0) + 1
            else:      co_scorer_stats["dynamic"] += 1
            cure_co_events.append({
                "Date": date_str, "Day": day, "Shift": shift, "Press": press,
                "From_SKU": _old, "Target_SKU": target, "CO_Type": "Dynamic",
            })

        def _wanting():
            _w = []
            for p in sorted(press_state):
                st = press_state[p]
                if st.get("status") != "RUNNING" or p in dynamic_co_tracker:
                    continue
                _cs  = st.get("sku")
                _rem = demand_remaining.get(_cs, 0.0)
                if _rem <= 0:                                  # demand done → free to CO
                    _reactive_surplus[p] = 0
                    _w.append(p); continue
                if not _RCO_SURPLUS:                            # B-3 off → only CO truly-idle presses
                    _reactive_surplus[p] = 0
                    continue
                _rate = _cure_qty_per_shift(cure_ct_map.get(_cs, DEFAULT_CURING_CT)) * 3
                _need = math.ceil(_rem / (_rate * _wdl)) if _rate > 0 else 10 ** 9
                if press_count.get(_cs, 0) - 1 >= max(1, _need):   # B-3 surplus + n-1 protection
                    _reactive_surplus[p] += 1
                    if _reactive_surplus[p] >= 3:              # 3-shift hysteresis
                        _w.append(p)
                else:
                    _reactive_surplus[p] = 0
            return _w

        def _legal(press):
            """(ready, forced) target lists. ready = supply-OK + 2 free moulds."""
            _cur = press_state[press]["sku"]; _ci = _r_inch(_cur)
            _bf = _bld_free_min_shift()
            _ready, _forced = [], []
            for t in (press_allow_skus.get(press) or ()):
                if t == _cur or demand_remaining.get(t, 0.0) <= 0 or _n_free_for(t, press) < 2:
                    continue
                _draw = _cure_qty_per_shift(cure_ct_map.get(t, DEFAULT_CURING_CT))
                _rec = (t, gt_inventory.get(t, 0.0) >= _draw, _r_inch(t) == _ci)
                (_ready if _supply_ok(t, _bf) else _forced).append(_rec)
            return _ready, _forced

        def _score(press, rec):
            t, gt_pool, same = rec
            _dd = (priority_deadline_map or {}).get(t) if _prio_active else None
            _edf = (0, float(_dd)) if _dd is not None else (1, 0.0)   # committed first, EDF
            return (_edf, 1 if gt_pool else 0, 1 if same else 0,
                    demand_remaining.get(t, 0.0), t)                  # NO horizon term

        def _swap_unblock(press):
            """Depth-1 machine-swap: free a target-eligible mould from a DEMAND-DONE donor
            press (CO it to an allowable SKU that does NOT reuse that mould), so `press` can
            then mount a currently mould-blocked target. Returns the newly-feasible target or
            None. Safe: routes ownership through _try_mount; only evicts demand-done donors."""
            _cur = press_state[press]["sku"]
            for t in (press_allow_skus.get(press) or ()):
                if t == _cur or demand_remaining.get(t, 0.0) <= 0:
                    continue
                _elig = _sku_moulds.get(t, set())
                if len(_elig) < 2 or _n_free_for(t, press) >= 2:
                    continue
                _deficit = 2 - _n_free_for(t, press)
                _freed = 0
                for _m in sorted(x for x in _elig if _mould_owner.get(x) not in (None, press)):
                    if _freed >= _deficit or daily_co_count[day] >= MAX_CHANGEOVERS_PER_DAY:
                        break
                    _q = _mould_owner.get(_m)
                    if _q is None or _q in dynamic_co_tracker or press_state.get(_q, {}).get("status") != "RUNNING":
                        continue
                    _qs = press_state[_q]["sku"]
                    if demand_remaining.get(_qs, 0.0) > 0:            # only evict demand-done donors
                        continue
                    _ht = next((s for s in (press_allow_skus.get(_q) or ())
                                if s != _qs and demand_remaining.get(s, 0.0) > 0
                                and _m not in _sku_moulds.get(s, set())
                                and _n_free_for(s, _q) >= 2), None)
                    if _ht is None:
                        continue
                    if _try_mount(_q, _ht):        # _m ∉ elig(_ht) → freed to the pool
                        _apply(_q, _ht, swap=True)
                        _freed += 1
                if _n_free_for(t, press) >= 2:
                    return t
            return None

        while daily_co_count[day] < MAX_CHANGEOVERS_PER_DAY:
            _cands = []                                       # (score, press, target, forced)
            for press in _wanting():
                if press in dynamic_co_tracker:
                    continue
                _ready, _forced = _legal(press)
                for _rec in _ready:
                    _cands.append((_score(press, _rec), press, _rec[0], False))
                if not _ready:
                    _sw = _swap_unblock(press)
                    if _sw is not None:
                        _draw = _cure_qty_per_shift(cure_ct_map.get(_sw, DEFAULT_CURING_CT))
                        _rec = (_sw, gt_inventory.get(_sw, 0.0) >= _draw,
                                _r_inch(_sw) == _r_inch(press_state[press]["sku"]))
                        _cands.append((_score(press, _rec), press, _sw, False))
                    elif _forced:
                        _f = max(_forced, key=lambda r: (demand_remaining.get(r[0], 0.0), r[0]))
                        _cands.append((_score(press, _f), press, _f[0], True))
            if not _cands:
                break
            _cands.sort(key=lambda x: x[0], reverse=True)
            _sc, press, target, forced = _cands[0]
            if press in dynamic_co_tracker or not _try_mount(press, target):
                break                                         # safety: avoid a stuck loop
            _apply(press, target, forced=forced)

    def _do_swap_for(press, target, day, shift, cur_shift_global, date_str):
        """B-1 point 1: depth-1 machine-swap for the MID-SHIFT path. `press` is mould-blocked
        for `target`; free target-eligible mould(s) from DEMAND-DONE donor presses (CO each to
        an allowable SKU that does NOT reuse the mould), so `press` can then mount `target`.
        Returns True iff press is now mould-feasible. Deterministic; ownership only via
        _try_mount; only evicts demand-done donors (never steals from a producing press)."""
        _elig = _sku_moulds.get(target, set())
        if len(_elig) < 2:
            return False
        _deficit = 2 - _n_free_for(target, press)
        if _deficit <= 0:
            return True
        _freed = 0
        for _m in sorted(x for x in _elig if _mould_owner.get(x) not in (None, press)):
            if _freed >= _deficit or daily_co_count[day] >= MAX_CHANGEOVERS_PER_DAY:
                break
            _q = _mould_owner.get(_m)
            if (_q is None or _q in dynamic_co_tracker or _q in co_press_map
                    or press_state.get(_q, {}).get("status") != "RUNNING"):
                continue
            _qs = press_state[_q]["sku"]
            if demand_remaining.get(_qs, 0.0) > 0:          # only evict demand-done donors
                continue
            _ht = next((s for s in (press_allow_skus.get(_q) or ())
                        if s != _qs and demand_remaining.get(s, 0.0) > 0
                        and _m not in _sku_moulds.get(s, set())
                        and _n_free_for(s, _q) >= 2), None)
            if _ht is None:
                continue
            if _try_mount(_q, _ht):                         # _m ∉ elig(_ht) → freed to the pool
                press_count[_qs] = max(0, press_count.get(_qs, 0) - 1)
                dynamic_co_tracker[_q] = (cur_shift_global, _ht)
                daily_co_count[day] += 1
                mould_life[_q]  = MOULD_CLEAN_CYCLES
                clean_carry[_q] = 0.0
                co_scorer_stats["swap"] = co_scorer_stats.get("swap", 0) + 1
                cure_co_events.append({"Date": date_str, "Day": day, "Shift": shift,
                                       "Press": _q, "From_SKU": _qs, "Target_SKU": _ht,
                                       "CO_Type": "Dynamic"})
                _freed += 1
        return _n_free_for(target, press) >= 2

    # ── BLD_CURABLE_CAP: per-SKU curable GT-on-hand STOCK cap ────────────────────
    # = (#eligible curing presses) × cure_rate/shift × GT_SHELF_LIFE_SHIFTS — the most GT
    # the SKU's presses can drain before it ages out (3-day shelf). Uses the ELIGIBLE-press
    # count from curing_allowable (not moulds, which the DB undercounts, and not the live
    # running draw, so a buffer for presses between runs is preserved). Bounds the captive-
    # max / sticky build so a single-source machine (ps2 → TUXPE, 1 press) cannot accumulate
    # GT beyond ~9 shifts of its lone press's throughput. None ⇒ feature off.
    _sku_curable_ceiling: dict[str, float] | None = None
    if _BLD_CURABLE_CAP:
        _sku_curable_ceiling = {}
        for _s in demand_dict:
            _np = len(curing_allowable.get(_s, ()) or ())
            if _np <= 0:
                continue   # no eligible-press signal → don't throttle (preserve baseline)
            _cr = _cure_qty_per_shift(cure_ct_map.get(_s, DEFAULT_CURING_CT))
            _sku_curable_ceiling[_s] = float(_np * _cr * GT_SHELF_LIFE_SHIFTS)

    _captured_state: dict | None = None            # mid-month: full state at snapshot_at_day start
    # ── SAME_GROUP soft home-group lever (env SAME_GROUP, default OFF) ──────────────
    # _sg_d12_units accumulates GT units built per (sku, finer-group) during the plant
    # 2-day replay (days 1-2); _sku_home_group is frozen ONCE at the start of day 3.
    _sg_d12_units: dict = defaultdict(float)
    _sku_home_group: dict | None = None            # {sku: home_group}; None until computed (inert)

    def _compute_home_groups(_from_day: int) -> dict:
        """Freeze each GT SKU's HOME group at day 3 (after the 2-day replay).
          1. SKU BUILT in days 1-2 → home = the finer group it built the MOST units in.
          2. Else, if a SINGLE allowable GT group has enough monthly building capacity
             (Σ floor(SHIFT_SECS/_bld_ct_sec)·working-shifts-left) to cover the SKU's
             remaining demand → home = that group (max-capacity group when several qualify,
             the SKU's only group when it has just one). If NO single group can finish it →
             home = None (cross-group allowed, no penalty). Also None if it has no GT group."""
        _homes: dict = {}
        _wshifts = max(1, _working_days_left(_from_day) * 3)   # working shifts from day 3 to month-end
        _shift_secs = float(SHIFT_MINS) * 60.0
        # (a) built during the plant replay
        _built_by_sku: dict = defaultdict(dict)
        for (_s, _g), _u in _sg_d12_units.items():
            if _u > 0 and _g in _SG_GT_GROUPS:
                _built_by_sku[_s][_g] = _built_by_sku[_s].get(_g, 0.0) + _u
        for _s, _gm in _built_by_sku.items():
            _homes[_s] = max(_gm.items(), key=lambda kv: (kv[1], kv[0]))[0]
        # (b) not built in days 1-2 → single-group monthly-capacity test
        for _s in set().union(*machine_skus.values()) if machine_skus else set():
            if _s in _homes:
                continue
            _grp_machs: dict = defaultdict(list)
            for _m in sku_machine_map.get(_s, ()):
                _g = _sku_group_of(_m)
                if _g in _SG_GT_GROUPS:
                    _grp_machs[_g].append(_m)
            if not _grp_machs:
                _homes[_s] = None
                continue
            if len(_grp_machs) == 1:
                _homes[_s] = next(iter(_grp_machs))
                continue
            _dem = float(demand_remaining.get(_s, 0.0))
            _cap: dict = {}
            for _g, _ms in _grp_machs.items():
                _c = 0.0
                for _m in _ms:
                    _ct = _bld_ct_sec(_m, _s)
                    if _ct > 0:
                        _c += math.floor(_shift_secs / _ct) * _wshifts
                _cap[_g] = _c
            _qual = [_g for _g, _c in _cap.items() if _c >= _dem]
            if _qual:
                _homes[_s] = max(_qual, key=lambda g: (_cap[g], g))   # natural home = most capable group
            else:
                _homes[_s] = None                                      # no single group can finish → free
        _npin = sum(1 for _v in _homes.values() if _v)
        print(f"  [SAME_GROUP] home groups frozen at day {_from_day}: "
              f"{_npin}/{len(_homes)} GT SKUs pinned to one group "
              f"(pen={_SAME_GROUP_PEN})")
        return _homes

    # ── SG_DELIB deliberate + stable group allocation state ─────────────────────────
    _sku_grp_target: dict | None = None            # {sku: frozenset(allowed groups)}; None until day 3
    _sku_cur_group: dict = {}                       # {sku: current finer group} (updated at commit)
    _sku_last_group_move: dict = {}                 # {sku: day of last admitted group MOVE} (cooldown)

    def _compute_grp_targets(_from_day: int):
        """_best_group over ALL GT SKUs: the DELIBERATE, STABLE target group SET.
          - single group when ONE group's remaining monthly building capacity completes the
            SKU's remaining demand (the most-capable such group, seeded by any days 1-2 build);
          - else the MINIMAL completing SET (greedily add the highest-capacity groups until
            Σ capacity ≥ remaining demand) — the high-demand class (e.g. LSTL0/SUNE1) that no
            single group can finish. Falls back to ALL its GT groups if even that can't finish.
          Returns (targets:{sku: frozenset}, cur:{sku: seed group}). SKUs with no GT group are
          left OUT of `targets` → unconstrained (no penalty)."""
        _wshifts = max(1, _working_days_left(_from_day) * 3)
        _shift_secs = float(SHIFT_MINS) * 60.0
        # days 1-2 plant-replay build per (sku, group) → seeds continuity of the assignment
        _built_by_sku: dict = defaultdict(dict)
        for (_s, _g), _u in _sg_d12_units.items():
            if _u > 0 and _g in _SG_GT_GROUPS:
                _built_by_sku[_s][_g] = _built_by_sku[_s].get(_g, 0.0) + _u
        _targets: dict = {}
        _cur: dict = {}
        _allsku = set().union(*machine_skus.values()) if machine_skus else set()
        # INCH-AWARE: total demand competing for each (GT group, inch). SKUs on an inch share the
        # group's inch-locked machines, so a SKU's realistic capacity from a group is its
        # demand-proportional share of the group's capacity (see _SG_INCH_AWARE_TARGETS).
        _grp_inch_dem: dict = defaultdict(float)
        if _SG_INCH_AWARE_TARGETS:
            for _s2 in _allsku:
                _d2 = float(demand_remaining.get(_s2, 0.0))
                if _d2 <= 0:
                    continue
                _i2 = sku_inch.get(_s2, "")
                for _g2 in ({_sku_group_of(_m2) for _m2 in sku_machine_map.get(_s2, ())}
                            & _SG_GT_GROUPS):
                    _grp_inch_dem[(_g2, _i2)] += _d2
        for _s in _allsku:
            _grp_machs: dict = defaultdict(list)
            for _m in sku_machine_map.get(_s, ()):
                _g = _sku_group_of(_m)
                if _g in _SG_GT_GROUPS:
                    _grp_machs[_g].append(_m)
            if not _grp_machs:
                continue                              # no GT group → unconstrained (free)
            _cap: dict = {}
            for _g, _ms in _grp_machs.items():
                _c = 0.0
                for _m in _ms:
                    _ct = _bld_ct_sec(_m, _s)
                    if _ct > 0:
                        _c += math.floor(_shift_secs / _ct) * _wshifts
                _cap[_g] = _c * _SG_GRP_CAP_DERATE   # deliverable (COs/contention/draw) < raw
            _dem = float(demand_remaining.get(_s, 0.0))
            _built = _built_by_sku.get(_s, {})
            # ── DAY2→DAY3 CONTINUITY (plant-seeded) ──────────────────────────────────
            # If the SKU was BUILT in days 1-2, its day-3 current group MUST be the plant's
            # days-1-2 DOMINANT group (most units) — carry the plant assignment forward, no
            # gratuitous boundary group-jump. This is exactly _compute_home_groups' "home".
            # Restrict to the plant groups that are ALSO allowable (day-3+ building must be
            # feasible there); if the plant built it ONLY in unallowable groups (§22 rows),
            # continuity is physically impossible → fall back to the best allowable group.
            _built_allow = {_g: _u for _g, _u in _built.items() if _g in _grp_machs}
            _home = (max(_built_allow.items(), key=lambda kv: (kv[1], kv[0]))[0]
                     if _built_allow else None)
            # order: HOME group first (continuity), then most days-1-2 units, capacity, id
            _order = sorted(_grp_machs, key=lambda g: (0 if g == _home else 1,
                                                       -_built.get(g, 0.0), -_cap[g], g))
            _inch_s = sku_inch.get(_s, "")
            _chosen: list = []
            _acc = 0.0
            for _g in _order:
                _chosen.append(_g)
                if _SG_INCH_AWARE_TARGETS:
                    # this SKU's demand-proportional SHARE of group g's capacity (inch contention)
                    _comp = _grp_inch_dem.get((_g, _inch_s), 0.0)
                    _acc += _cap[_g] * min(1.0, _dem / _comp) if _comp > _dem else _cap[_g]
                else:
                    _acc += _cap[_g]
                if _acc >= _dem:                      # minimal completing set reached
                    break
            _targets[_s] = frozenset(_chosen)
            # seed the day-3 current group = plant days-1-2 dominant group when built there;
            # else the deliberate best-allowable pick (SKU not built in days 1-2, or §22 forced).
            _cur[_s] = _home if _home is not None else _order[0]
        _mset = {_s: sorted(_v) for _s, _v in _targets.items() if len(_v) > 1}
        print(f"  [SG_DELIB] deliberate group targets frozen at day {_from_day}: "
              f"{len(_targets)} GT SKUs assigned "
              f"({len(_targets)-len(_mset)} single-group, {len(_mset)} minimal-SET) "
              f"(derate={_SG_GRP_CAP_DERATE}, band={_SG_MOVE_BAND}, cooldown={_SG_MOVE_COOLDOWN_DAYS}d, hard={_SG_HARD})")
        for _s, _v in _mset.items():
            print(f"      [SG_DELIB] minimal-SET: {_s}  demand={demand_remaining.get(_s,0.0):.0f}  groups={_v}")
        return _targets, _cur

    for day in range(1, planning_days + 1):
        _cur_day[0] = day                          # for the mould-movement log in _try_mount
        # Per-machine one-way locks (7002 14→16, 7501 12→13) — flip once on the switch day, never revert.
        if switch_day_7002 is not None and "7002" in machine_locked_inches:
            machine_locked_inches["7002"] = {"16"} if day >= switch_day_7002 else {"14"}
        if switch_day_7501 is not None and "7501" in machine_locked_inches:
            machine_locked_inches["7501"] = {"13"} if day >= switch_day_7501 else {"12"}
        # GENERAL one-way inch rule (subsumes the old per-machine 7003/7002/7501 switches): once a
        # machine's current (dominant) inch's servable demand is DONE, flip its lock ONE-WAY to the
        # NEEDIEST OTHER inch it is CT-allowable for. _inch_gate then routes a normal diff-CO there;
        # it never reverts (the machine is recorded in _oneway_switched).
        if _ONEWAY_INCH_GENERAL and day > 1:
            # BUILDING-done signal: an inch is "done" for a machine when every one of its SKUs on that
            # inch already has GT built for all remaining demand (demand_remaining - gt_inventory <= 0),
            # so no more BUILDING is needed there (curing lags, so demand_remaining alone is too late).
            # Skip UNBUILDABLE SKUs (no allowable-matrix master data → 0 eligible machines): their
            # demand can never be completed, so counting them would block the "inch done" check forever.
            def _bld_rem(_msk, _i):
                return sum(max(0.0, demand_remaining.get(_s, 0.0) - gt_inventory.get(_s, 0.0))
                           for _s in _msk
                           if sku_inch.get(_s, "") == _i and _s in _buildable_skus)
            for _m in list(machine_locked_inches):
                if _m in _oneway_switched:
                    continue
                _ci = machine_inch_now.get(_m) or next(iter(machine_locked_inches[_m]), None)
                if not _ci:
                    continue
                _msk = machine_skus.get(_m, ())
                if _bld_rem(_msk, _ci) > _ONEWAY_GEN_DONE_EPS:
                    continue                                   # current inch still needs building → stay
                _best = None; _best_rem = _ONEWAY_GEN_DONE_EPS
                for _oi in ({sku_inch.get(_s, "") for _s in _msk} - {_ci, ""}):
                    _r = _bld_rem(_msk, _oi)
                    if _r > _best_rem:
                        _best_rem = _r; _best = _oi
                if _best is not None:
                    machine_locked_inches[_m] = {_best}        # one diff-CO to the neediest other inch
                    machine_inch_now[_m] = _best
                    machine_used_inches.setdefault(_m, set()).add(_best)
                    _oneway_switched.add(_m)
        _holiday = _is_holiday(day)                 # this whole calendar day is a plant holiday
        date     = plan_start + timedelta(days=day - 1)
        date_str = date.strftime("%Y-%m-%d")
        # MID-MONTH snapshot: at the START of snapshot_at_day (before any mutation today), deep-copy
        # the full physical state so a second run can continue from here (see run_rolling_pipeline_2pass).
        if snapshot_at_day is not None and day == snapshot_at_day and _captured_state is None:
            _captured_state = {
                "_snap_day": day,
                "gt_inventory": dict(gt_inventory),
                "gt_lots": {s: [list(l) for l in lots] for s, lots in gt_lots.items()},
                "carcass_bank": {s: [list(l) for l in b] for s, b in _carcass_bank.items()},
                "press_state": {p: dict(v) for p, v in press_state.items()},
                "press_count": dict(press_count),
                "mould_life": dict(mould_life),
                "clean_carry": dict(clean_carry),
                "co_carry": dict(co_carry),
                "mould_owner": dict(_mould_owner),
                "press_moulds": {p: set(m) for p, m in _press_moulds.items()},
                "sku_moulds": {s: set(m) for s, m in _sku_moulds.items()},
                "mould_skus": {m: set(s) for m, s in _mould_skus.items()},
                "machine_current_sku": dict(machine_current_sku),
                "machine_minutes_on_sku": dict(machine_minutes_on_sku),
                "last_build_day": dict(last_build_day),
                "ps_built_sofar": dict(_ps_built_sofar),
                "demand_remaining": dict(demand_remaining),
                "demand_dict": dict(demand_dict),
            }
        # MID-MONTH injection: at the START of day 1, overwrite the 1st-of-month DB seeds with a
        # prior run's carried physical state (Run 1 at this run's start date). GT-lot build-days and
        # last_build_day are RE-BASED from Run-1 numbering to Run-2 (day1 = plan start) so the 3-day
        # GT shelf expires exactly what Run 1 would have. Demand/waste/CO trackers are NOT carried
        # (Run 2 re-plans the remaining period on its own reduced demand from a clean slate).
        if initial_state is not None and day == 1:
            _rb = int(initial_state.get("_snap_day", 1)) - 1        # Run-1 day → Run-2 day offset
            gt_inventory.clear(); gt_inventory.update(initial_state["gt_inventory"])
            gt_lots.clear()
            for _s, _lots in initial_state["gt_lots"].items():
                gt_lots[_s] = [[_bd - _rb, _q] for _bd, _q in _lots]
            _carcass_bank.clear()
            for _s, _b in initial_state["carcass_bank"].items():
                _carcass_bank[_s] = [list(_l) for _l in _b]
            press_state.clear()
            press_state.update({p: dict(v) for p, v in initial_state["press_state"].items()})
            press_count.clear(); press_count.update(initial_state["press_count"])
            mould_life.clear(); mould_life.update(initial_state["mould_life"])
            clean_carry.clear(); clean_carry.update(initial_state["clean_carry"])
            co_carry.clear(); co_carry.update(initial_state["co_carry"])
            _mould_owner.clear(); _mould_owner.update(initial_state["mould_owner"])
            _press_moulds.clear()
            _press_moulds.update({p: set(m) for p, m in initial_state["press_moulds"].items()})
            _sku_moulds.clear()
            _sku_moulds.update({s: set(m) for s, m in initial_state["sku_moulds"].items()})
            _mould_skus.clear()
            _mould_skus.update({m: set(s) for m, s in initial_state["mould_skus"].items()})
            machine_current_sku.clear(); machine_current_sku.update(initial_state["machine_current_sku"])
            machine_minutes_on_sku.clear()
            machine_minutes_on_sku.update(initial_state["machine_minutes_on_sku"])
            last_build_day.clear()
            last_build_day.update({s: d - _rb for s, d in initial_state["last_build_day"].items()})
            _ps_built_sofar.clear(); _ps_built_sofar.update(initial_state["ps_built_sofar"])
            # curing_allowable was built at :6494 from the DAY-0 press_state; resync it to the
            # carried day-K positions so a SKU cured by a carried press isn't falsely flagged
            # "missing curing allowable machine" (Eligible_Machines=0) in the Demand-Fulfillment sheet.
            for _p, _v in press_state.items():
                if _p not in curing_allowable[_v["sku"]]:
                    curing_allowable[_v["sku"]].append(_p)
            print(f"  [midmonth] injected carried state: {len(press_state)} presses, "
                  f"{sum(gt_inventory.values()):,.0f} GT on hand, {len(gt_lots)} GT-lot SKUs")
        # GT per-lot FIFO expiry at DAY START: drop lots older than the 3-day shelf as WASTE
        # BEFORE any curing today, so expired GT can never be cured (strict FIFO). Feeds
        # writeoff_cum (demand-cap fix) + the expired-GT waste column + expired_GT Shift rows.
        # Aging is calendar-day (runs on holidays too). Replaces the old end-of-day _writeoff_stale_gt.
        day_writeoff = _gt_expire_lots(day, date_str)
        writeoff_total += day_writeoff
        day_carcass_writeoff = 0.0                   # per-day expired carcass (accumulated in shift loop)
        # Reset the per-machine daily SKU set; the overnight carryover SKU counts as #1.
        machine_day_skus = {str(_m): ({str(_s)} if _s else set())
                            for _m, _s in machine_current_sku.items()}
        machine_day_diff_co = {}                    # Part 2: reset per-day diff-CO budget counter
        machine_day_co = {}                         # S2_CAMPAIGN: reset per-day total-CO budget counter
        if _ROLLING_HORIZON_CO_ENABLED:
            today_cos = _rolling_horizon_co_call(
                day=day, planning_days=planning_days,
                press_state=press_state, press_count=press_count,
                demand_remaining=demand_remaining, demand_dict=demand_dict,
                priority_score_map=priority_score_map,
                df_allowable=_df_curing_allow_static,
                ct_map=cc_result["ct_map"],
                dynamic_co_tracker=dynamic_co_tracker,
                scheduler=_rolling_co_scheduler,
                max_co_per_day=MAX_CHANGEOVERS_PER_DAY,
            )
            daily_co_count[day] += len(today_cos)
        elif _DYNAMIC_CO_PLANNER_ENABLED:
            today_cos = _plan_day_cos(
                day, press_state, demand_remaining, press_count,
                cure_ct_map, priority_score_map, demand_dict,
                press_to_demand_targets, press_total_demand,
                sku_campaign_tier, ri_skus, nri_skus,
                daily_co_count, planning_days,
            )
            daily_co_count[day] += len(today_cos)
            if os.environ.get("DYNCO_DEBUG"):
                print(f"    [DynCO-debug] Day {day}: {len(today_cos)} events "
                      f"{today_cos if today_cos else ''}")
            # Reset for today's shifts to populate fresh — sku_campaign_tier
            # must reflect the day just simulated, not accumulate forever.
            sku_campaign_tier = {}
        else:
            today_cos = co_by_day.get(day, [])
        # ── Holiday fix #1: defer EVERY CO planned on a holiday to the next working day ──
        # No new changeover starts on an idle day (decision B). Each CO rolls to the next
        # non-holiday day with CO budget; month-end overflow is dropped. Runs BEFORE the
        # HYBRID_CO_DEFER / mould gate / scorer so no mould is committed and no pre-build is
        # injected on the holiday. Inert unless this day is a holiday.
        if _HOLIDAY_CO_DEFER and today_cos and _is_holiday(day):
            _hol_moved, _hol_lost = 0, 0
            for (_p, _old, _new) in today_cos:
                daily_co_count[day] = max(0, daily_co_count.get(day, 0) - 1)
                _nwd = next((_x for _x in range(day + 1, planning_days + 1)
                             if _x not in _holiday_days
                             and daily_co_count.get(_x, 0) < MAX_CHANGEOVERS_PER_DAY), None)
                if _nwd is None:
                    _hol_lost += 1; continue
                co_by_day.setdefault(_nwd, []).append((_p, _old, _new))
                daily_co_count[_nwd] = daily_co_count.get(_nwd, 0) + 1
                _hol_moved += 1
            today_cos = []
            co_by_day[day] = []
            _VERBOSE and print(f"  [Rolling] Day {day} HOLIDAY: deferred {_hol_moved} CO(s) "
                               f"→ next working day, {_hol_lost} dropped (month-end)")
        # ── Item 2: DEFER a planned CO instead of preempting a fulfillable, needed SKU ──
        # A planned CO whose old SKU still has demand that WON'T finish today (>2 shifts) would
        # otherwise preempt in Shift A (co_shift_idx=0), abandoning that unmet demand. If the SKU
        # is still FULFILLABLE (building can supply it now, live _supply_ok) AND this press is NOT
        # surplus (n-1 presses can't cover it), push the CO to the next working day instead — the
        # press keeps producing its old SKU. Runs BEFORE the mould gate → no _try_mount residue.
        # HYBRID_CO_DEFER=0 = bit-for-bit.
        if _HYBRID_CO_DEFER and today_cos and not (_REACTIVE_ONLY and _RCO_ARBITER):
            _wdl_d   = max(1, _working_days_left(day))
            _bf_defer = _bld_free_min_shift()
            _kept_d, _deferred_d = [], []
            for (_p, _old, _new) in today_cos:
                _remd = demand_remaining.get(_old, 0.0)
                if _remd <= 0:
                    _kept_d.append((_p, _old, _new)); continue
                _odraw = (_cure_qty_per_shift(cure_ct_map.get(_old, DEFAULT_CURING_CT))
                          * max(1, press_count.get(_old, 1)))
                _nsh = math.ceil(_remd / _odraw) if _odraw > 0 else 99
                if _nsh <= 2:                                   # finishes today → fire as planned
                    _kept_d.append((_p, _old, _new)); continue
                _rate_d = _cure_qty_per_shift(cure_ct_map.get(_old, DEFAULT_CURING_CT)) * 3
                _need_d = math.ceil(_remd / (_rate_d * _wdl_d)) if _rate_d > 0 else 10 ** 9
                _surplus_d = press_count.get(_old, 0) - 1 >= max(1, _need_d)
                if _surplus_d or not _supply_ok(_old, _bf_defer):
                    _kept_d.append((_p, _old, _new)); continue  # surplus or unfulfillable → fire (preempt)
                _nwd = next((_x for _x in range(day + 1, planning_days + 1)
                             if _x not in _holiday_days
                             and daily_co_count.get(_x, 0) < MAX_CHANGEOVERS_PER_DAY), None)
                if _nwd is None:
                    _kept_d.append((_p, _old, _new)); continue  # nowhere to defer → fire
                _deferred_d.append((_p, _old, _new, _nwd))
            if _deferred_d:
                today_cos = _kept_d
                co_by_day[day] = _kept_d
                for (_p, _old, _new, _nwd) in _deferred_d:
                    daily_co_count[day] = max(0, daily_co_count.get(day, 0) - 1)
                    co_by_day.setdefault(_nwd, []).append((_p, _old, _new))
                    daily_co_count[_nwd] = daily_co_count.get(_nwd, 0) + 1
                _VERBOSE and print(f"  [Rolling] Day {day}: deferred {len(_deferred_d)} planned CO(s) "
                      f"(fulfillable + needed) → {[(c[0], c[3]) for c in _deferred_d]}")
        # ── Mould gate (planned COs) ──────────────────────────────────────────
        # A planned CO can only happen if 2 eligible moulds are free for the new
        # SKU. Gate HERE (day-start) — not at apply-time — because co_press_map
        # drives the curing sim THIS day; a CO blocked later would already have
        # been cured. Feasible COs get their moulds committed now (_try_mount);
        # blocked ones are dropped so the press keeps its old SKU all day.
        if (_mould_gate and _CO_SCORER_ENABLED and not (_REACTIVE_ONLY and _RCO_ARBITER)
                and not (_HOLIDAY_CO_DEFER and _is_holiday(day))):
            # Phase 3 — unified global CO solve (planned / pull-forward / retarget /
            # dynamic / idle under one utility + mould + building-feed gate). Runs even
            # when today_cos is empty (there may be idle presses to fill / pull-forwards).
            # DISABLED only under the once-per-shift arbiter (_RCO_ARBITER); the B-1
            # mid-shift base keeps it (idle-fill), like the original B-1 measurement.
            # Holiday fix #1: skipped on a holiday so no NEW idle-fill/scorer CO starts.
            today_cos = _solve_day_cos(day, today_cos)
        elif _mould_gate and today_cos and not (_HOLIDAY_CO_DEFER and _is_holiday(day)):
            # Phase 2a — scarce-first: claim moulds for the SCARCEST new-SKU first
            # (fewest eligible moulds) so a 2-mould SKU is not blocked by a 6-mould
            # SKU grabbing a shared mould first. Pure reordering — the SET of COs
            # attempted is unchanged, so this can only REDUCE blocks, never add them.
            # Deterministic tiebreak on (press, new_sku). MOULD_OPT=0 keeps input order.
            _co_order = today_cos
            if _mould_opt:
                _co_order = sorted(
                    today_cos,
                    key=lambda t: (len(_sku_moulds.get(t[2], set())), str(t[0]), str(t[2])),
                )
            _kept = []
            for _p, _os, _ns in _co_order:
                if _try_mount(_p, _ns, defer_free=True):
                    _kept.append((_p, _os, _ns))
                else:
                    # Phase 2b — retarget-on-block: the scheduled new-SKU has no free
                    # moulds; rather than idle the press on its old (usually demand-done)
                    # SKU, redirect the CO to the neediest allowable SKU that still has
                    # 2 free moulds. Recovers a wasted CO slot; demand cap still clips.
                    _alt = _pick_retarget(_p) if _mould_opt else None
                    if _alt is not None and _try_mount(_p, _alt, defer_free=True):
                        _kept.append((_p, _os, _alt))
                        mould_retargeted_cos += 1
                    else:
                        mould_blocked_cos += 1
            today_cos = _kept
        # Phase 4 — global mould optimiser: proactively move scarce moulds toward the
        # most-under-served SKUs (adds COs within the daily cap). No-op unless enabled.
        today_cos = _global_mould_boost(day, today_cos)

        # ── IDLE_PRESS_ACTIVATE (SUPPLY-AWARE, any working day) ───────────────────
        # Bring roster presses absent from the Day-0 snapshot online as DIRECT PRODUCTION (fresh
        # moulds, no curing CO — a fresh press is not a changeover). SUPPLY-AWARE: a press is
        # cold-started only when Building can realistically feed its target SKU (per-SKU marginal
        # test — see _IDLE_PRESS_SUPPLY_AWARE); otherwise it is DEFERRED and retried on a LATER
        # working day when supply exists, instead of blindly starting and immediately starving.
        # Runs each working day (was Day-1-only) so deferred presses activate when supply appears.
        # _idle_presses is the mutable pool of not-yet-activated presses. Exempt from the CO cap.
        def _idle_supply_ok(_sku: str) -> bool:
            """Per-SKU marginal supply test (day-level signals only): should we cold-start ANOTHER
            press on `_sku` now, or DEFER until Building can feed it? Rules (avoids the rejected
            inch-level over-skip): (1) no eligible building machine → never (unsupplyable). (2) the
            SKU has NO curing press yet → YES (must start one to cure it at all). (3) an ADDITIONAL
            press → only when Building is actually BANKING GT for it (on-hand ≥ one press-shift of
            draw) OR a known per-SKU building rate covers the extra press — i.e. real supply exists,
            so the new press is fed instead of immediately starving. Otherwise DEFER to a later day."""
            if not sku_machine_map.get(_sku):
                return False                        # no eligible building machine → cannot supply
            if press_count.get(_sku, 0) <= 0:
                return True                         # SKU uncured so far → activate the first press
            _per_press = _cure_qty_per_shift(cure_ct_map.get(_sku, DEFAULT_CURING_CT))
            if gt_inventory.get(_sku, 0.0) >= _per_press * _IDLE_PRESS_SUPPLY_MARGIN:
                return True                         # Building is banking GT → an extra press is fed
            _bs = _buildable_rate.get(_sku) if _buildable_rate is not None else None
            if _bs is not None and _bs > 0:
                _need = (press_count.get(_sku, 0) + 1) * _per_press * 3
                return _bs >= _need * _IDLE_PRESS_SUPPLY_MARGIN
            return False                            # no banked GT / no rate → DEFER, retry later
        if ((_IDLE_PRESS_ACTIVATE or _FULL_PRESS_ROSTER) and _idle_presses
                and day not in _holiday_days):
            _cold = []
            for _ip in list(_idle_presses):
                _tgt = _pick_retarget(_ip)          # neediest allowable in-demand SKU w/ 2 free moulds
                if _tgt is None:
                    continue                        # no feasible target today → keep in pool, retry later
                if _IDLE_PRESS_SUPPLY_AWARE and not _idle_supply_ok(_tgt):
                    continue                        # DEFER — Building can't feed this press yet
                if not _try_mount(_ip, _tgt, defer_free=False):
                    continue                        # no 2 free moulds → keep in pool, retry later
                press_state[_ip]  = {"sku": _tgt, "status": "RUNNING"}
                press_count[_tgt] = press_count.get(_tgt, 0) + 1
                curing_allowable[_tgt].append(_ip)
                mould_life[_ip]   = MOULD_CLEAN_CYCLES
                clean_carry[_ip]  = 0.0
                _idle_presses.remove(_ip)           # activated → out of the deferral pool
                _cold.append((_ip, _tgt))
            if _cold:
                _VERBOSE and print(f"  [Rolling] Day {day}: supply-aware cold-started "
                      f"{len(_cold)} idle press(es) as DIRECT PRODUCTION: {_cold}"
                      + (f" ({len(_idle_presses)} deferred — no supply yet)" if _idle_presses else ""))

        # ── Runner-Out Day-1 CO (Day 1 only) ──────────────────────────────────────
        # Plant rule: a press running a NO-DEMAND SKU at Day-0 (Runner-Out) must change over
        # on Day-1 Shift A to its preferred demand SKU and produce from Shift B — it must NOT
        # sit idle on the dead SKU. Force a Day-1 CO for any RO press not already scheduled one
        # today. Same contention-safe mount as the cold-start; exempt from the daily CO cap.
        # Toggle: bc_config.RUNNER_OUT_DAY1_CO_ENABLED (env RUNNER_OUT_DAY1_CO=0 disables).
        if day == _first_working_day and getattr(_bc_cfg, "RUNNER_OUT_DAY1_CO_ENABLED", True):
            _co_presses = {str(_p) for _p, _os, _ns in today_cos}
            _roco = []
            for _p, _st in list(press_state.items()):
                if str(_p) in _co_presses:
                    continue
                _psku = _st.get("sku", "")
                if demand_remaining.get(_psku, 0.0) > 0:      # running a demand SKU -> Runner-In, keep
                    continue
                _tgt = _pick_retarget(_p)                      # neediest allowable demand SKU w/ 2 free moulds
                if _tgt is None or _tgt == _psku:
                    continue
                if not _try_mount(_p, _tgt, defer_free=False):
                    continue
                press_state[_p] = {"sku": _tgt, "status": "RUNNING"}
                _roco.append((_p, _psku, _tgt))
            if _roco:
                today_cos = list(today_cos) + _roco
                _VERBOSE and print(f"  [Rolling] Day 1: forced Runner-Out CO on {len(_roco)} press(es) "
                      f"{[(c[0], c[2]) for c in _roco]}")

        # ── Point 2: light forward-looking PRE-POSITIONING (REACTIVE_ONLY) ──────────
        # Recreate the static plan's foresight from LIVE state: an UNDER-SERVED buildable SKU
        # (running presses < presses_needed for the remaining horizon) pulls a SURPLUS press
        # (its own SKU over-served, n-1 safe) via a proactive Day CO added to today_cos → it
        # flows through co_press_map so building PRE-BUILDS the target's GT (building couples to
        # the move). Need-gated on BOTH sides + buildable-checked + rate-limited → foresight
        # without the over-firing that killed the blind surplus release. Rides the CO cap.
        if _REACTIVE_ONLY and _RCO_PREPOS and day > 1:
            _wdl = max(1, _working_days_left(day))
            def _pneed(_s):
                _r = _cure_qty_per_shift(cure_ct_map.get(_s, DEFAULT_CURING_CT)) * 3
                _rem = demand_remaining.get(_s, 0.0)
                return math.ceil(_rem / (_r * _wdl)) if (_r > 0 and _rem > 0) else 0
            _cop = {str(_p) for _p, _o, _n in today_cos}
            _run: dict = defaultdict(int)
            for _p, _st in press_state.items():
                if (_st.get("status") == "RUNNING" and str(_p) not in _cop
                        and _p not in dynamic_co_tracker):
                    _run[_st["sku"]] += 1
            _under = sorted(                                      # under-served buildable, biggest deficit first
                ((_pneed(_s) - _run.get(_s, 0), _s) for _s in list(demand_remaining)
                 if demand_remaining.get(_s, 0.0) > 0 and _run.get(_s, 0) < _pneed(_s)
                 and _buildable_rate is not None and _buildable_rate.get(_s, 0.0) > 0),
                reverse=True)
            _limit = min(MAX_CHANGEOVERS_PER_DAY - daily_co_count[day], _RCO_PREPOS_MAX)
            _pp = []
            for _deficit, _t in _under:
                if len(_pp) >= _limit:
                    break
                if _run.get(_t, 0) >= _pneed(_t):
                    continue
                _draw = _cure_qty_per_shift(cure_ct_map.get(_t, DEFAULT_CURING_CT))
                if _bld_capacity(_t, _bld_free_min_shift()) < _draw:   # building can feed one more
                    continue
                _donor = None; _dsku = None
                for _p in sorted(press_state):                   # a SURPLUS press, n-1 safe, allowable+mountable
                    if str(_p) in _cop or _p in dynamic_co_tracker or press_state[_p].get("status") != "RUNNING":
                        continue
                    _ps = press_state[_p]["sku"]
                    if _run.get(_ps, 0) - 1 < max(1, _pneed(_ps)):
                        continue
                    if _t not in set(press_allow_skus.get(_p) or ()):
                        continue
                    if _n_free_for(_t, _p) >= 2 and _try_mount(_p, _t):
                        _donor = _p; _dsku = _ps
                        break
                if _donor is not None:
                    press_state[_donor] = {"sku": _t, "status": "RUNNING"}
                    _pp.append((_donor, _dsku, _t))
                    _cop.add(str(_donor))
                    _run[_dsku] -= 1; _run[_t] = _run.get(_t, 0) + 1
            if _pp:
                today_cos = list(today_cos) + _pp
                daily_co_count[day] += len(_pp)                  # reserve budget so mid-shift + R10 hold
                _VERBOSE and print(f"  [Rolling] Day {day}: pre-positioned {len(_pp)} press(es) "
                      f"{[(c[0], c[2]) for c in _pp]}")

        co_press_map: dict[str, str] = {p: ns for p, _, ns in today_cos}
        # SKUs that curing presses are switching TO today — building must pre-build for these
        co_target_skus_today: frozenset = frozenset(co_press_map.values())

        # ── Which SHIFT does each planned CO fire in? ──────────────────────────
        # Plant rule: a press goes to changeover as soon as it FINISHES its current
        # SKU (cap permitting) — not at a fixed 07:00. So place each CO in the shift
        # where its old SKU is projected to run out:
        #   already finished        → Shift A (index 0) — never make a free press wait
        #   finishes in n shifts    → that shift (clamped to C)
        #   will not finish today   → Shift A (preemptive Class-A CO — the static
        #                             scheduler booked it for this day deliberately)
        # OFF (_CO_SHIFT_SPREAD_ENABLED=False) ⇒ every CO in Shift A (old behaviour).
        co_shift_idx: dict[str, int] = {}
        for _p, _old, _new in today_cos:
            _idx = 0
            if _CO_SHIFT_SPREAD_ENABLED:
                _rem = demand_remaining.get(_old, 0.0)
                if _rem > 0:
                    _octt  = cure_ct_map.get(_old, DEFAULT_CURING_CT)
                    _odraw = _cure_qty_per_shift(_octt) * max(1, press_count.get(_old, 1))
                    _n     = math.ceil(_rem / _odraw) if _odraw > 0 else 99
                    _idx   = _n if _n <= 2 else 0      # >2 shifts ⇒ won't finish ⇒ preempt in A
            co_shift_idx[_p] = _idx

        # LOOKAHEAD_BUF: anticipated PEAK curing draw per SKU today = (presses running it now +
        # presses scheduled to CO onto it today) × cure-rate. Deterministic from the CO plan; lets
        # the dynamic + forward buffers pre-build for a KNOWN incoming draw spike. {} when OFF.
        _lookahead_draw: dict[str, float] = {}
        if _LOOKAHEAD_BUF:
            _joining: dict[str, int] = defaultdict(int)
            for _p, _os, _ns in today_cos:
                _joining[str(_ns)] += 1
            for _s in set(press_count) | set(_joining):
                _rate = _cure_qty_per_shift(cure_ct_map.get(_s, DEFAULT_CURING_CT))
                _lookahead_draw[_s] = (press_count.get(_s, 0) + _joining.get(_s, 0)) * _rate

        # Per-shift simulation: build → cure for each shift independently.
        # Building assignment runs once per shift (not once per day) so each
        # shift's build plan reacts to the actual GT inventory and curing demand
        # for that specific shift — matching the plant's actual scheduling approach.
        day_built:    dict[str, float] = defaultdict(float)  # all machines (GT + carcass)
        day_gt_built: dict[str, float] = defaultdict(float)  # GT machines only (no Stage-1)
        day_cured_d:  dict[str, float] = defaultdict(float)

        for shift in SHIFTS:
            key = (date_str, shift)
            shift_bld: dict[str, int] = defaultdict(int)
            build_by_shift_sku[key] = shift_bld

            # Global shift index: unique monotonic ID used by dynamic_co_tracker.
            # (day−1)×3 + {A=0, B=1, C=2}
            cur_shift_global = (day - 1) * 3 + SHIFTS.index(shift)

            # ── 0. Apply dynamic CO transitions ──────────────────────────────
            # When demand was met in shift X (CO started immediately), the press
            # starts RUNNING for the new SKU from shift X+1 onwards.
            for _press, (_co_idx, _new_sku) in list(dynamic_co_tracker.items()):
                if cur_shift_global == _co_idx + 1:
                    # CO used remaining time of shift _co_idx; press now produces.
                    if _PRESS_RETURN_BLOCK:
                        _left = press_state.get(_press, {}).get("sku")
                        if _left and _left != _new_sku:
                            press_ran[_press].add(_left)   # press has now left _left
                    press_count[_new_sku] = press_count.get(_new_sku, 0) + 1
                    press_state[_press]   = {"sku": _new_sku, "status": "RUNNING"}
                    curing_allowable[_new_sku].append(_press)
                    del dynamic_co_tracker[_press]

            # ── CAMPAIGN active-set: per-SKU presses beyond the day's target are IDLE this shift
            # (excess vs building supply). Deterministic (sorted). Skipped in draw + curing below.
            _camp_idle: set = set()
            if _campaign_on:
                _by_sku_now: dict = defaultdict(list)
                for _pr, _st in press_state.items():
                    if _st.get("status") == "RUNNING" and _pr not in co_press_map:
                        _by_sku_now[_st["sku"]].append(_pr)
                for _sk, _prs in _by_sku_now.items():
                    _tgt = _campaign_target_b2c.get(_sk, {}).get(day, 0)
                    for _pr in sorted(_prs)[_tgt:]:
                        _camp_idle.add(_pr)

            # ── 1. Per-shift curing demand (which SKUs need GT this shift) ──
            shift_cure_demand: dict[str, float] = defaultdict(float)
            for press, st in press_state.items():
                if press in _camp_idle:
                    continue                          # campaign-idle: no draw
                if press in co_press_map:
                    # Before its CO shift the press still draws its OLD SKU; after it,
                    # the NEW SKU. On the CO shift itself it is idle (no draw).
                    _cs = co_shift_idx.get(press, 0)
                    _si = SHIFTS.index(shift)
                    if _si < _cs:
                        _osku = st["sku"]
                        _oct  = cure_ct_map.get(_osku, DEFAULT_CURING_CT)
                        shift_cure_demand[_osku] += _cure_qty_per_shift(_oct)
                    elif _si > _cs:
                        new_sku = co_press_map[press]
                        new_ct  = cure_ct_map.get(new_sku, DEFAULT_CURING_CT)
                        shift_cure_demand[new_sku] += _cure_qty_per_shift(new_ct)
                elif press in dynamic_co_tracker:
                    # Dynamic CO: press already in CHANGEOVER this shift or started RUNNING
                    # (RUNNING case handled by press_state update at top of shift).
                    # No explicit demand signal needed — if RUNNING, press_state reflects it.
                    pass
                elif st["status"] == "RUNNING":
                    sku = st["sku"]
                    ct  = cure_ct_map.get(sku, DEFAULT_CURING_CT)
                    shift_cure_demand[sku] += _cure_qty_per_shift(ct)

            # Pre-build signal: in Shift A only, inject anticipated demand for all CO target
            # SKUs so building starts pre-building GT before their presses fire in Shift B.
            # Bug fix: original code injected 1× qty_per_shift per UNIQUE target SKU,
            # regardless of how many presses are CO'ing to it.  For SURL0 (3 CO presses),
            # this produced d = 1×56×2 = 112 — far too low to compete with a 10-press RI
            # SKU (d=1120).  Building never served SURL0 in Campaign 2.
            # Fix: count actual CO presses per target SKU and inject n_presses × qty,
            # matching how RUNNING presses accumulate their demand signal.
            # Injected on each press's OWN CO shift (not blanket Shift A), so building
            # still starts the new SKU's GT simultaneously with that press's changeover
            # — the simultaneity rule — now that COs are spread across A/B/C.
            _co_press_counts: dict[str, int] = defaultdict(int)
            for _cp, _cp_sku in co_press_map.items():
                if SHIFTS.index(shift) == co_shift_idx.get(_cp, 0):
                    _co_press_counts[_cp_sku] += 1
            for new_sku, n_co in _co_press_counts.items():
                new_ct = cure_ct_map.get(new_sku, DEFAULT_CURING_CT)
                shift_cure_demand[new_sku] += _cure_qty_per_shift(new_ct) * n_co

            # ── 2. Building assignment for this shift ──────────────────────
            # HOLIDAY: no machine builds (empty eligibility → empty shift_plan). State
            # (machine_current_sku etc.) is untouched → machines resume the same SKU with
            # no CO on the next working shift.
            _ms_capped = ({} if _holiday else
                          ({k: v for k, v in machine_skus.items()
                            if not (k in _PS_MAX_BUILD and _ps_built_sofar[k] >= _PS_MAX_BUILD[k])}
                           if _PS_MAX_BUILD else machine_skus))   # drop ps3/ps4 once at monthly cap
            # ── Holiday fix #2/#3 scalars (computed HERE where the holiday set is visible; the
            #    building fn is module-level and cannot see _holiday_days). Both inert unless a
            #    holiday falls inside the shelf window. ──
            _si_now = SHIFTS.index(shift)
            _fwd_work_shifts = None
            if _HOLIDAY_NO_PERISH:
                _fwd_work_shifts = 0
                for _k in range(GT_SHELF_LIFE_SHIFTS):          # working shifts in the next 9 calendar shifts
                    _abs = _si_now + 1 + _k
                    _dd  = day + (_abs // 3)
                    if _dd > planning_days:
                        break
                    if not _is_holiday(_dd):
                        _fwd_work_shifts += 1
            _bridge_shifts = 0
            if _HOLIDAY_BRIDGE:
                _k = 0                                          # count consecutive holiday shifts starting next shift
                while True:
                    _abs = _si_now + 1 + _k
                    _dd  = day + (_abs // 3)
                    if _dd > planning_days or not _is_holiday(_dd):
                        break
                    _bridge_shifts += 1; _k += 1
                # shelf guard: GT built now must survive to the first post-holiday shift (≤ 9 shifts ≈ 3 days)
                if _bridge_shifts + (2 - _si_now) > GT_SHELF_LIFE_SHIFTS:
                    _bridge_shifts = 0
            # #7 HOLIDAY_SHIFTC_CAP: cap day-D Shift C to the 23:00→00:00 pre-midnight window when
            # day D+1 is a holiday, so no building production / new CO lands on the holiday.
            _shift_budget = (
                _HOLIDAY_SHIFTC_CAP_MINS
                if (_HOLIDAY_SHIFTC_CAP and shift == "C"
                    and (day + 1) <= planning_days and (day + 1) in _holiday_days)
                else SHIFT_MINS)
            # MONOTONICITY FIX (BLD_DRAW_CAP): clamp the per-inch aggregate draw the building
            # assigner sees to the inch's building GT/shift capacity, so extra presses on a
            # saturated inch cannot inflate the signal and pull machines off productive inches.
            _scd_for_build = dict(shift_cure_demand)
            if _BLD_DRAW_CAP and _building_inch_capacity:
                _inch_tot: dict = defaultdict(float)
                for _s, _v in _scd_for_build.items():
                    if _v > 0:
                        _si = sku_inch.get(str(_s)) or (str(_s)[8:10] if len(str(_s)) >= 10 else "")
                        _inch_tot[_si] += _v
                for _i, _tot in _inch_tot.items():
                    _cap_shift = _building_inch_capacity.get(_i, 0.0) / 3.0   # GT/day → GT/shift
                    if _cap_shift > 0 and _tot > _cap_shift:
                        _scale = _cap_shift / _tot
                        for _s in _scd_for_build:
                            if _scd_for_build[_s] > 0 and (
                                    sku_inch.get(str(_s)) or (str(_s)[8:10] if len(str(_s)) >= 10 else "")) == _i:
                                _scd_for_build[_s] *= _scale
            # PM/MTC: minutes each building machine is in maintenance during THIS shift's wall-clock.
            # Under _PM_MTC_NO_OVERLAP (default ON), production resumes STRICTLY AFTER the last
            # maintenance window in the shift — the machine's usable minutes = the POST-maintenance
            # free time only (the pre-maintenance gap is NOT used; a deferred remainder carries to a
            # later shift as ordinary unmet deficit and is rebuilt there, so GT ages from its REAL
            # post-maintenance build day). `_mprod_start` records where emission must begin.
            _mdown = {}
            _mprod_start = {}
            if _PM_MTC_ENABLED and _BLD_DOWN:
                _sh_start = date + timedelta(hours=8 * _si_now)
                _sh_end   = _sh_start + timedelta(hours=8)
                for _dm, _win in _BLD_DOWN.items():
                    _dn = _down_mins(_win, _sh_start, _sh_end)
                    if _dn <= 0:
                        continue
                    if _PM_MTC_NO_OVERLAP:
                        # Usable = the largest window-free block in this shift (pre- OR post-window;
                        # PM_MTC_PREWINDOW). Emission starts at that block so no row overlaps a window.
                        _post_free, _post_start = _pm_maint_free(_win, _sh_start, _sh_end)
                        _mdown[_dm] = float(SHIFT_MINS) - _post_free   # usable = largest free block
                        _mprod_start[_dm] = _post_start
                    else:
                        _mdown[_dm] = _dn
            # ── 2-DAY PLANT PLAYBACK: for plan days 1-2, BYPASS the normal building
            # assignment and force the plant's exact day-0 snapshot (GT machines here;
            # Stage-1 carcass injected post-plan). Demand cap ignored; qty already capped
            # to the shift's physical capacity AND, where PM/MTC intrudes, to the
            # maintenance-free capacity (maintenance beats the plant seed — see
            # _plant_2day_gt_plan's docstring) in _plant_2day_gt_plan. The Day-1 Shift-A
            # seed pin is superseded here (its assigner arg / Stage-1 block are gated off).
            _replay_now = _PLANT_2DAY_REPLAY and day <= _PLANT_2DAY_DAYS and not _holiday
            if _replay_now:
                shift_plan = _plant_2day_gt_plan(day, shift, machine_current_sku,
                                                  machine_down_mins=_mdown)
                # accumulate ps monthly total so day-3+ still honours PS_MAX_BUILD (no clamp on days 1-2)
                for _pm in _PS_MAX_BUILD:
                    _pp = shift_plan.get(_pm)
                    if _pp:
                        _ps_built_sofar[_pm] += sum(_q for (_s, _q, _c) in _pp)
                # SAME_GROUP: record which finer group each GT SKU was built in during the
                # 2-day plant replay (drives the home-group choice, frozen at day 3).
                if _SAME_GROUP_SOFT:
                    for _mm, _rows in shift_plan.items():
                        _g = _sku_group_of(_mm)
                        if _g not in _SG_GT_GROUPS:
                            continue
                        for (_s, _q, _c) in _rows:
                            if _q > 0:
                                _sg_d12_units[(_s, _g)] += _q
            else:
                # SAME_GROUP: freeze home groups ONCE, at the first day-3 assignment (after the
                # full 2-day replay is accumulated). Inert (map stays None) when the lever is OFF.
                if _SAME_GROUP_SOFT and _sku_home_group is None and day > _PLANT_2DAY_DAYS:
                    _sku_home_group = _compute_home_groups(day)
                # SG_DELIB: freeze the DELIBERATE target group SETS once, at day 3 (after the
                # 2-day replay). Seeds _sku_cur_group from the per-SKU seed group. Inert when OFF.
                if _SG_DELIB and _sku_grp_target is None and day > _PLANT_2DAY_DAYS:
                    _sku_grp_target, _sg_seed = _compute_grp_targets(day)
                    _sku_cur_group.update(_sg_seed)
                shift_plan = _assign_building_shift(
                    shift_cure_demand=_scd_for_build,
                    machine_down_mins=_mdown,
                    machine_skus=_ms_capped,
                    machine_current_sku=machine_current_sku,
                    sku_inch=sku_inch,
                    demand_remaining=demand_remaining,
                    gt_inventory=gt_inventory,
                    machine_pool=machine_pool,
                    machine_minutes_on_sku=machine_minutes_on_sku,
                    cure_ct_map=cure_ct_map,
                    press_count=press_count,
                    co_target_skus=co_target_skus_today,
                    days_left=_working_days_left(day),   # holiday-aware urgency horizon
                    fwd_work_shifts=_fwd_work_shifts,     # #2 NO-PERISH: shelf window capped to working shifts
                    bridge_shifts=_bridge_shifts,         # #3 BRIDGE: pre-build through an imminent holiday
                    shift_budget_mins=_shift_budget,      # #7 HOLIDAY_SHIFTC_CAP: cap pre-holiday Shift C at midnight
                    writeoff_cum=writeoff_cum,           # R8B: cap tightener (built <= demand)
                    demand_dict=demand_dict,
                    machine_total_demand=machine_total_demand,
                    machine_anchor_inch=machine_anchor_inch,
                    machine_used_inches=machine_used_inches,
                    machine_left_skus=machine_left_skus,
                    machine_inch_since=machine_inch_since,
                    day=day,
                    machine_day_skus=machine_day_skus,
                    machine_plus3_used=machine_plus3_used,
                    machine_last_diff_co_day=machine_last_diff_co_day,
                    machine_locked_inches=machine_locked_inches,
                    machine_day_diff_co=machine_day_diff_co,
                    machine_day_co=machine_day_co,
                    fixed_escape_used=fixed_escape_used,
                    machine_step_drift=machine_step_drift,      # INCH_STEP_DRIFT state
                    machine_db_skus=_machine_db_skus,           # INCH_STEP_DRIFT: un-stripped DB eligibility
                    lookahead_draw=_lookahead_draw,             # LOOKAHEAD_BUF: anticipated peak draw
                    priority_deadline_map=(priority_deadline_map if _prio_active else None),
                    sku_curable_ceiling=_sku_curable_ceiling,   # BLD_CURABLE_CAP: no-waste-GT bound
                    shift_idx=_si_now,                           # BLD_SEED_PIN_D1A: Day-1 Shift-A seed pin
                    # PACING: GT built earlier today + SKUs already served today (flat cap + widen)
                    pacing_day_built=float(sum(day_gt_built.values())),
                    pacing_day_skus={_s for _s, _q in day_gt_built.items() if _q > 0},
                    sku_home_group=_sku_home_group,   # SAME_GROUP soft one-SKU→one-group lever
                    sku_grp_target=_sku_grp_target,   # SG_DELIB: deliberate target group SETS
                    sku_cur_group=_sku_cur_group,     # SG_DELIB: current group per SKU (mutable)
                    sku_last_group_move=_sku_last_group_move,  # SG_DELIB: last group-move day (cooldown)
                    machine_plant_set=_get_machine_plant_set(),  # PLANT_SET_LOCK: plant Days-1-2 SKU set per GT machine
                )

            # ps3/ps4 hard monthly cap: clamp this shift's ps build to the remaining room so the
            # month total never exceeds PS_MAX_BUILD, then accumulate. SKIPPED on days 1-2 replay
            # (the plant snapshot is followed exactly; its ps totals were already accumulated above).
            for _pm in ([] if _replay_now else _PS_MAX_BUILD):
                _room = _PS_MAX_BUILD[_pm] - _ps_built_sofar[_pm]
                _plan = shift_plan.get(_pm)
                if _plan:
                    if sum(_q for (_s, _q, _c) in _plan) > _room:
                        _new = []; _acc = 0.0
                        for (_s, _q, _c) in _plan:
                            if _acc >= _room:
                                break
                            _take = min(_q, _room - _acc)
                            if _take > 0:
                                _new.append((_s, _take, _c)); _acc += _take
                        shift_plan[_pm] = _new
                    _ps_built_sofar[_pm] += sum(_q for (_s, _q, _c) in shift_plan.get(_pm, []))

            _shift_start = _shift_start_dt(date_str, shift)

            def _s1_inch_ok(_m: str, _sku: str) -> bool:
                """Client inch rules applied to Stage-1 carcass machines too.

                Same anchor +/- band / single-inch lock as the GT machines. Used by
                BOTH the Stage-2 carcass gate (step 2c) and the post-hoc Step-3b pass.
                """
                if not _INCH_RULES_ENABLED:
                    return True
                if (_INCH_HIST_LOCK_ENABLED and _INCH_HIST_LOCK_STAGE1
                        and str(_m) in _MACHINE_ALLOWED_INCH_SET):
                    return sku_inch.get(_sku, "") in _MACHINE_ALLOWED_INCH_SET[str(_m)]
                if _STAGE1_SINGLE_INCH and _m in s1_locked_inch:
                    # S1_INCH_FLEX: honor the machine's CURRENT inch (which may have advanced
                    # one-way off a completed inch); else the fixed day-0 lock.
                    _lk = (s1_current_inch.get(_m, s1_locked_inch[_m])
                           if _S1_INCH_FLEX else s1_locked_inch[_m])
                    return sku_inch.get(_sku, "") == _lk
                return _inch_ok(sku_inch.get(_sku, ""),
                                s1_current_inch.get(_m, ""),
                                machine_anchor_inch.get(_m, ""),
                                machine_used_inches.get(_m, set()))

            # ── 2c. Stage-2 carcass GATE (hard constraint, aging-aware) ──────
            # Cap Stage-2 GT each shift by feasible carcass using a rolling per-SKU
            # carcass BANK (list of [age_shifts_left, qty]): Stage-1 builds this
            # shift's shortfall + PRE-BUILDS ahead with residual capacity, bounded by
            # the <=1-day shelf (CARCASS_SHELF_LIFE_DAYS x 3 shifts). Stage-2 draws
            # from the bank and WAITS only for carcass that truly cannot be built
            # within the aging window. Clamp ONLY (no rows/inch here) — the carcass
            # rows are finalized by the post-plan max-flow (_stage1_carcass_schedule)
            # and Step-3b; this runs BEFORE Step 3 so curing sees only backed GT.
            # SKIPPED on days 1-2 replay: the plant's exact Stage-2 GT must not be
            # carcass-clamped (its Stage-1 carcass is force-injected post-plan).
            if _STAGE2_CARCASS_GATE and not _replay_now:
                _cage = max(1, _STAGE1_CARCASS_LEAD + 1)      # usable window (shifts)
                for _bs in list(_carcass_bank):                  # age the bank 1 shift
                    _exp = sum(_q for (_a, _q) in _carcass_bank[_bs]      # lots aging out this shift = WASTE
                               if not (_a - 1 > 0 and _q > 1e-9))
                    if _exp > 0:
                        carcass_waste[_bs] += _exp
                        day_carcass_writeoff += _exp
                        if _exp >= 0.5:
                            _cst = _fmt_dt(_shift_start_dt(date_str, shift))
                            expiry_rows.append({"Machine": "—", "Date": date_str, "Shift": shift,
                                                "SKUCode": _bs, "Qty": int(round(_exp)), "CO_Mins": 0,
                                                "StartTime": _cst, "EndTime": _cst,
                                                "Machine_Group": "", "CO_Type": "expired_carcass"})
                    _kept = [[_a - 1, _q] for (_a, _q) in _carcass_bank[_bs]
                             if _a - 1 > 0 and _q > 1e-9]
                    if _kept:
                        _carcass_bank[_bs] = _kept
                    else:
                        del _carcass_bank[_bs]
                _s2_desired: dict[str, float] = defaultdict(float)
                for _gm, _gcps in shift_plan.items():
                    if _MACHINE_GROUP.get(_gm, "") == "STAGE2":
                        for _gs, _gq, _gct in _gcps:
                            _s2_desired[_gs] += _gq
                # Per-Stage-1-machine REMAINING carcass capacity this shift. A machine
                # may SPLIT its shift capacity across multiple SKUs (exactly like the
                # plant carcass max-flow, source->machine-shift capped at CAP[m]). The
                # old one-SKU-per-machine rule under-supplied carcass and made Stage-2
                # wait needlessly — that was the bulk of the avoidable loss.
                _gate_cap: dict = {}
                # min-carcass rule: cumulative carcass built per (Stage-1 machine, SKU) THIS
                # shift, across all _gate_build calls (PASS 1 + PASS 2). Reconciled to the
                # MIN floor right before the GT clamp below.
                _carc_ms: dict = defaultdict(float)

                def _s1cap(_m):
                    # BUGFIX: track remaining capacity in MINUTES (was units at the machine-DEFAULT CT,
                    # which over-assigned machines whose carcass SKU is slower than the default — e.g.
                    # 7701 default 163s vs real 261.8s → 133% util). Building a unit of SKU-s now costs
                    # its per-(SKU,machine) CT minutes, so no machine can exceed its true minute budget.
                    if _m not in _gate_cap:
                        _gate_cap[_m] = float(_shift_budget)          # remaining MINUTES this shift
                    return _gate_cap[_m]

                def _bank_avail(_sku):
                    return sum(_q for (_a, _q) in _carcass_bank.get(_sku, ()))

                def _co_units(_m, _sku):
                    """STAGE1_CO: capacity (in units) a Stage-1 machine loses to a
                    building CO when it switches TO _sku from its current carcass SKU.
                    0 when the machine is already on _sku or never assigned ("" = free
                    first start, like a GT machine's 'start'). 60 min same-inch / 180
                    diff-inch, converted to units via the machine's CT."""
                    _cur = machine_cur_carcass.get(_m, "")
                    if _cur in ("", _sku):
                        return 0.0
                    _comin = _co_cost(_m, sku_inch.get(_cur, ""), sku_inch.get(_sku, ""))
                    _ct = _bld_ct_sec(_m, _sku)
                    return (_comin * 60.0 / _ct) if _ct > 0 else 0.0

                def _gate_build(_sku, _target):
                    """Build <=_target carcass for _sku on its eligible, inch-OK Stage-1
                    machines, drawing each machine's REMAINING shift capacity (split
                    across SKUs); bank it fresh. When _STAGE1_CO, a machine switching to a
                    different carcass SKU first pays a building CO (no production during
                    it), and machines already set up for _sku are preferred (consolidation
                    → long carcass campaigns)."""
                    if _target <= 0:
                        return 0.0
                    _elig = [m for m in s1_sku_to_machines.get(_sku, ())
                             if _s1_inch_ok(m, _sku) and _s1cap(m) > 1e-9]
                    if _STAGE1_CO and _CV2_A2:
                        # A2: cheapest-changeover-first — continuation 0 / same-inch 60 /
                        # diff-inch 180 (real CO minutes), then most spare capacity, then id.
                        def _co_min_for(_m):
                            _p = machine_cur_carcass.get(_m, "")
                            if _p in ("", _sku):
                                return 0.0
                            return float(_co_cost(_m, sku_inch.get(_p, ""), sku_inch.get(_sku, "")))
                        _elig.sort(key=lambda m: (_co_min_for(m), -_s1cap(m), m))
                    elif _STAGE1_CO:
                        # co-free machines (already on _sku or unused) first, then most cap
                        _elig.sort(key=lambda m: (1 if _co_units(m, _sku) > 1e-9 else 0,
                                                  -_s1cap(m), m))
                    else:
                        _elig.sort(key=lambda m: (-_s1cap(m), m))
                    # BLD_SEED_PIN_D1A (Stage-1 best-effort): on Day-1 Shift-A prefer the Stage-1
                    # machine whose SEED SKU == this carcass SKU so it aligns with the plant Day-0
                    # snapshot; push a machine seeded to a DIFFERENT SKU to the back (freeing it to
                    # serve its own seed when that SKU is built this shift). Stable re-sort → keeps
                    # the prior ordering within each preference group. A Stage-1 machine can only
                    # align when its seed SKU actually has Stage-2 GT built this shift (carcass is
                    # derived) — where it isn't, alignment is physically impossible (best-effort).
                    if _BLD_SEED_PIN_D1A and day == 1 and shift == "A":
                        def _seed_pref(_m):
                            _sd = _BLD_SEED_SKU.get(str(_m), "")
                            if _sd == _sku:
                                return 0
                            return 2 if str(_m) in _BLD_SEED_MACHINES else 1
                        _elig.sort(key=_seed_pref)
                    _got = 0.0
                    for _m in _elig:
                        if _got >= _target:
                            break
                        _ctm = _bld_ct_sec(_m, _sku)               # per-(SKU,machine) CT (sec/unit)
                        if _ctm <= 0:
                            continue
                        if _STAGE1_CO:
                            _prev = machine_cur_carcass.get(_m, "")
                            if _prev not in ("", _sku):            # real switch → building CO (minutes)
                                _comin = float(_co_cost(_m, sku_inch.get(_prev, ""),
                                                        sku_inch.get(_sku, "")))
                                if _s1cap(_m) <= _comin + 1e-9:
                                    continue          # not enough time even for the CO
                                _gate_cap[_m] -= _comin  # charge CO minutes (no production)
                            machine_cur_carcass[_m] = _sku
                        _max_u = _s1cap(_m) * 60.0 / _ctm          # units the remaining MINUTES allow
                        _a = min(_max_u, _target - _got)
                        _gate_cap[_m] -= _a * _ctm / 60.0          # deduct production minutes
                        _got += _a
                        _carc_ms[(_m, _sku)] += _a                  # min-carcass: per-(machine,SKU) shift total
                        if _STAGE1_CO and _a > 0:      # log for Site 2 (rows built later)
                            _s1_prod_log.append({
                                "day": day, "date": date_str, "shift": shift,
                                "machine": _m, "sku": _sku, "qty": _a,
                            })
                    if _got > 0:
                        _carcass_bank[_sku].append([_cage, _got])
                    return _got

                # ── S1_INCH_FLEX: SURPLUS→SCARCE one-way inch advance (client rule) ──
                # A Stage-1 machine takes a DIFF-size CO to another inch only when (1) it is
                # SURPLUS on its own inch — the OTHER machines there can cover that inch's
                # remaining Stage-2 carcass demand for the rest of the horizon without it —
                # (2) a genuinely SCARCE eligible inch exists (whose machines cannot cover
                # their remaining demand), and (3) it has NOT been on that inch before
                # (one-way, never revert). It moves to the scarcest such inch. Because moves
                # are surplus-only, toward-scarcer-only, and never-revisit, each machine takes
                # ≤ (#eligible inches) diff-COs all month → Stage-1 CO stays bounded.
                if _S1_INCH_FLEX:
                    _dl = max(1, _working_days_left(day))          # working shifts left ≈ days_left×3 (holiday-aware)
                    _inch_rem: dict = defaultdict(float)
                    for _s2s in s1_sku_to_machines:
                        _ii = sku_inch.get(str(_s2s), "")
                        if _ii:
                            _inch_rem[_ii] += max(0.0, demand_remaining.get(str(_s2s), 0.0))
                    _mach_on: dict = defaultdict(list)
                    for _m in _S1_MACHINES:
                        _mach_on[s1_current_inch.get(_m, s1_locked_inch.get(_m, ""))].append(_m)

                    def _cap_left(_ms):
                        return sum(_bld_qty_per_shift(x) for x in _ms) * 3 * _dl

                    # S1_DIFF_CO loosens "scarce" to "TIGHT" (coverage < demand×margin, not just fully
                    # uncoverable) so idle Stage-1 relieves local carcass clamps; revisit allowed, but a
                    # per-machine monthly cap keeps diff-COs "optimal, not too much".
                    _s1_margin = _S1_TIGHT_MARGIN if _S1_DIFF_CO else 1.0
                    def _scarce(_i):                               # machines on _i can't cover it (×margin)
                        return _cap_left(_mach_on.get(_i, [])) < _inch_rem.get(_i, 0.0) * _s1_margin - 1e-9

                    for _m in sorted(_S1_MACHINES):
                        if _S1_DIFF_CO and _s1_diff_co_count[_m] >= _S1_MAX_DIFF_CO:
                            continue                               # spent its monthly diff-CO budget
                        _ci = s1_current_inch.get(_m, s1_locked_inch.get(_m, ""))
                        _peers = [x for x in _mach_on.get(_ci, []) if x != _m]
                        if _cap_left(_peers) < _inch_rem.get(_ci, 0.0) - 1e-9:
                            continue                               # NOT surplus → m needed here
                        _cands = [(_inch_rem.get(_i, 0.0) / max(1, len(_mach_on.get(_i, []))), _i)
                                  for _i in _s1_elig_inches.get(_m, ())
                                  if _i != _ci and (_S1_DIFF_CO or _i not in _s1_visited[_m])
                                  and _scarce(_i)]
                        if _cands:
                            _best = max(_cands, key=lambda t: (t[0], t[1]))[1]
                            s1_current_inch[_m] = _best
                            _s1_visited[_m].add(_best)
                            _s1_diff_co_count[_m] += 1
                            _mach_on[_ci].remove(_m); _mach_on[_best].append(_m)

                # PASS 1 — cover this shift's Stage-2 need beyond bank carry-in.
                # A1: most-constrained-SKU-first (fewest eligible Stage-1 machines) so scarce
                # SKUs claim their machines before flexible SKUs take them; else biggest-need-first.
                # HOLIDAY: no NEW Stage-1 carcass production on a holiday plan-day (0 building day,
                # same as GT) — gated by `_holiday` (`_is_holiday(day)`, the SAME plan-day-index
                # mechanism GT building/curing already use, NOT calendar date). The bank
                # aging/writeoff loop above still runs every shift (carcass shelf life stays
                # calendar-based across the holiday, matching GT). _s2_desired is normally already
                # empty on a holiday (Stage-2 building is capped via `_ms_capped`), but PASS 2's
                # `shift_cure_demand`-driven candidate set is NOT — this guard is what actually
                # stops that pre-build from firing on the holiday.
                if not _holiday:
                    if _CV2_A1:
                        _pass1_iter = sorted(_s2_desired.items(),
                                             key=lambda kv: (len(s1_sku_to_machines.get(kv[0], ())), kv[0]))
                    else:
                        _pass1_iter = sorted(_s2_desired.items(), key=lambda kv: (-kv[1], kv[0]))
                    for _gs, _gneed in _pass1_iter:
                        _short = _gneed - _bank_avail(_gs)
                        if _short > 0:
                            _gate_build(_gs, _short)
                # PASS 2 — pre-build residual capacity toward a <=1-day buffer for SKUs
                # the presses are actively pulling (so upcoming bursts stay backed).
                # A DISTRIBUTED buffer (rate x window, not full-capacity per SKU) banks
                # for more SKUs and beats concentrating capacity on the top SKU.
                if _STAGE2_CARCASS_PREBUILD and not _holiday:
                    # STAGE1_CO: bound the pre-build so the carcass held beyond this shift's
                    # own consumption (the buffer that can carry overnight) stays under
                    # MAX_ENDOFDAY_CARCASS_INVENTORY every shift. Measured NET-best across
                    # months vs an EOD-only bound (the every-shift bound keeps the carcass
                    # allocation balanced, which cascades to better Stage-2 coverage).
                    _desired_total = sum(_s2_desired.values())
                    _buf_cap = (_desired_total + _MAX_EOD_CARCASS) if _STAGE1_CO else float("inf")
                    def _bank_total():
                        return sum(_q for _es in _carcass_bank.values() for (_a, _q) in _es)
                    _cand = set(_s2_desired)
                    _cand |= {s for s in s1_sku_to_machines
                              if shift_cure_demand.get(s, 0.0) > 0 and demand_remaining.get(s, 0.0) > 0}
                    for _gs in sorted(_cand, key=lambda s: (-(shift_cure_demand.get(s, 0.0)
                                                             + _s2_desired.get(s, 0.0)), s)):
                        # Build-to-consumption: pre-build to the Stage-2 BUILD rate only (the true
                        # carcass consumer). The old max(...) added the full CURING DRAW, which
                        # counts BJ/Unistage GT that needs NO carcass → pre-built a surplus that
                        # aged out and churned. _s2_desired is exactly what Stage-2 will consume.
                        _rate = (_s2_desired.get(_gs, 0.0) if _CARCASS_NO_OVERBUILD
                                 else max(_s2_desired.get(_gs, 0.0), shift_cure_demand.get(_gs, 0.0)))
                        if _rate <= 0:
                            continue
                        # #2 NO-PERISH: cap the carcass pre-build LEAD to the WORKING shifts before a
                        # holiday. Carcass has a 1-day shelf, so anything banked for a holiday shift
                        # (no Stage-2 runs) strands and is written off. None-adjacent → full LEAD (parity).
                        _eff_lead = _STAGE1_CARCASS_LEAD
                        if _HOLIDAY_NO_PERISH:
                            _si_c = SHIFTS.index(shift); _work_life = 0
                            for _kk in range(_cage):
                                _absc = _si_c + 1 + _kk
                                _ddc  = day + (_absc // 3)
                                if _ddc > planning_days:
                                    break
                                if not _is_holiday(_ddc):
                                    _work_life += 1
                            _eff_lead = max(0, min(_STAGE1_CARCASS_LEAD, _work_life - 1))
                        _buf = min(_rate * (1 + _eff_lead),
                                   max(0.0, demand_remaining.get(_gs, 0.0)))
                        _extra = _buf - _bank_avail(_gs)
                        if _STAGE1_CO:                       # keep the carryable buffer ≤ cap
                            _extra = min(_extra, max(0.0, _buf_cap - _bank_total()))
                        if _extra > 0:
                            _gate_build(_gs, _extra)
                # ── HARD min-carcass floor (per machine, SKU, shift) ─────────────
                # Any per-(Stage-1 machine, SKU) carcass built < MIN this shift is a
                # sub-MIN fragment (over-production leftover) → DROP it: remove those units
                # from the carcass bank (freshest lots = this shift's build) AND from the
                # display log (_s1_prod_log) so no sub-MIN carcass row is emitted. The GT
                # clamp immediately below then makes Stage-2 GT only to the carcass that
                # REMAINS available → GT + curing reduce in sync (carcass-first order).
                if _CARCASS_MIN_ENFORCE and _CARCASS_MIN_QTY > 0:
                    _drop_ms = [(_m, _sku) for (_m, _sku), _q in _carc_ms.items()
                                if 1e-9 < _q < _CARCASS_MIN_QTY]
                    for (_m, _sku) in _drop_ms:
                        _need = _carc_ms[(_m, _sku)]
                        # remove from bank, freshest (highest age-left) lots first = this shift
                        for _lot in sorted(_carcass_bank.get(_sku, []),
                                           key=lambda l: -l[0]):
                            if _need <= 1e-9:
                                break
                            _d = min(_lot[1], _need); _lot[1] -= _d; _need -= _d
                        if _sku in _carcass_bank:
                            _carcass_bank[_sku] = [l for l in _carcass_bank[_sku] if l[1] > 1e-9]
                            if not _carcass_bank[_sku]:
                                del _carcass_bank[_sku]
                        _carc_ms[(_m, _sku)] = 0.0
                    if _drop_ms:
                        _dkeys = set(_drop_ms)
                        # drop this shift's display-log entries for the dropped (machine,SKU)
                        _s1_prod_log[:] = [
                            _e for _e in _s1_prod_log
                            if not (_e.get("day") == day and _e.get("shift") == shift
                                    and (_e.get("machine"), _e.get("sku")) in _dkeys)]
                # CLAMP Stage-2 GT per SKU to the bank; consume FIFO (oldest first).
                # HARD min-carcass: if the carcass a SKU can draw this shift is a sub-MIN sliver,
                # don't back GT with it → drop the GT (and thus curing) for that SKU this shift, so
                # no <MIN carcass is consumed and GT/carcass/curing stay in sync (carcass-first).
                _take: dict[str, float] = {}
                for _gs in _s2_desired:
                    _t = min(_s2_desired[_gs], _bank_avail(_gs))
                    if _CARCASS_MIN_ENFORCE and _CARCASS_MIN_QTY > 0 and 0 < _t < _CARCASS_MIN_QTY:
                        _t = 0.0
                    _take[_gs] = _t
                    _rem = _t
                    for _entry in _carcass_bank.get(_gs, []):
                        if _rem <= 0:
                            break
                        _use = min(_entry[1], _rem)
                        _entry[1] -= _use
                        _rem -= _use
                    if _gs in _carcass_bank:
                        _carcass_bank[_gs] = [[_a, _q] for (_a, _q) in _carcass_bank[_gs] if _q > 1e-9]
                for _gm in list(shift_plan):
                    if _MACHINE_GROUP.get(_gm, "") != "STAGE2":
                        continue
                    _gnew = []
                    for (_gs, _gq, _gct) in shift_plan[_gm]:
                        _des = _s2_desired.get(_gs, 0.0)
                        _gq2 = _gq * (_take.get(_gs, 0.0) / _des) if _des > 1e-9 else 0.0
                        _gnew.append((_gs, _gq2, _gct))
                    shift_plan[_gm] = _gnew
                if _STAGE1_CO and shift == "C":   # end-of-day carcass buffer carried overnight
                    # HARD cap (like the 8k GT cap): the carcass carried overnight must be
                    # ≤ MAX_ENDOFDAY_CARCASS_INVENTORY. Trim the soonest-to-expire excess
                    # (lowest age-remaining first — it would age out first anyway).
                    _tot = sum(_q for _es in _carcass_bank.values() for (_a, _q) in _es)
                    _excess = _tot - _MAX_EOD_CARCASS
                    if _excess > 1e-9:
                        # Trim the soonest-to-expire excess first (lowest age-remaining — it
                        # would age out first anyway). Rarely binds under the every-shift cap.
                        _order = sorted(((_sk, _i) for _sk, _es in _carcass_bank.items()
                                         for _i in range(len(_es))),
                                        key=lambda t: (_carcass_bank[t[0]][t[1]][0], t[0], t[1]))
                        for _sk, _i in _order:
                            if _excess <= 1e-9:
                                break
                            _drop = min(_carcass_bank[_sk][_i][1], _excess)
                            _carcass_bank[_sk][_i][1] -= _drop
                            _excess -= _drop
                        for _sk in list(_carcass_bank):
                            _carcass_bank[_sk] = [[_a, _q] for (_a, _q) in _carcass_bank[_sk]
                                                  if _q > 1e-9]
                            if not _carcass_bank[_sk]:
                                del _carcass_bank[_sk]
                    _carcass_eod.append(sum(_q for _es in _carcass_bank.values()
                                            for (_a, _q) in _es))

            # ── 2b. Idle-recoverability diagnostic (read-only, plan-neutral) ──
            if _IDLE_DIAG_ON:
                _SM = float(SHIFT_MINS)
                # GT built THIS shift per SKU (non-Stage-1 only → real curable GT).
                _built_shift: dict[str, float] = defaultdict(float)
                _used_min: dict[str, float] = {}
                for _m, _cps in shift_plan.items():
                    _pi = sku_inch.get(machine_current_sku.get(_m, ""), "")
                    _um = 0.0
                    for _s, _q, _ct_type in _cps:
                        if _ct_type not in ("start", "production"):
                            _um += _co_cost(_m, _pi, sku_inch.get(_s, ""))
                        _um += _q * _bld_ct_sec(_m, _s) / 60.0
                        _pi = sku_inch.get(_s, "")
                        if _m not in _S1_MACHINES:
                            _built_shift[_s] += _q
                    _used_min[_m] = _um
                # projected GT after this shift's building (entry inv + built this shift).
                def _pg_after(_s):
                    return gt_inventory.get(_s, 0.0) + _built_shift.get(_s, 0.0)
                # Idle machines (>= a legal campaign of idle time left this shift).
                _idle_machs = []
                for _m in machine_skus:
                    _um = _used_min.get(_m, 0.0)   # machines absent from shift_plan = fully idle
                    _idle = max(0.0, _SM - _um)
                    if _m in _S1_MACHINES:
                        _IDLE_DIAG["s1_idle_min"] += _idle
                        continue
                    _IDLE_DIAG["gt_idle_min"] += _idle
                    if _idle >= MIN_CAMPAIGN_MINS:
                        _cps = shift_plan.get(_m)
                        _end_inch = (sku_inch.get(_cps[-1][0], "") if _cps
                                     else sku_inch.get(machine_current_sku.get(_m, ""), ""))
                        _idle_machs.append((_m, _idle, _end_inch))
                # Per-SKU: how many units the drawing press(es) will fail to cure this shift
                # for lack of GT (the momentary shortfall), and is it reachable by an idle machine?
                _rec_shifts = set(); _ceil_shifts = set()
                for _s, _draw in shift_cure_demand.items():
                    if _draw <= 0:
                        continue
                    _short = min(_draw, demand_remaining.get(_s, 0.0)) - _pg_after(_s)
                    if _short <= 0:
                        continue
                    _to = sku_inch.get(_s, "")
                    _hit = None
                    for _m, _idle, _cur in _idle_machs:
                        if _s not in machine_skus.get(_m, ()):
                            continue
                        if _to == _cur:
                            _reach = True
                        else:
                            _anc = machine_anchor_inch.get(_m, "")
                            _band = _inch_ok(_to, _cur, _anc, machine_used_inches.get(_m, set()))
                            _dwell = ((day - machine_inch_since.get(_m, day)) >= MIN_INCH_DWELL_DAYS
                                      or _inch_demand_done(_m, _cur, machine_skus, sku_inch,
                                                           None, 0, demand_remaining=demand_remaining,
                                                           projected_gt={k: _pg_after(k)
                                                                         for k in machine_skus.get(_m, ())}))
                            _reach = _band and _dwell
                        if _reach:
                            _hit = _m; break
                    if _hit is not None:
                        _IDLE_DIAG["rec_units"] += _short
                        _IDLE_DIAG["rec_by_inch"][_to] = _IDLE_DIAG["rec_by_inch"].get(_to, 0.0) + _short
                        _rec_shifts.add(_hit)
                    else:
                        _IDLE_DIAG["ceil_units"] += _short
                        _IDLE_DIAG["ceil_by_inch"][_to] = _IDLE_DIAG["ceil_by_inch"].get(_to, 0.0) + _short
                _IDLE_DIAG["rec_shifts"] += len(_rec_shifts)
                _IDLE_DIAG["ceil_shifts"] += len(_ceil_shifts)

            # ── 3. Add GT to inventory; record building rows + CO events ───
            # StartTime/EndTime: per-machine wall-clock cursor within the shift.
            # It starts at the shift's clock start and advances by each event's
            # own duration (CO minutes, then production minutes = qty × CT), so a
            # downstream scheduler reads an exact, non-overlapping timeline of
            # what each machine does and when. CT-per-unit = _BLD_CT_SEC[machine].
            _shift_start = _shift_start_dt(date_str, shift)
            for machine, campaigns in shift_plan.items():
                prev_sku  = machine_current_sku.get(machine, "")
                prev_inch = sku_inch.get(prev_sku, "")
                # PM/MTC POST-ONLY: a machine under maintenance this shift starts its CO/production
                # only AFTER the maintenance window ends (its assigned qty was already capped to the
                # post-maintenance minutes above), so the emitted timeline never touches the window.
                _cursor   = _mprod_start.get(machine, _shift_start)
                # 2-day playback: emitted replay rows are CONFINED to their own shift wall-clock
                # window [shift_start, shift_start+480]. The plant qty (min(plant, full-shift cap))
                # can exceed the shift when it also carries a CO, so a spilling row is CLAMPED at the
                # shift end (its span then reads short → R18B benign) rather than bleeding into the
                # next shift/day — this keeps day>=3 (the normal pipeline) feasibility-clean. The
                # Shift/Date/Qty columns stay the exact plant snapshot for the qty audit.
                _replay_shift_end = _shift_start + timedelta(minutes=SHIFT_MINS)
                for _tier_idx, (sku, qty, co_type) in enumerate(campaigns):
                    # MPQ (shift-level floor): skip a sub-floor building block — do NOT CO or
                    # build it. The un-built quantity persists as unmet deficit (GT not credited)
                    # and is naturally rebuilt >= floor in a later shift; a residual sub-MPQ tail
                    # stays unbuilt (unmet). Skipping only reduces build → never over-produces.
                    # (replay days 1-2 must reproduce the plant qty exactly → no MPQ skip)
                    if _BUILDING_MPQ > 0 and 0 < qty < _BUILDING_MPQ and not _replay_now:
                        continue
                    _ct_sec = _bld_ct_sec(machine, sku)   # per-SKU CT for the timeline
                    if co_type != "start" and prev_sku:
                        co_mins = (INCH_PLUS3_CO_MINS if co_type == "plus3_CO"
                                   else _co_cost(machine, prev_inch, sku_inch.get(sku, ""),
                                                 from_sku=prev_sku, to_sku=sku))
                        # SHIFT-CONTAINMENT (client hard rule): a building CO must START and FINISH
                        # inside one shift. The per-shift assignment budget already guarantees this
                        # for GT machines (0 crossings observed, all months), so this is a defensive
                        # HARD invariant: if a CO would ever cross the shift boundary, defer the whole
                        # campaign (CO + its production) to the next shift instead of emitting a
                        # boundary-crossing CO. The un-built qty persists as deficit and rebuilds next
                        # shift; the machine idles this shift's remainder (no CO produces output).
                        if (_CO_SHIFT_CONTAINED and not _replay_now
                                and _cursor + timedelta(minutes=co_mins) > _shift_start + timedelta(minutes=SHIFT_MINS)):
                            continue
                        # Keep the DISPLAYED CO_Type in lock-step with the charged minutes by
                        # deriving it from the SAME prev_inch that _co_cost uses (prevents label↔
                        # minutes drift, e.g. a Day-1 seed labelling a real inch-change same_size).
                        # A free machine (blank prev_sku) skips the CO entirely — its first mount of
                        # the month is a setup ("start"), not a charged CO (honours BLD_START_FREE),
                        # which is what produced the reported same_size/45-min VMI mismatch.
                        _disp_co_type = ("plus3_CO" if co_mins == INCH_PLUS3_CO_MINS
                                         else "same_size_CO" if prev_inch == sku_inch.get(sku, "")
                                         else "diff_size_CO")
                        _co_start = min(_cursor, _replay_shift_end) if _replay_now else _cursor
                        _cursor   = _co_start + timedelta(minutes=co_mins)
                        if _replay_now:
                            _cursor = min(_cursor, _replay_shift_end)   # keep the CO inside the shift
                        bld_shift_rows.append({
                            "Machine":       machine,
                            "Date":          date_str,
                            "Shift":         shift,
                            "SKUCode":       "CHANGEOVER",
                            # A changeover produces NO tyres: Qty must be 0 so the
                            # column means one thing everywhere and stays summable.
                            # The CO duration lives in its own CO_Mins column (and is
                            # already implied by StartTime→EndTime).
                            "Qty":           0,
                            "CO_Mins":       co_mins,
                            "StartTime":     _fmt_dt(_co_start),
                            "EndTime":       _fmt_dt(_cursor),
                            "Machine_Group": _group_label(machine),
                            "CO_Type":       _disp_co_type,
                            # Replay days 1-2: exempt from the shift-boundary splitter so the
                            # plant qty stays in its exact plant shift (no cross-shift re-timing).
                            **({"_replay": True} if _replay_now else {}),
                        })
                        bld_co_events.append({
                            "Machine":      machine,
                            "Date":         date_str,
                            "Shift":        shift,
                            "Day":          day,
                            "CO_Day_Index": day,
                            "From_SKU":     prev_sku,
                            "Target_SKU":   sku,
                            "CO_Type":      _disp_co_type,
                            "CO_Cost_Mins": co_mins,
                            "Status":       f"Rolling CO ({_disp_co_type})",
                        })
                    _prod_start = min(_cursor, _replay_shift_end) if _replay_now else _cursor
                    _cursor     = _prod_start + timedelta(minutes=qty * _ct_sec / 60.0)
                    if _replay_now:
                        _cursor = min(_cursor, _replay_shift_end)     # keep production inside the shift
                    bld_shift_rows.append({
                        "Machine":       machine,
                        "Date":          date_str,
                        "Shift":         shift,
                        "SKUCode":       sku,
                        "Qty":           qty,
                        "CO_Mins":       0,
                        "StartTime":     _fmt_dt(_prod_start),
                        "EndTime":       _fmt_dt(_cursor),
                        "Machine_Group": _group_label(machine),
                        "CO_Type":       "production",
                        # Replay days 1-2: exempt from the shift-boundary splitter so the plant
                        # qty stays in its exact plant shift (splitter would else re-sequence the
                        # CO+production past 480 min and push the overflow into the next shift).
                        **({"_replay": True} if _replay_now else {}),
                    })
                    # Stage-1 produces carcass (not GT) → carcass feeds Stage-2,
                    # NOT curing presses.  Do NOT add to gt_inventory for Stage-1
                    # machines; curing must only draw real GT (Stage-2 / Unistage / VMI / BJ).
                    if machine not in _S1_MACHINES:
                        gt_inventory[sku]   = gt_inventory.get(sku, 0.0) + qty
                        if qty > 0:                                    # FIFO: new dated GT lot
                            gt_lots.setdefault(sku, []).append([day, float(qty)])
                            # SG_DELIB: record the finer group that actually built this SKU
                            # (mutable current-group state; day 3+ only, inert when OFF).
                            if _SG_DELIB and day > _PLANT_2DAY_DAYS:
                                _sku_cur_group[sku] = _sku_group_of(machine)
                        day_gt_built[sku]  += qty
                        if _DYNAMIC_CO_PLANNER_ENABLED:
                            _prev_tier = sku_campaign_tier.get(sku)
                            if _prev_tier is None or _tier_idx < _prev_tier:
                                sku_campaign_tier[sku] = _tier_idx
                    day_built[sku]    += qty
                    shift_bld[sku]    += qty
                    if qty > 0:
                        last_build_day[sku] = day
                    prev_sku = sku; prev_inch = sku_inch.get(sku, "")
                if campaigns:
                    # SKU_NO_REVERT: every SKU the machine held-then-left this shift (all non-final
                    # campaign SKUs + the entry SKU if it changed) is recorded as LEFT → never rebuilt.
                    if _SKU_NO_REVERT:
                        _entry_sku = machine_current_sku.get(machine, "")
                        _final_sku = campaigns[-1][0]
                        _left_now = {c[0] for c in campaigns if c[0] != _final_sku}
                        if _entry_sku and _entry_sku != _final_sku:
                            _left_now.add(_entry_sku)
                        if _left_now:
                            machine_left_skus.setdefault(machine, set()).update(_left_now)
                    # Update current SKU at end of THIS SHIFT (not end of day)
                    machine_current_sku[machine] = campaigns[-1][0]
                    # Record this shift's SKUs into the day set (4-SKU/day cap tracker).
                    machine_day_skus.setdefault(machine, set()).update(
                        _c[0] for _c in campaigns)

                    # ── Client inch rules: advance the machine's anchor + inch history.
                    # The FIRST inch a machine ever runs fixes its +/-2 band for the
                    # month (Rule 2). machine_inch_since restarts the 5-day-dwell clock
                    # whenever the machine's current inch actually changes.
                    if _INCH_RULES_ENABLED:
                        _used = machine_used_inches.setdefault(machine, set())
                        for _c_sku, _c_qty, _c_type in campaigns:
                            _ci = sku_inch.get(_c_sku, "")
                            if not _ci:
                                continue
                            if machine not in machine_anchor_inch:
                                machine_anchor_inch[machine] = _ci
                            _used.add(_ci)
                        _end_inch = sku_inch.get(campaigns[-1][0], "")
                        if _end_inch and machine_inch_now.get(machine) != _end_inch:
                            machine_inch_since[machine] = day      # new inch → reset dwell clock
                            machine_inch_now[machine]   = _end_inch

                    # Update cross-shift minute tracker.
                    # Accumulate production minutes after the LAST CO in this shift.
                    # If no CO occurred: add all production minutes to existing total.
                    _had_co = False
                    _mins_after_last_co = 0.0
                    for _csku, _q, _ct in campaigns:
                        _ct_sec = _bld_ct_sec(machine, _csku)
                        _prod_m = _q * _ct_sec / 60.0
                        if _ct != "start":           # CO event → reset counter
                            _had_co = True
                            _mins_after_last_co = _prod_m
                        else:
                            _mins_after_last_co += _prod_m
                    if _had_co:
                        machine_minutes_on_sku[machine] = _mins_after_last_co
                    else:
                        machine_minutes_on_sku[machine] = (
                            machine_minutes_on_sku.get(machine, 0.0) + _mins_after_last_co
                        )

            # ── 3b. Stage-1 carcass scheduling (utilization tracking only) ──
            # For every SKU that Stage-2 machines built this shift, allocate the
            # matching carcass qty (1:1 with GT) across that SKU's eligible Stage-1
            # machines, capped by each machine's per-shift capacity. Recorded in
            # bld_shift_rows (CO_Type="carcass") so Machine Utilization / Daily GT
            # & Carcass show real numbers instead of 0%. Deliberately does NOT
            # touch gt_inventory, demand_remaining, or machine_skus — Stage-1 is
            # assumed to always have spare capacity to supply Stage-2 (validated:
            # Stage-1 util stays well under 100% even at full Stage-2 output), so
            # this can never gate or double-count real GT output.
            # Carcass rows for THIS shift's Stage-2 output (post gate-clamp). The
            # post-plan max-flow (_stage1_carcass_schedule) later REPLACES these with
            # the aging-aware allocation; kept here for Stage-1 inch tracking + the
            # STAGE1_CARCASS_PASS=0 fallback. Runs on the clamped plan when gated.
            stage2_built_this_shift: dict[str, float] = defaultdict(float)
            for machine, campaigns in shift_plan.items():
                if _MACHINE_GROUP.get(machine, "") != "STAGE2":
                    continue
                for sku, qty, _co_type in campaigns:
                    stage2_built_this_shift[sku] += qty
            # One machine = one SKU per shift (same physical constraint as every
            # other building machine — see CLAUDE.md "One building machine always
            # produces for exactly one SKU at a time"). Once a Stage-1 machine is
            # allocated to a SKU this shift it's removed from the pool even if it
            # has leftover capacity, instead of splitting it across SKUs with no
            # changeover between them.
            s1_machines_used_this_shift: set = set()
            # _s1_inch_ok is defined once per shift above (step 2c) and reused here.

            for sku, need in sorted(stage2_built_this_shift.items(), key=lambda kv: (-kv[1], kv[0])):
                if need <= 0:
                    continue
                eligible = sorted(
                    (m for m in s1_sku_to_machines.get(sku, ())
                     if m not in s1_machines_used_this_shift and _s1_inch_ok(m, sku)),
                    key=lambda m: (-_bld_qty_per_shift(m), m),
                )
                remaining_need = need
                for m in eligible:
                    if remaining_need <= 0:
                        break
                    cap = _bld_qty_per_shift(m)
                    if cap <= 0:
                        continue
                    alloc = min(cap, remaining_need)
                    s1_machines_used_this_shift.add(m)
                    remaining_need -= alloc
                    # Advance the Stage-1 machine's inch history (first inch fixes
                    # its band; every inch run is closed once it moves on).
                    if _INCH_RULES_ENABLED:
                        _s1_i = sku_inch.get(sku, "")
                        if _s1_i:
                            if m not in machine_anchor_inch:
                                machine_anchor_inch[m] = _s1_i
                            machine_used_inches.setdefault(m, set()).add(_s1_i)
                            s1_current_inch[m] = _s1_i
                            if machine_inch_now.get(m) != _s1_i:   # dwell clock (Stage-1 follows Stage-2)
                                machine_inch_since[m] = day
                                machine_inch_now[m]   = _s1_i
                    # One SKU per Stage-1 machine per shift → carcass run starts
                    # at the shift clock start; duration = alloc × CT.
                    _c_start = _shift_start
                    _c_end   = _c_start + timedelta(
                        minutes=round(alloc) * _bld_ct_sec(m, sku) / 60.0
                    )
                    bld_shift_rows.append({
                        "Machine":       m,
                        "Date":          date_str,
                        "Shift":         shift,
                        "SKUCode":       sku,
                        "Qty":           round(alloc),
                        "CO_Mins":       0,
                        "StartTime":     _fmt_dt(_c_start),
                        "EndTime":       _fmt_dt(_c_end),
                        "Machine_Group": _group_label(m),
                        "CO_Type":       "carcass",
                    })

            # ── Part B: single reactive CO rule (once per shift, AFTER building) ──
            if _REACTIVE_ONLY and _RCO_ARBITER:
                _reactive_co(day, shift, cur_shift_global, date_str)

            # ── 4. Curing simulation ───────────────────────────────────────
            # HOLIDAY: no press cures (iterate no presses). press_state + CO/clean carries
            # are left intact → a press mid-campaign resumes the same SKU on the next
            # working shift. (Phase 3 upgrades in-flight COs to COMPLETE during the idle day.)
            # Decision #3: an in-flight changeover/mould-clean COMPLETES during a plant
            # holiday (the setup crew works even though production is idle), so the press
            # is ready to run on the FIRST working shift. Drain co_carry/clean_carry by a
            # full shift each holiday shift — but book NO press_stats minutes (the holiday
            # shift is dropped from the utilization denominator, decision #4) and produce
            # no GT. Parity: only runs when _holiday, so empty-holidays is bit-identical.
            if _holiday:
                for _hp in sorted(press_state):
                    _rem = float(SHIFT_MINS)
                    if co_carry.get(_hp, 0.0) > 0:
                        _dd = min(co_carry[_hp], _rem)
                        co_carry[_hp] -= _dd
                        _rem -= _dd
                    if _MOULD_CLEAN_ENABLED and _rem > 0 and clean_carry.get(_hp, 0.0) > 0:
                        _dd = min(clean_carry[_hp], _rem)
                        clean_carry[_hp] -= _dd
                        _rem -= _dd

            for press in (sorted(press_state) if not _holiday else []):
                st  = press_state[press]
                sku = st["sku"]

                if press in co_press_map:
                    # Shifts BEFORE the CO → press still runs its OLD SKU.
                    # The CO shift        → CHANGEOVER (full shift idle).
                    # Shifts AFTER        → the NEW SKU runs.
                    _cs = co_shift_idx.get(press, 0)
                    _si = SHIFTS.index(shift)
                    if _si < _cs:
                        status = st["status"]       # old SKU keeps producing until its CO
                    elif _si == _cs:
                        status = "CHANGEOVER"       # CO shift: full shift idle
                    else:
                        sku    = co_press_map[press]
                        status = "RUNNING"
                elif press in dynamic_co_tracker:
                    co_idx, _ = dynamic_co_tracker[press]
                    if cur_shift_global == co_idx:
                        status = "CHANGEOVER"      # CO started this shift (demand just met)
                    else:
                        status = st["status"]      # shouldn't occur; fallback
                else:
                    status = st["status"]

                # CAMPAIGN active-set: this press is EXCESS for its SKU vs the day's building-sized
                # target → hold it IDLE (no cure, no draw, NOT starved). Building concentrates on the
                # active presses; the excess doesn't fake-busy-starve.
                if _campaign_on and press in _camp_idle and status == "RUNNING":
                    status = "CAMPAIGN_IDLE"

                ct       = cure_ct_map.get(sku, DEFAULT_CURING_CT)
                ct_disp  = ct                       # per-SKU DISPLAY CT (baked at 0.95) — sheet only
                # Per-press efficiency (plant rule, always applied): scale THIS press's effective
                # cure CT by 0.95/eff(press) — 4-digit keeps 0.95/0.95 = 1.0 (unchanged, cure_ct_map
                # already baked at 0.95); 5-digit → 0.95/0.94 (true CT = raw/0.94, ~1% slower).
                # Effective ct flows into cap / cap_time / prod_mins / early-CO; ct_disp stays 0.95.
                ct = ct * (ConsumptionConfig.PRESS_EFFICIENCY / _press_efficiency(press))
                cap      = _cure_qty_per_shift(ct)
                _pm_down = 0.0                                       # PM/MTC press maintenance mins
                if _PM_MTC_ENABLED and _CUR_DOWN.get(press):
                    _ps = date + timedelta(hours=8 * SHIFTS.index(shift))
                    _pe = _ps + timedelta(hours=8)
                    _pm_down = _down_mins(_CUR_DOWN[press], _ps, _pe)
                    if _pm_down > 0:
                        if _PM_MTC_NO_OVERLAP:
                            # Usable minutes = the LARGEST window-free block in this shift (the
                            # pre-window gap is now used too — PM_MTC_PREWINDOW — since building
                            # before the window and going down does not overlap it; the row
                            # relocation below places the cure block inside that free interval).
                            _post_free, _ = _pm_maint_free(_CUR_DOWN[press], _ps, _pe)
                            _pm_down = SHIFT_MINS - _post_free       # effective downtime (drives _avail below)
                        cap = cap * max(0.0, (SHIFT_MINS - _pm_down) / SHIFT_MINS)
                gt_avail = max(0.0, gt_inventory.get(sku, 0.0))

                # ── Mould-clean carry-in: a clean that began mid-shift last shift
                # occupies the front of THIS shift. If it fills the whole shift the
                # press is in MOULD_CLEAN (no production); otherwise production runs
                # in the reduced remaining minutes _avail.
                _avail    = max(0.0, float(SHIFT_MINS) - _pm_down)   # PM/MTC: downtime removes shift minutes
                _busy_in  = 0.0    # CO/clean minutes consumed at the FRONT of this shift
                # Per-shift CO / mould-clean minutes — surfaced as CO_Mins /
                # Mould_Clean_Mins columns so every changeover (planned full-shift,
                # dynamic mid-shift trigger, and its overhang) is VISIBLE in the sheet,
                # not just in the Changeover Plan.
                _co_mins_shift    = 0.0
                _clean_mins_shift = 0.0
                _dyn_co_tgt       = None
                # Separate time-portions so production and CO/clean become SEPARATE
                # rows in the sheet (each with its own real wall-clock window):
                _seg_co_in    = 0.0   # CO overhang carried into the FRONT of this shift
                _seg_clean_in = 0.0   # mould-clean overhang carried into the front
                _seg_co_trig  = 0.0   # CO fired mid-shift AFTER production (this shift's part)
                _seg_clean_trig = 0.0 # mould clean fired mid-shift after production
                # (a) Changeover overhang: a dynamic CO that fired mid-shift last shift
                #     is still running. It blocks the front of this shift, so the new
                #     SKU's production starts mid-shift.
                if status == "RUNNING" and co_carry.get(press, 0.0) > 0:
                    _coin = min(co_carry[press], _avail)
                    co_carry[press] -= _coin
                    _avail  -= _coin
                    _busy_in += _coin
                    if _avail <= 0:
                        status = "CHANGEOVER"           # else-branch books SHIFT_MINS co
                    else:
                        press_stats[press]["co_mins"] += _coin
                        _co_mins_shift += _coin
                        _seg_co_in = _coin
                # (b) Mould-clean overhang (same pattern).
                if (_MOULD_CLEAN_ENABLED and status == "RUNNING"
                        and clean_carry.get(press, 0.0) > 0):
                    _cin = min(clean_carry[press], _avail)
                    clean_carry[press] -= _cin
                    _avail  -= _cin
                    _busy_in += _cin
                    if _avail <= 0:
                        status = "MOULD_CLEAN"          # else-branch books SHIFT_MINS clean
                    else:
                        press_stats[press]["clean_mins"] += _cin   # partial carry booked here
                        _clean_mins_shift += _cin
                        _seg_clean_in = _cin

                _cleaned  = False
                prod_mins = 0.0
                # CONTINUOUS CYCLE CARRY (plant rule): capture prior shift's in-progress fractional
                # cycle, then reset to 0 — a CO / mould-clean / idle / holiday press has no cycle to
                # carry; a RUNNING, TIME-limited press re-sets it after producing (below).
                _carry_prev = _cure_carry_min[press] if _CURE_CYCLE_CARRY else 0.0
                _cure_carry_min[press] = 0.0
                if status == "RUNNING":
                    # DEBUG self-check: a RUNNING press must own 2 moulds eligible
                    # for its SKU. Counts leaks where the gate failed to block.
                    if _mould_gate and os.environ.get("MOULD_DEBUG"):
                        _ow = [m for m in _press_moulds.get(press, set())
                               if sku in _mould_skus.get(m, set())]
                        if len(_ow) < 2:
                            _mould_selfcheck[0] += 1
                            _mould_selfcheck.append((day, shift, press, sku, len(_ow),
                                                     len(_sku_moulds.get(sku, set()))))
                        # targeted: dump the actual owned-mould state for SRBT0 presses
                        if (os.environ.get("MOULD_DUMP") and day == 22 and shift == "C"
                                and sku == "1325119015008SRBT0"):
                            print(f"    [DUMP] press {press} sku {sku} owns "
                                  f"{sorted(_press_moulds.get(press, set()))} | "
                                  f"eligible-of-those={sorted(_ow)} | "
                                  f"sku_moulds[SRBT0]={len(_sku_moulds.get(sku,set()))}")
                    # Cap curing at remaining demand — never over-produce.
                    demand_left = max(0.0, demand_remaining.get(sku, 0.0))
                    # CONTINUOUS CYCLE CARRY: add the fractional cycle carried from the prior shift
                    # to this shift's production minutes, so the cap recovers it (30,32,32,… not
                    # 30,30,30,…). _carry_prev captured + reset above; re-set below iff TIME-limited.
                    if _MOULD_CLEAN_ENABLED:
                        _avail_prod = _avail + _carry_prev
                        cap_time = int(_avail_prod / ct) * CURING_CAVITIES   # partial-shift cap (+carry)
                        cap_life = mould_life[press] * CURING_CAVITIES  # cycles left × cavities
                        cured = min(cap_time, int(gt_avail), int(demand_left), cap_life)
                        _time_cap = cap_time
                    else:
                        _avail_prod = float(SHIFT_MINS) + _carry_prev
                        _time_cap = int(_avail_prod / ct) * CURING_CAVITIES
                        cured = min(_time_cap, int(gt_avail), int(demand_left))
                    # MPQ (shift-level floor): don't emit a sub-floor cured block. The GT is
                    # LEFT in inventory (batches a later shift); a genuine sub-MPQ demand tail
                    # is dropped via the _demand_done release threshold below (press moves on,
                    # tail stays UNMET — no over-production, no coverage inflation).
                    if _CURING_MPQ > 0 and 0 < cured < _CURING_MPQ:
                        cured = 0
                    # CONTINUOUS CYCLE CARRY: if this press was TIME-limited (ran flat-out — cured
                    # hit the time cap, not GT / demand / mould-life / MPQ), a cure cycle is in
                    # progress at the shift boundary → carry its fractional minutes to the next
                    # shift. Otherwise the press idled/stopped at the boundary → no carry (stays 0).
                    if _CURE_CYCLE_CARRY and cured > 0 and cured >= _time_cap:
                        _cure_carry_min[press] = max(0.0, _avail_prod - (cured // CURING_CAVITIES) * ct)
                    gt_inventory[sku]      = gt_avail - cured
                    if cured > 0:                                     # FIFO: draw oldest lots first
                        _gt_consume_lots(sku, cured)
                    day_cured_d[sku]      += cured
                    sku_cured[sku]        += cured
                    daily_cured[date_str] += cured
                    demand_remaining[sku]  = max(0.0, demand_remaining.get(sku, 0.0) - cured)
                    # CURING_ADAPT_CO: track this press's consecutive zero-GT (starvation)
                    # run — starving iff it produced nothing but still owes demand.
                    if _CURING_ADAPT_CO:
                        if cured == 0 and demand_left > 0 and int(round(gt_avail)) == 0:
                            _consec_zero_gt[press] += 1
                        else:
                            _consec_zero_gt[press] = 0
                    prod_mins = cured * ct / CURING_CAVITIES
                    press_stats[press]["running_mins"] += prod_mins
                    press_stats[press]["skus"].add(sku)
                    press_stats[press]["cycles"] += cured // CURING_CAVITIES
                    press_stats[press]["units"]  += cured
                    press_sku_stats[(press, sku)]["cycles"]    += cured // CURING_CAVITIES
                    press_sku_stats[(press, sku)]["units"]     += cured
                    press_sku_stats[(press, sku)]["mins_used"] += prod_mins

                    # ── Mould-clean: decrement life; fire an immediate clean when it
                    # hits 0 (cured was capped by cap_life). The 8h clean fills the
                    # rest of THIS shift; the remainder bleeds into the next shift.
                    if _MOULD_CLEAN_ENABLED:
                        mould_life[press] -= cured // CURING_CAVITIES
                        if mould_life[press] <= 0:
                            clean_here = max(0.0, min(float(MOULD_CLEAN_MINS),
                                                      _avail - prod_mins))
                            press_stats[press]["clean_mins"] += clean_here
                            _clean_mins_shift += clean_here
                            _seg_clean_trig = clean_here
                            clean_carry[press] = MOULD_CLEAN_MINS - clean_here
                            mould_life[press]  = MOULD_CLEAN_CYCLES
                            _cleaned = True
                            # Record the mould-clean event (one per trigger, 8h/480 min
                            # total; may overhang into the next shift) — surfaced in the
                            # Changeover Plan sheet alongside curing COs.
                            mould_clean_events.append({
                                "Date":       date_str,
                                "Day":        day,
                                "Shift":      shift,
                                "Press":      press,
                                "From_SKU":   sku,
                                "Target_SKU": sku,          # clean keeps the same SKU
                                "CO_Type":    "Mould Clean",
                                "Mins":       float(MOULD_CLEAN_MINS),
                            })

                    # ── Instant CO when demand is met ─────────────────────────
                    # Demand just hit 0: CO starts immediately (remaining shift
                    # time = CHANGEOVER).  Next shift = PRODUCTION for new SKU.
                    # Conditions: no pre-planned CO today, not already in a
                    # dynamic CO, and no pre-planned CO tomorrow (avoid conflict).
                    # MPQ: a press is "demand-done" (releases / CO's away) once its SKU's
                    # remaining demand is at/below the curing floor — the sub-MPQ tail is left
                    # unmet rather than wasting further press-shifts on <MPQ cures.
                    _demand_done = demand_remaining.get(sku, 0.0) <= (_CURING_MPQ if _CURING_MPQ > 0 else 0)
                    # Early-CO (_EARLY_CO_ENABLED): a press may reassign BEFORE its
                    # SKU's demand is done, IFF that SKU's remaining demand is met by
                    # its OTHER (n-1) presses within the horizon — i.e. this press is
                    # surplus. Replicates the static proxy's early freeing, on real
                    # state. Only fires if _select_ratio_co_target finds a needy target.
                    _early_co = False
                    if (_EARLY_CO_ENABLED and not _demand_done
                            and press not in co_press_map
                            and press not in dynamic_co_tracker):
                        _n_cur = press_count.get(sku, 0)
                        if _n_cur > 1:
                            _rate_cur = _cure_qty_per_shift(ct) * 3
                            _rem_cur  = demand_remaining.get(sku, 0.0)
                            _hz_cur   = _working_days_left(day)   # holiday-aware horizon
                            if (_rate_cur > 0
                                    and _rem_cur / ((_n_cur - 1) * _rate_cur) <= _hz_cur):
                                _early_co = True
                    # CURING_ADAPT_CO: sustained-starvation switch. This press got 0 GT for
                    # N consecutive shifts on a SKU that still owes demand → the SKU is
                    # building-limited, so CO the press to a feedable in-demand SKU (a
                    # building machine can supply). Reuses the dynamic-CO firing below; the
                    # target is picked by the same dynamic selector (which prefers SKUs with
                    # GT on hand = feedable), and this press's STALE future planned CO is
                    # blocked (it's now committed to the new SKU).
                    _starv_thresh = (_RCO_STARV_SHIFTS if _REACTIVE_ONLY   # point 1: fast supply-aware
                                     else _CURING_STARV_SWITCH_SHIFTS)
                    _starv_switch = (_CURING_ADAPT_CO and not _demand_done and not _early_co
                                     and _consec_zero_gt[press] >= _starv_thresh
                                     and press not in co_press_map
                                     and press not in dynamic_co_tracker)
                    # Feed guard: don't switch a SKU building CAN sustain (its 0-GT run is
                    # transient — the dynamic buffer is busy elsewhere and will return),
                    # only one that is genuinely building-limited (buildable < curing draw).
                    # Point 1 (user): under _REACTIVE_ONLY, trust REAL starvation over the
                    # (optimistic) buildable_rate estimate — a press starved _RCO_STARV_SHIFTS
                    # shifts IS not being fed, so let it CO to a GT-on-hand SKU. Skip the guard.
                    if (_starv_switch and _CURING_ADAPT_FEED_GUARD and not _REACTIVE_ONLY
                            and _buildable_rate is not None
                            and _buildable_rate.get(sku, 0.0)
                                >= press_count.get(sku, 1) * cap * 3):
                        _starv_switch = False
                    if (not (_REACTIVE_ONLY and _RCO_ARBITER)   # arbiter owns all COs only under RCO_ARBITER
                            and (((_DYNAMIC_CO_TRACKER_ENABLED and _demand_done) or _early_co
                            or _starv_switch)
                            and not _cleaned          # a just-started clean defers any CO
                            and press not in co_press_map
                            and press not in dynamic_co_tracker)):
                        _next_day_cos = {p for p, _, _ in co_by_day.get(day + 1, [])}
                        # Phase 3 pull-forward: a press blocked here (planned CO booked
                        # TOMORROW) normally idles the rest of today. With the scorer on,
                        # bring tomorrow's planned CO FORWARD to now instead of idling.
                        _pf_target = None
                        if _CO_SCORER_ENABLED and _mould_gate and press in _next_day_cos:
                            _pf_target = next(
                                (ns for (p, _o, ns) in co_by_day.get(day + 1, [])
                                 if p == press), None)
                        if (press not in _next_day_cos) or (_pf_target is not None):
                            _slots_left = MAX_CHANGEOVERS_PER_DAY - daily_co_count[day]
                            if _slots_left > 0:
                                _horizon_left = _working_days_left(day)   # holiday-aware horizon
                                _pf_fired = False
                                _forced_co = False                        # P1: set if supply test bypassed
                                if _CO_SCORER_ENABLED and _mould_gate:
                                    # Pull-forward if tomorrow's planned CO is feasible now;
                                    # else fall back to the SAME tuned dynamic selector as
                                    # Phase-2 (don't disturb its behaviour — measured better
                                    # than the utility picker for reactive mid-shift COs).
                                    _bf = _bld_free_min_shift()
                                    if (_pf_target is not None
                                            and demand_remaining.get(_pf_target, 0.0) > 0
                                            and _n_free_for(_pf_target, press) >= 2
                                            and _bld_capacity(_pf_target, _bf)
                                            >= _cure_qty_per_shift(cure_ct_map.get(
                                                _pf_target, DEFAULT_CURING_CT))):
                                        _target = _pf_target
                                        _pf_fired = True
                                    elif _REACTIVE_CO or _REACTIVE_ONLY:
                                        # P1 / B-1 supply-gate (point 3): pick the best dynamic
                                        # target that PASSES the building-supply test, best-first
                                        # — re-query the selector excluding each supply-failed SKU
                                        # so the next-best is considered. If NONE is supply-feasible,
                                        # fire a FORCED CO on the plain best target (bypasses
                                        # ONLY the supply test; the allowable + mould gates
                                        # below still apply). Deterministic.
                                        _already = set(dynamic_co_tracker[p][1]
                                                       for p in dynamic_co_tracker)
                                        _excl = set(_already)
                                        _target = None
                                        while True:
                                            _cand = _select_dynamic_co_target(
                                                sku, demand_remaining, press_count,
                                                cure_ct_map, priority_score_map,
                                                gt_inventory, _horizon_left, _excl,
                                                priority_deadline_map=(priority_deadline_map if _prio_active else None),
                                                day=day,
                                            )
                                            if _cand is None:
                                                break
                                            if _supply_ok(_cand, _bf):
                                                _target = _cand
                                                break
                                            _excl.add(_cand)      # supply-infeasible → next-best
                                        if _target is None:
                                            # forced fallback — best allowable/mould target regardless of supply
                                            _target = _select_dynamic_co_target(
                                                sku, demand_remaining, press_count,
                                                cure_ct_map, priority_score_map,
                                                gt_inventory, _horizon_left, _already,
                                                priority_deadline_map=(priority_deadline_map if _prio_active else None),
                                                day=day,
                                            )
                                            if _target is not None:
                                                _forced_co = True
                                    else:
                                        _already = set(dynamic_co_tracker[p][1]
                                                       for p in dynamic_co_tracker)
                                        _target = _select_dynamic_co_target(
                                            sku, demand_remaining, press_count,
                                            cure_ct_map, priority_score_map,
                                            gt_inventory, _horizon_left, _already,
                                            priority_deadline_map=(priority_deadline_map if _prio_active else None),
                                            day=day,
                                        )
                                elif _RATIO_CO_ALLOCATION_ENABLED or _EARLY_CO_ENABLED:
                                    _pending_counts = Counter(
                                        dynamic_co_tracker[p][1] for p in dynamic_co_tracker
                                    )
                                    _target = _select_ratio_co_target(
                                        sku, press, demand_remaining, press_count,
                                        _pending_counts, demand_dict, cure_ct_map,
                                        press_to_demand_targets, press_total_demand,
                                        _horizon_left,
                                        sku_to_press_count=sku_to_press_count,
                                        rich_ranking=_RATIO_CO_RICH_RANKING_ENABLED,
                                    )
                                else:
                                    _already = set(dynamic_co_tracker[p][1]
                                                   for p in dynamic_co_tracker)
                                    _target = _select_dynamic_co_target(
                                        sku, demand_remaining, press_count,
                                        cure_ct_map, priority_score_map,
                                        gt_inventory, _horizon_left, _already,
                                        priority_deadline_map=(priority_deadline_map if _prio_active else None),
                                        day=day,
                                    )
                                # Allowable-press gate (R3C fix): _select_dynamic_co_target ranks
                                # by urgency across ALL demand SKUs and never sees this press's
                                # allowable matrix, so late-month (once the press's own demand is
                                # done) it can pick a mould-feasible but ALLOWABLE-INFEASIBLE SKU.
                                # Reject any target this press is not curing-allowable for. Fall
                                # back to press_to_demand_targets (always built) if press_allow_skus
                                # is empty, so the gate never silently no-ops.
                                if _target is not None:
                                    if _REACTIVE_ONLY:
                                        # STRICT (R3C): only the curing-allowable set — no
                                        # press_to_demand_targets fallback (which is a building
                                        # map and can contain non-curing-allowable SKUs). Empty
                                        # allowable set → don't fire (never CO to an unallowable
                                        # press,SKU pair).
                                        _allow_p = press_allow_skus.get(press)
                                        if not _allow_p or _target not in set(_allow_p):
                                            _target = None
                                    else:
                                        _allow_p = press_allow_skus.get(press) or press_to_demand_targets.get(press)
                                        if _allow_p and _target not in set(_allow_p):
                                            _target = None
                                # Mould gate: only fire the reactive CO if 2 eligible moulds
                                # are free for the target. B-1 point 1: if blocked, try a
                                # depth-1 machine-swap (free a mould from a demand-done donor)
                                # before giving up; else keep this SKU.
                                if _target is not None and not _try_mount(press, _target):
                                    if (_REACTIVE_ONLY
                                            and _do_swap_for(press, _target, day, shift,
                                                             cur_shift_global, date_str)
                                            and _try_mount(press, _target)):
                                        pass                    # swap unblocked the original target
                                    elif _REACTIVE_ONLY:
                                        # retarget-on-block: the contended moulds are held by
                                        # producing presses (swap can't free them) — instead of
                                        # idling, pick the neediest allowable SKU this press CAN
                                        # mount (2 free moulds) and CO there.
                                        _rt = _pick_retarget(press)
                                        if _rt is not None and _rt != _target and _try_mount(press, _rt):
                                            _target = _rt
                                        else:
                                            mould_blocked_cos += 1
                                            _target = None
                                    else:
                                        mould_blocked_cos += 1
                                        _target = None
                                if _target is not None and _CO_SCORER_ENABLED:
                                    if _pf_fired:
                                        # consume tomorrow's planned CO so it can't fire twice
                                        co_by_day[day + 1] = [
                                            (p, o, n) for (p, o, n) in co_by_day.get(day + 1, [])
                                            if p != press]
                                        daily_co_count[day + 1] = max(
                                            0, daily_co_count.get(day + 1, 0) - 1)
                                        co_scorer_stats["pullfwd"] += 1
                                    elif _forced_co:
                                        co_scorer_stats["forced"] += 1     # P1: supply test bypassed
                                    else:
                                        co_scorer_stats["dynamic"] += 1
                                if _target is not None:
                                    # CO starts NOW — mid-shift, the moment demand was
                                    # met. Charge the real CURING_CO_CHANGEOVER_MINS from
                                    # this point: it eats the rest of THIS shift, and the
                                    # overhang carries into the next shift (co_carry), so
                                    # the new SKU begins mid-shift rather than free at the
                                    # boundary. Without this a dynamic CO cost nothing.
                                    press_count[sku] = max(
                                        0, press_count.get(sku, 0) - 1
                                    )
                                    dynamic_co_tracker[press] = (cur_shift_global, _target)
                                    daily_co_count[day] += 1
                                    # CURING_ADAPT_CO: a CO resets this press's starvation
                                    # run; a sustained-starvation SWITCH also BLOCKS the
                                    # press's stale future planned CO(s) — it is now
                                    # committed to the new (feedable) SKU until that SKU
                                    # completes or itself starves N shifts.
                                    if _CURING_ADAPT_CO:
                                        _consec_zero_gt[press] = 0
                                    # Item 3: wipe this press's STALE future planned COs (booked
                                    # against its OLD sku). Baseline = starv-switch only (under
                                    # _CURING_ADAPT_CO); HYBRID_CO_CANCEL makes it comprehensive for
                                    # ANY dynamic CO. Presence-guarded → idempotent (pull-forward
                                    # already removed day+1).
                                    if _HYBRID_CO_CANCEL or (_CURING_ADAPT_CO and _starv_switch):
                                        for _fd in range(day + 1, planning_days + 1):
                                            _cofd = co_by_day.get(_fd)
                                            if _cofd and any(p == press for (p, _o, _n) in _cofd):
                                                co_by_day[_fd] = [
                                                    (p, o, n) for (p, o, n) in _cofd if p != press]
                                                daily_co_count[_fd] = max(
                                                    0, daily_co_count.get(_fd, 0) - 1)
                                    _co_here = max(0.0, min(float(CURING_CO_CHANGEOVER_MINS),
                                                            _avail - prod_mins))
                                    press_stats[press]["co_mins"] += _co_here
                                    _co_mins_shift += _co_here
                                    _seg_co_trig    = _co_here
                                    _dyn_co_tgt     = _target      # surfaced in Remarks
                                    co_carry[press] = float(CURING_CO_CHANGEOVER_MINS) - _co_here
                                    # Rule 2: a curing CO resets mould life (the CO
                                    # shift already includes a clean).
                                    mould_life[press]  = MOULD_CLEAN_CYCLES
                                    clean_carry[press] = 0.0
                                    cure_co_events.append({
                                        "Date":       date_str,
                                        "Day":        day,
                                        "Shift":      shift,
                                        "Press":      press,
                                        "From_SKU":   sku,
                                        "Target_SKU": _target,
                                        "CO_Type":    "Early-CO" if (_early_co and not _demand_done) else "Dynamic",
                                    })
                                    print(
                                        f"    [DynCO{'/FORCED' if _forced_co else ''}] "
                                        f"Day {day} Shift {shift}: "
                                        f"press {press} {sku}→{_target} "
                                        f"(slot {MAX_CHANGEOVERS_PER_DAY - _slots_left + 1}"
                                        f"/{MAX_CHANGEOVERS_PER_DAY})"
                                    )
                else:
                    cured = 0
                    if status == "CHANGEOVER":
                        press_stats[press]["co_mins"] += SHIFT_MINS
                        _co_mins_shift = float(SHIFT_MINS)
                    elif status == "MOULD_CLEAN":
                        press_stats[press]["clean_mins"] += SHIFT_MINS
                        _clean_mins_shift = float(SHIFT_MINS)

                # ── Emit ONE ROW PER SEGMENT ──────────────────────────────────
                # A press-shift is broken into its real time segments so production
                # and any changeover / mould clean appear as SEPARATE rows, each with
                # its own wall-clock window. Chronological order within the shift:
                #   [CO overhang carried in] → [clean overhang in] → [production] →
                #   [CO fired mid-shift] → [clean fired mid-shift].
                # A whole-shift CHANGEOVER / MOULD_CLEAN is a single row.
                # STARVED/IDLE production rows span the rest of the shift (real idle
                # window) so StartTime ≠ EndTime. The RUNNING segment is emitted exactly
                # once per press-shift, so the sheet's STARVED count still equals the KPI.
                _segs = []   # (seg_status, seg_sku, seg_mins, seg_qty, seg_remark)
                if status in ("CHANGEOVER", "MOULD_CLEAN"):
                    if status == "CHANGEOVER":
                        _co_tgt = co_press_map.get(press) or (
                            dynamic_co_tracker[press][1] if press in dynamic_co_tracker else None)
                        _r = f"CO → {_co_tgt}" if _co_tgt else "CHANGEOVER"
                    else:
                        _r = "MOULD_CLEAN"
                    _segs.append((status, sku, float(SHIFT_MINS), 0, _r))
                else:                                    # RUNNING shift
                    _dleft = demand_remaining.get(sku, 0.0)
                    _seg_run = "RUNNING"
                    if status == "CAMPAIGN_IDLE":    _prod_remark = "IDLE (campaign)"; _seg_run = "CAMPAIGN_IDLE"
                    elif cured > 0:                  _prod_remark = ""
                    elif _dleft <= 0:                _prod_remark = "IDLE (demand met)"
                    elif int(round(gt_avail)) == 0:  _prod_remark = "STARVED (no GT)"
                    else:                            _prod_remark = ""
                    if _seg_co_in > 0:
                        # Overhang of a mid-shift dynamic CO into the NEXT shift —
                        # a continuation of an already-counted changeover, NOT a new
                        # event. Labelled distinctly so counting "CO → " rows in this
                        # sheet gives the true event count (= curingChangeovers).
                        _segs.append(("CHANGEOVER", sku, _seg_co_in, 0, f"CO (cont.) → {sku}"))
                    if _seg_clean_in > 0:
                        _segs.append(("MOULD_CLEAN", sku, _seg_clean_in, 0, "MOULD_CLEAN"))
                    _prod_dur = prod_mins
                    if cured == 0 and _seg_co_trig == 0 and _seg_clean_trig == 0:
                        _prod_dur = max(0.0, SHIFT_MINS - _seg_co_in - _seg_clean_in)  # idle rest of shift
                    _segs.append((_seg_run, sku, _prod_dur, cured, _prod_remark))
                    if _seg_co_trig > 0:
                        _segs.append(("CHANGEOVER", sku, _seg_co_trig, 0, f"CO → {_dyn_co_tgt}"))
                    if _seg_clean_trig > 0:
                        _segs.append(("MOULD_CLEAN", sku, _seg_clean_trig, 0, "MOULD_CLEAN"))

                _s_start = _shift_start_dt(date_str, shift)
                _cursor  = 0.0
                for _sstat, _ssku, _smins, _sqty, _srem in _segs:
                    _st = _s_start + timedelta(minutes=_cursor)
                    _en = _st + timedelta(minutes=_smins)
                    _cursor += _smins
                    cure_shift_rows.append({
                        "Date":          date_str,
                        "Shift":         shift,
                        "Machine":       press,
                        "SKUCode":       _ssku,
                        "StartTime":     _fmt_dt(_st),
                        "EndTime":       _fmt_dt(_en),
                        "Qty":           _sqty,
                        "CO_Mins":          int(round(_smins)) if _sstat == "CHANGEOVER" else 0,
                        "Mould_Clean_Mins": int(round(_smins)) if _sstat == "MOULD_CLEAN" else 0,
                        "CycleTime_min": round(ct_disp, 1),
                        "GT_Inventory":  int(round(gt_avail)),
                        "Remarks":       _srem,
                        "_status":       _sstat,
                        "_demand_left":  demand_remaining.get(_ssku, 0.0) if _sstat == "RUNNING" else None,
                    })

        # ── 5. Pool replacement: swap out any finished SKUs ─────────────────
        # If a pool SKU's demand_remaining hit 0 this day, remove it and add
        # the next best same-inch eligible SKU (highest urgency, not yet in pool).
        days_left_now = _working_days_left(day + 1) if day < planning_days else 0  # working days after today (holiday-aware)
        for machine in list(machine_pool.keys()):
            pool = machine_pool[machine]
            finished = [s for s in pool if demand_remaining.get(s, 0.0) <= 0]
            if not finished:
                continue
            new_pool   = [s for s in pool if demand_remaining.get(s, 0.0) > 0]
            pool_set   = set(new_pool)
            dom_inch   = _MACHINE_DOMINANT_INCH.get(str(machine), "")
            eligible_m = machine_skus.get(machine, set())
            # Candidates: same dominant inch, has demand, has active presses, not already in pool
            replacements = sorted(
                [s for s in eligible_m
                 if s not in pool_set
                 and demand_remaining.get(s, 0.0) > 0
                 and press_count.get(s, 0) > 0
                 and sku_inch.get(s, "") == dom_inch],
                key=lambda s: -_urgency_score(
                    s, demand_remaining, press_count, cure_ct_map, days_left_now
                ),
            )
            slots = POOL_SIZE - len(new_pool)
            new_pool.extend(replacements[:slots])
            machine_pool[machine] = new_pool
            if finished:
                _VERBOSE and print(f"    [Pool] Day {day}: machine {machine} dropped {finished}, "
                      f"added {replacements[:slots]}")

        # ── 6. GT shelf-life writeoff — now done at DAY START via _gt_expire_lots
        # (strict per-lot FIFO); day_writeoff / writeoff_total already updated there.

        # DELIVERY_PRIORITY: snapshot cured-so-far for any committed SKU whose deadline
        # is TODAY (all of today's curing is already reflected in demand_remaining here;
        # today's CO transitions below only affect tomorrow). Read-only, plan-neutral.
        if _prio_active:
            for _psku, _pdd in priority_deadline_map.items():
                if _pdd == day and _psku not in cured_by_deadline:
                    cured_by_deadline[_psku] = (
                        demand_dict.get(_psku, 0.0) - max(0.0, demand_remaining.get(_psku, 0.0)))

        # End-of-day total GT inventory (after curing + writeoff) — audits the 10k
        # plant cap. Should never exceed MAX_ENDOFDAY_GT_INVENTORY when the cap is on.
        endday_gt_inv = sum(v for v in gt_inventory.values() if v > 0)
        if _ENDOFDAY_GT_CAP_ENABLED and endday_gt_inv > MAX_ENDOFDAY_GT_INVENTORY + 1:
            print(f"  [GT-CAP WARN] Day {day}: end-of-day GT inventory "
                  f"{endday_gt_inv:,.0f} > cap {MAX_ENDOFDAY_GT_INVENTORY:,}")

        # ── 6. Apply CO transitions ───────────────────────────────────────
        # today_cos was already mould-gated at day-start (moulds committed there),
        # so every entry here is feasible — just apply the transition.
        for press, old_sku, new_sku in today_cos:
            # old_sku is None for an IDLE_PRESS_ACTIVATE cold-start (no prior SKU to release).
            if old_sku is not None:
                press_count[old_sku] = max(0, press_count.get(old_sku, 0) - 1)
                if _PRESS_RETURN_BLOCK and old_sku != new_sku:
                    press_ran[press].add(old_sku)      # this press has now left old_sku
            press_count[new_sku] = press_count.get(new_sku, 0) + 1
            press_state[press]   = {"sku": new_sku, "status": "RUNNING"}
            curing_allowable[new_sku].append(press)
            # Rule 2: a curing CO resets mould life (CO includes a clean).
            mould_life[press]  = MOULD_CLEAN_CYCLES
            clean_carry[press] = 0.0
            # Planned COs execute the CHANGEOVER in the shift chosen by co_shift_idx
            # (the shift the press finishes its old SKU — Shift A when it is already
            # free or when the CO is preemptive). Record here (once per day) so both
            # static-schedule and rolling-horizon planned COs appear in the curing
            # Changeover Plan output sheet.
            cure_co_events.append({
                "Date":       date_str,
                "Day":        day,
                "Shift":      SHIFTS[co_shift_idx.get(press, 0)],
                "Press":      press,
                "From_SKU":   (old_sku if old_sku is not None else "IDLE-START"),
                "Target_SKU": new_sku,
                "CO_Type":    ("Planned" if old_sku is not None else "Cold-Start"),
            })

        # The planned COs have now completed for the day — release the old moulds
        # that were held reserved through their pre-CO shifts.
        if _mould_gate:
            _release_deferred()

        # Daily summary — report GT-only (not carcass) for "built" KPI
        d_gt_built = sum(day_gt_built.values())  # real GT (excludes Stage-1 carcass)
        d_built    = sum(day_built.values())      # all machines (for internal tracking)
        d_cured    = sum(day_cured_d.values())
        n_active   = sum(1 for st in press_state.values() if st["status"] == "RUNNING")
        dem_met    = total_demand - sum(max(0, v) for v in demand_remaining.values())
        cov        = dem_met / total_demand * 100 if total_demand > 0 else 0
        if day % 5 == 0 or day == 1 or day == planning_days:
            _VERBOSE and print(f"  Day {day:2d} | built {d_gt_built:6,.0f} | cured {d_cured:6,.0f} | "
                  f"presses {n_active} | COs {len(today_cos)} | "
                  f"writeoff {day_writeoff:,.0f} | coverage {cov:.1f}%")
        daily_summary.append({
            "Day": day, "Date": date_str, "Holiday": bool(_holiday),
            "GT_Built": int(round(d_gt_built)), "GT_Cured": int(round(d_cured)),
            "GT_Writeoff": int(round(day_writeoff)),
            "Carcass_Writeoff": int(round(day_carcass_writeoff)),
            "EndDay_GT_Inventory": int(round(endday_gt_inv)),
            "Active_Presses": n_active, "COs_Today": len(today_cos),
            "Demand_Coverage": round(cov, 2),
        })

    # ── Final KPIs ────────────────────────────────────────────────────────────
    total_built  = sum(r["GT_Built"]  for r in daily_summary)
    total_cured  = sum(r["GT_Cured"]  for r in daily_summary)
    dem_met      = total_demand - sum(max(0, v) for v in demand_remaining.values())
    final_cov    = dem_met / total_demand * 100 if total_demand > 0 else 0
    # A press only "starves" when it's RUNNING, has zero GT, produced zero units,
    # AND still has real demand left to cure. If demand_left is already 0, the
    # SKU's job is done — the press is correctly idle, not starved (see the
    # false-positive case this excluded: press finishes a SKU's demand, GT drains
    # to 0 naturally, and the old formula counted that success as a failure).
    starvation_n = sum(
        1 for r in cure_shift_rows
        if r.get("_status") == "RUNNING" and r.get("GT_Inventory", 1) == 0
        and r.get("Qty", 0) == 0 and (r.get("_demand_left") or 0) > 0
    )

    # Env-gated diagnostic (OFF by default; no effect on scheduling). Dumps every
    # starvation event to a CSV so it can be categorized by SKU/inch/press.
    _starv_dump_path = os.environ.get("STARV_DUMP", "")
    if _starv_dump_path:
        import csv as _csv
        _srows = [
            r for r in cure_shift_rows
            if r.get("_status") == "RUNNING" and r.get("GT_Inventory", 1) == 0
            and r.get("Qty", 0) == 0 and (r.get("_demand_left") or 0) > 0
        ]
        with open(_starv_dump_path, "w", newline="") as _f:
            _w = _csv.writer(_f)
            _w.writerow(["Date", "Shift", "Press", "SKUCode", "Inch",
                         "GT_Inventory", "Demand_Left"])
            for r in _srows:
                _sku = r.get("SKUCode", "")
                _w.writerow([r.get("Date", ""), r.get("Shift", ""),
                             r.get("Machine", ""), _sku,
                             sku_inch.get(_sku, _sku[8:10] if len(_sku) >= 10 else ""),
                             r.get("GT_Inventory", 0), r.get("_demand_left", 0)])
        print(f"  [STARV_DUMP] wrote {len(_srows)} starvation events → {_starv_dump_path}")

    # Curing CO breakdown (planned schedule + reactive dynamic) and mould cleans.
    _n_co_planned = sum(1 for e in cure_co_events if e.get("CO_Type") == "Planned")
    _n_co_dynamic = sum(1 for e in cure_co_events if e.get("CO_Type") in ("Dynamic", "Early-CO"))
    # Cold-start presses (Day-1 IDLE_PRESS_ACTIVATE) are DIRECT PRODUCTION — a fresh press with
    # newly mounted moulds is not a changeover, so it starts RUNNING in Shift A with NO curing CO
    # charged and emits no Cold-Start CO event. _n_co_cold is therefore 0 (kept for back-compat /
    # any legacy event) and excluded from the CO total.
    _n_co_cold    = sum(1 for e in cure_co_events if e.get("CO_Type") == "Cold-Start")
    _n_co_total   = _n_co_planned + _n_co_dynamic + _n_co_cold
    # One event per clean trigger = the authoritative clean count (matches the
    # Changeover Plan sheet exactly). Cross-checked against total clean-minutes/480.
    _n_mould_cleans = len(mould_clean_events)
    _n_cleans_by_mins = int(round(sum(s.get("clean_mins", 0.0)
                                      for s in press_stats.values()) / MOULD_CLEAN_MINS))

    print("\n" + "=" * 70)
    print("  ROLLING PIPELINE — Results")
    print("=" * 70)
    # NOTE: GT built / cured / coverage are printed AFTER the carcass SYNC below (which drops the
    # over-carcass Stage-2 GT), so the terminal matches the output Excel + the result dict exactly.
    if _IDLE_DIAG_ON:
        _d = _IDLE_DIAG
        _tot = _d["rec_units"] + _d["ceil_units"]
        print("  " + "-" * 62)
        print("  [IDLE_DIAG] momentary curing shortfall split (by reachability):")
        print(f"    recoverable (reachable idle machine existed) : {_d['rec_units']:>10,.0f} units"
              f"  ({_d['rec_units']/_tot*100 if _tot else 0:4.1f}%)")
        print(f"    ceiling (no reachable machine — curing-limit) : {_d['ceil_units']:>10,.0f} units"
              f"  ({_d['ceil_units']/_tot*100 if _tot else 0:4.1f}%)")
        print(f"    GT-machine idle: {_d['gt_idle_min']:>12,.0f} min   Stage-1 idle: {_d['s1_idle_min']:,.0f} min")
        _ri = sorted(_d["rec_by_inch"].items(), key=lambda z: -z[1])[:6]
        _ci = sorted(_d["ceil_by_inch"].items(), key=lambda z: -z[1])[:6]
        print(f"    recoverable by inch : " + "  ".join(f"{k}:{v:,.0f}" for k, v in _ri))
        print(f"    ceiling by inch     : " + "  ".join(f"{k}:{v:,.0f}" for k, v in _ci))
    print(f"  Expired GT           : {writeoff_total:>10,.0f}")
    print(f"  Expired carcass      : {sum(carcass_waste.values()):>10,.0f}")
    print(f"  Starvation events    : {starvation_n:>10,}")
    print(f"  Curing COs (total)   : {_n_co_total:>10,}"
          f"  (planned {_n_co_planned:,} + dynamic {_n_co_dynamic:,}"
          + (f" + cold-start {_n_co_cold:,}" if _n_co_cold else "") + ")")
    print(f"  Mould cleans taken   : {_n_mould_cleans:>10,}  "
          f"(events={_n_mould_cleans}, by-minutes={_n_cleans_by_mins}; "
          f"clean {'ON' if _MOULD_CLEAN_ENABLED else 'OFF'})")
    if _INCH_PLUS3_ENABLED:
        print(f"  +3/-3 inch escapes   : {len(machine_plus3_used):>10,}  "
              f"(one-time per machine, {INCH_PLUS3_CO_MINS} min/8h each)")
        if _P3DBG:
            print(f"    [P3-debug] machine-shifts [has-room, stranded, dwell-ok, "
                  f"has-±3-SKU] = {_PLUS3_DBG}")
    if _PM_MTC_NO_OVERLAP:
        print(f"  PM/MTC no-overlap    : ON  (curing rows relocated out of maintenance windows; see [PM_MTC] line)")
    print(f"  Mould-blocked COs    : {mould_blocked_cos:>10,}  "
          f"(mould gate {'ON' if _mould_gate else 'OFF'})")
    print(f"  Mould-retargeted COs : {mould_retargeted_cos:>10,}  "
          f"(Phase-2 opt {'ON' if (_mould_gate and _mould_opt) else 'OFF'})")
    if _INCH_RULES_ENABLED and os.environ.get("INCH_DEBUG"):
        print(f"  Inch leave-gate [deficit-done, dwell-pass, dwell-BLOCK]: {_INCH_DBG}")
    if _CO_SCORER_ENABLED:
        _cs = co_scorer_stats
        _forced_txt = f" forced={_cs['forced']}" if _REACTIVE_CO else ""
        print(f"  CO scorer ({'FULL' if _SCORER_FULL_REOPT else 'ADD'}) : "
              f"planned={_cs['planned']} pullfwd={_cs['pullfwd']} "
              f"retarget={_cs['retarget']} dynamic={_cs['dynamic']} "
              f"cancelled={_cs['cancelled']} build_blocked={_cs['build_blocked']}"
              f"{_forced_txt}")
    if _MOULD_GLOBAL_OPT_ENABLED:
        print(f"  Global mould opt ({_MOULD_GLOBAL_OPT_MODE}): "
              f"direct-add={mould_global_stats['add']} "
              f"liberations={mould_global_stats['lib']}")
    if os.environ.get("MOULD_DEBUG"):
        # Cross-press exclusivity: is any mould listed in >1 press's owned set?
        _own_by_mould: dict = {}
        for _pr, _ms in _press_moulds.items():
            for _m in _ms:
                _own_by_mould.setdefault(_m, set()).add(_pr)
        _dbl = {m: ps for m, ps in _own_by_mould.items() if len(ps) > 1}
        print(f"  [MOULD-DBL] moulds claimed by >1 press (final state): {len(_dbl)}")
        # also cross-check _mould_owner vs _press_moulds consistency
        _mismatch = sum(1 for _m, _ps in _own_by_mould.items()
                        if _mould_owner.get(_m) not in _ps)
        print(f"  [MOULD-DBL] _mould_owner/_press_moulds mismatches: {_mismatch}")
        if _mould_selfcheck[0]:
            print(f"  [MOULD-LEAK] {_mould_selfcheck[0]} running press-shifts own <2 moulds")
    _eod_inv = [r["EndDay_GT_Inventory"] for r in daily_summary if "EndDay_GT_Inventory" in r]
    if _eod_inv:
        _n_over = sum(1 for v in _eod_inv if v > MAX_ENDOFDAY_GT_INVENTORY)
        print(f"  End-day GT inventory : max {max(_eod_inv):>6,.0f} | "
              f"mean {sum(_eod_inv)/len(_eod_inv):>6,.0f} | "
              f"days>{MAX_ENDOFDAY_GT_INVENTORY//1000}k {_n_over}  (cap "
              f"{'ON' if _ENDOFDAY_GT_CAP_ENABLED else 'OFF'}, fwd-buf "
              f"{'ON' if _FORWARD_BUFFER_ENABLED else 'OFF'})")

    # ── DELIVERY_PRIORITY: committed-delivery fulfillment report ────────────────
    # Best-effort + relax-report (client decision): show, per committed SKU, how much
    # was cured BY its deadline vs demand, the shortfall, whether it was met, and — when
    # not fully feasible — the earliest date it could be completed. EDF-ordered.
    priority_report: list = []
    if _prio_active and priority_precheck:
        for _r in priority_precheck:
            _s   = _r["sku"]
            _dem = _r["demand"]
            _cbd = float(cured_by_deadline.get(_s,
                     _dem - max(0.0, demand_remaining.get(_s, 0.0))))  # fallback = final
            _fin = _dem - max(0.0, demand_remaining.get(_s, 0.0))       # cured by end of month
            _short = max(0.0, _dem - _cbd)
            _met   = _cbd >= _dem - 1e-6
            _rr = dict(_r)
            _rr.update({"cured_by_deadline": _cbd, "cured_final": _fin,
                        "shortfall": _short, "met": _met})
            priority_report.append(_rr)
        print("\n  ── PRIORITY FULFILLMENT (committed-delivery SKUs, EDF order) ──")
        print(f"  {'SKU':<20} {'Demand':>7} {'Date':>10} {'Dd':>3} "
              f"{'CuredByDl':>9} {'Short':>6} {'Met':>4} {'Presses':>7} {'EarliestFull':>12}")
        for _rr in priority_report:
            _flags = []
            if _rr["undated"]:      _flags.append("EOM")
            if _rr["past_start"]:   _flags.append("PAST")
            if _rr["beyond_month"]: _flags.append(">MO")
            if _rr["structurally_infeasible"]: _flags.append("NO-CAP")
            _ef = _rr["earliest_feasible_date"] or "n/a"
            if not _rr["on_time_possible"] and _rr["earliest_feasible_date"]:
                _ef = _ef + "*"          # * = later than the deadline (physically)
            print(f"  {_rr['sku']:<20} {_rr['demand']:>7,.0f} "
                  f"{(_rr['deadline_date'] or '—'):>10} {_rr['deadline_day']:>3} "
                  f"{_rr['cured_by_deadline']:>9,.0f} {_rr['shortfall']:>6,.0f} "
                  f"{('Y' if _rr['met'] else 'N'):>4} {_rr['simul_presses']:>7} {_ef:>12}"
                  + (("  [" + ",".join(_flags) + "]") if _flags else ""))
        _n_met = sum(1 for _rr in priority_report if _rr["met"])
        print(f"  → {_n_met}/{len(priority_report)} committed SKUs fully delivered by their deadline.")

    # ── Correct Stage-1 carcass schedule (replaces tracking-only Step-3b rows) ──
    # Full 1:1 carcass for every Stage-2 GT unit, allocated by an exact time-windowed
    # max-flow with a 1-2 shift pre-build (≤1-day aging). Does NOT touch gt_inventory /
    # cured — correct utilization/qty/time accounting + a feasibility flag. OFF restores
    # the old undercounting tracking rows. See _stage1_carcass_schedule / _STAGE1_CARCASS_PASS.
    if _STAGE1_CO and _STAGE2_CARCASS_GATE:
        # Site 2 (STAGE1_CO): carcass rows from the gate's own production, capped at real
        # Stage-2 consumption (keeps pre-built-and-consumed carcass within the 1-day aging
        # window; drops only the never-consumed aged-out over-production), CO recomputed.
        _s2_gt_per_sku: dict = defaultdict(float)
        _s2_gt_consume: dict = defaultdict(list)     # per-SKU (day, shift_ord, qty) for the FIFO match
        _SORD_C = {"A": 0, "B": 1, "C": 2}
        # Stage-2 GT-consume day must be PLAN-RELATIVE (1 = plan_start), so it shares an origin with
        # the Stage-1 prod_log's loop-day. Using the raw day-of-month broke every mid-month run
        # (plan_start.day≠1): supply cidx [0..N] never matched demand cidx [60..92] → 0 carcass.
        def _planday(_dstr):
            return (datetime.strptime(str(_dstr)[:10], "%Y-%m-%d")
                    - datetime(plan_start.year, plan_start.month, plan_start.day)).days + 1
        for r in bld_shift_rows:
            if (r.get("Machine_Group") == "TBM STAGE2" and str(r.get("SKUCode")) != "CHANGEOVER"
                    and (r.get("Qty", 0) or 0) > 0):
                _s = str(r["SKUCode"])
                _s2_gt_per_sku[_s] += r["Qty"]
                _day = _planday(r["Date"])
                _s2_gt_consume[_s].append((_day, _SORD_C.get(r.get("Shift"), 0), float(r["Qty"])))
        # A4: global per-SKU carcass->Stage-2 FIFO reconcile (1 calendar-day aging). Caps
        # Stage-2 GT carcass can't back + drops aged carcass, then cascades to cured. Only
        # reduces. Makes R5/R9C pass by construction; coverage drops by design.
        if _CV2_A4:
            _a4_red_sku, _a4_red_day = _fifo_reconcile_greedy(
                (opening_carcass if _CARCASS_INV_ENABLED else {}),
                bld_shift_rows, dict(_s2_gt_consume), _s1_prod_log, _SORD_C)
            if _a4_red_sku:
                _s2_gt_per_sku = defaultdict(float); _s2_gt_consume = defaultdict(list)
                for r in bld_shift_rows:
                    if (r.get("Machine_Group") == "TBM STAGE2" and str(r.get("SKUCode")) != "CHANGEOVER"
                            and (r.get("Qty", 0) or 0) > 0):
                        _s = str(r["SKUCode"]); _s2_gt_per_sku[_s] += r["Qty"]
                        _day = _planday(r["Date"])   # plan-relative (see _planday above)
                        _s2_gt_consume[_s].append((_day, _SORD_C.get(r.get("Shift"), 0), float(r["Qty"])))
                _cured_cut_day: dict = defaultdict(float)
                for _s in sorted(_a4_red_sku):
                    _avail = _s2_gt_per_sku.get(_s, 0.0) + float(opening_gt.get(_s, 0.0))
                    _exc = sku_cured.get(_s, 0.0) - _avail
                    for _cr in sorted((r for r in cure_shift_rows
                                       if str(r.get("SKUCode")) == _s and (r.get("Qty", 0) or 0) > 0),
                                      key=lambda x: (str(x["Date"]), _SORD_C.get(x.get("Shift"), 0)),
                                      reverse=True):
                        if _exc <= 0.5:
                            break
                        _q = float(_cr.get("Qty", 0.0)); _t = min(_q, _exc)
                        _cr["Qty"] = int(round(_q - _t)); _exc -= _t
                        _cured_cut_day[str(_cr["Date"])] += _t
                        sku_cured[_s] = max(0.0, sku_cured.get(_s, 0.0) - _t)
                        daily_cured[str(_cr["Date"])] = max(0.0, daily_cured.get(str(_cr["Date"]), 0.0) - _t)
                        demand_remaining[_s] = demand_remaining.get(_s, 0.0) + _t
                for _r in daily_summary:
                    _r["GT_Built"] = int(round(_r["GT_Built"] - _a4_red_day.get(_r["Date"], 0.0)))
                    _r["GT_Cured"] = int(round(_r["GT_Cured"] - _cured_cut_day.get(_r["Date"], 0.0)))
                total_built = sum(r["GT_Built"] for r in daily_summary)
                total_cured = sum(r["GT_Cured"] for r in daily_summary)
                dem_met   = total_demand - sum(max(0, v) for v in demand_remaining.values())
                final_cov = dem_met / total_demand * 100 if total_demand > 0 else 0
                print(f"  [A4 reconcile] Stage-2 GT capped {sum(_a4_red_sku.values()):,.0f} "
                      f"over {len(_a4_red_sku)} SKU(s); cured cascade "
                      f"{sum(_cured_cut_day.values()):,.0f} -> coverage {final_cov:.2f}%")
        _carc_rows, _carc_rep, _carc_co = _stage1_carcass_rows_co(
            _s1_prod_log, dict(_s2_gt_per_sku), sku_inch,
            opening_carcass=(opening_carcass if _CARCASS_INV_ENABLED else None),
            s2_gt_consume=dict(_s2_gt_consume),
            aging_shifts=max(1, _STAGE1_CARCASS_LEAD + 1))
        if _CARCASS_CONSOLIDATE:
            # Split to the FINAL wall-clock Date/Shift first, then consolidate WITHIN each day
            # (representation-only; per-(date, SKU) carcass preserved exactly). NOTE:
            # `_consolidate_carcass_rows` deliberately re-lays each (machine, date) as ONE
            # contiguous block from 07:00 and emits a SINGLE row per (machine, date, SKU) whose
            # Shift label is only the block's START shift — the row's Qty is the whole day's
            # total and can span 2-3 shifts (e.g. a day-total of 161 units on one SKU emitted as
            # one "Shift B" row spanning into C = 585 min of production in a 480-min shift).
            # The unconditional re-split below (right before extend into bld_shift_rows) is what
            # actually re-derives the FINAL per-shift Date/Shift/Qty and caps each emitted row at
            # its own shift's physical capacity — it MUST always run, not only inside the
            # BLD_SEED_PIN_D1A branch (that branch is skipped e.g. under PLANT_2DAY_REPLAY, which
            # used to leave these day-aggregated over-capacity rows unsplit in the output).
            # HOLIDAY FIX: pass holiday_windows so a machine whose wall-clock cursor is running
            # behind schedule (e.g. an earlier PM/MTC window it had to skip over) cannot have that
            # drift carry its carcass production into a later holiday's wall-clock window — the
            # holiday is skipped over exactly like a PM/MTC window (no quantity dropped, resumes
            # on the next working shift). Empty when PLANT_HOLIDAYS is empty (bit-for-bit parity).
            _carc_split = _split_rows_at_shift_boundaries(
                _carc_rows, "Machine", mpq=_BUILDING_MPQ,
                holiday_windows=_holiday_wallclock_windows)
            _carc_rows, _carc_co = _consolidate_carcass_rows(_carc_split, sku_inch)
        # ── BLD_SEED_PIN_D1A (Stage-1): pin ONLY Day-1 Shift-A onto the plant Day-0 seed carcass ──
        # Force each seeded Stage-1 machine's SHIFT-A carcass SKU to be its seed (100% D1-A alignment)
        # at a NORMAL single-shift quantity — WITHOUT touching its Shift-B/C rows, so the normal
        # demand-driven renderer still drives B/C and the rest of the month (Stage-1 running in B/C
        # stays at the pin-OFF level). We RE-SPLIT the (consolidated) carcass rows to per-shift first,
        # so the `Shift=="A"` filter touches ONLY Shift A (consolidation merges a machine's whole day
        # into one A-dated row spanning A/B/C AND may re-order its SKUs, which would both mis-place the
        # seed and drop B/C). After the replace we leave the rows per-shift; the Excel writer's own
        # re-split at shift boundaries is idempotent. Where the seed SKU has no Stage-2 draw the
        # un-consumed carcass is accepted expired waste. Carcass is a post-hoc DISPLAY render (≠ GT,
        # not in gt_inventory, does not feed cured) → this changes only the Stage-1 carcass display +
        # the benign R5/R9C/R9G rules; never cured/built, the GT demand cap, mould, allowable, PM/MTC.
        if _BLD_SEED_PIN_D1A and initial_state is None and not _PLANT_2DAY_REPLAY:
            _d1a_date = plan_start.strftime("%Y-%m-%d")
            _s1_seed = {str(_m): _BLD_SEED_SKU[str(_m)] for _m in _BLD_SEED_MACHINES
                        if str(_m) in _S1_MACHINES}
            if _s1_seed:
                # per-shift rows so we replace ONLY Shift A (undo any consolidation merge)
                _carc_rows = _split_rows_at_shift_boundaries(
                    _carc_rows, "Machine", mpq=_BUILDING_MPQ,
                    holiday_windows=_holiday_wallclock_windows)
                def _is_d1a_s1(_r):
                    return (str(_r.get("Machine")) in _s1_seed
                            and str(_r.get("Date"))[:10] == _d1a_date
                            and _r.get("Shift") == "A")
                # the machine's own planned Shift-A carcass amount (0 if it was idle in A)
                _a_qty = {}
                for _r in _carc_rows:
                    if _is_d1a_s1(_r) and str(_r.get("SKUCode")) != "CHANGEOVER":
                        _a_qty[str(_r["Machine"])] = _a_qty.get(str(_r["Machine"]), 0.0) + float(_r.get("Qty", 0) or 0)
                _carc_rows = [r for r in _carc_rows if not _is_d1a_s1(r)]   # drop only true Shift-A rows
                _carc_co   = [e for e in _carc_co   if not _is_d1a_s1(e)]
                _forced = 0
                for _m, _S in sorted(_s1_seed.items()):
                    _ct = _bld_ct_sec(_m, _S)
                    if _ct <= 0:
                        print(f"  [Rolling] BLD_SEED_PIN_D1A: Stage-1 {_m} seed {_S!r} "
                              f"has no building CT → cannot force carcass")
                        continue
                    # normal single-shift qty: reuse the machine's own planned Shift-A amount;
                    # if it was idle in A, use one shift of its seed-SKU build rate.
                    _q = int(round(_a_qty.get(_m, 0.0))) or int(_bld_qty_per_shift(_m, _S))
                    _q = max(1, min(_q, int(SHIFT_MINS * 60.0 / _ct)))   # keep the row within one shift
                    _st = _shift_start_dt(_d1a_date, "A")
                    _en = _st + timedelta(minutes=_q * _ct / 60.0)
                    _carc_rows.append({
                        "Machine": _m, "Date": _d1a_date, "Shift": "A", "SKUCode": _S,
                        "Qty": _q, "CO_Mins": 0,
                        "StartTime": _fmt_dt(_st), "EndTime": _fmt_dt(_en),
                        "Machine_Group": _group_label(_m), "CO_Type": "carcass",
                    })
                    _forced += 1
                print(f"  [Rolling] BLD_SEED_PIN_D1A: pinned {_forced}/{len(_s1_seed)} seeded Stage-1 "
                      f"machines' Shift-A carcass to their seed (Day-1 Shift-A only; B/C demand-driven)")
        # FINAL unconditional re-split at shift boundaries — the load-bearing fix. Whatever
        # path produced `_carc_rows` (raw `_stage1_carcass_rows_co` emit, `_consolidate_carcass_rows`
        # day-level re-layout, or the BLD_SEED_PIN_D1A branch above), this guarantees NO carcass
        # row's [StartTime,EndTime]/Qty extends past its own shift — re-deriving Date/Shift from
        # wall clock and apportioning Qty by shift, exactly like every other building/curing sheet
        # (`_split_rows_at_shift_boundaries` is idempotent on rows already confined to one shift,
        # so this is a no-op when the input is already correct). Previously this only ran inside
        # the BLD_SEED_PIN_D1A branch, which is skipped under PLANT_2DAY_REPLAY (default ON) and
        # mid-month runs — leaving `_consolidate_carcass_rows`'s day-aggregated rows (Qty = a
        # whole day's total, Shift = only the block's start shift) in the final output uncapped.
        if _carc_rows:
            _carc_rows = _split_rows_at_shift_boundaries(
                _carc_rows, "Machine", mpq=_BUILDING_MPQ,
                holiday_windows=_holiday_wallclock_windows)
        bld_shift_rows[:] = [r for r in bld_shift_rows if r.get("CO_Type") != "carcass"]
        bld_shift_rows.extend(_carc_rows)
        bld_co_events.extend(_carc_co)                # so Stage-1 occupancy counts the CO
        _cm = len({r["Machine"] for r in _carc_rows})
        _ou = _carc_rep.get("opening_used", 0)
        print(f"  [Stage-1 carcass] STAGE1_CO ON: {_carc_rep['produced']:,} carcass units"
              + (f" (+{_ou:,} opening)" if _ou else "")
              + f" / {_carc_rep['demand']:,} Stage-2 GT across {_cm} machines; "
              f"{_carc_rep['co_count']:,} carcass building COs = {_carc_rep['co_mins']:,} min "
              f"(no production during CO).")
        # FIX 1 (cosmetic, display-only, zero KPI risk): `_carc_rep["unmet"]` here is computed
        # from `prod_log` — the normal demand-derived Stage-1 gate's OWN production — BEFORE the
        # Days-1-2 plant carcass force-injection below (`_PLANT_2DAY_REPLAY`) replaces the
        # derived days-1-2 carcass rows with the plant's exact snapshot. Under replay, the
        # plant-forced Stage-2 GT on days 1-2 has no matching derived Stage-1 build in
        # `prod_log` yet (that's what the injection supplies), so this pre-injection number is
        # dominated by a reporting-timing artifact, not a real shortfall — skip it and print the
        # corrected post-injection residual after the injection block instead. When replay is
        # OFF nothing later touches carcass rows, so this number IS already final — print as
        # before (bit-for-bit unchanged for OFF/inert runs).
        if _PLANT_2DAY_REPLAY:
            pass
        elif _carc_rep["unmet"] > 0:
            print(f"  [Stage-1 carcass] ⚠ INFEASIBLE: {_carc_rep['unmet']:,} carcass units cannot "
                  f"be supplied within the aging window.")
        else:
            print("  [Stage-1 carcass] FEASIBLE: Stage-1 supplies 100% of Stage-2 carcass demand "
                  "(CO-charged, pre-build within 1-day aging).")
        if _carcass_eod:
            _n_over = sum(1 for v in _carcass_eod if v > _MAX_EOD_CARCASS + 1)
            print(f"  [Stage-1 carcass] EOD carcass buffer: max {max(_carcass_eod):>5,.0f} | "
                  f"mean {sum(_carcass_eod)/len(_carcass_eod):>5,.0f} | "
                  f"cap {_MAX_EOD_CARCASS:,} | days>cap {_n_over}")
        if _S1_INCH_FLEX:
            _adv = [(m, s1_locked_inch.get(m, ""), s1_current_inch.get(m, ""))
                    for m in sorted(_S1_MACHINES)
                    if s1_current_inch.get(m, "") != s1_locked_inch.get(m, "")]
            print(f"  [Stage-1 carcass] S1_INCH_FLEX: {len(_adv)} machines advanced inch "
                  f"(one-way): {[(m, a+'→'+b) for m, a, b in _adv]}")
    elif _STAGE1_CARCASS_PASS:
        _carc_rows, _carc_rep = _stage1_carcass_schedule(
            bld_shift_rows, s1_sku_to_machines, planning_days, lead=_STAGE1_CARCASS_LEAD,
            opening_carcass=(opening_carcass if _CARCASS_INV_ENABLED else None),
            shelf_days=int(getattr(_bc_cfg, "CARCASS_SHELF_LIFE_DAYS", 1)))
        bld_shift_rows[:] = [r for r in bld_shift_rows if r.get("CO_Type") != "carcass"]
        bld_shift_rows.extend(_carc_rows)
        _cq = sum(r["Qty"] for r in _carc_rows)
        _cm = len({r["Machine"] for r in _carc_rows})
        _ou = _carc_rep.get("opening_used", 0)
        print(f"  [Stage-1 carcass] 1:1 allocation (pre-build ≤{_STAGE1_CARCASS_LEAD} shifts, "
              f"≤1-day aging): {_cq:,} carcass units / {_carc_rep['demand']:,} Stage-2 GT "
              f"across {_cm} machines" + (f"; {_ou:,} covered by opening carcass" if _ou else ""))
        if _carc_rep["unmet"] > 0:
            print(f"  [Stage-1 carcass] ⚠ INFEASIBLE: {_carc_rep['unmet']:,} carcass units cannot "
                  f"be supplied within the aging window; {_carc_rep['no_elig_units']:,} on "
                  f"{len(_carc_rep['no_elig_skus'])} SKU(s) with NO eligible Stage-1 machine.")
        else:
            print("  [Stage-1 carcass] FEASIBLE: Stage-1 supplies 100% of Stage-2 carcass demand.")

    # ── 2-DAY PLANT PLAYBACK (Stage-1 carcass): force days 1-2 Stage-1 carcass to the plant ──
    # The post-plan carcass renderers above derive Stage-1 carcass from Stage-2 GT; for the
    # playback days (1-2) we REPLACE that with the plant's exact Stage-1 assignment so every
    # (Stage-1 machine, day, shift) carcass SKU matches the snapshot. qty = min(plant_qty,
    # shift_capacity); rows are pinned inside the shift. Removes only the days-1-2 Stage-1
    # carcass rows, then appends the plant ones. Inert (bit-for-bit) when replay is OFF.
    if _PLANT_2DAY_REPLAY:
        _sched2 = _load_plant_2day_schedule()
        _rp_dates = {(plan_start + timedelta(days=_d - 1)).strftime("%Y-%m-%d"): _d
                     for _d in range(1, _PLANT_2DAY_DAYS + 1)}
        # drop existing days-1-2 Stage-1 carcass rows (they were derived, not plant-exact)
        bld_shift_rows[:] = [r for r in bld_shift_rows
                             if not (r.get("CO_Type") == "carcass"
                                     and str(r.get("Date"))[:10] in _rp_dates
                                     and str(r.get("Machine")) in _S1_MACHINES)]
        _rp_carc_n = 0
        for (_d, _sh), _rows in sorted(_sched2.items()):
            if _d not in range(1, _PLANT_2DAY_DAYS + 1):
                continue
            _date_str = (plan_start + timedelta(days=_d - 1)).strftime("%Y-%m-%d")
            _sst = _shift_start_dt(_date_str, _sh)
            _send = _sst + timedelta(minutes=SHIFT_MINS)
            _rp_carc_cursor: dict = {}               # per-(machine) cursor, RESET each shift (confined)
            for (_m, _s, _q) in _rows:
                if _m not in _S1_MACHINES:
                    continue
                _ct = _bld_ct_sec(_m, _s)
                _cap = int(SHIFT_MINS * 60.0 / _ct) if _ct > 0 else 0
                _qi = int(min(round(_q), _cap))
                if _qi <= 0:
                    continue
                _cst = min(max(_sst, _rp_carc_cursor.get(_m, _sst)), _send)
                _en = min(_cst + timedelta(minutes=_qi * _ct / 60.0), _send)   # clamp inside the shift
                _rp_carc_cursor[_m] = _en
                bld_shift_rows.append({
                    "Machine": _m, "Date": _date_str, "Shift": _sh, "SKUCode": _s,
                    "Qty": _qi, "CO_Mins": 0,
                    "StartTime": _fmt_dt(_cst), "EndTime": _fmt_dt(_en),
                    "Machine_Group": _group_label(_m), "CO_Type": "carcass",
                    "_replay": True,           # exempt from the shift-boundary splitter
                })
                _rp_carc_n += 1
        print(f"  [Plant2Day] Stage-1 carcass: forced {_rp_carc_n} plant rows for days "
              f"1-{_PLANT_2DAY_DAYS} (exact snapshot replay)")

    # FIX 1 (cosmetic, display-only, zero KPI risk): the TRUE final Stage-1 carcass residual,
    # recomputed AFTER the Days-1-2 plant carcass force-injection above (see the guarded print
    # in the STAGE1_CO block for why the pre-injection number is stale under replay). Uses the
    # FINAL bld_shift_rows (post-injection) against the same per-SKU Stage-2 GT targets
    # `_stage1_carcass_rows_co` used, capped per SKU exactly like that function caps `produced`
    # at `T` — so this is the honest post-injection shortfall, not a re-derivation of the
    # planning logic. Does not touch any row Qty, gt_inventory, or the cured/built KPI.
    if _PLANT_2DAY_REPLAY and _STAGE1_CO and _STAGE2_CARCASS_GATE:
        _gt_int_f = {str(s): int(round(float(q))) for s, q in _s2_gt_per_sku.items()}
        _open_f = {str(k): float(v) for k, v in (opening_carcass or {}).items() if v and float(v) > 0}
        _open_used_f = {s: min(int(_open_f.get(s, 0.0)), q) for s, q in _gt_int_f.items()}
        _final_supplied: dict = defaultdict(float)
        for _r in bld_shift_rows:
            if _r.get("CO_Type") == "carcass" and str(_r.get("SKUCode")) != "CHANGEOVER":
                _final_supplied[str(_r["SKUCode"])] += float(_r.get("Qty", 0) or 0)
        _total_gt_f = float(sum(_gt_int_f.values()))
        _supplied_f = float(sum(_open_used_f.values()))
        for _s, _T in _gt_int_f.items():
            _cap_f = max(0, _T - _open_used_f.get(_s, 0))
            _supplied_f += min(_final_supplied.get(_s, 0.0), _cap_f)
        _true_unmet = round(max(0.0, _total_gt_f - _supplied_f))
        if _true_unmet > 0:
            print(f"  [Stage-1 carcass] ⚠ INFEASIBLE: {_true_unmet:,} carcass units cannot "
                  f"be supplied within the aging window (post Days-1-2 plant injection).")
        else:
            print("  [Stage-1 carcass] FEASIBLE: Stage-1 supplies 100% of Stage-2 carcass demand "
                  "(CO-charged, pre-build within 1-day aging; post Days-1-2 plant injection).")

    # ── Write Excel outputs (same format as legacy pipeline) ─────────────────
    closing_gt_bal = {sku: v for sku, v in gt_inventory.items() if v > 0}

    # RAW allowable-matrix membership sets for the Demand-Fulfillment Skip_Reason
    # (rule: present in allowable → blank; absent → "missing from allowable matrix").
    # Lock-blind / master-level, so inch-locked or press-starved SKUs are NOT flagged.
    try:
        from connection import B2C_ETL as _BETL_SR
        _bld_matrix_skus = {str(_r["SKUCode"]).strip()
                            for _, _r in _BETL_SR(engine).load_machine_allowable().iterrows()
                            if _r.get("Machines")}
    except Exception as _e_sr:   # noqa: BLE001
        print(f"  [Skip_Reason] building matrix set unavailable ({_e_sr}); legacy label")
        _bld_matrix_skus = None
    try:
        _cur_master_skus = {str(_r["SKUCode"]).strip()
                            for _, _r in cetl.load_curing_allowable().iterrows()
                            if _r.get("Machines")}
    except Exception as _e_sr2:  # noqa: BLE001
        print(f"  [Skip_Reason] curing master set unavailable ({_e_sr2}); legacy label")
        _cur_master_skus = None

    _carcass_cap_by_sku, _gt_drop_by_sku = _write_rolling_building_excel(
        output_path    = build_output,
        bld_shift_rows = bld_shift_rows,
        bld_co_events  = bld_co_events,
        df_day0        = df_day0,
        sku_machine_map = sku_machine_map,
        opening_gt     = opening_gt,
        demand_dict    = demand_dict,
        planning_days  = planning_days,
        working_days   = _working_days_count(plan_start, planning_days),  # holiday-adjusted util denom
        n_curing_cos   = _n_co_total,  # ACTUAL executed COs (planned+dynamic). co_events is the stale pre-computed plan — defer/pull-forward change the real count.
        endday_gt_by_date = {r["Date"]: r["EndDay_GT_Inventory"] for r in daily_summary},
        cure_ct_map    = cure_ct_map,
        curing_allowable = dict(curing_allowable),
        sku_moulds     = ({s: set(m) for s, m in _sku_moulds.items()}
                          if _sku_moulds else None),
        gt_waste_map      = dict(writeoff_cum),
        carcass_waste_map = dict(carcass_waste),
        expiry_rows       = expiry_rows,
        holiday_dates     = {(plan_start + timedelta(days=_d - 1)).strftime("%Y-%m-%d")
                             for _d in _holiday_days},
        pm_mtc_rows       = _pm_mtc_display_rows(plan_start, planning_days, "building"),
        bld_matrix_skus   = _bld_matrix_skus,
        cur_master_skus   = _cur_master_skus,
    )
    # ── SYNC cured to the dropped Stage-2 GT: the building writer capped Stage-2 GT rows to the
    # displayed carcass (R5), dropping _gt_drop_by_sku[sku] units of Stage-2 GT. That GT is no
    # longer built → it can't be cured → reduce cured by the SAME amount (bounded by the SKU's
    # cured), and mark it back as unmet demand. IMPORTANT: a SKU's GT also comes from carcass-free
    # groups (Unistage/VMI/BJ), so we reduce ONLY by the dropped Stage-2 GT — never cap cured to
    # total carcass. Non-Stage-2 GT is untouched.
    if _CARCASS_MIN_ENFORCE and _gt_drop_by_sku:
        _sync_cured_drop = 0.0
        for _sk, _gtd in _gt_drop_by_sku.items():
            if _gtd <= 1e-6:
                continue
            _drop = min(float(sku_cured.get(_sk, 0.0)), float(_gtd))
            if _drop <= 1e-6:
                continue
            _sync_cured_drop += _drop
            sku_cured[_sk] = float(sku_cured.get(_sk, 0.0)) - _drop
            demand_remaining[_sk] = demand_remaining.get(_sk, 0.0) + _drop
            # reduce this SKU's curing production rows, latest first, by _drop (2 tyres/cavity-cycle
            # but rows carry unit Qty already) — keep the curing sheet consistent with the KPI.
            _rem = _drop
            for _cr in sorted((r for r in cure_shift_rows if str(r.get("SKUCode")) == _sk
                               and float(r.get("Qty", 0) or 0) > 0),
                              key=lambda x: str(x.get("StartTime", "")), reverse=True):
                if _rem <= 1e-6:
                    break
                _cq = float(_cr.get("Qty", 0) or 0); _cd = min(_cq, _rem)
                _cr["Qty"] = int(round(_cq - _cd)); _rem -= _cd
            # reduce daily_cured (latest days first)
            _rem = _drop
            for _dk in sorted(daily_cured, reverse=True):
                if _rem <= 1e-6:
                    break
                _dv = float(daily_cured.get(_dk, 0.0)); _dd_ = min(_dv, _rem)
                daily_cured[_dk] = _dv - _dd_; _rem -= _dd_
        # reduce daily_summary GT_Built by the dropped GT (latest days), GT_Cured by dropped cured
        _gt_drop_total = sum(_gt_drop_by_sku.values())
        _rem = _gt_drop_total
        for _row in sorted(daily_summary, key=lambda r: r["Day"], reverse=True):
            if _rem <= 1e-6:
                break
            _dv = float(_row["GT_Built"]); _dd_ = min(_dv, _rem)
            _row["GT_Built"] = int(round(_dv - _dd_)); _rem -= _dd_
        _rem = _sync_cured_drop
        for _row in sorted(daily_summary, key=lambda r: r["Day"], reverse=True):
            if _rem <= 1e-6:
                break
            _dv = float(_row["GT_Cured"]); _dd_ = min(_dv, _rem)
            _row["GT_Cured"] = int(round(_dv - _dd_)); _rem -= _dd_
        # recompute headline KPIs
        total_built  = sum(r["GT_Built"] for r in daily_summary)
        total_cured  = sum(r["GT_Cured"] for r in daily_summary)
        dem_met      = total_demand - sum(max(0, v) for v in demand_remaining.values())
        final_cov    = dem_met / total_demand * 100 if total_demand > 0 else 0
        if _sync_cured_drop or _gt_drop_total:
            print(f"  [SYNC] dropped {_gt_drop_total:,.0f} GT / {_sync_cured_drop:,.0f} cured to match "
                  f"displayed carcass (GT/carcass/cured now in sync). New cured {total_cured:,.0f} "
                  f"/ coverage {final_cov:.1f}%.")

    # Headline KPIs printed here (AFTER the carcass SYNC) so the terminal matches the output Excel
    # + the result dict exactly — total_built/total_cured/final_cov are now the post-SYNC values.
    print(f"  Total GT built       : {total_built:>10,.0f}")
    print(f"  Total cured          : {total_cured:>10,.0f}")
    print(f"  Demand coverage      : {final_cov:>9.1f}%  ({dem_met:,.0f} / {total_demand:,.0f})")

    _write_rolling_curing_excel(
        output_path       = curing_output,
        cure_shift_rows   = cure_shift_rows,
        cure_co_events    = cure_co_events,
        mould_clean_events= mould_clean_events,
        press_stats       = dict(press_stats),
        press_sku_stats   = dict(press_sku_stats),
        daily_cured       = dict(daily_cured),
        sku_cured         = dict(sku_cured),
        closing_gt_bal    = closing_gt_bal,
        build_by_shift_sku= build_by_shift_sku,
        opening_gt        = opening_gt,
        demand_dict       = demand_dict,
        cure_ct_map       = cure_ct_map,
        curing_allowable  = dict(curing_allowable),
        planning_days     = planning_days,
        plan_start        = plan_start,
        df_day0           = df_day0,
        mould_life        = dict(mould_life),
        mould_info        = ({
            "press_moulds":  {p: sorted(ms) for p, ms in _press_moulds.items()},
            "final_sku":     {p: st["sku"] for p, st in press_state.items()},
            "mould_skus":    {m: set(s) for m, s in _mould_skus.items()},
            "sku_moulds":    {s: set(m) for s, m in _sku_moulds.items()},
            "press_count":   dict(press_count),
            "blocked":       mould_blocked_cos,
            "retargeted":    mould_retargeted_cos,
            "day0_topups":   mould_day0_topups,
            "scorer":        (co_scorer_stats if _CO_SCORER_ENABLED else None),
            "events":        mould_events,          # per-swap movement log (day, press, sku, added, removed)
            "assignments":   mould_assignments,     # full (press, sku, 2-moulds) timeline incl Day-0
        } if _mould_gate else None),
        sku_desc_map      = sku_desc_map,
        sku_machine_map   = sku_machine_map,
        pm_mtc_rows       = _pm_mtc_display_rows(plan_start, planning_days, "curing"),
        bld_matrix_skus   = _bld_matrix_skus,
        cur_master_skus   = _cur_master_skus,
    )

    # Client output rule: next to every SKU-code column write its description, and
    # next to every BUILDING machine-code column write the machine name. Building
    # workbook gets both; curing workbook gets descriptions only (its "Machine" is a
    # press id, not a building machine). Cosmetic post-pass — never alters the plan
    # (runs AFTER total_cured is computed). LABELS=0 reproduces label-free sheets.
    if _labels_on:
        _inject_label_columns(build_output,  sku_desc_map or {}, BUILDING_MACHINE_NAMES)
        _inject_label_columns(curing_output, sku_desc_map or {}, None)

    return {
        "total_built":       total_built,
        "total_cured":       total_cured,
        "gt_writeoff":       writeoff_total,             # expired GT (aged out)
        "carcass_writeoff":  sum(carcass_waste.values()),  # expired carcass (aged out)
        "starvation_events": starvation_n,
        "demand_coverage":   final_cov,
        "demand_remaining":  demand_remaining,
        "gt_inventory":      dict(gt_inventory),
        "daily_summary":     daily_summary,
        "co_events":         co_events,
        "n_co":              _n_co_total,      # planned + dynamic (was planned-only)
        "n_co_planned":      _n_co_planned,
        "n_co_dynamic":      _n_co_dynamic,
        "n_mould_cleans":    _n_mould_cleans,
        "mould_blocked_cos": mould_blocked_cos,
        "mould_retargeted_cos": mould_retargeted_cos,
        "co_scorer_stats": co_scorer_stats if _CO_SCORER_ENABLED else None,
        "build_output":      build_output,
        "curing_output":     curing_output,
        # Per-day capacity utilisation (30-31 rows) for jkt_plan_capacityUtilisation.
        # Curing daily = (production + mould-clean + CO) / available press-minutes.
        "daily_capacity_util": _daily_capacity_util(
            cure_shift_rows, bld_shift_rows, press_stats, planning_days),
        # DELIVERY_PRIORITY committed-delivery fulfillment (empty when inert).
        "priority_report":   priority_report,
        # MID-MONTH: full physical state captured at the start of snapshot_at_day (None if unused).
        "state_snapshot":    _captured_state,
    }


# ══════════════════════════════════════════════════════════════════════════════
# LEGACY PIPELINE (31-day LP — use --legacy flag)
# ══════════════════════════════════════════════════════════════════════════════

def run_pipeline(
    demand_path:   str | None = None,
    cc_output:     str | None = None,
    build_output:  str | None = None,
    curing_output: str | None = None,
    plan_start:    datetime | None = None,
    planning_days: int | None = None,
) -> dict:
    demand_path   = demand_path   or DEMAND_FILE
    cc_output     = cc_output     or CC_OUTPUT
    build_output  = build_output  or BUILD_OUTPUT
    curing_output = curing_output or CURING_OUTPUT
    plan_start    = plan_start    or PLAN_START
    planning_days = planning_days or PLANNING_DAYS

    print("\n" + "=" * 70)
    print("  PIPELINE — Step 1: Curing Consumption (Dynamic)")
    print("=" * 70)
    cc_result = run_dynamic_consumption(
        demand_path=demand_path, output_path=cc_output,
        plan_start=plan_start, planning_days=planning_days,
        max_co_per_day=MAX_CHANGEOVERS_PER_DAY,
    )
    co_events = cc_result["co_events"]
    df_day0   = cc_result["df_day0"]
    print(f"\n  [Pipeline] Step 1 complete — {len(co_events)} CO events → {os.path.basename(cc_output)}")

    df_cons = df_day0[df_day0["Category"].isin({"Runner-In", "Non-Runner-In"})].copy()
    if "Skip_Reason" in df_cons.columns:
        df_cons = df_cons[
            df_cons["Skip_Reason"].isna() | (df_cons["Skip_Reason"].astype(str).str.strip() == "")
        ].copy()
        df_cons = df_cons.drop(columns=["Skip_Reason"], errors="ignore")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=tempfile.gettempdir())
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        df_cons.to_excel(writer, sheet_name="Consumption Summary", index=False)

    print("\n" + "=" * 70)
    print("  PIPELINE — Step 2: Building Scheduler (B2C)")
    print("=" * 70)
    try:
        build_result = run_from_database_b2c(
            plan_start=plan_start, consumption_path=tmp.name,
            output_path=build_output, planning_days=planning_days,
            external_co_schedule=co_events,
            max_changeovers_per_day=MAX_CHANGEOVERS_PER_DAY,
            min_campaign_mins=MIN_CAMPAIGN_MINS,
            build_lead_shifts=BUILD_LEAD_SHIFTS,
        )
    finally:
        os.unlink(tmp.name)
    print(f"\n  [Pipeline] Step 2 complete → {os.path.basename(build_output)}")

    print("\n" + "=" * 70)
    print("  PIPELINE — Step 3: Curing Schedule (B2C)")
    print("=" * 70)
    curing_result = run_curing_b2c(
        building_path=build_output, output_path=curing_output,
        demand_path=demand_path, plan_start=plan_start, planning_days=planning_days,
    )
    total_cured = sum(curing_result["daily_cured"].values())
    print(f"\n  [Pipeline] Step 3 complete — Total cured: {total_cured:,.0f} → {os.path.basename(curing_output)}")

    return {
        "co_events": co_events, "n_co": len(co_events),
        "cc_output": cc_output, "build_output": build_output, "curing_output": curing_output,
        "build_result": build_result, "curing_result": curing_result,
    }


if __name__ == "__main__":
    _args   = sys.argv[1:]
    _legacy = "--legacy" in _args
    _demand = next((a for a in _args if not a.startswith("--")), None)

    if _legacy:
        print("[Pipeline] Legacy 31-day LP mode (--legacy flag)")
        result = run_pipeline(demand_path=_demand)
        print("\n" + "█" * 70)
        print("  LEGACY PIPELINE COMPLETE")
        print("█" * 70)
        print(f"  Changeovers scheduled : {result['n_co']}")
        print(f"  1. Curing consumption : {result['cc_output']}")
        print(f"  2. Building schedule  : {result['build_output']}")
        print(f"  3. Curing schedule    : {result['curing_output']}")
        total = sum(result["curing_result"]["daily_cured"].values())
        print(f"  Total cured (month)   : {total:,.0f} tyres")
        print("█" * 70)
    else:
        print("[Pipeline] Rolling day-by-day mode (new architecture)")
        result = run_rolling_pipeline(demand_path=_demand)
        print("\n" + "█" * 70)
        print("  ROLLING PIPELINE COMPLETE")
        print("█" * 70)
        print(f"  Curing COs (total)    : {result['n_co']:>10,}"
              f"  (planned {result.get('n_co_planned', 0):,} + dynamic {result.get('n_co_dynamic', 0):,})")
        print(f"  Mould cleans taken    : {result.get('n_mould_cleans', 0):>10,}")
        print(f"  GT built (month)      : {result['total_built']:>10,.0f}")
        print(f"  GT cured (month)      : {result['total_cured']:>10,.0f}")
        print(f"  GT written off        : {result['gt_writeoff']:>10,.0f}")
        print(f"  Starvation events     : {result['starvation_events']:>10,}")
        print(f"  Demand coverage       : {result['demand_coverage']:>9.1f}%")
        print(f"  Building output       : {result['build_output']}")
        print(f"  Curing  output        : {result['curing_output']}")
        print("█" * 70)
        print("\n  Worst 10 SKUs by remaining demand:")
        rem = sorted(result["demand_remaining"].items(), key=lambda x: -x[1])[:10]
        for sku, qty in rem:
            print(f"    {sku}: {qty:,.0f} units remaining")
