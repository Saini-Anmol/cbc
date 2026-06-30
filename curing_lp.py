"""
JK Tyre BTP — PCR Curing LP Scheduler v4  (standalone, single-file build)
Designed By — Paranjay Dodiya — Algo8 AI Pvt Ltd
===============================================
SELF-CONTAINED: this file carries the COMPLETE curing-schedule pipeline —
Config → MouldTracker → ETL → LP_Solver → Rounder → ScheduleBuilder →
JK_LP_Curing_Scheduler_v2 → ExcelExporter → run_from_excel / run_from_database
→ _post_process_schedule_excel — with no project-internal imports. It mirrors
V1/routes/schedule_route.py (the Flask pipeline's engine) including the
default-cycle-time handling and Excel post-processing, but reads its knobs from
the Config class + the jkt_plan_params row instead of config.yaml.

Run it directly:
    python3 curing.py                       # uses Config.PLAN_ID, reads the DB
or import run_from_excel(...) / run_from_database(plan_id=...) from it.
===============================================
v4 fixes (over v3):
-------------------
 [Fix 3] Rounder over-reserved CO budget when SKUs got trimmed
   - Old behaviour: co_cost = (n_skus - 1) * 300 was deducted up front,
     then SKUs were trimmed if capacity ran out — but the reserved CO for
     a dropped SKU was never given back, so production capacity was lost.
   - v4: walk the SKU list in priority order and charge a CO ONLY for
     SKUs actually kept on that press. Same accurate CO total when every
     SKU fits, but no wasted budget when the rounder has to drop one.

 [Fix 4] Mould tracker too strict for idle-press utilisation
   - Old behaviour: `can_assign(sku)` required ≥ MOULDS_PER_PRESS *free*
     moulds. If all of a SKU's moulds were locked to one continuity press
     and another idle press was physically compatible, the LP couldn't
     allocate to it — the bound was zeroed out.
   - v4: `Config.PERMISSIVE_MOULD_ELIGIBILITY` (default True). When set,
     `get_eligible_machines_with_moulds` accepts any compatible press
     for which the SKU has ≥ MOULDS_PER_PRESS *total* compatible moulds
     in the master (free or locked). Operationally this assumes moulds
     can be physically moved between presses if needed.

 [Misc]
   - Removed the inline hardcoded-machines hack inside _build_continuity
     (was redundant — locked_mins.get(str(m), 0.0) already defaults).
   - Removed debug print/dump statements in run().
   - OUTPUT_FILE typo: 311Days -> 31Days.

v3 fixes (carried forward):
---------------------------
 [Fix 1] Continuity → LP changeover row insertion (ScheduleBuilder)
 [Fix 2] continuity-aware mould eligibility (whitelist continuity press)
         + Rounder str(m) capacity-key bug
         + Topup CO accounting on continuity presses

Architecture (unchanged)
------------------------
Phase 1 : ETL          — load & clean all input data
Phase 2 : LP Solve     — globally optimal press-minute allocation
Phase 3 : Rounding     — convert continuous LP solution to integer cycles
Phase 4 : Schedule     — build shift-wise row-level schedule
Phase 5 : Export       — Excel output
"""
from __future__ import annotations  # Python 3.9 compatibility for `X | Y` annotations

import ast
import math
import warnings
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import linprog

try:
    from sqlalchemy import create_engine
except ImportError:
    create_engine = None

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
# DB credentials come from cbc_env (.env) — never hardcoded. If the helper or
# .env is missing, fail loudly rather than fall back to a baked-in secret.
from cbc_env import db_config as _cbc_db_config
_DB = _cbc_db_config()


class Config:
    # ── database (sourced from .env via cbc_env) ───────────────────────────────
    DB_SERVER   = _DB["host"]
    DB_NAME     = _DB["database"]
    DB_USER     = _DB["user"]
    DB_PASSWORD = _DB["password"]

    # ── plan identity ─────────────────────────────────────────────────────────
    # Primary key in jkt_plan_params. planStartDate and planEndDate are read
    # from that row at runtime and override PLAN_DATE / PLANNING_DAYS below.
    PLAN_ID = "BTP_June_Plan_V_384072"

    # Plan-params table name. Defaults to the planning table.
    PLAN_PARAMS_TABLE = "jkt_plan_params"

    # ── planning horizon ──────────────────────────────────────────────────────
    PLANNING_DAYS    = 30
    SHIFTS_PER_DAY   = 3
    HOURS_PER_SHIFT  = 8
    SHIFT_START_HOUR = 7

    # ── press & mould ─────────────────────────────────────────────────────────
    CAVITIES_PER_MOULD = 2
    MOULDS_PER_PRESS   = 2
    NEW_MOULD_LIFE     = 3000

    # ── downtime constants (minutes) ──────────────────────────────────────────
    CHANGEOVER_DURATION_MIN = 300
    CLEANING_DURATION_MIN   = 120
    LOAD_UNLOAD_BUFFER_MIN  = 2.3
    PRESS_EFFICIENCY        = 0.94
    # Default EFFECTIVE cycle time (min) for demand SKUs missing from the
    # cycle-time master. This is the FINAL per-cycle value — it ALREADY includes
    # the load/unload buffer and press efficiency, so it is used directly and NOT
    # run through (raw + buffer) / efficiency again.
    DEFAULT_CYCLE_TIME_MIN  = 17.0

    # ── changeover scheduling ─────────────────────────────────────────────────
    MAX_CHANGEOVERS_PER_SHIFT = 5
    CHANGEOVER_PENALTY_WEIGHT = 0.01

    # ── eligibility policy (v4) ───────────────────────────────────────────────
    # True  : count TOTAL compatible moulds (free + locked) when deciding
    #         whether a SKU can be scheduled on a compatible press. Lets
    #         idle presses be used even if the SKU's moulds are currently
    #         on another press (assumes moulds can be moved manually).
    # False : strict — require MOULDS_PER_PRESS free moulds in the global pool.
    PERMISSIVE_MOULD_ELIGIBILITY = True

    PLAN_DATE = datetime(2026, 6, 1, 7, 0, 0)

    # ── output ────────────────────────────────────────────────────────────────
    TYRE_TYPE   = "pcr"
    OUTPUT_FILE = f"BTP_PCR_Curing_LP_v4_PlanSchedule_May_{PLAN_DATE.date()}_31V8Days.xlsx"

    @classmethod
    def avail_mins(cls) -> float:
        return cls.PLANNING_DAYS * cls.SHIFTS_PER_DAY * cls.HOURS_PER_SHIFT * 60

    @classmethod
    def units_per_cleaning_cycle(cls) -> int:
        return cls.NEW_MOULD_LIFE * cls.CAVITIES_PER_MOULD * cls.MOULDS_PER_PRESS


# ══════════════════════════════════════════════════════════════════════════════
# MOULD TRACKER  (v4 — permissive eligibility for utilisation)
# ══════════════════════════════════════════════════════════════════════════════
class MouldTracker:
    """
    v4 change: get_eligible_machines_with_moulds now uses
    `total moulds for SKU` (free + locked) when
    Config.PERMISSIVE_MOULD_ELIGIBILITY is True. This unlocks idle
    compatible presses for SKUs whose moulds are currently on other
    presses, which is what you want when the operator can move moulds.
    """

    def __init__(self):
        self._ledger: dict[str, dict] = {}

    def load_from_df(self, df_mould: pd.DataFrame, df_running: pd.DataFrame):
        mould_col = "MouldNo" if "MouldNo" in df_mould.columns else "Mould"
        for _, row in df_mould.iterrows():
            mid = str(row[mould_col])
            sku = str(row["Matl.Code"])
            if mid not in self._ledger:
                self._ledger[mid] = {
                    "compatible_skus":   set(),
                    "life_remaining":    Config.NEW_MOULD_LIFE,
                    "assigned_machine":  None,
                }
            self._ledger[mid]["compatible_skus"].add(sku)

        for _, row in df_running.iterrows():
            machine = str(row["Machine"])
            life    = int(row.get("MouldLife_remaining", Config.NEW_MOULD_LIFE))
            raw = row.get("MouldNos", row.get("MouldNo", None))
            if raw is None:
                continue
            moulds = raw if isinstance(raw, list) else [str(raw)]
            for mould in moulds:
                mould = str(mould).strip()
                if mould in self._ledger:
                    self._ledger[mould]["life_remaining"]   = life
                    self._ledger[mould]["assigned_machine"] = machine

    def load_from_excel(self, mould_path: str, running_path: str = None):
        df_mould = pd.read_excel(mould_path)
        df_mould = df_mould[df_mould.get("Active Flag", pd.Series([True]*len(df_mould))).astype(bool)]
        df_mould = df_mould.rename(columns={"Matl.Code": "Matl.Code"})
        if running_path:
            df_running = pd.read_excel(running_path)
            if "MouldLife_remaining" not in df_running.columns:
                df_running["MouldLife_remaining"] = Config.NEW_MOULD_LIFE
        else:
            df_running = pd.DataFrame(columns=["Machine", "SKUCode", "MouldNo", "MouldLife_remaining"])
        self.load_from_df(df_mould, df_running)

    def available_moulds_for_sku(self, sku: str) -> list[str]:
        return [
            mid for mid, data in self._ledger.items()
            if sku in data["compatible_skus"]
            and data["assigned_machine"] is None
        ]

    def total_moulds_for_sku(self, sku: str) -> int:
        return sum(1 for d in self._ledger.values() if sku in d["compatible_skus"])

    def can_assign(self, sku: str) -> bool:
        """Strict: ≥ MOULDS_PER_PRESS free moulds."""
        return len(self.available_moulds_for_sku(sku)) >= Config.MOULDS_PER_PRESS

    def can_schedule(self, sku: str) -> bool:
        """Permissive: ≥ MOULDS_PER_PRESS total moulds (free or locked)."""
        return self.total_moulds_for_sku(sku) >= Config.MOULDS_PER_PRESS

    def get_eligible_machines_with_moulds(
        self,
        sku: str,
        candidate_machines: list,
        continuity_machines: set | list | None = None,
    ) -> list:
        """
        Permissive mode (Config.PERMISSIVE_MOULD_ELIGIBILITY=True):
            Allow any compatible press if ≥ MOULDS_PER_PRESS total
            compatible moulds exist for this SKU. Continuity presses
            always pass.
        Strict mode:
            Allow only continuity presses, OR any compatible press if
            ≥ MOULDS_PER_PRESS *free* moulds exist.
        """
        cont = set()
        for cm in (continuity_machines or []):
            cont.add(cm)
            cont.add(str(cm))
            try:
                cont.add(int(cm))
            except (ValueError, TypeError):
                pass

        if Config.PERMISSIVE_MOULD_ELIGIBILITY:
            if not self.can_schedule(sku):
                # Not enough physical moulds anywhere — only continuity machines work
                return [m for m in candidate_machines
                        if m in cont or str(m) in cont]
            return list(candidate_machines)

        # Strict
        has_free = self.can_assign(sku)
        eligible = []
        for m in candidate_machines:
            if m in cont or str(m) in cont:
                eligible.append(m)
            elif has_free:
                eligible.append(m)
        return eligible

    def assign_moulds(self, sku: str, machine: str) -> list[str]:
        avail = self.available_moulds_for_sku(sku)
        if len(avail) < Config.MOULDS_PER_PRESS:
            raise ValueError(
                f"Cannot assign {Config.MOULDS_PER_PRESS} moulds for "
                f"SKU={sku} on Machine={machine}: only {len(avail)} free"
            )
        chosen = sorted(avail, key=lambda m: -self._ledger[m]["life_remaining"])[: Config.MOULDS_PER_PRESS]
        for mid in chosen:
            self._ledger[mid]["assigned_machine"] = machine
        return chosen

    def release_moulds(self, mould_ids: list[str]):
        for mid in mould_ids:
            if mid in self._ledger:
                self._ledger[mid]["assigned_machine"] = None

    def mould_life(self, mould_id: str) -> int:
        return self._ledger.get(mould_id, {}).get("life_remaining", Config.NEW_MOULD_LIFE)

    def avg_life_remaining_for_sku(self, sku: str) -> float:
        moulds = self.available_moulds_for_sku(sku)
        if not moulds:
            return 0.0
        return sum(self._ledger[m]["life_remaining"] for m in moulds) / len(moulds)

    @property
    def summary(self) -> pd.DataFrame:
        rows = []
        for mid, d in self._ledger.items():
            rows.append({
                "MouldNo":          mid,
                "Compatible_SKUs":  ", ".join(sorted(d["compatible_skus"])),
                "Life_Remaining":   d["life_remaining"],
                "Assigned_Machine": d["assigned_machine"] or "FREE",
            })
        return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# ETL  (unchanged from v3)
# ══════════════════════════════════════════════════════════════════════════════
class ETL:
    def __init__(self, engine=None, tyre_type: str = "pcr"):
        self.engine = engine
        self.t = tyre_type

    def _sql(self, q): return pd.read_sql(q, self.engine)

    def load_demand(self, csv_path: str) -> pd.DataFrame:
        # Accept either .csv (iterative-demand outputs) or .xlsx (requirement_summary).
        if str(csv_path).lower().endswith(".csv"):
            df = pd.read_csv(csv_path)
        else:
            df = pd.read_excel(csv_path)
        # First-run files have "Requirement"; iterative re-plans have "Updated_Requirement".
        qty_col = "Updated_Requirement" if "Updated_Requirement" in df.columns else "Requirement"
        df = (df.groupby("SKUCode")
                .agg(Quantity=(qty_col, "sum"),
                     Priority=("ConsolidatedPriorityScore", "max"))
                .reset_index())
        df = df[df["Quantity"] > 0].copy()
        df.to_excel("load_demand.xlsx", index=False)
        return df

    def load_cycle_times(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT Sapcode AS SKUCode, `Cure Time` AS Raw "
            f"FROM {Config.DB_NAME}.Master_Curing_Design_CycleTime"
        )
        df["CycleTime_min"] = np.round(
            (df["Raw"] + Config.LOAD_UNLOAD_BUFFER_MIN) / Config.PRESS_EFFICIENCY
        )
        df = df[["SKUCode","CycleTime_min"]].drop_duplicates("SKUCode")
        df['SKUCode'] = df['SKUCode'].str.strip()
        df.to_excel("load_cycle_times.xlsx", index=False)
        return df

    def load_machine_allowable(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT * FROM {Config.DB_NAME}"
            f".Master_Curing_Allowable_Machines_source"
        )
        df = df.rename(columns={"SKU Code":"SKUCode"})
        mcols = [c for c in df.columns if str(c).isdigit()]
        df["Machines"] = df.apply(
            lambda r: [str(c) for c in mcols if str(r[c]).strip().lower()=="yes"], axis=1
        )
        df = df[["SKUCode","Machines"]]
        df['Machines'] = df['Machines'].apply(lambda lst: list(map(int, lst)))
        df.to_excel("load_machine_allowable.xlsx", index=False)
        return df

    def load_gt_inventory(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory"
            f" FROM {Config.DB_NAME}.gt_inventory_manual"
        )
        df.to_excel("load_gt_inventory.xlsx", index=False)
        return df

    def load_running_moulds(self) -> pd.DataFrame:
        wc_master = self._sql(f"SELECT * FROM {Config.DB_NAME}.Master_WC_Master")
        wc_master = wc_master[['wcID', 'WCNAME']]

        df = self._sql(f"SELECT * FROM {Config.DB_NAME}.testing_Daily_Running_Moulds")
        df = df.drop(columns=["updatedAt"])

        dff = df[['WCNAME', 'Side','Sapcode', 'Current MouldNo', 'Mould life']]
        dff['Mould life'] = 3000 - dff['Mould life']
        dff['Mould life'] = np.where(dff['Mould life']<0, 0, dff['Mould life'])

        dff = dff.merge(wc_master, on=['WCNAME'], how='left')
        dff['WCNAME'] = dff['WCNAME'].str.replace(r'(LH|RH)$', '', regex=True).str.strip()
        dff['curing_machine'] = dff['WCNAME'] + dff['Side']

        Running_Moulds = dff[['curing_machine', 'Current MouldNo', 'Sapcode', 'Mould life']]
        Running_Moulds.columns = ['WCNAME', 'Current MouldNo', 'Sapcode', 'Mould life']
        Running_Moulds['WCNAME'] = Running_Moulds['WCNAME'].str.strip('LH|RH')
        Running_Moulds['No'] = 1

        grouped = (
            Running_Moulds.groupby("WCNAME")
                .agg(
                    SKUCode=("Sapcode", "first"),
                    MouldNos=("Current MouldNo", list),
                    MouldLife_remaining=("Mould life", "min"),
                    Num_Moulds=("No", "count"),
                )
                .reset_index()
        )
        grouped.columns = ["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]
        grouped.to_excel("load_running_moulds.xlsx", index=False)
        return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]

    def load_mould_master(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT * FROM {Config.DB_NAME}.Master_Mapping_Mould_SKU "
            "WHERE `Active Flag`=True"
        )
        df.to_excel("load_mould_master.xlsx", index=False)
        return df

    @staticmethod
    def load_demand_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df = df.rename(columns={"Penetration%":"Priority"})
        return df[df["Quantity"] > 0].copy()

    @staticmethod
    def load_cycle_times_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df = df.rename(columns={"Sapcode":"SKUCode","Cure Time":"Raw"})
        df["CycleTime_min"] = np.round(
            (df["Raw"] + Config.LOAD_UNLOAD_BUFFER_MIN) / Config.PRESS_EFFICIENCY
        )
        return df[["SKUCode","CycleTime_min"]].drop_duplicates("SKUCode")

    @staticmethod
    def load_machine_allowable_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df["Machines"] = df["Machines"].apply(
            lambda x: ast.literal_eval(x) if isinstance(x, str) else []
        )
        return df[["SKUCode","Machines"]]

    @staticmethod
    def load_gt_inventory_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        return df.rename(columns={"Sapcode":"SKUCode","ctb_qty":"GT_Inventory"})[
            ["SKUCode","GT_Inventory"]
        ]

    @staticmethod
    def load_mould_master_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        if "Active Flag" in df.columns:
            df = df[df["Active Flag"].astype(bool)]
        return df

    @staticmethod
    def load_running_moulds_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        if "MouldLife_remaining" not in df.columns:
            df["MouldLife_remaining"] = Config.NEW_MOULD_LIFE
        wcname_col = next((c for c in ["WCNAME", "Machine"] if c in df.columns), None)
        mould_col  = next((c for c in ["Current MouldNo", "MouldNo", "MouldNos"] if c in df.columns), None)
        sku_col    = next((c for c in ["Sapcode", "SKUCode"] if c in df.columns), None)
        if wcname_col is None or mould_col is None or sku_col is None:
            raise ValueError(
                f"Running moulds Excel missing required columns. "
                f"Found: {list(df.columns)}"
            )
        df["Machine"] = (df[wcname_col].astype(str)
                         .str.replace(r"(LH|RH)$", "", regex=True).str.strip())
        df = df.rename(columns={sku_col: "SKUCode", mould_col: "MouldNo"})
        if df["MouldNo"].dtype == object and isinstance(df["MouldNo"].iloc[0], list):
            return df[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining"]]
        grouped = (
            df.groupby("Machine")
              .agg(
                  SKUCode=("SKUCode", "first"),
                  MouldNos=("MouldNo", list),
                  MouldLife_remaining=("MouldLife_remaining", "min"),
                  Num_Moulds=("MouldNo", "count"),
              )
              .reset_index()
        )
        single = (grouped["Num_Moulds"] == 1).sum()
        double = (grouped["Num_Moulds"] == 2).sum()
        print(f"  [ETL] Running moulds (Excel): {len(grouped)} machines | "
              f"2-mould: {double} | 1-mould: {single}")
        return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]


# ══════════════════════════════════════════════════════════════════════════════
# LP SOLVER  (unchanged from v3 — uses MouldTracker which is now permissive)
# ══════════════════════════════════════════════════════════════════════════════
class LP_Solver:
    def __init__(self):
        self.avail_mins = Config.avail_mins()
        self.co_mins    = Config.CHANGEOVER_DURATION_MIN
        self.penalty    = Config.CHANGEOVER_PENALTY_WEIGHT

    def solve(
        self,
        df_valid: pd.DataFrame,
        all_machines: list,
        mould_tracker: MouldTracker,
        locked_machine_mins: dict | None = None,
        continuity_last_sku: dict | None = None,
    ) -> tuple[np.ndarray, dict]:
        locked = locked_machine_mins or {}
        continuity_last_sku = continuity_last_sku or {}

        cont_machs_by_sku: dict[str, set] = defaultdict(set)
        for mach_str, sku in continuity_last_sku.items():
            cont_machs_by_sku[sku].add(mach_str)
            try:
                cont_machs_by_sku[sku].add(int(mach_str))
            except (ValueError, TypeError):
                pass

        S = len(df_valid)
        M = len(all_machines)
        midx = {m: i for i, m in enumerate(all_machines)}
        sku_rows = list(df_valid.itertuples(index=False))

        def xidx(s, mi): return s * M + mi

        eff_cap = {}
        for m in all_machines:
            cap = self.avail_mins - locked.get(str(m), 0.0)
            eff_cap[m] = max(cap, 0.0)

        n_vars = S * M + S
        c = np.zeros(n_vars)
        for s in range(S):
            c[S * M + s] = 1.0
        for s, row in enumerate(sku_rows):
            if row.Demand_Mins > 0:
                for mi in range(M):
                    c[xidx(s, mi)] += self.penalty / row.Demand_Mins

        bounds = [(0.0, None)] * n_vars
        for si, row in enumerate(sku_rows):
            sku = row.SKUCode
            eligible = set(row.Machines)
            mould_eligible = set(
                mould_tracker.get_eligible_machines_with_moulds(
                    sku, list(eligible), cont_machs_by_sku.get(sku, set())
                )
            )
            for mi, mach in enumerate(all_machines):
                if mach not in mould_eligible:
                    bounds[xidx(si, mi)] = (0.0, 0.0)
                else:
                    bounds[xidx(si, mi)] = (0.0, float(row.Demand_Mins))

        A_cap = np.zeros((M, n_vars))
        b_cap = np.array([eff_cap[m] for m in all_machines])
        for mi in range(M):
            for si in range(S):
                A_cap[mi, xidx(si, mi)] = 1.0

        A_dem = np.zeros((S, n_vars))
        b_dem = np.zeros(S)
        for si, row in enumerate(sku_rows):
            for mi in range(M):
                A_dem[si, xidx(si, mi)] = -1.0
            A_dem[si, S * M + si] = -1.0
            b_dem[si] = -float(row.Demand_Mins)

        A_ub = np.vstack([A_cap, A_dem])
        b_ub = np.concatenate([b_cap, b_dem])

        print(f"  [LP] {n_vars:,} vars | {len(b_ub):,} constraints | "
              f"Penalty weight: {self.penalty}")
        print(f"  [LP] Eff capacity range: "
              f"{min(eff_cap.values()):,.0f}–{max(eff_cap.values()):,.0f} min/press")

        # Guard: when continuity blocks have already covered all remaining demand,
        # n_vars==0 → c is the empty 1-D array np.zeros(0). scipy.linprog (HiGHS)
        # refuses empty problems, so short-circuit with a trivially-empty solution.
        # Capacity constraints A_cap·x <= b_cap are vacuously satisfied (x is empty);
        # there is nothing left for the LP to allocate. Rounder/ScheduleBuilder handle
        # an empty allocation correctly (continuity rows are added downstream regardless).
        if n_vars == 0:
            print("  [LP] SKIPPED — 0 vars (continuity blocks cover all remaining demand)")
            result_x = np.zeros(0)
        else:
            result = linprog(c, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
            if result.status != 0:
                raise RuntimeError(f"LP did not converge: {result.message}")
            result_x = result.x
            unmet = sum(result.x[S * M + s] for s in range(S))
            print(f"  [LP] OPTIMAL | Unmet demand-mins: {unmet:,.0f} ({unmet/60:.1f} hrs)")

        meta = {"S":S,"M":M,"midx":midx,"all_machines":all_machines,
                "sku_rows":sku_rows,"xidx":xidx}
        return result_x, meta


# ══════════════════════════════════════════════════════════════════════════════
# ROUNDER  (v4 — accurate per-SKU CO accounting, less wasted capacity)
# ══════════════════════════════════════════════════════════════════════════════
class Rounder:
    """
    v4 change: walks each press's SKU list in priority order and charges
    CO ONLY for SKUs actually kept. Previously the rounder reserved
    (n_skus - 1) * 300 min up front, then trimmed SKUs if needed — but
    the reserved CO time for a dropped SKU was never recovered. Now if
    a SKU is dropped during trimming, no CO is charged for it, leaving
    that capacity available for production or the topup pass.
    """

    def __init__(self):
        self.avail_mins = Config.avail_mins()
        self.co_mins    = Config.CHANGEOVER_DURATION_MIN

    def _row(self, mach, a_or_row, cycles, ct, priority):
        actual_min = cycles * ct
        return {
            "Machine":       mach,
            "SKUCode":       a_or_row,
            "Priority":      round(priority, 4),
            "CycleTime_min": ct,
            "Cycles":        cycles,
            "Units_Planned": cycles * Config.CAVITIES_PER_MOULD,
            "Mins_Used":     round(actual_min, 1),
            "Days_Used":     round(actual_min / (Config.SHIFTS_PER_DAY
                                  * Config.HOURS_PER_SHIFT * 60), 2),
        }

    def round(
        self,
        x: np.ndarray,
        meta: dict,
        df_valid: pd.DataFrame,
        locked_machine_mins: dict,
        continuity_last_sku: dict | None = None,
    ) -> tuple[pd.DataFrame, dict]:
        continuity_last_sku = continuity_last_sku or {}

        S        = meta["S"]
        M        = meta["M"]
        machines = meta["all_machines"]
        xidx     = meta["xidx"]
        sku_rows = meta["sku_rows"]

        machine_cap = {
            m: self.avail_mins - locked_machine_mins.get(str(m), 0.0)
            for m in machines
        }

        # First pass: floor LP allocations to integer cycles
        raw = []
        for si, row in enumerate(sku_rows):
            ct = row.CycleTime_min
            for mi, mach in enumerate(machines):
                mins_lp = x[xidx(si, mi)]
                if mins_lp < ct:
                    continue
                cycles = int(mins_lp / ct)
                raw.append({"si": si, "mi": mi, "mach": mach,
                            "sku": row.SKUCode, "ct": ct,
                            "cycles": cycles, "priority": row.Priority})

        by_machine: dict = defaultdict(list)
        for a in raw:
            by_machine[a["mach"]].append(a)

        final = []
        machine_sku_order: dict = {}

        # Process each machine with ACCURATE per-kept-SKU CO charging
        for mach, assignments in by_machine.items():
            cont_sku = continuity_last_sku.get(str(mach))

            # Run order: continuity SKU first if also in LP allocation,
            # then by priority (avoids spurious CO when LP keeps same SKU)
            if cont_sku:
                cont_idx = next(
                    (i for i, a in enumerate(assignments) if a["sku"] == cont_sku),
                    None,
                )
                if cont_idx is not None:
                    cont_a = assignments.pop(cont_idx)
                    assignments.sort(key=lambda a: -a["priority"])
                    assignments.insert(0, cont_a)
                else:
                    assignments.sort(key=lambda a: -a["priority"])
            else:
                assignments.sort(key=lambda a: -a["priority"])

            prev_sku   = cont_sku
            used       = 0.0
            actual_co  = 0.0
            mach_skus  = []

            for a in assignments:
                co_for_this = (self.co_mins
                              if (prev_sku is not None and prev_sku != a["sku"])
                              else 0.0)
                budget = machine_cap[mach] - used - actual_co - co_for_this
                if budget < a["ct"]:
                    continue
                max_cyc = int(budget / a["ct"])
                cycles  = min(a["cycles"], max_cyc)
                if cycles <= 0:
                    continue
                used      += cycles * a["ct"]
                actual_co += co_for_this
                prev_sku   = a["sku"]
                mach_skus.append(a["sku"])
                final.append(self._row(mach, a["sku"], cycles, a["ct"], a["priority"]))

            machine_sku_order[mach] = mach_skus
            machine_cap[mach] -= used + actual_co

        # ── Top-up pass (greedy, priority-sorted) ─────────────────────────────
        planned = defaultdict(int)
        for a in final:
            planned[a["SKUCode"]] += a["Units_Planned"]

        sorted_rows = sorted(sku_rows, key=lambda r: -r.Priority)
        for row in sorted_rows:
            sku    = row.SKUCode
            ct     = row.CycleTime_min
            needed = int(row.Demand) - planned[sku]
            if needed <= 0:
                continue
            sku_compatible = set(row.Machines)
            eligible = sorted(
                [m for m in machines
                 if machine_cap[m] >= ct and m in sku_compatible],
                key=lambda m: -machine_cap[m],
            )
            for mach in eligible:
                if needed <= 0:
                    break
                cap = machine_cap[mach]
                existing = machine_sku_order.get(mach, [])
                cont_sku_m = continuity_last_sku.get(str(mach))
                prev = existing[-1] if existing else cont_sku_m
                co = self.co_mins if (prev is not None and prev != sku) else 0.0
                available = cap - co
                if available < ct:
                    continue
                extra_c = min(int(available / ct),
                              math.ceil(needed / Config.CAVITIES_PER_MOULD))
                if extra_c <= 0:
                    continue
                machine_cap[mach] -= extra_c * ct + co
                units = extra_c * Config.CAVITIES_PER_MOULD
                planned[sku] += units
                needed       -= units
                machine_sku_order.setdefault(mach, []).append(sku)
                final.append(self._row(mach, sku, extra_c, ct, row.Priority))

        df_sched = pd.DataFrame(final)

        # CO statistics (count actual transitions)
        total_co_mins = 0.0
        co_count      = 0
        for m, skus in machine_sku_order.items():
            cont = continuity_last_sku.get(str(m))
            prev = cont
            for s in skus:
                if prev is not None and prev != s:
                    co_count += 1
                    total_co_mins += self.co_mins
                prev = s

        residual = sum(machine_cap.values())
        print(f"  [Round] Rows: {len(df_sched)} | "
              f"Units: {df_sched['Units_Planned'].sum() if not df_sched.empty else 0:,.0f} | "
              f"Total CO time: {total_co_mins/60:.1f} hrs | "
              f"Changeovers: {co_count} | "
              f"Residual press-mins: {residual:,.0f}")
        return df_sched, machine_sku_order


# ══════════════════════════════════════════════════════════════════════════════
# SCHEDULE BUILDER  (unchanged from v3)
# ══════════════════════════════════════════════════════════════════════════════
class ScheduleBuilder:
    def __init__(self, plan_start: datetime):
        self.plan_start   = plan_start
        self.plan_end     = plan_start + timedelta(days=Config.PLANNING_DAYS)
        self.max_co_shift = Config.MAX_CHANGEOVERS_PER_SHIFT
        self._co_shift_counter: dict[tuple, int] = defaultdict(int)

    @staticmethod
    def _get_shift(dt: datetime) -> tuple[str, datetime]:
        return _get_shift_fn(dt)

    def _shift_key(self, dt: datetime) -> tuple:
        shift, _ = self._get_shift(dt)
        return (dt.date(), shift)

    def _next_co_slot(self, earliest: datetime) -> datetime:
        dt = earliest
        for _ in range(Config.PLANNING_DAYS * Config.SHIFTS_PER_DAY + 1):
            key = self._shift_key(dt)
            if self._co_shift_counter[key] < self.max_co_shift:
                self._co_shift_counter[key] += 1
                return dt
            _, shift_end = self._get_shift(dt)
            dt = shift_end
        return dt

    def _make_row(self, start: datetime, end: datetime, machine,
                  sku: str, qty: int, ct: float, remarks: str, gt_inv: int = 0):
        shift, _ = self._get_shift(start)
        return {
            "Date":          start.date(),
            "Shift":         shift,
            "Machine":       machine,
            "SKUCode":       sku,
            "StartTime":     start,
            "EndTime":       end,
            "Qty":           qty,
            "CycleTime_min": round(ct, 2),
            "GT_Inventory":  gt_inv,
            "Remarks":       remarks,
        }

    def _split_block(
        self, start: datetime, end: datetime,
        machine, sku: str, ct: float,
        gt_inv: int, remarks: str, units_so_far: int,
    ) -> tuple[list[dict], int]:
        rows = []
        cleaning_cycle = Config.units_per_cleaning_cycle()
        total_mins  = (end - start).total_seconds() / 60
        total_units = int(total_mins / ct) * Config.CAVITIES_PER_MOULD
        produced    = 0
        curr        = start

        while curr < end and produced < total_units:
            units_to_clean = cleaning_cycle - (units_so_far % cleaning_cycle)
            units_this_run = min(total_units - produced, units_to_clean)
            mins_this_run  = math.ceil(units_this_run / Config.CAVITIES_PER_MOULD) * ct
            run_end        = min(curr + timedelta(minutes=mins_this_run), end)

            inner = curr
            run_produced = 0
            run_total = int((run_end - curr).total_seconds() / 60 / ct) * Config.CAVITIES_PER_MOULD

            while inner < run_end:
                _, shift_end = self._get_shift(inner)
                slice_end    = min(shift_end, run_end)
                dur          = (slice_end - inner).total_seconds() / 60
                if dur <= 0:
                    inner = slice_end
                    continue
                if slice_end == run_end:
                    qty = run_total - run_produced
                else:
                    qty = int(dur / ct) * Config.CAVITIES_PER_MOULD
                rows.append(self._make_row(inner, slice_end, machine, sku,
                                           qty, ct, remarks, gt_inv))
                run_produced += qty
                inner = slice_end

            produced     += run_produced
            units_so_far += run_produced
            curr          = run_end

            if (units_so_far % cleaning_cycle == 0
                    and produced < total_units and curr < end):
                clean_end = curr + timedelta(minutes=Config.CLEANING_DURATION_MIN)
                rows.append(self._make_row(curr, clean_end, machine,
                                           "MOULD_CLEAN", 0, 0.0,
                                           f"Mould cleaning after {units_so_far} units"))
                curr = clean_end

        return rows, units_so_far

    def build(
        self,
        df_sched: pd.DataFrame,
        machine_sku_order: dict,
        df_gt: pd.DataFrame,
        continuity_rows: list[dict],
        continuity_last_sku: dict | None = None,
    ) -> pd.DataFrame:
        continuity_last_sku = continuity_last_sku or {}
        gt_map = dict(zip(df_gt["SKUCode"], df_gt["GT_Inventory"]))

        machine_free: dict = {}
        for r in continuity_rows:
            m = r["Machine"]
            if m not in machine_free or r["EndTime"] > machine_free[m]:
                machine_free[m] = r["EndTime"]

        alloc: dict = {}
        for _, row in df_sched.iterrows():
            key = (row["Machine"], row["SKUCode"])
            # If topup added a second row for the same (machine, sku), accumulate
            if key in alloc:
                alloc[key]["Cycles"]        += row["Cycles"]
                alloc[key]["Units_Planned"] += row["Units_Planned"]
                alloc[key]["Mins_Used"]     += row["Mins_Used"]
            else:
                alloc[key] = row.to_dict()

        all_rows = []

        if continuity_rows:
            con_dataframe = con_split_into_shifts(pd.DataFrame(continuity_rows))
            con_dataframe['Date'] = con_dataframe['StartTime'].dt.date
            con_dataframe['Date'] = pd.to_datetime(con_dataframe['Date'])
            con_dataframe['Date'] = np.where(
                con_dataframe['StartTime'].dt.hour.isin([0,1,2,3,4,5,6]),
                con_dataframe['Date'] - pd.Timedelta(days=1),
                con_dataframe['Date'],
            )
        else:
            con_dataframe = pd.DataFrame()

        # Process each machine in run order; deduplicate consecutive same-SKU
        # entries in machine_sku_order (topup may append duplicates) so we
        # don't insert spurious COs between same-SKU runs.
        for mach, sku_order in machine_sku_order.items():
            dedup_order = []
            for s in sku_order:
                if not dedup_order or dedup_order[-1] != s:
                    dedup_order.append(s)

            cursor       = machine_free.get(str(mach), self.plan_start)
            cont_sku     = continuity_last_sku.get(str(mach))
            units_so_far = 0

            for idx, sku in enumerate(dedup_order):
                key = (mach, sku)
                if key not in alloc:
                    continue
                a  = alloc[key]
                ct = a["CycleTime_min"]
                gt = int(gt_map.get(sku, 0))

                prev_sku = dedup_order[idx - 1] if idx > 0 else cont_sku
                if prev_sku is not None and prev_sku != sku:
                    co_start = self._next_co_slot(cursor)
                    co_end   = co_start + timedelta(minutes=Config.CHANGEOVER_DURATION_MIN)
                    all_rows.append(self._make_row(
                        co_start, co_end, mach, "CHANGEOVER", 0, 0.0,
                        f"C/O from {prev_sku} to {sku}"
                    ))
                    cursor = co_end

                total_mins = a["Mins_Used"]
                block_end  = min(
                    cursor + timedelta(minutes=total_mins),
                    self.plan_end,
                )
                if block_end <= cursor:
                    continue

                prod_rows, units_so_far = self._split_block(
                    cursor, block_end, mach, sku, ct,
                    gt, "LP Scheduled", units_so_far,
                )
                all_rows.extend(prod_rows)
                cursor = block_end
            machine_free[mach] = cursor

        df_out = pd.DataFrame(all_rows)
        if not con_dataframe.empty:
            df_out = pd.concat([df_out, con_dataframe], axis=0)
        if not df_out.empty:
            df_out = df_out.sort_values(["Machine","StartTime"]).reset_index(drop=True)

        co_count = (df_out["SKUCode"] == "CHANGEOVER").sum() if not df_out.empty else 0
        cl_count = (df_out["SKUCode"] == "MOULD_CLEAN").sum() if not df_out.empty else 0
        prod_qty = (df_out.loc[~df_out["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"]),"Qty"].sum()
                    if not df_out.empty else 0)
        print(f"  [Build] Total rows: {len(df_out)} | "
              f"Prod qty: {prod_qty:,.0f} | "
              f"Changeovers: {co_count} | Cleanings: {cl_count}")
        return df_out


# ══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATOR (v4)
# ══════════════════════════════════════════════════════════════════════════════
class JK_LP_Curing_Scheduler_v2:   # name kept for backwards compatibility
    def __init__(self):
        self.avail_mins = Config.avail_mins()

    def _prepare_skus(self, df_demand, df_cycles, df_allow, df_gt,
                      mould_tracker: MouldTracker, df_running):
        cycle_map = dict(zip(df_cycles["SKUCode"], df_cycles["CycleTime_min"]))
        mach_map  = dict(zip(df_allow["SKUCode"],  df_allow["Machines"]))
        gt_map    = dict(zip(df_gt["SKUCode"],     df_gt["GT_Inventory"]))

        running_skus = (set(df_running['SKUCode'].astype(str).unique())
                        if df_running is not None and not df_running.empty
                        else set())

        rows = []
        for _, r in df_demand.iterrows():
            sku      = r["SKUCode"]
            qty      = int(r["Quantity"])
            priority = r["Priority"]
            ct       = cycle_map.get(sku, Config.DEFAULT_CYCLE_TIME_MIN)
            machines = mach_map.get(sku, [])
            gt       = int(gt_map.get(sku, 0))
            dm       = math.ceil(qty / Config.CAVITIES_PER_MOULD) * ct if ct else 0

            if str(sku) in running_skus:
                has_mould = True
            elif Config.PERMISSIVE_MOULD_ELIGIBILITY:
                has_mould = (mould_tracker.can_schedule(sku)
                             if (ct and machines) else False)
            else:
                has_mould = (mould_tracker.can_assign(sku)
                             if (ct and machines) else False)

            schedulable = bool(ct and machines and has_mould)
            skip = ("" if schedulable
                    else ("No cycle time" if not ct
                          else ("No machine mapping" if not machines
                                else "No compatible mould available")))
            rows.append({
                "SKUCode":        sku,
                "Demand":         qty,
                "Priority":       priority,
                "GT_Inventory":   gt,
                "CycleTime_min":  ct,
                "Machines":       machines,
                "Num_Machines":   len(machines),
                "Demand_Mins":    dm,
                "Presses_Needed": round(dm / self.avail_mins, 2) if ct else 0,
                "Schedulable":    schedulable,
                "Skip_Reason":    skip,
            })

        df_all   = pd.DataFrame(rows)
        df_valid = df_all[df_all["Schedulable"]].copy().reset_index(drop=True)
        df_valid = df_valid.sort_values(["Priority","Num_Machines"],
                                        ascending=[False,True]).reset_index(drop=True)
        all_machines = sorted({m for ml in df_allow["Machines"] for m in ml})
        print(f"  [Prep] Schedulable: {len(df_valid)}/{len(df_all)} | "
              f"Machines: {len(all_machines)} | "
              f"Demand: {df_valid['Demand'].sum():,.0f} | "
              f"Mould eligibility: {'PERMISSIVE' if Config.PERMISSIVE_MOULD_ELIGIBILITY else 'STRICT'}")
        return df_valid, df_all, all_machines

    def _build_continuity(self, df_running: pd.DataFrame,
                          df_valid: pd.DataFrame,
                          df_gt: pd.DataFrame,
                          plan_start: datetime):
        """v3 returns continuity_last_sku; v4 cleaned-up internals."""
        continuity_last_sku: dict = {}

        if df_running is None or df_running.empty:
            return [], {}, {}, continuity_last_sku

        cycle_map  = dict(zip(df_valid["SKUCode"], df_valid["CycleTime_min"]))
        demand_map = dict(zip(df_valid["SKUCode"], df_valid["Demand"]))
        gt_map     = dict(zip(df_gt["SKUCode"],    df_gt["GT_Inventory"]))
        plan_end     = plan_start + timedelta(days=Config.PLANNING_DAYS)
        horizon_mins = Config.avail_mins()

        continuity_rows  = []
        locked_mins      = {}
        demand_remainder = {}

        sku_groups: dict = defaultdict(list)
        for _, row in df_running.iterrows():
            mach = str(row["Machine"])
            sku  = str(row["SKUCode"])
            ct   = cycle_map.get(sku, Config.DEFAULT_CYCLE_TIME_MIN)
            if not ct:
                locked_mins[mach] = 0.0
                continue
            sku_groups[sku].append({
                "machine":  mach,
                "life":     int(row.get("MouldLife_remaining", Config.NEW_MOULD_LIFE)),
                "cavities": int(row.get("Num_Moulds", Config.MOULDS_PER_PRESS)),
                "ct":       ct,
            })

        for sku, group in sku_groups.items():
            demand = demand_map.get(sku, 0)
            gt     = int(gt_map.get(sku, 0))

            if demand <= 0:
                for m in group:
                    locked_mins[m["machine"]] = 0.0
                continue

            for m in group:
                m["max_units"] = int(horizon_mins / m["ct"]) * m["cavities"]
            group_total_cap = sum(m["max_units"] for m in group)
            can_meet = group_total_cap >= demand

            group_sorted = sorted(group, key=lambda m: -m["max_units"])
            remaining_demand = demand
            for i, m in enumerate(group_sorted):
                if not can_meet:
                    alloc = m["max_units"]
                elif remaining_demand <= 0:
                    alloc = 0
                elif i == len(group_sorted) - 1:
                    alloc = min(remaining_demand, m["max_units"])
                else:
                    share = m["max_units"] / group_total_cap if group_total_cap else 0
                    alloc = min(math.ceil(demand * share), m["max_units"], remaining_demand)
                m["alloc_units"] = max(alloc, 0)
                remaining_demand -= m["alloc_units"]

            if not can_meet:
                unmet = demand - group_total_cap
                demand_remainder[sku] = max(int(unmet), 0)
                if unmet > 0:
                    print(f"  [Cont] SKU {sku}: group cap {group_total_cap:,} < "
                          f"demand {demand:,} -> {unmet:,} units to LP")
            else:
                demand_remainder[sku] = 0

            for m in group:
                mach        = m["machine"]
                ct          = m["ct"]
                life        = m["life"]
                cavities    = m["cavities"]
                alloc_units = m["alloc_units"]

                if alloc_units <= 0:
                    locked_mins[mach] = 0.0
                    continue

                continuity_last_sku[mach] = sku

                alloc_mins         = math.ceil(alloc_units / cavities) * ct
                units_before_clean = life * cavities
                cursor             = plan_start
                remaining_mins     = alloc_mins

                while remaining_mins > 0 and cursor < plan_end:
                    mins_to_clean = (units_before_clean / cavities) * ct
                    block_mins    = min(mins_to_clean, remaining_mins)
                    block_end     = min(cursor + timedelta(minutes=block_mins), plan_end)
                    actual_mins   = (block_end - cursor).total_seconds() / 60
                    qty           = int(actual_mins / ct) * cavities

                    shift, _ = _get_shift_fn(cursor)
                    continuity_rows.append({
                        "Date":          cursor.date(),
                        "Shift":         shift,
                        "Machine":       mach,
                        "SKUCode":       sku,
                        "StartTime":     cursor,
                        "EndTime":       block_end,
                        "Qty":           qty,
                        "CycleTime_min": ct,
                        "GT_Inventory":  gt,
                        "Remarks":       f"Continuity ({cavities}-mould press)",
                    })
                    cursor        = block_end
                    remaining_mins -= block_mins

                    if remaining_mins > 0 and cursor < plan_end:
                        clean_end = min(
                            cursor + timedelta(minutes=Config.CLEANING_DURATION_MIN),
                            plan_end,
                        )
                        shift, _ = _get_shift_fn(cursor)
                        continuity_rows.append({
                            "Date":          cursor.date(),
                            "Shift":         shift,
                            "Machine":       mach,
                            "SKUCode":       "MOULD_CLEAN",
                            "StartTime":     cursor,
                            "EndTime":       clean_end,
                            "Qty":           0,
                            "CycleTime_min": 0.0,
                            "GT_Inventory":  0,
                            "Remarks":       "Mould cleaning (continuity)",
                        })
                        cursor             = clean_end
                        units_before_clean = Config.NEW_MOULD_LIFE * cavities

                locked_mins[mach] = (cursor - plan_start).total_seconds() / 60
            
            machines = [4401, 4402, 4403, 4404, 4405, 4406, 4407, 4408, 4409, 4410, 4411, 4412, 4413, 4414, 4415, 4416, 4417, 4801, 4802, 4901, 4902, 4903, 4904, 4905, 4906, 4907, 4908, 4909, 4910, 4911, 4912, 4913, 4914, 4915, 4916, 4917, 4918, 4919, 4920, 4921, 4922, 4923, 4924, 4925, 8601, 8602, 8603, 8604, 8605, 8606, 8701, 8702, 9001, 9002, 9201, 9202, 9203, 9401, 9402, 9403, 9404, 9405, 9406, 9407, 9408, 9409, 9410, 9411, 9412, 9413, 9414, 9502, 9503, 9504, 9505, 9506, 9507, 9701, 9702, 9703, 9704, 9802, 9803, 9804, 14801, 14802, 14803, 14804, 14805, 14806, 14807, 14808, 14809, 14810, 14811, 14812, 14813, 15201, 15202, 15203, 15204, 15205, 15206, 15207, 15208, 15209, 15210, 15211, 15212, 15213, 24801, 24802, 24803, 24804, 24805, 24806, 24807, 24808, 24809, 24810, 24811, 24812, 24813, 24814, 24815, 24816, 24817, 24818, 24819, 24820, 24821, 24822, 24823, 24824, 24825, 24826, 54801, 54802, 54803, 54804, 75201, 75202, 75203, 75204, 75205, 75206, 75207, 75208, 75209, 75210, 75211, 75212, 75213, 75214, 75215, 75216, 75217, 75218, 75219, 75220, 75221, 75222, 75223, 75224, 75225, 75226, 75227, 75228, 75229, 75230]

            mmm = []
            for mm in machines:
                mmm.append(str(mm))

            dict_keys = set(locked_mins.keys())
            extra_in_list = set(mmm) - set(dict_keys)
            
            for key in extra_in_list:
                locked_mins[str(key)] = 0.0

        total_locked = sum(locked_mins.values())
        print(f"  [Cont] {len(continuity_rows)} rows | "
              f"Machines committed: {len(locked_mins)} | "
              f"Locked mins: {total_locked:,.0f} | "
              f"SKUs with LP remainder: {sum(1 for v in demand_remainder.values() if v > 0)} | "
              f"Continuity SKUs tracked: {len(continuity_last_sku)}")
        return continuity_rows, locked_mins, demand_remainder, continuity_last_sku

    def _build_summary(self, df_all, df_sched, df_shift):
        if not df_shift.empty:
            prod = df_shift[~df_shift["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"])]
            planned = prod.groupby("SKUCode")["Qty"].sum().to_dict()
        else:
            planned = {}

        rows = []
        for _, r in df_all.iterrows():
            sku  = r["SKUCode"]
            d    = r["Demand"]
            plan = int(planned.get(sku, 0))
            gap  = max(d - plan, 0)
            pct  = round(plan / d * 100, 1) if d > 0 else 100.0
            if not r["Schedulable"]:
                status = "UNSCHEDULABLE"
            elif gap <= 0:
                status = "FULLY MET"
            elif plan > 0:
                status = "PARTIAL"
            else:
                status = "UNMET"
            rows.append({"SKUCode": sku, "Priority": r["Priority"],
                         "Demand": d, "GT_Inventory": r["GT_Inventory"],
                         "Planned_Units": plan, "Gap": gap,
                         "Fulfillment_Pct": pct, "Status": status,
                         "CycleTime_min": r["CycleTime_min"],
                         "Eligible_Machines": r["Num_Machines"],
                         "Presses_Needed": r["Presses_Needed"],
                         "Skip_Reason": r["Skip_Reason"]})
        return pd.DataFrame(rows).sort_values("Priority", ascending=False)

    def _build_util(self, df_sched, df_shift, all_machines):
        if not df_shift.empty:
            prod = df_shift[~df_shift["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])].copy()
            try:
                prod['Machine'] = prod['Machine'].astype('int64')
            except (ValueError, TypeError):
                pass
            prod["Elapsed"] = (pd.to_datetime(prod["EndTime"]) - pd.to_datetime(prod["StartTime"])).dt.total_seconds() / 60
            grp = prod.groupby("Machine").agg(
                Used_Mins=("Elapsed",  "sum"),
                Total_Units=("Qty",    "sum"),
                SKUs_Count=("SKUCode", "nunique"),
            ).reset_index()
            if not df_sched.empty:
                cyc = df_sched.groupby("Machine")["Cycles"].sum().reset_index()
                grp = grp.merge(cyc, on="Machine", how="left").fillna(0)
                grp.rename(columns={"Cycles":"Total_Cycles"}, inplace=True)
            else:
                grp["Total_Cycles"] = 0
        else:
            grp = pd.DataFrame(columns=["Machine","Used_Mins","Total_Units","SKUs_Count","Total_Cycles"])
        df_u = pd.DataFrame({"Machine": all_machines}).merge(grp, on="Machine", how="left").fillna(0)
        df_u["Available_Mins"]  = self.avail_mins
        df_u["Idle_Mins"]       = self.avail_mins - df_u["Used_Mins"]
        df_u["Utilization_Pct"] = ((df_u["Used_Mins"] / df_u['Available_Mins']) * 100).round(2)
        return df_u[["Machine","Available_Mins","Used_Mins","Idle_Mins",
                     "Utilization_Pct","SKUs_Count","Total_Cycles","Total_Units"]] \
                  .sort_values("Utilization_Pct", ascending=False)

    def _print_results(self, df_summary, df_util, df_shift):
        td  = df_summary["Demand"].sum()
        tp  = df_summary["Planned_Units"].sum()
        co  = (df_shift["SKUCode"]=="CHANGEOVER").sum() if not df_shift.empty else 0
        cl  = (df_shift["SKUCode"]=="MOULD_CLEAN").sum() if not df_shift.empty else 0
        print(f"\n{'='*64}")
        print(f"  Total demand    : {td:>10,.0f}")
        print(f"  Units planned   : {tp:>10,.0f}  ({tp/td*100:.1f}%)")
        print(f"  Gap             : {td-tp:>10,.0f}")
        print(f"  Avg press util  : {df_util['Utilization_Pct'].mean():.1f}%")
        print(f"  Changeover rows : {co:>10}")
        print(f"  Mould clean rows: {cl:>10}")
        print(f"  Fully met SKUs  : {(df_summary['Status']=='FULLY MET').sum():>10}")
        print(f"  Partial SKUs    : {(df_summary['Status']=='PARTIAL').sum():>10}")
        print(f"  Unmet SKUs      : {(df_summary['Status']=='UNMET').sum():>10}")
        print(f"  Unschedulable   : {(df_summary['Status']=='UNSCHEDULABLE').sum():>10}")
        print(f"{'='*64}")

    def run(
        self,
        df_demand:    pd.DataFrame,
        df_cycles:    pd.DataFrame,
        df_allow:     pd.DataFrame,
        df_gt:        pd.DataFrame,
        mould_tracker: MouldTracker,
        df_running:   pd.DataFrame = None,
        plan_start:   datetime     = None,
    ) -> dict:

        if plan_start is None:
            plan_start = datetime.now().replace(
                hour=Config.SHIFT_START_HOUR, minute=0, second=0, microsecond=0
            )

        print("\n" + "="*64)
        print(f"  JK Tyre PCR Curing LP Scheduler v4")
        print(f"  Plan start  : {plan_start:%Y-%m-%d %H:%M}")
        print(f"  Horizon     : {Config.PLANNING_DAYS} days ({self.avail_mins:,.0f} min/press)")
        print(f"  Changeover  : {Config.CHANGEOVER_DURATION_MIN} min | "
              f"Max/shift: {Config.MAX_CHANGEOVERS_PER_SHIFT}")
        print(f"  Mould clean : {Config.CLEANING_DURATION_MIN} min every "
              f"{Config.units_per_cleaning_cycle():,} units")
        print(f"  Mould policy: {'PERMISSIVE (free+locked)' if Config.PERMISSIVE_MOULD_ELIGIBILITY else 'STRICT (free only)'}")
        print("="*64)

        print("\n[Phase 1] Preparing SKU table...")
        df_valid, df_all, all_machines = self._prepare_skus(
            df_demand, df_cycles, df_allow, df_gt, mould_tracker, df_running
        )

        print("\n[Phase 2] Building continuity blocks...")
        continuity_rows, locked_mins, demand_remainder, continuity_last_sku = \
            self._build_continuity(df_running, df_valid, df_gt, plan_start)

        df_lp = df_valid.copy()
        for sku, remainder in demand_remainder.items():
            mask = df_lp["SKUCode"] == sku
            if remainder == 0:
                df_lp = df_lp[~mask]
            else:
                df_lp.loc[mask, "Demand"] = remainder
                df_lp.loc[mask, "Demand_Mins"] = (
                    math.ceil(remainder / Config.CAVITIES_PER_MOULD)
                    * df_lp.loc[mask, "CycleTime_min"].values[0]
                )
        print(f"  [LP Input] {len(df_lp)} SKUs | "
              f"Remaining demand: {df_lp['Demand'].sum():,.0f} units")

        print("\n[Phase 3] Solving LP...")
        solver  = LP_Solver()
        x, meta = solver.solve(df_lp, all_machines, mould_tracker,
                               locked_mins, continuity_last_sku)

        print("\n[Phase 4] Rounding to integer cycles...")
        rounder = Rounder()
        df_mach, machine_sku_order = rounder.round(
            x, meta, df_lp, locked_mins, continuity_last_sku
        )

        print("\n[Phase 5] Building shift-wise schedule...")
        builder  = ScheduleBuilder(plan_start)
        df_shift = builder.build(df_mach, machine_sku_order, df_gt,
                                 continuity_rows, continuity_last_sku)
        df_shift.to_excel("df_shiftv1.xlsx", index=False)

        df_summary = self._build_summary(df_all, df_mach, df_shift)
        df_util    = self._build_util(df_mach, df_shift, all_machines)

        self._print_results(df_summary, df_util, df_shift)

        return {
            "machine_schedule":    df_mach,
            "shift_schedule":      df_shift,
            "demand_fulfillment":  df_summary,
            "machine_utilization": df_util,
            "mould_tracker":       mould_tracker.summary,
        }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER  (unchanged)
# ══════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    _C = {"navy":"1F3864","blue":"2E75B6","teal":"1F6B75",
          "green":"C6EFCE","amber":"FFEB9C","red":"FFC7CE",
          "grey":"F2F2F2","white":"FFFFFF","lgrey":"E8E8E8","orange":"F4B942"}

    def __init__(self, path: str): self.path = path

    def F(self, c): return PatternFill("solid", fgColor=self._C.get(c, c))
    def _b(self):
        s=Side(style="thin",color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    def _hf(self): return Font(bold=True,name="Arial",size=10,color="FFFFFF")
    def _bf(self, bold=False): return Font(bold=bold,name="Arial",size=9)

    def _cell(self, ws, r, c, v, fmt=None, fc="white", bold=False, aln="center"):
        cell=ws.cell(r,c,v); cell.font=self._bf(bold); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal=aln,vertical="center",wrap_text=True)
        if fmt: cell.number_format=fmt

    def _hdr(self, ws, r, c, v, fc="navy"):
        cell=ws.cell(r,c,v); cell.font=self._hf(); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    def _title(self, ws, text, sub, n):
        ws.insert_rows(1); ws.insert_rows(1)
        cl=get_column_letter(n)
        ws.merge_cells(f"A1:{cl}1"); ws["A1"]=text
        ws["A1"].font=Font(bold=True,name="Arial",size=13,color="FFFFFF")
        ws["A1"].fill=self.F("navy"); ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=26
        ws.merge_cells(f"A2:{cl}2"); ws["A2"]=sub
        ws["A2"].font=Font(italic=True,name="Arial",size=9,color="FFFFFF")
        ws["A2"].fill=self.F("teal"); ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[2].height=16

    def _hdr_row(self, ws, row, n):
        for c in range(1,n+1):
            cell=ws.cell(row,c)
            cell.font=self._hf(); cell.fill=self.F("navy"); cell.border=self._b()
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[row].height=30

    STATUS_FC = {"FULLY MET":"green","PARTIAL":"amber",
                 "UNMET":"red","UNSCHEDULABLE":"lgrey"}

    def export(self, results: dict):
        df_mach  = results["machine_schedule"]
        df_shift = results["shift_schedule"]
        df_sum   = results["demand_fulfillment"]
        df_util  = results["machine_utilization"]
        df_mould = results["mould_tracker"]

        td   = int(df_sum["Demand"].sum())
        tp   = int(df_sum["Planned_Units"].sum())
        tg   = int(df_sum["Gap"].sum())
        pct  = round(tp/td*100,1) if td else 0
        avg  = round(df_util["Utilization_Pct"].mean(),1)
        co_n = (df_shift["SKUCode"]=="CHANGEOVER").sum() if not df_shift.empty else 0
        cl_n = (df_shift["SKUCode"]=="MOULD_CLEAN").sum() if not df_shift.empty else 0
        kpi  = (f"Demand: {td:,}  |  Planned: {tp:,}  |  Gap: {tg:,}  |  "
                f"Fulfillment: {pct}%  |  Avg Util: {avg}%  |  "
                f"Changeovers: {co_n}  |  Mould Cleans: {cl_n}")

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            cols1=["SKUCode","Priority","Demand","GT_Inventory","Planned_Units",
                   "Gap","Fulfillment_Pct","Status","CycleTime_min",
                   "Eligible_Machines","Presses_Needed","Skip_Reason"]
            df_sum[cols1].to_excel(writer,sheet_name="Demand Fulfillment",index=False)
            ws=writer.book["Demand Fulfillment"]
            self._title(ws,"PCR CURING v4 — DEMAND FULFILLMENT",kpi,len(cols1))
            self._hdr_row(ws,3,len(cols1))
            for ci,w in enumerate([26,10,13,12,13,12,10,14,12,16,13,22],1):
                ws.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws.max_row+1):
                st=str(ws.cell(ri,8).value)
                sf=self.STATUS_FC.get(st,"white"); bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols1)+1):
                    self._cell(ws,ri,ci,ws.cell(ri,ci).value,
                               fc=sf if ci in(7,8) else bf,
                               bold=(ci==5),aln="left" if ci==1 else "center")
                fp=ws.cell(ri,7)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"
            tr=ws.max_row+1
            for ci in range(1,len(cols1)+1):
                c=ws.cell(tr,ci); c.fill=self.F("navy"); c.font=self._hf()
                c.border=self._b(); c.alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(tr,1).value="TOTAL"
            for ci,v,fmt in [(3,td,"#,##0"),(5,tp,"#,##0"),(6,tg,"#,##0"),(7,pct/100,"0.0%")]:
                ws.cell(tr,ci).value=v; ws.cell(tr,ci).number_format=fmt

            cols2=["Machine","SKUCode","Priority","CycleTime_min",
                   "Cycles","Units_Planned","Mins_Used","Days_Used"]
            df_mach.sort_values(["Machine","SKUCode"]).to_excel(
                writer,sheet_name="Machine Schedule",index=False)
            ws2=writer.book["Machine Schedule"]
            self._title(ws2,"MACHINE-WISE SCHEDULE — PCR v4",kpi,len(cols2))
            self._hdr_row(ws2,3,len(cols2))
            for ci,w in enumerate([12,26,10,14,12,14,16,14],1):
                ws2.column_dimensions[get_column_letter(ci)].width=w
            prev=None
            for ri in range(4,ws2.max_row+1):
                m=ws2.cell(ri,1).value; bf="lgrey" if m!=prev else ("grey" if ri%2==0 else "white")
                prev=m
                for ci in range(1,len(cols2)+1):
                    self._cell(ws2,ri,ci,ws2.cell(ri,ci).value,
                               fc=bf,bold=(ci in(1,6)),aln="left" if ci==2 else "center")

            cols3=["Date","Shift","Machine","SKUCode","StartTime","EndTime",
                   "Qty","CycleTime_min","GT_Inventory","Remarks"]
            df_shift[cols3].to_excel(writer,sheet_name="Shift Schedule",index=False)
            ws3=writer.book["Shift Schedule"]
            self._title(ws3,"SHIFT-WISE SCHEDULE — PCR v4",kpi,len(cols3))
            self._hdr_row(ws3,3,len(cols3))
            for ci,w in enumerate([12,8,12,26,18,18,10,12,12,26],1):
                ws3.column_dimensions[get_column_letter(ci)].width=w
            ROW_FC={"CHANGEOVER":"orange","MOULD_CLEAN":"amber",
                    "A":"E8F4F8","B":"FFF8E8","C":"F0F0F0"}
            for ri in range(4,ws3.max_row+1):
                sku=str(ws3.cell(ri,4).value); shift=str(ws3.cell(ri,2).value)
                fc=ROW_FC.get(sku,ROW_FC.get(shift,"white"))
                for ci in range(1,len(cols3)+1):
                    self._cell(ws3,ri,ci,ws3.cell(ri,ci).value,
                               fc=fc,bold=(sku in("CHANGEOVER","MOULD_CLEAN")),
                               aln="left" if ci==4 else "center")

            cols4=["Machine","Available_Mins","Used_Mins","Idle_Mins",
                   "Utilization_Pct","SKUs_Count","Total_Cycles","Total_Units"]
            df_util.to_excel(writer,sheet_name="Machine Utilization",index=False)
            ws4=writer.book["Machine Utilization"]
            idle_c=int((df_util["Utilization_Pct"]==0).sum())
            high_c=int((df_util["Utilization_Pct"]>=90).sum())
            self._title(ws4,"PRESS UTILIZATION — PCR v4",
                        f"Avg: {avg}% | High(≥90%): {high_c} | Idle: {idle_c} | Total: {len(df_util)}",
                        len(cols4))
            self._hdr_row(ws4,3,len(cols4))
            for ci,w in enumerate([12,15,14,14,14,12,14,14],1):
                ws4.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws4.max_row+1):
                u=ws4.cell(ri,5).value or 0
                uf="green" if u>=90 else ("amber" if u>=60 else "red")
                bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols4)+1):
                    self._cell(ws4,ri,ci,ws4.cell(ri,ci).value,
                               fc=uf if ci==5 else bf,bold=(ci in(1,5)))
                fp=ws4.cell(ri,5)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"

            df_mould.to_excel(writer,sheet_name="Mould Tracker",index=False)
            ws5=writer.book["Mould Tracker"]
            self._title(ws5,"MOULD AVAILABILITY TRACKER",
                        f"Total moulds: {len(df_mould)} | "
                        f"Free: {(df_mould['Assigned_Machine']=='FREE').sum()} | "
                        f"Assigned: {(df_mould['Assigned_Machine']!='FREE').sum()}",
                        len(df_mould.columns))
            self._hdr_row(ws5,3,len(df_mould.columns))
            for ci,w in enumerate([22,30,14,16],1):
                ws5.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws5.max_row+1):
                assigned=str(ws5.cell(ri,4).value)
                bf="C6EFCE" if assigned=="FREE" else ("FFEB9C" if assigned!="FREE" else "white")
                for ci in range(1,len(df_mould.columns)+1):
                    self._cell(ws5,ri,ci,ws5.cell(ri,ci).value,fc=bf)

        print(f"\n  [Export] Saved → {self.path}")


def con_split_into_shifts(df):
    rows = []
    for _, r in df.iterrows():
        start = r['StartTime']
        end   = r['EndTime']
        total_qty = r['Qty']
        total_minutes = (end - start).total_seconds() / 60
        current = start
        while current < end:
            shift, shift_end = _get_shift_fn(current)
            slice_end = min(shift_end, end)
            slice_minutes = (slice_end - current).total_seconds() / 60
            qty = (slice_minutes / total_minutes) * total_qty if total_minutes > 0 else 0
            new_row = r.copy()
            new_row['StartTime'] = current
            new_row['EndTime']   = slice_end
            new_row['Shift']     = shift
            new_row['Qty']       = round(qty)
            rows.append(new_row)
            current = slice_end
    return pd.DataFrame(rows)


def _get_shift_fn(dt: datetime) -> tuple[str, datetime]:
    h    = dt.hour
    base = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    sh   = Config.SHIFT_START_HOUR
    if sh <= h < sh + 8:
        return "A", base + timedelta(hours=sh + 8)
    elif sh + 8 <= h < sh + 16:
        return "B", base + timedelta(hours=sh + 16)
    else:
        if h >= sh + 16:
            return "C", base + timedelta(days=1, hours=sh)
        else:
            return "C", base + timedelta(hours=sh)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINTS
# ══════════════════════════════════════════════════════════════════════════════
def run_from_excel(
    demand_path:  str = "Demand_for_Curing_Schedule3_pcr.xlsx",
    cycles_path:  str = "Master_Curing_Design_CycleTime_pcr.xlsx",
    allow_path:   str = "curing_pcr_machine_allowable.xlsx",
    gt_path:      str = "GT_Inventory_pcr.xlsx",
    mould_path:   str = "Master_Mapping_Mould_SKU.xlsx",
    running_path: str = None,
    plan_start:   datetime = None,
    output_path:  str = "PCR_Curing_LP_v4_Schedule.xlsx",
) -> dict:
    print("\n[Phase 0] ETL from Excel files...")
    df_demand  = ETL.load_demand_from_excel(demand_path)
    df_cycles  = ETL.load_cycle_times_from_excel(cycles_path)
    df_allow   = ETL.load_machine_allowable_from_excel(allow_path)
    df_gt      = ETL.load_gt_inventory_from_excel(gt_path)
    df_running = ETL.load_running_moulds_from_excel(running_path) if running_path else None

    tracker = MouldTracker()
    tracker.load_from_excel(mould_path, running_path)

    scheduler = JK_LP_Curing_Scheduler_v2()
    results   = scheduler.run(df_demand, df_cycles, df_allow, df_gt,
                              tracker, df_running, plan_start)
    ExcelExporter(output_path).export(results)
    return results


# Default cycle time for SKUs missing in master now lives in
# Config.DEFAULT_CYCLE_TIME_MIN (EFFECTIVE minutes, used directly).


def _post_process_schedule_excel(path: str, default_ct_skus: set) -> None:
    """Two corrections applied to the schedule output Excel after the LP runs:

    1. Machine Utilization sheet — Used_Mins, Idle_Mins, Utilization_Pct, etc.
       are recomputed EXCLUDING CHANGEOVER + mould-clean rows. The headline
       summary line in R2 is also rewritten. Matches the productive-util
       definition we use everywhere else.

    2. CycleTime_min column on the Demand Fulfillment sheet shows "NA" for any
       SKU that received the default cycle time (i.e. the SKU had no row in
       Master_Curing_Design_CycleTime). Internally the LP used the default cure
       time of 15 min run through (raw + buffer) / efficiency — same as every
       other SKU — but the user-facing report shows "NA" to make the
       missing-master-data status explicit.
    """
    import openpyxl
    from collections import defaultdict

    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        # ---------- (1) Machine Utilization — exclude CO / clean ----------
        ws_s = wb["Shift Schedule"]
        used_min:    dict = defaultdict(float)
        skus_set:    dict = defaultdict(set)
        cycles_cnt:  dict = defaultdict(float)   # productive cycles only
        units_cnt:   dict = defaultdict(float)
        for r in range(4, ws_s.max_row + 1):
            machine = ws_s.cell(row=r, column=3).value
            sku     = ws_s.cell(row=r, column=4).value
            s       = ws_s.cell(row=r, column=5).value
            e       = ws_s.cell(row=r, column=6).value
            qty     = ws_s.cell(row=r, column=7).value or 0
            ct      = ws_s.cell(row=r, column=8).value
            remarks = (ws_s.cell(row=r, column=10).value or "")
            if machine is None or s is None or e is None or e <= s:
                continue
            sku_u, rem_u = (str(sku).upper() if sku else ""), str(remarks).upper()
            if sku_u == "CHANGEOVER" or "CLEAN" in rem_u or "CHANGEOVER" in rem_u:
                continue                            # productive only
            # IMPORTANT: Shift Schedule uses both int and str machine IDs in the
            # same column; Machine Utilization uses all int. Coerce to str on
            # both sides to make the lookup match (same trick as capacity_writer).
            m_key = str(machine)
            mins = (e - s).total_seconds() / 60
            used_min[m_key]   += mins
            skus_set[m_key].add(sku)
            units_cnt[m_key]  += float(qty or 0)
            try:
                if ct: cycles_cnt[m_key] += float(qty) / 2.0   # 2 tyres/cycle
            except (TypeError, ValueError):
                pass

        ws_u = wb["Machine Utilization"]
        avail_per_mach = None
        utils = []
        high  = 0; idle  = 0; total = 0
        for r in range(4, ws_u.max_row + 1):
            m = ws_u.cell(row=r, column=1).value
            if m is None: continue
            avail = ws_u.cell(row=r, column=2).value or 0
            avail_per_mach = avail
            m_key = str(m)                                    # match Shift-Schedule keys
            u = used_min.get(m_key, 0.0)
            i = max(0.0, avail - u)
            util = (u / avail) if avail else 0.0
            ws_u.cell(row=r, column=3).value = round(u, 0)
            ws_u.cell(row=r, column=4).value = round(i, 0)
            ws_u.cell(row=r, column=5).value = round(util, 3)
            ws_u.cell(row=r, column=6).value = len(skus_set.get(m_key, set()))
            ws_u.cell(row=r, column=7).value = round(cycles_cnt.get(m_key, 0), 0)
            ws_u.cell(row=r, column=8).value = round(units_cnt.get(m_key, 0), 0)
            utils.append(util); total += 1
            if util >= 0.90: high += 1
            if util == 0.0:  idle += 1
        avg_util = (sum(utils) / total * 100) if total else 0.0
        ws_u.cell(row=2, column=1).value = (
            f"Avg: {avg_util:.1f}% | High(≥90%): {high} | Idle: {idle} | Total: {total}"
        )

        # ---------- (2) Demand Fulfillment — "NA" for defaulted CT SKUs ----------
        if default_ct_skus:
            ws_df = wb["Demand Fulfillment"]
            for r in range(4, ws_df.max_row + 1):
                sku = ws_df.cell(row=r, column=1).value
                if sku in default_ct_skus:
                    ws_df.cell(row=r, column=9).value = "NA"      # CycleTime_min col

        # ---------- (3) Machine Schedule — rebuild from Shift Schedule ----------
        # Bug: the legacy df_mach only captures the LP rounder's FIRST-PASS
        # allocation. The ScheduleBuilder then adds continuity rows + extra
        # runs that don't propagate back. Result: Σ Units_Planned in this
        # sheet ≠ Σ Qty in Shift Schedule ≠ Σ Planned_Units in Demand
        # Fulfillment. Fix: re-derive (Machine, SKU) totals from Shift
        # Schedule (the authoritative final schedule).
        from collections import defaultdict as _dd
        units_by_pair: dict = _dd(float)      # (machine, sku) → Σ Qty
        mins_by_pair:  dict = _dd(float)      # (machine, sku) → Σ elapsed min
        ct_by_pair:    dict = {}              # first non-null CT seen
        for r in range(4, ws_s.max_row + 1):
            machine = ws_s.cell(row=r, column=3).value
            sku     = ws_s.cell(row=r, column=4).value
            s       = ws_s.cell(row=r, column=5).value
            e       = ws_s.cell(row=r, column=6).value
            qty     = ws_s.cell(row=r, column=7).value or 0
            ct      = ws_s.cell(row=r, column=8).value
            remarks = (ws_s.cell(row=r, column=10).value or "")
            if not (machine and sku and s and e and e > s): continue
            sku_u, rem_u = str(sku).upper(), str(remarks).upper()
            if sku_u == "CHANGEOVER" or sku_u == "MOULD_CLEAN" \
               or "CLEAN" in rem_u or "CHANGEOVER" in rem_u:
                continue                                # productive only
            key = (machine, sku)
            units_by_pair[key] += float(qty)
            mins_by_pair[key]  += (e - s).total_seconds() / 60
            if key not in ct_by_pair and ct: ct_by_pair[key] = ct

        # Per-SKU priority from Demand Fulfillment (col 2 = Priority).
        ws_df = wb["Demand Fulfillment"]
        prio_by_sku: dict = {}
        for r in range(4, ws_df.max_row + 1):
            sku = ws_df.cell(row=r, column=1).value
            if sku: prio_by_sku[sku] = ws_df.cell(row=r, column=2).value

        # Wipe old Machine Schedule rows and rewrite.
        ws_m = wb["Machine Schedule"]
        # Columns: 1=Machine 2=SKUCode 3=Priority 4=CycleTime_min
        #          5=Cycles 6=Units_Planned 7=Mins_Used 8=Days_Used
        if ws_m.max_row > 3:
            ws_m.delete_rows(4, ws_m.max_row - 3)

        # Sort: machine asc, then by priority desc (high-priority SKUs first).
        rows_sorted = sorted(units_by_pair.keys(),
                             key=lambda k: (str(k[0]), -float(prio_by_sku.get(k[1], 0) or 0)))
        SLOT_MIN_PER_DAY = 3 * 8 * 60                  # 3 shifts × 8h × 60
        for i, (machine, sku) in enumerate(rows_sorted, start=4):
            units = units_by_pair[(machine, sku)]
            mins  = mins_by_pair[(machine, sku)]
            ct    = ct_by_pair.get((machine, sku))
            cycles = units / 2.0                       # 2 tyres/cycle
            ws_m.cell(row=i, column=1).value = machine
            ws_m.cell(row=i, column=2).value = sku
            ws_m.cell(row=i, column=3).value = prio_by_sku.get(sku)
            ws_m.cell(row=i, column=4).value = "NA" if sku in default_ct_skus else ct
            ws_m.cell(row=i, column=5).value = round(cycles, 0)
            ws_m.cell(row=i, column=6).value = round(units, 0)
            ws_m.cell(row=i, column=7).value = round(mins, 1)
            ws_m.cell(row=i, column=8).value = round(mins / SLOT_MIN_PER_DAY, 2)

        # Update Machine Schedule R2 summary.
        total_units = sum(units_by_pair.values())
        ws_m.cell(row=2, column=1).value = (
            f"Pairs: {len(units_by_pair)} | Σ Units: {int(total_units):,} | "
            f"Σ Cycles: {int(total_units/2):,}"
        )

        wb.save(path)
    finally:
        wb.close()


def run_from_database_simple(
    demand_path:  str,
    plan_start:   datetime,
    output_path:  str = "PCR_Curing_CBC_FeedAware.xlsx",
    tyre_type:    str = Config.TYRE_TYPE,
    engine=None,
) -> dict:
    """
    CBC curing entry: pull every input from the DB EXCEPT demand (which comes
    from `demand_path`, e.g. Book4.xlsx), with NO jkt_plan_params/plan_id
    dependency. Mirrors run_from_database's ETL block but uses an explicit
    plan_start. Feed-awareness (if installed via feed_aware_curing) is already
    active on MouldTracker by the time scheduler.run() is called.
    """
    if create_engine is None and engine is None:
        raise ImportError("sqlalchemy not installed. Use run_from_excel() instead.")
    if engine is None:
        engine = create_engine(
            f"mysql+pymysql://{Config.DB_USER}:{Config.DB_PASSWORD}"
            f"@{Config.DB_SERVER}/{Config.DB_NAME}",
            pool_pre_ping=True, pool_recycle=280,
        )

    etl = ETL(engine, tyre_type)
    print("\n[Phase 0] ETL from database (demand from file)...")
    df_demand  = etl.load_demand(demand_path)
    df_cycles  = etl.load_cycle_times()
    df_allow   = etl.load_machine_allowable()
    df_gt      = etl.load_gt_inventory()
    df_running = etl.load_running_moulds()
    df_mould_m = etl.load_mould_master()

    # Default cycle time for demand SKUs missing in the cycle-time master.
    # Config.DEFAULT_CYCLE_TIME_MIN is already EFFECTIVE (post buffer+efficiency),
    # so it is used directly.
    have_ct     = set(df_cycles["SKUCode"].astype(str))
    demand_skus = set(df_demand["SKUCode"].astype(str))
    missing_ct  = demand_skus - have_ct
    if missing_ct:
        default_ct = float(Config.DEFAULT_CYCLE_TIME_MIN)
        df_cycles = pd.concat(
            [df_cycles, pd.DataFrame({"SKUCode": sorted(missing_ct),
                                      "CycleTime_min": default_ct})],
            ignore_index=True)
        print(f"[Phase 0] CT default {default_ct:.0f} min (effective) applied "
              f"to {len(missing_ct)} SKU(s).")

    tracker = MouldTracker()
    tracker.load_from_df(df_mould_m, df_running)

    scheduler = JK_LP_Curing_Scheduler_v2()
    results   = scheduler.run(df_demand, df_cycles, df_allow, df_gt,
                              tracker, df_running, plan_start)
    ExcelExporter(output_path).export(results)
    _post_process_schedule_excel(output_path, missing_ct)
    return results


def run_from_database(
    plan_id:      str = None,
    demand_csv:   str = None,
    tyre_type:    str = Config.TYRE_TYPE,
    output_path:  str = None,
) -> dict:
    if create_engine is None:
        raise ImportError("sqlalchemy not installed. Use run_from_excel() instead.")
    engine = create_engine(
        f"mysql+pymysql://{Config.DB_USER}:{Config.DB_PASSWORD}"
        f"@{Config.DB_SERVER}/{Config.DB_NAME}"
    )

    # Fetch plan params and override the date-dependent Config knobs.
    pid = plan_id or Config.PLAN_ID
    plan_params = pd.read_sql(
        f"SELECT planStartDate, planEndDate, efficiency, noOfChangeOver "
        f"FROM {Config.DB_NAME}.{Config.PLAN_PARAMS_TABLE} "
        f"WHERE plan_id = %(pid)s",
        engine, params={"pid": pid},
    )
    if plan_params.empty:
        raise ValueError(f"plan_id={pid!r} not found in {Config.PLAN_PARAMS_TABLE}")
    row0 = plan_params.iloc[0]
    ps_date = pd.to_datetime(row0["planStartDate"]).date()
    pe_date = pd.to_datetime(row0["planEndDate"]).date()
    plan_start = datetime(ps_date.year, ps_date.month, ps_date.day,
                          Config.SHIFT_START_HOUR, 0, 0)
    Config.PLAN_DATE     = plan_start
    Config.PLANNING_DAYS = (pe_date - ps_date).days + 1  # inclusive
    print(f"[Plan] plan_id={pid}  start={plan_start}  "
          f"end={pe_date}  PLANNING_DAYS={Config.PLANNING_DAYS}")

    # ── Per-plan DB overrides (win over Config defaults when set) ──────────────
    # noOfChangeOver: stored AS-IS as max changeovers per SHIFT. NULL/0 → default.
    db_co = row0.get("noOfChangeOver")
    if db_co is not None and not pd.isna(db_co) and int(db_co) > 0:
        Config.MAX_CHANGEOVERS_PER_SHIFT = int(db_co)
        print(f"[Plan] DB override: noOfChangeOver -> "
              f"MAX_CHANGEOVERS_PER_SHIFT={Config.MAX_CHANGEOVERS_PER_SHIFT}/shift")
    # efficiency: stored as a percentage (94 = 94%). Config wants a fraction.
    db_eff = row0.get("efficiency")
    if db_eff is not None and not pd.isna(db_eff) and float(db_eff) > 0:
        Config.PRESS_EFFICIENCY = float(db_eff) / 100.0
        print(f"[Plan] DB override: efficiency={db_eff}% -> "
              f"PRESS_EFFICIENCY={Config.PRESS_EFFICIENCY}")

    # Derive paths from plan_id if not overridden.
    if demand_csv is None:
        demand_csv = f"requirement_summary_{pid}.xlsx"
    if output_path is None:
        output_path = f"PCR_Schedule_{pid}_{ps_date}_{Config.PLANNING_DAYS}days.xlsx"

    etl = ETL(engine, tyre_type)
    print("\n[Phase 0] ETL from database...")
    df_demand  = etl.load_demand(demand_csv)
    df_cycles  = etl.load_cycle_times()
    df_allow   = etl.load_machine_allowable()
    df_gt      = etl.load_gt_inventory()
    df_running = etl.load_running_moulds()
    df_mould_m = etl.load_mould_master()

    # ── Inject default cycle time for demand SKUs missing in master ──
    # Config.DEFAULT_CYCLE_TIME_MIN is an EFFECTIVE cycle time (already includes
    # the load/unload buffer and press efficiency), so it is used directly — not
    # re-run through (raw + buffer) / efficiency. Their CycleTime_min still
    # displays as "NA" in the final Excel via _post_process_schedule_excel to
    # flag the missing-master-data status.
    have_ct      = set(df_cycles["SKUCode"].astype(str))
    demand_skus  = set(df_demand["SKUCode"].astype(str))
    missing_ct   = demand_skus - have_ct
    if missing_ct:
        default_ct = float(Config.DEFAULT_CYCLE_TIME_MIN)
        extra = pd.DataFrame(
            {"SKUCode": sorted(missing_ct), "CycleTime_min": default_ct}
        )
        df_cycles = pd.concat([df_cycles, extra], ignore_index=True)
        print(f"[Phase 0] CT default applied to {len(missing_ct)} SKU(s) "
              f"({default_ct:.0f} min effective — shown as 'NA' in output)")

    tracker = MouldTracker()
    tracker.load_from_df(df_mould_m, df_running)

    scheduler = JK_LP_Curing_Scheduler_v2()
    results   = scheduler.run(df_demand, df_cycles, df_allow, df_gt,
                              tracker, df_running, plan_start)
    ExcelExporter(output_path).export(results)

    # Post-process the Excel: fix Machine Utilization (productive only) +
    # show 'NA' for SKUs that received the default cycle time.
    _post_process_schedule_excel(output_path, missing_ct)
    return results


# ------- Standalone execution --------------------------------------------------
if __name__ == "__main__":
    results = run_from_database(plan_id=Config.PLAN_ID)
