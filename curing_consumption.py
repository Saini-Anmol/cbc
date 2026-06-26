"""
B2C Pipeline — Phase 0: Curing Consumption Table
=================================================
Builds the GT consumption table: for each SKU currently running on curing presses,
computes how many GTs each press consumes per shift, then aggregates to a total
consumption figure per shift per SKU across the full planning horizon.

Also classifies every SKU into one of three categories:
  Runner-In      : in demand AND actively running on curing presses
  Runner-Out     : NOT in demand AND actively running on curing presses
  Non-Runner-In  : in demand AND NOT running on any curing press

Standalone usage:
    python curing_consumption.py

Output: data/output/curing_consumption_table.xlsx
"""

from __future__ import annotations

import math
import os
import sys
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── venv re-exec (same pattern as cbc.py) ────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

import cbc_env

HERE     = os.path.dirname(os.path.abspath(__file__))
IN_DIR   = cbc_env.INPUT_DIR
OUT_DIR  = cbc_env.OUTPUT_DIR

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

class ConsumptionConfig:
    # ── planning horizon ──────────────────────────────────────────────────────
    PLANNING_DAYS      = 30
    SHIFTS_PER_DAY     = 3
    HOURS_PER_SHIFT    = 8
    SHIFT_START_HOUR   = 7          # Shift A starts at 07:00
    SHIFT_MINS         = 480        # minutes per shift

    # ── press physics ─────────────────────────────────────────────────────────
    CAVITIES_PER_MOULD    = 2
    MOULDS_PER_PRESS      = 2
    LOAD_UNLOAD_BUFFER_MIN = 2.3
    PRESS_EFFICIENCY       = 0.94
    # Default EFFECTIVE cycle time for SKUs missing from the CT master.
    # Already includes buffer + efficiency — do NOT re-apply the formula.
    DEFAULT_CYCLE_TIME_MIN = 17.0

    # ── downtime (minutes) — used for press exclusion logic ───────────────────
    CURING_CO_MIN     = 300
    MOULD_CLEAN_MIN   = 120

    # ── changeover cap ────────────────────────────────────────────────────────
    MAX_CHANGEOVERS_PER_DAY = 8

    # ── database ──────────────────────────────────────────────────────────────
    DB_NAME = cbc_env.ENV.get("JKT_DB_DATABASE", "jkplanningV1")


# ══════════════════════════════════════════════════════════════════════════════
# ETL  (adapted from curing_lp.ETL — loads the three DB tables we need)
# ══════════════════════════════════════════════════════════════════════════════

class ConsumptionETL:
    """Load data from DB needed to build the curing consumption table."""

    def __init__(self, engine):
        self.engine = engine
        self.db = ConsumptionConfig.DB_NAME

    def _sql(self, q: str) -> pd.DataFrame:
        return pd.read_sql(q, self.engine)

    # -- adapted from curing_lp.ETL.load_demand (lines 316-330) ---------------
    def load_demand(self, path: str) -> pd.DataFrame:
        """Load demand file. Returns [SKUCode, Quantity, Priority]."""
        if str(path).lower().endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path)
        # Already normalised format (SKUCode, Quantity, Priority)
        if "Quantity" in df.columns and "Priority" in df.columns:
            df = df.groupby("SKUCode").agg(
                Quantity=("Quantity", "sum"),
                Priority=("Priority", "max"),
            ).reset_index()
        else:
            qty_col = "Updated_Requirement" if "Updated_Requirement" in df.columns else "Requirement"
            df = (df.groupby("SKUCode")
                    .agg(Quantity=(qty_col, "sum"),
                         Priority=("ConsolidatedPriorityScore", "max"))
                    .reset_index())
        return df[df["Quantity"] > 0].copy()

    # -- adapted from curing_lp.ETL.load_cycle_times (lines 332-343) ----------
    def load_cycle_times(self) -> pd.DataFrame:
        """Load effective cycle times from DB. Returns [SKUCode, CycleTime_min]."""
        df = self._sql(
            f"SELECT Sapcode AS SKUCode, `Cure Time` AS Raw "
            f"FROM {self.db}.Master_Curing_Design_CycleTime"
        )
        df["CycleTime_min"] = np.round(
            (df["Raw"] + ConsumptionConfig.LOAD_UNLOAD_BUFFER_MIN)
            / ConsumptionConfig.PRESS_EFFICIENCY
        )
        df = df[["SKUCode", "CycleTime_min"]].drop_duplicates("SKUCode")
        df["SKUCode"] = df["SKUCode"].str.strip()
        return df

    # -- adapted from curing_lp.ETL.load_gt_inventory (lines 360-366) ---------
    def load_gt_inventory(self) -> pd.DataFrame:
        """Load opening GT inventory from DB. Returns [SKUCode, GT_Inventory]."""
        return self._sql(
            f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory"
            f" FROM {self.db}.gt_inventory_manual"
        )

    # -- adapted from curing_lp.ETL.load_running_moulds (lines 368-400) -------
    def load_running_moulds(self) -> pd.DataFrame:
        """
        Load currently running moulds per curing press.
        Returns [Machine, SKUCode, MouldNos, MouldLife_remaining, Num_Moulds].
        Excludes presses where SKUCode is blank/NULL (in changeover or idle).
        """
        wc_master = self._sql(f"SELECT * FROM {self.db}.Master_WC_Master")
        wc_master = wc_master[["wcID", "WCNAME"]]

        df = self._sql(f"SELECT * FROM {self.db}.Daily_Running_Moulds")
        if "updatedAt" in df.columns:
            df = df.drop(columns=["updatedAt"])

        dff = df[["WCNAME", "Side", "Sapcode", "Current MouldNo", "Mould life"]].copy()
        dff["Mould life"] = 3000 - dff["Mould life"]
        dff["Mould life"] = np.where(dff["Mould life"] < 0, 0, dff["Mould life"])

        dff = dff.merge(wc_master, on=["WCNAME"], how="left")
        dff["WCNAME"] = dff["WCNAME"].str.replace(r"(LH|RH)$", "", regex=True).str.strip()
        dff["curing_machine"] = dff["WCNAME"] + dff["Side"]

        running = dff[["curing_machine", "Current MouldNo", "Sapcode", "Mould life"]].copy()
        running.columns = ["WCNAME", "Current MouldNo", "Sapcode", "Mould life"]
        running["WCNAME"] = running["WCNAME"].str.strip("LH|RH")
        running["No"] = 1

        # Exclude presses that are in CO or mould clean (sentinel SKU codes)
        _co_sentinels = {"CHANGEOVER", "MOULD_CLEAN", "MOULDCLEAN", "CO", "CLEAN", ""}
        running = running[
            running["Sapcode"].notna()
            & (~running["Sapcode"].str.strip().str.upper().isin(_co_sentinels))
        ]

        grouped = (
            running.groupby("WCNAME")
                .agg(
                    SKUCode=("Sapcode", "first"),
                    MouldNos=("Current MouldNo", list),
                    MouldLife_remaining=("Mould life", "min"),
                    Num_Moulds=("No", "count"),
                )
                .reset_index()
        )
        grouped.columns = ["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]
        return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]

    def load_curing_allowable(self) -> pd.DataFrame:
        """Load SKU → eligible curing presses. Returns [SKUCode, Machines (list of str)]."""
        df = self._sql(
            f"SELECT * FROM {self.db}.Master_Curing_Allowable_Machines_source"
        )
        df = df.rename(columns={"SKU Code": "SKUCode"})
        mcols = [c for c in df.columns if str(c).isdigit()]
        df["Machines"] = df.apply(
            lambda r: [str(c) for c in mcols if str(r[c]).strip().lower() == "yes"], axis=1
        )
        return df[["SKUCode", "Machines"]]

    def load_building_allowable_skus(self) -> set:
        """SKUs that appear in building allowable master with at least one machine."""
        try:
            df = self._sql(
                f"SELECT `SKU Code` AS SKUCode "
                f"FROM {self.db}.Master_Building_Allowable_Machines_source"
            )
            return set(df["SKUCode"].astype(str).str.strip())
        except Exception as exc:
            print(f"  ⚠  Building allowable master unavailable: {exc}")
            return set()

    def load_building_history_skus(self) -> set:
        """SKUs appearing in Stage-1 or Stage-2 building history tables."""
        skus: set = set()
        for tbl in ("Building_Stage1_Best_Machines", "Building_Stage2_Best_Machines"):
            try:
                df = self._sql(
                    f"SELECT DISTINCT sizeCode AS SKUCode FROM {self.db}.{tbl}"
                )
                skus |= set(df["SKUCode"].astype(str).str.strip())
            except Exception as exc:
                print(f"  ⚠  {tbl} unavailable: {exc}")
        return skus

    def load_curing_allowable_skus(self) -> set:
        """SKUs that have at least one curing press in master data."""
        try:
            df = self._sql(
                f"SELECT `SKU Code` AS SKUCode "
                f"FROM {self.db}.Master_Curing_Allowable_Machines_source"
            )
            return set(df["SKUCode"].astype(str).str.strip())
        except Exception as exc:
            print(f"  ⚠  Curing allowable master unavailable: {exc}")
            return set()

    def load_curing_history_skus(self) -> set:
        """SKUs seen in Daily_Running_Moulds (current/historical curing press state)."""
        _sentinels = {"CHANGEOVER", "MOULD_CLEAN", "MOULDCLEAN", "CO", "CLEAN", "NAN", ""}
        try:
            df = self._sql(
                f"SELECT DISTINCT Sapcode AS SKUCode "
                f"FROM {self.db}.Daily_Running_Moulds "
                f"WHERE Sapcode IS NOT NULL AND Sapcode != ''"
            )
            raw = set(df["SKUCode"].astype(str).str.strip())
            return {s for s in raw if s.upper() not in _sentinels}
        except Exception as exc:
            print(f"  ⚠  Curing history (Daily_Running_Moulds) unavailable: {exc}")
            return set()


# ══════════════════════════════════════════════════════════════════════════════
# SKU CLASSIFIER
# ══════════════════════════════════════════════════════════════════════════════

class SKUClassifier:
    """Classify SKUs into Runner-In, Runner-Out, and Non-Runner-In."""

    def classify(
        self,
        df_demand: pd.DataFrame,
        df_running_moulds: pd.DataFrame,
    ) -> pd.DataFrame:
        """
        Returns DataFrame with columns:
          [SKUCode, Category, RunningPressCount, MouldLife_min]
        """
        demand_skus  = set(df_demand["SKUCode"].str.strip())
        # Group running moulds by SKU to count active presses
        press_count = (
            df_running_moulds.groupby("SKUCode")
            .agg(
                RunningPressCount=("Machine", "count"),
                MouldLife_min=("MouldLife_remaining", "min"),
            )
            .reset_index()
        )
        press_count["SKUCode"] = press_count["SKUCode"].str.strip()
        running_skus = set(press_count["SKUCode"])

        all_skus = demand_skus | running_skus
        rows = []
        for sku in sorted(all_skus):
            in_demand  = sku in demand_skus
            is_running = sku in running_skus
            if in_demand and is_running:
                cat = "Runner-In"
            elif not in_demand and is_running:
                cat = "Runner-Out"
            else:
                cat = "Non-Runner-In"

            pc_row = press_count[press_count["SKUCode"] == sku]
            run_count  = int(pc_row["RunningPressCount"].values[0]) if len(pc_row) else 0
            mould_life = int(pc_row["MouldLife_min"].values[0])     if len(pc_row) else 0
            rows.append({
                "SKUCode":           sku,
                "Category":          cat,
                "RunningPressCount": run_count,
                "MouldLife_min":     mould_life,
            })

        return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# CYCLE TIME RESOLVER
# ══════════════════════════════════════════════════════════════════════════════

class CycleTimeResolver:
    """Resolve effective cycle time per SKU; fall back to DEFAULT_CYCLE_TIME_MIN."""

    def resolve(
        self,
        skus: list[str],
        df_cycle_times: pd.DataFrame,
    ) -> dict[str, float]:
        """Returns {SKUCode: effective_CT_min}."""
        ct_lookup = dict(zip(
            df_cycle_times["SKUCode"].str.strip(),
            df_cycle_times["CycleTime_min"].astype(float),
        ))
        result = {}
        for sku in skus:
            ct = ct_lookup.get(sku)
            if ct is None or (isinstance(ct, float) and math.isnan(ct)) or ct <= 0:
                ct = ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN
            result[sku] = float(ct)
        return result


# ══════════════════════════════════════════════════════════════════════════════
# CONSUMPTION CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

class ConsumptionCalculator:
    """Compute GT consumption per shift per SKU and build the planning-horizon table."""

    def compute(
        self,
        df_classify: pd.DataFrame,
        ct_map: dict[str, float],
        df_demand: pd.DataFrame,
        plan_start: datetime,
        planning_days: int = 30,
    ) -> pd.DataFrame:
        """
        Returns the full consumption DataFrame covering all planning shifts.

        Columns:
          SKUCode, Category, Running_Press_Count, Effective_CT_Min,
          Qty_Per_Press_Per_Shift, Total_GT_Per_Shift_Day0,
          Demand_Qty, Priority_Score
        """
        # Merge demand info
        demand_lookup  = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Quantity"]))
        priority_lookup = dict(zip(df_demand["SKUCode"].str.strip(), df_demand["Priority"]))

        records = []
        for _, row in df_classify.iterrows():
            sku        = row["SKUCode"]
            category   = row["Category"]
            press_count = int(row["RunningPressCount"])

            ct = ct_map.get(sku, ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
            qty_per_press = math.floor(ConsumptionConfig.SHIFT_MINS / ct) \
                            * ConsumptionConfig.CAVITIES_PER_MOULD
            total_gt = press_count * qty_per_press

            records.append({
                "SKUCode":                  sku,
                "Category":                 category,
                "Running_Press_Count":       press_count,
                "MouldLife_min":             int(row["MouldLife_min"]),
                "Effective_CT_Min":          ct,
                "Qty_Per_Press_Per_Shift":   qty_per_press,
                "Total_GT_Per_Shift_Day0":   total_gt,
                "Demand_Qty":                demand_lookup.get(sku, 0),
                "Priority_Score":            priority_lookup.get(sku, 0),
            })

        df = pd.DataFrame(records)
        # Sort: Runner-In first, then Runner-Out, then Non-Runner-In; within each by priority desc
        cat_order = {"Runner-In": 0, "Runner-Out": 1, "Non-Runner-In": 2}
        df["_cat_ord"] = df["Category"].map(cat_order)
        df = df.sort_values(["_cat_ord", "Priority_Score"], ascending=[True, False]) \
               .drop(columns=["_cat_ord"]) \
               .reset_index(drop=True)
        return df

    def build_shift_index(
        self,
        plan_start: datetime,
        planning_days: int,
    ) -> list[tuple[datetime, str, int]]:
        """
        Returns list of (shift_start_dt, shift_label, shift_idx) for all planning shifts.
        Shift labels: A (07-15), B (15-23), C (23-07+1).
        """
        shifts = []
        shift_hours = [7, 15, 23]
        shift_labels = ["A", "B", "C"]
        # Start one shift before plan_start (building pre-start shift)
        pre_start = plan_start - timedelta(hours=ConsumptionConfig.HOURS_PER_SHIFT)
        for day_offset in range(-1, planning_days):  # -1 = pre-start day
            base_date = plan_start.date() + timedelta(days=day_offset)
            for sh, label in zip(shift_hours, shift_labels):
                dt = datetime(base_date.year, base_date.month, base_date.day, sh, 0, 0)
                if dt >= pre_start:
                    shifts.append((dt, label, len(shifts)))
        return shifts


# ══════════════════════════════════════════════════════════════════════════════
# SKU ELIGIBILITY FILTER
# ══════════════════════════════════════════════════════════════════════════════

class SKUEligibilityFilter:
    """
    Checks each demand SKU against building and curing master + history data.

    Eligibility rules:
      - Building OK  : SKU in (Master_Building_Allowable OR Building_Stage1/2_History)
      - Curing OK    : SKU in (Master_Curing_Allowable  OR Daily_Running_Moulds history)
      - BOTH must be OK; failing either → excluded with remark

    CT missing is NOT an exclusion criterion — default CT = 17 min is used instead.
    """

    def filter(
        self,
        df_demand: pd.DataFrame,
        bld_master_skus: set,
        bld_history_skus: set,
        cur_master_skus: set,
        cur_history_skus: set,
    ) -> tuple:
        """
        Returns (df_eligible, df_excluded).

        df_excluded columns:
          SKUCode, Demand_Qty, Priority_Score, Remark
        """
        bld_pool = {s.upper() for s in (bld_master_skus | bld_history_skus)}
        cur_pool = {s.upper() for s in (cur_master_skus | cur_history_skus)}

        eligible_rows: list[dict] = []
        excluded_rows: list[dict] = []

        for _, row in df_demand.iterrows():
            sku    = str(row["SKUCode"]).strip()
            sku_up = sku.upper()
            in_bld = sku_up in bld_pool
            in_cur = sku_up in cur_pool

            if in_bld and in_cur:
                eligible_rows.append(row.to_dict())
            else:
                missing = []
                if not in_bld:
                    missing.append("building machine")
                if not in_cur:
                    missing.append("curing mould")
                excluded_rows.append({
                    "SKUCode":        sku,
                    "Demand_Qty":     float(row.get("Quantity", 0)),
                    "Priority_Score": float(row.get("Priority", 0)),
                    "Remark": f"No PDE & master data- {', '.join(missing)}",
                })

        df_eligible = (
            pd.DataFrame(eligible_rows)
            if eligible_rows
            else df_demand.iloc[0:0].copy()
        )
        df_excluded = (
            pd.DataFrame(excluded_rows)
            if excluded_rows
            else pd.DataFrame(
                columns=["SKUCode", "Demand_Qty", "Priority_Score", "Remark"]
            )
        )
        return df_eligible, df_excluded


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════

_NAVY   = "1F3864"
_WHITE  = "FFFFFF"
_GREEN  = "E2EFDA"
_AMBER  = "FFF2CC"
_RED    = "FFE0E0"
_BLUE   = "DDEEFF"

def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _bold_white() -> Font:
    return Font(bold=True, color=_WHITE)

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1):
    """Write a DataFrame to a worksheet starting at start_row."""
    hdr_fill  = _hdr_fill(_NAVY)
    hdr_font  = _bold_white()
    border    = _thin_border()
    cat_colors = {
        "Runner-In":     _GREEN,
        "Runner-Out":    _AMBER,
        "Non-Runner-In": _BLUE,
    }

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        category = row.get("Category", "")
        row_fill = _hdr_fill(cat_colors.get(category, _WHITE)) if category else None
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_fill:
                cell.fill = row_fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 35)


def export_consumption_table(
    df_consumption: pd.DataFrame,
    df_gt_inventory: pd.DataFrame,
    output_path: str,
    df_excluded: "pd.DataFrame | None" = None,
):
    """Write curing_consumption_table.xlsx with two sheets."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # ── Sheet 1: Consumption Summary ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Consumption Summary"

    # Category legend
    legend_text = (
        "Runner-In (green): in demand + running    "
        "Runner-Out (amber): not in demand + running    "
        "Non-Runner-In (blue): in demand + not running"
    )
    ws1.cell(row=1, column=1, value=legend_text).font = Font(italic=True, size=9)
    ws1.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=len(df_consumption.columns))

    _write_dataframe(ws1, df_consumption, start_row=3)

    # Summary stats below
    n_ri  = (df_consumption["Category"] == "Runner-In").sum()
    n_ro  = (df_consumption["Category"] == "Runner-Out").sum()
    n_nri = (df_consumption["Category"] == "Non-Runner-In").sum()
    total_gt = df_consumption.loc[
        df_consumption["Category"] == "Runner-In", "Total_GT_Per_Shift_Day0"
    ].sum()

    last_row = len(df_consumption) + 5
    ws1.cell(row=last_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws1.cell(row=last_row + 1, column=1, value=f"Runner-In SKUs: {n_ri}")
    ws1.cell(row=last_row + 2, column=1, value=f"Runner-Out SKUs: {n_ro}")
    ws1.cell(row=last_row + 3, column=1, value=f"Non-Runner-In SKUs: {n_nri}")
    ws1.cell(row=last_row + 4, column=1,
             value=f"Total GT consumed/shift (Runner-In): {total_gt:,}")

    # ── Sheet 2: GT Inventory ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("GT Inventory")
    ws2.cell(row=1, column=1, value="Opening GT Inventory (from DB: gt_inventory_manual)") \
       .font = Font(bold=True, italic=True)
    _write_dataframe(ws2, df_gt_inventory, start_row=3)

    # ── Sheet 3: Excluded SKUs ────────────────────────────────────────────────
    if df_excluded is not None and len(df_excluded) > 0:
        ws3 = wb.create_sheet("Excluded SKUs")
        ws3.cell(
            row=1, column=1,
            value=(
                "SKUs excluded from planning — not found in building/curing "
                "master data OR historical data. CT missing → default 17 min used (not an exclusion)."
            ),
        ).font = Font(italic=True, size=9)
        ws3.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(df_excluded.columns),
        )
        _write_dataframe(ws3, df_excluded, start_row=3)

    wb.save(output_path)
    print(f"  [Consumption] Saved → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def build_consumption_table(
    demand_path: str,
    engine,
    output_path: str,
    plan_start: datetime | None = None,
    planning_days: int = ConsumptionConfig.PLANNING_DAYS,
) -> dict:
    """
    Build the curing consumption table and export to Excel.

    Returns:
        {
          "consumption_df":       pd.DataFrame,   # full per-SKU consumption table
          "ct_map":               dict[str, float], # {SKUCode: effective_CT_min}
          "gt_inventory":         pd.DataFrame,   # opening GT inventory
          "demand_df":            pd.DataFrame,   # loaded demand
          "classify_df":          pd.DataFrame,   # SKU classification
        }
    """
    if plan_start is None:
        plan_start = datetime(2026, 6, 1, 7, 0, 0)

    print("\n" + "=" * 70)
    print("  B2C Phase 0 — Curing Consumption Table")
    print("=" * 70)

    etl = ConsumptionETL(engine)

    print("  [ETL] Loading demand …")
    df_demand = etl.load_demand(demand_path)
    print(f"        {len(df_demand)} demanded SKUs")

    print("  [ETL] Loading cycle times …")
    df_ct = etl.load_cycle_times()
    print(f"        {len(df_ct)} SKUs with CT from DB")

    print("  [ETL] Loading running moulds (active curing presses) …")
    df_running = etl.load_running_moulds()
    print(f"        {len(df_running)} active curing press rows")

    print("  [ETL] Loading GT inventory …")
    df_gt_inv = etl.load_gt_inventory()
    print(f"        {len(df_gt_inv)} SKUs with opening GT inventory")

    # ── SKU eligibility: must appear in building AND curing master or history ──
    print("  [Eligible] Loading building/curing master & history for eligibility check …")
    bld_master  = etl.load_building_allowable_skus()
    bld_history = etl.load_building_history_skus()
    cur_master  = etl.load_curing_allowable_skus()
    cur_history = etl.load_curing_history_skus()
    print(f"        Bld master: {len(bld_master)} | Bld history: {len(bld_history)} | "
          f"Cur master: {len(cur_master)} | Cur history: {len(cur_history)}")

    eligibility = SKUEligibilityFilter()
    df_demand, df_excluded = eligibility.filter(
        df_demand, bld_master, bld_history, cur_master, cur_history
    )
    if len(df_excluded) > 0:
        print(f"  ⚠  {len(df_excluded)} SKU(s) excluded (no PDE/master data):")
        for _, r in df_excluded.iterrows():
            print(f"       {r['SKUCode']}: {r['Remark']}")
    print(f"  [Eligible] {len(df_demand)} SKU(s) proceed to planning")

    # Classify SKUs
    print("  [Classify] Classifying SKUs …")
    classifier = SKUClassifier()
    df_classify = classifier.classify(df_demand, df_running)
    ri  = (df_classify["Category"] == "Runner-In").sum()
    ro  = (df_classify["Category"] == "Runner-Out").sum()
    nri = (df_classify["Category"] == "Non-Runner-In").sum()
    print(f"        Runner-In: {ri}  |  Runner-Out: {ro}  |  Non-Runner-In: {nri}")

    # Resolve cycle times
    all_skus = df_classify["SKUCode"].tolist()
    ct_resolver = CycleTimeResolver()
    ct_map = ct_resolver.resolve(all_skus, df_ct)
    n_default = sum(1 for v in ct_map.values() if v == ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN)
    print(f"  [CT] {len(ct_map)} SKUs resolved  |  {n_default} using default {ConsumptionConfig.DEFAULT_CYCLE_TIME_MIN} min")

    # Compute consumption
    print("  [Compute] Computing per-shift GT consumption …")
    calc = ConsumptionCalculator()
    df_consumption = calc.compute(df_classify, ct_map, df_demand, plan_start, planning_days)

    total_gt_ri = df_consumption.loc[
        df_consumption["Category"] == "Runner-In", "Total_GT_Per_Shift_Day0"
    ].sum()
    print(f"  [Compute] Total GT consumed/shift by Runner-In presses: {total_gt_ri:,}")

    # Sanity check
    default_check = math.floor(480 / 17.0) * 2  # = 56
    print(f"  [Check] floor(480/17.0)*2 = {default_check} (expected 56)")

    # Export
    print(f"  [Export] Writing → {output_path}")
    export_consumption_table(df_consumption, df_gt_inv, output_path, df_excluded)

    print("=" * 70)
    print("  Phase 0 complete.")
    print("=" * 70 + "\n")

    return {
        "consumption_df":  df_consumption,
        "ct_map":          ct_map,
        "gt_inventory":    df_gt_inv,
        "demand_df":       df_demand,
        "classify_df":     df_classify,
        "excluded_df":     df_excluded,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STANDALONE  RUN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from cbc_env import make_engine

    _DEMAND = os.path.join(IN_DIR, "demand_tomerji_june_normalized.xlsx")
    _OUTPUT = os.path.join(OUT_DIR, "curing_consumption_table.xlsx")
    _PLAN_START = datetime(2026, 6, 1, 7, 0, 0)

    engine = make_engine()
    result = build_consumption_table(
        demand_path   = _DEMAND,
        engine        = engine,
        output_path   = _OUTPUT,
        plan_start    = _PLAN_START,
        planning_days = ConsumptionConfig.PLANNING_DAYS,
    )

    df = result["consumption_df"]
    print("\nTop 10 Runner-In SKUs by Total GT Per Shift:")
    top = df[df["Category"] == "Runner-In"] \
          .sort_values("Total_GT_Per_Shift_Day0", ascending=False) \
          .head(10)
    print(top[["SKUCode", "Running_Press_Count", "Effective_CT_Min",
               "Qty_Per_Press_Per_Shift", "Total_GT_Per_Shift_Day0"]].to_string(index=False))
