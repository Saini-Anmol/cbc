"""
feasibility_test.py
====================
Comprehensive, INDEPENDENT feasibility / business-rule validator for the JK Tyre
BTP B2C Building + Curing schedules. It does NOT run or trust the planner — it
reads the two generated output workbooks and re-derives every reference value from
the authoritative sources (DB, the building-CT CSV, bc_config constants), then
validates the schedule row-by-row / shift-by-shift / machine-by-machine /
press-by-press / SKU-by-SKU.

Run:
    myenv/bin/python feasibility_test.py \
        --building data/output/main_output/bc_building_schedule_2026-07-01.xlsx \
        --curing   data/output/main_output/bc_curing_b2c_2026-07-01.xlsx \
        --demand   data/input/july_correct_plan.xlsx \
        --plan-month 2026-07 \
        --holidays 2026-07-15,2026-07-16          (optional; else reads bc_config.PLANT_HOLIDAYS) \
        --out feasibility_report.xlsx

Exit code 1 if any rule FAILs, else 0.

================================================================================
HARD BUSINESS RULES VALIDATED (exactly as agreed with the plant planner)
================================================================================
R1.  CYCLE TIMES.
       - BUILDING CT is taken ONLY from data/input/Cycle_time_Building.xlsx,
         per (SKU Code, machine), in seconds/unit. NO 94% efficiency divisor on
         building. A built (SKU,machine) pair with no CSV entry is a FAIL
         (missing-source), never silently defaulted.
       - CURING CT is taken from the DB ("Cure Time") and divided by the 94% plant
         efficiency: CT = round(RawCureTime / 0.94). The curing sheet's
         CycleTime_min must match this DB-derived value.

R2.  CHANGEOVER / CLEAN TIMES come from config, not invented:
       - Building same_size_CO / diff_size_CO minutes = bc_config
         BUILDING_CO_SAME_SIZE / BUILDING_CO_DIFF_SIZE for that machine group
         (a DB "+10" changeover variant is tolerated).
       - Curing CO time  = CURING_CO_CHANGEOVER_MINS (480).
       - Mould clean time = MOULD_CLEAN_MINS (480).

R3.  ALLOWABLE MATRICES strictly followed:
       - Every built (SKU,machine) is in Master_Building_Allowable_Machines
         (Stage-1 carcass rows exempt — carcass is upstream of the GT SKU).
       - Every cured (SKU,press) is in Master_Curing_Allowable_Machines.
       - Mould mapping (Master_Mapping_Mould_SKU) respected — see R17.

R4.  HISTORICAL INCH-SIZE respected: each building GT row's SKU inch is inside the
       machine's historical allowed-inch set (b2c_pipeline._MACHINE_ALLOWED_INCH_SET).
       Stage-1 carcass rows exempt.

R5.  STAGE-2 vs STAGE-1: Stage-2 GT production for a SKU must never exceed the
       cumulative Stage-1 carcass produced for that SKU available at that time
       (carcass 1-day aging window). Stage-2 cannot start before its carcass exists.

R6.  OPENING INVENTORY: opening GT and opening carcass are taken from the DB for
       the given plan_month; the curing sheet's opening GT_Inventory must match.

R7.  UTILIZATION <= 100%: no building machine and no curing press may exceed 100%
       occupancy. HOLIDAY-AWARE: available minutes = working_days x 3 x 480 (holiday
       shifts dropped from the denominator).

R8.  DEMAND CAP: per SKU, total GT built <= demand AND total cured <= demand (from
       the demand file). Excess production is flagged.

R9.  AGING (STRICT FIFO, per SKU, calendar days incl. holidays):
       - GT aging = 3 days: GT built on day D must be cured by day D+3, else the
         un-cured remainder EXPIRES = FAIL.
       - Carcass aging = 1 day: carcass built on day D must be consumed by Stage-2
         by day D+1, else it EXPIRES = FAIL.
       Each SKU keeps its OWN GT/carcass lots (FIFO oldest-first consumption).

R10. CURING CO CAP: curing changeovers per calendar day <= MAX_CHANGEOVERS_PER_DAY
       (12). Also curing CO time = mould clean time = 480 minutes.

R11. SHIFT TIME: one shift = 480 min, 3 shifts/day. Production + CO minutes must
       never exceed the available shift time (per machine/press/shift).

R12. MISSING CURING CT falls back to the default 17.0 minutes.

R13. CURING PRESS ROSTER: only the 170 presses in the curing allowable matrix may
       be used. A press appearing in Daily_Running_Moulds but NOT in the 170 set
       must NOT be used.

R14. BUILDING MACHINE ROSTER: only the 38 building machines may appear. No other
       building machine is allowed in the schedule.

R15. CAPACITY PER MACHINE/PRESS/SHIFT:
         Production_Mins + CO_Mins + Mould_Clean_Mins <= 480.

R16. ONE SKU AT A TIME: no building machine and no curing press may PRODUCE two
       different SKUs in the same shift simultaneously.

R17. MOULD FEASIBILITY (bipartite): the output does not emit a mould-id per row, so
       for each shift we verify a valid DISJOINT assignment exists — every running
       press gets 2 eligible moulds and no mould is shared by two presses in the
       same shift (Day-0 mounted moulds folded in). A SKU with <2 eligible moulds
       that is nonetheless run is a FAIL (structural). If the curing Mould Movement
       / Mould Tracker sheets carry real mould-ids they are additionally checked for
       the same physical mould appearing on two presses in one shift.

R18. CT<->QTY RECONCILIATION:
       - Building: Production_Mins = Qty x CT_sec / 60.
       - Curing:   Production_Mins = (Qty / CURING_CAVITIES) x CT   (cavities = 2).
       Reconciled per row; a single production row can never need > 480 min.

R19. NO PRODUCTION DURING A CHANGEOVER: a building CO row and a curing CO/clean row
       must carry Qty = 0. No row mixes changeover time with production.
================================================================================

NOTES
- Reference values are DB/CSV/config-derived; the validator never trusts the sheet
  for a value it can independently source.
- Missing sheets, missing columns, missing data and ambiguous cases are reported
  loudly (DATA-ISSUE), never silently ignored.
- Holiday-aware: aging counts calendar days (GT/carcass degrade during a shutdown);
  utilization denominators drop holiday shifts; no production is expected on a
  holiday shift.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import defaultdict, deque
from datetime import datetime, timedelta

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  edit these defaults for the month you want to validate, then
#  just run:   myenv/bin/python feasibility_test.py
#  (every value can still be overridden on the command line — flags mirror the keys).
# ══════════════════════════════════════════════════════════════════════════════
CONFIG = {
    # ── INPUT FILES ─────────────────────────────────────────────────────────────
    "building":      "data/output/main_output/bc_building_schedule_2026-07-01.xlsx",
    "curing":        "data/output/main_output/bc_curing_b2c_2026-07-01.xlsx",
    "demand":        "data/input/july_correct_plan.xlsx",   # None → bc_config.DEMAND_FILE
    "bld_ct_file":   None,                                   # None → bc_config.BLD_CT_FILE (Cycle_time_Building.xlsx)
    # ── MONTH / HORIZON ─────────────────────────────────────────────────────────
    "plan_month":    "2026-07",     # YYYY-MM — DB plan_month (moulds / opening GT / running-moulds)
    "plan_start":    "2026-07-01",  # YYYY-MM-DD — anchors the aging / horizon clock
    "planning_days": 32,            # horizon length in days (None → infer from the sheet dates)
    "holidays":      "",            # "" = none;  e.g. "2026-07-15,2026-07-16"
    # ── PLANT CONSTANTS  (None → take the bc_config default) ────────────────────
    "gt_shelf":        None,   # GT shelf life days       (bc_config 3)
    "carcass_shelf":   None,   # carcass shelf life days  (1)
    "max_co_day":      None,   # max curing COs / day     (12)
    "cavities":        None,   # curing cavities / press  (2 → 2 tyres per cycle)
    "shift_mins":      None,   # minutes per shift        (480)
    "default_cure_ct": None,   # fallback curing CT       (17.0)
    "curing_co_mins":  None,   # curing CO minutes        (480)
    "clean_mins":      None,   # mould-clean minutes      (480)
    # ── REPORT / TOLERANCES ─────────────────────────────────────────────────────
    "out":       "feasibility_report_july.xlsx",
    "max_rows":  30,           # max sample violation rows printed per rule
    "ct_tol":    0.75,         # minute tolerance for the CT (sheet-vs-DB) match
    "qty_tol":   1.0,          # unit tolerance for demand-cap / reconciliation
}

# ── CLI  (parsed BEFORE the DB modules import; each flag defaults to CONFIG above) ──
_ap = argparse.ArgumentParser(description="Independent B2C Building+Curing feasibility validator")
_ap.add_argument("--building", default=CONFIG["building"], help="Building schedule xlsx")
_ap.add_argument("--curing", default=CONFIG["curing"], help="Curing schedule xlsx")
_ap.add_argument("--demand", default=CONFIG["demand"], help="demand xlsx (None → bc_config.DEMAND_FILE)")
_ap.add_argument("--bld-ct-file", default=CONFIG["bld_ct_file"], help="building-CT CSV (None → bc_config.BLD_CT_FILE)")
_ap.add_argument("--plan-month", default=os.environ.get("PLAN_MONTH", CONFIG["plan_month"]), help="YYYY-MM DB plan_month")
_ap.add_argument("--plan-start", default=CONFIG["plan_start"], help="YYYY-MM-DD plan start (aging/horizon anchor)")
_ap.add_argument("--planning-days", type=int, default=CONFIG["planning_days"], help="horizon days (None → infer)")
_ap.add_argument("--holidays", default=CONFIG["holidays"], help="comma YYYY-MM-DD holidays ('' = none)")
_ap.add_argument("--gt-shelf", type=int, default=CONFIG["gt_shelf"], help="GT shelf life days")
_ap.add_argument("--carcass-shelf", type=int, default=CONFIG["carcass_shelf"], help="carcass shelf life days")
_ap.add_argument("--max-co-day", type=int, default=CONFIG["max_co_day"], help="max curing COs/day")
_ap.add_argument("--cavities", type=int, default=CONFIG["cavities"], help="curing cavities/press")
_ap.add_argument("--shift-mins", type=int, default=CONFIG["shift_mins"], help="minutes per shift")
_ap.add_argument("--default-cure-ct", type=float, default=CONFIG["default_cure_ct"], help="fallback curing CT")
_ap.add_argument("--curing-co-mins", type=int, default=CONFIG["curing_co_mins"], help="curing CO minutes")
_ap.add_argument("--clean-mins", type=int, default=CONFIG["clean_mins"], help="mould-clean minutes")
_ap.add_argument("--out", default=CONFIG["out"], help="Excel report path")
_ap.add_argument("--max-rows", type=int, default=CONFIG["max_rows"], help="max sample violation rows printed per rule")
_ap.add_argument("--ct-tol", type=float, default=CONFIG["ct_tol"], help="minute tolerance for CT match")
_ap.add_argument("--qty-tol", type=float, default=CONFIG["qty_tol"], help="unit tolerance for demand-cap / reconciliation")
_ap.add_argument("--midmonth-opening", default=None,
                 help="JSON of the plan's day-start opening (aged GT lots + carcass totals) written by "
                      "run_rolling_pipeline_2pass; seeds R9G/R9C/R5 from the Run-2 start-date inventory "
                      "instead of the 1st-of-month DB. Omit for a normal 1st-of-month audit.")
ARGS = _ap.parse_args()

# plan_month must be set in the env BEFORE bc_config / the ETLs import (imported by value)
if ARGS.plan_month:
    os.environ["PLAN_MONTH"] = ARGS.plan_month
    os.environ["RUNNING_MOULDS_MONTH"] = ARGS.plan_month
if ARGS.holidays is not None:
    os.environ["PLANT_HOLIDAYS"] = ARGS.holidays

import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════════════════════════
#  Reference sources (DB / CSV / config) — the authoritative side of every check
# ══════════════════════════════════════════════════════════════════════════════
SRC: dict = {}
DATA_ISSUES: list = []   # missing sheets/columns/data / ambiguous — reported loudly


def _issue(kind: str, detail: str):
    DATA_ISSUES.append({"kind": kind, "detail": detail})


def _try(key, fn):
    try:
        SRC[key] = fn()
    except Exception as e:  # noqa: BLE001
        SRC[key] = None
        _issue("source-load-failed", f"{key}: {type(e).__name__}: {e}")


# machine-group fallback (used if the b2c import fails)
_MG_FALLBACK: dict = {}
for _g, _ms in {
    "VMI": ["6001", "6002", "6003", "6004", "7001", "7002", "7003", "7004"],
    "BJ": ["7101", "7102", "7103", "7104", "7105", "7106", "7201"],
    "UNISTAGE": ["7501", "7502", "7503"],
    "STAGE2": ["8201", "8301", "8302", "8501", "8502", "7301"],
    "STAGE1": ["6802", "6803", "6909", "6911", "7601", "7701", "7801", "7802",
               "7803", "7804", "8001", "8002", "8003", "8101"],
}.items():
    for _m in _ms:
        _MG_FALLBACK[_m] = _g


def load_sources():
    import bc_config as bc
    SRC["bc"] = bc
    if not ARGS.plan_month:
        os.environ.setdefault("PLAN_MONTH", bc.PLAN_START.strftime("%Y-%m"))
    # plan_start anchors the aging/horizon clock — CLI override lets ANY month validate
    # without editing bc_config.
    if ARGS.plan_start:
        try:
            SRC["plan_start"] = datetime.strptime(ARGS.plan_start[:10], "%Y-%m-%d")
        except Exception:
            _issue("bad-plan-start", f"--plan-start '{ARGS.plan_start}' not YYYY-MM-DD; using bc_config")
            SRC["plan_start"] = bc.PLAN_START
    else:
        SRC["plan_start"] = bc.PLAN_START

    # config constants (R2, R10, R11, R15) — each takes the CLI override when given, else bc_config
    SRC["CO_SAME"] = dict(bc.BUILDING_CO_SAME_SIZE)
    SRC["CO_DIFF"] = dict(bc.BUILDING_CO_DIFF_SIZE)
    SRC["CURING_CO"] = ARGS.curing_co_mins if ARGS.curing_co_mins is not None else int(getattr(bc, "CURING_CO_CHANGEOVER_MINS", 480))
    SRC["CLEAN_MINS"] = ARGS.clean_mins if ARGS.clean_mins is not None else int(getattr(bc, "MOULD_CLEAN_MINS", 480))
    SRC["MAX_CO_DAY"] = ARGS.max_co_day if ARGS.max_co_day is not None else int(getattr(bc, "MAX_CHANGEOVERS_PER_DAY", 12))
    SRC["GT_AGE_DAYS"] = ARGS.gt_shelf if ARGS.gt_shelf is not None else int(getattr(bc, "GT_SHELF_LIFE_DAYS", 3))
    SRC["CARC_AGE_DAYS"] = ARGS.carcass_shelf if ARGS.carcass_shelf is not None else int(getattr(bc, "CARCASS_SHELF_LIFE_DAYS", 1))
    SRC["SHIFT_MINS"] = ARGS.shift_mins if ARGS.shift_mins is not None else int(getattr(bc, "SHIFT_MINS", 480))
    SRC["CAVITIES"] = ARGS.cavities if ARGS.cavities is not None else int(getattr(bc, "CURING_CAVITIES", 2))
    SRC["DEFAULT_CURE_CT"] = ARGS.default_cure_ct if ARGS.default_cure_ct is not None else float(getattr(bc, "DEFAULT_CURING_CT", 17.0))

    # holidays → set of 'YYYY-MM-DD'
    if ARGS.holidays is not None:
        hol = {h.strip() for h in ARGS.holidays.split(",") if h.strip()}
    else:
        hol = {str(h).strip() for h in (getattr(bc, "PLANT_HOLIDAYS", []) or [])}
    SRC["holidays"] = hol

    # building machine-group + inch sets (from b2c_pipeline — the CSV/inch truth)
    def _bld_helpers():
        import b2c_pipeline as bp
        inch = getattr(bp, "_MACHINE_ALLOWED_INCH_SET", {}) or {}
        return {
            "MG": dict(getattr(bp, "_MACHINE_GROUP", {})),
            "inch_set": {str(k): set(str(x) for x in v) for k, v in inch.items()},
        }
    _try("bld", _bld_helpers)

    # building CT from the CSV ONLY (R1) — per (SKU Code, machine), seconds. No 0.94.
    def _bld_ct_csv():
        path = ARGS.bld_ct_file or getattr(bc, "BLD_CT_FILE", os.path.join("data", "input", "Cycle_time_Building.xlsx"))
        df = (pd.read_excel(path, dtype=str) if str(path).lower().endswith((".xlsx", ".xls"))
              else pd.read_csv(path, dtype=str))
        if "SKU Code" not in df.columns:
            raise ValueError(f"'SKU Code' column missing in {path}")
        mach_cols = [c for c in df.columns[2:]]
        out: dict = {}
        for _, r in df.iterrows():
            sku = str(r.get("SKU Code", "")).strip()
            if not sku:
                continue
            for m in mach_cols:
                v = r.get(m)
                if pd.notna(v) and str(v).strip() != "":
                    try:
                        out[(sku, str(m).strip())] = float(v)
                    except (TypeError, ValueError):
                        pass
        SRC["bld_ct_path"] = path
        return out
    _try("bld_ct", _bld_ct_csv)

    # DB engine
    def _engine():
        try:
            from connection import get_engine
            return get_engine()
        except Exception:
            from bc_config import make_engine
            return make_engine()
    _try("engine", _engine)
    eng = SRC.get("engine")

    if eng is not None:
        from connection import ConsumptionETL
        etl = ConsumptionETL(eng)
        SRC["etl"] = etl

        # curing CT (DB / 0.94, default 17) — R1, R12
        def _cure_ct():
            df = etl.load_cycle_times()
            scol = df.columns[0]
            ccol = "CycleTime_min" if "CycleTime_min" in df.columns else df.columns[-1]
            return {str(r[scol]): float(r[ccol]) for _, r in df.iterrows()}
        _try("cure_ct", _cure_ct)

        _try("presses170", lambda: set(map(str, etl.load_allowable_press_ids())))

        def _cure_allow():
            m = etl.load_curing_allowable()
            out: dict = {}
            if isinstance(m, pd.DataFrame):
                pcol = "Machines" if "Machines" in m.columns else ("Presses" if "Presses" in m.columns else m.columns[-1])
                scol = "SKUCode" if "SKUCode" in m.columns else m.columns[0]
                for _, r in m.iterrows():
                    mm = r[pcol]
                    if isinstance(mm, str):
                        mm = [t for t in re.split(r"[ ,]+", mm) if t.strip()]
                    out[str(r[scol])] = set(str(x).strip() for x in (mm or []))
            else:
                out = {str(k): set(map(str, v)) for k, v in dict(m).items()}
            return out
        _try("cure_allow", _cure_allow)

        def _mould():
            m = etl.load_mould_eligibility()
            sm = {str(k): set(map(str, v)) for k, v in m["sku_moulds"].items()}
            ms = {str(k): set(map(str, v)) for k, v in m["mould_skus"].items()}
            return {"sku_moulds": sm, "mould_skus": ms}
        _try("mould", _mould)

        # Day-0 running moulds → press set + (press,SKU) mounted pairs (R3C exception)
        def _rm():
            df = etl.load_running_moulds()
            pcol = "Machine" if "Machine" in df.columns else df.columns[0]
            presses = set(str(x) for x in df[pcol].tolist())
            scol = next((c for c in df.columns if str(c).upper() in ("SKUCODE", "SKU", "SAPCODE")
                         or "SKU" in str(c).upper()), None)
            pairs = set()
            if scol is not None:
                for _, r in df.iterrows():
                    pairs.add((str(r[pcol]).strip(), str(r[scol]).strip()))
            return {"df": df, "presses": presses, "pairs": pairs, "pcol": pcol}
        _try("rm", _rm)

        def _open_gt():
            m = etl.load_gt_inventory()
            if isinstance(m, dict):
                return {str(k): float(v) for k, v in m.items()}
            scol = "SKUCode" if "SKUCode" in m.columns else m.columns[0]
            vcol = "GT_Inventory" if "GT_Inventory" in m.columns else m.columns[-1]
            return {str(r[scol]): float(r[vcol]) for _, r in m.iterrows()}
        _try("open_gt", _open_gt)

        def _open_carc():
            from curing_b2c import _load_opening_carcass
            m = _load_opening_carcass(eng)
            return {str(k): float(v) for k, v in dict(m).items()}
        _try("open_carc", _open_carc)

        # MID-MONTH: seed opening GT (aged lots) + carcass from the plan's day-start snapshot so
        # R9G/R9C/R5 audit the Run-2 start-date inventory, not the 1st-of-month DB. R6 keeps the DB
        # opening (unchanged). age_days per lot → the FIFO seeds it at (−age) inside _fifo_expiry.
        if getattr(ARGS, "midmonth_opening", None):
            import json as _json
            with open(ARGS.midmonth_opening) as _f:
                _mm = _json.load(_f)
            SRC["mm_gt_lots"] = {str(s): [(int(a), float(q)) for a, q in lots]
                                 for s, lots in _mm.get("gt_lots", {}).items()}
            SRC["mm_carc"] = {str(s): float(v) for s, v in _mm.get("carcass", {}).items()}
            print(f"  [midmonth-audit] opening seeded from {ARGS.midmonth_opening}: "
                  f"{len(SRC['mm_gt_lots'])} GT-lot SKUs, {len(SRC['mm_carc'])} carcass SKUs")

        def _bld_allow():
            from connection import ETL as BETL
            df = BETL(eng).load_machine_allowable()
            out: dict = {}
            for _, r in df.iterrows():
                mm = r["Machines"]
                if isinstance(mm, str):
                    mm = [t for t in re.split(r"[ ,]+", mm) if t.strip()]
                out[str(r["SKUCode"])] = set(str(x).strip() for x in mm)
            return out
        _try("bld_allow", _bld_allow)

        def _sku_inch():
            try:
                m = etl.load_sku_sizes()
                return {str(k): str(v).replace('"', "").strip() for k, v in dict(m).items()}
            except Exception:
                return {}
        _try("sku_inch", _sku_inch)

    # demand per SKU
    def _demand():
        path = ARGS.demand or bc.DEMAND_FILE
        SRC["demand_path"] = path
        raw = pd.read_excel(path)
        scol = next((c for c in raw.columns
                     if str(c).lower() in ("skucode", "sku code", "sku_code", "sapcode")),
                    raw.columns[0])
        qcol = next((c for c in ("Quantity", "Updated_Requirement", "Requirement", "requirement", "Demand")
                     if c in raw.columns), None)
        if qcol is None:
            qcol = next((c for c in raw.columns if "req" in str(c).lower() or "qty" in str(c).lower()),
                        raw.columns[-1])
        g = raw.groupby(raw[scol].astype(str))[qcol].sum()
        return {str(k): float(v) for k, v in g.items() if float(v) > 0}
    _try("demand", _demand)


# ── small accessors ───────────────────────────────────────────────────────────
def mgroup(m):
    mg = (SRC.get("bld") or {}).get("MG") if SRC.get("bld") else None
    if mg and str(m) in mg:
        return mg[str(m)]
    return _MG_FALLBACK.get(str(m), "?")


def sku_inch(sku: str) -> str:
    d = SRC.get("sku_inch") or {}
    v = d.get(str(sku))
    if v:
        return str(v).replace('"', "").strip()
    s = str(sku)
    return s[8:10] if len(s) >= 10 else ""


S1_MACHINES = {m for m, g in _MG_FALLBACK.items() if g == "STAGE1"}
S2_MACHINES = {m for m, g in _MG_FALLBACK.items() if g == "STAGE2"}
GT_MACHINES = {m for m, g in _MG_FALLBACK.items() if g in ("VMI", "BJ", "UNISTAGE", "STAGE2")}
ALL_38 = set(_MG_FALLBACK.keys())
SHIFT_ORDER = {"A": 0, "B": 1, "C": 2}


# ── shift-window clipping (plant shifts A=07:00 B=15:00 C=23:00) ───────────────
# Output rows are a CONTINUOUS run labelled by their START shift; a run/CO can cross
# shift boundaries (e.g. 07:00->15:50). To validate capacity per REAL shift interval
# we clip each row's [StartTime,EndTime] to the shift windows it overlaps and charge
# each shift only its own minutes. This is the "shift-level interpretation" fix.
def _shift_of(dt):
    """(plan_date_str, shift) that a datetime belongs to."""
    h = dt.hour
    if 7 <= h < 15:
        return dt.strftime("%Y-%m-%d"), "A"
    if 15 <= h < 23:
        return dt.strftime("%Y-%m-%d"), "B"
    if h >= 23:
        return dt.strftime("%Y-%m-%d"), "C"
    return (dt - timedelta(days=1)).strftime("%Y-%m-%d"), "C"   # 00:00-06:59 → prev-day C


def _shift_window_end(dt):
    """Datetime at which dt's current shift window ends (next 07/15/23 boundary)."""
    h = dt.hour
    d = dt.replace(minute=0, second=0, microsecond=0)
    if 7 <= h < 15:
        return d.replace(hour=15)
    if 15 <= h < 23:
        return d.replace(hour=23)
    if h >= 23:
        return (d + timedelta(days=1)).replace(hour=7)
    return d.replace(hour=7)    # 00:00-06:59 → 07:00 same day


def _shift_segments(start, end):
    """Split [start,end] into [(plan_date, shift, minutes)] clipped at shift boundaries."""
    segs = []
    if start is None or end is None or pd.isna(start) or pd.isna(end) or end <= start:
        return segs
    cur, guard = start, 0
    while cur < end and guard < 40:
        guard += 1
        seg_end = min(_shift_window_end(cur), end)
        pdte, sh = _shift_of(cur)
        mins = (seg_end - cur).total_seconds() / 60.0
        if mins > 0:
            segs.append((pdte, sh, mins))
        cur = seg_end
    return segs


def _dur_min(start, end):
    """Duration in minutes of a row's [start,end], or None if unparseable."""
    if start is None or end is None or pd.isna(start) or pd.isna(end) or end < start:
        return None
    return (end - start).total_seconds() / 60.0


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet loading (exact header rows per the verified schema)
# ══════════════════════════════════════════════════════════════════════════════
BLD_HDR = {"Shift Schedule": 2, "Changeover Plan": 0, "SKU Classification": 0,
           "Daily GT & Carcass": 0, "Demand Fulfillment (B2C)": 0, "Machine Utilization": 1}
CUR_HDR = {"Demand Fulfillment": 0, "Machine Utilization": 1, "Shift Schedule": 0,
           "Changeover Plan": 0, "Mould Tracker": 1, "Mould Movement": 1, "MouldInUse": 1,
           "Machine Schedule": 1, "Daily Cured tyres": 0, "GT Gap Diagnostic": 0}


def load_sheet(path, sheet, hdr, required_cols=None, tag=""):
    try:
        df = pd.read_excel(path, sheet_name=sheet, header=hdr)
    except Exception as e:  # noqa: BLE001
        _issue("missing-sheet", f"{tag}[{sheet}]: {type(e).__name__}: {e}")
        return None
    df.columns = [str(c).strip() for c in df.columns]
    if required_cols:
        miss = [c for c in required_cols if c not in df.columns]
        if miss:
            _issue("missing-columns", f"{tag}[{sheet}] missing columns {miss}; have {list(df.columns)}")
    return df


def num(x, default=0.0):
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return default
        return float(x)
    except (TypeError, ValueError):
        return default


def dstr(x):
    """Normalize a Date cell to 'YYYY-MM-DD'."""
    try:
        return pd.to_datetime(x).strftime("%Y-%m-%d")
    except Exception:  # noqa: BLE001
        return str(x).strip()[:10]


# ══════════════════════════════════════════════════════════════════════════════
#  Findings
# ══════════════════════════════════════════════════════════════════════════════
COLUMNS = ["Rule", "Date", "Shift", "Machine/Press", "SKU", "Qty", "CT",
           "Prod_Mins", "CO_Mins", "Clean_Mins", "Expected", "Actual", "Detail"]
FINDINGS: list = []
RULE_STATUS: dict = {}   # rule_id -> dict(title, status, checked, violations, skipped_reason)


def add(rule, detail, date="", shift="", mp="", sku="", qty="", ct="",
        prod="", co="", clean="", expected="", actual=""):
    FINDINGS.append({
        "Rule": rule, "Date": date, "Shift": shift, "Machine/Press": mp, "SKU": sku,
        "Qty": qty, "CT": ct, "Prod_Mins": prod, "CO_Mins": co, "Clean_Mins": clean,
        "Expected": expected, "Actual": actual, "Detail": detail,
    })


def rule_result(rule, title, checked, skipped_reason=None):
    v = sum(1 for f in FINDINGS if f["Rule"] == rule)
    status = "SKIP" if skipped_reason else ("PASS" if v == 0 else "FAIL")
    RULE_STATUS[rule] = {"title": title, "status": status, "checked": checked,
                         "violations": v, "skipped_reason": skipped_reason or ""}


# ══════════════════════════════════════════════════════════════════════════════
#  Parse both Shift Schedules into normalized row lists
# ══════════════════════════════════════════════════════════════════════════════
def parse_building(path):
    df = load_sheet(path, "Shift Schedule", BLD_HDR["Shift Schedule"],
                    required_cols=["Machine", "Date", "Shift", "SKUCode", "Qty", "CO_Mins", "CO_Type"],
                    tag="BLD")
    rows = []
    if df is None:
        return rows
    for _, r in df.iterrows():
        m = str(r.get("Machine", "")).strip()
        if not m or m.lower() in ("nan", "none") or "total" in m.lower() or "average" in m.lower():
            continue
        cotype = str(r.get("CO_Type", "")).strip()
        # Expired GT / carcass are WASTE display markers (Machine "—", CO_Mins 0), NOT
        # production or changeovers — skip entirely so they never count toward build/CO/
        # shift-minute rules (R5/R8/R10/R11/R18).
        if cotype in ("expired_GT", "expired_carcass"):
            continue
        qty = num(r.get("Qty"))
        co = num(r.get("CO_Mins"))
        rows.append({
            "date": dstr(r.get("Date")), "shift": str(r.get("Shift", "")).strip(),
            "machine": m, "sku": str(r.get("SKUCode", "")).strip(), "qty": qty,
            "co": co, "cotype": cotype, "group": mgroup(m),
            "_start": pd.to_datetime(r.get("StartTime"), errors="coerce"),
            "_end": pd.to_datetime(r.get("EndTime"), errors="coerce"),
            "is_carcass": cotype == "carcass",
            "is_co": (cotype in ("same_size_CO", "diff_size_CO")) or (co > 0 and qty == 0 and cotype != "carcass"),
        })
    for x in rows:
        x["is_prod"] = (not x["is_co"]) and (not x["is_carcass"]) and x["qty"] > 0
    return rows


def parse_curing(path):
    df = load_sheet(path, "Shift Schedule", CUR_HDR["Shift Schedule"],
                    required_cols=["Date", "Shift", "Machine", "SKUCode", "Qty", "CO_Mins",
                                   "Mould_Clean_Mins", "CycleTime_min"], tag="CUR")
    rows = []
    if df is None:
        return rows
    for _, r in df.iterrows():
        p = str(r.get("Machine", "")).strip()
        if not p or p.lower() in ("nan", "none") or "total" in p.lower() or "average" in p.lower():
            continue
        qty = num(r.get("Qty"))
        co = num(r.get("CO_Mins"))
        clean = num(r.get("Mould_Clean_Mins"))
        rows.append({
            "date": dstr(r.get("Date")), "shift": str(r.get("Shift", "")).strip(),
            "press": p, "sku": str(r.get("SKUCode", "")).strip(), "qty": qty,
            "co": co, "clean": clean, "ct_sheet": num(r.get("CycleTime_min"), np.nan),
            "remarks": str(r.get("Remarks", "")).strip(),
            "_start": pd.to_datetime(r.get("StartTime"), errors="coerce"),
            "_end": pd.to_datetime(r.get("EndTime"), errors="coerce"),
            "is_co": co > 0, "is_clean": clean > 0,
        })
    for x in rows:
        x["is_prod"] = (not x["is_co"]) and (not x["is_clean"]) and x["qty"] > 0
    return rows


# ── horizon (calendar-day aligned to plan_start; holiday-aware) ────────────────
def build_horizon(bld, cur):
    dates = sorted({r["date"] for r in bld} | {r["date"] for r in cur})
    ps = SRC["plan_start"].date() if hasattr(SRC["plan_start"], "date") else SRC["plan_start"]
    try:
        dts = [datetime.strptime(d, "%Y-%m-%d").date() for d in dates if re.match(r"\d{4}-\d\d-\d\d", d)]
    except Exception:  # noqa: BLE001
        dts = []
    if not dts and not ARGS.planning_days:
        return {"all_days": [], "planning_days": 0, "working_days": 0, "hol_idx": set(), "ps": ps}
    start = min([ps] + dts) if dts else ps
    # plan length = explicit --planning-days if given, else inferred from the date range.
    plan_len = ARGS.planning_days if ARGS.planning_days else ((max(dts) - start).days + 1)
    # all_days must cover BOTH the intended horizon AND any actual (spilled) dates so aging
    # FIFO still sees every lot; R7's denominator uses the intended plan length only.
    end = max((max(dts) if dts else start), start + timedelta(days=plan_len - 1))
    horizon = (end - start).days + 1
    all_days = [(start + timedelta(days=k)).strftime("%Y-%m-%d") for k in range(horizon)]
    hol = SRC.get("holidays") or set()
    hol_idx = {i for i, d in enumerate(all_days) if d in hol}
    working_days = plan_len - sum(1 for d in all_days[:plan_len] if d in hol)
    return {"all_days": all_days, "planning_days": plan_len,
            "working_days": working_days, "hol_idx": hol_idx, "ps": start}


# ══════════════════════════════════════════════════════════════════════════════
#  RULE CHECKS
# ══════════════════════════════════════════════════════════════════════════════
def r14_building_roster(bld):
    for r in bld:
        if r["machine"] not in ALL_38:
            add("R14", "building machine not in the 38-machine roster",
                r["date"], r["shift"], r["machine"], r["sku"], r["qty"],
                expected="one of 38", actual=r["machine"])
    rule_result("R14", "Only the 38 building machines are used", len(bld))


def r13_curing_roster(cur):
    p170 = SRC.get("presses170")
    if not p170:
        rule_result("R13", "Only the 170 allowable curing presses are used", 0,
                    skipped_reason="curing allowable press list not loaded")
        return
    seen = set()
    for r in cur:
        if r["press"] in seen:
            continue
        seen.add(r["press"])
        if r["press"] not in p170:
            add("R13", "curing press not in the 170 allowable set (even if in running-moulds)",
                r["date"], r["shift"], r["press"], r["sku"],
                expected="one of 170 allowable", actual=r["press"])
    rule_result("R13", "Only the 170 allowable curing presses are used", len(seen))


def r1_r18_building_ct(bld):
    ctmap = SRC.get("bld_ct")
    if ctmap is None:
        rule_result("R1B", "Building CT sourced from Cycle_time_Building.xlsx (no 0.94)", 0,
                    skipped_reason="building CT CSV not loaded")
        rule_result("R18B", "Building Production_Mins = Qty x CT_sec/60", 0,
                    skipped_reason="building CT CSV not loaded")
        return
    n = 0
    for r in bld:
        if not r["is_prod"] and not r["is_carcass"]:
            continue
        n += 1
        ct = ctmap.get((r["sku"], r["machine"]))
        if ct is None:
            add("R1B", "no building CT in Cycle_time_Building.xlsx for this (SKU,machine)",
                r["date"], r["shift"], r["machine"], r["sku"], r["qty"],
                expected="CSV CT (sec)", actual="MISSING")
            continue
        # R18 = CT<->Qty reconciliation: Qty x CT must match the row's actual time SPAN
        # (StartTime->EndTime), which may cross shift boundaries. Per-shift capacity is R11/R15.
        prod = r["qty"] * ct / 60.0
        dur = _dur_min(r["_start"], r["_end"])
        if dur is not None and abs(prod - dur) > max(2.0, 0.03 * prod):   # 2-min floor absorbs Qty-floor rounding
            add("R18B", "Qty x CT_sec/60 does not match the row's time span (CT/qty reconciliation)",
                r["date"], r["shift"], r["machine"], r["sku"], r["qty"], round(ct, 1),
                round(prod, 1), expected=f"span {round(dur, 1)} min", actual=round(prod, 1))
    rule_result("R1B", "Building CT sourced from Cycle_time_Building.xlsx (no 0.94)", n)
    rule_result("R18B", "Building Prod_Mins = Qty x CT_sec/60 matches time span", n)


def r1_r12_r18_curing_ct(cur):
    ctmap = SRC.get("cure_ct")
    if ctmap is None:
        for rr in ("R1C", "R12", "R18C"):
            rule_result(rr, "Curing CT (DB/0.94, default 17) & reconciliation", 0,
                        skipped_reason="curing CT map not loaded")
        return
    dflt = SRC["DEFAULT_CURE_CT"]
    cav = SRC["CAVITIES"]
    n = 0
    press_prodmin = defaultdict(float)   # sum (Qty/2) x own-CT per press  (R18C)
    press_span = defaultdict(float)      # sum production-row time span per press
    for r in cur:
        if not r["is_prod"]:
            continue
        n += 1
        db_ct = ctmap.get(r["sku"], dflt)
        sheet_ct = r["ct_sheet"]
        # R1C / R12: the SHEET CT must equal the DB-derived CT (or default 17 when DB missing).
        if not np.isnan(sheet_ct) and abs(sheet_ct - db_ct) > ARGS.ct_tol:
            miss = r["sku"] not in ctmap
            add("R12" if miss else "R1C",
                "curing CT mismatch vs DB" + (" (should default to 17)" if miss else " (DB/0.94)"),
                r["date"], r["shift"], r["press"], r["sku"], r["qty"], sheet_ct,
                expected=round(db_ct, 2), actual=round(sheet_ct, 2))
        # R18C accumulator — reconcile PER PRESS TOTAL (not per row). The even-Qty split credits
        # a boundary cycle to its COMPLETION shift, so a single row's Qty need not match its own
        # time span; but each press's TOTAL cured×CT must equal its TOTAL production time.
        ct_use = sheet_ct if not np.isnan(sheet_ct) else db_ct
        press_prodmin[r["press"]] += (r["qty"] / cav) * ct_use
        d = _dur_min(r["_start"], r["_end"])
        if d is not None:
            press_span[r["press"]] += d
    for p in press_prodmin:
        pm, sp = press_prodmin[p], press_span[p]
        if sp > 0 and abs(pm - sp) > max(5.0, 0.02 * sp):
            add("R18C", "press total (Qty/2)xCT does not match its total production time",
                mp=p, expected=f"total span {round(sp, 1)} min", actual=round(pm, 1))
    rule_result("R1C", "Curing CT = DB Cure Time / 0.94 matches sheet", n)
    rule_result("R12", "Missing curing CT falls back to default 17.0", n)
    rule_result("R18C", "Curing (Qty/2) x CT matches total production time PER PRESS", len(press_prodmin))


def r2_building_co(bld_path):
    """Building CO time = config per group/type, read from the Changeover Plan (ONE row per
    CO event with the full CO_Cost_Mins) — the Shift Schedule splits a boundary-crossing CO
    across two rows (60 → 40+20), so per-row checks there mis-fire."""
    same, diff = SRC["CO_SAME"], SRC["CO_DIFF"]
    df = load_sheet(bld_path, "Changeover Plan", BLD_HDR["Changeover Plan"],
                    required_cols=["Machine", "CO_Type", "CO_Cost_Mins"], tag="R2B")
    if df is None:
        rule_result("R2B", "Building CO time = config per group/type", 0,
                    skipped_reason="building Changeover Plan sheet missing")
        return
    n = 0
    for _, r in df.iterrows():
        m = str(r.get("Machine", "")).strip()
        if not m or "total" in m.lower():
            continue
        cotype = str(r.get("CO_Type", "")).strip()
        if cotype not in ("same_size_CO", "diff_size_CO"):
            continue
        n += 1
        g = mgroup(m)
        exp = same.get(g) if cotype == "same_size_CO" else diff.get(g)
        if exp is None:
            continue
        co = num(r.get("CO_Cost_Mins"))
        if abs(co - exp) > 0.5 and abs(co - (exp + 10)) > 0.5:   # +10 DB changeover variant tolerated
            add("R2B", f"building {cotype} time != config for group {g}",
                dstr(r.get("Date")), "", m, str(r.get("Target_SKU", "")).strip(), co=co,
                expected=f"{exp} (or {exp+10})", actual=co)
    rule_result("R2B", "Building CO time = config per group/type (per event)", n)


def r2_r10_curing_co(cur_path):
    """R2C (CO time = clean time = 480) and R10 (curing COs/day <= cap) from the curing
    Changeover Plan sheet — the authoritative ONE-ROW-PER-EVENT list. The Shift Schedule
    splits a mid-shift CO/clean across two rows (e.g. 38 + 442 = 480), so per-row checks
    there mis-fire; the Changeover Plan carries the full 480 per event with a CO_Type."""
    co_t, clean_t = SRC["CURING_CO"], SRC["CLEAN_MINS"]
    cap = SRC["MAX_CO_DAY"]
    df = load_sheet(cur_path, "Changeover Plan", CUR_HDR["Changeover Plan"],
                    required_cols=["Date", "Press", "CO_Type", "Mins"], tag="R2C/R10")
    if df is None:
        rule_result("R2C", "Curing CO time = mould clean time = 480", 0,
                    skipped_reason="curing Changeover Plan sheet missing")
        rule_result("R10", f"Curing CO per day <= {cap}", 0,
                    skipped_reason="curing Changeover Plan sheet missing")
        return
    n = 0
    co_by_day = defaultdict(int)
    KNOWN = {"Planned", "Dynamic", "Cold-Start", "Mould Clean"}
    for _, r in df.iterrows():
        cotype = str(r.get("CO_Type", "")).strip()
        if cotype not in KNOWN:          # skip footer / summary / blank rows
            continue
        press = str(r.get("Press", "")).strip()
        if not press or "total" in press.lower():
            continue
        mins = num(r.get("Mins"))
        d = dstr(r.get("Date"))
        sh = str(r.get("Shift", "")).strip()
        n += 1
        # R2C: every CO and every mould clean event must be exactly 480 minutes
        exp = clean_t if cotype.lower().startswith("mould") else co_t
        if abs(mins - exp) > 0.5:
            add("R2C", f"curing {cotype} event time != {exp}", d, sh, press,
                str(r.get("Target_SKU", "")).strip(), co=mins if not cotype.lower().startswith("mould") else "",
                clean=mins if cotype.lower().startswith("mould") else "", expected=exp, actual=mins)
        # R10: count demand-driven curing COs per day (Planned + Dynamic). Mould cleans are not
        # COs; Day-0/1 Cold-Starts are one-time press activations EXEMPT from the daily CO cap.
        if cotype in ("Planned", "Dynamic"):
            co_by_day[d] += 1
    rule_result("R2C", "Curing CO time = mould clean time = 480 (per event)", n)
    for d, c in sorted(co_by_day.items()):
        if c > cap:
            add("R10", "demand-driven curing changeovers in a day exceed the cap",
                d, expected=f"<={cap}", actual=c)
    rule_result("R10", f"Curing CO (Planned+Dynamic) per day <= {cap}", len(co_by_day))


def r3_building_allow(bld):
    allow = SRC.get("bld_allow")
    if not allow:
        rule_result("R3B", "Building (SKU,machine) in allowable matrix", 0,
                    skipped_reason="building allowable matrix not loaded")
        return
    n = 0
    for r in bld:
        # Only real GT production rows. CO rows carry SKUCode="CHANGEOVER" (a sentinel, not a
        # SKU); Stage-1 carcass rows are upstream of the GT SKU and exempt (agreed R3 scope).
        if not r["is_prod"] or r["is_carcass"]:
            continue
        if not r["sku"] or r["sku"].upper() == "CHANGEOVER":
            continue
        n += 1
        elig = allow.get(r["sku"])
        if elig is None:
            add("R3B", "SKU absent from building allowable matrix", r["date"], r["shift"],
                r["machine"], r["sku"], r["qty"], expected="in matrix", actual="SKU missing")
        elif r["machine"] not in elig:
            add("R3B", "machine not allowable for this SKU", r["date"], r["shift"],
                r["machine"], r["sku"], r["qty"], expected=f"{sorted(elig)[:6]}...", actual=r["machine"])
    rule_result("R3B", "Building (SKU,machine) in allowable matrix", n)


def r3_curing_allow(cur):
    allow = SRC.get("cure_allow")
    if not allow:
        rule_result("R3C", "Curing (SKU,press) in allowable matrix", 0,
                    skipped_reason="curing allowable matrix not loaded")
        return
    # Exception: a (press,SKU) that is in Daily_Running_Moulds (Day-0 mounted) but not in the
    # allowable matrix is NOT a violation — the press legitimately continues its mounted SKU.
    rm_pairs = (SRC.get("rm") or {}).get("pairs") or set()
    n = 0
    for r in cur:
        if not r["is_prod"] or not r["sku"]:
            continue
        n += 1
        if (r["press"], r["sku"]) in rm_pairs:
            continue   # Day-0 running-moulds exception
        elig = allow.get(r["sku"])
        if elig is None:
            add("R3C", "SKU absent from curing allowable matrix (and not in Daily_Running_Moulds)",
                r["date"], r["shift"], r["press"], r["sku"], r["qty"], expected="in matrix", actual="SKU missing")
        elif r["press"] not in elig:
            add("R3C", "press not allowable for this SKU (and not in Daily_Running_Moulds)",
                r["date"], r["shift"], r["press"], r["sku"], r["qty"], expected="allowable press", actual=r["press"])
    rule_result("R3C", "Curing (SKU,press) allowable, or Day-0 running-moulds exception", n)


def r4_inch(bld):
    inch_sets = (SRC.get("bld") or {}).get("inch_set") or {}
    if not inch_sets:
        rule_result("R4", "Historical inch respected (building)", 0,
                    skipped_reason="historical inch sets not loaded")
        return
    n = 0
    for r in bld:
        if not r["is_prod"] or r["machine"] in S1_MACHINES:
            continue
        allowed = inch_sets.get(r["machine"])
        if not allowed:
            continue
        n += 1
        inch = sku_inch(r["sku"])
        if inch and inch not in allowed:
            add("R4", "SKU inch not in machine's historical allowed-inch set",
                r["date"], r["shift"], r["machine"], r["sku"], r["qty"],
                expected=f"{sorted(allowed)}", actual=inch)
    rule_result("R4", "Historical inch respected (building GT)", n)


def r19_no_prod_during_co(bld, cur):
    nb = nc = 0
    for r in bld:
        if r["is_co"] and r["qty"] > 0:
            nb += 1
            add("R19B", "production during a building CO row (Qty>0 with CO)",
                r["date"], r["shift"], r["machine"], r["sku"], r["qty"], co=r["co"],
                expected="Qty=0", actual=r["qty"])
    for r in cur:
        if (r["is_co"] or r["is_clean"]) and r["qty"] > 0:
            nc += 1
            add("R19C", "production during a curing CO/clean row (Qty>0)",
                r["date"], r["shift"], r["press"], r["sku"], r["qty"], co=r["co"],
                clean=r["clean"], expected="Qty=0", actual=r["qty"])
    rule_result("R19B", "No production during a building CO", len(bld))
    rule_result("R19C", "No production during a curing CO/clean", len(cur))


def r16_one_sku(bld, cur):
    """True SIMULTANEITY via time-interval overlap. Now that rows carry real
    StartTime/EndTime, two DIFFERENT-SKU production rows on the same physical unit that
    overlap in time = producing two SKUs at once = FAIL. Sequential same-machine SKUs
    (A finishes, then B starts) do NOT overlap and are legal."""
    def _overlaps(rows, label, rid, mp_id):
        prod = sorted([r for r in rows if r["is_prod"] and r["_start"] is not None
                       and not pd.isna(r["_start"]) and r["_end"] is not None and not pd.isna(r["_end"])],
                      key=lambda r: r["_start"])
        n = 0
        for i in range(len(prod) - 1):
            a, b = prod[i], prod[i + 1]
            if a["sku"] != b["sku"] and a["_end"] > b["_start"] + timedelta(minutes=1):
                n += 1
                add(rid, f"{label} produces two SKUs simultaneously (time intervals overlap)",
                    a["date"], a["shift"], mp_id,
                    f'{a["sku"]} | {b["sku"]}',
                    expected="no overlap",
                    actual=f'{a["_end"]:%m-%d %H:%M} vs {b["_start"]:%m-%d %H:%M}')
        return n
    # group building rows by machine, curing by press
    by_m = defaultdict(list)
    for r in bld:
        by_m[r["machine"]].append(r)
    for m, rows in by_m.items():
        _overlaps(rows, "building machine", "R16B", m)
    by_p = defaultdict(list)
    for r in cur:
        by_p[r["press"]].append(r)
    for p, rows in by_p.items():
        _overlaps(rows, "curing press", "R16C", p)
    rule_result("R16B", "No building machine produces 2 SKUs simultaneously (interval overlap)", len(by_m))
    rule_result("R16C", "No curing press produces 2 SKUs simultaneously (interval overlap)", len(by_p))


def r11_r15_capacity(bld, cur):
    """SHIFT-CLIPPED capacity. A row's [StartTime,EndTime] is clipped to each real shift
    window it overlaps, and each shift is charged only its own minutes (prod / CO / clean).
    A continuous run that crosses a shift boundary therefore no longer looks like >480 in
    one shift. Occupancy per real shift = Prod + CO + Clean minutes must be <= 480."""
    smin = SRC["SHIFT_MINS"]
    # Building: kind = co (changeover) else prod (production or Stage-1 carcass build)
    bagg = defaultdict(lambda: {"prod": 0.0, "co": 0.0, "clean": 0.0})
    for r in bld:
        kind = "co" if r["is_co"] else "prod"
        for (d, s, mins) in _shift_segments(r["_start"], r["_end"]):
            bagg[(r["machine"], d, s)][kind] += mins
    for (m, d, s), a in bagg.items():
        tot = a["prod"] + a["co"] + a["clean"]
        if tot > smin + 1.0:
            for rr in ("R11B", "R15B"):
                add(rr, "building busy-minutes in a REAL shift exceed 480", d, s, m,
                    prod=round(a["prod"], 1), co=round(a["co"], 1), clean=round(a["clean"], 1),
                    expected=f"<={smin}", actual=round(tot, 1))
    # Curing: kind = co / clean / prod
    cagg = defaultdict(lambda: {"prod": 0.0, "co": 0.0, "clean": 0.0})
    for r in cur:
        kind = "co" if r["is_co"] else ("clean" if r["is_clean"] else "prod")
        for (d, s, mins) in _shift_segments(r["_start"], r["_end"]):
            cagg[(r["press"], d, s)][kind] += mins
    for (p, d, s), a in cagg.items():
        tot11 = a["prod"] + a["co"]
        tot15 = a["prod"] + a["co"] + a["clean"]
        if tot11 > smin + 1.0:
            add("R11C", "curing Prod+CO in a REAL shift exceed 480", d, s, p, prod=round(a["prod"], 1),
                co=round(a["co"], 1), clean=round(a["clean"], 1), expected=f"<={smin}", actual=round(tot11, 1))
        if tot15 > smin + 1.0:
            add("R15C", "curing Prod+CO+Clean in a REAL shift exceed 480", d, s, p, prod=round(a["prod"], 1),
                co=round(a["co"], 1), clean=round(a["clean"], 1), expected=f"<={smin}", actual=round(tot15, 1))
    rule_result("R11B", "Building Prod+CO <= 480 per REAL shift (clipped)", len(bagg))
    rule_result("R11C", "Curing Prod+CO <= 480 per REAL shift (clipped)", len(cagg))
    rule_result("R15B", "Building Prod+CO+Clean <= 480 per REAL shift (clipped)", len(bagg))
    rule_result("R15C", "Curing Prod+CO+Clean <= 480 per REAL shift (clipped)", len(cagg))


def r7_utilization(bld_path, cur_path):
    smin = SRC["SHIFT_MINS"]
    # Robust denominator: every machine must share the SAME Available (working_days x 3 x 480).
    # We check occupancy <= 100 and Available CONSISTENCY (all equal to the sheet's own modal
    # value) rather than asserting a recomputed working-day count — the latter is fragile to a
    # tiny last-day time spill inflating an inferred horizon.
    for tag, path, sheet, hdr in [("R7B", bld_path, "Machine Utilization", BLD_HDR["Machine Utilization"]),
                                  ("R7C", cur_path, "Machine Utilization", CUR_HDR["Machine Utilization"])]:
        df = load_sheet(path, sheet, hdr, tag=tag)
        if df is None:
            rule_result(tag, "Occupancy <= 100% + Available consistent", 0,
                        skipped_reason="Machine Utilization sheet missing")
            continue
        occ_col = next((c for c in df.columns if "Occupancy" in c), None)
        av_col = next((c for c in df.columns if "Available" in c), None)
        mcol = df.columns[0]
        rows = []
        for _, r in df.iterrows():
            m = str(r.get(mcol, "")).strip()
            if not m or m.lower() in ("nan", "none") or "total" in m.lower() or "average" in m.lower():
                continue
            rows.append((m, num(r.get(occ_col)) if occ_col else None,
                         num(r.get(av_col)) if av_col else None))
        # modal Available = the consistent expected value
        avs = [a for _, _, a in rows if a]
        modal_av = max(set(avs), key=avs.count) if avs else None
        for m, occ, av in rows:
            if occ is not None and occ > 100.0 + 0.5:
                add(tag, "occupancy exceeds 100%", mp=m, expected="<=100%", actual=f"{occ:.1f}%")
            if av is not None and modal_av is not None and abs(av - modal_av) > 1.0:
                add(tag, "Available_Mins inconsistent across machines", mp=m,
                    expected=modal_av, actual=av)
        rule_result(tag, "Occupancy <= 100% + Available consistent", len(rows))


def r8_demand_cap(bld, cur):
    dem = SRC.get("demand")
    if not dem:
        for rr in ("R8B", "R8C"):
            rule_result(rr, "Production <= demand per SKU", 0, skipped_reason="demand not loaded")
        return
    tol = ARGS.qty_tol
    built = defaultdict(float)
    for r in bld:
        if r["is_prod"]:
            built[r["sku"]] += r["qty"]
    for sku, q in built.items():
        d = dem.get(sku)
        if d is not None and q > d + tol:
            add("R8B", "GT built exceeds demand", sku=sku, qty=round(q, 1),
                expected=f"<={d:.0f}", actual=round(q, 1))
    cured = defaultdict(float)
    for r in cur:
        if r["is_prod"]:
            cured[r["sku"]] += r["qty"]
    for sku, q in cured.items():
        d = dem.get(sku)
        if d is not None and q > d + tol:
            add("R8C", "cured qty exceeds demand", sku=sku, qty=round(q, 1),
                expected=f"<={d:.0f}", actual=round(q, 1))
    rule_result("R8B", "GT built <= demand per SKU", len(built))
    rule_result("R8C", "Cured <= demand per SKU", len(cured))


def r6_opening(cur_path):
    og = SRC.get("open_gt")
    if og is None:
        rule_result("R6", "Opening GT from DB per plan_month matches sheet", 0,
                    skipped_reason="opening GT not loaded")
        return
    df = load_sheet(cur_path, "Demand Fulfillment", CUR_HDR["Demand Fulfillment"], tag="R6")
    if df is None:
        rule_result("R6", "Opening GT from DB per plan_month matches sheet", 0,
                    skipped_reason="curing Demand Fulfillment sheet missing")
        return
    scol = next((c for c in df.columns if c in ("SKUCode", "SKU Code")), df.columns[0])
    gcol = next((c for c in df.columns if "GT_Inventory" in c or c == "GT_Inventory"), None)
    if gcol is None:
        rule_result("R6", "Opening GT from DB per plan_month matches sheet", 0,
                    skipped_reason="GT_Inventory column missing in curing Demand Fulfillment")
        return
    n = 0
    for _, r in df.iterrows():
        sku = str(r.get(scol, "")).strip()
        if not sku or sku.lower() in ("nan", "total"):
            continue
        n += 1
        sheet_gt = num(r.get(gcol))
        db_gt = og.get(sku, 0.0)
        if abs(sheet_gt - db_gt) > 1.0:
            add("R6", "opening GT_Inventory != DB value for plan_month", sku=sku,
                expected=round(db_gt, 1), actual=round(sheet_gt, 1))
    rule_result("R6", "Opening GT from DB per plan_month matches sheet", n)


def _fifo_expiry(built_by_day, consumed_by_day, opening, all_days, shelf, opening_lots=None):
    lot = deque()
    if opening_lots:                                 # MID-MONTH: aged opening lots (age_days, qty)
        for _age, _q in opening_lots:                # seed each at index (−age) so its remaining
            if _q > 0:                               # shelf equals shelf−age (day-start carry).
                lot.append((-int(_age), float(_q)))
    elif opening and opening > 0:
        lot.append((0, float(opening)))
    exp = []
    for i, d in enumerate(all_days):
        b = built_by_day.get(d, 0.0)
        if b > 0:
            lot.append((i, float(b)))
        c = consumed_by_day.get(d, 0.0)
        while c > 1e-6 and lot:
            ld, lq = lot[0]
            take = min(lq, c)
            lq -= take
            c -= take
            if lq <= 1e-6:
                lot.popleft()
            else:
                lot[0] = (ld, lq)
        keep = deque()
        for ld, lq in lot:
            if (i - ld) > shelf and lq > 1e-6:
                _bd = all_days[ld] if ld >= 0 else "opening"   # aged opening lot has ld<0
                exp.append((_bd, all_days[i], lq))
            else:
                keep.append((ld, lq))
        lot = keep
    return exp


def r9_aging(bld, cur, hz):
    all_days = hz["all_days"]
    if not all_days:
        for rr in ("R9G", "R9C"):
            rule_result(rr, "FIFO aging", 0, skipped_reason="no dated rows to build a horizon")
        return
    og = SRC.get("open_gt") or {}
    oc = SRC.get("open_carc") or {}
    _mm_gt = SRC.get("mm_gt_lots")      # MID-MONTH: aged GT opening lots (else None → DB scalar)
    _mm_carc = SRC.get("mm_carc")       # MID-MONTH: carcass opening totals (else None → DB scalar)
    # GT: built by GT machines (production) vs cured
    gt_built = defaultdict(lambda: defaultdict(float))
    gt_cured = defaultdict(lambda: defaultdict(float))
    carc_built = defaultdict(lambda: defaultdict(float))
    carc_used = defaultdict(lambda: defaultdict(float))
    for r in bld:
        if r["is_prod"]:
            gt_built[r["sku"]][r["date"]] += r["qty"]
            if r["machine"] in S2_MACHINES:
                carc_used[r["sku"]][r["date"]] += r["qty"]   # Stage-2 GT consumes carcass
        if r["is_carcass"]:
            carc_built[r["sku"]][r["date"]] += r["qty"]
    for r in cur:
        if r["is_prod"]:
            gt_cured[r["sku"]][r["date"]] += r["qty"]
    # GT aging = 3 days
    gt_skus = set(gt_built) | set(gt_cured) | set(og) | set(_mm_gt or {})
    for sku in gt_skus:
        exp = _fifo_expiry(gt_built[sku], gt_cured[sku], og.get(sku, 0.0), all_days,
                           SRC["GT_AGE_DAYS"], opening_lots=(_mm_gt or {}).get(sku) if _mm_gt else None)
        for built_d, exp_d, q in exp:
            add("R9G", f"GT expired: built {built_d}, not cured within {SRC['GT_AGE_DAYS']} days",
                date=exp_d, sku=sku, qty=round(q, 1), expected=f"cure by +{SRC['GT_AGE_DAYS']}d",
                actual=f"expired {exp_d}")
    rule_result("R9G", "GT aging <= 3 days (strict FIFO per SKU)", len(gt_skus))
    # Carcass aging = 1 day. Mid-month: seed the carried carcass total as opening (age-0; the 1-day
    # shelf makes lot-age immaterial vs the DB scalar it replaces).
    _oc_use = _mm_carc if _mm_carc is not None else oc
    carc_skus = set(carc_built) | set(carc_used) | set(_oc_use)
    for sku in carc_skus:
        exp = _fifo_expiry(carc_built[sku], carc_used[sku], _oc_use.get(sku, 0.0), all_days, SRC["CARC_AGE_DAYS"])
        for built_d, exp_d, q in exp:
            add("R9C", f"carcass expired: built {built_d}, not used within {SRC['CARC_AGE_DAYS']} day",
                date=exp_d, sku=sku, qty=round(q, 1), expected=f"use by +{SRC['CARC_AGE_DAYS']}d",
                actual=f"expired {exp_d}")
    rule_result("R9C", "Carcass aging <= 1 day (strict FIFO per SKU)", len(carc_skus))


def r5_stage2_vs_stage1(bld, hz):
    all_days = hz["all_days"]
    if not all_days:
        rule_result("R5", "Stage-2 <= available Stage-1 carcass", 0, skipped_reason="no horizon")
        return
    oc = SRC.get("mm_carc")                 # MID-MONTH day-start carcass, else DB 1st-of-month
    oc = oc if oc is not None else (SRC.get("open_carc") or {})
    carc_built = defaultdict(lambda: defaultdict(float))
    s2 = defaultdict(lambda: defaultdict(float))
    for r in bld:
        if r["is_carcass"]:
            carc_built[r["sku"]][r["date"]] += r["qty"]
        if r["is_prod"] and r["machine"] in S2_MACHINES:
            s2[r["sku"]][r["date"]] += r["qty"]
    skus = set(carc_built) | set(s2)
    n = 0
    for sku in skus:
        n += 1
        cum_carc = float(oc.get(sku, 0.0))
        cum_s2 = 0.0
        for d in all_days:
            cum_carc += carc_built[sku].get(d, 0.0)
            cum_s2 += s2[sku].get(d, 0.0)
            if cum_s2 > cum_carc + 1.0:
                add("R5", "Stage-2 GT exceeds cumulative Stage-1 carcass available",
                    date=d, sku=sku, qty=round(s2[sku].get(d, 0.0), 1),
                    expected=f"cum carcass {cum_carc:.0f}", actual=f"cum S2 {cum_s2:.0f}")
                break
    rule_result("R5", "Stage-2 GT <= available Stage-1 carcass (1-day)", n)


def r17_mould(cur, cur_path):
    mould = SRC.get("mould")
    if not mould:
        rule_result("R17", "Mould feasibility (bipartite per shift)", 0,
                    skipped_reason="mould eligibility not loaded")
        return
    try:
        from scipy.sparse import csr_matrix
        from scipy.sparse.csgraph import maximum_bipartite_matching
    except Exception as e:  # noqa: BLE001
        rule_result("R17", "Mould feasibility (bipartite per shift)", 0,
                    skipped_reason=f"scipy unavailable: {e}")
        return
    sku_moulds = mould["sku_moulds"]
    # fold Day-0 mounted moulds into eligibility (a mounted mould is usable for its SKU)
    rm = SRC.get("rm")
    if rm and isinstance(rm.get("df"), pd.DataFrame):
        df0 = rm["df"]
        scol = next((c for c in df0.columns if "SKU" in c or "Sapcode" in c or "Matl" in c), None)
        mcol = next((c for c in df0.columns if "Mould" in c or "Mold" in c), None)
        if scol and mcol:
            for _, r in df0.iterrows():
                s = str(r[scol]).strip()
                mo = str(r[mcol]).strip()
                if s and mo and mo.lower() != "nan":
                    sku_moulds.setdefault(s, set()).add(mo)
    # structural: SKU run with <2 eligible moulds
    struct_bad = set()
    for r in cur:
        if r["is_prod"]:
            if len(sku_moulds.get(r["sku"], set())) < 2:
                if r["sku"] not in struct_bad:
                    struct_bad.add(r["sku"])
                    add("R17", "SKU run on a press but has <2 eligible moulds (structural)",
                        r["date"], r["shift"], r["press"], r["sku"], r["qty"],
                        expected=">=2 eligible moulds", actual=len(sku_moulds.get(r["sku"], set())))
    # per-shift exact bipartite feasibility (2 slots/press)
    shifts = defaultdict(list)   # (date,shift) -> list of (press,sku)
    for r in cur:
        if r["is_prod"] and r["qty"] > 0:
            shifts[(r["date"], r["shift"])].append((r["press"], r["sku"]))
    n_infeasible = 0
    for (d, s), pairs in sorted(shifts.items()):
        slots = []
        for p, sku in pairs:
            slots.append((p, sku)); slots.append((p, sku))   # 2 moulds/press
        moulds = sorted({m for _, sku in pairs for m in sku_moulds.get(sku, set())})
        if not moulds:
            continue
        midx = {m: i for i, m in enumerate(moulds)}
        rows, cols = [], []
        for si, (p, sku) in enumerate(slots):
            for m in sku_moulds.get(sku, set()):
                rows.append(si); cols.append(midx[m])
        if not rows:
            matched = 0
        else:
            g = csr_matrix((np.ones(len(rows)), (rows, cols)), shape=(len(slots), len(moulds)))
            match = maximum_bipartite_matching(g, perm_type="column")
            matched = int((match >= 0).sum())
        if matched < len(slots):
            n_infeasible += 1
            add("R17", "shift not mould-feasible (a disjoint 2-moulds/press assignment does not exist)",
                date=d, shift=s, mp=f"{len(pairs)} presses",
                expected=f"{len(slots)} mould-slots fillable", actual=f"{matched} fillable")
    # actual mould-id conflict from Mould Movement / Tracker sheets (if populated)
    _mould_id_conflict_check(cur_path)
    rule_result("R17", "Mould feasibility (bipartite per shift) + structural + id-conflict",
                len(shifts))


def _mould_id_conflict_check(cur_path):
    for sheet in ("Mould Movement", "Mould Tracker", "MouldInUse"):
        df = load_sheet(cur_path, sheet, CUR_HDR.get(sheet, 1), tag="R17")
        if df is None or df.empty:
            continue
        # need per-(date,shift) mould->press; only if the sheet carries those columns
        mcol = next((c for c in df.columns if "Mould" in c or "Mold" in c), None)
        pcol = next((c for c in df.columns if c in ("Machine", "Press", "Assigned_Machine")), None)
        dcol = next((c for c in df.columns if c == "Date"), None)
        scol = next((c for c in df.columns if c == "Shift"), None)
        if not (mcol and pcol and dcol and scol):
            _issue("mould-id-sheet-not-usable",
                   f"{sheet}: cannot cross-check mould-ids (need Mould/Press/Date/Shift; have {list(df.columns)})")
            continue
        seen = defaultdict(set)
        for _, r in df.iterrows():
            key = (dstr(r.get(dcol)), str(r.get(scol)).strip(), str(r.get(mcol)).strip())
            p = str(r.get(pcol)).strip()
            if not str(r.get(mcol)).strip():
                continue
            seen[key].add(p)
        for (d, s, m), presses in seen.items():
            if len(presses) > 1:
                add("R17", f"same mould-id used by >1 press in one shift (from {sheet})",
                    date=d, shift=s, mp=",".join(sorted(presses)),
                    expected="1 press/mould/shift", actual=f"mould {m} on {len(presses)} presses")
        return  # first usable sheet wins


# ══════════════════════════════════════════════════════════════════════════════
#  Excel report
# ══════════════════════════════════════════════════════════════════════════════
def write_report(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    skip_fill = PatternFill("solid", fgColor="FFEB9C")

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["JK Tyre BTP — Building + Curing Feasibility Validation"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append([f"Building : {ARGS.building}"])
    ws.append([f"Curing   : {ARGS.curing}"])
    ws.append([f"Demand   : {SRC.get('demand_path', ARGS.demand)}"])
    ws.append([f"Plan month : {os.environ.get('PLAN_MONTH', '')}   |   Holidays: {sorted(SRC.get('holidays') or [])}"])
    tot_v = sum(v["violations"] for v in RULE_STATUS.values())
    n_fail = sum(1 for v in RULE_STATUS.values() if v["status"] == "FAIL")
    n_pass = sum(1 for v in RULE_STATUS.values() if v["status"] == "PASS")
    n_skip = sum(1 for v in RULE_STATUS.values() if v["status"] == "SKIP")
    ws.append([f"RESULT: {n_pass} PASS | {n_fail} FAIL | {n_skip} SKIP | {tot_v} total violations"])
    ws["A7"].font = Font(bold=True, size=11)
    ws.append([])
    head_row = ws.max_row + 1
    head = ["Rule", "Status", "Violations", "Rows Checked", "Description", "Skipped Reason"]
    ws.append(head)
    style_header(ws, head_row, len(head))
    for rid in sorted(RULE_STATUS.keys(), key=_rule_sort):
        v = RULE_STATUS[rid]
        ws.append([rid, v["status"], v["violations"], v["checked"], v["title"], v["skipped_reason"]])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.fill = {"PASS": pass_fill, "FAIL": fail_fill, "SKIP": skip_fill}[v["status"]]
        cell.font = Font(bold=True)
    for col, w in zip("ABCDEF", [8, 8, 11, 13, 60, 45]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{head_row + 1}"

    # ── Data issues ──
    wi = wb.create_sheet("Data Issues")
    wi.append(["Kind", "Detail"])
    style_header(wi, 1, 2)
    if DATA_ISSUES:
        for it in DATA_ISSUES:
            wi.append([it["kind"], it["detail"]])
    else:
        wi.append(["none", "no missing sheets/columns/data or ambiguous cases"])
    wi.column_dimensions["A"].width = 26
    wi.column_dimensions["B"].width = 100

    # ── All violations ──
    wa = wb.create_sheet("All Violations")
    wa.append(COLUMNS)
    style_header(wa, 1, len(COLUMNS))
    for f in FINDINGS:
        wa.append([f[c] for c in COLUMNS])
    for i, c in enumerate(COLUMNS, 1):
        wa.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(c) + 6))
    if FINDINGS:
        wa.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{wa.max_row}"
    wa.freeze_panes = "A2"

    # ── one detail sheet per FAILED rule ──
    for rid in sorted(RULE_STATUS.keys(), key=_rule_sort):
        if RULE_STATUS[rid]["status"] != "FAIL":
            continue
        rows = [f for f in FINDINGS if f["Rule"] == rid]
        wd = wb.create_sheet(f"{rid} ({len(rows)})"[:31])
        wd.append([f"{rid} — {RULE_STATUS[rid]['title']}  |  {len(rows)} violation(s)"])
        wd["A1"].font = Font(bold=True, size=11)
        wd.append(COLUMNS)
        style_header(wd, 2, len(COLUMNS))
        for f in rows:
            wd.append([f[c] for c in COLUMNS])
        for i, c in enumerate(COLUMNS, 1):
            wd.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(c) + 6))
        wd.freeze_panes = "A3"

    wb.save(path)


def _rule_sort(rid):
    m = re.match(r"R(\d+)([A-Z]?)", rid)
    return (int(m.group(1)), m.group(2)) if m else (999, rid)


# ══════════════════════════════════════════════════════════════════════════════
#  main
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 78)
    print("  FEASIBILITY VALIDATION — Building + Curing")
    print("=" * 78)
    load_sources()
    bld = parse_building(ARGS.building)
    cur = parse_curing(ARGS.curing)
    print(f"  Building rows: {len(bld):>6}  |  Curing rows: {len(cur):>6}")
    hz = build_horizon(bld, cur)
    print(f"  Horizon: {hz['planning_days']} calendar days | {hz['working_days']} working "
          f"| holidays {sorted(SRC.get('holidays') or [])}")

    # run every rule
    r14_building_roster(bld)
    r13_curing_roster(cur)
    r1_r18_building_ct(bld)
    r1_r12_r18_curing_ct(cur)
    r2_building_co(ARGS.building)
    r2_r10_curing_co(ARGS.curing)
    r3_building_allow(bld)
    r3_curing_allow(cur)
    r4_inch(bld)
    r5_stage2_vs_stage1(bld, hz)
    r6_opening(ARGS.curing)
    r7_utilization(ARGS.building, ARGS.curing)
    r8_demand_cap(bld, cur)
    r9_aging(bld, cur, hz)
    r11_r15_capacity(bld, cur)
    r16_one_sku(bld, cur)
    r17_mould(cur, ARGS.curing)
    r19_no_prod_during_co(bld, cur)

    # terminal report
    print("\n  RULE                                                          STATUS  VIOL  CHECKED")
    print("  " + "-" * 84)
    for rid in sorted(RULE_STATUS.keys(), key=_rule_sort):
        v = RULE_STATUS[rid]
        print(f"  {rid:<5} {v['title'][:52]:<52}  {v['status']:<6} {v['violations']:>5} {v['checked']:>8}")
    n_fail = sum(1 for v in RULE_STATUS.values() if v["status"] == "FAIL")
    n_pass = sum(1 for v in RULE_STATUS.values() if v["status"] == "PASS")
    n_skip = sum(1 for v in RULE_STATUS.values() if v["status"] == "SKIP")
    tot_v = sum(v["violations"] for v in RULE_STATUS.values())
    print("  " + "-" * 84)
    print(f"  TOTAL: {n_pass} PASS | {n_fail} FAIL | {n_skip} SKIP | {tot_v} violations")
    if DATA_ISSUES:
        print(f"\n  DATA ISSUES ({len(DATA_ISSUES)}) — not silently ignored:")
        for it in DATA_ISSUES[:20]:
            print(f"    [{it['kind']}] {it['detail'][:110]}")

    # sample violations per failing rule
    for rid in sorted(RULE_STATUS.keys(), key=_rule_sort):
        if RULE_STATUS[rid]["status"] != "FAIL":
            continue
        rows = [f for f in FINDINGS if f["Rule"] == rid]
        print(f"\n  ── {rid} — {RULE_STATUS[rid]['title']}  [{len(rows)} violation(s)] ──")
        for f in rows[:ARGS.max_rows]:
            print(f"     date={f['Date']} shift={f['Shift']} m/p={f['Machine/Press']} "
                  f"sku={f['SKU']} qty={f['Qty']} ct={f['CT']} prod={f['Prod_Mins']} "
                  f"co={f['CO_Mins']} clean={f['Clean_Mins']} exp={f['Expected']} act={f['Actual']}")
        if len(rows) > ARGS.max_rows:
            print(f"     … and {len(rows) - ARGS.max_rows} more (see {ARGS.out})")

    write_report(ARGS.out)
    print(f"\n  Excel report → {ARGS.out}")
    sys.exit(1 if n_fail else 0)


if __name__ == "__main__":
    main()
