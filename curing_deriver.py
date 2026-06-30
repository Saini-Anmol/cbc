"""
B2C Pipeline — Phase 2: Curing Schedule Deriver
================================================
Derives the curing schedule 100% deterministically from the building output.
No LP, no GA — pure rolling GT balance check.

Algorithm (per SKU, rolling across all planning shifts):
    GT_balance(SKU, S=0) = opening_GT_inventory(SKU)
    For each shift S in chronological order:
        GT_balance += building_output(SKU, S-1)  ← STRICT: prior-shift GT only
        cured_qty   = min(GT_balance, consumption_needed)
        GT_balance -= cured_qty
        record (Press_ID, Date, Shift, SKU_Code, Status, Qty_Produced, GT_Source,
                GT_Consumed, GT_Inventory_Remaining, Active_Press_Count)

Hard constraint satisfied by design: curing presses never run more than the
GT available in the balance tracker. If GT_balance < consumption_needed, the
press produces at reduced output (proportional to available GT).

Standalone usage:
    python curing_deriver.py

Inputs:
    data/main_output/bc_building_schedule.xlsx   (Phase 1 output)
    data/output/curing_consumption_table.xlsx    (Phase 0 output)
    DB: gt_inventory_manual                       (opening GT inventory)

Output:
    data/main_output/bc_curing_schedule.xlsx
"""

from __future__ import annotations

import math
import os
import sys
import warnings
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

import cbc_env

HERE    = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = cbc_env.OUTPUT_DIR
MAIN_OUT = os.path.join(OUT_DIR, "main_output")
os.makedirs(MAIN_OUT, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

class DeriverConfig:
    SHIFT_MINS      = 480
    SHIFTS_PER_DAY  = 3
    HOURS_PER_SHIFT = 8
    SHIFT_START_H   = 7   # Shift A starts 07:00
    DB_NAME         = cbc_env.ENV.get("JKT_DB_DATABASE", "jkplanningV1")

    # Stages in building.py that produce GTs (not carcass)
    GT_STAGES = frozenset({
        "8201", "8301", "8302", "8501", "8502", "7301",   # STAGE2
        "7001", "7002", "7003", "7004",
        "6001", "6002", "6003", "6004",
        "7101", "7102", "7103", "7104", "7105", "7106",
        "7201", "7501", "7502", "7503",                    # UNISTAGE
    })


# ══════════════════════════════════════════════════════════════════════════════
# BUILDING OUTPUT READER
# ══════════════════════════════════════════════════════════════════════════════

class BuildingOutputReader:
    """
    Reads the building shift schedule from bc_building_schedule.xlsx and
    builds a lookup {(SKUCode, shift_key): qty} for GT-producing shifts.

    shift_key = (date, shift_label) e.g. (datetime.date(2026,6,1), 'A')
    """

    # Columns expected in the Shift Schedule sheet
    _REQUIRED_COLS = {"SKUCode", "Date", "Shift", "Qty", "Machine"}
    _CO_SENTINELS  = {"CHANGEOVER", "MOULD_CLEAN", "C/O", "CLEANING"}

    def read(self, building_path: str) -> tuple[dict, list]:
        """
        Returns:
            gt_output:   {(SKUCode, date, shift_label): qty}
            shift_order: [(date, shift_label)] in chronological order
        """
        xl = pd.ExcelFile(building_path)
        sheet = ("Shift Schedule" if "Shift Schedule" in xl.sheet_names
                 else xl.sheet_names[0])

        df = None
        for hdr in range(4):
            try:
                cand = pd.read_excel(building_path, sheet_name=sheet, header=hdr)
                low = {str(c).strip().lower() for c in cand.columns}
                if {"skucode", "date", "shift", "qty"}.issubset(low):
                    df = cand
                    break
            except Exception:
                continue

        if df is None:
            print(f"  ⚠️  Could not parse Shift Schedule sheet in {building_path}")
            return {}, []

        # Normalise columns
        col_map = {}
        for c in df.columns:
            cl = str(c).strip().lower()
            if cl == "skucode":   col_map[c] = "SKUCode"
            elif cl == "date":    col_map[c] = "Date"
            elif cl == "shift":   col_map[c] = "Shift"
            elif cl == "qty":     col_map[c] = "Qty"
            elif cl == "machine": col_map[c] = "Machine"
        df = df.rename(columns=col_map)

        df["SKUCode"] = df["SKUCode"].astype(str).str.strip()
        df["Qty"]     = pd.to_numeric(df.get("Qty", 0), errors="coerce").fillna(0)
        df["Machine"] = df.get("Machine", pd.Series([""] * len(df))).astype(str).str.strip()

        # Filter: GT-producing rows only
        df = df[
            df["SKUCode"].apply(lambda s: s not in self._CO_SENTINELS)
            & df["Machine"].apply(lambda m: m in DeriverConfig.GT_STAGES)
            & (df["Qty"] > 0)
        ].copy()

        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"]).dt.date

        # Aggregate by (SKUCode, Date, Shift) — multiple machines may produce same SKU
        agg = (
            df.groupby(["SKUCode", "Date", "Shift"], as_index=False)["Qty"].sum()
        )

        gt_output = {}
        for _, row in agg.iterrows():
            key = (str(row["SKUCode"]), row["Date"], str(row["Shift"]))
            gt_output[key] = gt_output.get(key, 0) + float(row["Qty"])

        # Chronological shift order
        shift_order = []
        seen = set()
        for _, row in df.sort_values(["Date", "Shift"]).iterrows():
            k = (row["Date"], str(row["Shift"]))
            if k not in seen:
                seen.add(k)
                shift_order.append(k)

        print(f"  [BuildingReader] {len(gt_output)} (SKU, shift) GT-output entries")
        return gt_output, shift_order

    def get_gt(self, gt_output: dict, sku: str, date, shift: str) -> float:
        return float(gt_output.get((sku, date, shift), 0.0))


# ══════════════════════════════════════════════════════════════════════════════
# PRESS STATE TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class PressStateTracker:
    """
    Tracks curing press state (RUNNING / CHANGEOVER / MOULD_CLEAN / IDLE)
    for each shift.

    Initialised from:
      - df_running_moulds : active presses with SKU assignment
      - all_press_ids     : ALL curing presses (running + idle)
      - runner_out_skus   : SKUs classified Runner-Out → presses IDLE by default
      - df_co_plan        : scheduled changeover events

    Runner-Out presses → IDLE immediately; RUNNING target SKU after CO.
    Presses not in testing_Daily_Running_Moulds → IDLE (no initial SKU).
    """

    def __init__(
        self,
        df_running_moulds: pd.DataFrame,
        df_co_plan: pd.DataFrame | None,
        shift_order: list,
        runner_out_skus: set | None = None,
        all_press_ids: set | None = None,
    ):
        self._initial: dict[str, str] = {}          # {press: initial_sku}
        self._runner_out_initial: set[str] = set()   # presses running Runner-Out SKUs
        self._idle_presses: set[str] = set()         # presses with no initial assignment
        self._co_events: dict[str, list] = defaultdict(list)
        self._shift_order = shift_order
        self._shift_idx: dict[tuple, int] = {s: i for i, s in enumerate(shift_order)}

        _ro_skus = runner_out_skus or set()

        # Initialise from running moulds
        if df_running_moulds is not None and len(df_running_moulds) > 0:
            for _, row in df_running_moulds.iterrows():
                m   = str(row["Machine"]).strip()
                sku = str(row.get("SKUCode", "")).strip()
                if m and sku:
                    self._initial[m] = sku
                    if sku in _ro_skus:
                        self._runner_out_initial.add(m)

        # Presses in all_press_ids but NOT in running moulds → IDLE
        if all_press_ids:
            for press in all_press_ids:
                if press and press not in self._initial:
                    self._idle_presses.add(str(press).strip())

        # Load CO events
        if df_co_plan is not None and len(df_co_plan) > 0:
            for _, co in df_co_plan.iterrows():
                if co.get("Status") != "SCHEDULED":
                    continue
                press = str(co["Press"]).strip()
                day   = int(co["CO_Day_Index"])
                tgt   = str(co["Target_SKU"]).strip()
                self._co_events[press].append((day, tgt))

    def _base_state(self, press: str) -> dict:
        """State before any CO event applies."""
        if press in self._idle_presses or press in self._runner_out_initial:
            return {"status": "IDLE", "sku": self._initial.get(press, "")}
        return {"status": "RUNNING", "sku": self._initial.get(press, "")}

    def get_state(self, press: str, date, shift: str) -> dict:
        """Returns {"status": str, "sku": str} for the given (press, date, shift)."""
        shift_key = (date, shift)
        si = self._shift_idx.get(shift_key, -1)
        if si < 0:
            return self._base_state(press)

        day_idx    = si // DeriverConfig.SHIFTS_PER_DAY
        shift_in_d = si % DeriverConfig.SHIFTS_PER_DAY   # 0=A, 1=B, 2=C

        # Walk CO events chronologically; find today's CO or last past CO
        sorted_events = sorted(self._co_events.get(press, []), key=lambda x: x[0])
        latest_past_sku: str | None = None

        for (co_day, target_sku) in sorted_events:
            if co_day == day_idx:
                if shift_in_d == 0:
                    return {"status": "CHANGEOVER",  "sku": ""}
                if shift_in_d == 1:
                    return {"status": "MOULD_CLEAN", "sku": ""}
                return {"status": "RUNNING", "sku": target_sku}
            if co_day < day_idx:
                latest_past_sku = target_sku
            # co_day > day_idx → future CO, ignore

        if latest_past_sku is not None:
            return {"status": "RUNNING", "sku": latest_past_sku}

        return self._base_state(press)

    @property
    def all_presses(self) -> list[str]:
        """All tracked presses: running moulds + idle presses."""
        all_p = set(self._initial.keys()) | self._idle_presses
        return sorted(all_p)


# ══════════════════════════════════════════════════════════════════════════════
# GT BALANCE TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class GTBalanceTracker:
    """
    Rolling per-SKU GT balance.

    Balance is shared across ALL presses running the same SKU (not per-press),
    because GTs are fungible inventory — any press can consume from the pool.
    """

    def __init__(self, opening_gt: dict[str, float]):
        # opening_gt: {SKUCode: GT_Inventory}
        self._balance: dict[str, float] = {k: float(v) for k, v in opening_gt.items()}

    def get_balance(self, sku: str) -> float:
        return self._balance.get(sku, 0.0)

    def process_shift(
        self,
        sku: str,
        building_output_qty: float,
        consumption_needed: float,   # qty a SINGLE press needs this shift
    ) -> dict:
        """
        Add building output, then consume. Returns result dict.

        consumption_needed is per-press; caller passes total for all presses
        running this SKU this shift.
        """
        # GT arrives from building (same-shift S — minimum case; S-1 preferred
        # but handled by caller already having added previous shift's output first)
        balance = self._balance.get(sku, 0.0) + building_output_qty

        # How much can we cure?
        cured_qty = min(consumption_needed, balance)
        balance   -= cured_qty

        self._balance[sku] = balance

        if cured_qty <= 0 and consumption_needed > 0:
            status = "WAITING_GT"
        elif consumption_needed <= 0:
            status = "IDLE"
        else:
            status = "RUNNING"

        return {
            "cured_qty":          cured_qty,
            "gt_consumed":        cured_qty,
            "gt_balance":         balance,
            "building_output":    building_output_qty,
            "consumption_needed": consumption_needed,
            "status":             status,
        }


# ══════════════════════════════════════════════════════════════════════════════
# CURING SCHEDULE DERIVER
# ══════════════════════════════════════════════════════════════════════════════

class CuringScheduleDeriver:
    """
    Iterates all (press, shift) pairs chronologically and derives curing output.
    Delegates GT tracking to GTBalanceTracker.
    """

    def derive(
        self,
        press_state_tracker: PressStateTracker,
        gt_balance_tracker:  GTBalanceTracker,
        building_output_reader: BuildingOutputReader,
        gt_output: dict,
        shift_order: list,
        df_consumption: pd.DataFrame,
    ) -> tuple:
        """
        Returns 4 DataFrames:
          df_shift_schedule, df_gt_balance, df_demand_fulfillment, df_daily_summary

        STRICT GT timing: curing shift S uses only building output from shift S-1
        (prior-shift). Same-shift building output does NOT count for curing.
        """
        # Build lookup: {SKUCode: qty_per_press_per_shift}
        qty_per_press_map: dict[str, float] = {}
        category_map:      dict[str, str]   = {}
        demand_map:        dict[str, float] = {}
        for _, row in df_consumption.iterrows():
            sku = str(row["SKUCode"]).strip()
            qty_per_press_map[sku] = float(row.get("Qty_Per_Press_Per_Shift", 0))
            category_map[sku]      = str(row.get("Category", ""))
            demand_map[sku]        = float(row.get("Demand_Qty", 0))

        all_presses = press_state_tracker.all_presses

        ss_rows  = []   # Shift Schedule rows (one per press)
        gtb_rows = []   # GT Balance rows (one per SKU per shift)

        # Track per-SKU totals for demand fulfillment
        sku_cured: dict[str, float] = defaultdict(float)
        sku_starv: dict[str, int]   = defaultdict(int)

        for i, (date, shift) in enumerate(shift_order):
            # ── STRICT prior-shift GT rule ──────────────────────────────────
            # GT from building shift S-1 is available at start of curing shift S.
            # Same-shift building output is NOT counted (not yet produced).
            if i > 0:
                prev_date, prev_shift = shift_order[i - 1]
            else:
                prev_date, prev_shift = None, None

            # Collect all presses and their states this shift
            active_skus:    dict[str, list[str]] = defaultdict(list)  # sku → [presses]
            press_statuses: dict[str, dict]      = {}
            for press in all_presses:
                state = press_state_tracker.get_state(press, date, shift)
                press_statuses[press] = state
                if state["status"] == "RUNNING" and state["sku"]:
                    active_skus[state["sku"]].append(press)

            # Process GT balance once per SKU (GT pool is shared across presses)
            sku_gt_results: dict[str, dict] = {}
            sku_gt_source:  dict[str, str]  = {}

            for sku, presses in active_skus.items():
                n_presses = len(presses)

                # Previous shift's building output only (STRICT prior-shift rule)
                if prev_date is not None:
                    prev_bld = building_output_reader.get_gt(
                        gt_output, sku, prev_date, prev_shift
                    )
                else:
                    prev_bld = 0.0

                # Snapshot balance before adding building output (determines GT_Source)
                pre_balance = gt_balance_tracker.get_balance(sku)

                consumption_total = qty_per_press_map.get(sku, 0.0) * n_presses
                result = gt_balance_tracker.process_shift(sku, prev_bld, consumption_total)
                sku_gt_results[sku] = result

                # GT_Source: what fed curing this shift?
                if prev_bld > 0:
                    gt_source = "BUILDING_S-1"
                elif pre_balance > 0:
                    gt_source = "INVENTORY"
                else:
                    gt_source = "NONE"
                sku_gt_source[sku] = gt_source

                # GT Balance row (one per SKU per shift)
                gtb_rows.append({
                    "Date":                date,
                    "Shift":               shift,
                    "SKUCode":             sku,
                    "Active_Press_Count":  n_presses,
                    "Building_Output_S1":  round(prev_bld, 0),
                    "Curing_Consumption":  round(consumption_total, 0),
                    "Cured_Qty":           round(result["cured_qty"], 0),
                    "GT_Balance":          round(result["gt_balance"], 0),
                    "GT_Source":           gt_source,
                    "Status":              result["status"],
                })

                if result["status"] == "WAITING_GT":
                    sku_starv[sku] += 1

                sku_cured[sku] += result["cured_qty"]

            # Shift Schedule rows — one per press (bc.md §12.4 schema)
            for press in all_presses:
                state = press_statuses[press]
                sku   = state["sku"]
                st    = state["status"]

                if st == "RUNNING" and sku:
                    gt_res    = sku_gt_results.get(sku, {})
                    n_presses = len(active_skus.get(sku, [press]))
                    # Each press shares the GT pool equally → split cured qty
                    press_cured = round(gt_res.get("cured_qty", 0) / max(n_presses, 1), 0)
                    ss_rows.append({
                        "Press_ID":               press,
                        "Date":                   date,
                        "Shift":                  shift,
                        "SKU_Code":               sku,
                        "Status":                 gt_res.get("status", "RUNNING"),
                        "Qty_Produced":           press_cured,
                        "GT_Source":              sku_gt_source.get(sku, "NONE"),
                        "GT_Consumed":            press_cured,
                        "GT_Inventory_Remaining": round(gt_res.get("gt_balance", 0), 0),
                        "Active_Press_Count":      n_presses,
                    })
                else:
                    ss_rows.append({
                        "Press_ID":               press,
                        "Date":                   date,
                        "Shift":                  shift,
                        "SKU_Code":               sku if sku else "",
                        "Status":                 st,
                        "Qty_Produced":           0,
                        "GT_Source":              "",
                        "GT_Consumed":            0,
                        "GT_Inventory_Remaining": round(
                            gt_balance_tracker.get_balance(sku) if sku else 0, 0
                        ),
                        "Active_Press_Count":      0,
                    })

        df_shift_schedule = pd.DataFrame(ss_rows)
        df_gt_balance     = pd.DataFrame(gtb_rows)

        df_demand_fulfillment = self._build_demand_fulfillment(
            df_consumption, sku_cured, sku_starv
        )
        df_daily_summary = self._build_daily_summary(df_gt_balance)

        return df_shift_schedule, df_gt_balance, df_demand_fulfillment, df_daily_summary

    @staticmethod
    def _build_demand_fulfillment(
        df_consumption: pd.DataFrame,
        sku_cured: dict,
        sku_starv: dict,
    ) -> pd.DataFrame:
        _valid_cats = {"Runner-In", "Runner-Out", "Non-Runner-In"}
        rows = []
        for _, row in df_consumption.iterrows():
            sku = str(row.get("SKUCode", "")).strip()
            if not sku or sku.lower() in {"nan", "none", ""}:
                continue
            cat = str(row.get("Category", "")).strip()
            if cat not in _valid_cats:
                continue
            dem_raw = row.get("Demand_Qty", 0)
            dem     = 0.0 if (dem_raw is None or (isinstance(dem_raw, float) and math.isnan(dem_raw))) else float(dem_raw)
            cured   = sku_cured.get(sku, 0.0)
            starv   = sku_starv.get(sku, 0)
            pct     = round(100 * cured / dem, 1) if dem > 0 else 0.0

            if pct >= 100:
                status = "FULLY MET"
                skip   = "-"
            elif cured > 0:
                status = "PARTIAL"
                skip   = f"GT shortage — {starv} starvation shifts out of 90"
            else:
                status = "UNMET"
                if starv >= 80:
                    skip = "No GT from building — check building schedule for this SKU"
                elif cat == "Non-Runner-In":
                    skip = "NRI: curing press not yet active (no CO completed within plan)"
                else:
                    skip = "Zero curing output — GT supply or press state issue"

            rows.append({
                "SKUCode":           sku,
                "Category":          cat,
                "Demand_Qty":        int(dem),
                "Total_Cured":       int(cured),
                "Fulfillment_Pct":   pct,
                "Status":            status,
                "Skip_Reason":       skip,
                "Starvation_Shifts": starv,
            })
        return (
            pd.DataFrame(rows)
            .sort_values(["Category", "Fulfillment_Pct"], ascending=[True, False])
            .reset_index(drop=True)
        )

    @staticmethod
    def _build_daily_summary(df_gt_balance: pd.DataFrame) -> pd.DataFrame:
        if df_gt_balance.empty:
            return pd.DataFrame(columns=["Date", "Total_Cured", "Distinct_SKUs",
                                         "Active_Presses", "Starvation_Events"])
        daily = (
            df_gt_balance.groupby("Date")
            .agg(
                Total_Cured=("Cured_Qty", "sum"),
                Distinct_SKUs=("SKUCode", "nunique"),
                Active_Presses=("Active_Press_Count", "sum"),
                Starvation_Events=(
                    "Status",
                    lambda x: (x == "WAITING_GT").sum(),
                ),
            )
            .reset_index()
        )
        daily["Total_Cured"]    = daily["Total_Cured"].astype(int)
        daily["Active_Presses"] = daily["Active_Presses"].astype(int)
        return daily


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════

_NAVY  = "1F3864"
_WHITE = "FFFFFF"
_GREEN = "E2EFDA"
_RED   = "FFE0E0"
_AMBER = "FFF2CC"
_BLUE  = "DDEEFF"

_STATUS_COLORS = {
    "RUNNING":     _GREEN,
    "WAITING_GT":  _RED,
    "CHANGEOVER":  _AMBER,
    "MOULD_CLEAN": _AMBER,
    "IDLE":        _BLUE,
}


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor=_NAVY)
    hdr_font = Font(bold=True, color=_WHITE)
    bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))

    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = bd
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        status_col = row.get("Status", "")
        row_color  = _STATUS_COLORS.get(str(status_col), "")
        row_fill   = PatternFill("solid", fgColor=row_color) if row_color else None
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bd
            cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 30)


def export_curing_schedule(
    df_shift_schedule: pd.DataFrame,
    df_gt_balance: pd.DataFrame,
    df_demand_fulfillment: pd.DataFrame,
    df_daily_summary: pd.DataFrame,
    output_path: str,
):
    """Write bc_curing_schedule.xlsx with four sheets."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # Sheet 1: Shift Schedule
    ws1 = wb.active
    ws1.title = "Shift Schedule"
    ws1.cell(row=1, column=1, value="Curing Shift Schedule — B2C Derived").font = \
        Font(bold=True, size=12)
    _write_df_to_sheet(ws1, df_shift_schedule, start_row=3)

    # Sheet 2: GT Balance
    ws2 = wb.create_sheet("GT Balance")
    ws2.cell(row=1, column=1, value="GT Balance Per SKU Per Shift").font = \
        Font(bold=True, size=12)
    _write_df_to_sheet(ws2, df_gt_balance, start_row=3)

    # Sheet 3: Demand Fulfillment
    ws3 = wb.create_sheet("Demand Fulfillment")
    ws3.cell(row=1, column=1, value="Demand Fulfillment by SKU").font = \
        Font(bold=True, size=12)
    _write_df_to_sheet(ws3, df_demand_fulfillment, start_row=3)
    # Color-code by Status column
    if "Status" in df_demand_fulfillment.columns:
        status_col_idx = list(df_demand_fulfillment.columns).index("Status") + 1
        color_map = {"FULLY MET": "E2EFDA", "PARTIAL": "FFF2CC",
                     "UNMET": "FFE0E0", "EXCLUDED": "D3D3D3"}
        for row_i, status_val in enumerate(df_demand_fulfillment["Status"], start=4):
            fill_hex = color_map.get(str(status_val), "FFFFFF")
            fill = PatternFill("solid", fgColor=fill_hex)
            for col_i in range(1, len(df_demand_fulfillment.columns) + 1):
                ws3.cell(row=row_i, column=col_i).fill = fill

    # Sheet 4: Daily Summary
    ws4 = wb.create_sheet("Daily Summary")
    ws4.cell(row=1, column=1, value="Daily Curing Summary").font = \
        Font(bold=True, size=12)
    _write_df_to_sheet(ws4, df_daily_summary, start_row=3)

    # KPI summary block on Daily Summary sheet
    if len(df_daily_summary) > 0:
        last_row = len(df_daily_summary) + 6
        total_cured   = int(df_daily_summary["Total_Cured"].sum())
        avg_daily     = total_cured / max(len(df_daily_summary), 1)
        starv         = int(df_daily_summary["Starvation_Events"].sum())

        # Customer demand = sum of Demand_Qty from fulfillment sheet
        dem_col = next((c for c in ["Demand_Qty", "Customer_Demand"] if c in df_demand_fulfillment.columns), None)
        total_demand = int(df_demand_fulfillment[dem_col].sum()) if dem_col else 0
        kpi_pct = round(100 * total_cured / total_demand, 1) if total_demand else 0.0

        ws4.cell(row=last_row, column=1, value="KPI SUMMARY").font = Font(bold=True, size=11)
        kpi_label = ws4.cell(row=last_row + 1, column=1,
                             value=f"KPI — Cured / Customer Demand: {kpi_pct}%")
        kpi_label.font = Font(bold=True)
        ws4.cell(row=last_row + 2, column=1, value=f"Total Customer Demand (units): {total_demand:,}")
        ws4.cell(row=last_row + 3, column=1, value=f"Total Cured Tyres (Month): {total_cured:,}")
        ws4.cell(row=last_row + 4, column=1, value=f"Average Daily Cured Tyres: {avg_daily:,.0f}")
        ws4.cell(row=last_row + 5, column=1, value=f"Starvation Events (shifts): {starv}")
        print(f"  [Curing KPI] Cured / Customer Demand = {total_cured:,} / {total_demand:,} = {kpi_pct}%")

    wb.save(output_path)
    print(f"  [DeriverExport] Saved → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# ETL: load GT inventory from DB
# ══════════════════════════════════════════════════════════════════════════════

def _load_gt_inventory(engine) -> dict[str, float]:
    """Load opening GT inventory from DB. Returns {SKUCode: qty}."""
    try:
        df = pd.read_sql(
            f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory "
            f"FROM {DeriverConfig.DB_NAME}.gt_inventory_manual",
            engine,
        )
        return dict(zip(
            df["SKUCode"].astype(str).str.strip(),
            df["GT_Inventory"].astype(float)
        ))
    except Exception as exc:
        print(f"  ⚠️  Could not load GT inventory: {exc}")
        return {}


def _load_running_moulds(engine) -> pd.DataFrame:
    """Load curing running moulds for press state initialisation."""
    try:
        from curing_consumption import ConsumptionETL
        cetl = ConsumptionETL(engine)
        return cetl.load_running_moulds()
    except Exception as exc:
        print(f"  ⚠️  Could not load running moulds: {exc}")
        return pd.DataFrame(columns=["Machine", "SKUCode", "MouldNos",
                                     "MouldLife_remaining", "Num_Moulds"])


# ══════════════════════════════════════════════════════════════════════════════
# ETL: load ALL curing press IDs from DB
# ══════════════════════════════════════════════════════════════════════════════

def _load_all_curing_press_ids(engine) -> set:
    """
    Return the set of ALL curing press IDs (running + idle) from:
      1. testing_Daily_Running_Moulds  — same machine-ID derivation as ConsumptionETL
      2. Master_Curing_Allowable_Machines_source — numeric column names are press IDs
    """
    db = DeriverConfig.DB_NAME
    press_ids: set = set()

    # ── From testing_Daily_Running_Moulds (all rows, not just those with active SKU) ──
    try:
        import numpy as np
        wc_master = pd.read_sql(f"SELECT wcID, WCNAME FROM {db}.Master_WC_Master", engine)
        df_rm = pd.read_sql(f"SELECT WCNAME, Side FROM {db}.testing_Daily_Running_Moulds", engine)
        df_rm["WCNAME"] = (
            df_rm["WCNAME"].str.replace(r"(LH|RH)$", "", regex=True).str.strip()
        )
        df_rm["curing_machine"] = df_rm["WCNAME"] + df_rm["Side"]
        press_ids.update(
            str(m).strip() for m in df_rm["curing_machine"].dropna() if str(m).strip()
        )
    except Exception as exc:
        print(f"  ⚠️  Could not load all testing_Daily_Running_Moulds presses: {exc}")

    # ── From Master_Curing_Allowable_Machines_source (press IDs = numeric cols) ──
    try:
        df_allow = pd.read_sql(
            f"SELECT * FROM {db}.Master_Curing_Allowable_Machines_source LIMIT 1", engine
        )
        mcols = [str(c) for c in df_allow.columns if str(c).strip().isdigit()]
        press_ids.update(mcols)
    except Exception as exc:
        print(f"  ⚠️  Could not load Master_Curing_Allowable_Machines_source presses: {exc}")

    press_ids.discard("")
    press_ids.discard("nan")
    return press_ids


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def derive_curing_schedule(
    building_path: str,
    consumption_path: str,
    output_path: str,
    engine=None,
    co_plan: pd.DataFrame | None = None,
) -> dict:
    """
    Derive the B2C curing schedule from building output.

    Args:
        building_path:    Path to bc_building_schedule.xlsx (Phase 1 output).
        consumption_path: Path to curing_consumption_table.xlsx (Phase 0 output).
        output_path:      Where to write bc_curing_schedule.xlsx.
        engine:           SQLAlchemy engine (created from .env if None).
        co_plan:          Optional changeover plan DataFrame from building_b2c.

    Returns dict with keys: shift_schedule, gt_balance, demand_fulfillment, daily_summary
    """
    from cbc_env import make_engine as _mk

    if engine is None:
        engine = _mk()

    print("\n" + "=" * 70)
    print("  B2C Phase 2 — Curing Schedule Deriver")
    print("=" * 70)

    # ── Load consumption table ────────────────────────────────────────────────
    print("  [ETL] Loading consumption table …")
    xl = pd.ExcelFile(consumption_path)
    sheet = ("Consumption Summary" if "Consumption Summary" in xl.sheet_names
             else xl.sheet_names[0])
    df_consumption = None
    for hdr in range(3):
        try:
            cand = pd.read_excel(consumption_path, sheet_name=sheet, header=hdr)
            if "SKUCode" in cand.columns and "Category" in cand.columns:
                df_consumption = cand
                break
        except Exception:
            continue
    if df_consumption is None:
        raise ValueError(f"Cannot parse consumption table from {consumption_path}")
    df_consumption["SKUCode"] = df_consumption["SKUCode"].astype(str).str.strip()
    print(f"        {len(df_consumption)} SKUs in consumption table")

    # ── Load opening GT inventory ─────────────────────────────────────────────
    print("  [ETL] Loading opening GT inventory …")
    opening_gt = _load_gt_inventory(engine)
    print(f"        {len(opening_gt)} SKUs with GT inventory")

    # ── Extract Runner-Out SKUs from consumption table ────────────────────────
    runner_out_skus: set = set()
    if df_consumption is not None and "Category" in df_consumption.columns:
        runner_out_skus = set(
            df_consumption.loc[df_consumption["Category"] == "Runner-Out", "SKUCode"]
            .astype(str).str.strip()
        )
    print(f"        {len(runner_out_skus)} Runner-Out SKUs (presses → IDLE by default)")

    # ── Load running moulds (press → SKU snapshot) ────────────────────────────
    print("  [ETL] Loading curing running moulds …")
    df_running_moulds = _load_running_moulds(engine)
    print(f"        {len(df_running_moulds)} active curing presses in testing_Daily_Running_Moulds")

    # ── Load ALL curing press IDs ─────────────────────────────────────────────
    print("  [ETL] Loading all curing press IDs …")
    all_press_ids = _load_all_curing_press_ids(engine)
    print(f"        {len(all_press_ids)} total curing presses (running + idle)")

    # ── Read building output ──────────────────────────────────────────────────
    print("  [ETL] Reading building shift schedule …")
    reader = BuildingOutputReader()
    gt_output, shift_order = reader.read(building_path)

    if not shift_order:
        print("  ⚠️  No shifts found in building output. Deriving shift order from consumption.")
        # Fallback: generate shift order from a 30-day horizon
        from datetime import date as _date
        base = datetime(2026, 6, 1, 7, 0)
        for d in range(30):
            for sh in ["A", "B", "C"]:
                day_date = (base + timedelta(days=d)).date()
                shift_order.append((day_date, sh))

    print(f"        {len(shift_order)} shifts in timeline | {len(gt_output)} GT output entries")

    # ── Initialise trackers ───────────────────────────────────────────────────
    print("  [Derive] Initialising press state tracker …")
    press_tracker = PressStateTracker(
        df_running_moulds,
        co_plan,
        shift_order,
        runner_out_skus=runner_out_skus,
        all_press_ids=all_press_ids,
    )
    print(f"        {len(press_tracker.all_presses)} curing presses tracked")

    gt_tracker = GTBalanceTracker(opening_gt)

    # ── Derive curing schedule ────────────────────────────────────────────────
    print("  [Derive] Rolling through shifts …")
    deriver = CuringScheduleDeriver()
    df_ss, df_gtb, df_ful, df_day = deriver.derive(
        press_tracker, gt_tracker, reader, gt_output, shift_order, df_consumption
    )

    # ── KPI summary ───────────────────────────────────────────────────────────
    if len(df_day) > 0:
        total = int(df_day["Total_Cured"].sum())
        starv = int(df_day["Starvation_Events"].sum())
        avg   = total / max(len(df_day), 1)
        dem_col = next((c for c in ["Demand_Qty", "Customer_Demand"] if c in df_ful.columns), None)
        total_demand_console = int(df_ful[dem_col].sum()) if dem_col else 0
        kpi_pct_console = round(100 * total / total_demand_console, 1) if total_demand_console else 0.0
        print(f"\n  KPI — Customer Demand:       {total_demand_console:,}")
        print(f"  KPI — Total Cured Tyres:     {total:,}")
        print(f"  KPI — Cured / Demand:        {kpi_pct_console}%")
        print(f"  KPI — Avg Daily Cured:       {avg:,.0f}")
        print(f"  KPI — Starvation Events:     {starv}")
        if total > 0:
            ri_cured = df_ful[df_ful["Category"] == "Runner-In"]["Total_Cured"].sum()
            print(f"  KPI — Runner-In Cured:       {ri_cured:,.0f} ({100*ri_cured/total:.1f}%)")

    # ── Export ────────────────────────────────────────────────────────────────
    print(f"\n  [Export] Writing → {output_path}")
    export_curing_schedule(df_ss, df_gtb, df_ful, df_day, output_path)

    print("=" * 70)
    print("  Phase 2 complete.")
    print("=" * 70 + "\n")

    return {
        "shift_schedule":     df_ss,
        "gt_balance":         df_gtb,
        "demand_fulfillment": df_ful,
        "daily_summary":      df_day,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STANDALONE RUN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    _BUILDING_PATH    = os.path.join(MAIN_OUT, "bc_building_schedule.xlsx")
    _CONSUMPTION_PATH = os.path.join(OUT_DIR,  "curing_consumption_table.xlsx")
    _OUTPUT_PATH      = os.path.join(MAIN_OUT, "bc_curing_schedule.xlsx")

    # Try to load the curing CO plan from the building schedule Excel
    _co_plan = None
    try:
        _xl = pd.ExcelFile(_BUILDING_PATH)
        if "Changeover Plan" in _xl.sheet_names:
            _co_plan = pd.read_excel(_BUILDING_PATH, sheet_name="Changeover Plan")
            print(f"  [Standalone] Loaded CO plan from building schedule: {len(_co_plan)} rows")
        else:
            print("  [Standalone] No 'Changeover Plan' sheet found — running without CO plan")
    except Exception as _exc:
        print(f"  [Standalone] Could not load CO plan: {_exc}")

    result = derive_curing_schedule(
        building_path    = _BUILDING_PATH,
        consumption_path = _CONSUMPTION_PATH,
        output_path      = _OUTPUT_PATH,
        co_plan          = _co_plan,
    )

    df_day = result["daily_summary"]
    if len(df_day) > 0:
        print("\nFirst 5 days:")
        print(df_day.head(5).to_string(index=False))
