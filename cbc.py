"""
JK Tyre BTP — CBC ORCHESTRATOR  (Curing -> Building -> Curing)
=============================================================
Wires the three stages together:

    Phase 0  FEED map      build the press->building-machine map from history
                           (feed_map_builder.py) — run periodically, not daily.

    Phase C  CURING        run the curing LP scheduler (curing_lp.py) with the
             (feed-aware)   FEED filter installed, sourcing all inputs from the
                           planning DB EXCEPT demand (Book4.xlsx).

    Phase B  BUILDING      run the building GA+LP scheduler (building.py) on the
                           feed-aware curing plan (handed off as a bridge file).

Data layout:  inputs in data/input/, outputs in data/output/.
DB creds:     data from .env via cbc_env.py.

Run:
    python cbc.py
"""

from __future__ import annotations

import os
import sys

# ── Run under the project venv even if launched as `python3 cbc.py` ──────────
# The system python3 has no pandas/scipy; re-exec with myenv's interpreter so
# the user can run `python3 cbc.py` directly. (No-op when already in myenv.)
# NB: myenv/bin/python is a symlink to the system interpreter, so compare
# sys.prefix (which differs: venv root vs system) — NOT the binary path.
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("CBC_REEXEC")):
    os.environ["CBC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd

import curing_lp
import cbc_env
from feed_map_builder import FeedMapConfig, build_feed_map
from feed_aware_curing import (
    load_feed_map,
    install_feed_awareness,
)

HERE = os.path.dirname(os.path.abspath(__file__))
IN = cbc_env.INPUT_DIR
OUT = cbc_env.OUTPUT_DIR
# The three final schedules go here; intermediates stay in OUT.
MAIN_OUT = os.path.join(OUT, "main_output")
os.makedirs(MAIN_OUT, exist_ok=True)

# Building PDE (production-event) history — the monthly exports under
# data/input/building pde/. These supply both the feed-map builder (Phase 0)
# and the Phase-C2 building-history feasibility check. Replaces the old single
# productionevent_data.csv. Each file has workcenter / RecipeCode / DtandTime.
BUILDING_PDE_FILES = [
    os.path.join("building pde", "april_pde.csv"),
    os.path.join("building pde", "maypde.csv"),
    os.path.join("building pde", "junepde.csv"),
]


def _feed_cfg() -> FeedMapConfig:
    """Feed-map config pointed at the real data files (data/input) with the
    column names those files actually use."""
    return FeedMapConfig(
        INPUT_DIR=IN,
        CURING_EVENTS_FILE="CURING_PCR 1.csv",     # wcID, recipeID, dtandTime
        BUILDING_EVENTS_FILE=BUILDING_PDE_FILES,    # monthly building-pde exports
        RECIPE_MASTER_FILE="RECIPE_MASTER.csv",    # iD -> description (recipe code)
        WC_MASTER_FILE="WCMASTER.csv",             # iD -> name (curing press)
        NAME_MAP_FILE=None,                        # use built-in workcenter->code map
        # column overrides to match our files
        RM_RECIPE_ID_COLS=("iD", "recipe id", "recipeID"),
        RM_RECIPE_CODE_COLS=("description", "recipecode", "RecipeCode"),
        WCM_WCID_COLS=("iD", "wcID", "WCID"),
        WCM_PRESS_COLS=("name", "press", "WCNAME"),
        # write feed-map outputs into data/output
        OUT_MAP=os.path.join("..", "output", "feed_map.xlsx"),
        OUT_INVERSE=os.path.join("..", "output", "feed_map_inverse.xlsx"),
        OUT_JSON=os.path.join("..", "output", "feed_map.json"),
        OUT_COVERAGE=os.path.join("..", "output", "feed_coverage.xlsx"),
    )


# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════
@dataclass
class CBCConfig:
    # FEED map: prebuilt once from the 2.5 GB building log and reused (stable /
    # periodic). Flip to True to rebuild when the building-events data changes.
    REBUILD_FEED_MAP: bool = False
    FEED_MAP_JSON: str = os.path.join(OUT, "feed_map.json")
    feed_cfg: FeedMapConfig = field(default_factory=_feed_cfg)

    # ── Curing inputs ───────────────────────────────────────────────────
    # Demand comes from a file (Book4); everything else from the DB.
    CUR_DEMAND: str = os.path.join(IN, "Book4.xlsx")
    CUR_OUTPUT: str = os.path.join(MAIN_OUT, "PCR_CBC_Curing_Initial.xlsx")
    PLAN_START: datetime = datetime(2026, 5, 1, 7, 0, 0)

    # Feed-filter behaviour for missing data (see feed_aware_curing).
    KEEP_UNKNOWN_PRESS: bool = True
    KEEP_UNKNOWN_SKU: bool = True

    # ── Building stage ──────────────────────────────────────────────────
    RUN_BUILDING: bool = True
    # Align building's horizon to curing's so it supplies GT for the WHOLE month
    # (was 8 → final curing was capped at an 8-day build). Matches curing's
    # PLANNING_DAYS=31. Building GA runs longer at this horizon.
    BUILDING_PLANNING_DAYS: int = 31
    BUILDING_OUTPUT: str = os.path.join(MAIN_OUT, "PCR_CBC_Building.xlsx")
    BRIDGE_FILE: str = os.path.join(OUT, "curing_plan_for_building.xlsx")

    # ── Final curing (Phase C2): re-run curing capped by building's GT supply ──
    RUN_FINAL_CURING: bool = True
    FINAL_CURING_OUTPUT: str = os.path.join(MAIN_OUT, "PCR_CBC_Curing_Final.xlsx")
    FINAL_DEMAND_FILE: str = os.path.join(OUT, "Book4_final_capped.xlsx")


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════
def _append_daily_production(output_path: str, df_shift) -> None:
    """Add a 'Daily Total Production' sheet to the curing output workbook:
    per calendar day, cured units by shift (A/B/C) + day total, distinct SKUs
    and active presses, with a grand-total row. Production = rows with Qty>0
    (excludes CHANGEOVER and zero-qty mould-clean rows)."""
    if df_shift is None or df_shift.empty:
        print("  [Phase C] no shift_schedule — skipped Daily Total Production sheet")
        return
    df = df_shift.copy()
    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)
    prod = df[df["Qty"] > 0]

    piv = prod.pivot_table(index="Date", columns="Shift", values="Qty",
                           aggfunc="sum", fill_value=0)
    for s in ("A", "B", "C"):
        if s not in piv.columns:
            piv[s] = 0
    piv = piv[["A", "B", "C"]]
    piv["Total_Production_Units"] = piv.sum(axis=1)
    piv["Distinct_SKUs"] = prod.groupby("Date")["SKUCode"].nunique()
    piv["Active_Presses"] = prod.groupby("Date")["Machine"].nunique()
    daily = piv.reset_index().sort_values("Date")

    total = {"Date": "TOTAL", "A": daily["A"].sum(), "B": daily["B"].sum(),
             "C": daily["C"].sum(),
             "Total_Production_Units": daily["Total_Production_Units"].sum(),
             "Distinct_SKUs": prod["SKUCode"].nunique(),
             "Active_Presses": prod["Machine"].nunique()}
    daily = pd.concat([daily, pd.DataFrame([total])], ignore_index=True)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as xw:
        daily.to_excel(xw, sheet_name="Daily Total Production", index=False)
    print(f"  [Phase C] added 'Daily Total Production' sheet "
          f"({len(daily) - 1} days, {int(total['Total_Production_Units']):,} units)")


def _building_allowable_from_db(engine) -> dict[str, set[str]]:
    """SKU -> set(building machine codes) from Master_Building_Allowable_Machines_source."""
    df = pd.read_sql(
        "SELECT * FROM Master_Building_Allowable_Machines_source", engine)
    sku_col = "SKU Code" if "SKU Code" in df.columns else (
        "SKUCode" if "SKUCode" in df.columns else df.columns[0])
    mcols = [c for c in df.columns if str(c).strip().isdigit()]
    yes = {"Y", "YES", "1", "ONLY ONE M/C RUN", "TRUE"}
    out: dict[str, set[str]] = {}
    for _, r in df.iterrows():
        machines = {str(c).strip() for c in mcols
                    if str(r[c]).strip().upper() in yes}
        out[str(r[sku_col]).strip()] = machines
    print(f"  [FeedAware] building-allowable for {len(out)} SKUs (from DB)")
    return out


# ══════════════════════════════════════════════════════════════════════════
# PHASES
# ══════════════════════════════════════════════════════════════════════════
def phase0_feed_map(cfg: CBCConfig) -> dict:
    print("\n" + "=" * 70)
    print("  PHASE 0 — FEED MAP")
    print("=" * 70)
    if cfg.REBUILD_FEED_MAP:
        return build_feed_map(cfg.feed_cfg)
    return {p: ms for p, ms in load_feed_map(cfg.FEED_MAP_JSON).items()}


def phaseC_curing(cfg: CBCConfig, feed_map, engine) -> dict:
    print("\n" + "=" * 70)
    print("  PHASE C — FEED-AWARE CURING (DB inputs + Book4 demand)")
    print("=" * 70)

    bld_allow = _building_allowable_from_db(engine)
    install_feed_awareness(
        curing_lp, feed_map, bld_allow,
        keep_unknown_press=cfg.KEEP_UNKNOWN_PRESS,
        keep_unknown_sku=cfg.KEEP_UNKNOWN_SKU,
    )

    results = curing_lp.run_from_database_simple(
        demand_path=cfg.CUR_DEMAND,
        plan_start=cfg.PLAN_START,
        output_path=cfg.CUR_OUTPUT,
        engine=engine,
    )
    _append_daily_production(cfg.CUR_OUTPUT, results.get("shift_schedule"))
    print(f"\n  [Phase C] curing schedule -> {cfg.CUR_OUTPUT}")
    return results


def phaseB_building(cfg: CBCConfig, curing_results: dict) -> dict | None:
    print("\n" + "=" * 70)
    print("  PHASE B — BUILDING")
    print("=" * 70)
    if not cfg.RUN_BUILDING:
        print("  [Phase B] skipped (RUN_BUILDING=False).")
        return None
    try:
        import building
    except Exception as e:  # noqa: BLE001
        print(f"  [Phase B] building.py not importable ({e}). "
              f"Curing output ready at {cfg.CUR_OUTPUT}.")
        return None

    # Align building horizon to curing's so it supplies GT for the whole month.
    building.Config.PLANNING_DAYS = cfg.BUILDING_PLANNING_DAYS
    print(f"  [Phase B] building horizon set to "
          f"{building.Config.PLANNING_DAYS} days (matches curing)")

    # Hand the feed-aware curing schedule to building as a flat bridge file and
    # point building's curing-plan loader at it.
    df_shift = curing_results.get("shift_schedule")
    if df_shift is not None and not df_shift.empty:
        df_shift.to_excel(cfg.BRIDGE_FILE, index=False)
        building.Config.CURING_PLAN_FILE = cfg.BRIDGE_FILE
        print(f"  [Phase B] curing plan handed off -> {cfg.BRIDGE_FILE}")
    else:
        print("  [Phase B] WARNING: no shift_schedule from curing; "
              "building will use its configured fallback.")

    # NB: run_from_database_hybrid writes the Excel but returns None — return the
    # output path so callers can tell building succeeded (and trigger Phase C2).
    building.run_from_database_hybrid(
        plan_start=cfg.PLAN_START, output_path=cfg.BUILDING_OUTPUT)
    return cfg.BUILDING_OUTPUT


def _building_gt_supply(building_output: str) -> dict[str, float]:
    """Per-SKU GT available to curing = opening GT_Inventory + Planned_GT (what
    building actually built), read from building's 'Demand Summary' sheet."""
    df = pd.read_excel(building_output, sheet_name="Demand Summary", header=2)
    df = df.dropna(subset=["SKUCode"])
    df["SKUCode"] = df["SKUCode"].astype(str).str.strip()
    for c in ("GT_Inventory", "Planned_GT"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return {r.SKUCode: float(r.GT_Inventory + r.Planned_GT)
            for r in df.itertuples()}


def _rebase_final_report_to_true_demand(final_path: str, demand_path: str,
                                        planned_by_sku: dict | None = None,
                                        infeasible_reasons: dict | None = None) -> None:
    """The final-curing 'Demand Fulfillment' sheet reports vs the CAPPED demand,
    so it always looks ~100%. Rewrite it to report vs the TRUE demand (Book4):
      • per-SKU Demand/Gap/Fulfillment recomputed against true demand,
      • per-SKU Planned overridden with the feasible (trimmed) total when given,
      • SKUs not scheduled added as rows with their infeasibility reason,
      • headline KPI line rebased (keeps Util/Changeover/Clean as-is).
    """
    import openpyxl
    infeasible_reasons = infeasible_reasons or {}
    raw = pd.read_excel(demand_path)
    qcol = "Updated_Requirement" if "Updated_Requirement" in raw.columns else "Requirement"
    pcol = "ConsolidatedPriorityScore"
    raw["SKUCode"] = raw["SKUCode"].astype(str).str.strip()
    raw[qcol] = pd.to_numeric(raw[qcol], errors="coerce").fillna(0)
    agg = raw.groupby("SKUCode").agg(d=(qcol, "sum"), p=(pcol, "max"))
    true_d = agg["d"].to_dict()
    prio = agg["p"].to_dict()

    wb = openpyxl.load_workbook(final_path)
    ws = wb["Demand Fulfillment"]
    hdr = next(r for r in range(1, 9)
               if str(ws.cell(r, 1).value).strip() == "SKUCode")
    total_row = next(r for r in range(hdr + 1, ws.max_row + 1)
                     if str(ws.cell(r, 1).value).strip().upper() == "TOTAL")

    seen, total_planned = set(), 0.0
    for r in range(hdr + 1, total_row):
        sku = ws.cell(r, 1).value
        if sku is None or str(sku).strip() == "":
            continue
        sku = str(sku).strip(); seen.add(sku)
        # planned_by_sku is authoritative when given (absent SKU = 0 produced),
        # so a SKU fully trimmed to zero correctly shows 0, not its untrimmed value.
        planned = (float(planned_by_sku.get(sku, 0.0))
                   if planned_by_sku is not None else float(ws.cell(r, 5).value or 0))
        ws.cell(r, 5).value = round(planned)   # reflect feasible (trimmed) planned
        total_planned += planned
        td = float(true_d.get(sku, ws.cell(r, 3).value or 0))
        ws.cell(r, 3).value = round(td)
        ws.cell(r, 6).value = round(td - planned)
        ful = planned / td if td > 0 else 0
        ws.cell(r, 7).value = round(ful, 3)
        ws.cell(r, 8).value = ("FULLY MET" if ful >= 0.999
                               else "PARTIAL" if ful > 0 else "NOT BUILT")

    dropped = sorted((s for s in true_d if s not in seen and true_d[s] > 0),
                     key=lambda x: -true_d[x])
    if dropped:
        ws.insert_rows(total_row, len(dropped))
        for i, s in enumerate(dropped):
            r = total_row + i
            ws.cell(r, 1).value = s
            ws.cell(r, 2).value = round(prio.get(s, 0), 3)
            ws.cell(r, 3).value = round(true_d[s])
            ws.cell(r, 4).value = 0
            ws.cell(r, 5).value = 0
            ws.cell(r, 6).value = round(true_d[s])
            ws.cell(r, 7).value = 0
            ws.cell(r, 8).value = infeasible_reasons.get(s, "NOT FED BY BUILDING")
        total_row += len(dropped)

    total_true = sum(true_d.values())
    gap = total_true - total_planned
    ful = 100.0 * total_planned / total_true if total_true else 0
    ws.cell(total_row, 3).value = round(total_true)
    ws.cell(total_row, 5).value = round(total_planned)
    ws.cell(total_row, 6).value = round(gap)
    ws.cell(total_row, 7).value = round(total_planned / total_true, 3) if total_true else 0
    # rebase headline, preserving Util/Changeover/Clean segments
    old = str(ws.cell(hdr - 1, 1).value or "")
    tail = " | ".join(p.strip() for p in old.split("|")[4:])
    ws.cell(hdr - 1, 1).value = (
        f"Demand: {round(total_true):,}  |  Planned: {round(total_planned):,}  |  "
        f"Gap: {round(gap):,}  |  Fulfillment: {ful:.1f}% (vs TRUE demand)"
        + (f"  |  {tail}" if tail else ""))
    wb.save(final_path)
    print(f"  [Phase C2] report rebased to TRUE demand: {ful:.1f}% "
          f"({int(total_planned):,}/{int(total_true):,}); "
          f"{len(dropped)} SKUs not fed by building")


def _historical_overlap(cfg: CBCConfig):
    """Which demand SKUs have BOTH curing history (CURING_PCR) and building
    history (productionevent_data)? Returns (overlap:set, reasons:dict) where
    reasons maps each non-overlap SKU to its infeasibility text."""
    rm = pd.read_csv(cbc_env.in_path("RECIPE_MASTER.csv"))
    id2code = dict(zip(rm["iD"].astype(str).str.strip(),
                       rm["description"].astype(str).str.strip().str.upper()))
    cur = pd.read_csv(cbc_env.in_path("CURING_PCR 1.csv"), usecols=["recipeID"])
    cur_hist = set(cur["recipeID"].astype(str).str.strip().map(id2code).dropna())
    bld_hist: set[str] = set()
    for f in BUILDING_PDE_FILES:
        bh = pd.read_csv(cbc_env.in_path(f), usecols=["RecipeCode"],
                         low_memory=False)
        bld_hist |= set(bh["RecipeCode"].astype(str).str.strip().str.upper())

    raw = pd.read_excel(cfg.CUR_DEMAND)
    dem_skus = set(raw["SKUCode"].astype(str).str.strip().str.upper())
    overlap = {s for s in dem_skus if s in cur_hist and s in bld_hist}
    reasons = {}
    for s in dem_skus - overlap:
        has_c, has_b = s in cur_hist, s in bld_hist
        reasons[s] = ("No historical data (building & curing)" if not has_c and not has_b
                      else "No historical data (building)" if not has_b
                      else "No historical data (curing)")
    print(f"  [Phase C2] historical overlap: {len(overlap)} schedulable SKUs | "
          f"{len(reasons)} infeasible (no history)")
    return overlap, reasons


def _building_gt_timeline(cfg: CBCConfig) -> dict:
    """{(SKUCode, Date, Shift): GT units} built by GT machines (excludes carcass
    stage-1) from building's Shift Schedule — the time-phased GT supply."""
    import building
    S1 = set(building.Config.STAGE1)
    ss = pd.read_excel(cfg.BUILDING_OUTPUT, sheet_name="Shift Schedule", header=2)
    ss.columns = [str(c) for c in ss.columns]
    ss = ss[~ss["SKUCode"].astype(str).isin(["CHANGEOVER", "MOULD_CLEAN"])]
    ss = ss[~ss["Machine"].astype(str).isin(S1)]
    ss = ss.copy()
    ss["Qty"] = pd.to_numeric(ss["Qty"], errors="coerce").fillna(0)
    ss["SKUCode"] = ss["SKUCode"].astype(str).str.strip()
    ss["Date"] = pd.to_datetime(ss["Date"])
    return ss.groupby(["SKUCode", "Date", "Shift"])["Qty"].sum().to_dict()


def _opening_gt_inventory(engine) -> dict:
    df = curing_lp.ETL(engine, curing_lp.Config.TYRE_TYPE).load_gt_inventory()
    return dict(zip(df.iloc[:, 0].astype(str).str.strip(),
                    pd.to_numeric(df.iloc[:, 1], errors="coerce").fillna(0)))


def _apply_feasibility(shift_df, gt_built: dict, opening: dict):
    """Enforce day/shift feasibility: per SKU, cumulative cured cannot exceed
    opening GT inventory + cumulative GT built (same-day GT allowed). Curing that
    runs ahead of supply is trimmed. Returns (feasible_shift_df, records_df)."""
    from collections import defaultdict
    sho = {"A": 0, "B": 1, "C": 2}
    df = shift_df.copy()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)
    df["SKUCode"] = df["SKUCode"].astype(str).str.strip()
    df["Date"] = pd.to_datetime(df["Date"])
    cured = (df[df["Qty"] > 0].groupby(["SKUCode", "Date", "Shift"])["Qty"]
             .sum().to_dict())
    sku_keys = defaultdict(set)
    for (s, d, sh) in set(cured) | set(gt_built):
        sku_keys[s].add((d, sh))

    feasible, records = {}, []
    for s, keys in sku_keys.items():
        avail = float(opening.get(s, 0.0)); cum = 0.0
        for d, sh in sorted(keys, key=lambda x: (x[0], sho.get(x[1], 9))):
            avail += float(gt_built.get((s, d, sh), 0.0))
            c = float(cured.get((s, d, sh), 0.0))
            if c <= 0:
                continue
            f = min(c, max(0.0, avail - cum)); cum += f
            feasible[(s, d, sh)] = f
            if f < c - 1e-6:
                records.append({"SKUCode": s, "Date": d.date(), "Shift": sh,
                                "GT_available_cum": round(avail),
                                "Cured_requested": round(c),
                                "Cured_feasible": round(f),
                                "Trimmed_infeasible": round(c - f)})

    def _scale(r):
        if r["Qty"] <= 0:
            return r["Qty"]
        key = (r["SKUCode"], r["Date"], r["Shift"]); raw = cured.get(key, 0.0)
        return r["Qty"] * feasible.get(key, raw) / raw if raw > 0 else r["Qty"]

    df["Qty"] = df.apply(_scale, axis=1).round()
    feas = df[(df["Qty"] > 0) | df["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"])].copy()
    return feas, pd.DataFrame(records)


def phaseC2_final_curing(cfg: CBCConfig, engine) -> dict | None:
    """
    Phase C2 — the final C in C->B->C, made physically realistic:

      #2  schedule only SKUs that have BOTH curing history (CURING_PCR) and
          building history (productionevent_data); the rest are reported with an
          explicit 'No historical data (...)' infeasibility reason.

      #1  the schedule is made DAY/SHIFT feasible — per SKU, cumulative cured may
          not exceed opening GT inventory + cumulative GT built by building up to
          that shift. Curing that ran ahead of building's supply is trimmed.
    """
    print("\n" + "=" * 70)
    print("  PHASE C2 — FINAL CURING (day/shift feasible, historical-overlap SKUs)")
    print("=" * 70)
    if not os.path.exists(cfg.BUILDING_OUTPUT):
        print(f"  [Phase C2] building output not found ({cfg.BUILDING_OUTPUT}); "
              f"run building first. Skipped.")
        return None

    # ── #2 historical overlap ────────────────────────────────────────────
    overlap, infeasible_reasons = _historical_overlap(cfg)

    # ── building time-phased GT supply + opening inventory ───────────────
    gt_built = _building_gt_timeline(cfg)
    from collections import defaultdict
    supply_total = defaultdict(float)
    for (s, _d, _sh), q in gt_built.items():
        supply_total[s] += q
    opening = _opening_gt_inventory(engine)

    # ── demand: Book4 → overlap SKUs only → cap at building GT total + opening
    raw = pd.read_excel(cfg.CUR_DEMAND)
    qcol = "Updated_Requirement" if "Updated_Requirement" in raw.columns else "Requirement"
    pcol = "ConsolidatedPriorityScore"
    raw["SKUCode"] = raw["SKUCode"].astype(str).str.strip()
    raw[qcol] = pd.to_numeric(raw[qcol], errors="coerce").fillna(0)
    dem = (raw.groupby("SKUCode")
              .agg(**{qcol: (qcol, "sum"), pcol: (pcol, "max")})
              .reset_index())
    orig_total = dem[qcol].sum()
    orig_skus = int((dem[qcol] > 0).sum())
    dem = dem[dem["SKUCode"].str.upper().isin(overlap)].copy()          # #2 filter
    cap = dem["SKUCode"].map(lambda s: supply_total.get(s, 0.0) + opening.get(s, 0.0))
    dem[qcol] = pd.concat([dem[qcol], cap], axis=1).min(axis=1)
    dem = dem[dem[qcol] > 0]
    print(f"  [Phase C2] demand {int(orig_total):,} -> {int(dem[qcol].sum()):,} "
          f"across {len(dem)} overlap SKUs (capped at building GT supply)")
    dem.to_excel(cfg.FINAL_DEMAND_FILE, index=False, sheet_name="requirement_summary")

    results = curing_lp.run_from_database_simple(
        demand_path=cfg.FINAL_DEMAND_FILE,
        plan_start=cfg.PLAN_START,
        output_path=cfg.FINAL_CURING_OUTPUT,
        engine=engine,
    )

    # ── #1 day/shift feasibility trim ────────────────────────────────────
    feas_df, feas_records = _apply_feasibility(results["shift_schedule"], gt_built, opening)
    prod = feas_df[(pd.to_numeric(feas_df["Qty"], errors="coerce").fillna(0) > 0)
                   & (~feas_df["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"]))]
    planned_by_sku = prod.groupby("SKUCode")["Qty"].sum().to_dict()
    planned = float(prod["Qty"].sum())
    raw_planned = float(pd.to_numeric(
        results["shift_schedule"]["Qty"], errors="coerce").fillna(0).clip(lower=0).sum())
    trimmed = raw_planned - planned
    print(f"  [Phase C2] day/shift feasibility: trimmed {int(trimmed):,} infeasible "
          f"units ({100*trimmed/raw_planned:.1f}%) → feasible planned {int(planned):,}")

    # ── rewrite the affected sheets from the FEASIBLE schedule ───────────
    with pd.ExcelWriter(cfg.FINAL_CURING_OUTPUT, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as xw:
        feas_df.to_excel(xw, sheet_name="Shift Schedule", index=False)
        (feas_records if not feas_records.empty
         else pd.DataFrame(columns=["SKUCode", "Date", "Shift", "GT_available_cum",
                                    "Cured_requested", "Cured_feasible",
                                    "Trimmed_infeasible"])
         ).to_excel(xw, sheet_name="Shift-wise Feasibility", index=False)
    _append_daily_production(cfg.FINAL_CURING_OUTPUT, feas_df)
    _rebase_final_report_to_true_demand(cfg.FINAL_CURING_OUTPUT, cfg.CUR_DEMAND,
                                        planned_by_sku=planned_by_sku,
                                        infeasible_reasons=infeasible_reasons)

    # ── summary ──────────────────────────────────────────────────────────
    summary = pd.DataFrame([
        ("True curing demand (Book4)", round(orig_total)),
        ("SKUs in true demand", orig_skus),
        ("Schedulable SKUs (curing∩building history)", len(overlap)),
        ("SKUs actually scheduled (feasible)", len(planned_by_sku)),
        ("SKUs infeasible: no historical data", len(infeasible_reasons)),
        ("Curing planned before feasibility trim", round(raw_planned)),
        ("Trimmed: cured-before-built (day/shift)", round(trimmed)),
        ("Final feasible curing planned", round(planned)),
        ("True fulfilment % (feasible vs true demand)",
         f"{100.0 * planned / orig_total:.1f}%" if orig_total else "n/a"),
    ], columns=["Metric", "Value"])
    with pd.ExcelWriter(cfg.FINAL_CURING_OUTPUT, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as xw:
        summary.to_excel(xw, sheet_name="CBC True-Demand Summary", index=False)
    print(f"  [Phase C2] feasible final curing → {100.0*planned/orig_total:.1f}% "
          f"of true demand ({int(planned):,}/{int(orig_total):,})")
    print(f"\n  [Phase C2] final curing schedule -> {cfg.FINAL_CURING_OUTPUT}")
    return results


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════
def run_cbc(cfg: CBCConfig | None = None):
    cfg = cfg or CBCConfig()
    engine = cbc_env.make_engine()
    feed_map = phase0_feed_map(cfg)
    curing_results = phaseC_curing(cfg, feed_map, engine)
    building_results = phaseB_building(cfg, curing_results)
    final_results = None
    if cfg.RUN_FINAL_CURING and building_results is not None:
        final_results = phaseC2_final_curing(cfg, engine)
    print("\n" + "=" * 70)
    print("  CBC RUN COMPLETE")
    print(f"    curing (initial) : {cfg.CUR_OUTPUT}")
    if building_results is not None:
        print(f"    building         : {cfg.BUILDING_OUTPUT}")
    if final_results is not None:
        print(f"    curing (final)   : {cfg.FINAL_CURING_OUTPUT}")
    print("=" * 70)
    return {"curing": curing_results, "building": building_results,
            "final_curing": final_results}


def run_final_only(cfg: CBCConfig | None = None):
    """Regenerate ONLY the final curing schedule from the existing building
    output (no full pipeline re-run). Installs feed-awareness from the saved
    feed map so eligibility matches the main run."""
    cfg = cfg or CBCConfig()
    engine = cbc_env.make_engine()
    try:
        feed_map = load_feed_map(cfg.FEED_MAP_JSON)
        bld_allow = _building_allowable_from_db(engine)
        install_feed_awareness(curing_lp, feed_map, bld_allow,
                               keep_unknown_press=cfg.KEEP_UNKNOWN_PRESS,
                               keep_unknown_sku=cfg.KEEP_UNKNOWN_SKU)
    except Exception as e:  # noqa: BLE001
        print(f"  [final-only] feed-awareness not installed ({e}); "
              f"running curing without feed filter.")
    return phaseC2_final_curing(cfg, engine)


if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    if "final" in args:
        run_final_only()
    else:
        # `python cbc.py rebuild` (or REBUILD_FEED_MAP=1) rebuilds the feed map
        # from the building-pde files before running the pipeline. Use this after
        # adding/refreshing the building-pde exports; otherwise the saved
        # feed_map.json is reused.
        rebuild = "rebuild" in args or os.environ.get("REBUILD_FEED_MAP") in ("1", "true", "True")
        run_cbc(CBCConfig(REBUILD_FEED_MAP=True) if rebuild else None)
