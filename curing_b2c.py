"""
curing_b2c.py  —  B2C Curing Schedule Generator
=================================================
Derives a 31-day shift-level curing schedule directly from building output.

Algorithm (GT-balance simulation):
  • For each shift S, the GT produced by building in shift S-1 is added to
    per-SKU inventory.
  • Each active curing press cures up to floor(480/CT) × 2 units (2 cavities),
    limited by available GT inventory.
  • CO events from the building schedule's "Changeover Plan" sheet drive
    press-state transitions:
        Shift A of CO day → CHANGEOVER (press occupied, no production)
        Shift B of CO day → MOULD_CLEAN (press occupied, no production)
        Shift C of CO day → RUNNING on new SKU
  • Press state initialised from testing_Daily_Running_Moulds.

Output Excel (6 sheets — same column format as curing_lp.py):
  1. Demand Fulfillment
  2. Machine Utilization
  3. Shift Schedule
  4. Mould Tracker
  5. Machine Schedule
  6. Daily Cured tyres

Run:
    python curing_b2c.py                                    # uses defaults
    python curing_b2c.py <building_path> <output_path>      # custom paths
"""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import cbc_env

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

# ── All scheduling params imported from bc_config (single source of truth) ───
from bc_config import (
    PLAN_START,
    PLANNING_DAYS,
    SHIFT_MINS,
    SHIFT_NAMES  as SHIFTS,
    SHIFT_STARTS,
    SHIFT_ENDS,
    CAVITIES_PER_PRESS as CAVITIES,
    DEFAULT_CURING_CT  as DEFAULT_CT,
    GT_MACHINES,
)

DB = cbc_env.ENV.get("JKT_DB_DATABASE", "jkplanningV1")

# ── Excel colour palette (same as curing_lp.py) ───────────────────────────────
_GREEN  = "C6EFCE"
_AMBER  = "FFEB9C"
_RED    = "FFC7CE"
_LGREY  = "D9D9D9"
_NAVY   = "1F3864"
_WHITE  = "FFFFFF"
_BLUE   = "DCE6F1"   # shift A
_LYELL  = "FFF2CC"   # shift B
_DGREY  = "F2F2F2"   # shift C
_ORANGE = "FFC000"   # changeover


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _bold(size: int = 10, color: str = "000000") -> Font:
    return Font(bold=True, size=size, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header(ws, row: int, cols: list[str], bg: str = _NAVY, fg: str = _WHITE):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _fill(bg)
        cell.font = Font(bold=True, size=10, color=fg)
        cell.alignment = _center()


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADERS
# ══════════════════════════════════════════════════════════════════════════════

def _sql(engine, q: str) -> pd.DataFrame:
    with engine.connect() as conn:
        return pd.read_sql(q, conn)


def _load_building_output(path: str) -> tuple[dict, dict]:
    """
    Returns:
        build_by_shift: (date_str, shift) -> {sku: qty}
        co_events:      list of (press_str, co_day_0based, old_sku, new_sku)
    """
    xl = pd.ExcelFile(path)

    # ── Building output ───────────────────────────────────────────────────────
    if "Shift Schedule (Clean)" in xl.sheet_names:
        df = xl.parse("Shift Schedule (Clean)")
    else:
        df = xl.parse("Shift Schedule", header=1)

    df.columns = [str(c).strip() for c in df.columns]
    col = {c.lower().replace(" ", "_"): c for c in df.columns}
    sku_c  = col.get("skucode",  col.get("sku_code",  "SKUCode"))
    date_c = col.get("date",     col.get("date2",     "Date"))
    shft_c = col.get("shift",    "Shift")
    qty_c  = col.get("qty",      "Qty")
    mach_c = col.get("machine",  "Machine")

    df[mach_c] = df[mach_c].astype(str).str.strip()
    df[sku_c]  = df[sku_c].astype(str).str.strip()
    df[qty_c]  = pd.to_numeric(df[qty_c], errors="coerce").fillna(0)
    df[date_c] = pd.to_datetime(df[date_c], errors="coerce").dt.strftime("%Y-%m-%d")

    df_gt = df[
        df[mach_c].isin(GT_MACHINES) &
        (~df[sku_c].str.upper().isin({"CHANGEOVER","MOULD_CLEAN","C/O","CLEANING","NAN",""})) &
        (df[qty_c] > 0) &
        df[date_c].notna()
    ]

    build_by_shift: dict[tuple, dict] = defaultdict(lambda: defaultdict(float))
    for _, r in df_gt.iterrows():
        build_by_shift[(str(r[date_c]), str(r[shft_c]).strip())][str(r[sku_c])] += float(r[qty_c])

    # ── CO plan ───────────────────────────────────────────────────────────────
    co_events: list[tuple] = []
    if "Changeover Plan" in xl.sheet_names:
        dfc = xl.parse("Changeover Plan")
        dfc.columns = [str(c).strip() for c in dfc.columns]
        if "Status" in dfc.columns:
            dfc = dfc[dfc["Status"].astype(str).str.strip().str.upper() == "SCHEDULED"]
        for _, r in dfc.iterrows():
            press   = str(r.get("Press", "")).strip()
            new_sku = str(r.get("Target_SKU", "")).strip()
            old_sku = str(r.get("Old_SKU", "")).strip()
            try:
                day_idx = int(r.get("CO_Day_Index", 0))
            except (ValueError, TypeError):
                day_idx = 0
            if press and new_sku:
                co_events.append((press, day_idx, old_sku, new_sku))

    return dict({k: dict(v) for k, v in build_by_shift.items()}), co_events


def _load_demand(demand_path: Optional[str]) -> pd.DataFrame:
    if not demand_path or not os.path.exists(demand_path):
        return pd.DataFrame(columns=["SKUCode", "Demand", "Priority"])
    df = pd.read_excel(demand_path)
    df.columns = [str(c).strip() for c in df.columns]
    col = {c.lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    sku_c = col.get("skucode", df.columns[0])
    qty_c = col.get("updatedrequirement", col.get("requirement", col.get("quantity", df.columns[1])))
    pri_c = col.get("consolidatedpriorityscore", col.get("priority", None))
    out = pd.DataFrame()
    out["SKUCode"]  = df[sku_c].astype(str).str.strip()
    out["Demand"]   = pd.to_numeric(df[qty_c], errors="coerce").fillna(0)
    out["Priority"] = (pd.to_numeric(df[pri_c], errors="coerce").fillna(1.0)
                       if pri_c else 1.0)
    return out[out["Demand"] > 0].reset_index(drop=True)


def _load_cycle_times(engine) -> dict:
    try:
        df = _sql(engine,
            f"SELECT Sapcode AS sku, `Cure Time` AS raw_ct "
            f"FROM {DB}.Master_Curing_Design_CycleTime")
        df["sku"] = df["sku"].astype(str).str.strip()
        df["ct"]  = (pd.to_numeric(df["raw_ct"], errors="coerce") + 2.3) / 0.94
        return {r["sku"]: float(r["ct"]) for _, r in df.iterrows() if pd.notna(r["ct"])}
    except Exception as e:
        print(f"  ⚠  Cycle times: {e}")
        return {}


def _load_opening_gt(engine) -> dict:
    try:
        df = _sql(engine,
            f"SELECT sizeCode AS sku, gtInventory AS qty "
            f"FROM {DB}.gt_inventory_manual")
        df["sku"] = df["sku"].astype(str).str.strip()
        df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
        return {r["sku"]: float(r["qty"]) for _, r in df.iterrows()}
    except Exception as e:
        print(f"  ⚠  Opening GT inventory: {e}")
        return {}


def _load_press_state(engine) -> pd.DataFrame:
    """
    Returns DataFrame with columns: press, sku, mould_life
    press = WCNAME_clean (e.g. "75206") — same format as CO events from
    curing_consumption_dynamic.py. MUST NOT use wcID — that format never
    matches CO event press IDs and silently breaks all CO transitions.
    """
    try:
        rm = _sql(engine, f"SELECT * FROM {DB}.testing_Daily_Running_Moulds")
        if "updatedAt" in rm.columns:
            rm = rm.drop(columns=["updatedAt"])

        rm["press"]      = rm["WCNAME"].str.replace(r"(LH|RH)$", "", regex=True).str.strip()
        rm["sku"]        = rm["Sapcode"].astype(str).str.strip()
        rm["mould_life"] = pd.to_numeric(
            rm["Mould life"] if "Mould life" in rm.columns else 6000,
            errors="coerce",
        ).fillna(6000)

        valid = rm[
            rm["press"].notna() & (rm["press"] != "") &
            rm["sku"].notna()   & (rm["sku"]   != "") & (rm["sku"] != "nan")
        ].copy()

        # One row per press (LH+RH both strip to same WCNAME_clean with same SKU)
        return valid[["press", "sku", "mould_life"]].drop_duplicates("press").reset_index(drop=True)
    except Exception as e:
        print(f"  ⚠  Press state: {e}")
        return pd.DataFrame(columns=["press", "sku", "mould_life"])


def _load_mould_tracker(engine) -> pd.DataFrame:
    try:
        rm  = _sql(engine, f"SELECT * FROM {DB}.testing_Daily_Running_Moulds")
        mms = _sql(engine,
            f"SELECT MouldNo, `Matl.Code` AS sku, `Active Flag` AS flag "
            f"FROM {DB}.Master_Mapping_Mould_SKU")

        active = mms[mms["flag"].astype(str).str.strip().str.upper() == "X"]
        compat = (active.groupby("MouldNo")["sku"]
                        .apply(lambda x: ", ".join(x.astype(str).str.strip()))
                        .reset_index()
                        .rename(columns={"sku": "Compatible_SKUs"}))

        assigned: dict[str, str] = {}
        life_map: dict[str, int] = {}
        if "Current MouldNo" in rm.columns:
            for _, r in rm.iterrows():
                life = int(pd.to_numeric(r.get("Mould life", 6000), errors="coerce") or 6000)
                for mn in str(r.get("Current MouldNo", "")).split(","):
                    mn = mn.strip()
                    if mn:
                        assigned[mn] = str(r.get("WCNAME", "FREE")).strip()
                        life_map[mn] = life

        rows = []
        for _, row in compat.iterrows():
            mn = str(row["MouldNo"]).strip()
            rows.append({
                "MouldNo":          mn,
                "Compatible_SKUs":  row["Compatible_SKUs"],
                "Life_Remaining":   life_map.get(mn, 6000),
                "Assigned_Machine": assigned.get(mn, "FREE"),
            })
        return pd.DataFrame(rows)
    except Exception as e:
        print(f"  ⚠  Mould tracker: {e}")
        return pd.DataFrame(columns=["MouldNo", "Compatible_SKUs", "Life_Remaining", "Assigned_Machine"])


def _load_curing_allowable(engine) -> dict:
    """SKUCode -> list of eligible press IDs (strings)."""
    try:
        df = _sql(engine, f"SELECT * FROM {DB}.Master_Curing_Allowable_Machines_source")
        df.columns = [str(c).strip() for c in df.columns]
        # First column is SKUCode; remaining columns are press IDs (as column names or values)
        sku_col = df.columns[0]
        result: dict[str, list] = {}
        for _, row in df.iterrows():
            sku = str(row[sku_col]).strip()
            presses = []
            for c in df.columns[1:]:
                val = str(row[c]).strip()
                if val not in ("", "nan", "None", "0"):
                    try:
                        presses.append(str(int(float(val))))
                    except (ValueError, TypeError):
                        pass
            if presses:
                result[sku] = presses
        return result
    except Exception as e:
        print(f"  ⚠  Curing allowable: {e}")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# CORE SIMULATION
# ══════════════════════════════════════════════════════════════════════════════

def _all_shifts(plan_start: datetime, planning_days: int) -> list[tuple[str, str]]:
    result = []
    for d in range(planning_days):
        ds = (plan_start + timedelta(days=d)).strftime("%Y-%m-%d")
        for s in SHIFTS:
            result.append((ds, s))
    return result


def _qty_per_shift(ct: float) -> int:
    return int(SHIFT_MINS / ct) * CAVITIES


def simulate(
    build_by_shift: dict,
    co_events: list[tuple],
    opening_gt: dict,
    press_df: pd.DataFrame,
    ct_map: dict,
    demand_df: pd.DataFrame,
    plan_start: datetime,
    planning_days: int,
) -> dict:
    """
    GT-balance shift-by-shift curing simulation.

    Returns:
        shift_rows:      list of row dicts (for Shift Schedule)
        daily_cured:     {date_str: total_cured}
        sku_cured:       {sku: total_cured}
        press_stats:     {press: {running_mins, co_mins, clean_mins, idle_mins,
                                  skus, units, cycles}}
        press_sku_stats: {(press, sku): {units, mins_used, cycles}}
    """
    all_shifts = _all_shifts(plan_start, planning_days)
    demand_skus = set(demand_df["SKUCode"].astype(str).str.strip()) if not demand_df.empty else set()

    # ── Init press state ──────────────────────────────────────────────────────
    #   status: RUNNING | IDLE | CHANGEOVER | MOULD_CLEAN
    press_state: dict[str, dict] = {}
    for _, r in press_df.iterrows():
        press = str(r["press"]).strip()
        sku   = str(r["sku"]).strip()
        if press and sku and sku != "nan":
            press_state[press] = {
                "sku":    sku,
                "status": "RUNNING" if sku in demand_skus else "IDLE",
            }

    # Add presses that appear in CO events but not in running moulds
    for press, day_idx, old_sku, new_sku in co_events:
        if press not in press_state and old_sku:
            press_state[press] = {"sku": old_sku, "status": "IDLE"}

    # ── GT balance ────────────────────────────────────────────────────────────
    gt_bal: dict[str, float] = defaultdict(float)
    for sku, qty in opening_gt.items():
        gt_bal[str(sku).strip()] += float(qty)

    # Credit pre-plan-start building shifts (PRE_START_SHIFTS) to opening balance.
    # These shifts (e.g. Apr 30 Shift B/C for a May 1 plan) appear in build_by_shift
    # but not in all_shifts, so the idx-1 logic in the main loop never adds them.
    for (ds, _sh), sku_qty in build_by_shift.items():
        try:
            if datetime.strptime(ds, "%Y-%m-%d") < plan_start:
                for sku, qty in sku_qty.items():
                    gt_bal[str(sku).strip()] += float(qty)
        except (ValueError, TypeError):
            pass

    # ── CO transitions: (date_str, shift) -> list of (press, status, sku) ────
    co_trans: dict[tuple, list] = defaultdict(list)
    for press, day_idx, old_sku, new_sku in co_events:
        co_date = (plan_start + timedelta(days=day_idx)).strftime("%Y-%m-%d")
        co_trans[(co_date, "A")].append((press, "CHANGEOVER",  new_sku))
        co_trans[(co_date, "B")].append((press, "MOULD_CLEAN", new_sku))
        co_trans[(co_date, "C")].append((press, "RUNNING",     new_sku))

    # ── Accumulators ─────────────────────────────────────────────────────────
    shift_rows: list[dict] = []
    daily_cured: dict[str, float] = defaultdict(float)
    sku_cured:   dict[str, float] = defaultdict(float)
    press_stats:     dict[str, dict]         = defaultdict(lambda: {
        "running_mins": 0.0, "co_mins": 0.0, "clean_mins": 0.0,
        "idle_mins": 0.0, "skus": set(), "units": 0, "cycles": 0,
    })
    press_sku_stats: dict[tuple, dict] = defaultdict(lambda: {"units": 0, "mins_used": 0.0, "cycles": 0})

    # ── Simulation ────────────────────────────────────────────────────────────
    for idx, (date_str, shift) in enumerate(all_shifts):

        # 1. Add prior-shift building output to GT balance
        if idx > 0:
            prev_date, prev_shift = all_shifts[idx - 1]
            for sku, qty in build_by_shift.get((prev_date, prev_shift), {}).items():
                gt_bal[sku] += qty

        # 2. Apply CO transitions for this shift
        for press, new_status, new_sku in co_trans.get((date_str, shift), []):
            press_state[press] = {"sku": new_sku, "status": new_status}

        # 3. Simulate each press
        for press in sorted(press_state):
            state  = press_state[press]
            sku    = state["sku"]
            status = state["status"]
            ct     = ct_map.get(sku, DEFAULT_CT)
            cap    = _qty_per_shift(ct)
            cured  = 0
            gt_after = gt_bal.get(sku, 0.0)
            remark = status

            if status == "RUNNING":
                avail = max(0.0, gt_bal.get(sku, 0.0))
                cured = min(cap, int(avail))
                gt_bal[sku] = avail - cured
                gt_after    = gt_bal[sku]
                prod_mins   = cured * ct / CAVITIES
                press_stats[press]["running_mins"] += prod_mins
                press_stats[press]["skus"].add(sku)
                press_stats[press]["units"]  += cured
                press_stats[press]["cycles"] += cured // CAVITIES
                press_sku_stats[(press, sku)]["units"]    += cured
                press_sku_stats[(press, sku)]["mins_used"] += prod_mins
                press_sku_stats[(press, sku)]["cycles"]   += cured // CAVITIES
                remark = "PRODUCTION" if cured > 0 else "WAITING_GT"
                if avail <= 0:
                    press_stats[press]["idle_mins"] += SHIFT_MINS
            elif status == "CHANGEOVER":
                press_stats[press]["co_mins"] += SHIFT_MINS
                remark = f"C/O → {sku}"
            elif status == "MOULD_CLEAN":
                press_stats[press]["clean_mins"] += SHIFT_MINS
                remark = "MOULD_CLEAN"
            else:  # IDLE
                press_stats[press]["idle_mins"] += SHIFT_MINS
                remark = "IDLE"

            shift_rows.append({
                "Date":         date_str,
                "Shift":        shift,
                "Machine":      press,
                "SKUCode":      sku if status not in ("IDLE",) else "-",
                "StartTime":    SHIFT_STARTS[shift],
                "EndTime":      SHIFT_ENDS[shift],
                "Qty":          cured,
                "CycleTime_min": round(ct, 2),
                "GT_Inventory": round(gt_after, 0),
                "Remarks":      remark,
                "_status":      status,
            })

            if cured > 0:
                daily_cured[date_str] += cured
                sku_cured[sku] += cured

    return {
        "shift_rows":      shift_rows,
        "daily_cured":     dict(daily_cured),
        "sku_cured":       dict(sku_cured),
        "press_stats":     dict(press_stats),
        "press_sku_stats": dict(press_sku_stats),
        "closing_gt_bal":  dict(gt_bal),      # GT remaining at end of horizon
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITERS  (format matches curing_lp.py sheet layout)
# ══════════════════════════════════════════════════════════════════════════════

def _write_demand_fulfillment(ws, demand_df, sku_cured, ct_map, curing_allowable, planning_days):
    cols = ["SKUCode", "Priority", "Demand", "GT_Inventory", "Planned_Units",
            "Gap", "Fulfillment_Pct", "Status", "CycleTime_min",
            "Eligible_Machines", "Presses_Needed", "Skip_Reason"]
    _header(ws, 1, cols)

    status_fill = {
        "FULLY MET": _fill(_GREEN),
        "PARTIAL":   _fill(_AMBER),
        "UNMET":     _fill(_RED),
        "NO DATA":   _fill(_LGREY),
    }

    rows = []
    for _, r in demand_df.iterrows():
        sku     = str(r["SKUCode"]).strip()
        demand  = float(r.get("Demand", 0))
        pri     = float(r.get("Priority", 1.0))
        gt_inv  = float(r.get("GT_Inventory", 0))
        planned = float(sku_cured.get(sku, 0))
        gap     = max(0.0, demand - planned)
        pct     = planned / demand if demand > 0 else 0.0
        ct      = ct_map.get(sku, DEFAULT_CT)
        machines= curing_allowable.get(sku, [])
        cap_day = _qty_per_shift(ct) * 3 * planning_days if ct else 1
        p_needed= max(1, round(demand / cap_day)) if cap_day > 0 else "-"

        if demand <= 0:
            status = "NO DATA"
        elif planned >= demand * 0.999:
            status = "FULLY MET"
        elif planned > 0:
            status = "PARTIAL"
        else:
            status = "UNMET"

        rows.append({
            "SKUCode": sku, "Priority": round(pri, 4), "Demand": int(demand),
            "GT_Inventory": int(gt_inv), "Planned_Units": int(planned),
            "Gap": int(gap), "Fulfillment_Pct": pct, "Status": status,
            "CycleTime_min": round(ct, 2), "Eligible_Machines": len(machines),
            "Presses_Needed": p_needed, "Skip_Reason": "",
        })

    for ri, r in enumerate(rows, 2):
        fill = status_fill.get(r["Status"], _fill(_WHITE))
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r[h])
            cell.fill = fill
            cell.alignment = _center()
            if h == "Fulfillment_Pct":
                cell.number_format = "0.0%"

    # Totals
    n = len(rows)
    tr = n + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = _bold(11)
    ws.cell(row=tr, column=3, value=sum(r["Demand"]        for r in rows)).font = _bold(11)
    ws.cell(row=tr, column=5, value=sum(r["Planned_Units"] for r in rows)).font = _bold(11)
    ws.cell(row=tr, column=6, value=sum(r["Gap"]           for r in rows)).font = _bold(11)
    tot_d = sum(r["Demand"] for r in rows)
    tot_p = sum(r["Planned_Units"] for r in rows)
    pct_c = ws.cell(row=tr, column=7, value=tot_p / tot_d if tot_d else 0)
    pct_c.font = _bold(11)
    pct_c.number_format = "0.0%"

    ws.column_dimensions["A"].width = 34
    for ltr in "BCDEFGHIJKL":
        ws.column_dimensions[ltr].width = 15
    ws.freeze_panes = "A2"


def _write_machine_utilization(ws, press_stats, planning_days, total_press_count: int = 0):
    avail = planning_days * 3 * SHIFT_MINS  # per press

    # Summary row first
    all_presses = sorted(press_stats)
    if all_presses:
        avg_u = sum(
            press_stats[p]["running_mins"] / avail
            for p in all_presses
        ) / len(all_presses)
        high = sum(1 for p in all_presses if press_stats[p]["running_mins"] / avail >= 0.90)
        low  = sum(1 for p in all_presses if press_stats[p]["running_mins"] / avail < 0.05)
        sim_count = len(all_presses)
        plant_note = f"  |  Plant total: {total_press_count}" if total_press_count else ""
        ws.cell(row=1, column=1,
                value=(f"Avg util: {avg_u:.1%}  |  High(≥90%): {high}  |  "
                       f"Idle(<5%): {low}  |  In simulation: {sim_count}{plant_note}")
                ).font = _bold(10)

    cols = ["Machine", "Available_Mins", "Used_Mins", "Idle_Mins",
            "Utilization_Pct", "SKUs_Count", "Total_Cycles", "Total_Units"]
    _header(ws, 2, cols)

    for ri, press in enumerate(all_presses, 3):
        s    = press_stats[press]
        used = s["running_mins"]
        idle = avail - used - s["co_mins"] - s["clean_mins"]
        pct  = used / avail if avail else 0.0
        color = _GREEN if pct >= 0.90 else (_AMBER if pct >= 0.60 else _RED)
        fill  = _fill(color)
        vals  = [press, avail, round(used), round(max(0, idle)), pct,
                 len(s["skus"]), s["cycles"], s["units"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = fill
            cell.alignment = _center()
            if ci == 5:  # Utilization_Pct
                cell.number_format = "0.0%"

    for ltr in "ABCDEFGH":
        ws.column_dimensions[ltr].width = 17
    ws.freeze_panes = "A3"


def _write_shift_schedule(ws, shift_rows):
    cols = ["Date", "Shift", "Machine", "SKUCode", "StartTime", "EndTime",
            "Qty", "CycleTime_min", "GT_Inventory", "Remarks"]
    _header(ws, 1, cols)

    s_fill = {"A": _fill(_BLUE), "B": _fill(_LYELL), "C": _fill(_DGREY)}

    for ri, r in enumerate(shift_rows, 2):
        st = r.get("_status", "")
        if st == "CHANGEOVER":
            fill = _fill(_ORANGE)
        elif st == "MOULD_CLEAN":
            fill = _fill(_AMBER)
        elif st == "IDLE":
            fill = _fill(_LGREY)
        else:
            fill = s_fill.get(r["Shift"], _fill(_WHITE))

        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r.get(h, ""))
            cell.fill = fill
            cell.alignment = _center()
            if st in ("CHANGEOVER", "MOULD_CLEAN"):
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["J"].width = 20
    ws.freeze_panes = "A2"


def _write_mould_tracker(ws, mould_df):
    cols = ["MouldNo", "Compatible_SKUs", "Life_Remaining", "Assigned_Machine"]
    _header(ws, 1, cols)

    for ri, (_, r) in enumerate(mould_df.iterrows(), 2):
        is_free = str(r.get("Assigned_Machine", "FREE")).upper() == "FREE"
        fill    = _fill(_GREEN) if is_free else _fill(_AMBER)
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r.get(h, ""))
            cell.fill = fill
            cell.alignment = _center()

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22


def _write_machine_schedule(ws, press_sku_stats, ct_map, demand_df, planning_days):
    pri_map = (dict(zip(demand_df["SKUCode"].astype(str).str.strip(),
                        demand_df["Priority"].astype(float)))
               if not demand_df.empty else {})

    rows = []
    for (press, sku), s in sorted(press_sku_stats.items()):
        if s["units"] == 0:
            continue
        ct       = ct_map.get(sku, DEFAULT_CT)
        days_used = s["mins_used"] / (3 * SHIFT_MINS) if s["mins_used"] else 0
        rows.append({
            "Machine": press, "SKUCode": sku,
            "Priority": round(pri_map.get(sku, 1.0), 4),
            "CycleTime_min": round(ct, 2),
            "Cycles": s["cycles"],
            "Units_Planned": s["units"],
            "Mins_Used": round(s["mins_used"]),
            "Days_Used": round(days_used, 2),
        })

    rows.sort(key=lambda r: (r["Machine"], -r["Units_Planned"]))
    tot_u = sum(r["Units_Planned"] for r in rows)
    tot_c = sum(r["Cycles"] for r in rows)

    ws.cell(row=1, column=1,
            value=f"Press-SKU pairs: {len(rows)}  |  Total Units: {tot_u:,}  |  Total Cycles: {tot_c:,}"
            ).font = _bold(10)

    cols = ["Machine", "SKUCode", "Priority", "CycleTime_min",
            "Cycles", "Units_Planned", "Mins_Used", "Days_Used"]
    _header(ws, 2, cols)

    for ri, r in enumerate(rows, 3):
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r[h])
            cell.alignment = _center()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    for ltr in "CDEFGH":
        ws.column_dimensions[ltr].width = 15
    ws.freeze_panes = "A3"


def _write_daily_cured(ws, daily_cured, plan_start, planning_days):
    cols = ["Date", "Cured_Qty"]
    _header(ws, 1, cols)

    total = 0
    for d in range(planning_days):
        date_str = (plan_start + timedelta(days=d)).strftime("%Y-%m-%d")
        qty      = int(daily_cured.get(date_str, 0))
        ri       = d + 2
        ws.cell(row=ri, column=1, value=date_str).alignment = _center()
        c = ws.cell(row=ri, column=2, value=qty)
        c.alignment = _center()
        c.fill = _fill(_BLUE) if qty > 0 else _fill(_RED)
        total += qty

    tr = planning_days + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = _bold(11)
    t = ws.cell(row=tr, column=2, value=total)
    t.font = _bold(11)
    t.fill = _fill(_GREEN)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14


def _write_gt_balance_diagnostic(
    ws, closing_gt_bal, sku_cured, build_by_shift, opening_gt, press_state_skus, demand_df,
):
    """
    GT Balance Diagnostic sheet — explains the building→curing gap.

    For each SKU with non-zero closing GT balance, shows:
      GT_Built | GT_Cured | Closing_Balance | Reason
    Reason codes:
      NO_PRESS        — no curing press assigned (NRI-no-CO or IDLE press)
      DEMAND_MET      — press ran; closing balance is carry-over inventory
      RESIDUAL        — small rounding / last-shift lag
    """
    cols = ["SKUCode", "GT_Built", "GT_Cured", "Closing_Balance", "Reason"]
    _header(ws, 1, cols)

    # Compute total GT built per SKU from build_by_shift
    built_per_sku: dict[str, float] = {}
    for sku_qty in build_by_shift.values():
        for sku, qty in sku_qty.items():
            built_per_sku[str(sku).strip()] = built_per_sku.get(str(sku).strip(), 0.0) + float(qty)
    for sku, qty in opening_gt.items():
        built_per_sku[str(sku).strip()] = built_per_sku.get(str(sku).strip(), 0.0) + float(qty)

    demand_set = set(demand_df["SKUCode"].astype(str).str.strip()) if not demand_df.empty else set()

    ri = 2
    total_built = total_cured = total_closing = 0.0
    for sku in sorted(closing_gt_bal, key=lambda s: -closing_gt_bal[s]):
        bal = closing_gt_bal[sku]
        if bal < 0.5:
            continue   # negligible — skip
        built  = built_per_sku.get(sku, 0.0)
        cured  = float(sku_cured.get(sku, 0))
        if sku not in press_state_skus:
            reason = "NO_PRESS"
            fill   = _fill(_RED)
        elif cured > 0:
            reason = "DEMAND_MET"
            fill   = _fill(_AMBER)
        else:
            reason = "RESIDUAL"
            fill   = _fill(_LGREY)

        ws.cell(row=ri, column=1, value=sku)
        ws.cell(row=ri, column=2, value=round(built))
        ws.cell(row=ri, column=3, value=round(cured))
        c = ws.cell(row=ri, column=4, value=round(bal))
        c.fill = fill
        ws.cell(row=ri, column=5, value=reason)
        total_built   += built
        total_cured   += cured
        total_closing += bal
        ri += 1

    # Summary row
    ri += 1
    ws.cell(row=ri, column=1, value="TOTAL").font = _bold(11)
    ws.cell(row=ri, column=2, value=round(total_built)).font  = _bold(11)
    ws.cell(row=ri, column=3, value=round(total_cured)).font  = _bold(11)
    t = ws.cell(row=ri, column=4, value=round(total_closing))
    t.font = _bold(11)
    t.fill = _fill(_RED)
    note = ws.cell(row=ri + 2, column=1,
        value=("NO_PRESS = built but no curing press (main gap cause)  |  "
               "DEMAND_MET = carry-over to next month  |  "
               "RESIDUAL = last-shift lag (next month opening inventory)"))
    note.font = _bold(9)

    for col, w in zip("ABCDE", [28, 12, 12, 18, 16]):
        ws.column_dimensions[chr(ord("A") + "ABCDE".index(col))].width = w


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def run_curing_b2c(
    building_path:  str,
    output_path:    str,
    demand_path:    Optional[str] = None,
    engine=None,
    plan_start:     datetime = PLAN_START,
    planning_days:  int      = PLANNING_DAYS,
) -> dict:
    """
    Generate the B2C curing schedule from building output.

    Args:
        building_path: path to bc_building_schedule.xlsx
        output_path:   where to write the curing schedule Excel
        demand_path:   optional demand Excel (for Priority + Demand columns)
        engine:        SQLAlchemy engine (created automatically if None)
        plan_start:    first day of plan horizon
        planning_days: number of days

    Returns dict with keys: shift_rows, daily_cured, sku_cured, press_stats
    """
    if engine is None:
        engine = cbc_env.make_engine()

    print("\n" + "=" * 70)
    print("  B2C Curing Schedule Generator")
    print("=" * 70)

    print("  [ETL] Reading building schedule …")
    build_by_shift, co_events = _load_building_output(building_path)
    total_gt_input = sum(q for ds in build_by_shift.values() for q in ds.values())
    print(f"        GT entries: {len(build_by_shift)} shift-buckets  |  "
          f"CO events: {len(co_events)}  |  Total GT built: {total_gt_input:,.0f}")

    print("  [ETL] Loading cycle times …")
    ct_map = _load_cycle_times(engine)
    print(f"        {len(ct_map)} SKUs with cycle time")

    print("  [ETL] Loading opening GT inventory …")
    opening_gt = _load_opening_gt(engine)
    print(f"        {len(opening_gt)} SKUs with opening inventory, "
          f"total = {sum(opening_gt.values()):,.0f}")

    print("  [ETL] Loading press state (testing_Daily_Running_Moulds) …")
    press_df = _load_press_state(engine)
    print(f"        {len(press_df)} press rows loaded")

    print("  [ETL] Loading demand file …")
    demand_df = _load_demand(demand_path)
    print(f"        {len(demand_df)} SKUs in demand")

    print("  [ETL] Loading mould tracker …")
    mould_df = _load_mould_tracker(engine)

    print("  [ETL] Loading curing allowable machines …")
    curing_allowable = _load_curing_allowable(engine)

    # Attach opening GT inventory to demand_df for Demand Fulfillment sheet
    demand_df["GT_Inventory"] = demand_df["SKUCode"].map(opening_gt).fillna(0)

    print(f"\n  [Simulate] Running GT-balance curing simulation "
          f"({planning_days} days × 3 shifts) …")
    sim = simulate(
        build_by_shift = build_by_shift,
        co_events      = co_events,
        opening_gt     = opening_gt,
        press_df       = press_df,
        ct_map         = ct_map,
        demand_df      = demand_df,
        plan_start     = plan_start,
        planning_days  = planning_days,
    )
    total_cured = sum(sim["daily_cured"].values())
    print(f"  [Simulate] Total cured: {total_cured:,.0f} tyres")

    n_press = len(sim["press_stats"])
    avg_u   = (sum(s["running_mins"] for s in sim["press_stats"].values())
               / max(n_press * planning_days * 3 * SHIFT_MINS, 1))
    print(f"  [Simulate] Presses active: {n_press}  |  Avg utilization: {avg_u:.1%}")

    print(f"\n  [Output] Writing {os.path.basename(output_path)} …")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_demand_fulfillment(
        wb.create_sheet("Demand Fulfillment"),
        demand_df, sim["sku_cured"], ct_map, curing_allowable, planning_days,
    )
    _write_machine_utilization(
        wb.create_sheet("Machine Utilization"),
        sim["press_stats"], planning_days,
        total_press_count=len(press_df),
    )
    _write_shift_schedule(
        wb.create_sheet("Shift Schedule"),
        sim["shift_rows"],
    )
    _write_mould_tracker(
        wb.create_sheet("Mould Tracker"),
        mould_df,
    )
    _write_machine_schedule(
        wb.create_sheet("Machine Schedule"),
        sim["press_sku_stats"], ct_map, demand_df, planning_days,
    )
    _write_daily_cured(
        wb.create_sheet("Daily Cured tyres"),
        sim["daily_cured"], plan_start, planning_days,
    )

    # GT Balance Diagnostic — shows exactly where the building→curing gap goes
    # NO_PRESS rows = main cause of gap (GT built but no press to cure it)
    press_state_skus = {str(r["sku"]).strip() for _, r in press_df.iterrows()
                        if str(r.get("sku", "nan")).strip() not in ("", "nan")}
    _write_gt_balance_diagnostic(
        wb.create_sheet("GT Gap Diagnostic"),
        sim["closing_gt_bal"],
        sim["sku_cured"],
        build_by_shift,
        opening_gt,
        press_state_skus,
        demand_df,
    )

    # Summary print
    total_closing_gt = sum(v for v in sim["closing_gt_bal"].values() if v >= 0.5)
    no_press_gt = sum(
        v for sku, v in sim["closing_gt_bal"].items()
        if v >= 0.5 and sku not in press_state_skus
    )
    print(f"  [Gap] Closing GT balance: {total_closing_gt:,.0f} units "
          f"(NO_PRESS: {no_press_gt:,.0f}  |  carry-over: {total_closing_gt - no_press_gt:,.0f})")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"  ✓ Saved → {output_path}")
    print("=" * 70 + "\n")

    return sim


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import bc_config as _cfg

    _BUILDING = sys.argv[1] if len(sys.argv) > 1 else _cfg.BUILDING_OUTPUT
    _OUTPUT   = sys.argv[2] if len(sys.argv) > 2 else _cfg.CURING_B2C_OUTPUT

    run_curing_b2c(
        building_path = _BUILDING,
        output_path   = _OUTPUT,
        demand_path   = _cfg.DEMAND_FILE,
    )
