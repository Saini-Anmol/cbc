"""
B2C Pipeline — Phase 1: Building Scheduler (B2C mode)
======================================================
Schedules building machines driven by curing CONSUMPTION rather than a curing LP plan.

Key differences from CBC building.py:
  • Input:   curing_consumption_table.xlsx instead of a curing LP schedule
  • Demand:  derived from active press counts × qty_per_press_per_shift
  • GT inv:  opened from DB (not zeroed — real opening inventory used)
  • Start:   1 shift before curing starts (not 1 full day)
  • Cap:     building output ≤ consumption per shift (strict, OVERBUILD_BUFFER_FRAC=0.0)
  • Phasing: Runner-In scheduled first; Non-Runner-In + Runner-Out via joint priority pool

Reuses from building.py (imported):
  Config, ETL, LPMinuteSolver, GeneticOptimiser, CampaignSequencer,
  ScheduleBuilder, StarvationValidator, ExcelExporter, HybridDailyScheduler,
  build_summary, build_util, build_daywise_report, build_daily_sku_counts

Standalone usage:
    python building_b2c.py

Output: data/output/main_output/bc_building_schedule.xlsx
"""

from __future__ import annotations

import math
import os
import sys
import warnings
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ── venv re-exec ──────────────────────────────────────────────────────────────
_VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "myenv")
_VENV_PY  = os.path.join(_VENV_DIR, "bin", "python")
if (os.path.exists(_VENV_PY)
        and os.path.realpath(sys.prefix) != os.path.realpath(_VENV_DIR)
        and not os.environ.get("BC_REEXEC")):
    os.environ["BC_REEXEC"] = "1"
    os.execv(_VENV_PY, [_VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])

# ── Import reusable machinery from existing building.py ───────────────────────
import building as _bld
from building import (
    Config,
    ETL,
    LPMinuteSolver,
    GeneticOptimiser,
    CampaignSequencer,
    ScheduleBuilder,
    StarvationValidator,
    ExcelExporter,
    HybridDailyScheduler,
    build_summary,
    build_util,
    build_daywise_report,
    build_daily_sku_counts,
    _shift_fn,
    _shift_start,
)

import cbc_env

HERE    = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = cbc_env.OUTPUT_DIR
MAIN_OUT = os.path.join(OUT_DIR, "main_output")
os.makedirs(MAIN_OUT, exist_ok=True)

# ── B2C-specific config overrides ────────────────────────────────────────────
# MAX_CHANGEOVERS_PER_DAY, MIN_CAMPAIGN_MINS, BUILD_LEAD_SHIFTS are no longer
# hardcoded here. They are passed as parameters to run_from_database_b2c() and
# set on Config inside that function. Edit them in bc.py (single source of truth).
# OVERBUILD_BUFFER_FRAC: inherits 0.2 from building.py Config — do NOT override.
Config.CURING_PLAN_FILE = None  # not used in B2C (we use consumption table)

# ══════════════════════════════════════════════════════════════════════════════
# B2C ETL  (extends building ETL with consumption table reader)
# ══════════════════════════════════════════════════════════════════════════════

class B2C_ETL(ETL):
    """B2C variant of the building ETL — replaces load_curing_schedule()."""

    def load_consumption_table(self, path: str) -> pd.DataFrame:
        """
        Load the curing consumption table produced by curing_consumption.py.
        Returns DataFrame: [SKUCode, Category, Running_Press_Count,
                            Effective_CT_Min, Qty_Per_Press_Per_Shift,
                            Total_GT_Per_Shift_Day0, Demand_Qty, Priority_Score]
        """
        xl = pd.ExcelFile(path)
        sheet = "Consumption Summary" if "Consumption Summary" in xl.sheet_names else xl.sheet_names[0]
        # Try multiple header rows in case a legend row is present
        for hdr in (0, 1, 2):
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=hdr)
                if "SKUCode" in df.columns and "Category" in df.columns:
                    break
            except Exception:
                continue
        df["SKUCode"] = df["SKUCode"].astype(str).str.strip()
        # Drop legend/summary rows that the Excel reader picks up as data
        df = df[df["SKUCode"].notna() & (df["SKUCode"] != "") & (df["SKUCode"] != "nan")]
        _valid_cats = {"Runner-In", "Non-Runner-In"}
        if "Category" in df.columns:
            df = df[df["Category"].isin(_valid_cats)]
        df = df.reset_index(drop=True)
        print(f"  [B2C ETL] Consumption table: {len(df)} rows from {os.path.basename(path)}")
        return df

    def load_gt_inventory_for_b2c(self) -> pd.DataFrame:
        """Load REAL opening GT inventory (not zeroed out as in CBC cold-start)."""
        return self._sql(
            f"SELECT sizeCode AS SKUCode, gtInventory AS GT_Inventory "
            f"FROM {Config.DB_NAME}.gt_inventory_manual"
        )


# ══════════════════════════════════════════════════════════════════════════════
# B2C DEMAND DERIVER
# ══════════════════════════════════════════════════════════════════════════════

class B2C_DemandDeriver:
    """
    Converts the curing consumption table into the per-(SKU, shift) demand matrix
    that HybridDailyScheduler/LPMinuteSolver expects.

    Runner-In:      curing_matrix[s, t] = Total_GT_Per_Shift_Day0[s]  (constant demand)
    Non-Runner-In:  curing_matrix[s, t] = 0 initially (no press yet; demand from file)
    """

    def derive_from_consumption(
        self,
        df_consumption: pd.DataFrame,
        df_gt_inv: pd.DataFrame,
        plan_start: datetime,
        planning_days: int = 1,
    ) -> tuple:
        """
        Returns the same tuple as building.DemandDeriver.derive():
          (df_sku_demand, curing_matrix, all_skus, df_shift_demand)
        """
        T = planning_days * Config.SHIFTS_PER_DAY
        gt_inv_map = dict(zip(
            df_gt_inv["SKUCode"].astype(str).str.strip(),
            df_gt_inv["GT_Inventory"].astype(float)
        ))

        shift_starts = [
            _shift_start(plan_start, t) for t in range(T)
        ]
        plan_end = plan_start + timedelta(days=planning_days)

        # Build per-(SKU, shift) demand
        sku_shift_demand: dict[tuple, float] = {}

        for _, row in df_consumption.iterrows():
            sku      = str(row["SKUCode"]).strip()
            cat      = str(row.get("Category", "")).strip()
            qty_per  = float(row.get("Total_GT_Per_Shift_Day0", 0))
            dem_qty  = float(row.get("Demand_Qty", 0))

            if cat == "Runner-In":
                # Cap at customer demand spread evenly — building must not exceed demand.
                # Press consumption rate is the physical ceiling; demand is the business cap.
                # If demand < press consumption, build only to demand (press will idle/CO after).
                if dem_qty > 0:
                    demand_per_shift = dem_qty / (planning_days * Config.SHIFTS_PER_DAY)
                    capped = min(qty_per, demand_per_shift)
                else:
                    capped = qty_per  # no demand info — fall back to press consumption
                for t in range(T):
                    sku_shift_demand[(sku, t)] = capped
            elif cat == "Non-Runner-In" and dem_qty > 0:
                # Spread total demand evenly across shifts
                per_shift = dem_qty / (planning_days * Config.SHIFTS_PER_DAY)
                for t in range(T):
                    sku_shift_demand[(sku, t)] = per_shift
            # Runner-Out: excluded from building demand (they'll changeover)

        if not sku_shift_demand:
            # Fallback: empty — shouldn't happen in practice
            return (
                pd.DataFrame(columns=["SKUCode", "GT_Demand", "GT_Inventory",
                                      "Net_GT_Demand", "LP_Demand",
                                      "First_Curing_Start", "Burn_Rate_Per_Shift",
                                      "Active_Shifts"]),
                np.zeros((0, T)), [], pd.DataFrame()
            )

        all_skus = sorted(set(k[0] for k in sku_shift_demand))
        S = len(all_skus)
        sku_idx = {s: i for i, s in enumerate(all_skus)}
        curing_matrix = np.zeros((S, T))
        for (sku, t), qty in sku_shift_demand.items():
            if sku in sku_idx:
                curing_matrix[sku_idx[sku], t] = qty

        total_gt = curing_matrix.sum(axis=1)
        rows = []
        for si, sku in enumerate(all_skus):
            gt_inv   = gt_inv_map.get(sku, 0.0)
            net_dem  = max(0.0, total_gt[si] - gt_inv)
            first_t  = next((t for t in range(T) if curing_matrix[si, t] > 0), T)
            first_start = shift_starts[first_t] if first_t < T else plan_end
            rows.append({
                "SKUCode":             sku,
                "GT_Demand":           int(total_gt[si]),
                "GT_Inventory":        int(gt_inv),
                "Net_GT_Demand":       int(net_dem),
                "LP_Demand":           int(net_dem),
                "First_Curing_Start":  first_start,
                "Burn_Rate_Per_Shift": round(total_gt[si] / max(T, 1), 1),
                "Active_Shifts":       int((curing_matrix[si] > 0).sum()),
            })

        df_sku_demand = pd.DataFrame(rows).sort_values("First_Curing_Start")

        sd_rows = []
        for (sku, t), qty in sku_shift_demand.items():
            if qty > 0:
                s_start = shift_starts[t]
                shift_lbl, _ = _shift_fn(s_start)
                sd_rows.append({
                    "Date": s_start.date(), "Shift": shift_lbl,
                    "SKUCode": sku, "Curing_Qty": int(qty),
                })
        df_shift_demand = pd.DataFrame(sd_rows)

        print(f"  [B2C Demand] SKUs: {S} | Shifts: {T} | "
              f"Total consumption: {total_gt.sum():,.0f}")
        return df_sku_demand, curing_matrix, all_skus, df_shift_demand


# ══════════════════════════════════════════════════════════════════════════════
# SYNTHETIC CURING SCHEDULE BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def _make_synthetic_curing(
    df_consumption: pd.DataFrame,
    plan_start: datetime,
    planning_days: int,
    nri_skus_to_build: set | None = None,
    co_day_map: dict | None = None,
) -> pd.DataFrame:
    """
    Convert the consumption table into a df_curing-compatible DataFrame so
    the existing HybridDailyScheduler can consume it without modification.

    Produces one row per (SKU, shift) for the full planning horizon.

    nri_skus_to_build: if provided, only NRI SKUs in this set get LP demand;
    inactive NRI presses (no CO scheduled) are excluded to avoid wasting
    building capacity on uncurable green tyres.
    """
    rows = []
    shift_hours = [
        Config.SHIFT_START_HOUR,
        Config.SHIFT_START_HOUR + Config.HOURS_PER_SHIFT,
        Config.SHIFT_START_HOUR + Config.HOURS_PER_SHIFT * 2,
    ]

    for _, row in df_consumption.iterrows():
        sku      = str(row["SKUCode"]).strip()
        cat      = str(row.get("Category", "")).strip()
        qty_ps   = float(row.get("Total_GT_Per_Shift_Day0", 0))
        dem_qty  = float(row.get("Demand_Qty", 0))

        if cat == "Runner-In":
            # Cap synthetic curing target at customer demand — same cap as demand deriver.
            if dem_qty > 0:
                demand_spread = dem_qty / (planning_days * Config.SHIFTS_PER_DAY)
                target_qty = min(qty_ps, demand_spread)
            else:
                target_qty = qty_ps
        elif cat == "Non-Runner-In" and dem_qty > 0:
            # Only build NRI SKUs with a planned CO (active press); skip the rest
            # to concentrate building capacity on Runner-In curing demand.
            if nri_skus_to_build is not None and sku not in nri_skus_to_build:
                continue
            # Front-load building demand before the CO day so the LP pre-builds
            # a buffer of GT before the curing press fires on CO Shift C.
            # 70% of demand distributed over shifts 0 … co_day*3-1 (pre-CO),
            # 30% over shifts co_day*3 … end (post-CO running phase).
            # If CO day is unknown / Day 0, use flat distribution.
            co_day = (co_day_map.get(sku) if co_day_map else None)
            total_shifts = planning_days * Config.SHIFTS_PER_DAY
            if co_day is not None and 0 < co_day < planning_days:
                n_pre  = co_day * Config.SHIFTS_PER_DAY
                n_post = total_shifts - n_pre
                pre_rate  = (0.7 * dem_qty) / n_pre   if n_pre  > 0 else 0.0
                post_rate = (0.3 * dem_qty) / n_post  if n_post > 0 else 0.0
            else:
                pre_rate = post_rate = dem_qty / total_shifts

            for day_offset in range(planning_days):
                base_date = plan_start + timedelta(days=day_offset)
                shift_qty = pre_rate if (co_day is not None and day_offset < co_day) else post_rate
                for sh_h in shift_hours:
                    sh_start = datetime(
                        base_date.year, base_date.month, base_date.day, sh_h % 24, 0, 0
                    )
                    if sh_h >= 24:
                        sh_start += timedelta(days=1)
                    sh_end = sh_start + timedelta(hours=Config.HOURS_PER_SHIFT)
                    rows.append({
                        "SKUCode":   sku,
                        "StartTime": sh_start,
                        "EndTime":   sh_end,
                        "Qty":       shift_qty,
                    })
            continue   # NRI loop handled inline above — skip shared append below
        else:
            continue  # Runner-Out: no building demand

        for day_offset in range(planning_days):
            base_date = plan_start + timedelta(days=day_offset)
            for sh_h in shift_hours:
                sh_start = datetime(
                    base_date.year, base_date.month, base_date.day, sh_h % 24, 0, 0
                )
                if sh_h >= 24:
                    sh_start += timedelta(days=1)
                sh_end = sh_start + timedelta(hours=Config.HOURS_PER_SHIFT)
                rows.append({
                    "SKUCode":   sku,
                    "StartTime": sh_start,
                    "EndTime":   sh_end,
                    "Qty":       target_qty,
                })

    if not rows:
        return pd.DataFrame(columns=["SKUCode", "StartTime", "EndTime", "Qty"])
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# CHANGEOVER SCHEDULER  (Phase 2b — joint priority pool assignments)
# ══════════════════════════════════════════════════════════════════════════════

class ChangeoverScheduler:
    """
    Assigns changeover days from the joint priority pool:
      1. Combine Non-Runner-In SKUs + Runner-Out press candidates.
      2. Sort by Priority_Score DESC, then MouldLife ASC.
      3. Greedy assign to days respecting MAX_CHANGEOVERS_PER_DAY.

    CO_Day Shift A = CHANGEOVER (300 min)
    CO_Day Shift B = MOULD_CLEAN (120 min)
    CO_Day Shift C = new SKU production begins
    """

    def schedule(
        self,
        df_consumption: pd.DataFrame,
        df_running_moulds_curing,   # from curing_consumption.py ETL
        plan_start: datetime,
        planning_days: int,
        max_co_per_day: int = 8,
    ) -> pd.DataFrame:
        """
        Returns df_co_plan: [Press, Old_SKU, Target_SKU, CO_Day_Index, Status]
        CO_Day_Index is 0-based day offset from plan_start.
        """
        # NRI SKUs: need a curing press assignment
        nri_skus = df_consumption[df_consumption["Category"] == "Non-Runner-In"].copy()
        nri_skus = nri_skus.sort_values("Priority_Score", ascending=False)

        # CO-candidate presses: any curing press currently running a non-demand SKU.
        # Runner-Out SKUs are no longer in df_consumption (only demand SKUs are
        # classified), so we find candidate presses directly from the live mould data.
        demand_skus_in_plan = set(df_consumption["SKUCode"].astype(str))

        # Build a candidate list: (priority, press_or_sku, old_sku, target_sku, mould_life)
        candidates = []

        # Simple pairing: assign highest priority NRI SKU to each available CO press
        ro_presses = {}
        if df_running_moulds_curing is not None and len(df_running_moulds_curing) > 0:
            ro_press_df = df_running_moulds_curing[
                ~df_running_moulds_curing["SKUCode"].isin(demand_skus_in_plan)
            ]
            for _, pr in ro_press_df.iterrows():
                ro_presses[str(pr["Machine"])] = {
                    "old_sku": str(pr["SKUCode"]),
                    "mould_life": int(pr["MouldLife_remaining"]),
                }

        nri_queue = list(nri_skus["SKUCode"])
        for press, info in ro_presses.items():
            if not nri_queue:
                break
            target = nri_queue.pop(0)
            pri = float(
                nri_skus.loc[nri_skus["SKUCode"] == target, "Priority_Score"]
                .values[0] if len(nri_skus.loc[nri_skus["SKUCode"] == target]) > 0 else 0
            )
            candidates.append({
                "Press":          press,
                "Old_SKU":        info["old_sku"],
                "Target_SKU":     target,
                "Priority_Score": pri,
                "MouldLife_min":  info["mould_life"],
            })

        if not candidates:
            return pd.DataFrame(columns=["Press", "Old_SKU", "Target_SKU",
                                         "CO_Day_Index", "Status"])

        # Sort by priority DESC, then mould life ASC (near-expiry goes first)
        cand_df = pd.DataFrame(candidates).sort_values(
            ["Priority_Score", "MouldLife_min"],
            ascending=[False, True],
        )

        co_per_day: dict[int, int] = defaultdict(int)
        plan_rows = []
        for _, c in cand_df.iterrows():
            assigned = False
            for day in range(planning_days):
                if co_per_day[day] < max_co_per_day:
                    co_per_day[day] += 1
                    plan_rows.append({
                        "Press":        c["Press"],
                        "Old_SKU":      c["Old_SKU"],
                        "Target_SKU":   c["Target_SKU"],
                        "CO_Day_Index": day,
                        "Status":       "SCHEDULED",
                    })
                    assigned = True
                    break
            if not assigned:
                plan_rows.append({
                    "Press":        c["Press"],
                    "Old_SKU":      c["Old_SKU"],
                    "Target_SKU":   c["Target_SKU"],
                    "CO_Day_Index": -1,
                    "Status":       "DEFERRED",
                })

        df_co_plan = pd.DataFrame(plan_rows)
        sched = (df_co_plan["Status"] == "SCHEDULED").sum()
        deferred = (df_co_plan["Status"] == "DEFERRED").sum()
        print(f"  [CO Scheduler] {sched} changeovers scheduled | {deferred} deferred")
        return df_co_plan


# ══════════════════════════════════════════════════════════════════════════════
# DYNAMIC TARGET LOCK  (Phase 3 — per-shift building cap from CO plan)
# ══════════════════════════════════════════════════════════════════════════════

class DynamicTargetLock:
    """
    Updates active_press_count(SKU, shift_idx) based on the changeover plan.
    Returns a dict {(SKUCode, shift_idx): building_target_units}.
    """

    def lock(
        self,
        df_consumption: pd.DataFrame,
        df_co_plan: pd.DataFrame,
        plan_start: datetime,
        planning_days: int,
    ) -> dict:
        """Returns {(SKUCode, shift_idx): target_units}."""
        T = planning_days * Config.SHIFTS_PER_DAY

        # Initial active press count from consumption table
        press_count: dict[str, list[int]] = {}
        qty_per_press: dict[str, float]   = {}
        for _, row in df_consumption.iterrows():
            sku     = str(row["SKUCode"]).strip()
            pc_val  = row.get("Running_Press_Count", 0)
            pc      = 0 if (pc_val is None or (isinstance(pc_val, float) and math.isnan(pc_val))) else int(pc_val)
            qpp_val = row.get("Qty_Per_Press_Per_Shift", 0)
            qpp     = 0.0 if (qpp_val is None or (isinstance(qpp_val, float) and math.isnan(qpp_val))) else float(qpp_val)
            press_count[sku] = [pc] * T
            qty_per_press[sku] = qpp

        # Apply changeover events
        if df_co_plan is not None and len(df_co_plan) > 0:
            for _, co in df_co_plan.iterrows():
                if co["Status"] != "SCHEDULED":
                    continue
                day_idx     = int(co["CO_Day_Index"])
                target_sku  = str(co["Target_SKU"])
                old_sku     = str(co["Old_SKU"])
                # Building starts simultaneously with CO (Shift A of CO day).
                # Curing press: Shift A = CO, Shift B = Mould Clean, Shift C = production.
                # Building pre-builds 2 full shifts of GT by the time curing fires up.
                first_prod_shift = day_idx * Config.SHIFTS_PER_DAY

                # Old SKU loses one press from CO shift A onward
                old_first_blocked = day_idx * Config.SHIFTS_PER_DAY
                if old_sku in press_count:
                    for t in range(old_first_blocked, T):
                        press_count[old_sku][t] = max(0, press_count[old_sku][t] - 1)

                # Target SKU gains one press from production shift onward
                if target_sku not in press_count:
                    press_count[target_sku] = [0] * T
                    qty_per_press[target_sku] = 0.0
                    # Try to look up from consumption table
                    row_match = df_consumption[df_consumption["SKUCode"] == target_sku]
                    if len(row_match) > 0:
                        qty_per_press[target_sku] = float(
                            row_match.iloc[0].get("Qty_Per_Press_Per_Shift", 0)
                        )
                for t in range(first_prod_shift, T):
                    press_count[target_sku][t] += 1

        # Build target dict
        targets = {}
        for sku, counts in press_count.items():
            qpp = qty_per_press.get(sku, 0.0)
            for t, cnt in enumerate(counts):
                targets[(sku, t)] = int(cnt * qpp)

        return targets


# ══════════════════════════════════════════════════════════════════════════════
# B2C ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_from_database_b2c(
    plan_start: datetime | None = None,
    consumption_path: str | None = None,
    output_path: str | None = None,
    engine=None,
    planning_days: int | None = None,
    external_co_schedule: list | None = None,
    max_changeovers_per_day: int = 10,
    min_campaign_mins: int = 120,
    build_lead_shifts: int = 3,
) -> dict:
    """
    Run the B2C building scheduler.

    Args:
        plan_start:            First curing shift (building starts 1 shift earlier).
        consumption_path:      Path to curing_consumption_table.xlsx from Phase 0.
        output_path:           Where to write bc_building_schedule.xlsx.
        engine:                SQLAlchemy engine (created from .env if None).
        planning_days:         Override Config.PLANNING_DAYS if provided.
        external_co_schedule:  Optional list of CO events from curing_consumption_dynamic.
                               Each dict: {day: int (1-based), press: str,
                               old_sku: str, new_sku: str}.
                               When provided, skips ChangeoverScheduler entirely.

    Returns dict with same keys as building.run_from_database_hybrid().
    """
    from cbc_env import make_engine as _mk

    if plan_start is None:
        plan_start = datetime(2026, 5, 1, 7, 0, 0)
    if output_path is None:
        output_path = os.path.join(
            MAIN_OUT,
            f"bc_building_schedule_{plan_start.date()}.xlsx",
        )
    if consumption_path is None:
        consumption_path = os.path.join(cbc_env.OUTPUT_DIR, "curing_consumption_table.xlsx")
    if engine is None:
        engine = _mk()
    if planning_days is not None:
        Config.PLANNING_DAYS = planning_days
    Config.MAX_CHANGEOVERS_PER_DAY = max_changeovers_per_day
    Config.MIN_CAMPAIGN_MINS       = min_campaign_mins
    Config.BUILD_LEAD_SHIFTS       = build_lead_shifts

    print("\n" + "=" * 70)
    print("  B2C Phase 1 — Building Scheduler")
    print("=" * 70)

    # ── Load building CT from DB (authoritative source) ───────────────────────
    try:
        df_bct = pd.read_sql(
            "SELECT `SAP Machine Code` AS machine, `Cycle Time (Minutes)` AS ct_min "
            "FROM Master_Building_Machine_Design_cycleTime",
            engine,
        )
        db_ct = {
            str(r["machine"]).strip(): round(float(r["ct_min"]) * 60, 4)
            for _, r in df_bct.iterrows()
            if r["machine"] is not None
        }
        Config._CT_SEC.update(db_ct)
        print(f"  [Config] Building CT loaded from DB: {len(db_ct)} machines")
    except Exception as _e:
        print(f"  [Config] CT load from DB failed ({_e}); using hardcoded fallback")

    # ── B2C pre-start: 2 shifts before first curing shift ───────────────────
    # Building starts April 30 Shift B (15:00) — two shifts before May 1 curing.
    # This gives the LP one full extra shift to pre-build GT for RI SKUs that
    # enter May with zero opening inventory but active curing presses.
    # With PRE_START_SHIFTS=1 (Apr 30 Shift C), those SKUs starved in the very
    # first shift because there was no pre-build time. With PRE_START_SHIFTS=2,
    # Shift B (15:00-23:00) builds the buffer, eliminating Shift C starvation.
    # plan_start = May 1 07:00 (Shift A) → build_start = Apr 30 15:00 (Shift B)
    PRE_START_SHIFTS = 2
    build_start = plan_start - timedelta(hours=Config.HOURS_PER_SHIFT * PRE_START_SHIFTS)
    print(f"  [Config] Plan start:  {plan_start}  |  Build start: {build_start} ({PRE_START_SHIFTS} shifts early)")
    print(f"  [Config] Planning days: {Config.PLANNING_DAYS}")

    # ── ETL ──────────────────────────────────────────────────────────────────
    etl = B2C_ETL(engine)

    print("\n  [ETL] Loading consumption table …")
    df_consumption = etl.load_consumption_table(consumption_path)

    print("  [ETL] Loading real GT inventory (not zeroed) …")
    df_gt_inv = etl.load_gt_inventory_for_b2c()
    print(f"        {len(df_gt_inv)} SKUs with opening GT inventory")

    print("  [ETL] Loading carcass inventory (zeroing — B2C cold start) …")
    df_carcass_inv = etl.load_carcass_inventory()
    # B2C assumption: opening carcass inventory = 0.
    # Stage-2 GT machines can only consume carcasses produced by Stage-1 in this plan.
    # This makes the Stage-1 → Stage-2 dependency a hard binding constraint.
    if df_carcass_inv is not None and not df_carcass_inv.empty:
        df_carcass_inv = df_carcass_inv.copy()
        df_carcass_inv["Carcass_Inventory"] = 0
    print(f"        Carcass inventory zeroed — Stage-2 waits for Stage-1 output")

    print("  [ETL] Loading building allowable machines …")
    df_allow = etl.load_machine_allowable()

    print("  [ETL] Loading changeover times …")
    co_map = etl.load_changeover_map()

    print("  [ETL] Loading SKU sizes …")
    sku_to_size = etl.load_sku_sizes()

    print("  [ETL] Loading running building machines (for continuity locks) …")
    df_running = etl.load_running_machines()

    print("  [ETL] Loading history map (heuristic scoring) …")
    history_map = etl.load_history_map()

    # ── Union: merge historical production machines into allow_map ────────────
    # Master_Building_Allowable_Machines_source = current master data (hard allow_map).
    # Building_Stage1/2_Best_Machines = 3-month historical runs (actual production).
    # Union ensures SKUs that were historically built on a machine but are missing
    # from master data still get scheduled on that machine.
    print("  [Allow] Merging historical machine-SKU pairs into allow_map …")
    hist_by_sku: dict = {}
    for (machine, sku), count in history_map.items():
        if count > 0:
            hist_by_sku.setdefault(sku, set()).add(machine)

    allow_sku_idx = {str(r["SKUCode"]): i for i, r in df_allow.iterrows()}
    extra_pairs = 0
    new_hist_rows = []
    for sku, hist_machs in hist_by_sku.items():
        if sku in allow_sku_idx:
            idx = allow_sku_idx[sku]
            cur_set = set(df_allow.at[idx, "Machines"] or [])
            added = hist_machs - cur_set
            if added:
                df_allow.at[idx, "Machines"] = list(cur_set | added)
                extra_pairs += len(added)
        else:
            new_hist_rows.append({"SKUCode": sku, "Machines": list(hist_machs)})
            extra_pairs += len(hist_machs)

    if new_hist_rows:
        df_allow = pd.concat(
            [df_allow, pd.DataFrame(new_hist_rows)], ignore_index=True
        )
    print(f"  [Allow] +{extra_pairs} machine-SKU pairs from historical data "
          f"({len(new_hist_rows)} new SKUs unlocked via production history)")

    # ── Inch-group restriction for UNISTAGE machines (Inch-Run Study) ───────
    # Source: CLAUDE.md §Inch-Run Study — Machine Group Inch Policies
    #
    # Three UNISTAGE sub-groups each have hard or soft inch constraints:
    #   VMIMAXX  (6001-6004, 7001-7004) : 14"-18" group range
    #   BJ       (7101-7106, 7201)       : 13", 15", 16" only
    #   UNISTAGE (7501-7503)             : 12", 13" HARD — never 14"+
    #
    # Machines 7004/7003 reached 173/120 COs in a month because they were
    # assigned 25/23 different SKUs spanning multiple inch sizes.  Restricting
    # each group to its allowed inches eliminates cross-inch (diff-size) COs
    # and bounds daily SKU variety per machine.
    #
    # Stage-1 and Stage-2 are NOT restricted here — their inch mix is managed
    # by the LP and the Stage-2 CO-time multiplier (Fix 3).
    _UNISTAGE_INCH_POLICY: dict[str, set] = {
        # VMIMAXX — 14" to 18" allowed
        "6001": {"14","15","16","17","18"},
        "6002": {"14","15","16","17","18"},
        "6003": {"14","15","16","17","18"},
        "6004": {"14","15","16","17","18"},
        "7001": {"14","15","16","17","18"},
        "7002": {"14","15","16","17","18"},
        "7003": {"14","15","16","17","18"},
        "7004": {"14","15","16","17","18"},
        # BJ — 13", 14", 15", 16" allowed (14" added: plant master data confirms
        # 14" SKUs run on BJ machines; "well-locked" means 83-99% dominant inch,
        # not a hard prohibition on 14")
        "7101": {"13","14","15","16"},
        "7102": {"13","14","15","16"},
        "7103": {"13","14","15","16"},
        "7104": {"13","14","15","16"},
        "7105": {"13","14","15","16"},
        "7106": {"13","14","15","16"},
        "7201": {"13","14","15","16"},
        # UNISTAGE — 12" and 13" HARD (never 14"+)
        "7501": {"12","13"},
        "7502": {"12","13"},
        "7503": {"12","13"},
    }

    removed_pairs = 0
    for idx, row in df_allow.iterrows():
        sku       = str(row["SKUCode"])
        mach_list = list(row.get("Machines", []) or [])
        sku_inch  = str(sku_to_size.get(sku, "")).strip().replace('"', "")
        if not sku_inch:
            continue  # unknown inch — keep as-is
        filtered = []
        for m in mach_list:
            allowed = _UNISTAGE_INCH_POLICY.get(str(m))
            if allowed is None:
                filtered.append(m)   # Stage-1/Stage-2 — no inch restriction
            elif sku_inch in allowed:
                filtered.append(m)   # inch within group policy → keep
            else:
                removed_pairs += 1   # inch violates group policy → drop
        df_allow.at[idx, "Machines"] = filtered

    print(f"  [Inch] Removed {removed_pairs} machine-SKU pairs violating "
          f"group inch policies (VMIMAXX:14-18, BJ:13/14/15/16, UNISTAGE:12-13)")

    # ── Per-machine hard inch allocation (demand-driven, replaces history bias) ──
    # Demand analysis (May): 14"=77k VMI-eligible, 15"=70k VMI + 142k BJ,
    # 16"=57k VMI, 17"=26k VMI, 13"=86k BJ, 12"=22k UNISTAGE.
    # Each machine is locked to its dominant inch (± one overflow inch) so that:
    #   - Same-inch demand is shared fairly across sibling machines (no history winner)
    #   - No machine bleeds into a neighbour's inch pool and dilutes their demand
    #   - 7102/7104 retain 14" so the 2 BJ-exclusive 14" RI SKUs keep machines
    _MACHINE_HARD_INCH: dict[str, set] = {
        # VMIMAXX — strict dominant inch; demand analysis shows each inch pool
        # is under-capacity when shared, so sibling machines must not dilute each other
        "7001": {"16"},               # dominant=16"; strict {16} so TIER-3 elig_count
                                      # matches 6004/7201 — enables fair 3-way split
        "6001": {"14"},               # dominant=14"
        "7002": {"14"},               # dominant=14"
        "7004": {"14"},               # dominant=14"
        "6002": {"15"},               # dominant=15"
        "7003": {"15"},               # dominant=15"
        "6003": {"17", "18"},         # dominant=17"; 18" demand (840 units) added
        "6004": {"16"},               # dominant=16"
        # BJ — strict dominant inch; 15" demand (142k BJ-eligible) > BJ capacity → all
        # 15"-dominant BJ machines run full; 13" demand (86k) fills 3 BJ machines
        "7101": {"15"},               # dominant=15"
        "7102": {"14", "15"},         # dominant=15"; 14" kept for BJ-exclusive 14" RI SKUs
        "7103": {"13"},               # dominant=13"
        "7104": {"14", "15"},         # dominant=15"; 14" kept for BJ-exclusive 14" RI SKUs
        "7105": {"13"},               # dominant=13"
        "7106": {"13"},               # dominant=13"
        "7201": {"16"},               # dominant=16"
        # UNISTAGE — 12" and 13" split cleanly; 7501 takes all 12" demand
        "7501": {"12"},               # dominant=12" HARD (EXISTING)
        "7502": {"13"},               # dominant=13"
        "7503": {"13"},               # dominant=13"
    }
    removed_hard = 0
    for idx, row in df_allow.iterrows():
        sku       = str(row["SKUCode"])
        mach_list = list(row.get("Machines", []) or [])
        sku_inch  = str(sku_to_size.get(sku, "")).strip().replace('"', "")
        if not sku_inch:
            continue
        filtered = [
            m for m in mach_list
            if str(m) not in _MACHINE_HARD_INCH
            or sku_inch in _MACHINE_HARD_INCH[str(m)]
        ]
        removed_hard += len(mach_list) - len(filtered)
        df_allow.at[idx, "Machines"] = filtered
    print(f"  [Inch] Removed {removed_hard} machine-SKU pairs via per-machine "
          f"hard inch allocation (VMI+BJ+UNISTAGE dominant-inch locking)")

    # ── Diagnostic: RI SKUs with zero eligible building machines ─────────────
    # These will produce 0 GT even if they have curing presses — starvation source.
    # Root cause: either master data gap OR inch filter removed all their machines.
    _allow_map_diag = {str(r["SKUCode"]): list(r.get("Machines", []) or [])
                       for _, r in df_allow.iterrows()}
    _ri_zero_bld = []
    for _, _r in df_consumption.iterrows():
        if str(_r.get("Category", "")).strip() != "Runner-In":
            continue
        _s = str(_r["SKUCode"])
        _mlist = _allow_map_diag.get(_s, [])
        _press = float(_r.get("Running_Press_Count", 0) or 0)
        if _press > 0 and len(_mlist) == 0:
            _inch = sku_to_size.get(_s, "?")
            _ri_zero_bld.append((_s, _inch, _press))
    if _ri_zero_bld:
        print(f"  [WARN] {len(_ri_zero_bld)} Runner-In SKUs have 0 eligible building "
              f"machines after inch filter — these will cause starvation:")
        for _s, _inch, _press in _ri_zero_bld:
            _orig = [str(m) for _, _row in df_allow.iterrows()
                     if str(_row["SKUCode"]) == _s
                     for m in (_row.get("Machines") or [])]
            print(f"    SKU {_s}  inch={_inch}  presses={int(_press)}  "
                  f"machines_after_filter={len(_allow_map_diag.get(_s, []))}")
    else:
        print("  [OK] All Runner-In SKUs with active presses have ≥1 eligible building machine")

    print("  [ETL] Loading running curing moulds (for changeover planning) …")
    try:
        from curing_consumption import ConsumptionETL
        cetl = ConsumptionETL(engine)
        df_running_curing = cetl.load_running_moulds()
    except Exception as exc:
        print(f"  ⚠️  Could not load curing running moulds: {exc}")
        df_running_curing = None

    # ── Changeover planning ───────────────────────────────────────────────────
    if external_co_schedule is not None:
        # Use the pre-computed CO schedule from curing_consumption_dynamic.py.
        # Convert from {day (1-based), press, old_sku, new_sku} to
        # the DynamicTargetLock format: [Press, Old_SKU, Target_SKU,
        # CO_Day_Index (0-based), Status].
        print(f"\n  [CO Plan] Using external CO schedule ({len(external_co_schedule)} events) …")
        if external_co_schedule:
            df_co_plan = pd.DataFrame([
                {
                    "Press":        ev["press"],
                    "Old_SKU":      ev["old_sku"],
                    "Target_SKU":   ev["new_sku"],
                    "CO_Day_Index": int(ev["day"]) - 1,   # convert 1-based → 0-based
                    "Status":       "SCHEDULED",
                }
                for ev in external_co_schedule
            ])
        else:
            df_co_plan = pd.DataFrame(columns=["Press", "Old_SKU", "Target_SKU", "CO_Day_Index", "Status"])
        print(f"  [CO Plan] {len(df_co_plan)} CO rows loaded from external schedule")
    else:
        print("\n  [CO Plan] Scheduling changeovers for Non-Runner-In SKUs …")
        co_scheduler = ChangeoverScheduler()
        df_co_plan = co_scheduler.schedule(
            df_consumption,
            df_running_curing,
            plan_start,
            Config.PLANNING_DAYS,
            max_co_per_day=getattr(Config, "MAX_CHANGEOVERS_PER_DAY", 8),
        )

    # ── Dynamic target lock ───────────────────────────────────────────────────
    print("  [Lock] Computing dynamic per-shift building targets …")
    target_lock = DynamicTargetLock()
    dynamic_targets = target_lock.lock(
        df_consumption, df_co_plan, plan_start, Config.PLANNING_DAYS
    )

    # ── Synthetic curing schedule ─────────────────────────────────────────────
    # Build a df_curing-compatible DataFrame that HybridDailyScheduler can use.
    # Incorporate dynamic targets: update consumption column from dynamic_targets.
    print("  [Synthetic] Building synthetic curing schedule from consumption table …")
    df_consumption_updated = df_consumption.copy()

    # Only build NRI SKUs that have a scheduled CO (press activation planned).
    # Inactive NRI presses waste building capacity on uncurable green tyres.
    nri_with_co = (
        set(df_co_plan["Target_SKU"].astype(str).tolist())
        if df_co_plan is not None and not df_co_plan.empty and "Target_SKU" in df_co_plan.columns
        else set()
    )
    # Build CO day map (NRI SKU → 0-based CO day) for front-loading NRI demand
    co_day_map: dict = {}
    if df_co_plan is not None and not df_co_plan.empty:
        for _, _co in df_co_plan.iterrows():
            tgt = str(_co.get("Target_SKU", "")).strip()
            if tgt and str(_co.get("Status", "")) == "SCHEDULED":
                co_day_map[tgt] = int(_co.get("CO_Day_Index", 0))

    df_curing_synthetic = _make_synthetic_curing(
        df_consumption_updated, plan_start, Config.PLANNING_DAYS,
        nri_skus_to_build=nri_with_co,
        co_day_map=co_day_map,
    )
    n_skus = df_curing_synthetic["SKUCode"].nunique()
    print(f"  [Synthetic] {len(df_curing_synthetic)} rows | {n_skus} SKUs")

    # Shift synthetic curing times back by PRE_START_SHIFTS so the scheduler's
    # day-0 window aligns with build_start (not plan_start).
    df_curing_synthetic["StartTime"] = df_curing_synthetic["StartTime"] - timedelta(
        hours=Config.HOURS_PER_SHIFT * PRE_START_SHIFTS
    )
    df_curing_synthetic["EndTime"] = df_curing_synthetic["EndTime"] - timedelta(
        hours=Config.HOURS_PER_SHIFT * PRE_START_SHIFTS
    )

    # ── UNISTAGE priority steering & idle-fill demand ────────────────────────
    # priority_map: steer high-priority SKUs onto UNISTAGE machines in assigner.
    if "Priority_Score" in df_consumption_updated.columns:
        priority_map = dict(zip(
            df_consumption_updated["SKUCode"].astype(str),
            df_consumption_updated["Priority_Score"].fillna(0),
        ))
    else:
        priority_map = {}

    # Mould-constrained RI SKU priority boost.
    # A Runner-In SKU with few presses relative to its demand needs more building
    # days (current_days > horizon) — it must start building FIRST to avoid being
    # discovered late. Multiply its priority_map value by a constraint factor so
    # the LP heuristic assigns building machines to it preferentially.
    # Multiplier = 1 + current_days / PLANNING_DAYS, clamped to [1, 4].
    _mould_boosted = 0
    for _, _row in df_consumption_updated.iterrows():
        _sku = str(_row["SKUCode"])
        if str(_row.get("Category", "")).strip() != "Runner-In":
            continue
        _n_press = float(_row.get("Running_Press_Count", 0) or 0)
        _qpp     = float(_row.get("Qty_Per_Press_Per_Shift", 0) or 0)
        _dem     = float(_row.get("Demand_Qty", 0) or 0)
        if _n_press > 0 and _qpp > 0 and _dem > 0:
            _rate_day   = _qpp * Config.SHIFTS_PER_DAY
            _curr_days  = _dem / (_n_press * _rate_day)
            _multiplier = min(1.0 + _curr_days / Config.PLANNING_DAYS, 4.0)
            if _multiplier > 1.01:
                priority_map[_sku] = priority_map.get(_sku, 1.0) * _multiplier
                _mould_boosted += 1
    print(f"  [Priority] Mould-constrained boost applied to {_mould_boosted} RI SKUs")

    # NRI SKUs with a confirmed CO in co_day_map can have very low
    # ConsolidatedPriorityScore (e.g. 0.008), causing the heuristic to process
    # them last — by which time all building machines are at MAX_SKUS_PER_MACHINE
    # cap and these NRI SKUs get no machine assignment despite having a CO.
    # Guarantee a minimum priority so they compete for machines before the cap bites.
    _MIN_NRI_CO_PRIORITY = 0.05   # above typical median; tune if too aggressive
    _nri_co_boosted = 0
    for _sku_nri, _co_day in co_day_map.items():
        _cat_rows = df_consumption_updated[
            df_consumption_updated["SKUCode"].astype(str) == _sku_nri
        ]
        if _cat_rows.empty:
            continue
        if str(_cat_rows.iloc[0].get("Category", "")).strip() != "Non-Runner-In":
            continue
        if priority_map.get(_sku_nri, 0.0) < _MIN_NRI_CO_PRIORITY:
            priority_map[_sku_nri] = _MIN_NRI_CO_PRIORITY
            _nri_co_boosted += 1
    print(f"  [Priority] NRI-with-CO minimum priority ({_MIN_NRI_CO_PRIORITY}) "
          f"applied to {_nri_co_boosted} low-priority NRI SKUs")

    # extra_topup_demand: ALL demand SKUs eligible for UNISTAGE that are NOT
    # already in the active LP synthetic curing plan (i.e., not in _gt_remaining).
    # LP covers Runner-In + NRI-with-CO. Everything else with UNISTAGE eligibility
    # goes here so TopUp can fill idle UNISTAGE tails with them.
    # This maximises UNISTAGE utilisation without changing the LP problem size.
    unistage_set = set(map(str, Config.UNISTAGE))
    allow_map_lookup = {}
    for _, r in df_allow.iterrows():
        allow_map_lookup[str(r["SKUCode"])] = set(map(str, r.get("Machines", [])))

    qty_col = "Demand_Qty" if "Demand_Qty" in df_consumption_updated.columns else None

    # SKUs already in active LP plan (Runner-In active + NRI-with-CO + Runner-Out)
    lp_active_skus = set(
        df_curing_synthetic["SKUCode"].astype(str).unique()
    ) if not df_curing_synthetic.empty else set()

    extra_topup_demand = {}
    for _, row in df_consumption_updated.iterrows():
        sku = str(row["SKUCode"])
        if sku in lp_active_skus:
            continue  # already tracked in _gt_remaining via synthetic curing
        if not (allow_map_lookup.get(sku, set()) & unistage_set):
            continue  # not eligible for any UNISTAGE machine
        demand = float(row[qty_col]) if qty_col else 0.0
        if demand > 0:
            extra_topup_demand[sku] = demand

    n_extra = len(extra_topup_demand)
    extra_total = sum(extra_topup_demand.values())
    print(f"  [UNISTAGE] {n_extra} SKUs ({extra_total:,.0f} units) added to "
          f"UNISTAGE idle-fill pool (not in active LP plan)")

    # Customer demand cap: prevents TopUp from building beyond what the
    # customer ordered, freeing machine time for under-served SKUs instead.
    demand_cap: dict = {}
    if qty_col:
        for _, row in df_consumption_updated.iterrows():
            sku = str(row["SKUCode"])
            qty = float(row.get(qty_col, 0) or 0)
            if qty > 0:
                demand_cap[sku] = qty
    print(f"  [Cap] Demand cap set for {len(demand_cap)} SKUs — "
          f"building will not exceed customer demand per SKU")

    # ── Run the existing HybridDailyScheduler ────────────────────────────────
    # Pass real GT/carcass inventory (NOT zeroed — B2C uses opening inventory).
    # Pass real running machine state (continuity locks active on Day 0).
    print("\n  [Scheduler] Launching HybridDailyScheduler in B2C mode …")
    scheduler = HybridDailyScheduler()
    results = scheduler.run(
        df_curing_synthetic,
        df_gt_inv,          # REAL GT inventory (CBC zeroed this out)
        df_carcass_inv,     # REAL carcass inventory
        df_allow,
        co_map,
        sku_to_size,
        df_running,         # REAL running machine state
        build_start,        # PRE_START_SHIFTS before plan_start
        history_map=history_map,
        priority_map=priority_map,
        extra_topup_demand=extra_topup_demand if extra_topup_demand else None,
        demand_cap=demand_cap if demand_cap else None,
    )

    # ── Patch GT_Demand → always use Demand_Qty from demand file ────────────
    # The scheduler derives GT_Demand from curing_matrix.sum() = min(press_rate,
    # demand_rate) × shifts. For slow presses this is < customer demand, causing
    # the KPI to undercount. Override with Demand_Qty so the KPI always shows the
    # true customer demand (694,973) regardless of press burn rate.
    _dem_qty_map: dict = {}
    if df_consumption_updated is not None and not df_consumption_updated.empty:
        for _, _cr in df_consumption_updated.iterrows():
            _s = str(_cr["SKUCode"])
            _q = float(_cr.get("Demand_Qty", 0) or 0)
            if _q > 0:
                _dem_qty_map[_s] = int(_q)

    if "demand_summary" in results and not results["demand_summary"].empty and _dem_qty_map:
        _ds = results["demand_summary"].copy()
        _old_gt = _ds.set_index("SKUCode")["GT_Demand"].to_dict()
        _ds["GT_Demand"] = _ds["SKUCode"].astype(str).map(
            lambda s: _dem_qty_map.get(s) or int(_old_gt.get(s, 0))
        )
        _ds["Net_GT_Demand"] = (
            _ds["GT_Demand"] - _ds["GT_Inventory"]
        ).clip(lower=0).astype(int)
        _ds["Gap"] = (_ds["GT_Demand"] - _ds["Planned_GT"]).clip(lower=0)
        _ds["Fulfillment_Pct"] = (
            _ds["Planned_GT"] / _ds["GT_Demand"].replace(0, 1) * 100
        ).round(1)
        _ds["Status"] = _ds.apply(
            lambda r: ("FULLY MET" if r["Planned_GT"] >= r["Net_GT_Demand"]
                       else "PARTIAL" if r["Planned_GT"] > 0 else "UNMET"),
            axis=1,
        )
        results["demand_summary"] = _ds
        print(f"  [Summary] GT_Demand patched from demand file for "
              f"{len(_dem_qty_map)} SKUs")

    # ── Augment demand_summary with ALL demand SKUs ───────────────────────────
    # scheduler.run() only includes LP-active SKUs (Runner-In + NRI-with-CO).
    # Add remaining demand SKUs so the Demand Summary sheet shows all 89.
    _GT_MACH      = {str(m) for m in Config.STAGE2 | Config.UNISTAGE}
    _sentinels    = {"CHANGEOVER", "MOULD_CLEAN", "C/O", "CLEANING"}
    _ss           = results.get("shift_schedule", pd.DataFrame())
    _prod_by_sku  : dict = {}
    if not _ss.empty:
        _pr = _ss[
            ~_ss["SKUCode"].astype(str).str.upper().isin(_sentinels)
            & _ss["Machine"].astype(str).isin(_GT_MACH)
        ]
        _prod_by_sku = _pr.groupby("SKUCode")["Qty"].sum().to_dict()

    _existing_skus = set(
        results["demand_summary"]["SKUCode"].astype(str)
    ) if "demand_summary" in results and not results["demand_summary"].empty else set()

    _gt_inv_dict  = dict(zip(
        df_gt_inv["SKUCode"].astype(str).str.strip(),
        df_gt_inv["GT_Inventory"].astype(float),
    )) if df_gt_inv is not None and not df_gt_inv.empty else {}

    _extra_rows = []
    for _, cr in df_consumption_updated.iterrows():
        sku    = str(cr["SKUCode"])
        if sku in _existing_skus:
            continue
        demand = float(cr.get("Demand_Qty", 0) or 0)
        if demand <= 0:
            continue
        gt_inv  = float(_gt_inv_dict.get(sku, 0))
        net_dem = max(0, int(demand) - int(gt_inv))
        planned = float(_prod_by_sku.get(sku, 0))
        gap     = max(0, int(demand) - int(planned))
        pct     = round(planned / demand * 100, 1) if demand > 0 else 0.0
        _extra_rows.append({
            "SKUCode":             sku,
            "GT_Demand":           int(demand),
            "GT_Inventory":        int(gt_inv),
            "Net_GT_Demand":       int(net_dem),
            "Planned_GT":          int(planned),
            "Gap":                 gap,
            "Fulfillment_Pct":     pct,
            "Status":              ("FULLY MET" if planned >= net_dem and net_dem >= 0
                                    else "PARTIAL" if planned > 0 else "UNMET"),
            "Burn_Rate_Per_Shift": 0.0,
            "Active_Shifts":       0,
            "First_Curing_Start":  "",
        })
    if _extra_rows:
        results["demand_summary"] = pd.concat(
            [results["demand_summary"], pd.DataFrame(_extra_rows)],
            ignore_index=True,
        )
        print(f"  [Summary] Added {len(_extra_rows)} demand SKUs to Demand Summary "
              f"(NRI-no-CO + excluded)")

    # ── Attach B2C-specific data to results ───────────────────────────────────
    results["consumption_table"]   = df_consumption
    results["co_plan"]             = df_co_plan
    results["dynamic_targets"]     = pd.DataFrame(
        [(k[0], k[1], v) for k, v in dynamic_targets.items()],
        columns=["SKUCode", "Shift_Idx", "Building_Target"],
    )

    # ── Export ────────────────────────────────────────────────────────────────
    print(f"\n  [Export] Writing → {output_path}")
    exporter = ExcelExporter(output_path)
    exporter.export(results)

    # ── Append B2C-specific sheets ────────────────────────────────────────────
    # Build SKU-level allow_map and avg-CT map for the Demand Fulfillment sheet
    _allow_map_final = {}
    for _, r in df_allow.iterrows():
        _allow_map_final[str(r["SKUCode"])] = set(map(str, r.get("Machines", []) or []))
    gt_inv_map = dict(zip(
        df_gt_inv["SKUCode"].astype(str).str.strip(),
        df_gt_inv["GT_Inventory"].astype(float),
    )) if df_gt_inv is not None and not df_gt_inv.empty else {}

    _append_b2c_sheets(
        output_path, df_co_plan, results["dynamic_targets"],
        df_consumption, consumption_path,
        gt_inv_map=gt_inv_map,
        allow_map_lookup=_allow_map_final,
        planning_days=planning_days,
    )

    print("=" * 70)
    print("  Phase 1 complete.")
    print("=" * 70 + "\n")

    return results


def _skip_reason(
    sku: str,
    status: str,
    category: str,
    co_targets: dict,
    planned_gt: float,
    eligible_machines: int = 0,
) -> str:
    """Return a human-readable Skip_Reason for a given demand SKU."""
    if status == "FULLY MET":
        return "-"
    if category == "Runner-Out":
        return "Runner-Out: SKU not in customer demand — excluded from building plan"
    if category == "Runner-In":
        if planned_gt == 0:
            return "Runner-In: zero build — check allowable machines master data"
        return "Runner-In: partial — demand cap applied (build ≤ customer demand)"
    # Non-Runner-In
    co = co_targets.get(sku)
    if co:
        if co["status"] == "SCHEDULED":
            d = co["day"] + 1
            return (f"NRI: curing CO Day {d} — building ahead; "
                    f"curing starts Shift C Day {d}")
        if co["status"] == "DEFERRED":
            return ("NRI: curing CO deferred — 8 CO/day cap reached; "
                    "expand horizon or rebalance CO budget")
    if planned_gt == 0:
        if eligible_machines > 0:
            return ("NRI: machines available (historical) but no curing CO scheduled — "
                    "add curing mould/CO to activate this SKU")
        return ("NRI: no building machine in master data or historical production — "
                "add to Master_Building_Allowable_Machines_source")
    return "NRI: partial — building capacity shared with higher-priority SKUs"


def _append_b2c_sheets(
    output_path: str,
    df_co_plan: pd.DataFrame,
    df_dynamic_targets: pd.DataFrame,
    df_consumption: pd.DataFrame,
    consumption_path: str | None = None,
    gt_inv_map: dict | None = None,
    allow_map_lookup: dict | None = None,
    planning_days: int = 31,
):
    """Append B2C-specific sheets to the building output workbook."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)

    _NAVY   = "1F3864"
    _WHITE  = "FFFFFF"
    _GREEN  = "E2EFDA"
    _AMBER  = "FFF2CC"
    _RED    = "FFE0E0"
    _GREY   = "D3D3D3"

    def _add_sheet(wb, name, df):
        # Remove sheet if it already exists (avoid duplicate errors on re-run)
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        hdr_fill = PatternFill("solid", fgColor=_NAVY)
        hdr_font = Font(bold=True, color=_WHITE)
        bd = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        for ci, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = bd
            cell.alignment = Alignment(horizontal="center")
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = bd
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 35)
        return ws

    # ── Build CO lookup for Skip_Reason ──────────────────────────────────────
    co_targets: dict = {}
    if df_co_plan is not None and len(df_co_plan) > 0:
        for _, co_row in df_co_plan.iterrows():
            tgt = str(co_row.get("Target_SKU", "")).strip()
            if tgt:
                co_targets[tgt] = {
                    "status": str(co_row.get("Status", "")),
                    "day":    int(co_row.get("CO_Day_Index", -1)),
                }

    # ── Category + demand lookup from consumption table ───────────────────────
    cat_map:  dict = {}
    dem_map:  dict = {}
    qty_map:  dict = {}
    if df_consumption is not None and not df_consumption.empty:
        for _, cr in df_consumption.iterrows():
            s = str(cr["SKUCode"]).strip()
            cat_map[s] = str(cr.get("Category", ""))
            dem_raw = cr.get("Demand_Qty", 0)
            dem_map[s] = 0.0 if (dem_raw is None or (isinstance(dem_raw, float) and math.isnan(dem_raw))) else float(dem_raw)

    # ── Load excluded SKUs from consumption table (Phase 0 output) ────────────
    df_excluded_skus = pd.DataFrame()
    if consumption_path and os.path.exists(consumption_path):
        try:
            xl_cons = pd.ExcelFile(consumption_path)
            if "Excluded SKUs" in xl_cons.sheet_names:
                df_excluded_skus = pd.read_excel(consumption_path, sheet_name="Excluded SKUs")
                df_excluded_skus["SKUCode"] = df_excluded_skus["SKUCode"].astype(str).str.strip()
                print(f"  [B2C Sheets] Loaded {len(df_excluded_skus)} excluded SKUs from consumption table")
        except Exception as _e:
            print(f"  ⚠  Could not load Excluded SKUs: {_e}")

    # ── Read Shift Schedule for daily totals + CO count ───────────────────────
    _sentinel_skus = {"CHANGEOVER", "MOULD_CLEAN", "C/O", "CLEANING", "MOULDCLEAN", "CO"}
    df_ss       = pd.DataFrame()
    df_clean    = pd.DataFrame()
    total_co_count = 0

    try:
        df_ss = pd.read_excel(output_path, sheet_name="Shift Schedule", header=2)
    except Exception as _e:
        print(f"  ⚠  Could not read Shift Schedule: {_e}")

    if not df_ss.empty:
        prod_mask = ~df_ss["SKUCode"].astype(str).str.strip().str.upper().isin(_sentinel_skus)
        df_prod    = df_ss[prod_mask].copy()
        df_nonprod = df_ss[~prod_mask].copy()

        total_co_count = int(
            df_ss["SKUCode"].astype(str).str.strip().str.upper()
            .isin({"CHANGEOVER", "C/O"}).sum()
        )

        grp_cols = [c for c in ["Machine", "Date", "Shift", "SKUCode"] if c in df_prod.columns]
        if grp_cols and "Qty" in df_prod.columns:
            other_cols = [c for c in df_prod.columns if c not in grp_cols + ["Qty"]]
            agg_spec   = {"Qty": "sum", **{c: "first" for c in other_cols}}
            df_merged  = df_prod.groupby(grp_cols, as_index=False).agg(agg_spec)
            df_clean   = pd.concat([df_merged, df_nonprod], ignore_index=True)
            sort_keys  = [c for c in ["Date", "Machine", "Shift"] if c in df_clean.columns]
            if sort_keys:
                df_clean = df_clean.sort_values(sort_keys).reset_index(drop=True)
        else:
            df_clean = df_ss

    # ── Sheet: Changeover Plan ────────────────────────────────────────────────
    try:
        if df_co_plan is not None and len(df_co_plan) > 0:
            _add_sheet(wb, "Changeover Plan", df_co_plan)
            print(f"  [B2C Sheets] Changeover Plan: {len(df_co_plan)} rows")
    except Exception as _e:
        print(f"  ⚠  Changeover Plan sheet failed: {_e}")

    # ── Sheet: Dynamic Targets ────────────────────────────────────────────────
    try:
        if df_dynamic_targets is not None and len(df_dynamic_targets) > 0:
            _add_sheet(wb, "Dynamic Targets", df_dynamic_targets.head(1000))
            print(f"  [B2C Sheets] Dynamic Targets: {len(df_dynamic_targets)} rows")
    except Exception as _e:
        print(f"  ⚠  Dynamic Targets sheet failed: {_e}")

    # ── Sheet: SKU Classification Summary ────────────────────────────────────
    try:
        agg_dict: dict = {"SKU_Count": ("SKUCode", "count")}
        if "Total_GT_Per_Shift_Day0" in df_consumption.columns:
            agg_dict["Total_GT_Per_Shift"] = ("Total_GT_Per_Shift_Day0", "sum")
        if "Priority_Score" in df_consumption.columns:
            agg_dict["Avg_Priority"] = ("Priority_Score", "mean")
        if "Demand_Qty" in df_consumption.columns:
            agg_dict["Total_Customer_Demand"] = ("Demand_Qty", "sum")

        cat_summary = df_consumption.groupby("Category").agg(**agg_dict).reset_index()
        ws_cat = _add_sheet(wb, "SKU Classification", cat_summary)
        footer_row = len(cat_summary) + 3
        ws_cat.cell(row=footer_row,     column=1, value="KPI").font = Font(bold=True)
        ws_cat.cell(row=footer_row + 1, column=1, value="Total Building COs (no daily limit)")
        ws_cat.cell(row=footer_row + 1, column=2, value=total_co_count)
        ws_cat.cell(row=footer_row + 2, column=1, value="Curing Press CO limit / day")
        ws_cat.cell(row=footer_row + 2, column=2, value=8)
        ws_cat.cell(row=footer_row + 3, column=1, value="Curing COs scheduled (this plan)")
        ws_cat.cell(row=footer_row + 3, column=2, value=len(df_co_plan) if df_co_plan is not None else 0)
        print(f"  [B2C Sheets] SKU Classification: {len(cat_summary)} categories")
    except Exception as _e:
        print(f"  ⚠  SKU Classification sheet failed: {_e}")

    # ── Sheet: Shift Schedule (Clean) ─────────────────────────────────────────
    try:
        if not df_clean.empty:
            _add_sheet(wb, "Shift Schedule (Clean)", df_clean)
            print(f"  [B2C Sheets] Shift Schedule (Clean): {len(df_clean)} rows")
    except Exception as _e:
        print(f"  ⚠  Shift Schedule (Clean) sheet failed: {_e}")

    # ── Sheet: Daily GT & Carcass ─────────────────────────────────────────────
    _S1_MACHINES = {
        "6801","6802","6803","6909","6911",
        "7601","7701","7801","7802","7803","7804",
        "8001","8002","8003","8101",
    }
    try:
        if not df_clean.empty and "Qty" in df_clean.columns:
            df_pc = df_clean[
                ~df_clean["SKUCode"].astype(str).str.strip().str.upper().isin(_sentinel_skus)
            ].copy()
            df_pc["_mach"] = df_pc["Machine"].astype(str).str.strip()
            if "Stage" in df_pc.columns:
                df_pc["_is_s1"] = df_pc["Stage"].astype(str).str.upper().str.contains("1")
            elif "MachineType" in df_pc.columns:
                df_pc["_is_s1"] = df_pc["MachineType"].astype(str).str.upper() == "STAGE1"
            else:
                df_pc["_is_s1"] = df_pc["_mach"].isin(_S1_MACHINES)
            df_pc["GT_Qty"]      = df_pc["Qty"].where(~df_pc["_is_s1"], 0)
            df_pc["Carcass_Qty"] = df_pc["Qty"].where( df_pc["_is_s1"], 0)
            if "Date" in df_pc.columns:
                daily = (
                    df_pc.groupby("Date", as_index=False)
                    .agg(
                        GT_Produced=("GT_Qty", "sum"),
                        Carcass_Produced=("Carcass_Qty", "sum"),
                        Total_Units=("Qty", "sum"),
                        Active_SKUs=("SKUCode", "nunique"),
                    )
                    .sort_values("Date")
                )
                daily["Cumulative_GT"] = daily["GT_Produced"].cumsum()
                _add_sheet(wb, "Daily GT & Carcass", daily)
                print(f"  [B2C Sheets] Daily GT & Carcass: {len(daily)} days | "
                      f"Total GT: {daily['GT_Produced'].sum():,.0f} | "
                      f"Total Carcass: {daily['Carcass_Produced'].sum():,.0f}")
    except Exception as _e:
        print(f"  ⚠  Daily GT & Carcass sheet failed: {_e}")

    # ── Sheet: Demand Fulfillment (B2C) — ALL demand SKUs + Skip_Reason ───────
    # Built from scratch so it is never broken by the ExcelExporter header format.
    # Shows all 89 demand SKUs; unscheduled ones carry a Skip_Reason.
    try:
        # 1. Actual GT production per SKU (STAGE2 + UNISTAGE only — not carcass).
        #    Stage1 carcass is an intermediate product, not a finished tyre, so
        #    it must not be counted toward customer demand fulfillment.
        _GT_MACHINES = (
            {str(m) for m in Config.STAGE2} | {str(m) for m in Config.UNISTAGE}
        )
        prod_by_sku: dict = {}
        if not df_clean.empty and "Qty" in df_clean.columns:
            _prod_rows = df_clean[
                ~df_clean["SKUCode"].astype(str).str.strip().str.upper()
                .isin(_sentinel_skus)
                & df_clean["Machine"].astype(str).str.strip().isin(_GT_MACHINES)
            ]
            prod_by_sku = _prod_rows.groupby("SKUCode")["Qty"].sum().to_dict()

        # Helper: avg cycle time (min) for a SKU across its eligible machines
        _gt_inv_map    = gt_inv_map or {}
        _allow_lkp     = allow_map_lookup or {}
        _plan_mins     = planning_days * 3 * 8 * 60  # total shift-minutes per press
        # Presses_Needed normalisation: 86,400 min = 60 days × 24h × 60min
        # This gives an intuitive count of how many continuously-running presses
        # are required to satisfy the monthly demand for a given SKU.
        _PRESS_NORM    = 86_400.0

        def _sku_ct(sku: str) -> float | None:
            machs = _allow_lkp.get(sku, set())
            if not machs:
                return None
            cts = [Config.ct_min(m) for m in machs]
            return round(sum(cts) / len(cts), 1) if cts else None

        # 2. All demand SKUs from df_consumption (Demand_Qty > 0 = in demand file)
        dem_rows: list = []
        active_skus: set = set()
        if df_consumption is not None and not df_consumption.empty:
            for _, cr in df_consumption.iterrows():
                sku    = str(cr["SKUCode"]).strip()
                demand = float(cr.get("Demand_Qty", 0) or 0)
                if demand <= 0:
                    continue
                active_skus.add(sku)
                produced   = float(prod_by_sku.get(sku, 0))
                gt_inv     = float(_gt_inv_map.get(sku, 0))
                fill_pct   = round(100 * produced / demand, 1) if demand > 0 else 0.0
                gap        = max(0, int(demand) - int(produced))
                status     = (
                    "FULLY MET" if produced >= demand * 0.95
                    else "PARTIAL"  if produced > 0
                    else "UNMET"
                )
                avg_ct     = _sku_ct(sku)
                elig_count = len(_allow_lkp.get(sku, set()))
                presses    = (
                    round(demand * avg_ct / _PRESS_NORM, 2)
                    if avg_ct else None
                )
                dem_rows.append({
                    "SKUCode":           sku,
                    "Category":          str(cr.get("Category", "")),
                    "Priority":          round(float(cr.get("Priority_Score", 0) or 0), 7),
                    "Demand":            int(demand),
                    "GT_Inventory":      int(gt_inv),
                    "Planned_Units":     int(produced),
                    "Gap":               gap,
                    "Fulfillment_Pct":   f"{fill_pct}%",
                    "Status":            status,
                    "CycleTime_min":     avg_ct if avg_ct is not None else "NA",
                    "Eligible_Machines": elig_count if elig_count else "NA",
                    "Presses_Needed":    presses if presses is not None else "NA",
                    "Skip_Reason":       "",
                })

        # 3. Excluded SKUs (in demand file but not in consumption table).
        #    The "Excluded SKUs" sheet has a title in row 0, blank in row 1,
        #    actual column headers in row 2 — read with header=2 to fix that.
        if consumption_path and os.path.exists(consumption_path):
            try:
                _xl = pd.ExcelFile(consumption_path)
                if "Excluded SKUs" in _xl.sheet_names:
                    _df_ex = pd.read_excel(consumption_path,
                                           sheet_name="Excluded SKUs", header=2)
                    _df_ex.columns = [str(c).strip() for c in _df_ex.columns]
                    if "SKUCode" in _df_ex.columns:
                        _df_ex["SKUCode"] = _df_ex["SKUCode"].astype(str).str.strip()
                        for _, er in _df_ex.iterrows():
                            sku = str(er["SKUCode"]).strip()
                            if not sku or sku.lower() == "nan" or sku in active_skus:
                                continue
                            _demand = float(er.get("Demand_Qty", 0) or 0)
                            _remark = str(er.get("Remark",
                                "No master data — building machine or curing mould missing"))
                            avg_ct_ex  = _sku_ct(sku)
                            elig_ex    = len(_allow_lkp.get(sku, set()))
                            dem_rows.append({
                                "SKUCode":           sku,
                                "Category":          "Excluded",
                                "Priority":          0.0,
                                "Demand":            int(_demand),
                                "GT_Inventory":      int(float(_gt_inv_map.get(sku, 0))),
                                "Planned_Units":     0,
                                "Gap":               int(_demand),
                                "Fulfillment_Pct":   "0.0%",
                                "Status":            "EXCLUDED",
                                "CycleTime_min":     avg_ct_ex if avg_ct_ex is not None else "NA",
                                "Eligible_Machines": elig_ex if elig_ex else "NA",
                                "Presses_Needed":    (
                                    round(_demand * avg_ct_ex / _PRESS_NORM, 2)
                                    if avg_ct_ex and _demand > 0 else "NA"
                                ),
                                "Skip_Reason":       _remark,
                            })
                            active_skus.add(sku)
            except Exception as _ee:
                print(f"  ⚠  Could not load Excluded SKUs sheet: {_ee}")

        # 4. Fill Skip_Reason and write sheet
        df_dem = pd.DataFrame(dem_rows)
        if not df_dem.empty:
            df_dem["Skip_Reason"] = df_dem.apply(
                lambda r: (
                    r["Skip_Reason"] if str(r["Status"]) == "EXCLUDED"
                    else _skip_reason(
                        sku               = str(r["SKUCode"]),
                        status            = str(r["Status"]),
                        category          = str(r["Category"]),
                        co_targets        = co_targets,
                        planned_gt        = float(r["Planned_Units"]),
                        eligible_machines = int(r["Eligible_Machines"])
                            if str(r["Eligible_Machines"]) not in ("NA", "nan") else 0,
                    )
                ),
                axis=1,
            )
            df_dem = df_dem.sort_values(
                ["Category", "Priority"], ascending=[True, False]
            ).reset_index(drop=True)

            ws_dem = _add_sheet(wb, "Demand Fulfillment (B2C)", df_dem)

            # Bold Planned_Units column
            _col_idx = {c: i+1 for i, c in enumerate(df_dem.columns)}
            _pu_col  = _col_idx.get("Planned_Units")
            if _pu_col:
                for _ri in range(2, len(df_dem) + 2):
                    ws_dem.cell(row=_ri, column=_pu_col).font = Font(bold=True)

            # Colour-code rows by status
            status_colors = {
                "FULLY MET": _GREEN, "PARTIAL": _AMBER,
                "UNMET": _RED,       "EXCLUDED": _GREY,
            }
            for ri, (_, row) in enumerate(df_dem.iterrows(), start=2):
                fill_hex = status_colors.get(str(row.get("Status", "")), "")
                if fill_hex:
                    _fill = PatternFill("solid", fgColor=fill_hex)
                    for ci in range(1, len(df_dem.columns) + 1):
                        ws_dem.cell(row=ri, column=ci).fill = _fill

            # KPI footer
            n_full  = int((df_dem["Status"] == "FULLY MET").sum())
            n_part  = int((df_dem["Status"] == "PARTIAL").sum())
            n_unmet = int((df_dem["Status"] == "UNMET").sum())
            n_excl  = int((df_dem["Status"] == "EXCLUDED").sum())
            total_built  = int(df_dem["Planned_Units"].sum())
            total_demand = int(df_dem["Demand"].sum())
            kpi_pct = round(100 * total_built / total_demand, 1) if total_demand else 0.0

            footer = len(df_dem) + 3
            ws_dem.cell(row=footer,   column=1, value="KPI SUMMARY").font = Font(bold=True)
            ws_dem.cell(row=footer+1, column=1, value="Total Customer Demand (units)  ← Net GT demand = this value")
            ws_dem.cell(row=footer+1, column=2, value=total_demand)
            ws_dem.cell(row=footer+2, column=1, value="Total GT Built (units)")
            ws_dem.cell(row=footer+2, column=2, value=total_built)
            _kpi = ws_dem.cell(row=footer+3, column=1, value="KPI — GT Built / Customer Demand")
            _kpi.font = Font(bold=True)
            _kpiv = ws_dem.cell(row=footer+3, column=2, value=f"{kpi_pct}%")
            _kpiv.font = Font(bold=True)
            ws_dem.cell(row=footer+4, column=1,
                        value="Note: GT Built > Customer Demand possible — TopUp pre-builds inventory for next period")
            ws_dem.cell(row=footer+6, column=1, value="Total SKUs in demand file")
            ws_dem.cell(row=footer+6, column=2, value=len(df_dem))
            ws_dem.cell(row=footer+7, column=1, value="Fully Met (≥95% of demand built)")
            ws_dem.cell(row=footer+7, column=2, value=n_full)
            ws_dem.cell(row=footer+8, column=1, value="Partial (0 < built < 95%)")
            ws_dem.cell(row=footer+8, column=2, value=n_part)
            ws_dem.cell(row=footer+9, column=1, value="Unmet (0 built)")
            ws_dem.cell(row=footer+9, column=2, value=n_unmet)
            ws_dem.cell(row=footer+10, column=1, value="Excluded (no machine / mould data)")
            ws_dem.cell(row=footer+10, column=2, value=n_excl)
            ws_dem.cell(row=footer+11, column=1, value="Total Building COs")
            ws_dem.cell(row=footer+11, column=2, value=total_co_count)
            ws_dem.cell(row=footer+12, column=1, value="Curing COs scheduled (≤8/day)")
            ws_dem.cell(row=footer+12, column=2, value=len(df_co_plan) if df_co_plan is not None else 0)

            print(f"  [B2C KPI] GT Built / Customer Demand = {total_built:,} / {total_demand:,} = {kpi_pct}%")
            print(f"  [B2C Sheets] Demand Fulfillment (B2C): {len(df_dem)} SKUs "
                  f"(FULLY MET={n_full}, PARTIAL={n_part}, UNMET={n_unmet}, EXCLUDED={n_excl})")
    except Exception as _e:
        import traceback
        print(f"  ⚠  Demand Fulfillment (B2C) sheet failed: {_e}")
        print(traceback.format_exc())

    wb.save(output_path)
    print(f"  [B2C Sheets] Saved → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# STANDALONE RUN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    _PLAN_START       = datetime(2026, 6, 1, 7, 0, 0)
    _CONSUMPTION_PATH = os.path.join(cbc_env.OUTPUT_DIR, "curing_consumption_table.xlsx")
    _OUTPUT_PATH      = os.path.join(MAIN_OUT, "bc_building_schedule.xlsx")

    results = run_from_database_b2c(
        plan_start       = _PLAN_START,
        consumption_path = _CONSUMPTION_PATH,
        output_path      = _OUTPUT_PATH,
    )

    # Quick summary
    ss = results.get("shift_schedule")
    if ss is not None and len(ss) > 0:
        prod = ss[~ss["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"])]
        total_gt = prod[prod["Machine"].isin(
            [str(m) for m in (Config.STAGE2 | Config.UNISTAGE)]
        )]["Qty"].sum()
        print(f"\n  Total GT produced across schedule: {total_gt:,.0f}")
